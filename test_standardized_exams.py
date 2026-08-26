import os
import sys
import django
import datetime
from decimal import Decimal
import io
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

from django.test import Client
from apps.accounts.models import User
from apps.academics.models import AcademicYear, Classroom, Subject, GradeLevelRule
from apps.students.models import Student
from apps.examinations.models import (
    StandardizedExam, ExamRoom, ExamSubject, ExamCandidate, CandidateSubjectScore
)

def test_standardized_exams():
    print("🚀 Starting Automated Test for Cambodian High School Standardized Exams...")

    # 1. Setup Admin & Academic Year
    admin_user = User.objects.filter(is_superuser=True).first()
    if not admin_user:
        admin_user = User.objects.create_superuser('admin_test', 'admin@test.com', 'adminpass123')

    ay, _ = AcademicYear.objects.get_or_create(
        name='2025-2026',
        defaults={'start_date': '2025-09-01', 'end_date': '2026-07-15', 'is_current': True}
    )

    # 2. Setup Classrooms and Students (Create 30 students to test 25/room partition)
    cls_12a, _ = Classroom.objects.get_or_create(
        code='12A1',
        academic_year=ay,
        defaults={'name': 'ថ្នាក់ទី១២A1', 'grade_level': 12, 'track': 'SCIENCE'}
    )

    # Clear old test students if any
    Student.objects.filter(student_id__startswith='TEST_STD_').delete()
    for i in range(1, 31):
        Student.objects.create(
            student_id=f"TEST_STD_{i:03d}",
            khmer_name=f"សិស្ស តេស្ត {i:02d}",
            latin_name=f"Student Test {i:02d}",
            gender='F' if i % 2 == 0 else 'M',
            date_of_birth=datetime.date(2008, (i % 12) + 1, (i % 25) + 1),
            classroom=cls_12a,
            academic_year=ay,
            status='ACTIVE'
        )


    # 3. Create Standardized Exam
    StandardizedExam.objects.filter(name='ការប្រឡងតេស្តស្តង់ដា ឆមាសទី១ ថ្នាក់ទី១២ (Test)').delete()
    exam = StandardizedExam.objects.create(
        name='ការប្រឡងតេស្តស្តង់ដា ឆមាសទី១ ថ្នាក់ទី១២ (Test)',
        academic_year=ay,
        grade_level=12,
        track='ALL',
        exam_date=datetime.date.today(),
        candidates_per_room=25,
        is_published=True
    )
    print(f"✅ 1. Created StandardizedExam: {exam.name}")

    # 4. Setup Exam Subjects with Coefficients & Max Scores
    sub_math, _ = Subject.objects.get_or_create(code='M_TEST', defaults={'name_kh': 'គណិតវិទ្យា', 'name_en': 'Math', 'credit': 2})
    sub_khmer, _ = Subject.objects.get_or_create(code='K_TEST', defaults={'name_kh': 'ភាសាខ្មែរ', 'name_en': 'Khmer', 'credit': 2})
    sub_phys, _ = Subject.objects.get_or_create(code='P_TEST', defaults={'name_kh': 'រូបវិទ្យា', 'name_en': 'Physics', 'credit': 1})

    es_math = ExamSubject.objects.create(exam=exam, subject=sub_math, max_score=Decimal('100.00'), coefficient=Decimal('2.00'), order=1)
    es_khmer = ExamSubject.objects.create(exam=exam, subject=sub_khmer, max_score=Decimal('100.00'), coefficient=Decimal('2.00'), order=2)
    es_phys = ExamSubject.objects.create(exam=exam, subject=sub_phys, max_score=Decimal('50.00'), coefficient=Decimal('1.00'), order=3)
    print(f"✅ 2. Created 3 ExamSubjects (Math x2.0, Khmer x2.0, Phys x1.0). Total Coefficients = 5.0")

    # 5. Test 1-Click Pull Candidates
    client = Client()
    client.force_login(admin_user)

    res_pull = client.post(f'/examinations/standardized/{exam.id}/pull-candidates/', follow=True)
    assert res_pull.status_code == 200
    candidate_count = exam.candidates.count()
    assert candidate_count >= 30, f"Expected >= 30 candidates, got {candidate_count}"
    print(f"✅ 3. Pulled {candidate_count} candidates from Grade 12 into ExamCandidate records successfully.")

    # 6. Test Auto-Generate Rooms (25 candidates per room)
    res_rooms = client.post(f'/examinations/standardized/{exam.id}/generate-rooms/', follow=True)
    assert res_rooms.status_code == 200
    rooms = list(exam.rooms.all().order_by('room_number'))
    assert len(rooms) >= 2, f"Expected at least 2 rooms for 30 candidates (25 in Room 1, 5 in Room 2), got {len(rooms)}"
    
    room_1 = rooms[0]
    room_2 = rooms[1]
    assert room_1.candidates.count() == 25, f"Room 1 should have 25 candidates, got {room_1.candidates.count()}"
    assert room_2.candidates.count() >= 5, f"Room 2 should have >= 5 candidates, got {room_2.candidates.count()}"

    # Verify Desk Numbers 01 to 25
    desk_nums = list(room_1.candidates.values_list('desk_number', flat=True))
    assert desk_nums == list(range(1, 26)), f"Desk numbers must be 1..25 sequentially! Got: {desk_nums}"
    print(f"✅ 4. Auto-generated {len(rooms)} rooms. Room 01 has {room_1.candidates.count()} candidates (Desks 01-25). Room 02 has {room_2.candidates.count()} candidates.")

    # 7. Test Entering Scores and Auto-Calculations (Total, Weighted Average, Grade A-F, Ranks)
    # Give Candidate 1 top scores (Math=95, Khmer=90, Phys=48) -> Percentage = (233/250)*100 = 93.2% -> Grade A
    c1 = room_1.candidates.first()
    CandidateSubjectScore.objects.filter(candidate=c1, exam_subject=es_math).update(score=Decimal('95.00'))
    CandidateSubjectScore.objects.filter(candidate=c1, exam_subject=es_khmer).update(score=Decimal('90.00'))
    CandidateSubjectScore.objects.filter(candidate=c1, exam_subject=es_phys).update(score=Decimal('48.00'))

    # Give Candidate 2 middle scores (Math=50, Khmer=55, Phys=25) -> Percentage = (130/250)*100 = 52.0% -> Grade E
    c2 = room_1.candidates.all()[1]
    CandidateSubjectScore.objects.filter(candidate=c2, exam_subject=es_math).update(score=Decimal('50.00'))
    CandidateSubjectScore.objects.filter(candidate=c2, exam_subject=es_khmer).update(score=Decimal('55.00'))
    CandidateSubjectScore.objects.filter(candidate=c2, exam_subject=es_phys).update(score=Decimal('25.00'))

    # Give Candidate 3 low scores (Math=20, Khmer=20, Phys=10) -> Percentage = (50/250)*100 = 20.0% -> Grade F
    c3 = room_1.candidates.all()[2]
    CandidateSubjectScore.objects.filter(candidate=c3, exam_subject=es_math).update(score=Decimal('20.00'))
    CandidateSubjectScore.objects.filter(candidate=c3, exam_subject=es_khmer).update(score=Decimal('20.00'))
    CandidateSubjectScore.objects.filter(candidate=c3, exam_subject=es_phys).update(score=Decimal('10.00'))

    exam.recalculate_all_ranks()

    c1.refresh_from_db()
    c2.refresh_from_db()
    c3.refresh_from_db()

    assert c1.total_score == Decimal('233.00'), f"C1 total expected 233.00, got {c1.total_score}"
    assert c1.grade_letter == 'A', f"C1 Grade expected 'A', got {c1.grade_letter}"
    assert c1.rank_overall == 1, f"C1 Rank overall expected 1, got {c1.rank_overall}"
    assert c1.rank_in_room == 1, f"C1 Rank in room expected 1, got {c1.rank_in_room}"

    assert c2.grade_letter == 'E', f"C2 Grade expected 'E', got {c2.grade_letter}"
    assert c3.grade_letter == 'F', f"C3 Grade expected 'F', got {c3.grade_letter}"
    print(f"✅ 5. Score calculation validated: C1 Total={c1.total_score} (Avg={c1.average_score}, Grade={c1.grade_letter}, Rank=#{c1.rank_overall})")

    # 8. Test HTTP Views & Official Print Sheets
    # A. Dashboard
    res_dash = client.get(f'/examinations/standardized/{exam.id}/manage/')
    assert res_dash.status_code == 200 and 'បញ្ជីឈ្មោះបេក្ខជនប្រឡង' in res_dash.content.decode('utf-8')
    print("✅ 6. GET /examinations/standardized/<id>/manage/ -> 200 OK")

    # B. Official Room Notice Board Postings Sheet (បញ្ជីបិទផ្សាយតាមបន្ទប់)
    res_postings = client.get(f'/examinations/standardized/{exam.id}/room-postings/')
    assert res_postings.status_code == 200 and 'បញ្ជីឈ្មោះបេក្ខជនតាមបន្ទប់ប្រឡង' in res_postings.content.decode('utf-8')
    print("✅ 7. GET /examinations/standardized/<id>/room-postings/ -> 200 OK (MoEYS Room Notice Posting Sheet)")

    # C. Official Subject Attendance Signature Sheet (បញ្ជីវត្តមានចុះហត្ថលេខា)
    res_att = client.get(f'/examinations/standardized/{exam.id}/attendance-sheets/')
    assert res_att.status_code == 200 and 'បញ្ជីវត្តមាន និងហត្ថលេខាបេក្ខជនប្រឡង' in res_att.content.decode('utf-8')
    print("✅ 8. GET /examinations/standardized/<id>/attendance-sheets/ -> 200 OK (MoEYS Attendance & Signature Sheet)")

    # D. Room Scores Entry Matrix
    res_scores = client.get(f'/examinations/standardized/{exam.id}/scores-entry/?room_id={room_1.id}')
    assert res_scores.status_code == 200 and 'តារាងបញ្ចូលពិន្ទុ' in res_scores.content.decode('utf-8')
    print("✅ 9. GET /examinations/standardized/<id>/scores-entry/ -> 200 OK (Rapid Score Entry Matrix)")

    # E. Master Provisional Results Board (តារាងបិទផ្សាយបណ្តោះអាសន្ន)
    res_prov_rank = client.get(f'/examinations/standardized/{exam.id}/provisional-results/?sort=rank')
    assert res_prov_rank.status_code == 200 and 'តារាងបិទផ្សាយបណ្តោះអាសន្ន' in res_prov_rank.content.decode('utf-8')

    res_prov_name = client.get(f'/examinations/standardized/{exam.id}/provisional-results/?sort=name')
    assert res_prov_name.status_code == 200
    print("✅ 10. GET /examinations/standardized/<id>/provisional-results/ -> 200 OK (Tested Sort by Rank & Sort by Name)")

    # 9. Test Excel Exports
    res_cand_excel = client.get(f'/examinations/standardized/{exam.id}/export-candidates/')
    assert res_cand_excel.status_code == 200
    wb_cand = openpyxl.load_workbook(io.BytesIO(res_cand_excel.content))
    assert 'បញ្ជីបេក្ខជនប្រឡង' in wb_cand.sheetnames
    print("✅ 11. Candidates Excel Export (.xlsx) validated successfully.")

    res_prov_excel = client.get(f'/examinations/standardized/{exam.id}/export-provisional-excel/')
    assert res_prov_excel.status_code == 200
    wb_prov = openpyxl.load_workbook(io.BytesIO(res_prov_excel.content))
    assert 'លទ្ធផលបណ្តោះអាសន្ន' in wb_prov.sheetnames
    print("✅ 12. Provisional Results Excel Export (.xlsx) validated successfully.")

    # Cleanup test exam
    exam.delete()
    print("\n🎉 ALL TESTS PASSED! Cambodian High School Standardized Examination System is 100% OPERATIONAL & VERIFIED!")

if __name__ == '__main__':
    test_standardized_exams()
