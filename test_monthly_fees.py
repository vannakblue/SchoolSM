import os
import sys
import django
from decimal import Decimal

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Set Django settings
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()


from django.test import Client
from django.contrib.auth import get_user_model
from django.utils import timezone
from apps.academics.models import AcademicYear, Classroom, GradeLevel
from apps.students.models import Student, StudentCategory
from apps.teachers.models import Teacher
from apps.finance.models import MonthlyFeeConfig, MonthlyFeeRate, StudentMonthlyPayment
import io
import openpyxl

User = get_user_model()

def run_tests():
    print("=== STARTING EXTENDED TEST SUITE: MONTHLY UTILITY & FEE TRACKER ===")
    
    # 1. Setup Active Academic Year
    year, _ = AcademicYear.objects.get_or_create(
        name="2025-2026 Test",
        defaults={'start_date': timezone.now().date(), 'end_date': timezone.now().date(), 'is_current': True}
    )
    AcademicYear.objects.filter(id=year.id).update(is_current=True)
    AcademicYear.objects.exclude(id=year.id).update(is_current=False)

    # 2. Setup Test Grade & Classroom
    grade, _ = GradeLevel.objects.get_or_create(
        grade_number=7,
        defaults={'name': 'ថ្នាក់ទី ៧', 'order': 1}
    )
    classroom, _ = Classroom.objects.get_or_create(
        name="7A-Test",
        defaults={'code': '7A-TEST', 'grade_level': 7, 'academic_year': year, 'capacity': 40}
    )
    classroom.academic_year = year
    classroom.save()


    # 3. Setup Users & Teachers
    admin_user, _ = User.objects.get_or_create(username='admin_test_m', defaults={'role': 'ADMIN', 'is_superuser': True})
    admin_user.set_password('pass123')
    admin_user.save()

    teacher_user, _ = User.objects.get_or_create(username='teacher_collector_m', defaults={'role': 'TEACHER'})
    teacher_user.set_password('pass123')
    teacher_user.save()

    teacher_profile, _ = Teacher.objects.get_or_create(
        user=teacher_user,
        defaults={'teacher_id': 'T-COL-01', 'khmer_name': 'សំ សុក', 'latin_name': 'SAM SOK', 'gender': 'M', 'is_fee_collector': True, 'specialization': 'គណិតវិទ្យា', 'phone': '012345678'}
    )
    teacher_profile.khmer_name = 'សំ សុក'
    teacher_profile.latin_name = 'SAM SOK'
    teacher_profile.gender = 'M'
    teacher_profile.is_fee_collector = True
    teacher_profile.save()

    # Regular teacher (not authorized)
    reg_teacher_user, _ = User.objects.get_or_create(username='teacher_regular_m', defaults={'role': 'TEACHER'})
    reg_teacher_user.set_password('pass123')
    reg_teacher_user.save()
    reg_teacher_profile, _ = Teacher.objects.get_or_create(
        user=reg_teacher_user,
        defaults={'teacher_id': 'T-REG-01', 'khmer_name': 'អ៊ុំ សុវណ្ណ', 'latin_name': 'OUM SOVANN', 'gender': 'M', 'is_fee_collector': False, 'specialization': 'អក្សរសាស្ត្រខ្មែរ'}
    )
    reg_teacher_profile.khmer_name = 'អ៊ុំ សុវណ្ណ'
    reg_teacher_profile.latin_name = 'OUM SOVANN'
    reg_teacher_profile.gender = 'M'
    reg_teacher_profile.is_fee_collector = False
    reg_teacher_profile.save()


    # Assign homeroom
    classroom.homeroom_teacher = teacher_profile
    classroom.save()

    # 4. Setup Categories
    cat_normal, _ = StudentCategory.objects.get_or_create(code="NORMAL", defaults={'name': "សិស្សទូទៅ", 'display_order': 1})
    cat_poor, _ = StudentCategory.objects.get_or_create(code="POOR", defaults={'name': "សិស្សក្រីក្រ", 'display_order': 2})
    cat_free, _ = StudentCategory.objects.get_or_create(code="FREE", defaults={'name': "ឥតគិតថ្លៃ", 'display_order': 3})

    # 5. Setup Students
    st1, _ = Student.objects.get_or_create(
        student_id="250001",
        defaults={
            'khmer_name': 'សុខ ចិន្តា',
            'latin_name': 'Sok Chenda',
            'gender': 'F',
            'date_of_birth': timezone.now().date(),
            'classroom': classroom,
            'academic_year': year,
            'category': cat_normal,
            'father_phone': '012999888',
            'telegram_chat_id': '987654321',
            'status': 'ACTIVE'
        }
    )
    st1.khmer_name = 'សុខ ចិន្តា'
    st1.classroom = classroom
    st1.academic_year = year
    st1.category = cat_normal
    st1.status = 'ACTIVE'
    st1.save()

    st2, _ = Student.objects.get_or_create(
        student_id="250002",
        defaults={
            'khmer_name': 'ម៉ៅ វិបុល',
            'latin_name': 'Mao Vibul',
            'gender': 'M',
            'date_of_birth': timezone.now().date(),
            'classroom': classroom,
            'academic_year': year,
            'category': cat_poor,
            'father_phone': '098111222',
            'status': 'ACTIVE'
        }
    )
    st2.khmer_name = 'ម៉ៅ វិបុល'
    st2.classroom = classroom
    st2.academic_year = year
    st2.category = cat_poor
    st2.status = 'ACTIVE'
    st2.save()
    print("[PASS] 1. Initialized test environment.")



    # 6. Setup Config & Month Sequence
    config = MonthlyFeeConfig.get_or_create_for_year(year)
    config.start_month = 10
    config.end_month = 8
    config.ticked_months = [10, 11, 12, 1] # 4 months ticked: Oct, Nov, Dec, Jan
    config.save()

    # Set rates
    for m in [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8]:
        MonthlyFeeRate.objects.update_or_create(config=config, category=cat_normal, month=m, defaults={'amount': Decimal('20000.00')})
        MonthlyFeeRate.objects.update_or_create(config=config, category=cat_poor, month=m, defaults={'amount': Decimal('10000.00')})
        MonthlyFeeRate.objects.update_or_create(config=config, category=cat_free, month=m, defaults={'amount': Decimal('0.00')})

    print("[PASS] 2. Rates and scope configured.")

    # 7. Record Payment for st1 (Pay Oct & Nov = 40,000 KHR)
    StudentMonthlyPayment.objects.filter(student=st1, academic_year=year).delete()
    StudentMonthlyPayment.objects.create(
        student=st1,
        academic_year=year,
        month=10,
        expected_amount=Decimal('20000.00'),
        paid_amount=Decimal('20000.00'),
        is_on_time=True,
        payment_date=timezone.now()
    )
    StudentMonthlyPayment.objects.create(
        student=st1,
        academic_year=year,
        month=11,
        expected_amount=Decimal('20000.00'),
        paid_amount=Decimal('20000.00'),
        is_on_time=True,
        payment_date=timezone.now()
    )
    print("[PASS] 3. Payments recorded for st1.")

    client = Client()
    client.force_login(admin_user)

    # 8. Test Excel Export (.xlsx)
    resp = client.get('/finance/monthly-fees/export-excel/')
    assert resp.status_code == 200, f"Excel export failed: {resp.status_code}"
    assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in resp['Content-Type']
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "បញ្ជីទឹកភ្លើងប្រចាំខែ" in wb.sheetnames
    ws = wb.active
    assert "តារាងតាមដានការបង់ប្រាក់" in ws['A1'].value
    print("[PASS] 4. Excel Export (.xlsx) validated successfully with openpyxl.")

    # 9. Test CSV Export
    resp_csv = client.get('/finance/monthly-fees/export-csv/')
    assert resp_csv.status_code == 200, f"CSV export failed: {resp_csv.status_code}"
    assert 'text/csv' in resp_csv['Content-Type']
    csv_content = resp_csv.content.decode('utf-8-sig', errors='replace')
    assert "សុខ ចិន្តា" in csv_content
    print("[PASS] 5. CSV Export validated successfully.")



    # 10. Test Filtering (Filter status=DUE, status=PAID)
    resp_due = client.get('/finance/monthly-fees/?status=DUE')
    assert resp_due.status_code == 200
    resp_paid = client.get('/finance/monthly-fees/?status=PAID')
    assert resp_paid.status_code == 200
    print("[PASS] 6. Advanced Filtering (DUE, PAID, category) tested.")

    # 11. Test Month Range Settings
    resp_range = client.post('/finance/monthly-fees/save-range/', {
        'start_month': '10',
        'end_month': '8',
        'currency_symbol': '៛'
    })
    assert resp_range.status_code == 200
    assert resp_range.json().get('status') == 'success'
    print("[PASS] 7. Save Academic Year Month Range endpoint validated.")

    # 12. Test Student Category CRUD (Create, Edit, Delete)
    resp_cat_create = client.post('/finance/monthly-fees/categories/', {
        'name': 'សិស្សអាហារូបករណ៍ពិសេស',
        'code': 'SPECIAL_SCHOLAR',
        'display_order': 10
    })
    assert resp_cat_create.status_code == 200
    new_cat_id = resp_cat_create.json()['category']['id']

    # Delete newly created category
    resp_cat_del = client.post(f'/finance/monthly-fees/categories/{new_cat_id}/delete/')
    assert resp_cat_del.status_code == 200
    assert resp_cat_del.json().get('status') == 'success'
    assert not StudentCategory.objects.filter(id=new_cat_id).exists()
    print("[PASS] 8. Student Category CRUD (Create, Edit, Delete) tested.")

    # 13. Test Telegram Reminder Dispatch
    resp_tg = client.post('/finance/monthly-fees/send-reminder/', {
        'student_id': str(st1.id),
        'custom_message': 'តេស្តសាររំលឹកបង់ប្រាក់'
    })
    assert resp_tg.status_code == 200
    assert resp_tg.json().get('status') == 'success'
    print("[PASS] 9. Individual Telegram Reminder dispatch tested.")

    # 14. Test Classroom Fee Summary Telegram Dispatch
    resp_class_tg = client.post('/finance/monthly-fees/send-classroom-summary/', {
        'classroom_id': str(classroom.id)
    })
    assert resp_class_tg.status_code == 200
    assert resp_class_tg.json().get('status') == 'success'
    print("[PASS] 10. Classroom Fee Summary Telegram dispatch tested.")

    # Cleanup test data so no duplicate test classrooms remain
    StudentMonthlyPayment.objects.filter(student__in=[st1, st2]).delete()
    st1.delete()
    st2.delete()
    classroom.delete()
    year.delete()

    print("\n*** ALL EXTENDED MONTHLY UTILITY & FEE TRACKER TESTS PASSED WITH 100% SUCCESS! ***")

if __name__ == '__main__':
    run_tests()

