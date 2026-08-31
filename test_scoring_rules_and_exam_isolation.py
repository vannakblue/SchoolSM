import os
import django
import io
import openpyxl
from decimal import Decimal
from datetime import timedelta

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

from django.utils import timezone
from django.test import Client
from django.core.files.uploadedfile import SimpleUploadedFile
from apps.accounts.models import User
from apps.teachers.models import Teacher
from apps.academics.models import AcademicYear, GradeLevel, Subject, GradeLevelRule
from apps.examinations.models import (
    StandardizedExam, ExamRoom, ExamSubject, ExamRoomSubjectCode,
    ExamCandidate, CandidateSubjectScore
)

def run_tests():
    print("=== STARTING SCORING RULES EXPORT/IMPORT & EXAM SESSION ISOLATION TESTS ===")

    admin_user = User.objects.filter(role='ADMIN').first()
    if not admin_user:
        admin_user = User.objects.create_superuser('admin_iso_test', 'admin_iso@school.com', 'adminpass123')

    client = Client()
    client.force_login(admin_user)

    active_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.first()

    # ----------------- 1. TEST SCORING RULES EXPORT EXCEL & CSV -----------------
    # 1.1 Excel Export
    resp_xl = client.get('/academics/scoring-rules/export-excel/')
    assert resp_xl.status_code == 200, f"Expected 200 for Excel export, got {resp_xl.status_code}"
    assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in resp_xl['Content-Type']
    wb = openpyxl.load_workbook(io.BytesIO(resp_xl.content))
    assert 'Scoring Rules Matrix' in wb.sheetnames
    ws = wb['Scoring Rules Matrix']
    print(f"1. [PASS] Scoring Rules Excel Export validated successfully! ({ws.max_row} rows, {ws.max_column} columns)")

    # 1.2 CSV Export
    resp_csv = client.get('/academics/scoring-rules/export-csv/')
    assert resp_csv.status_code == 200, f"Expected 200 for CSV export, got {resp_csv.status_code}"
    assert 'text/csv' in resp_csv['Content-Type']
    csv_text = resp_csv.content.decode('utf-8-sig')
    assert "Grade_Level_Name" in csv_text and "Total_Max_Score" in csv_text
    print("2. [PASS] Scoring Rules CSV Export (UTF-8 BOM) validated successfully!")

    # ----------------- 2. TEST SCORING RULES IMPORT -----------------
    # Create a test Excel workbook to import
    sub_math = Subject.objects.filter(code='M').first() or Subject.objects.filter(name_kh__contains='គណិត').first()
    assert sub_math, "Math subject required for test"

    import_wb = openpyxl.Workbook()
    import_ws = import_wb.active
    import_ws.append(["No", "Grade_Level_Name", "Grade_Number", "Track", f"{sub_math.name_kh} ({sub_math.code})", "Total_Max_Score"])
    import_ws.append([1, "ថ្នាក់ទី ៧", 7, "GENERAL", 125.0, 125.0])
    import_ws.append([2, "ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រ", 12, "SCIENCE", 175.0, 175.0])

    out_file = io.BytesIO()
    import_wb.save(out_file)
    out_file.seek(0)

    uploaded_xl = SimpleUploadedFile("test_scoring_rules_import.xlsx", out_file.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    resp_import = client.post('/academics/scoring-rules/import/', {'file': uploaded_xl}, follow=True)
    assert resp_import.status_code == 200, f"Expected 200 for Import, got {resp_import.status_code}"

    # Verify rules were updated in DB
    rule_g7 = GradeLevelRule.objects.filter(grade_level=7, track='GENERAL', subject=sub_math).first()
    assert rule_g7 and rule_g7.max_score == Decimal('125.00'), f"Expected 125.00 for G7 Math, got {rule_g7.max_score if rule_g7 else 'None'}"

    rule_g12 = GradeLevelRule.objects.filter(grade_level=12, track='SCIENCE', subject=sub_math).first()
    assert rule_g12 and rule_g12.max_score == Decimal('175.00'), f"Expected 175.00 for G12 Science Math, got {rule_g12.max_score if rule_g12 else 'None'}"

    print("3. [PASS] Scoring Rules Import (.xlsx) successfully parsed and updated GradeLevelRules in DB!")

    # ----------------- 3. TEST STANDARDIZED EXAM SESSION ISOLATION & AUDIT TRAIL -----------------
    now = timezone.now()

    # Create Exam Session 1 (May Exam)
    exam_may = StandardizedExam.objects.create(
        name="តេស្តសមត្ថភាព ប្រចាំខែឧសភា",
        academic_year=active_year,
        grade_level=12,
        track='SCIENCE',
        exam_date=now.date() - timedelta(days=30),
        candidates_per_room=25
    )
    import uuid
    uid_str = uuid.uuid4().hex[:5].upper()
    code_may_str = f"MAY{uid_str}"
    code_june_str = f"JUN{uid_str}"

    room_may = ExamRoom.objects.create(exam=exam_may, room_number=1, room_name="បន្ទប់ ១ (May)", secret_code=f"SEC-{code_may_str}")
    es_may = ExamSubject.objects.create(exam=exam_may, subject=sub_math, max_score=Decimal('50.00'), coefficient=Decimal('1.00'))
    code_may = ExamRoomSubjectCode.objects.create(exam_room=room_may, exam_subject=es_may, secret_code=code_may_str)

    cand_may = ExamCandidate.objects.create(
        exam=exam_may, room=room_may, roll_number="R-001", desk_number=1,
        candidate_name_kh="បេក្ខជន ខែឧសភា", gender='M'
    )

    # Create Exam Session 2 (June Exam)
    exam_june = StandardizedExam.objects.create(
        name="តេស្តសមត្ថភាព ប្រចាំខែមិថុនា",
        academic_year=active_year,
        grade_level=12,
        track='SCIENCE',
        exam_date=now.date(),
        candidates_per_room=25
    )
    room_june = ExamRoom.objects.create(exam=exam_june, room_number=1, room_name="បន្ទប់ ១ (June)", secret_code=f"SEC-{code_june_str}")
    es_june = ExamSubject.objects.create(exam=exam_june, subject=sub_math, max_score=Decimal('50.00'), coefficient=Decimal('1.00'))
    code_june = ExamRoomSubjectCode.objects.create(exam_room=room_june, exam_subject=es_june, secret_code=code_june_str)

    cand_june = ExamCandidate.objects.create(
        exam=exam_june, room=room_june, roll_number="R-001", desk_number=1,
        candidate_name_kh="បេក្ខជន ខែមិថុនា", gender='F'
    )

    # 3.1 Save Blind Scores for May Exam
    payload_may = {
        'exam_id': exam_may.id,
        'subject_id': es_may.id,
        'secret_code': code_may_str,
        'scores': [{'desk_number': 1, 'score': '42.50'}]
    }
    resp_sm = client.post('/examinations/standardized/api/save-blind-scores/', payload_may, content_type='application/json')
    assert resp_sm.status_code == 200, f"May save failed: {resp_sm.status_code}"

    # 3.2 Save Blind Scores for June Exam
    payload_june = {
        'exam_id': exam_june.id,
        'subject_id': es_june.id,
        'secret_code': code_june_str,
        'scores': [{'desk_number': 1, 'score': '49.00'}]
    }
    resp_sj = client.post('/examinations/standardized/api/save-blind-scores/', payload_june, content_type='application/json')
    assert resp_sj.status_code == 200, f"June save failed: {resp_sj.status_code}"

    # 3.3 Verify Complete Isolation between May and June Exams
    score_may = CandidateSubjectScore.objects.filter(candidate=cand_may, exam_subject=es_may).first()
    score_june = CandidateSubjectScore.objects.filter(candidate=cand_june, exam_subject=es_june).first()

    assert score_may and score_may.score == Decimal('42.50'), "May score must be 42.50"
    assert score_may.entered_by == admin_user, "Entered by user must be tracked in audit trail"
    assert score_may.secret_code_used == code_may_str, f"Secret code used must be {code_may_str}"
    assert score_may.entered_at is not None, "Entered at timestamp must be populated"

    assert score_june and score_june.score == Decimal('49.00'), "June score must be 49.00"
    assert score_june.entered_by == admin_user
    assert score_june.secret_code_used == code_june_str, f"Secret code used must be {code_june_str}"
    assert score_june.entered_at is not None

    print("4. [PASS] Exam Sessions are 100% isolated with independent secret codes, rooms, and candidate scores!")
    print("5. [PASS] Grader Audit Trail successfully verified (entered_by, entered_at, secret_code_used)!")

    print("\n=== ALL SCORING RULES EXPORT/IMPORT & EXAM ISOLATION TESTS PASSED 100%! ===")

if __name__ == '__main__':
    run_tests()
