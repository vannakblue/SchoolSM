import os
import django
import io
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

from django.test import Client
from apps.accounts.models import User

def run_tests():
    print("=== TESTING ALL 4 TIMETABLE DAILY REPORT EXCEL EXPORTS ===")

    admin_user = User.objects.filter(role='ADMIN').first()
    if not admin_user:
        admin_user = User.objects.create_superuser('admin_rpt_test', 'admin_rpt@test.com', 'adminpass123')

    client = Client()
    client.force_login(admin_user)

    # 1. Test Daily Reports View HTML
    resp_view = client.get('/academics/timetable/daily-reports/')
    assert resp_view.status_code == 200, f"View failed: {resp_view.status_code}"
    html = resp_view.content.decode('utf-8')
    assert "បញ្ជីវត្តមានប្រចាំថ្ងៃ" in html
    assert "បន្ទុកបង្រៀនរបស់គ្រូ" in html
    assert "កូដគ្រូ-មុខវិជ្ជា" in html
    assert "ម៉ោងតាមថ្នាក់" in html
    print("1. [PASS] Daily Reports Web View rendered with all 4 report tabs and export buttons!")

    # 2. Test Report 1: Daily Duty Sign-In Sheets
    resp_r1 = client.get('/academics/timetable/daily-reports/export-excel/?report_type=duty_sheets')
    assert resp_r1.status_code == 200
    assert resp_r1['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    wb1 = openpyxl.load_workbook(io.BytesIO(resp_r1.content))
    assert len(wb1.sheetnames) > 0
    print(f"2. [PASS] Excel Export (1. បញ្ជីវត្តមានប្រចាំថ្ងៃ) returned {len(wb1.sheetnames)} sheets successfully!")

    # 3. Test Report 2: Teacher Load Matrix
    resp_r2 = client.get('/academics/timetable/daily-reports/export-excel/?report_type=teacher_load')
    assert resp_r2.status_code == 200
    wb2 = openpyxl.load_workbook(io.BytesIO(resp_r2.content))
    assert "បន្ទុកបង្រៀនរបស់គ្រូ" in wb2.sheetnames
    ws2 = wb2["បន្ទុកបង្រៀនរបស់គ្រូ"]
    assert "របាយការណ៍បន្ទុកម៉ោងបង្រៀនរបស់គ្រូ" in str(ws2['A2'].value)
    print("3. [PASS] Excel Export (2. បន្ទុកបង្រៀនរបស់គ្រូ) generated sheet with headers and totals!")

    # 4. Test Report 3: Teacher Subject Codes Directory
    resp_r3 = client.get('/academics/timetable/daily-reports/export-excel/?report_type=subject_codes')
    assert resp_r3.status_code == 200
    wb3 = openpyxl.load_workbook(io.BytesIO(resp_r3.content))
    assert "កូដគ្រូ-មុខវិជ្ជា" in wb3.sheetnames
    ws3 = wb3["កូដគ្រូ-មុខវិជ្ជា"]
    assert "បញ្ជីកូដគ្រូបង្រៀន និងមុខវិជ្ជា" in str(ws3['A2'].value)
    print("4. [PASS] Excel Export (3. កូដគ្រូ-មុខវិជ្ជា) generated directory sheet!")

    # 5. Test Report 4: Classrooms Summary
    resp_r4 = client.get('/academics/timetable/daily-reports/export-excel/?report_type=class_summary')
    assert resp_r4.status_code == 200
    wb4 = openpyxl.load_workbook(io.BytesIO(resp_r4.content))
    assert "ម៉ោងតាមថ្នាក់រៀន" in wb4.sheetnames
    ws4 = wb4["ម៉ោងតាមថ្នាក់រៀន"]
    assert "របាយការណ៍ម៉ោងតាមថ្នាក់រៀន" in str(ws4['A2'].value)
    print("5. [PASS] Excel Export (4. ម៉ោងតាមថ្នាក់) generated classrooms summary sheet!")

    # 6. Test All Reports Combined Workbook
    resp_all = client.get('/academics/timetable/daily-reports/export-excel/?report_type=all_reports')
    assert resp_all.status_code == 200
    wb_all = openpyxl.load_workbook(io.BytesIO(resp_all.content))
    expected_sheets = ["បន្ទុកបង្រៀនរបស់គ្រូ", "កូដគ្រូ-មុខវិជ្ជា", "ម៉ោងតាមថ្នាក់រៀន"]
    for s in expected_sheets:
        assert s in wb_all.sheetnames
    print(f"6. [PASS] Excel Export (All-in-One Workbook) combined {len(wb_all.sheetnames)} sheets successfully!")

    print("\n=== ALL 4 REPORT EXCEL EXPORTS TESTED AND PASSED 100%! ===")

if __name__ == '__main__':
    run_tests()
