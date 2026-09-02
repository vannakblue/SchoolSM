import json
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.db import models, transaction
from django.utils import timezone
from decimal import Decimal
from datetime import datetime, timedelta
from apps.accounts.decorators import role_required
from apps.accounts.utils import send_telegram_notification
from .models import (
    FeeCategory, Invoice, PaymentTransaction, Expense, Payroll,
    MonthlyFeeConfig, MonthlyFeeRate, StudentMonthlyPayment, StudentMonthlyCategory
)
from .forms import FeeCategoryForm, InvoiceForm, PaymentTransactionForm, ExpenseForm
from apps.students.models import Student, StudentCategory
from apps.teachers.models import Teacher, TeacherAttendance
from apps.academics.models import Classroom, AcademicYear, GradeLevel
from apps.academics.utils import get_active_academic_year

@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def fee_category_list(request):
    categories = FeeCategory.objects.all()
    if request.method == 'POST':
        form = FeeCategoryForm(request.POST)
        if form.is_valid():
            cat = form.save()
            messages.success(request, f"បានបង្កើតប្រភេទកម្រៃ {cat.name} ជោគជ័យ!")
            return redirect('fee_category_list')
    else:
        form = FeeCategoryForm()
    return render(request, 'finance/fee_category_list.html', {'categories': categories, 'form': form})


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def invoice_list(request):
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    status_filter = request.GET.get('status', '')
    query = request.GET.get('q', '').strip()

    invoices = Invoice.objects.select_related('student', 'fee_category', 'academic_year').all()
    if active_year:
        invoices = invoices.filter(academic_year=active_year)

    if status_filter:
        invoices = invoices.filter(status=status_filter)
    if query:
        invoices = invoices.filter(
            Q(invoice_no__icontains=query) |
            Q(student__khmer_name__icontains=query) |
            Q(student__latin_name__icontains=query) |
            Q(student__student_id__icontains=query)
        )

    # Summary metrics
    total_billed = invoices.aggregate(s=Sum('final_amount'))['s'] or Decimal('0.00')
    total_collected = invoices.aggregate(s=Sum('paid_amount'))['s'] or Decimal('0.00')
    total_due = max(Decimal('0.00'), total_billed - total_collected)

    return render(request, 'finance/invoice_list.html', {
        'invoices': invoices,
        'status_filter': status_filter,
        'query': query,
        'total_billed': total_billed,
        'total_collected': total_collected,
        'total_due': total_due,
        'statuses': Invoice.Status.choices,
        'active_year': active_year,
    })


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def invoice_create(request):
    from apps.academics.utils import get_active_academic_year
    current_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save()
            messages.success(request, f"បានចេញវិក្កយបត្រ {invoice.invoice_no} ជោគជ័យ!")
            return redirect('invoice_detail', pk=invoice.pk)
    else:
        due_default = (datetime.now() + timedelta(days=15)).date()
        form = InvoiceForm(initial={'academic_year': current_year, 'due_date': due_default})

    return render(request, 'finance/invoice_form.html', {'form': form, 'title': 'ចេញវិក្កយបត្រថ្មី / Create Invoice'})


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def invoice_batch_create(request):
    """
    Batch invoice generation for an entire class
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    classrooms = Classroom.objects.filter(academic_year=active_year) if active_year else Classroom.objects.all()
    categories = FeeCategory.objects.all()
    academic_years = AcademicYear.objects.all().order_by('-start_date')

    if request.method == 'POST':
        classroom_id = request.POST.get('classroom')
        category_id = request.POST.get('fee_category')
        year_id = request.POST.get('academic_year')
        due_date_str = request.POST.get('due_date')
        base_amount_str = request.POST.get('amount')

        classroom = get_object_or_404(Classroom, pk=classroom_id)
        fee_cat = get_object_or_404(FeeCategory, pk=category_id)
        academic_year = get_object_or_404(AcademicYear, pk=year_id)
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        base_amount = Decimal(base_amount_str)

        students = Student.objects.filter(classroom=classroom, status='ACTIVE')
        created_count = 0

        with transaction.atomic():
            for stu in students:
                # Calculate discount based on scholarship
                discount_rate = Decimal('0.00')
                if stu.scholarship_type == Student.ScholarshipType.SCHOLARSHIP_50:
                    discount_rate = Decimal('50.00')
                elif stu.scholarship_type == Student.ScholarshipType.SCHOLARSHIP_100:
                    discount_rate = Decimal('100.00')

                final_amt = base_amount * ((Decimal('100.00') - discount_rate) / Decimal('100.00'))

                Invoice.objects.create(
                    student=stu,
                    fee_category=fee_cat,
                    academic_year=academic_year,
                    original_amount=base_amount,
                    discount_percent=discount_rate,
                    final_amount=final_amt,
                    due_date=due_date,
                    notes=f"ចេញវិក្កយបត្រស្វ័យប្រវត្តិតាមថ្នាក់ {classroom.name}"
                )
                created_count += 1

        messages.success(request, f"🎉 បានចេញវិក្កយបត្រជូនសិស្សថ្នាក់ {classroom.name} ចំនួន {created_count} សន្លឹកដោយជោគជ័យ!")
        return redirect('invoice_list')

    return render(request, 'finance/invoice_batch_form.html', {
        'classrooms': classrooms,
        'categories': categories,
        'academic_years': academic_years,
        'default_due': (datetime.now() + timedelta(days=15)).strftime('%Y-%m-%d'),
        'active_year': active_year,
    })


@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice.objects.select_related('student', 'fee_category', 'academic_year'), pk=pk)
    
    # Check student role permission
    if request.user.role == 'STUDENT':
        if not hasattr(request.user, 'student_profile') or request.user.student_profile.id != invoice.student.id:
            messages.error(request, "លោកអ្នកអាចមើលបានតែវិក្កយបត្រផ្ទាល់ខ្លួនប៉ុណ្ណោះ!")
            return redirect('student_dashboard')

    payment_form = PaymentTransactionForm(initial={'amount': invoice.remaining_balance})
    payments = invoice.payments.all()

    return render(request, 'finance/invoice_detail.html', {
        'invoice': invoice,
        'payment_form': payment_form,
        'payments': payments,
    })


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def record_payment(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        form = PaymentTransactionForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.invoice = invoice
            payment.received_by = request.user
            payment.save()
            messages.success(request, f"✅ បានកត់ត្រាការបង់ប្រាក់ចំនួន ${payment.amount} (បង្កាន់ដៃ: {payment.receipt_number}) ជោគជ័យ!")
            return redirect('official_receipt', pk=payment.pk)
    return redirect('invoice_detail', pk=invoice.pk)


@login_required
def official_receipt(request, pk):
    payment = get_object_or_404(PaymentTransaction.objects.select_related('invoice__student', 'invoice__fee_category', 'received_by'), pk=pk)
    
    # Check student permission
    if request.user.role == 'STUDENT':
        if not hasattr(request.user, 'student_profile') or request.user.student_profile.id != payment.invoice.student.id:
            messages.error(request, "លោកអ្នកអាចមើលបានតែបង្កាន់ដៃផ្ទាល់ខ្លួនប៉ុណ្ណោះ!")
            return redirect('student_dashboard')

    return render(request, 'finance/official_receipt.html', {
        'payment': payment,
        'invoice': payment.invoice,
        'student': payment.invoice.student,
    })


@login_required
def due_fees_list(request):
    """
    Shows unpaid / overdue invoices
    """
    due_invoices = Invoice.objects.filter(
        status__in=[Invoice.Status.PENDING, Invoice.Status.PARTIAL, Invoice.Status.OVERDUE]
    ).select_related('student', 'fee_category', 'academic_year')
    return render(request, 'finance/due_fees_list.html', {'due_invoices': due_invoices})


# -------------------------------------------------------------------------
# MONTHLY UTILITY & MAINTENANCE FEE SYSTEM (ប្រព័ន្ធគ្រប់គ្រងថ្លៃទឹក-ភ្លើងប្រចាំខែ)
# -------------------------------------------------------------------------

def is_monthly_fee_admin(user):
    """Check if user has full config permission for monthly fees"""
    if getattr(user, 'role', '') in ['ADMIN', 'DIRECTOR', 'ACCOUNTANT']:
        return True
    return False


def is_fee_collector(user):
    """Check if user is a teacher assigned as fee collector or admin"""
    if getattr(user, 'role', '') in ['ADMIN', 'DIRECTOR', 'ACCOUNTANT']:
        return True
    if getattr(user, 'role', '') == 'TEACHER' and hasattr(user, 'teacher_profile') and user.teacher_profile.is_fee_collector:
        return True
    return False


MONTH_NAMES_KM = {
    1: 'មករា (Jan)',
    2: 'កុម្ភៈ (Feb)',
    3: 'មីនា (Mar)',
    4: 'មេសា (Apr)',
    5: 'ឧសភា (May)',
    6: 'មិថុនា (Jun)',
    7: 'កក្កដា (Jul)',
    8: 'សីហា (Aug)',
    9: 'កញ្ញា (Sep)',
    10: 'តុលា (Oct)',
    11: 'វិច្ឆិកា (Nov)',
    12: 'ធ្នូ (Dec)',
}

ALL_12_MONTHS = [
    {'number': 1, 'name': 'ខែ ១ - មករា (January)'},
    {'number': 2, 'name': 'ខែ ២ - កុម្ភៈ (February)'},
    {'number': 3, 'name': 'ខែ ៣ - មីនា (March)'},
    {'number': 4, 'name': 'ខែ ៤ - មេសា (April)'},
    {'number': 5, 'name': 'ខែ ៥ - ឧសភា (May)'},
    {'number': 6, 'name': 'ខែ ៦ - មិថុនា (June)'},
    {'number': 7, 'name': 'ខែ ៧ - កក្កដា (July)'},
    {'number': 8, 'name': 'ខែ ៨ - សីហា (August)'},
    {'number': 9, 'name': 'ខែ ៩ - កញ្ញា (September)'},
    {'number': 10, 'name': 'ខែ ១០ - តុលា (October)'},
    {'number': 11, 'name': 'ខែ ១១ - វិច្ឆិកា (November)'},
    {'number': 12, 'name': 'ខែ ១២ - ធ្នូ (December)'},
]

def check_fee_collector_access(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or getattr(user, 'role', '') in ['ADMIN', 'ACCOUNTANT']:
        return True
    if getattr(user, 'role', '') == 'TEACHER' and hasattr(user, 'teacher_profile') and user.teacher_profile.is_fee_collector:
        return True
    return False


def get_monthly_fees_data(request):
    """
    Unified calculation helper for Monthly Utility & Fee Tracker,
    Excel export, CSV export, and Telegram summaries.
    Strictly isolated per Academic Year!
    """
    from apps.students.models import Student, StudentCategory
    from apps.academics.models import Classroom, AcademicYear, GradeLevel
    from apps.academics.utils import get_active_academic_year

    active_year = get_active_academic_year(request)
    if not active_year:
        active_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-id').first()
    config = MonthlyFeeConfig.get_or_create_for_year(active_year)

    # Ensure default student categories exist if database is empty
    if not StudentCategory.objects.exists():
        StudentCategory.objects.create(name="សិស្សទូទៅ (Normal)", code="NORMAL", display_order=1)
        StudentCategory.objects.create(name="សិស្សក្រីក្រ (Poor)", code="POOR", display_order=2)
        StudentCategory.objects.create(name="កូនគ្រូបង្រៀន (Teacher's Child)", code="TEACHER_CHILD", display_order=3)
        StudentCategory.objects.create(name="អាហារូបករណ៍ (Scholarship)", code="SCHOLARSHIP", display_order=4)
        StudentCategory.objects.create(name="ឥតគិតថ្លៃ (Free 100%)", code="FREE", display_order=5)

    categories = StudentCategory.objects.filter(is_active=True).order_by('display_order', 'id')
    default_category = categories.first()

    month_numbers = config.get_month_sequence() if config else [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8]
    ticked_months_set = set(config.ticked_months or []) if config else set()

    months_data = []
    for m in month_numbers:
        months_data.append({
            'number': m,
            'name': MONTH_NAMES_KM.get(m, f'ខែ {m}'),
            'short_name': MONTH_NAMES_KM.get(m, f'M{m}').split(' ')[0],
            'is_ticked': m in ticked_months_set,
        })

    rates = MonthlyFeeRate.objects.filter(config=config) if config else []
    rate_map = {(r.category_id, r.month): r.amount for r in rates}

    matrix_rates = []
    for cat in categories:
        cat_row = {'category': cat, 'months': []}
        for m in month_numbers:
            amt = rate_map.get((cat.id, m))
            if amt is None:
                if 'FREE' in cat.code or '100' in cat.name:
                    amt = Decimal('0.00')
                elif 'SCHOLAR' in cat.code or '50' in cat.name or 'TEACHER' in cat.code:
                    amt = Decimal('10000.00')
                else:
                    amt = Decimal('20000.00')
            cat_row['months'].append({'month': m, 'amount': amt})
        matrix_rates.append(cat_row)

    classroom_filter = request.GET.get('classroom', '')
    grade_level_filter = request.GET.get('grade_level', '')
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', 'ALL')
    search_q = request.GET.get('q', '').strip()

    student_status_filter = request.GET.get('student_status', '')
    students_qs = Student.objects.all().select_related('classroom', 'category')
    if active_year:
        students_qs = students_qs.filter(Q(academic_year=active_year) | Q(classroom__academic_year=active_year))

    if student_status_filter and student_status_filter != 'ALL':
        students_qs = students_qs.filter(status=student_status_filter)
    elif not student_status_filter:
        # Default: Show active students, plus any dropped/suspended students who have a fee_end_month set
        students_qs = students_qs.filter(Q(status='ACTIVE') | Q(fee_end_month__isnull=False))

    if classroom_filter:
        students_qs = students_qs.filter(classroom_id=classroom_filter)
    if grade_level_filter:
        if str(grade_level_filter).isdigit():
            gl = GradeLevel.objects.filter(id=grade_level_filter).first()
            if gl:
                if gl.track and gl.track != 'GENERAL':
                    students_qs = students_qs.filter(classroom__grade_level=gl.grade_number, classroom__track=gl.track)
                else:
                    students_qs = students_qs.filter(classroom__grade_level=gl.grade_number)
            else:
                students_qs = students_qs.filter(classroom__grade_level=int(grade_level_filter))
        else:
            students_qs = students_qs.filter(classroom__grade_level=grade_level_filter)
    if category_filter:
        if category_filter == 'NONE':
            students_qs = students_qs.filter(category__isnull=True)
        else:
            students_qs = students_qs.filter(category_id=category_filter)
    if search_q:
        students_qs = students_qs.filter(
            Q(khmer_name__icontains=search_q) |
            Q(latin_name__icontains=search_q) |
            Q(student_id__icontains=search_q) |
            Q(phone__icontains=search_q) |
            Q(father_phone__icontains=search_q) |
            Q(mother_phone__icontains=search_q)
        )

    payments = StudentMonthlyPayment.objects.filter(academic_year=active_year).select_related('collected_by') if active_year else []
    payment_map = {(p.student_id, p.month): p for p in payments}

    from apps.finance.models import StudentMonthlyCategory
    monthly_cat_records = StudentMonthlyCategory.objects.filter(academic_year=active_year).select_related('category') if active_year else []
    monthly_cat_map = {(mc.student_id, mc.month): mc.category for mc in monthly_cat_records}

    students_data = []
    kpi_total_expected = Decimal('0.00')
    kpi_total_collected = Decimal('0.00')
    kpi_fully_paid_count = 0
    kpi_due_count = 0

    for st in students_qs:
        st_cat = st.category or default_category
        st_cat_id = st_cat.id if st_cat else None

        # Determine pre-enrollment and post-stop/dropout months for this student
        fee_start = st.fee_start_month
        fee_end = st.fee_end_month

        start_idx = 0
        if fee_start and fee_start in month_numbers:
            start_idx = month_numbers.index(fee_start)

        end_idx = len(month_numbers) - 1
        if fee_end and fee_end in month_numbers:
            end_idx = month_numbers.index(fee_end)

        st_months = []
        st_expected_ticked = Decimal('0.00')
        st_paid_ticked = Decimal('0.00')
        st_total_paid = Decimal('0.00')
        unpaid_months_names = []

        for idx, m in enumerate(month_numbers):
            is_pre_enrollment = (idx < start_idx)
            is_post_stop = (idx > end_idx)
            is_attending = (not is_pre_enrollment) and (not is_post_stop)
            is_ticked = (m in ticked_months_set) and is_attending
            expected = Decimal('0.00')

            m_cat = monthly_cat_map.get((st.id, m), st_cat)
            m_cat_id = m_cat.id if m_cat else None

            if m_cat_id and is_attending:
                expected = rate_map.get((m_cat_id, m))
                if expected is None:
                    if m_cat and ('FREE' in m_cat.code or '100' in m_cat.name):
                        expected = Decimal('0.00')
                    elif m_cat and ('SCHOLAR' in m_cat.code or '50' in m_cat.name or 'TEACHER' in m_cat.code):
                        expected = Decimal('10000.00')
                    else:
                        expected = Decimal('20000.00')
            elif is_attending:
                expected = Decimal('20000.00')

            payment = payment_map.get((st.id, m))
            paid = payment.paid_amount if payment else Decimal('0.00')
            st_total_paid += paid

            if is_ticked:
                st_expected_ticked += expected
                st_paid_ticked += paid

            if paid >= expected and expected > 0:
                cell_status = 'PAID'
                cell_label = 'បង់គ្រប់'
                badge_class = 'badge-success-soft'
            elif expected == 0 and is_ticked:
                cell_status = 'PAID'
                cell_label = 'ឥតគិតថ្លៃ'
                badge_class = 'badge-info-soft'
            elif paid > 0:
                cell_status = 'PARTIAL'
                cell_label = f'បង់បាន {paid:,.0f}'
                badge_class = 'badge-warning-soft'
                unpaid_months_names.append(MONTH_NAMES_KM.get(m, f'ខែ {m}'))
            elif is_pre_enrollment or is_post_stop or not (m in ticked_months_set):
                cell_status = 'NOT_DUE'
                if is_pre_enrollment:
                    cell_label = 'មិនទាន់ចូល'
                elif is_post_stop:
                    if st.status == 'DROPPED':
                        cell_label = 'ឈប់រៀន'
                    elif st.status == 'SUSPENDED':
                        cell_label = 'ផ្អាក'
                    elif st.status == 'TRANSFERRED':
                        cell_label = 'ផ្ទេរ'
                    else:
                        cell_label = 'ផ្អាកបង់'
                else:
                    cell_label = 'មិនទាន់ដល់ខែ'
                badge_class = 'badge-secondary-soft'
            else:
                cell_status = 'UNPAID'
                cell_label = 'ជំពាក់'
                badge_class = 'badge-danger-soft'
                unpaid_months_names.append(MONTH_NAMES_KM.get(m, f'ខែ {m}'))

            st_months.append({
                'month': m,
                'month_name': MONTH_NAMES_KM.get(m, f'ខែ {m}'),
                'is_ticked': is_ticked,
                'is_pre_enrollment': is_pre_enrollment,
                'is_post_stop': is_post_stop,
                'expected': expected,
                'paid': paid,
                'remaining': max(Decimal('0.00'), expected - paid) if is_ticked else Decimal('0.00'),
                'cell_status': cell_status,
                'cell_label': cell_label,
                'badge_class': badge_class,
                'payment': payment,
                'is_on_time': payment.is_on_time if payment else True,
                'payment_date': payment.payment_date if payment else None,
            })

        st_remaining_ticked = max(Decimal('0.00'), st_expected_ticked - st_paid_ticked)
        is_fully_paid = (st_remaining_ticked == Decimal('0.00'))

        if is_fully_paid:
            kpi_fully_paid_count += 1
        else:
            kpi_due_count += 1

        kpi_total_expected += st_expected_ticked
        kpi_total_collected += st_total_paid

        # Status filter check
        if status_filter == 'PAID' and not is_fully_paid:
            continue
        elif status_filter == 'DUE' and is_fully_paid:
            continue
        elif status_filter == 'UNPAID' and (st_total_paid > 0 or st_expected_ticked == 0):
            continue
        elif status_filter == 'PARTIAL' and (st_total_paid == 0 or is_fully_paid):
            continue

        guardian_phone = st.father_phone or st.mother_phone or st.phone or ''
        guardian_name = st.father_name or st.mother_name or st.guardian_name or st.khmer_name

        import json
        months_status_map = {m['month']: {'status': m['cell_status'], 'paid': float(m['paid']), 'expected': float(m['expected'])} for m in st_months}
        paid_month_numbers = [m['month'] for m in st_months if m['cell_status'] == 'PAID']
        unpaid_month_numbers = [m['month'] for m in st_months if m['is_ticked'] and m['cell_status'] in ['UNPAID', 'PARTIAL']]

        students_data.append({
            'student': st,
            'category': st_cat,
            'fee_start_month': st.fee_start_month,
            'fee_start_month_name': MONTH_NAMES_KM.get(st.fee_start_month, 'ដើមឆ្នាំ') if st.fee_start_month else 'ដើមឆ្នាំ',
            'fee_end_month': st.fee_end_month,
            'fee_end_month_name': MONTH_NAMES_KM.get(st.fee_end_month, 'ចប់ឆ្នាំ') if st.fee_end_month else 'ចប់ឆ្នាំ',
            'status': st.status,
            'status_display': st.get_status_display(),
            'months': st_months,
            'total_expected': st_expected_ticked,
            'total_paid': st_total_paid,
            'remaining_balance': st_remaining_ticked,
            'is_fully_paid': is_fully_paid,
            'unpaid_months_str': ', '.join(unpaid_months_names) if unpaid_months_names else 'គ្មាន',
            'guardian_phone': guardian_phone,
            'guardian_name': guardian_name,
            'paid_month_numbers': paid_month_numbers,
            'unpaid_month_numbers': unpaid_month_numbers,
            'months_status_json': json.dumps(months_status_map),
        })

    kpi_total_remaining = max(Decimal('0.00'), kpi_total_expected - kpi_total_collected)
    classrooms = Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code')
    grade_levels = GradeLevel.objects.all().order_by('order', 'grade_number')
    teachers = Teacher.objects.filter(status='ACTIVE').order_by('teacher_id')


    return {
        'active_year': active_year,
        'config': config,
        'categories': categories,
        'months_data': months_data,
        'all_12_months': ALL_12_MONTHS,
        'matrix_rates': matrix_rates,
        'students_data': students_data,
        'students_count': len(students_data),
        'kpi_total_expected': kpi_total_expected,
        'kpi_total_collected': kpi_total_collected,
        'kpi_total_remaining': kpi_total_remaining,
        'kpi_fully_paid_count': kpi_fully_paid_count,
        'kpi_due_count': kpi_due_count,
        'classrooms': classrooms,
        'grade_levels': grade_levels,
        'teachers': teachers,
        'selected_classroom': classroom_filter,
        'selected_grade_level': grade_level_filter,
        'selected_category': category_filter,
        'selected_status': status_filter,
        'selected_student_status': student_status_filter,
        'search_q': search_q,
    }


@login_required
def monthly_fees_tracker(request):
    """
    Main tracker & manager for Monthly Utilities & Fees (បញ្ជីទឹកភ្លើងប្រចាំខែ)
    """
    if not check_fee_collector_access(request.user):
        messages.error(request, "⛔ លោកអ្នកគ្មានសិទ្ធិចូលមើលផ្ទាំងប្រមូលថវិកានេះទេ! មានតែគ្រូដែល Admin បានអនុញ្ញាតប៉ុណ្ណោះ។")
        return redirect('teacher_dashboard' if request.user.role == 'TEACHER' else 'login')

    context = get_monthly_fees_data(request)
    context['is_admin'] = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'
    context['is_accountant'] = getattr(request.user, 'role', '') == 'ACCOUNTANT'
    context['is_teacher'] = getattr(request.user, 'role', '') == 'TEACHER'

    return render(request, 'finance/monthly_fees_tracker.html', context)


@login_required
def export_monthly_fees_excel(request):
    """
    Export filtered Monthly Fees & Utilities report to styled Excel (.xlsx)
    """
    if not check_fee_collector_access(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied")

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    data = get_monthly_fees_data(request)
    config = data['config']
    currency_symbol = config.currency_symbol if config else '៛'
    months_data = data['months_data']
    students_data = data['students_data']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "បញ្ជីទឹកភ្លើងប្រចាំខែ"

    title_font = Font(name='Khmer OS Siemreap', size=13, bold=True, color='1E3A8A')
    meta_font = Font(name='Khmer OS Siemreap', size=9, italic=True, color='475569')
    header_font = Font(name='Khmer OS Siemreap', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    
    data_font = Font(name='Khmer OS Siemreap', size=9)
    bold_font = Font(name='Khmer OS Siemreap', size=9, bold=True)
    green_font = Font(name='Khmer OS Siemreap', size=9, bold=True, color='15803D')
    red_font = Font(name='Khmer OS Siemreap', size=9, bold=True, color='DC2626')

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    footer_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    # Row 1: Title
    total_cols = 8 + len(months_data) + 4
    col_end_letter = get_column_letter(total_cols)
    ws.merge_cells(f'A1:{col_end_letter}1')
    title_cell = ws['A1']
    title_cell.value = "តារាងតាមដានការបង់ប្រាក់ថ្លៃទឹកភ្លើង និងសេវាសិក្សាប្រចាំខែ (Monthly Utilities & Fee Report)"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Row 2: Metadata Filters
    active_year_name = data['active_year'].name if data['active_year'] else 'បច្ចុប្បន្ន'
    ws.merge_cells(f'A2:{col_end_letter}2')
    meta_cell = ws['A2']
    meta_cell.value = f"ឆ្នាំសិក្សា៖ {active_year_name} | ស្ថានភាព៖ {data['selected_status']} | កាលបរិច្ឆេទ Export៖ {datetime.now().strftime('%d/%m/%Y %H:%M')} | សរុបសិស្ស៖ {len(students_data)} នាក់"
    meta_cell.font = meta_font
    meta_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # Row 4: Column Headers
    headers = ["ល.រ", "អត្តលេខ", "ឈ្មោះខ្មែរ", "ឈ្មោះឡាតាំង", "ភេទ", "ថ្នាក់", "ប្រភេទសិស្ស", "លេខទូរស័ព្ទ"]
    for m in months_data:
        m_tag = " (Tick)" if m['is_ticked'] else ""
        headers.append(f"{m['short_name']}{m_tag}")
    headers.extend(["សរុបត្រូវបង់", "សរុបបានបង់", "សរុបនៅជំពាក់", "ស្ថានភាព"])

    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 26

    # Data Rows
    current_row = 5
    for idx, item in enumerate(students_data, 1):
        st = item['student']
        row_vals = [
            idx,
            st.student_id,
            st.khmer_name,
            st.latin_name,
            "ប្រុស" if st.gender == 'M' else "ស្រី",
            st.classroom.name if st.classroom else "-",
            item['category'].name if item['category'] else "ទូទៅ",
            item['guardian_phone'] or "-",
        ]

        for m_item in item['months']:
            if not m_item['is_ticked']:
                row_vals.append("-")
            elif m_item['cell_status'] == 'PAID':
                row_vals.append(f"{m_item['paid']:,.0f}")
            elif m_item['cell_status'] == 'PARTIAL':
                row_vals.append(f"បង់ {m_item['paid']:,.0f}")
            else:
                row_vals.append("ជំពាក់")

        row_vals.append(item['total_expected'])
        row_vals.append(item['total_paid'])
        row_vals.append(item['remaining_balance'])
        row_vals.append("បង់គ្រប់" if item['is_fully_paid'] else "នៅជំពាក់")

        for col_idx, val in enumerate(row_vals, 1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = data_font
            c.border = thin_border

            if col_idx in [1, 2, 5, 6, 8]:
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [3, 4, 7]:
                c.alignment = Alignment(horizontal='left', vertical='center')
            else:
                c.alignment = Alignment(horizontal='center', vertical='center')

            if isinstance(val, (int, float, Decimal)):
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right', vertical='center')

            if col_idx == len(headers) - 1:
                c.font = red_font if item['remaining_balance'] > 0 else green_font
            elif col_idx == len(headers):
                c.font = green_font if item['is_fully_paid'] else red_font

        current_row += 1

    # Summary Footer Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers) - 4)
    total_label = ws.cell(row=current_row, column=1, value="សរុបថវិកាទាំងអស់ (TOTAL)")
    total_label.font = bold_font
    total_label.alignment = Alignment(horizontal='right', vertical='center')
    total_label.fill = footer_fill
    total_label.border = thin_border

    for c_i in range(2, len(headers) - 3):
        ws.cell(row=current_row, column=c_i).border = thin_border
        ws.cell(row=current_row, column=c_i).fill = footer_fill

    exp_c = ws.cell(row=current_row, column=len(headers) - 3, value=data['kpi_total_expected'])
    exp_c.font = bold_font
    exp_c.number_format = '#,##0'
    exp_c.alignment = Alignment(horizontal='right', vertical='center')
    exp_c.fill = footer_fill
    exp_c.border = thin_border

    col_c = ws.cell(row=current_row, column=len(headers) - 2, value=data['kpi_total_collected'])
    col_c.font = green_font
    col_c.number_format = '#,##0'
    col_c.alignment = Alignment(horizontal='right', vertical='center')
    col_c.fill = footer_fill
    col_c.border = thin_border

    rem_c = ws.cell(row=current_row, column=len(headers) - 1, value=data['kpi_total_remaining'])
    rem_c.font = red_font
    rem_c.number_format = '#,##0'
    rem_c.alignment = Alignment(horizontal='right', vertical='center')
    rem_c.fill = footer_fill
    rem_c.border = thin_border

    status_c = ws.cell(row=current_row, column=len(headers), value="")
    status_c.fill = footer_fill
    status_c.border = thin_border
    ws.row_dimensions[current_row].height = 22

    # Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 2 and cell.value is not None:
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Monthly_Fees_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_monthly_fees_csv(request):
    """
    Export filtered Monthly Fees & Utilities report to UTF-8 CSV (with BOM)
    """
    if not check_fee_collector_access(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied")

    import csv
    from django.http import HttpResponse

    data = get_monthly_fees_data(request)
    months_data = data['months_data']
    students_data = data['students_data']

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f"Monthly_Fees_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')

    writer = csv.writer(response)


    headers = ["ល.រ", "អត្តលេខ", "ឈ្មោះខ្មែរ", "ឈ្មោះឡាតាំង", "ភេទ", "ថ្នាក់", "ប្រភេទសិស្ស", "លេខទូរស័ព្ទ"]
    for m in months_data:
        m_tag = " (Tick)" if m['is_ticked'] else ""
        headers.append(f"{m['short_name']}{m_tag}")
    headers.extend(["សរុបត្រូវបង់", "សរុបបានបង់", "សរុបនៅជំពាក់", "ស្ថានភាព"])
    writer.writerow(headers)

    for idx, item in enumerate(students_data, 1):
        st = item['student']
        row = [
            idx,
            st.student_id,
            st.khmer_name,
            st.latin_name,
            "ប្រុស" if st.gender == 'M' else "ស្រី",
            st.classroom.name if st.classroom else "-",
            item['category'].name if item['category'] else "ទូទៅ",
            item['guardian_phone'] or "-",
        ]
        for m_item in item['months']:
            if not m_item['is_ticked']:
                row.append("-")
            elif m_item['cell_status'] == 'PAID':
                row.append(f"{m_item['paid']:,.0f}")
            elif m_item['cell_status'] == 'PARTIAL':
                row.append(f"បង់ {m_item['paid']:,.0f}")
            else:
                row.append("ជំពាក់")
        row.append(f"{item['total_expected']:,.0f}")
        row.append(f"{item['total_paid']:,.0f}")
        row.append(f"{item['remaining_balance']:,.0f}")
        row.append("បង់គ្រប់" if item['is_fully_paid'] else "នៅជំពាក់")
        writer.writerow(row)

    # Footer total row
    footer = ["សរុបទាំងអស់ (TOTAL)", "", "", "", "", "", "", ""]
    for _ in months_data:
        footer.append("")
    footer.append(f"{data['kpi_total_expected']:,.0f}")
    footer.append(f"{data['kpi_total_collected']:,.0f}")
    footer.append(f"{data['kpi_total_remaining']:,.0f}")
    footer.append("")
    writer.writerow(footer)

    return response


@login_required
def save_monthly_fee_scope(request):
    """
    AJAX handler for Admin to tick/untick active due months
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        active_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-id').first()
        config = MonthlyFeeConfig.get_or_create_for_year(active_year)
        
        raw_months = request.POST.getlist('months[]') or request.POST.getlist('months')
        if not raw_months and request.content_type == 'application/json':
            import json
            try:
                data = json.loads(request.body.decode('utf-8'))
                raw_months = data.get('months', [])
            except Exception:
                raw_months = []

        try:
            ticked = [int(m) for m in raw_months if m]
            config.ticked_months = ticked
            config.save()
            return JsonResponse({'status': 'success', 'ticked_months': config.ticked_months})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def save_monthly_fee_range(request):
    """
    Admin tool to configure academic year payment month range (start_month to end_month)
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        import json

        start_month = request.POST.get('start_month')
        end_month = request.POST.get('end_month')
        currency_symbol = request.POST.get('currency_symbol', '៛')

        if not start_month and request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                start_month = data.get('start_month')
                end_month = data.get('end_month')
                currency_symbol = data.get('currency_symbol', '៛')
            except Exception:
                pass

        if not start_month or not end_month:
            return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសខែចាប់ផ្តើម និងខែបញ្ចប់!'}, status=400)

        active_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-id').first()
        config = MonthlyFeeConfig.get_or_create_for_year(active_year)
        config.start_month = int(start_month)
        config.end_month = int(end_month)
        config.currency_symbol = currency_symbol
        config.save()

        return JsonResponse({
            'status': 'success',
            'message': f'🎉 បានរក្សាទុកការកំណត់ខែបង់ប្រាក់ពីខែ {config.start_month} ដល់ខែ {config.end_month} ជោគជ័យ!'
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def record_student_monthly_payment(request):
    """
    AJAX / POST handler to record fee payment for single or multiple months
    """
    if not check_fee_collector_access(request.user):
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        import json
        
        student_id = request.POST.get('student_id')
        action = request.POST.get('action', 'SAVE').upper()
        payment_method = request.POST.get('payment_method', 'CASH')
        notes = request.POST.get('notes', '')
        custom_amount = request.POST.get('amount')
        is_on_time = request.POST.get('is_on_time', 'true') == 'true'

        raw_months = request.POST.getlist('months[]') or request.POST.getlist('months')

        if not student_id and request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                student_id = data.get('student_id')
                action = data.get('action', 'SAVE').upper()
                raw_months = data.get('months', [])
                payment_method = data.get('payment_method', 'CASH')
                notes = data.get('notes', '')
                custom_amount = data.get('amount')
                is_on_time = data.get('is_on_time', True)
            except Exception:
                pass

        if not student_id:
            return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសសិស្ស!'}, status=400)

        student = get_object_or_404(Student, pk=student_id)
        active_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-id').first()
        config = MonthlyFeeConfig.get_or_create_for_year(active_year)

        # Handle Revert / Reset to Unpaid (or if 0 months provided)
        if action in ['REVERT', 'DELETE', 'RESET'] or not raw_months:
            target_months = []
            for m_str in raw_months:
                try:
                    target_months.append(int(m_str))
                except ValueError:
                    pass
            
            if target_months:
                deleted_count, _ = StudentMonthlyPayment.objects.filter(
                    student=student,
                    academic_year=active_year,
                    month__in=target_months
                ).delete()
            else:
                deleted_count, _ = StudentMonthlyPayment.objects.filter(
                    student=student,
                    academic_year=active_year
                ).delete()

            return JsonResponse({
                'status': 'success',
                'message': f'🔄 បានដោះ Tick និងកែប្រែត្រឡប់ទៅជា «ជំពាក់» វិញសម្រាប់សិស្ស {student.khmer_name} ជោគជ័យ!'
            })

        # When saving: parse selected target months
        target_months = []
        for m_str in raw_months:
            try:
                target_months.append(int(m_str))
            except ValueError:
                pass

        # Existing payments for this student in current academic year
        existing_paid_months = list(StudentMonthlyPayment.objects.filter(
            student=student,
            academic_year=active_year
        ).values_list('month', flat=True))

        # Any month that was previously recorded as paid but is now UNTICKED by Admin/Teacher must be removed/deleted
        months_to_remove = [m for m in existing_paid_months if m not in target_months]
        if months_to_remove:
            StudentMonthlyPayment.objects.filter(
                student=student,
                academic_year=active_year,
                month__in=months_to_remove
            ).delete()

        rates = MonthlyFeeRate.objects.filter(config=config)
        rate_map = {(r.category_id, r.month): r.amount for r in rates}

        default_cat = StudentCategory.objects.filter(is_active=True).order_by('display_order', 'id').first()
        st_cat = student.category or default_cat

        recorded_count = 0
        receipt_nos = []

        with transaction.atomic():
            for m in target_months:
                cat_id = st_cat.id if st_cat else None
                expected = Decimal('20000.00')
                if cat_id:
                    expected = rate_map.get((cat_id, m))
                    if expected is None:
                        if st_cat and ('FREE' in st_cat.code or '100' in st_cat.name):
                            expected = Decimal('0.00')
                        elif st_cat and ('SCHOLAR' in st_cat.code or '50' in st_cat.name or 'TEACHER' in st_cat.code):
                            expected = Decimal('10000.00')
                        else:
                            expected = Decimal('20000.00')

                paid = Decimal(str(custom_amount)) if custom_amount and len(target_months) == 1 else expected

                from django.utils import timezone
                now_dt = timezone.now()

                payment_rec, created = StudentMonthlyPayment.objects.get_or_create(
                    student=student,
                    academic_year=active_year,
                    month=m,
                    defaults={
                        'expected_amount': expected,
                        'paid_amount': paid,
                        'payment_method': payment_method,
                        'collected_by': request.user,
                        'payment_date': now_dt,
                        'is_on_time': is_on_time,
                        'notes': notes,
                    }
                )
                if not created:
                    payment_rec.expected_amount = expected
                    payment_rec.paid_amount = paid
                    payment_rec.payment_method = payment_method
                    payment_rec.collected_by = request.user
                    payment_rec.is_on_time = is_on_time
                    payment_rec.notes = notes
                    payment_rec.save()

                recorded_count += 1
                if payment_rec.receipt_no:
                    receipt_nos.append(payment_rec.receipt_no)

        msg = f'✅ បានរក្សាទុកការបង់ប្រាក់ចំនួន {recorded_count} ខែ សម្រាប់សិស្ស {student.khmer_name} ជោគជ័យ!'
        if months_to_remove:
            msg += f' (បានដោះ Tick ចំនួន {len(months_to_remove)} ខែត្រឡប់ជាជំពាក់)'

        return JsonResponse({
            'status': 'success',
            'message': msg,
            'receipt_no': receipt_nos[0] if receipt_nos else ''
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)



@login_required
def update_student_fee_start_month(request):
    """
    AJAX handler for Admin to update a student's fee_start_month
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        student_id = request.POST.get('student_id')
        fee_start_month = request.POST.get('fee_start_month')

        if not student_id:
            return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសសិស្ស!'}, status=400)

        student = get_object_or_404(Student, pk=student_id)
        if fee_start_month and str(fee_start_month).isdigit():
            student.fee_start_month = int(fee_start_month)
        else:
            student.fee_start_month = None
        student.save()

        start_name = MONTH_NAMES_KM.get(student.fee_start_month, 'ដើមឆ្នាំសិក្សា') if student.fee_start_month else 'ដើមឆ្នាំសិក្សា'
        return JsonResponse({
            'status': 'success',
            'message': f'🎉 បានកំណត់ខែចាប់ផ្តើមបង់ប្រាក់សម្រាប់សិស្ស {student.khmer_name} ត្រឹមខែ «{start_name}» ជោគជ័យ!'
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def bulk_assign_student_category(request):
    """
    Admin tool to bulk assign category to students by class, grade level, or individual selection
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        from apps.students.models import StudentCategory
        import json

        category_id = request.POST.get('category_id')
        assign_type = request.POST.get('assign_type') # 'CLASS', 'GRADE', 'STUDENTS', 'INDIVIDUAL'
        classroom_id = request.POST.get('classroom_id')
        grade_level_id = request.POST.get('grade_level_id')
        student_ids = request.POST.getlist('student_ids[]') or request.POST.getlist('student_ids')
        fee_start_month = request.POST.get('fee_start_month')
        fee_end_month = request.POST.get('fee_end_month')
        student_status = request.POST.get('student_status')
        effective_from_month = request.POST.get('effective_from_month')

        if not category_id and request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                category_id = data.get('category_id')
                assign_type = data.get('assign_type')
                classroom_id = data.get('classroom_id')
                grade_level_id = data.get('grade_level_id')
                student_ids = data.get('student_ids', [])
                fee_start_month = data.get('fee_start_month')
                fee_end_month = data.get('fee_end_month')
                student_status = data.get('student_status')
                effective_from_month = data.get('effective_from_month')
            except Exception:
                pass

        if category_id is None or str(category_id).strip() == '':
            return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសប្រភេទសិស្ស!'}, status=400)

        category = None
        cat_name = "សិស្សទូទៅ (Normal)"
        if str(category_id).upper() not in ['NONE', '0', '']:
            category = get_object_or_404(StudentCategory, pk=category_id)
            cat_name = category.name

        start_month_val = int(fee_start_month) if fee_start_month and str(fee_start_month).isdigit() else None
        end_month_val = int(fee_end_month) if fee_end_month and str(fee_end_month).isdigit() else None

        active_year = AcademicYear.objects.filter(is_current=True).first()
        from apps.finance.models import MonthlyFeeConfig, StudentMonthlyCategory
        config = MonthlyFeeConfig.get_or_create_for_year(active_year) if active_year else None
        month_seq = config.get_month_sequence() if config else []

        updated_count = 0

        with transaction.atomic():
            if assign_type == 'CLASS' and classroom_id:
                updated_count = Student.objects.filter(classroom_id=classroom_id, status='ACTIVE').update(category=category)
            elif assign_type == 'GRADE' and grade_level_id:
                updated_count = Student.objects.filter(classroom__grade_level=grade_level_id, status='ACTIVE').update(category=category)
            elif assign_type in ['STUDENTS', 'INDIVIDUAL', 'SINGLE'] and student_ids:
                update_fields = {'category': category}
                if fee_start_month is not None:
                    update_fields['fee_start_month'] = start_month_val
                if fee_end_month is not None:
                    update_fields['fee_end_month'] = end_month_val
                if student_status and student_status in ['ACTIVE', 'SUSPENDED', 'DROPPED', 'TRANSFERRED', 'GRADUATED']:
                    update_fields['status'] = student_status

                target_students = list(Student.objects.filter(id__in=student_ids))
                
                # Check if mid-year transition is requested
                if effective_from_month and str(effective_from_month).isdigit() and active_year and month_seq:
                    eff_m = int(effective_from_month)
                    if eff_m in month_seq:
                        eff_idx = month_seq.index(eff_m)
                        for target_st in target_students:
                            old_cat = target_st.category
                            for idx, m in enumerate(month_seq):
                                if idx < eff_idx:
                                    # Months prior to effective date retain the old category
                                    StudentMonthlyCategory.objects.update_or_create(
                                        student=target_st,
                                        academic_year=active_year,
                                        month=m,
                                        defaults={'category': old_cat}
                                    )
                                else:
                                    # Months from effective date onward switch to the new category
                                    StudentMonthlyCategory.objects.update_or_create(
                                        student=target_st,
                                        academic_year=active_year,
                                        month=m,
                                        defaults={'category': category}
                                    )
                elif active_year:
                    # Full year assignment -> remove per-month transition overrides
                    StudentMonthlyCategory.objects.filter(student_id__in=student_ids, academic_year=active_year).delete()

                updated_count = Student.objects.filter(id__in=student_ids).update(**update_fields)
            else:
                return JsonResponse({'status': 'error', 'message': 'សូមបញ្ជាក់ទិន្នន័យគោលដៅ (ថ្នាក់ ឬកម្រិត ឬសិស្ស)!'}, status=400)

        return JsonResponse({
            'status': 'success',
            'message': f'🎉 បានកែប្រែទិន្នន័យ និងប្រភេទ "{cat_name}" ជូនសិស្សចំនួន {updated_count} នាក់ជោគជ័យ!',
            'count': updated_count
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def get_classroom_students_ajax(request):
    """
    AJAX helper to get active students of a classroom for category assignment
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    from django.http import JsonResponse
    classroom_id = request.GET.get('classroom_id')
    if not classroom_id:
        return JsonResponse({'status': 'success', 'students': []})

    students = Student.objects.filter(classroom_id=classroom_id, status='ACTIVE').select_related('category').order_by('khmer_name')
    data = []
    for s in students:
        data.append({
            'id': s.id,
            'student_id': s.student_id,
            'khmer_name': s.khmer_name,
            'latin_name': s.latin_name or '',
            'gender': s.gender,
            'category_id': s.category_id,
            'category_name': s.category.name if s.category else 'ទូទៅ',
        })
    return JsonResponse({'status': 'success', 'students': data})


@login_required
def configure_monthly_rates_matrix(request):
    """
    Admin tool to set monthly fee rates per student category and per month
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        import json
        from apps.students.models import StudentCategory

        active_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-id').first()
        config = MonthlyFeeConfig.get_or_create_for_year(active_year)

        rates_data = request.POST.get('rates_json')
        if not rates_data and request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                rates_data = data.get('rates', [])
            except Exception:
                rates_data = []
        elif rates_data:
            try:
                rates_data = json.loads(rates_data)
            except Exception:
                rates_data = []

        if not isinstance(rates_data, list):
            rates_data = []

        updated_count = 0
        with transaction.atomic():
            for item in rates_data:
                cat_id = item.get('category_id')
                month = item.get('month')
                amount = item.get('amount')
                if cat_id and month and amount is not None:
                    MonthlyFeeRate.objects.update_or_create(
                        config=config,
                        category_id=cat_id,
                        month=int(month),
                        defaults={'amount': Decimal(str(amount))}
                    )
                    updated_count += 1

        return JsonResponse({
            'status': 'success',
            'message': f'🎉 បានរក្សាទុកការកំណត់តម្លៃប្រចាំខែចំនួន {updated_count} ចំណុចជោគជ័យ!'
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def manage_fee_collectors(request):
    """
    Admin tool to authorize/de-authorize teachers as fee collectors
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        import json

        raw_ids = request.POST.getlist('teacher_ids[]') or request.POST.getlist('teacher_ids')
        if not raw_ids and request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                raw_ids = data.get('teacher_ids', [])
            except Exception:
                raw_ids = []

        collector_ids = [int(i) for i in raw_ids if i]

        with transaction.atomic():
            Teacher.objects.all().update(is_fee_collector=False)
            if collector_ids:
                Teacher.objects.filter(id__in=collector_ids).update(is_fee_collector=True)

        return JsonResponse({
            'status': 'success',
            'message': f'🎉 បានធ្វើបច្ចុប្បន្នភាពសិទ្ធិគ្រូប្រមូលថវិកាចំនួន {len(collector_ids)} នាក់ជោគជ័យ!'
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def manage_student_categories(request):
    """
    CRUD management for Student Categories (Create / Edit / List)
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    from apps.students.models import StudentCategory
    from django.http import JsonResponse

    if request.method == 'POST':
        cat_id = request.POST.get('category_id')
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        description = request.POST.get('description', '').strip()
        display_order = int(request.POST.get('display_order', 1))

        if not name:
            return JsonResponse({'status': 'error', 'message': 'សូមបញ្ចូលឈ្មោះប្រភេទសិស្ស!'}, status=400)

        if cat_id:
            cat = get_object_or_404(StudentCategory, pk=cat_id)
            cat.name = name
            if code:
                cat.code = code
            cat.description = description
            cat.display_order = display_order
            cat.save()
            msg = f"បានកែប្រែប្រភេទសិស្ស {cat.name} ជោគជ័យ!"
        else:
            cat = StudentCategory.objects.create(
                name=name,
                code=code,
                description=description,
                display_order=display_order
            )
            msg = f"បានបង្កើតប្រភេទសិស្សថ្មី {cat.name} ជោគជ័យ!"

        return JsonResponse({'status': 'success', 'message': msg, 'category': {'id': cat.id, 'name': cat.name, 'code': cat.code, 'display_order': cat.display_order}})

    categories = list(StudentCategory.objects.all().values('id', 'name', 'code', 'description', 'display_order', 'is_active'))
    return JsonResponse({'status': 'success', 'categories': categories})


@login_required
def delete_student_category(request, pk):
    """
    Admin tool to delete a Student Category
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        from apps.students.models import StudentCategory
        category = get_object_or_404(StudentCategory, pk=pk)
        name = category.name
        Student.objects.filter(category=category).update(category=None)
        MonthlyFeeRate.objects.filter(category=category).delete()
        category.delete()
        return JsonResponse({'status': 'success', 'message': f'🗑️ បានលុបប្រភេទសិស្ស "{name}" ជោគជ័យ!'})

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def send_monthly_fee_reminder(request):
    """
    Send payment reminder via Telegram to student's guardian
    """
    if not check_fee_collector_access(request.user):
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        import json
        student_id = request.POST.get('student_id')
        custom_message = request.POST.get('custom_message', '').strip()
        custom_chat_id = request.POST.get('custom_chat_id', '').strip() or None

        if not student_id and request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                student_id = data.get('student_id')
                custom_message = data.get('custom_message', '').strip()
                custom_chat_id = data.get('custom_chat_id', '').strip() or None
            except Exception:
                pass

        if not student_id:
            return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសសិស្ស!'}, status=400)

        student = get_object_or_404(Student, pk=student_id)
        
        # Calculate student's due details
        active_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-id').first()
        config = MonthlyFeeConfig.get_or_create_for_year(active_year)
        ticked_months = set(config.ticked_months or [])
        rates = MonthlyFeeRate.objects.filter(config=config)
        rate_map = {(r.category_id, r.month): r.amount for r in rates}
        payments = {p.month: p.paid_amount for p in StudentMonthlyPayment.objects.filter(student=student, academic_year=active_year)}

        cat_id = student.category_id
        unpaid_months = []
        total_due = Decimal('0.00')

        for m in config.get_month_sequence():
            if m in ticked_months:
                expected = rate_map.get((cat_id, m), Decimal('20000.00')) if cat_id else Decimal('20000.00')
                paid = payments.get(m, Decimal('0.00'))
                if paid < expected:
                    unpaid_months.append(MONTH_NAMES_KM.get(m, f'ខែ {m}'))
                    total_due += (expected - paid)

        if not custom_message:
            months_str = ', '.join(unpaid_months) if unpaid_months else 'គ្រប់ខែ'
            custom_message = (
                f"សួស្តីលោក/លោកស្រីអាណាព្យាបាលសិស្ស {student.khmer_name}!\n"
                f"🏫 ថ្នាក់រៀន៖ {student.classroom.name if student.classroom else '-'}\n"
                f"🏷️ ប្រភេទសិស្ស៖ {student.category.name if student.category else 'ទូទៅ'}\n"
                f"📅 ខែដែលត្រូវបង់៖ {months_str}\n"
                f"💰 ចំនួនទឹកប្រាក់នៅសល់សរុប៖ {total_due:,.0f} {config.currency_symbol}\n\n"
                f"សូមមេត្តាទូទាត់តាមរយៈ KHQR ឬមកកាន់ការិយាល័យរដ្ឋបាលសាលា។ សូមអរគុណ!"
            )

        recipient_phone = student.father_phone or student.mother_phone or student.phone
        recipient_name = student.father_name or student.mother_name or student.guardian_name or student.khmer_name
        chat_id_to_use = custom_chat_id or student.telegram_chat_id

        log = send_telegram_notification(
            title=f"🔔 ជូនដំណឹងបង់ថ្លៃទឹកភ្លើង: {student.khmer_name}",
            message=custom_message,
            recipient_name=recipient_name,
            recipient_phone=recipient_phone,
            custom_chat_id=chat_id_to_use
        )

        return JsonResponse({
            'status': 'success',
            'message': f'✅ បានផ្ញើសាររំលឹកបង់ប្រាក់ទៅកាន់អាណាព្យាបាលសិស្ស {student.khmer_name} ({recipient_phone or "Telegram"}) ជោគជ័យ!',
            'log_id': log.id if log else None
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
def send_classroom_fee_summary_telegram(request):
    """
    Send full classroom monthly fees summary report to Homeroom Teacher or Custom Telegram Group
    """
    if not check_fee_collector_access(request.user):
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    if request.method == 'POST':
        from django.http import JsonResponse
        import json
        classroom_id = request.POST.get('classroom_id')
        custom_chat_id = request.POST.get('custom_chat_id', '').strip() or None
        custom_message = request.POST.get('custom_message', '').strip()

        if not classroom_id and request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
                classroom_id = data.get('classroom_id')
                custom_chat_id = data.get('custom_chat_id', '').strip() or None
                custom_message = data.get('custom_message', '').strip()
            except Exception:
                pass

        if not classroom_id:
            return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសថ្នាក់រៀន!'}, status=400)

        classroom = get_object_or_404(Classroom, pk=classroom_id)
        
        request.GET = request.GET.copy()
        request.GET['classroom'] = str(classroom_id)
        request.GET['status'] = 'ALL'
        data = get_monthly_fees_data(request)
        config = data['config']
        currency = config.currency_symbol if config else '៛'

        due_students = [item for item in data['students_data'] if item['remaining_balance'] > 0]
        
        if not custom_message:
            due_list_lines = []
            for i, it in enumerate(due_students[:15], 1):
                due_list_lines.append(f"{i}. {it['student'].khmer_name}: នៅជំពាក់ {it['remaining_balance']:,.0f} {currency} ({it['unpaid_months_str']})")
            if len(due_students) > 15:
                due_list_lines.append(f"... និងសិស្ស {len(due_students) - 15} នាក់ផ្សេងទៀត")

            due_section = "\n".join(due_list_lines) if due_list_lines else "🎉 សិស្សទាំងអស់បានបង់រួចរាល់គ្រប់ចំនួន!"

            custom_message = (
                f"📋 *របាយការណ៍ប្រមូលថវិកាទឹកភ្លើង៖ {classroom.name}*\n"
                f"👨‍🏫 គ្រូបន្ទុកថ្នាក់៖ {classroom.homeroom_teacher.khmer_name if classroom.homeroom_teacher else 'មិនទាន់ចាត់តាំង'}\n"
                f"👥 សិស្សសរុប៖ {data['students_count']} នាក់ (បង់គ្រប់៖ {data['kpi_fully_paid_count']}, នៅជំពាក់៖ {data['kpi_due_count']})\n"
                f"💰 ប្រាក់ត្រូវប្រមូល៖ {data['kpi_total_expected']:,.0f} {currency}\n"
                f"🟢 ប្រមូលបាន៖ {data['kpi_total_collected']:,.0f} {currency}\n"
                f"🔴 នៅជំពាក់សរុប៖ {data['kpi_total_remaining']:,.0f} {currency}\n\n"
                f"📌 *បញ្ជីសិស្សនៅជំពាក់៖*\n{due_section}"
            )

        homeroom = classroom.homeroom_teacher
        target_chat_id = custom_chat_id
        recipient_name = f"ថ្នាក់ {classroom.name}"
        recipient_phone = homeroom.phone if homeroom else None

        log = send_telegram_notification(
            title=f"📋 របាយការណ៍ថ្នាក់ {classroom.name}",
            message=custom_message,
            recipient_name=recipient_name,
            recipient_phone=recipient_phone,
            custom_chat_id=target_chat_id
        )

        return JsonResponse({
            'status': 'success',
            'message': f'✅ បានផ្ញើរបាយការណ៍ថ្នាក់ {classroom.name} ទៅកាន់ Telegram ជោគជ័យ!',
            'log_id': log.id if log else None
        })

    from django.http import JsonResponse
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def expense_list(request):
    expenses = Expense.objects.select_related('recorded_by').all()
    category_filter = request.GET.get('category', '')
    if category_filter:
        expenses = expenses.filter(category=category_filter)

    total_expense = expenses.aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            exp = form.save(commit=False)
            exp.recorded_by = request.user
            exp.save()
            messages.success(request, f"បានកត់ត្រាចំណាយ {exp.title} (${exp.amount}) ជោគជ័យ!")
            return redirect('expense_list')
    else:
        form = ExpenseForm(initial={'date': datetime.now().date()})

    return render(request, 'finance/expense_list.html', {
        'expenses': expenses,
        'total_expense': total_expense,
        'form': form,
        'categories': Expense.Category.choices,
        'selected_category': category_filter,
    })


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def payroll_list(request):
    now = datetime.now()
    selected_month = int(request.GET.get('month', now.month))
    selected_year = int(request.GET.get('year', now.year))

    payrolls = Payroll.objects.filter(month=selected_month, year=selected_year).select_related('teacher')
    
    total_payroll = sum(p.net_salary for p in payrolls)

    return render(request, 'finance/payroll_list.html', {
        'payrolls': payrolls,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'total_payroll': total_payroll,
        'months': range(1, 13),
        'years': [now.year - 1, now.year, now.year + 1],
    })


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def payroll_generate(request):
    now = datetime.now()
    selected_month = int(request.GET.get('month', now.month))
    selected_year = int(request.GET.get('year', now.year))

    teachers = Teacher.objects.filter(status='ACTIVE')
    count = 0

    with transaction.atomic():
        for teacher in teachers:
            payroll, created = Payroll.objects.get_or_create(
                teacher=teacher,
                month=selected_month,
                year=selected_year,
                defaults={'base_salary': teacher.base_salary}
            )
            payroll.calculate()
            payroll.save()
            count += 1

    messages.success(request, f"🎉 បានគណនា និងបង្កើតតារាងប្រាក់ខែគ្រូចំនួន {count} នាក់សម្រាប់ខែ {selected_month:02d}/{selected_year} ដោយស្វ័យប្រវត្តិ (កាត់កងតាមថ្ងៃអវត្តមានរួចរាល់)!")
    return redirect(f"/finance/payroll/?month={selected_month}&year={selected_year}")


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def payroll_mark_paid(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    payroll.status = Payroll.Status.PAID
    payroll.payment_date = datetime.now().date()
    payroll.save()
    messages.success(request, f"បានកត់ត្រាបើកប្រាក់ខែជូនគ្រូ {payroll.teacher.khmer_name} (${payroll.net_salary}) ជោគជ័យ!")
    return redirect(f"/finance/payroll/?month={payroll.month}&year={payroll.year}")


# ==============================================================================
# TRI-CHANNEL MOBILE FEE COLLECTION SYSTEM (ប្រព័ន្ធប្រមូលថវិកា ៣ ឆានែល លើទូរស័ព្ទ)
# ==============================================================================

def mobile_fee_collector_portal(request):
    """
    Channel 1 & 2: Dedicated Mobile Web App for Teachers & Collectors
    - Supports Magic QR Pass token authentication (?token=<collector_token>)
    - Touch-friendly cards, live camera QR scanner, 1-tap payment, Bakong KHQR
    """
    token = request.GET.get('token')
    if token:
        from apps.teachers.models import Teacher
        teacher = Teacher.objects.filter(collector_token=token, status='ACTIVE').first()
        if teacher and teacher.user:
            from django.contrib.auth import login
            login(request, teacher.user, backend='django.contrib.auth.backends.ModelBackend')
        elif not request.user.is_authenticated:
            messages.error(request, "❌ Token ឬ QR Pass ចូលប្រើមិនត្រឹមត្រូវ ឬផុតកំណត់!")
            return redirect('login')

    if not request.user.is_authenticated:
        return redirect(f"/accounts/login/?next={request.path}")

    if not check_fee_collector_access(request.user):
        messages.error(request, "⛔ លោកអ្នកគ្មានសិទ្ធិចូលមើលផ្ទាំងប្រមូលថវិកានេះទេ!")
        return redirect('teacher_dashboard' if request.user.role == 'TEACHER' else 'login')

    context = get_monthly_fees_data(request)
    context['is_admin'] = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'
    
    teacher_obj = getattr(request.user, 'teacher_profile', None)
    context['teacher_obj'] = teacher_obj

    # Calculate amount collected by this user today
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_collected = StudentMonthlyPayment.objects.filter(
        collected_by=request.user,
        updated_at__gte=today_start,
        paid_amount__gt=0
    ).aggregate(models.Sum('paid_amount'))['paid_amount__sum'] or Decimal('0.00')
    context['today_collected'] = today_collected

    # Filter classrooms if teacher is not admin and has assigned homeroom classes
    if teacher_obj and not context['is_admin']:
        teacher_classes = list(teacher_obj.homeroom_classes.all())
        if teacher_classes:
            context['my_classrooms'] = teacher_classes

    return render(request, 'finance/mobile_collector_portal.html', context)


@login_required
def get_teacher_collector_passes(request):
    """
    Channel 2 Admin Tool: Returns all fee collectors with their tokens and portal URLs
    """
    if request.user.role not in ['ADMIN', 'ACCOUNTANT'] and not request.user.is_superuser:
        from django.http import JsonResponse
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    from django.http import JsonResponse
    from apps.teachers.models import Teacher

    collectors = Teacher.objects.filter(is_fee_collector=True, status='ACTIVE').select_related('user')
    passes = []
    base_url = request.build_absolute_uri('/finance/monthly-fees/collector-portal/')

    for t in collectors:
        token = t.get_or_create_collector_token()
        magic_url = f"{base_url}?token={token}"
        homeroom_classes = list(t.homeroom_classes.values_list('name', flat=True))
        passes.append({
            'id': t.id,
            'teacher_id': t.teacher_id,
            'name': t.khmer_name,
            'latin_name': t.latin_name,
            'phone': t.phone,
            'classes': homeroom_classes or ['គ្រប់ថ្នាក់'],
            'token': token,
            'portal_url': magic_url,
        })

    return JsonResponse({'status': 'success', 'collectors': passes})


@login_required
def api_search_student_qr(request):
    """
    API for Live QR Scanner to find student by scanned QR code (student_id or ID)
    """
    from django.http import JsonResponse
    from apps.students.models import Student
    import json

    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'status': 'error', 'message': 'សូមស្កេន ឬបញ្ចូលកូដសិស្ស'}, status=400)

    st = Student.objects.filter(
        Q(student_id__iexact=query) | Q(student_id__icontains=query) | Q(id__iexact=query if query.isdigit() else None)
    ).select_related('classroom', 'category').first()

    if not st:
        return JsonResponse({'status': 'error', 'message': f'រកមិនឃើញសិស្សដែលមានកូដ "{query}" ឡើយ'}, status=404)

    active_year = AcademicYear.objects.filter(is_current=True).first()
    config = MonthlyFeeConfig.get_or_create_for_year(active_year) if active_year else None
    if not active_year or not config:
        return JsonResponse({'status': 'error', 'message': 'គ្មានឆ្នាំសិក្សាសកម្ម'}, status=400)

    month_seq = config.get_month_sequence()
    ticked_set = set(config.ticked_months or [])
    payments = {p.month: p for p in StudentMonthlyPayment.objects.filter(student=st, academic_year=active_year)}
    monthly_cats = {mc.month: mc.category for mc in StudentMonthlyCategory.objects.filter(student=st, academic_year=active_year)}
    rates = {(r.category_id, r.month): r.amount for r in MonthlyFeeRate.objects.filter(config=config)}

    fee_start_idx = month_seq.index(st.fee_start_month) if st.fee_start_month in month_seq else 0
    fee_end_idx = month_seq.index(st.fee_end_month) if st.fee_end_month in month_seq else len(month_seq) - 1

    months_list = []
    total_expected = Decimal('0.00')
    total_paid = Decimal('0.00')

    for idx, m in enumerate(month_seq):
        is_attending = (fee_start_idx <= idx <= fee_end_idx)
        is_ticked = (m in ticked_set) and is_attending
        m_cat = monthly_cats.get(m, st.category)
        m_cat_id = m_cat.id if m_cat else None

        expected = Decimal('0.00')
        if m_cat_id and is_attending:
            expected = rates.get((m_cat_id, m))
            if expected is None:
                if m_cat and ('FREE' in m_cat.code or '100' in m_cat.name):
                    expected = Decimal('0.00')
                elif m_cat and ('SCHOLAR' in m_cat.code or '50' in m_cat.name or 'TEACHER' in m_cat.code):
                    expected = Decimal('10000.00')
                else:
                    expected = Decimal('20000.00')
        elif is_attending:
            expected = Decimal('20000.00')

        p = payments.get(m)
        paid = p.paid_amount if p else Decimal('0.00')
        total_paid += paid

        if is_ticked:
            total_expected += expected

        months_list.append({
            'month': m,
            'month_name': MONTH_NAMES_KM.get(m, f'ខែ {m}'),
            'is_ticked': is_ticked,
            'expected': float(expected),
            'paid': float(paid),
            'is_paid': (paid >= expected and expected > 0) or (expected == 0 and is_ticked),
        })

    remaining = max(Decimal('0.00'), total_expected - total_paid)

    return JsonResponse({
        'status': 'success',
        'student': {
            'id': st.id,
            'student_id': st.student_id,
            'khmer_name': st.khmer_name,
            'latin_name': st.latin_name,
            'classroom': st.classroom.name if st.classroom else 'គ្មានថ្នាក់',
            'category': st.category.name if st.category else 'ទូទៅ',
            'phone': st.father_phone or st.mother_phone or st.phone or '',
            'telegram_chat_id': st.telegram_chat_id or '',
            'total_expected': float(total_expected),
            'total_paid': float(total_paid),
            'remaining_balance': float(remaining),
            'months': months_list,
        }
    })


# ==============================================================================
# PAYMENT LOGS, FIRESTORE CLOUD SYNC & ANY-TIME BACKUP PORTAL
# ==============================================================================

@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def payment_logs_dashboard(request):
    """
    Live Payment Audit Logs & Firebase Firestore Sync Dashboard:
    - Real-time audit logs of Telegram inquiries, QR scans, and confirmed payments.
    - Review and 1-click Approve/Reject for parent payment slips.
    - Bank Account & ABA / Bakong QR Code management.
    - 1-Click Any-Time Backup Suite (Excel, JSON, Cloud Sync, Telegram dispatch).
    """
    from apps.finance.models import (
        SchoolPaymentMethod,
        PaymentSlipSubmission,
        FirestorePaymentAuditLog,
        StudentMonthlyPayment,
        Invoice
    )
    from apps.finance.firebase_service import get_firestore_db

    # Filter params
    query = request.GET.get('q', '').strip()
    event_filter = request.GET.get('event', '').strip()
    status_filter = request.GET.get('status', '').strip()

    audit_logs = FirestorePaymentAuditLog.objects.all().select_related('student', 'student__classroom')
    if query:
        audit_logs = audit_logs.filter(
            models.Q(student_id_str__icontains=query) |
            models.Q(student_name__icontains=query) |
            models.Q(fee_category_name__icontains=query) |
            models.Q(telegram_user_info__icontains=query)
        )
    if event_filter:
        audit_logs = audit_logs.filter(event_type=event_filter)

    logs_list = audit_logs[:200]

    # Payment Slips
    slips = PaymentSlipSubmission.objects.all().select_related('student', 'student__classroom', 'reviewed_by')
    if status_filter:
        slips = slips.filter(status=status_filter)
    slips_list = slips[:50]
    pending_slips_count = PaymentSlipSubmission.objects.filter(status=PaymentSlipSubmission.Status.PENDING).count()

    # Bank Payment Methods
    payment_methods = SchoolPaymentMethod.objects.all()

    # Firestore status
    db = get_firestore_db()
    firestore_connected = (db is not None)

    # Summary KPIs
    active_year = AcademicYear.objects.filter(is_current=True).first()
    total_monthly_khr = StudentMonthlyPayment.objects.filter(paid_amount__gt=0).aggregate(models.Sum('paid_amount'))['paid_amount__sum'] or Decimal('0.00')
    total_invoices_usd = Invoice.objects.filter(paid_amount__gt=0).aggregate(models.Sum('paid_amount'))['paid_amount__sum'] or Decimal('0.00')
    total_inquiries_count = FirestorePaymentAuditLog.objects.filter(event_type=FirestorePaymentAuditLog.EventType.INQUIRY).count()
    synced_logs_count = FirestorePaymentAuditLog.objects.filter(is_synced_to_firestore=True).count()
    unsynced_logs_count = FirestorePaymentAuditLog.objects.filter(is_synced_to_firestore=False).count()

    context = {
        'active_tab': 'payment_logs',
        'logs_list': logs_list,
        'slips_list': slips_list,
        'pending_slips_count': pending_slips_count,
        'payment_methods': payment_methods,
        'firestore_connected': firestore_connected,
        'total_monthly_khr': total_monthly_khr,
        'total_invoices_usd': total_invoices_usd,
        'total_inquiries_count': total_inquiries_count,
        'synced_logs_count': synced_logs_count,
        'unsynced_logs_count': unsynced_logs_count,
        'query': query,
        'event_filter': event_filter,
        'status_filter': status_filter,
        'active_year': active_year,
    }
    return render(request, 'finance/payment_logs_dashboard.html', context)


@login_required
@role_required(['ADMIN'])
def api_manage_payment_methods(request):
    """
    Handles creating, updating, toggling, or deleting School Bank / ABA QR payment methods.
    """
    from apps.finance.models import SchoolPaymentMethod

    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        method_id = request.POST.get('method_id')

        if action == 'delete' and method_id:
            SchoolPaymentMethod.objects.filter(id=method_id).delete()
            messages.success(request, "បានលុបគណនីធនាគារជោគជ័យ!")
            return redirect('payment_logs_dashboard')

        if action == 'set_default' and method_id:
            SchoolPaymentMethod.objects.exclude(id=method_id).update(is_default=False)
            SchoolPaymentMethod.objects.filter(id=method_id).update(is_default=True)
            messages.success(request, "បានកំណត់ជាគណនីលំនាំដើម (Default) ជោគជ័យ!")
            return redirect('payment_logs_dashboard')

        bank_name = request.POST.get('bank_name', 'ABA Bank').strip()
        account_name = request.POST.get('account_name', '').strip()
        account_number = request.POST.get('account_number', '').strip()
        currency = request.POST.get('currency', 'KHR')
        instructions = request.POST.get('instructions', '').strip()
        is_default = request.POST.get('is_default') == 'on'
        is_active = request.POST.get('is_active') != 'off'

        if method_id:
            pm = get_object_or_404(SchoolPaymentMethod, id=method_id)
            pm.bank_name = bank_name
            pm.account_name = account_name
            pm.account_number = account_number
            pm.currency = currency
            pm.instructions = instructions
            pm.is_default = is_default
            pm.is_active = is_active
            if 'qr_image' in request.FILES:
                pm.qr_image = request.FILES['qr_image']
            pm.save()
            messages.success(request, f"បានកែប្រែគណនី {pm.bank_name} ជោគជ័យ!")
        else:
            pm = SchoolPaymentMethod.objects.create(
                bank_name=bank_name,
                account_name=account_name,
                account_number=account_number,
                currency=currency,
                instructions=instructions,
                is_default=is_default,
                is_active=is_active,
                qr_image=request.FILES.get('qr_image')
            )
            messages.success(request, f"បានបង្កើតគណនីធនាគារ {pm.bank_name} ជោគជ័យ!")

    return redirect('payment_logs_dashboard')


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def api_review_payment_slip(request, pk):
    """
    Approves or rejects a parent payment slip from the web dashboard.
    """
    from apps.finance.models import (
        PaymentSlipSubmission,
        MonthlyFeeConfig,
        MonthlyFeeRate,
        StudentMonthlyPayment,
        StudentMonthlyCategory
    )
    from apps.finance.firebase_service import log_payment_slip_to_firestore, log_payment_transaction_to_firestore
    from apps.accounts.utils import send_telegram_notification

    slip = get_object_or_404(PaymentSlipSubmission, pk=pk)
    action = request.POST.get('action', 'approve')
    notes = request.POST.get('notes', '').strip()

    student = slip.student
    active_year = slip.academic_year or AcademicYear.objects.filter(is_current=True).first()

    if action == 'approve':
        config = MonthlyFeeConfig.get_or_create_for_year(active_year)
        month_seq = config.get_month_sequence() if config else []
        ticked_set = set(config.ticked_months or []) if config else set()
        rates = {(r.category_id, r.month): r.amount for r in MonthlyFeeRate.objects.filter(config=config)} if config else {}
        monthly_cats = {mc.month: mc.category for mc in StudentMonthlyCategory.objects.filter(student=student, academic_year=active_year)}

        fee_start_idx = month_seq.index(student.fee_start_month) if student.fee_start_month in month_seq else 0
        fee_end_idx = month_seq.index(student.fee_end_month) if student.fee_end_month in month_seq else len(month_seq) - 1

        approved_months = []
        total_paid_rec = Decimal('0.00')

        with transaction.atomic():
            for idx, m in enumerate(month_seq):
                is_attending = (fee_start_idx <= idx <= fee_end_idx)
                if not ((m in ticked_set) and is_attending):
                    continue

                m_cat = monthly_cats.get(m, student.category)
                m_cat_id = m_cat.id if m_cat else None
                expected = rates.get((m_cat_id, m), Decimal('20000.00')) if m_cat_id else Decimal('20000.00')

                p, _ = StudentMonthlyPayment.objects.get_or_create(
                    student=student,
                    academic_year=active_year,
                    month=m,
                    defaults={'expected_amount': expected, 'paid_amount': Decimal('0.00')}
                )
                if p.paid_amount < expected:
                    p.expected_amount = expected
                    p.paid_amount = expected
                    p.status = StudentMonthlyPayment.Status.PAID
                    p.payment_date = timezone.now()
                    p.payment_method = StudentMonthlyPayment.PaymentMethod.ABA_BANK
                    p.collected_by = request.user
                    p.notes = f"ផ្ទៀងផ្ទាត់ និងអនុម័តបង្កាន់ដៃ #SLIP-{slip.id} ដោយ {request.user.get_full_name() or request.user.username}"
                    p.save()
                    approved_months.append(MONTH_NAMES_KM.get(m, f"ខែ {m}"))
                    total_paid_rec += expected
                    log_payment_transaction_to_firestore(p, user_disp=f"Web Admin: {request.user.username}")

            slip.status = PaymentSlipSubmission.Status.APPROVED
            slip.claimed_amount = total_paid_rec
            slip.reviewed_by = request.user
            slip.reviewed_at = timezone.now()
            slip.notes = notes or f"បានអនុម័តដោយ {request.user.get_full_name() or request.user.username}"
            slip.save()
            log_payment_slip_to_firestore(slip)

        if slip.telegram_chat_id:
            parent_msg = (
                f"🎉 *បង្កាន់ដៃបង់ប្រាក់ត្រូវបានផ្ទៀងផ្ទាត់ និងយល់ព្រមរួចរាល់!*\n\n"
                f"• សិស្ស៖ *{student.khmer_name}* ({student.student_id})\n"
                f"• ខែដែលបានបង់៖ {', '.join(approved_months) if approved_months else 'គ្រប់ខែ'}\n"
                f"• ចំនួនទឹកប្រាក់៖ *{total_paid_rec:,.0f} ៛*\n"
                f"• កាលបរិច្ឆេទ៖ {timezone.now().strftime('%d/%m/%Y %H:%M')}\n\n"
                f"🙏 *សូមអរគុណចំពោះការបង់ប្រាក់ទាន់ពេលវេលា!*"
            )
            send_telegram_notification(
                title="✅ បញ្ជាក់ការបង់ប្រាក់ជោគជ័យ",
                message=parent_msg,
                custom_chat_id=slip.telegram_chat_id
            )

        messages.success(request, f"បានយល់ព្រមបង្កាន់ដៃ #{slip.id} (ចំនួន {total_paid_rec:,.0f} ៛) ជោគជ័យ!")
    else:
        slip.status = PaymentSlipSubmission.Status.REJECTED
        slip.reviewed_by = request.user
        slip.reviewed_at = timezone.now()
        slip.notes = notes or f"បានបដិសេធដោយ {request.user.get_full_name() or request.user.username}"
        slip.save()
        log_payment_slip_to_firestore(slip)

        if slip.telegram_chat_id:
            send_telegram_notification(
                title="❌ បង្កាន់ដៃបង់ប្រាក់មិនត្រឹមត្រូវ",
                message=(
                    f"⚠️ បង្កាន់ដៃបង់ប្រាក់របស់សិស្ស *{student.khmer_name}* ({student.student_id}) មិនទាន់ត្រឹមត្រូវឡើយ。\n"
                    f"មូលហេតុ៖ {notes or 'សូមពិនិត្យចំនួនទឹកប្រាក់ ឬផ្ញើរូបភាពច្បាស់ឡើងវិញ'}"
                ),
                custom_chat_id=slip.telegram_chat_id
            )

        messages.warning(request, f"បានបដិសេធបង្កាន់ដៃ #{slip.id}!")

    return redirect('payment_logs_dashboard')


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def export_payment_logs_excel(request):
    """
    Exports full financial payment ledger, monthly utilities, and Firestore logs to Excel (.xlsx).
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from apps.finance.models import StudentMonthlyPayment, Invoice, PaymentSlipSubmission, FirestorePaymentAuditLog

    wb = openpyxl.Workbook()
    
    # Sheet 1: Monthly Utility Payments
    ws1 = wb.active
    ws1.title = "ថ្លៃទឹកភ្លើងប្រចាំខែ"

    # Header style
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers1 = [
        "ល.រ", "អត្តលេខ", "ឈ្មោះសិស្ស", "ភេទ", "ថ្នាក់", "ខែ", "ចំនួនត្រូវបង់ (៛)",
        "ចំនួនបានបង់ (៛)", "ស្ថានភាព", "វិធីសាស្ត្រ", "លេខបង្កាន់ដៃ", "ថ្ងៃបង់ប្រាក់", "អ្នកប្រមូល"
    ]
    ws1.append(headers1)
    for col_idx, col in enumerate(ws1.iter_cols(min_row=1, max_row=1), 1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

    monthly_payments = StudentMonthlyPayment.objects.all().select_related('student', 'student__classroom', 'collected_by').order_by('student__classroom__name', 'student__student_id', 'month')
    for idx, p in enumerate(monthly_payments, 1):
        st = p.student
        m_name = MONTH_NAMES_KM.get(p.month, f"ខែ {p.month}")
        ws1.append([
            idx,
            st.student_id,
            st.khmer_name,
            st.get_gender_display(),
            st.classroom.name if st.classroom else '-',
            m_name,
            float(p.expected_amount),
            float(p.paid_amount),
            p.get_status_display(),
            p.get_payment_method_display(),
            p.receipt_no or '-',
            p.payment_date.strftime('%Y-%m-%d %H:%M') if p.payment_date else '-',
            str(p.collected_by.username) if p.collected_by else 'Admin'
        ])

    # Sheet 2: Early Year & Standard Invoices
    ws2 = wb.create_sheet(title="ថវិកាដើមឆ្នាំ & វិក្កយបត្រ")
    headers2 = [
        "ល.រ", "លេខវិក្កយបត្រ", "អត្តលេខ", "ឈ្មោះសិស្ស", "ថ្នាក់", "ប្រភេទកម្រៃ",
        "តម្លៃដើម ($)", "បញ្ចុះតម្លៃ (%)", "តម្លៃសុទ្ធ ($)", "បានបង់ ($)", "នៅជំពាក់ ($)", "ស្ថានភាព", "កាលបរិច្ឆេទ"
    ]
    ws2.append(headers2)
    for col_idx, col in enumerate(ws2.iter_cols(min_row=1, max_row=1), 1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

    invoices = Invoice.objects.all().select_related('student', 'student__classroom', 'fee_category').order_by('-created_at')
    for idx, inv in enumerate(invoices, 1):
        st = inv.student
        ws2.append([
            idx,
            inv.invoice_no,
            st.student_id,
            st.khmer_name,
            st.classroom.name if st.classroom else '-',
            inv.fee_category.name,
            float(inv.original_amount),
            float(inv.discount_percent),
            float(inv.final_amount),
            float(inv.paid_amount),
            float(inv.remaining_balance),
            inv.get_status_display(),
            inv.created_at.strftime('%Y-%m-%d')
        ])

    # Sheet 3: Telegram & Firestore Audit Logs
    ws3 = wb.create_sheet(title="កំណត់ត្រា Firestore & Telegram")
    headers3 = [
        "ល.រ", "ប្រភេទព្រឹត្តិការណ៍", "អត្តលេខ", "ឈ្មោះសិស្ស", "ថ្នាក់", "ចំនួនទឹកប្រាក់", "រូបិយប័ណ្ណ",
        "កម្រៃ/ចំណងជើង", "ប៉ុស្តិ៍ទាក់ទង", "ព័ត៌មាន Telegram", "Firestore Synced", "កាលបរិច្ឆេទ"
    ]
    ws3.append(headers3)
    for col_idx, col in enumerate(ws3.iter_cols(min_row=1, max_row=1), 1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

    logs = FirestorePaymentAuditLog.objects.all().order_by('-created_at')[:500]
    for idx, l in enumerate(logs, 1):
        ws3.append([
            idx,
            l.get_event_type_display(),
            l.student_id_str or '-',
            l.student_name or '-',
            l.classroom_name or '-',
            float(l.amount),
            l.currency,
            l.fee_category_name or '-',
            l.channel,
            l.telegram_user_info or '-',
            "YES" if l.is_synced_to_firestore else "NO",
            l.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"SchoolSM_Payment_Logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def export_payment_logs_json(request):
    """
    Exports structured JSON snapshot of all payment logs and Firestore records.
    """
    from apps.finance.models import StudentMonthlyPayment, Invoice, PaymentSlipSubmission, FirestorePaymentAuditLog

    data = {
        'exported_at': timezone.now().isoformat(),
        'exported_by': request.user.username,
        'monthly_payments': [
            {
                'id': p.id,
                'student_id': p.student.student_id,
                'student_name': p.student.khmer_name,
                'month': p.month,
                'expected_amount': float(p.expected_amount),
                'paid_amount': float(p.paid_amount),
                'status': p.status,
                'payment_method': p.payment_method,
                'receipt_no': p.receipt_no,
                'payment_date': p.payment_date.isoformat() if p.payment_date else None,
            }
            for p in StudentMonthlyPayment.objects.all().select_related('student')
        ],
        'invoices': [
            {
                'invoice_no': inv.invoice_no,
                'student_id': inv.student.student_id,
                'fee_category': inv.fee_category.name,
                'final_amount': float(inv.final_amount),
                'paid_amount': float(inv.paid_amount),
                'status': inv.status,
                'created_at': inv.created_at.isoformat(),
            }
            for inv in Invoice.objects.all().select_related('student', 'fee_category')
        ],
        'payment_slips': [
            {
                'id': s.id,
                'student_id': s.student.student_id,
                'fee_type': s.fee_type,
                'claimed_amount': float(s.claimed_amount),
                'status': s.status,
                'telegram_username': s.telegram_username,
                'submitted_at': s.created_at.isoformat(),
            }
            for s in PaymentSlipSubmission.objects.all().select_related('student')
        ],
        'firestore_logs': [
            {
                'id': l.id,
                'event_type': l.event_type,
                'student_id': l.student_id_str,
                'student_name': l.student_name,
                'amount': float(l.amount),
                'currency': l.currency,
                'fee_category': l.fee_category_name,
                'firestore_doc_id': l.firestore_doc_id,
                'is_synced': l.is_synced_to_firestore,
                'created_at': l.created_at.isoformat(),
            }
            for l in FirestorePaymentAuditLog.objects.all()
        ]
    }

    response = HttpResponse(json.dumps(data, indent=2, ensure_ascii=False), content_type='application/json')
    filename = f"SchoolSM_Payment_Snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def api_sync_firestore(request):
    """
    AJAX endpoint to trigger full synchronization to Firebase Firestore.
    """
    from apps.finance.firebase_service import sync_all_local_payments_to_firestore
    res = sync_all_local_payments_to_firestore()
    if res.get('success'):
        messages.success(request, res.get('message', 'បាន Sync ទៅ Firestore ជោគជ័យ!'))
    else:
        messages.error(request, res.get('message', 'ការ Sync ទៅ Firestore បានបរាជ័យ!'))
    return redirect('payment_logs_dashboard')


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def api_send_payment_backup_telegram(request):
    """
    Generates Excel ledger snapshot and dispatches directly to Admin's Telegram Channel.
    """
    from apps.accounts.utils import send_telegram_document
    from apps.accounts.models import TelegramConfig

    tconfig = TelegramConfig.objects.first()
    if not (tconfig and tconfig.is_active and tconfig.bot_token):
        messages.error(request, "⚠️ Telegram Bot មិនទាន់ត្រូវបានកំណត់ ឬមិនទាន់សកម្មឡើយ!")
        return redirect('payment_logs_dashboard')

    # Generate Excel in memory
    excel_response = export_payment_logs_excel(request)
    excel_bytes = excel_response.content
    filename = f"SchoolSM_Payment_Audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    caption = (
        f"📊 *របាយការណ៍បង់ប្រាក់ & ទឹកភ្លើង (Financial Backup)*\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"• ស្ថាប័ន៖ សាលារៀន SM\n"
        f"• ទម្រង់ឯកសារ៖ Microsoft Excel (.xlsx)\n"
        f"• កាលបរិច្ឆេទ Backup៖ {timezone.now().strftime('%d/%m/%Y %H:%M')}\n"
        f"• បង្កើតដោយ៖ {request.user.get_full_name() or request.user.username}\n\n"
        f"🔒 _ទិន្នន័យត្រូវបានផ្ទៀងផ្ទាត់ និង Sync ជាមួយ Google Firebase Firestore រួចរាល់។_"
    )

    send_telegram_document(
        document_bytes=excel_bytes,
        filename=filename,
        caption=caption,
        recipient_name="Admin Management",
        custom_chat_id=tconfig.chat_id
    )

    messages.success(request, f"បានផ្ញើឯកសារ Backup របាយការណ៍បង់ប្រាក់ទៅកាន់ Telegram Channel ជោគជ័យ!")
    return redirect('payment_logs_dashboard')


