from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Avg
from decimal import Decimal
from datetime import datetime, date, timedelta

import csv
import json

from apps.accounts.decorators import role_required
from apps.accounts.models import User, NotificationLog
from apps.students.models import Student
from apps.teachers.models import Teacher, TeacherAttendance
from apps.academics.models import Classroom, Subject, Timetable, AcademicYear
from apps.attendance.models import StudentAttendance
from apps.examinations.models import ExamTerm, Grade
from apps.finance.models import Invoice, PaymentTransaction, Expense, Payroll
from apps.extras.models import Book, BookBorrowing, InventoryItem, Announcement

# ----------------- 1. SUPER ADMIN DASHBOARD -----------------

@login_required
@role_required(['ADMIN'])
def admin_dashboard(request):
    from apps.academics.utils import get_active_academic_year
    today = datetime.now().date()
    current_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()
    
    students_base = Student.objects.filter(status='ACTIVE')
    classrooms_base = Classroom.objects.all()
    if current_year:
        students_base = students_base.filter(Q(academic_year=current_year) | Q(classroom__academic_year=current_year))
        classrooms_base = classrooms_base.filter(academic_year=current_year)

    total_students = students_base.count()
    total_teachers = Teacher.objects.filter(status='ACTIVE').count()
    total_classes = classrooms_base.count()
    
    # Today student attendance
    today_attendances = StudentAttendance.objects.filter(date=today)
    if current_year:
        today_attendances = today_attendances.filter(classroom__academic_year=current_year)
    today_present = today_attendances.filter(status=StudentAttendance.Status.PRESENT).count()
    today_absent = today_attendances.filter(status=StudentAttendance.Status.ABSENT).count()
    today_att_total = today_attendances.count()
    today_att_rate = round((today_present / today_att_total) * 100, 1) if today_att_total > 0 else 100.0

    # Finance this month
    this_month = today.month
    this_year = today.year
    
    month_revenue = PaymentTransaction.objects.filter(
        payment_date__year=this_year,
        payment_date__month=this_month
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    month_expense = Expense.objects.filter(
        date__year=this_year,
        date__month=this_month
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    due_invoices = Invoice.objects.filter(status__in=['UNPAID', 'OVERDUE', 'PARTIAL'])
    if current_year:
        due_invoices = due_invoices.filter(academic_year=current_year)
    total_due_amount = sum(inv.remaining_balance for inv in due_invoices)

    # Class size breakdown for chart
    classrooms = classrooms_base.annotate(student_cnt=Count('students', filter=Q(students__status='ACTIVE')))
    class_labels = [c.name for c in classrooms]
    class_counts = [c.student_cnt for c in classrooms]

    # Gender ratio
    male_count = students_base.filter(gender='M').count()
    female_count = students_base.filter(gender='F').count()

    # Recent activity logs & announcements
    recent_logs = NotificationLog.objects.all()[:5]
    recent_payments = PaymentTransaction.objects.select_related('invoice__student').all()[:5]
    low_stock_items = InventoryItem.objects.filter(stock_quantity__lte=10)
    announcements = Announcement.objects.filter(is_published=True)[:4]

    # Monthly revenue & expense chart data for past 6 months
    months_labels = []
    revenue_chart_data = []
    expense_chart_data = []

    for i in range(5, -1, -1):
        d = today.replace(day=1) - timedelta(days=i*30)
        m_label = d.strftime('%b %Y')
        months_labels.append(m_label)

        rev = PaymentTransaction.objects.filter(
            payment_date__year=d.year,
            payment_date__month=d.month
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
        revenue_chart_data.append(float(rev))

        exp = Expense.objects.filter(
            date__year=d.year,
            date__month=d.month
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
        expense_chart_data.append(float(exp))

    # Pending Teacher Leave Requests for Admin Dashboard Alert
    from apps.teachers.models import TeacherLeaveRequest
    pending_leaves = TeacherLeaveRequest.objects.filter(status=TeacherLeaveRequest.Status.PENDING).select_related('teacher').order_by('-created_at')[:4]

    return render(request, 'dashboard/admin_dashboard.html', {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_classes': total_classes,
        'today_att_rate': today_att_rate,
        'today_present': today_present,
        'today_absent': today_absent,
        'month_revenue': month_revenue,
        'month_expense': month_expense,
        'total_due_amount': total_due_amount,
        'class_labels_json': json.dumps(class_labels),
        'class_counts_json': json.dumps(class_counts),
        'male_count': male_count,
        'female_count': female_count,
        'months_labels_json': json.dumps(months_labels),
        'revenue_chart_json': json.dumps(revenue_chart_data),
        'expense_chart_json': json.dumps(expense_chart_data),
        'recent_logs': recent_logs,
        'recent_payments': recent_payments,
        'low_stock_items': low_stock_items,
        'announcements': announcements,
        'pending_leaves': pending_leaves,
        'active_year': current_year,
    })



# ----------------- 2. ACCOUNTANT DASHBOARD -----------------

@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def finance_dashboard(request):
    today = datetime.now().date()
    this_month = today.month
    this_year = today.year

    month_collected = PaymentTransaction.objects.filter(
        payment_date__year=this_year,
        payment_date__month=this_month
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    month_expense = Expense.objects.filter(
        date__year=this_year,
        date__month=this_month
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')

    unpaid_invoices_count = Invoice.objects.filter(status='UNPAID').count()
    overdue_invoices_count = Invoice.objects.filter(status='OVERDUE').count()
    
    due_invoices = Invoice.objects.filter(status__in=['UNPAID', 'OVERDUE', 'PARTIAL']).select_related('student__classroom', 'fee_category')[:10]
    total_due_amount = sum(inv.remaining_balance for inv in due_invoices)

    recent_transactions = PaymentTransaction.objects.select_related('invoice__student', 'received_by').all()[:10]
    recent_expenses = Expense.objects.select_related('recorded_by').all()[:5]

    return render(request, 'dashboard/accountant_dashboard.html', {
        'month_collected': month_collected,
        'month_expense': month_expense,
        'unpaid_invoices_count': unpaid_invoices_count,
        'overdue_invoices_count': overdue_invoices_count,
        'total_due_amount': total_due_amount,
        'due_invoices': due_invoices,
        'recent_transactions': recent_transactions,
        'recent_expenses': recent_expenses,
    })


# ----------------- 3. TEACHER DASHBOARD -----------------

@login_required
@role_required(['ADMIN', 'TEACHER'])
def teacher_dashboard(request):
    from apps.academics.utils import get_active_academic_year
    current_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()
    user = request.user
    teacher = getattr(user, 'teacher_profile', None)
    
    if not teacher and user.is_superuser:
        teacher = Teacher.objects.first()

    today_weekday = datetime.now().weekday() + 1
    
    today_schedule = []
    my_classes = []
    my_students_count = 0
    my_homeroom = None

    if teacher:
        tt_qs = Timetable.objects.filter(
            teacher=teacher,
            day_of_week=today_weekday
        )
        if current_year:
            tt_qs = tt_qs.filter(classroom__academic_year=current_year)
        today_schedule = tt_qs.select_related('classroom', 'subject').order_by('start_time')

        cls_qs = Classroom.objects.filter(homeroom_teacher=teacher)
        if current_year:
            cls_qs = cls_qs.filter(academic_year=current_year)
        my_homeroom = cls_qs.first()

        if my_homeroom:
            my_students_count = my_homeroom.students.filter(status='ACTIVE').count()

        taught_class_ids = Timetable.objects.filter(
            teacher=teacher,
            classroom__academic_year=current_year
        ).order_by().values_list('classroom_id', flat=True).distinct() if current_year else Timetable.objects.filter(teacher=teacher).order_by().values_list('classroom_id', flat=True).distinct()
        
        my_classes = Classroom.objects.filter(id__in=taught_class_ids)

    today = datetime.now().date()
    today_att_done = False
    if teacher and my_homeroom:
        today_att_done = StudentAttendance.objects.filter(classroom=my_homeroom, date=today).exists()

    # Personal Teacher Attendance & Check-in status
    from apps.teachers.models import TeacherAttendance, TeacherPunchLog, TeacherAttendanceConfig
    today_teacher_att = None
    today_punch = None
    my_month_present = 0
    my_month_late = 0
    att_config = TeacherAttendanceConfig.get_settings()

    if teacher:
        today_teacher_att = TeacherAttendance.objects.filter(teacher=teacher, date=today).first()
        today_punch = TeacherPunchLog.objects.filter(teacher=teacher, date=today).order_by('-punch_time').first()
        first_day_month = date(today.year, today.month, 1)
        month_atts = TeacherAttendance.objects.filter(teacher=teacher, date__gte=first_day_month, date__lte=today)
        my_month_present = month_atts.filter(status=TeacherAttendance.Status.PRESENT).count()
        my_month_late = month_atts.filter(status=TeacherAttendance.Status.LATE).count()

    recent_terms = ExamTerm.objects.filter(academic_year=current_year)[:3] if current_year else ExamTerm.objects.all()[:3]
    announcements = Announcement.objects.filter(
        Q(target_audience__in=['ALL', 'TEACHERS']),
        is_published=True
    )[:4]

    today_duty_list = []
    if teacher:
        try:
            from apps.academics.models import TeacherDutySchedule, TeacherDutyType
            raw_duty_types = TeacherDutyType.get_all_duty_types()
            duty_types_dict = {dt.code: dt.name for dt in raw_duty_types}
            duty_qs = TeacherDutySchedule.objects.filter(
                teacher=teacher,
                day_of_week=today_weekday
            )
            if current_year:
                duty_qs = duty_qs.filter(academic_year=current_year)
            for d in duty_qs.order_by('period_number'):
                today_duty_list.append({
                    'period_number': d.period_number,
                    'duty_name': duty_types_dict.get(d.duty_type, d.duty_type),
                    'notes': d.notes or '',
                })
        except Exception:
            today_duty_list = []

    return render(request, 'dashboard/teacher_dashboard.html', {
        'teacher': teacher,
        'today_schedule': today_schedule,
        'today_duty_list': today_duty_list,
        'my_classes': my_classes,
        'my_homeroom': my_homeroom,
        'my_students_count': my_students_count,
        'today_att_done': today_att_done,
        'today_teacher_att': today_teacher_att,
        'today_punch': today_punch,
        'my_month_present': my_month_present,
        'my_month_late': my_month_late,
        'att_config': att_config,
        'recent_terms': recent_terms,
        'announcements': announcements,
        'active_year': current_year,
    })



# ----------------- 4. STUDENT & PARENT DASHBOARD -----------------

@login_required
@role_required(['ADMIN', 'STUDENT'])
def student_dashboard(request):
    user = request.user
    student = getattr(user, 'student_profile', None)
    
    if not student and user.is_superuser:
        student = Student.objects.first()

    today_weekday = datetime.now().weekday() + 1
    today_schedule = []
    
    if student and student.classroom:
        today_schedule = Timetable.objects.filter(
            classroom=student.classroom,
            day_of_week=today_weekday
        ).select_related('subject', 'teacher').order_by('start_time')

    recent_grades = []
    if student:
        recent_grades = Grade.objects.filter(student=student).select_related('subject', 'exam_term').order_by('-exam_term__start_date')[:6]

    invoices = []
    if student:
        invoices = Invoice.objects.filter(student=student).order_by('-created_at')[:3]

    announcements = Announcement.objects.filter(
        Q(target_audience__in=['ALL', 'STUDENTS', 'PARENTS']),
        is_published=True
    )[:4]

    return render(request, 'dashboard/student_dashboard.html', {
        'student': student,
        'today_schedule': today_schedule,
        'recent_grades': recent_grades,
        'invoices': invoices,
        'announcements': announcements,
    })


# ----------------- 5. MOEYS STATISTICAL REPORTING -----------------

@login_required
@role_required(['ADMIN', 'TEACHER', 'ACCOUNTANT'])
def moeys_statistics_view(request):
    """
    Official MoEYS Educational Statistics & EMIS Dashboard
    Computes Grade 7-12 stats, Gender breakdown, Academic Tracks, Scholarships, and Exam passing rates.
    Strictly isolated per Academic Year!
    """
    from apps.academics.utils import get_active_academic_year
    academic_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()
    
    students_base = Student.objects.filter(status='ACTIVE')
    classrooms_base = Classroom.objects.all()
    if academic_year:
        students_base = students_base.filter(Q(academic_year=academic_year) | Q(classroom__academic_year=academic_year))
        classrooms_base = classrooms_base.filter(academic_year=academic_year)

    total_students = students_base.count()
    female_students = students_base.filter(gender='F').count()
    male_students = total_students - female_students
    female_percent = round((female_students / total_students) * 100, 1) if total_students > 0 else 0.0

    total_teachers = Teacher.objects.filter(status='ACTIVE').count()
    female_teachers = Teacher.objects.filter(status='ACTIVE', gender='F').count()

    # 1. Grade 7 to 12 Breakdown
    grade_levels = [7, 8, 9, 10, 11, 12]
    grade_stats = []
    for g in grade_levels:
        classes = classrooms_base.filter(grade_level=g)
        stu_in_grade = students_base.filter(classroom__in=classes)
        t_cnt = stu_in_grade.count()
        f_cnt = stu_in_grade.filter(gender='F').count()
        m_cnt = t_cnt - f_cnt
        f_rate = round((f_cnt / t_cnt) * 100, 1) if t_cnt > 0 else 0.0
        
        # Track info for upper grades
        science_cnt = stu_in_grade.filter(classroom__track='SCIENCE').count()
        social_cnt = stu_in_grade.filter(classroom__track='SOCIAL').count()

        grade_stats.append({
            'grade': g,
            'name': f"ថ្នាក់ទី{g}",
            'classes_count': classes.count(),
            'total_students': t_cnt,
            'male_count': m_cnt,
            'female_count': f_cnt,
            'female_percent': f_rate,
            'science_count': science_cnt,
            'social_count': social_cnt,
        })

    # 2. Scholarship / Social Equity statistics
    scholarship_stats = {
        'full_pay': students_base.filter(scholarship_type=Student.ScholarshipType.FULL_PAY).count(),
        'sch_50': students_base.filter(scholarship_type=Student.ScholarshipType.SCHOLARSHIP_50).count(),
        'sch_100': students_base.filter(scholarship_type=Student.ScholarshipType.SCHOLARSHIP_100).count(),
        'installment': students_base.filter(scholarship_type=Student.ScholarshipType.INSTALLMENT).count(),
    }

    # 3. Dropout & Status statistics
    dropout_count = Student.objects.filter(status=Student.Status.DROPPED).count()
    transferred_count = Student.objects.filter(status=Student.Status.TRANSFERRED).count()
    graduated_count = Student.objects.filter(status=Student.Status.GRADUATED).count()

    # 4. Exam Performance Summary (Latest Term in this academic year)
    latest_term_qs = ExamTerm.objects.filter(is_published=True)
    if academic_year:
        latest_term_qs = latest_term_qs.filter(academic_year=academic_year)
    latest_term = latest_term_qs.first()
    exam_perf = []
    if latest_term:
        for g in grade_levels:
            classes = classrooms_base.filter(grade_level=g)
            stu_in_grade = students_base.filter(classroom__in=classes)
            
            # Grade stats for this term
            stu_tested = 0
            stu_passed = 0
            stu_failed = 0
            for s in stu_in_grade:
                grades = Grade.objects.filter(student=s, exam_term=latest_term)
                if grades.exists():
                    stu_tested += 1
                    avg_sc = grades.aggregate(a=Avg('score'))['a'] or 0
                    if avg_sc >= 50.0:
                        stu_passed += 1
                    else:
                        stu_failed += 1

            pass_rate = round((stu_passed / stu_tested) * 100, 1) if stu_tested > 0 else 0.0
            exam_perf.append({
                'grade': f"ថ្នាក់ទី{g}",
                'tested': stu_tested,
                'passed': stu_passed,
                'failed': stu_failed,
                'pass_rate': pass_rate,
            })

    return render(request, 'dashboard/moeys_reports.html', {
        'academic_year': academic_year,
        'total_students': total_students,
        'female_students': female_students,
        'male_students': male_students,
        'female_percent': female_percent,
        'total_teachers': total_teachers,
        'female_teachers': female_teachers,
        'grade_stats': grade_stats,
        'scholarship_stats': scholarship_stats,
        'dropout_count': dropout_count,
        'transferred_count': transferred_count,
        'graduated_count': graduated_count,
        'latest_term': latest_term,
        'exam_perf': exam_perf,
    })


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def export_moeys_excel(request):
    """
    Exports full MoEYS Educational Statistics report in multi-sheet Excel format
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    
    # Sheet 1: MoEYS Student Stats
    ws1 = wb.active
    ws1.title = "MoEYS EMIS Statistics"

    # Ministry Header
    ws1.merge_cells('A1:G1')
    ws1['A1'] = "ព្រះរាជាណាចក្រកម្ពុជា ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal="center")

    ws1.merge_cells('A2:G2')
    ws1['A2'] = "ក្រសួងអប់រំ យុវជន និងកីឡា - របាយការណ៍ស្ថិតិអប់រំប្រចាំឆ្នាំ (EMIS Report)"
    ws1['A2'].font = Font(bold=True, size=12, color="1E3A8A")
    ws1['A2'].alignment = Alignment(horizontal="center")

    headers = ['កម្រិតថ្នាក់ (Grade)', 'ចំនួនថ្នាក់ (Classes)', 'សិស្សសរុប (Total)', 'ប្រុស (Male)', 'ស្រី (Female)', 'អត្រាសិស្សស្រី (%)', 'ជំនាញ (Track)']
    ws1.append([])
    ws1.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws1.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for g in [7, 8, 9, 10, 11, 12]:
        classes = Classroom.objects.filter(grade_level=g)
        stu_in_grade = Student.objects.filter(classroom__in=classes, status='ACTIVE')
        t_cnt = stu_in_grade.count()
        f_cnt = stu_in_grade.filter(gender='F').count()
        m_cnt = t_cnt - f_cnt
        f_rate = round((f_cnt / t_cnt) * 100, 1) if t_cnt > 0 else 0.0

        track_info = "ទូទៅ"
        if g in [11, 12]:
            sci = stu_in_grade.filter(classroom__track='SCIENCE').count()
            soc = stu_in_grade.filter(classroom__track='SOCIAL').count()
            track_info = f"វិទ្យាសាស្ត្រ: {sci} | សង្គម: {soc}"

        ws1.append([
            f"ថ្នាក់ទី{g}",
            classes.count(),
            t_cnt,
            m_cnt,
            f_cnt,
            f"{f_rate}%",
            track_info
        ])

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="moeys_statistics_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ----------------- DATA EXPORT (STUDENTS & FINANCE) -----------------

@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def export_students_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="students_list_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Student ID', 'Khmer Name', 'Latin Name', 'Gender', 'DOB', 'Class', 'Phone', 'Father Name', 'Father Phone', 'Status', 'Scholarship Type'])
    
    for s in Student.objects.select_related('classroom').all():
        writer.writerow([
            s.student_id,
            s.khmer_name,
            s.latin_name,
            s.get_gender_display(),
            s.date_of_birth.strftime('%Y-%m-%d') if s.date_of_birth else '',
            s.classroom.name if s.classroom else '',
            s.phone or '',
            s.father_name or '',
            s.father_phone or '',
            s.get_status_display(),
            s.get_scholarship_type_display(),
        ])
    return response


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def export_students_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Directory"

    headers = ['Student ID', 'Khmer Name', 'Latin Name', 'Gender', 'DOB', 'Class', 'Phone', 'Father Name', 'Father Phone', 'Status', 'Scholarship Type']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for s in Student.objects.select_related('classroom').all():
        ws.append([
            s.student_id,
            s.khmer_name,
            s.latin_name,
            s.get_gender_display(),
            s.date_of_birth.strftime('%Y-%m-%d') if s.date_of_birth else '',
            s.classroom.name if s.classroom else '',
            s.phone or '',
            s.father_name or '',
            s.father_phone or '',
            s.get_status_display(),
            s.get_scholarship_type_display(),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="students_list_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def export_finance_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    
    # Sheet 1: Invoices
    ws1 = wb.active
    ws1.title = "Invoices & Revenue"
    ws1.append(['Invoice No', 'Student ID', 'Student Name', 'Category', 'Original ($)', 'Discount (%)', 'Final ($)', 'Paid ($)', 'Remaining ($)', 'Due Date', 'Status'])

    for inv in Invoice.objects.select_related('student', 'fee_category').all():
        ws1.append([
            inv.invoice_no,
            inv.student.student_id,
            inv.student.khmer_name,
            inv.fee_category.name,
            float(inv.original_amount),
            float(inv.discount_percent),
            float(inv.final_amount),
            float(inv.paid_amount),
            float(inv.remaining_balance),
            inv.due_date.strftime('%Y-%m-%d'),
            inv.get_status_display(),
        ])

    # Sheet 2: Expenses
    ws2 = wb.create_sheet(title="Expenses")
    ws2.append(['Expense Title', 'Category', 'Amount ($)', 'Date', 'Recorded By', 'Notes'])
    for exp in Expense.objects.select_related('recorded_by').all():
        ws2.append([
            exp.title,
            exp.get_category_display(),
            float(exp.amount),
            exp.date.strftime('%Y-%m-%d'),
            exp.recorded_by.username if exp.recorded_by else '',
            exp.notes or '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="finance_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
