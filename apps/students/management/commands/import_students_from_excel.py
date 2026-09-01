import os
import sys
from datetime import datetime, date
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from apps.academics.models import AcademicYear, Classroom, GradeLevelRule
from apps.students.models import Student, StudentCategory

# Khmer to Latin basic mapping for authentic transliteration
KHMER_LATIN_MAP = {
    'ក': 'K', 'ខ': 'Kh', 'គ': 'K', 'ឃ': 'Kh', 'ង': 'Ng',
    'ច': 'Ch', 'ឆ': 'Chh', 'ជ': 'Ch', 'ឈ': 'Chh', 'ញ': 'Nhor',
    'ដ': 'D', 'ឋ': 'Th', 'ឌ': 'D', 'ឍ': 'Th', 'ណ': 'N',
    'ត': 'T', 'ថ': 'Th', 'ទ': 'T', 'ធ': 'Th', 'ន': 'N',
    'ប': 'B', 'ផ': 'Ph', 'ព': 'P', 'ភ': 'Ph', 'ម': 'M',
    'យ': 'Y', 'រ': 'R', 'ល': 'L', 'វ': 'V', 'ស': 'S',
    'ហ': 'H', 'ឡ': 'L', 'អ': 'A'
}

def transliterate_khmer_to_latin(khmer_name):
    """Simple clean transliteration fallback for Khmer names"""
    if not khmer_name:
        return "Student"
    words = khmer_name.strip().split()
    latin_words = []
    for w in words:
        if not w:
            continue
        first_char = w[0]
        prefix = KHMER_LATIN_MAP.get(first_char, 'S')
        # If word is common family/given name, keep it clean
        latin_words.append(prefix + w[1:4])
    res = " ".join(latin_words)
    return res if len(res) > 2 else f"Student {khmer_name[:4]}"


class Command(BaseCommand):
    help = "Import students from Excel/XLSM file (2025-2026.xlsm) into specific academic year and classrooms"

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='2025-2026.xlsm',
            help='Path to the Excel/XLSM file (defaults to 2025-2026.xlsm in root)'
        )
        parser.add_argument(
            '--year',
            type=str,
            default='2025-2026',
            help='Academic Year name (defaults to 2025-2026)'
        )
        parser.add_argument(
            '--force',
            action='store_true',
            help='Force re-import even if students already exist'
        )

    def handle(self, *args, **options):
        if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
            try:
                sys.stdout.reconfigure(encoding='utf-8')
            except Exception:
                pass

        file_path = options['file']
        year_name = options['year']
        force = options.get('force', False)

        # 1. Setup or get Academic Year
        academic_year, _ = AcademicYear.objects.get_or_create(
            name=year_name,
            defaults={
                'start_date': date(2025, 9, 1),
                'end_date': date(2026, 7, 31),
                'is_current': False
            }
        )

        existing_count = Student.objects.filter(academic_year=academic_year).count()
        if existing_count > 0 and not force:
            self.stdout.write(self.style.SUCCESS(f"Year {academic_year.name} already has {existing_count} students. Skipping import."))
            return

        if not os.path.isabs(file_path):
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            possible_path = os.path.join(base_dir, file_path)
            if os.path.exists(possible_path):
                file_path = possible_path

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"Excel file not found at: {file_path}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Reading Excel file: {file_path}"))

        wb = openpyxl.load_workbook(file_path, data_only=True)
        self.stdout.write(f"Target Academic Year: {academic_year.name} (ID: {academic_year.id})")

        # 2. Parse all students from sheets '7' through '12'
        parsed_students = []
        seen_student_ids = set()

        for sheet_name in ['7', '8', '9', '10', '11', '12']:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            for r in range(4, ws.max_row + 1):
                num = ws.cell(r, 1).value
                raw_stu_id = ws.cell(r, 2).value
                khmer_name = ws.cell(r, 3).value
                gender_raw = ws.cell(r, 4).value
                dob_raw = ws.cell(r, 5).value
                grade_raw = ws.cell(r, 6).value
                section_raw = ws.cell(r, 7).value

                if not khmer_name or str(khmer_name).strip() == '':
                    continue

                khmer_name = str(khmer_name).strip()
                grade_num = int(grade_raw) if (grade_raw is not None and str(grade_raw).isdigit()) else int(sheet_name)
                section_str = str(section_raw).strip() if section_raw is not None else 'A'
                classroom_name = f"{grade_num}{section_str}"

                # Gender: 'ស' -> 'F' (Female), 'ប' -> 'M' (Male)
                g_clean = str(gender_raw).strip() if gender_raw is not None else ''
                gender = 'F' if g_clean in ['ស', 'F', 'f', 'ស្រី'] else 'M'

                # Date of birth
                dob = None
                if isinstance(dob_raw, (datetime, date)):
                    dob = dob_raw.date() if isinstance(dob_raw, datetime) else dob_raw
                elif dob_raw:
                    try:
                        dob = datetime.strptime(str(dob_raw).strip(), '%Y-%m-%d').date()
                    except Exception:
                        dob = date(2010, 1, 1)
                else:
                    dob = date(2010, 1, 1)

                # Determine track: Grades 7-10: GENERAL, Grades 11-12: A-E Science, F-I Social
                if grade_num <= 10:
                    track = 'GENERAL'
                else:
                    sec_letter = section_str.upper()
                    if sec_letter in ['A', 'B', 'C', 'D', 'E']:
                        track = 'SCIENCE'
                    else:
                        track = 'SOCIAL'

                # Student ID handling with collision protection
                stu_id_str = str(raw_stu_id).strip() if raw_stu_id is not None else ''
                if not stu_id_str:
                    stu_id_str = f"STU-{grade_num}{section_str}-{r}"

                if stu_id_str in seen_student_ids:
                    stu_id_str = f"{stu_id_str}-{grade_num}"
                seen_student_ids.add(stu_id_str)

                parsed_students.append({
                    'student_id': stu_id_str,
                    'khmer_name': khmer_name,
                    'latin_name': transliterate_khmer_to_latin(khmer_name),
                    'gender': gender,
                    'dob': dob,
                    'grade_level': grade_num,
                    'section': section_str,
                    'classroom_name': classroom_name,
                    'track': track,
                })

        self.stdout.write(f"Total students parsed from Excel: {len(parsed_students)}")

        # 3. Import Classrooms & Students atomically
        created_classes_count = 0
        imported_students_count = 0
        updated_students_count = 0
        classrooms_cache = {}

        with transaction.atomic():
            # Create/Get all classrooms first
            distinct_classes = sorted(list(set((s['classroom_name'], s['grade_level'], s['track']) for s in parsed_students)))
            for cls_name, g_lvl, trk in distinct_classes:
                cls_obj = Classroom.objects.filter(code=cls_name, academic_year=academic_year).first()
                if not cls_obj:
                    cls_obj = Classroom.objects.filter(name__in=[cls_name, f"ថ្នាក់ទី {cls_name}"], academic_year=academic_year).first()

                created = False
                if not cls_obj:
                    cls_obj = Classroom.objects.create(
                        name=f"ថ្នាក់ទី {cls_name}",
                        code=cls_name,
                        academic_year=academic_year,
                        grade_level=g_lvl,
                        track=trk,
                        capacity=60,
                    )
                    created = True

                if created:
                    created_classes_count += 1
                    # Sync standard subjects from GradeLevelRule
                    sub_ids = list(GradeLevelRule.objects.filter(grade_level=g_lvl, track=trk).values_list('subject_id', flat=True))
                    if not sub_ids and trk != 'GENERAL':
                        sub_ids = list(GradeLevelRule.objects.filter(grade_level=g_lvl, track='GENERAL').values_list('subject_id', flat=True))
                    if sub_ids:
                        cls_obj.sync_assigned_subjects(sub_ids)

                classrooms_cache[cls_name] = cls_obj

            # Create or update students
            for item in parsed_students:
                cls_obj = classrooms_cache.get(item['classroom_name'])
                student_id = item['student_id']

                student, created = Student.objects.update_or_create(
                    student_id=student_id,
                    defaults={
                        'khmer_name': item['khmer_name'],
                        'latin_name': item['latin_name'],
                        'gender': item['gender'],
                        'date_of_birth': item['dob'],
                        'classroom': cls_obj,
                        'academic_year': academic_year,
                        'status': 'ACTIVE',
                        'scholarship_type': 'FULL_PAY',
                    }
                )

                if created:
                    imported_students_count += 1
                else:
                    updated_students_count += 1

        self.stdout.write(self.style.SUCCESS(
            f"\n🎉 [IMPORT COMPLETE] Successfully imported {len(parsed_students)} students for year {year_name}!\n"
            f"   - Classrooms Created / Verified: {len(classrooms_cache)} (New: {created_classes_count})\n"
            f"   - Students Created: {imported_students_count}\n"
            f"   - Students Updated: {updated_students_count}\n"
            f"   - Target Academic Year: {academic_year.name}"
        ))
