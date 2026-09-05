import os
import openpyxl
from datetime import datetime, date
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from apps.academics.models import AcademicYear, Classroom, GradeLevelRule
from apps.students.models import Student, StudentPromotionRecord
from apps.examinations.models import ExamCandidate, ExamStudentExclusion
from apps.attendance.models import StudentAttendance
from apps.finance.models import Invoice, StudentMonthlyPayment, PaymentSlipSubmission
from apps.students.khmer_romanizer import romanize_khmer_name


class Command(BaseCommand):
    help = "Clean-wipe old students and import official 2026-2027 master roster (1,998 students across 40 classrooms) from 2026-2027.xlsm"

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default=os.path.join(settings.BASE_DIR, '2026-2027.xlsm'),
            help='Path to the 2026-2027.xlsm file'
        )

    def handle(self, *args, **options):
        xlsm_path = options['file']
        if not os.path.exists(xlsm_path):
            self.stderr.write(self.style.ERROR(f"File not found at: {xlsm_path}"))
            return

        self.stdout.write(self.style.NOTICE("=== STEP 1: Setting Academic Year 2026-2027 ==="))
        ay, _ = AcademicYear.objects.get_or_create(
            name='2026-2027',
            defaults={
                'start_date': date(2026, 11, 1),
                'end_date': date(2027, 8, 31),
                'is_current': True
            }
        )
        AcademicYear.objects.exclude(id=ay.id).update(is_current=False)
        ay.is_current = True
        ay.save()
        self.stdout.write(self.style.SUCCESS(f"Academic Year: {ay.name} is now ACTIVE."))

        self.stdout.write(self.style.NOTICE("=== STEP 2: Ensuring 40 Standard Classrooms ==="))
        official_40 = [
            ('7A', 7, 'GENERAL'), ('7B', 7, 'GENERAL'), ('7C', 7, 'GENERAL'), ('7D', 7, 'GENERAL'), ('7E', 7, 'GENERAL'),
            ('8A', 8, 'GENERAL'), ('8B', 8, 'GENERAL'), ('8C', 8, 'GENERAL'), ('8D', 8, 'GENERAL'),
            ('9A', 9, 'GENERAL'), ('9B', 9, 'GENERAL'), ('9C', 9, 'GENERAL'), ('9D', 9, 'GENERAL'),
            ('10A', 10, 'GENERAL'), ('10B', 10, 'GENERAL'), ('10C', 10, 'GENERAL'), ('10D', 10, 'GENERAL'),
            ('10E', 10, 'GENERAL'), ('10F', 10, 'GENERAL'), ('10G', 10, 'GENERAL'), ('10H', 10, 'GENERAL'), ('10I', 10, 'GENERAL'),
            ('11A', 11, 'SCIENCE'), ('11B', 11, 'SCIENCE'), ('11C', 11, 'SCIENCE'), ('11D', 11, 'SCIENCE'), ('11E', 11, 'SCIENCE'),
            ('11F', 11, 'SOCIAL'), ('11G', 11, 'SOCIAL'), ('11H', 11, 'SOCIAL'), ('11I', 11, 'SOCIAL'),
            ('12A', 12, 'SCIENCE'), ('12B', 12, 'SCIENCE'), ('12C', 12, 'SCIENCE'), ('12D', 12, 'SCIENCE'),
            ('12E', 12, 'SOCIAL'), ('12F', 12, 'SOCIAL'), ('12G', 12, 'SOCIAL'), ('12H', 12, 'SOCIAL'), ('12I', 12, 'SOCIAL'),
        ]
        valid_codes = {item[0].upper() for item in official_40}

        classroom_map = {}
        with transaction.atomic():
            for code, grade, track in official_40:
                c_obj, _ = Classroom.objects.update_or_create(
                    academic_year=ay,
                    code=code,
                    defaults={
                        'name': f"ថ្នាក់ទី {code}",
                        'grade_level': grade,
                        'track': track,
                        'capacity': 50
                    }
                )
                c_obj.name = f"ថ្នាក់ទី {code}"
                c_obj.grade_level = grade
                c_obj.track = track
                c_obj.save()

                # Sync default subjects
                default_sub_ids = list(GradeLevelRule.objects.filter(
                    grade_level=grade,
                    track=track
                ).values_list('subject_id', flat=True))
                if default_sub_ids:
                    c_obj.sync_assigned_subjects(default_sub_ids)

                classroom_map[code.upper()] = c_obj

            # Delete redundant classrooms in 2026-2027
            redundant = Classroom.objects.filter(academic_year=ay).exclude(code__in=valid_codes)
            for rc in redundant:
                rc.students.all().delete()
                rc.delete()

        self.stdout.write(self.style.SUCCESS(f"40 classrooms configured and synced with subjects."))

        self.stdout.write(self.style.NOTICE("=== STEP 3: Clean Wipe of Old Student Data ==="))
        with transaction.atomic():
            ExamCandidate.objects.all().delete()
            ExamStudentExclusion.objects.all().delete()
            StudentAttendance.objects.all().delete()
            StudentMonthlyPayment.objects.all().delete()
            Invoice.objects.all().delete()
            PaymentSlipSubmission.objects.all().delete()
            StudentPromotionRecord.objects.all().delete()
            deleted_count, _ = Student.objects.all().delete()

        self.stdout.write(self.style.SUCCESS(f"Deleted old student data cleanly."))

        self.stdout.write(self.style.NOTICE(f"=== STEP 4: Reading and Importing from {xlsm_path} ==="))
        wb = openpyxl.load_workbook(xlsm_path, data_only=True)

        def parse_gender(val):
            v = str(val).strip().upper() if val else ''
            if v in ['ស', 'ស្រី', 'F', 'FEMALE', 'GIRL']:
                return 'F'
            return 'M'

        def parse_dob(val):
            if not val:
                return None
            if isinstance(val, (datetime, date)):
                return val.date() if isinstance(val, datetime) else val
            val_str = str(val).strip()
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
                try:
                    return datetime.strptime(val_str.split()[0], fmt).date()
                except ValueError:
                    pass
            return None

        to_create = []
        seen_ids = set()

        for sname in ['7', '8', '9', '10', '11', '12']:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            sheet_count = 0
            for r in rows[1:]:
                if not r or not any(r):
                    continue
                st_id = str(r[1]).strip() if r[1] is not None else ''
                st_name = str(r[2]).strip() if r[2] is not None else ''
                if not st_id or not st_name or st_id.lower() in ['none', '', 'ល.រ', 'អត្តលេខ']:
                    continue

                gender = parse_gender(r[3])
                dob = parse_dob(r[4]) or date(2010, 1, 1)
                grade_val = str(r[5]).strip() if r[5] is not None else sname
                class_let = str(r[6]).strip() if r[6] is not None else ''
                class_code = f"{grade_val}{class_let}".upper().replace(' ', '')

                target_class = classroom_map.get(class_code)
                if not target_class:
                    continue

                clean_id = st_id
                if clean_id.lower() in seen_ids:
                    clean_id = f"{clean_id}-{grade_val}"
                while clean_id.lower() in seen_ids:
                    clean_id = f"{clean_id}-A"
                seen_ids.add(clean_id.lower())

                latin_name = romanize_khmer_name(st_name)

                student = Student(
                    student_id=clean_id,
                    khmer_name=st_name,
                    latin_name=latin_name,
                    gender=gender,
                    date_of_birth=dob,
                    classroom=target_class,
                    academic_year=ay,
                    status='ACTIVE'
                )
                to_create.append(student)
                sheet_count += 1

            self.stdout.write(f"  - Sheet {sname}: {sheet_count} students processed")

        with transaction.atomic():
            Student.objects.bulk_create(to_create, batch_size=500)

        total_imported = Student.objects.count()
        self.stdout.write(self.style.SUCCESS(f"\nSuccessfully imported {total_imported} students into 2026-2027!"))
