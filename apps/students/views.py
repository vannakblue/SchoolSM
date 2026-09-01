import csv
import io
import re
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.db import transaction
from apps.accounts.decorators import role_required
from apps.accounts.models import User
from .models import Student, ScholarshipType, StudentStatusConfig
from .forms import StudentEnrollmentForm
from apps.academics.models import Classroom, AcademicYear
from apps.attendance.models import StudentAttendance
from apps.examinations.models import Grade, ExamTerm
from apps.finance.models import Invoice
from apps.extras.models import BookBorrowing

@login_required
def student_list(request):
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    
    selected_year = ''
    if 'academic_year' in request.GET or 'year' in request.GET:
        raw_year = (request.GET.get('academic_year') if 'academic_year' in request.GET else request.GET.get('year') or '').strip()
        if raw_year == '' or raw_year.lower() == 'all':
            active_year = None
            selected_year = ''
        elif raw_year.isdigit():
            found_year = AcademicYear.objects.filter(id=int(raw_year)).first()
            if found_year:
                active_year = found_year
                selected_year = str(found_year.id)
            else:
                active_year = None
                selected_year = ''
        else:
            found_year = AcademicYear.objects.filter(name=raw_year).first()
            if found_year:
                active_year = found_year
                selected_year = str(found_year.id)
            else:
                active_year = None
                selected_year = ''
    elif active_year:
        selected_year = str(active_year.id)

    query = request.GET.get('q', '').strip()
    class_id = request.GET.get('classroom', '').strip()
    status_filter = request.GET.get('status', '').strip()
    scholarship_filter = request.GET.get('scholarship', '').strip()
    exam_status_filter = request.GET.get('exam_status', '').strip()

    students = Student.objects.select_related('classroom', 'academic_year').all()

    if active_year:
        students = students.filter(Q(academic_year=active_year) | Q(classroom__academic_year=active_year))

    if query:
        students = students.filter(
            Q(student_id__icontains=query) |
            Q(khmer_name__icontains=query) |
            Q(latin_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(father_name__icontains=query) |
            Q(mother_name__icontains=query)
        )

    if class_id and class_id.isdigit():
        students = students.filter(classroom_id=int(class_id))

    if status_filter:
        students = students.filter(status=status_filter)

    if scholarship_filter:
        students = students.filter(scholarship_type=scholarship_filter)

    if exam_status_filter == 'disqualified':
        students = students.filter(
            Q(is_exam_suspended=True) |
            Q(status__in=[Student.Status.SUSPENDED, Student.Status.DROPPED, Student.Status.TRANSFERRED])
        )
    elif exam_status_filter == 'eligible':
        students = students.filter(
            status=Student.Status.ACTIVE,
            is_exam_suspended=False
        )

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    classrooms = Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code')

    StudentStatusConfig.ensure_default_statuses()
    available_statuses = StudentStatusConfig.objects.filter(is_active=True).order_by('order', 'id')

    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    total_count = students.count()
    per_page_param = request.GET.get('per_page', '50').strip()

    if per_page_param == 'all':
        students_page = students
        paginator = None
        is_paginated = False
    else:
        try:
            per_page = int(per_page_param)
            if per_page not in [25, 50, 100, 200]:
                per_page = 50
        except (ValueError, TypeError):
            per_page = 50

        paginator = Paginator(students, per_page)
        page = request.GET.get('page', 1)
        try:
            students_page = paginator.page(page)
        except PageNotAnInteger:
            students_page = paginator.page(1)
        except EmptyPage:
            students_page = paginator.page(paginator.num_pages)
        is_paginated = paginator.num_pages > 1

    return render(request, 'students/student_list.html', {
        'students': students_page,
        'paginator': paginator,
        'page_obj': students_page if is_paginated else None,
        'is_paginated': is_paginated,
        'per_page': per_page_param,
        'classrooms': classrooms,
        'academic_years': academic_years,
        'active_year': active_year,
        'selected_year': str(active_year.id) if active_year else (selected_year if selected_year == 'all' else ''),
        'query': query,
        'selected_class': class_id,
        'selected_status': status_filter,
        'selected_scholarship': scholarship_filter,
        'selected_exam_status': exam_status_filter,
        'available_statuses': available_statuses,
        'exam_reasons': Student.ExamExclusionReason.choices,
        'total_count': total_count,
    })


def _extract_grade_options(request, classroom, existing_data=None):
    """Helper to extract and validate dynamic grade-level custom fields from request"""
    from apps.academics.models import GradeLevel, GradeEnrollmentOption
    from django.core.files.storage import default_storage

    enrollment_data = dict(existing_data) if existing_data else {}
    if not classroom:
        return enrollment_data

    gl = GradeLevel.objects.filter(grade_number=classroom.grade_level, track=classroom.track).first()
    if not gl:
        gl = GradeLevel.objects.filter(grade_number=classroom.grade_level).first()

    if gl:
        options = gl.enrollment_options.filter(is_active=True)
        for opt in options:
            if opt.field_type == GradeEnrollmentOption.FieldType.SECTION:
                continue

            post_key = f"grade_opt_{opt.field_name}"
            file_key = f"grade_opt_{opt.field_name}"

            if opt.field_type == GradeEnrollmentOption.FieldType.FILE:
                if file_key in request.FILES:
                    uploaded_file = request.FILES[file_key]
                    path = default_storage.save(f"students/docs/{uploaded_file.name}", uploaded_file)
                    enrollment_data[opt.field_name] = {
                        'label': opt.label,
                        'type': opt.field_type,
                        'value': path,
                        'url': default_storage.url(path)
                    }
            elif opt.field_type == GradeEnrollmentOption.FieldType.MULTISELECT:
                vals = request.POST.getlist(post_key)
                if vals:
                    enrollment_data[opt.field_name] = {
                        'label': opt.label,
                        'type': opt.field_type,
                        'value': ", ".join(vals),
                        'list': vals
                    }
                elif opt.field_name in enrollment_data:
                    enrollment_data.pop(opt.field_name, None)
            elif post_key in request.POST:
                val = request.POST.get(post_key, '').strip()
                if val:
                    enrollment_data[opt.field_name] = {
                        'label': opt.label,
                        'type': opt.field_type,
                        'value': val
                    }
                elif opt.field_name in enrollment_data:
                    enrollment_data.pop(opt.field_name, None)
    return enrollment_data



@login_required
@role_required(['ADMIN'])
def student_enroll(request):
    from apps.academics.utils import get_active_academic_year
    current_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()
    
    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST, request.FILES, academic_year=current_year)
        if form.is_valid():
            with transaction.atomic():
                student = form.save(commit=False)
                if not student.academic_year:
                    student.academic_year = current_year
                student.enrollment_data = _extract_grade_options(request, student.classroom)
                student.save()
                
                # Create user account for student/parent login if username doesn't exist
                username = student.student_id.lower().replace('-', '_')
                user = User.objects.filter(username=username).first()
                if not user:
                    user = User.objects.create_user(
                        username=username,
                        password='p123456',
                        role=User.Role.STUDENT,
                        khmer_name=student.khmer_name,
                        latin_name=student.latin_name,
                        phone=student.phone or student.father_phone or ''
                    )
                student.user = user
                student.save(update_fields=['user'])

            messages.success(request, f"🎉 បានចុះឈ្មោះសិស្ស {student.khmer_name} (ID: {student.student_id}) ក្នុងឆ្នាំសិក្សា {current_year.name if current_year else ''} ដោយជោគជ័យ! ពាក្យសម្ងាត់ដំបូងគឺ 'p123456'")
            return redirect('student_detail', pk=student.pk)
        else:
            messages.error(request, "សូមពិនិត្យទម្រង់ដែលបានបំពេញឡើងវិញ!")
    else:
        form = StudentEnrollmentForm(initial={'academic_year': current_year}, academic_year=current_year)

    return render(request, 'students/student_form.html', {
        'form': form,
        'current_year': current_year,
        'title': 'ចុះឈ្មោះសិស្សថ្មី / Student Enrollment'
    })


def public_student_enroll(request):
    """
    Public online self-registration for students & parents via smartphone or computer.
    Supports pre-selecting target classroom or filtering by grade level via query params (?classroom=<id> or ?grade=<grade>&track=<track>).
    No login required.
    """
    from apps.academics.utils import get_active_academic_year
    current_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()
    classrooms = Classroom.objects.filter(academic_year=current_year).select_related('academic_year').order_by('grade_level', 'code') if current_year else Classroom.objects.select_related('academic_year').order_by('grade_level', 'code')

    classroom_id = request.GET.get('classroom')
    grade_param = request.GET.get('grade')
    track_param = request.GET.get('track')
    
    target_classroom = None
    target_grade_name = None

    if classroom_id:
        target_classroom = classrooms.filter(id=classroom_id).first()
    elif grade_param:
        try:
            g_num = int(grade_param)
            classrooms = classrooms.filter(grade_level=g_num)
            track_name = ""
            if track_param:
                classrooms = classrooms.filter(track=track_param)
                if track_param == 'SCIENCE':
                    track_name = " វិទ្យាសាស្ត្រ (Science Track)"
                elif track_param == 'SOCIAL':
                    track_name = " វិទ្យាសាស្ត្រសង្គម (Social Track)"
            target_grade_name = f"ថ្នាក់ទី {g_num}{track_name}"
        except (ValueError, TypeError):
            pass

    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST, request.FILES, academic_year=current_year)
        if form.is_valid():
            with transaction.atomic():
                student = form.save(commit=False)
                if not student.academic_year:
                    student.academic_year = current_year
                student.status = Student.Status.ACTIVE
                student.enrollment_data = _extract_grade_options(request, student.classroom)
                student.save()

                # Create user account for student login
                username = student.student_id.lower().replace('-', '_')
                user = User.objects.filter(username=username).first()
                if not user:
                    user = User.objects.create_user(
                        username=username,
                        password='p123456',
                        role=User.Role.STUDENT,
                        khmer_name=student.khmer_name,
                        latin_name=student.latin_name,
                        phone=student.phone or student.father_phone or ''
                    )
                student.user = user
                student.save(update_fields=['user'])

            messages.success(request, f"🎉 ការចុះឈ្មោះសិស្ស {student.khmer_name} បានជោគជ័យ!")
            return redirect('public_enroll_success', pk=student.pk)
        else:
            messages.error(request, "សូមពិនិត្យព័ត៌មានដែលបានបំពេញឡើងវិញ!")
    else:
        initial_data = {'academic_year': current_year}
        if target_classroom:
            initial_data['classroom'] = target_classroom
            if target_classroom.academic_year:
                initial_data['academic_year'] = target_classroom.academic_year
        form = StudentEnrollmentForm(initial=initial_data, academic_year=current_year)
        if grade_param:
            form.fields['classroom'].queryset = classrooms

    return render(request, 'students/public_enroll.html', {
        'form': form,
        'current_year': current_year,
        'classrooms': classrooms,
        'target_classroom': target_classroom,
        'target_grade_name': target_grade_name,
    })


def api_get_grade_options(request):
    """AJAX API to return custom enrollment options for a classroom or grade level"""
    from apps.academics.models import GradeLevel
    from django.http import JsonResponse

    classroom_id = request.GET.get('classroom_id')
    grade_level_id = request.GET.get('grade_level_id')
    
    gl = None
    if classroom_id:
        c = Classroom.objects.filter(id=classroom_id).first()
        if c:
            gl = GradeLevel.objects.filter(grade_number=c.grade_level, track=c.track).first()
            if not gl:
                gl = GradeLevel.objects.filter(grade_number=c.grade_level).first()
    elif grade_level_id:
        gl = GradeLevel.objects.filter(id=grade_level_id).first()
        
    if not gl:
        return JsonResponse({'status': 'success', 'data': [], 'grade_name': ''})
        
    options = gl.enrollment_options.filter(is_active=True).order_by('order', 'id')
    data = []
    for opt in options:
        data.append({
            'id': opt.id,
            'label': opt.label,
            'field_name': opt.field_name,
            'field_type': opt.field_type,
            'col_width': opt.col_width or 6,
            'choices': opt.get_choices_list(),
            'placeholder': opt.placeholder or '',
            'is_required': opt.is_required,
            'order': opt.order
        })
        
    return JsonResponse({
        'status': 'success',
        'grade_name': gl.name,
        'grade_number': gl.grade_number,
        'track': gl.track,
        'data': data
    })


def api_check_student_id(request):
    """
    AJAX API: Real-time Live Check for Student ID Uniqueness.
    Params:
      - student_id: String (The student ID being tested)
      - exclude_id: Optional Int (Current student PK if editing)
      - academic_year_id: Optional Int (Target academic year)
    Returns:
      - is_available: Boolean
      - message: String (Descriptive feedback in Khmer)
      - suggested_id: String (Next available collision-free ID)
    """
    from django.http import JsonResponse
    from apps.academics.utils import get_active_academic_year

    sid = request.GET.get('student_id', '').strip()
    exclude_id = request.GET.get('exclude_id')
    year_id = request.GET.get('academic_year_id') or request.GET.get('year_id')

    target_year = None
    if year_id:
        target_year = AcademicYear.objects.filter(id=year_id).first()
    if not target_year:
        target_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()

    suggested_id = Student.generate_unique_student_id(target_year, exclude_pk=exclude_id)

    if not sid:
        return JsonResponse({
            'status': 'success',
            'is_blank': True,
            'is_available': True,
            'message': f'⚡ ប្រព័ន្ធនឹងបង្កើតអត្តលេខស្វ័យប្រវត្តិតាមឆ្នាំសិក្សា (ឧ. {suggested_id})',
            'suggested_id': suggested_id
        })

    qs = Student.objects.filter(student_id__iexact=sid)
    if exclude_id:
        try:
            qs = qs.exclude(pk=int(exclude_id))
        except (ValueError, TypeError):
            pass

    if qs.exists():
        existing = qs.first()
        class_str = f" ({existing.classroom.name})" if existing.classroom else ""
        return JsonResponse({
            'status': 'duplicate',
            'is_blank': False,
            'is_available': False,
            'message': f"❌ អត្តលេខ '{sid}' ត្រូវបានប្រើប្រាស់រួចហើយដោយសិស្ស {existing.khmer_name}{class_str}!",
            'existing_student': {
                'id': existing.id,
                'name': existing.khmer_name,
                'classroom': existing.classroom.name if existing.classroom else 'គ្មានថ្នាក់'
            },
            'suggested_id': suggested_id
        })

    return JsonResponse({
        'status': 'available',
        'is_blank': False,
        'is_available': True,
        'message': f"✅ អត្តលេខ '{sid}' ទំនេរ អាចប្រើប្រាស់បាន!",
        'suggested_id': suggested_id
    })


def api_generate_student_id(request):
    """
    AJAX API: Returns next guaranteed collision-free Student ID.
    """
    from django.http import JsonResponse
    from apps.academics.utils import get_active_academic_year

    year_id = request.GET.get('academic_year_id') or request.GET.get('year_id')
    target_year = None
    if year_id:
        target_year = AcademicYear.objects.filter(id=year_id).first()
    if not target_year:
        target_year = get_active_academic_year(request) or AcademicYear.objects.filter(is_current=True).first()

    student_id = Student.generate_unique_student_id(target_year)
    return JsonResponse({
        'status': 'success',
        'student_id': student_id
    })


# ----------------- SCHOLARSHIP / FEE TYPES CRUD (ADMIN ONLY) -----------------

@login_required
@role_required(['ADMIN'])
def scholarship_type_list(request):
    """Lists all configurable Scholarship / Fee Types for Admin with student stats"""
    from .forms import ScholarshipTypeForm
    scholarships = ScholarshipType.objects.all().order_by('order', 'id')
    form = ScholarshipTypeForm()
    
    scholarship_stats = []
    for st in scholarships:
        cnt = Student.objects.filter(scholarship_type=st.code).count()
        scholarship_stats.append({
            'obj': st,
            'student_count': cnt
        })
        
    return render(request, 'students/scholarship_types.html', {
        'scholarships': scholarship_stats,
        'form': form
    })


@login_required
@role_required(['ADMIN'])
def scholarship_type_save(request, pk=None):
    """Create or update a Scholarship / Fee Type"""
    from .forms import ScholarshipTypeForm
    st = get_object_or_404(ScholarshipType, pk=pk) if pk else None
    if request.method == 'POST':
        form = ScholarshipTypeForm(request.POST, instance=st)
        if form.is_valid():
            saved_st = form.save()
            messages.success(request, f"🎉 បានរក្សាទុកប្រភេទកម្រៃ '{saved_st.name}' ដោយជោគជ័យ!")
        else:
            for f, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"កំហុស [{f}]: {e}")
    return redirect('scholarship_type_list')


@login_required
@role_required(['ADMIN'])
def scholarship_type_delete(request, pk):
    """Delete a Scholarship / Fee Type if not in active use"""
    st = get_object_or_404(ScholarshipType, pk=pk)
    if request.method == 'POST':
        student_count = Student.objects.filter(scholarship_type=st.code).count()
        if student_count > 0:
            messages.warning(request, f"⚠️ មិនអាចលុប '{st.name}' បានទេ ដោយសារមានសិស្សចំនួន {student_count} នាក់កំពុងប្រើប្រាស់!")
        else:
            name = st.name
            st.delete()
            messages.success(request, f"🗑️ បានលុបប្រភេទកម្រៃ '{name}' ដោយជោគជ័យ!")
    return redirect('scholarship_type_list')


# ----------------- STUDENT STATUSES CONFIG CRUD (ADMIN ONLY) -----------------

@login_required
@role_required(['ADMIN'])
def student_status_list(request):
    """Lists all configurable Academic Statuses for Admin with student counts & behavior configs"""
    from .forms import StudentStatusConfigForm
    StudentStatusConfig.ensure_default_statuses()
    statuses = StudentStatusConfig.objects.all().order_by('order', 'id')
    form = StudentStatusConfigForm()

    status_stats = []
    for sc in statuses:
        cnt = Student.objects.filter(status=sc.code).count()
        status_stats.append({
            'obj': sc,
            'student_count': cnt
        })

    return render(request, 'students/student_status_list.html', {
        'statuses': status_stats,
        'form': form
    })


@login_required
@role_required(['ADMIN'])
def student_status_save(request, pk=None):
    """Create or update an Academic Status"""
    from .forms import StudentStatusConfigForm
    sc = get_object_or_404(StudentStatusConfig, pk=pk) if pk else None
    if request.method == 'POST':
        form = StudentStatusConfigForm(request.POST, instance=sc)
        if form.is_valid():
            saved_sc = form.save()
            messages.success(request, f"🎉 បានរក្សាទុកស្ថានភាពសិក្សា '{saved_sc.name}' ដោយជោគជ័យ!")
        else:
            for f, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"កំហុស [{f}]: {e}")
    return redirect('student_status_list')


@login_required
@role_required(['ADMIN'])
def student_status_delete(request, pk):
    """Delete an Academic Status if not a system default and not in active use"""
    sc = get_object_or_404(StudentStatusConfig, pk=pk)
    if request.method == 'POST':
        if sc.is_system_default:
            messages.error(request, f"⚠️ មិនអាចលុបស្ថានភាពគោលរបស់ប្រព័ន្ធ '{sc.name}' បានទេ! (លោកអ្នកអាចប្តូរឈ្មោះ ឬពណ៌បាន)")
            return redirect('student_status_list')

        student_count = Student.objects.filter(status=sc.code).count()
        if student_count > 0:
            messages.warning(request, f"⚠️ មិនអាចលុប '{sc.name}' បានទេ ដោយសារមានសិស្សចំនួន {student_count} នាក់កំពុងស្ថិតក្នុងស្ថានភាពនេះ!")
        else:
            name = sc.name
            sc.delete()
            messages.success(request, f"🗑️ បានលុបស្ថានភាពសិក្សា '{name}' ដោយជោគជ័យ!")
    return redirect('student_status_list')


@login_required
@role_required(['ADMIN'])
def api_quick_set_student_status(request, pk):
    """
    1-Click Quick Status update endpoint for a student from Student List or Detail.
    """
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status', '').strip()
        fee_end_month = request.POST.get('fee_end_month')

        if new_status:
            student.status = new_status
            if fee_end_month and str(fee_end_month).isdigit():
                student.fee_end_month = int(fee_end_month)
            elif fee_end_month == 'none' or fee_end_month == '':
                student.fee_end_month = None
            student.save(update_fields=['status', 'fee_end_month', 'updated_at'])

            msg = f"🎉 បានប្តូរស្ថានភាពសិស្ស «{student.khmer_name}» ទៅជា៖ {student.get_status_display()} ដោយជោគជ័យ!"
            from django.http import JsonResponse
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json':
                return JsonResponse({
                    'success': True,
                    'status': student.status,
                    'status_display': student.get_status_display(),
                    'badge_color': student.status_badge_color,
                    'message': msg
                })
            messages.success(request, msg)

    redirect_url = request.META.get('HTTP_REFERER') or 'student_list'
    return redirect(redirect_url)



def public_enroll_success(request, pk):
    """
    Public registration receipt & credentials confirmation page with print/PDF options.
    """
    student = get_object_or_404(Student.objects.select_related('classroom', 'academic_year', 'user'), pk=pk)
    username = student.user.username if student.user else student.student_id.lower().replace('-', '_')
    
    return render(request, 'students/public_enroll_success.html', {
        'student': student,
        'username': username,
        'initial_password': 'p123456',
    })


def enrollment_qr_code(request):
    """
    Printable and shareable QR Code Poster for school admission marketing/banners.
    Supports:
    1. General (All Grades)
    2. Grade-Specific (ថ្នាក់ទី ៧, ៨, ៩, ១០, ១១-SCI, ១១-SOC, ១២-SCI, ១២-SOC)
    3. Classroom-Specific (7A, 7B, 10A, 11-SCI...)
    """
    import socket
    from urllib.parse import quote as url_quote
    
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    current_year = active_year or AcademicYear.objects.filter(is_current=True).first()
    classrooms = Classroom.objects.filter(academic_year=current_year).select_related('academic_year').order_by('grade_level', 'code') if current_year else Classroom.objects.select_related('academic_year').order_by('grade_level', 'code')

    
    local_ip = '127.0.0.1'
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        pass

    host = request.get_host()
    port = request.get_port()
    
    if host.startswith('127.0.0.1') or host.startswith('localhost'):
        port_suffix = f":{port}" if port and str(port) not in ['80', '443'] else ""
        base_url = f"http://{local_ip}{port_suffix}/students/enroll/online/"
    else:
        base_url = request.build_absolute_uri('/students/enroll/online/')

    # 1. Build Grade Level Data (ថ្នាក់ទី ៧, ៨, ៩, ១០, ១១-វិទ្យាសាស្ត្រ, ១១-សង្គម...)
    grade_data = []
    grade_tracks = classrooms.values('grade_level', 'track').distinct().order_by('grade_level', 'track')
    for gt in grade_tracks:
        g_num = gt['grade_level']
        g_track = gt['track']
        track_name = ""
        if g_track == 'SCIENCE':
            track_name = " វិទ្យាសាស្ត្រ (Science)"
        elif g_track == 'SOCIAL':
            track_name = " វិទ្យាសាស្ត្រសង្គម (Social)"
        
        name_kh = f"ថ្នាក់ទី {g_num}{track_name}"
        direct_url = f"{base_url}?grade={g_num}" + (f"&track={g_track}" if g_track != 'GENERAL' else "")
        
        classes_in_grade = classrooms.filter(grade_level=g_num, track=g_track)
        class_codes = ", ".join(classes_in_grade.values_list('code', flat=True))

        grade_data.append({
            'grade_level': g_num,
            'track': g_track,
            'name': name_kh,
            'class_codes': class_codes,
            'classes_count': classes_in_grade.count(),
            'url': direct_url,
            'qr_src': f"https://api.qrserver.com/v1/create-qr-code/?size=260x260&margin=10&data={url_quote(direct_url)}",
        })

    # 2. Build Specific Classroom Data (7A, 7B, 10A...)
    classroom_data = []
    for c in classrooms:
        direct_url = f"{base_url}?classroom={c.id}"
        classroom_data.append({
            'id': c.id,
            'code': c.code,
            'name': c.name,
            'grade_level': c.grade_level,
            'track': c.get_track_display(),
            'room_number': c.room_number or '',
            'url': direct_url,
            'qr_src': f"https://api.qrserver.com/v1/create-qr-code/?size=260x260&margin=10&data={url_quote(direct_url)}",
        })

    return render(request, 'students/enrollment_qr_modal.html', {
        'public_url': base_url,
        'base_url': base_url,
        'local_ip': local_ip,
        'current_year': current_year,
        'classrooms': classrooms,
        'grade_data': grade_data,
        'classroom_data': classroom_data,
    })


@login_required
def student_detail(request, pk):
    student = get_object_or_404(Student.objects.select_related('classroom', 'academic_year', 'user'), pk=pk)
    
    # Check permission for student role: can only view own profile
    if request.user.role == User.Role.STUDENT:
        if not hasattr(request.user, 'student_profile') or request.user.student_profile.id != student.id:
            messages.error(request, "លោកអ្នកអាចចូលមើលបានតែប្រវត្តិរូបផ្ទាល់ខ្លួនប៉ុណ្ណោះ!")
            return redirect('student_dashboard')

    attendances = StudentAttendance.objects.filter(student=student).order_by('-date')[:30]
    grades = Grade.objects.filter(student=student).select_related('subject', 'exam_term')
    invoices = Invoice.objects.filter(student=student).select_related('fee_category', 'academic_year').order_by('-created_at')
    borrowings = BookBorrowing.objects.filter(student=student).select_related('book').order_by('-borrow_date')

    # Attendance stats
    total_att = StudentAttendance.objects.filter(student=student).count()
    present_att = StudentAttendance.objects.filter(student=student, status='PRESENT').count()
    attendance_rate = round((present_att / total_att) * 100, 1) if total_att > 0 else 100.0

    return render(request, 'students/student_detail.html', {
        'student': student,
        'attendances': attendances,
        'grades': grades,
        'invoices': invoices,
        'borrowings': borrowings,
        'attendance_rate': attendance_rate,
        'total_att': total_att,
    })


@login_required
@role_required(['ADMIN'])
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST, request.FILES, instance=student, academic_year=student.academic_year)
        if form.is_valid():
            with transaction.atomic():
                updated_student = form.save(commit=False)
                updated_student.enrollment_data = _extract_grade_options(request, updated_student.classroom, existing_data=student.enrollment_data)
                updated_student.save()
            messages.success(request, f"បានកែប្រែព័ត៌មានសិស្ស {student.khmer_name} ជោគជ័យ!")
            return redirect('student_detail', pk=student.pk)
    else:
        form = StudentEnrollmentForm(instance=student, academic_year=student.academic_year)

    return render(request, 'students/student_form.html', {
        'form': form,
        'current_year': student.academic_year,
        'title': f'កែប្រែព័ត៌មានសិស្ស {student.khmer_name}',
        'student': student
    })


@login_required
def student_id_card(request, pk):
    student = get_object_or_404(Student, pk=pk)
    return render(request, 'students/student_id_card.html', {'student': student})


# -------------------------------------------------------------
# BULK STUDENT IMPORT & TEMPLATE DOWNLOAD HELPERS & VIEWS
# -------------------------------------------------------------

def _clean_str(val):
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        return str(int(val)).strip()
    val_str = str(val).strip()
    if val_str.lower() in ['none', 'nan', 'null']:
        return ''
    return val_str


def _normalize_header(header):
    if not header:
        return ''
    # Remove text in parentheses, asterisks, brackets, colons, trim and lowercase
    cleaned = re.sub(r'[\(\[\{].*?[\)\]\}]', '', str(header))
    cleaned = cleaned.replace('*', '').replace(':', '').strip().lower()
    cleaned = re.sub(r'\s+', ' ', cleaned)
    
    mapping = {
        # Khmer Name
        'ឈ្មោះខ្មែរ': 'khmer_name',
        'ឈ្មោះជាភាសាខ្មែរ': 'khmer_name',
        'ឈ្មោះសិស្ស': 'khmer_name',
        'ឈ្មោះ': 'khmer_name',
        'khmer_name': 'khmer_name',
        'khmer name': 'khmer_name',
        'name_kh': 'khmer_name',
        'student_name': 'khmer_name',
        'name': 'khmer_name',
        
        # Latin Name
        'ឈ្មោះឡាតាំង': 'latin_name',
        'ឈ្មោះជាអក្សរឡាតាំង': 'latin_name',
        'ឈ្មោះអង់គ្លេស': 'latin_name',
        'latin_name': 'latin_name',
        'latin name': 'latin_name',
        'english_name': 'latin_name',
        'name_en': 'latin_name',
        'full_name_en': 'latin_name',

        # Gender
        'ភេទ': 'gender',
        'gender': 'gender',
        'sex': 'gender',

        # Date of Birth
        'ថ្ងៃខែឆ្នាំកំណើត': 'date_of_birth',
        'ថ្ងៃកំណើត': 'date_of_birth',
        'កាលបរិច្ឆេទកំណើត': 'date_of_birth',
        'date_of_birth': 'date_of_birth',
        'date of birth': 'date_of_birth',
        'dob': 'date_of_birth',
        'birth_date': 'date_of_birth',
        'birthdate': 'date_of_birth',

        # Place of Birth
        'ទីកន្លែងកំណើត': 'place_of_birth',
        'ទីកន្លែងកើត': 'place_of_birth',
        'place_of_birth': 'place_of_birth',
        'pob': 'place_of_birth',

        # Current Address
        'អាសយដ្ឋានបច្ចុប្បន្ន': 'current_address',
        'អាសយដ្ឋាន': 'current_address',
        'current_address': 'current_address',
        'address': 'current_address',

        # Phone
        'លេខទូរស័ព្ទ': 'phone',
        'លេខទូរស័ព្ទសិស្ស': 'phone',
        'ទូរស័ព្ទ': 'phone',
        'phone': 'phone',
        'phone_number': 'phone',
        'student_phone': 'phone',
        'tel': 'phone',

        # Classroom
        'ថ្នាក់': 'classroom',
        'ថ្នាក់រៀន': 'classroom',
        'កូដថ្នាក់': 'classroom',
        'ឈ្មោះថ្នាក់': 'classroom',
        'classroom': 'classroom',
        'class': 'classroom',
        'class_code': 'classroom',
        'grade': 'classroom',

        # Scholarship
        'ប្រភេទកម្រៃ': 'scholarship_type',
        'ប្រភេទកម្រៃសិក្សា': 'scholarship_type',
        'អាហារូបករណ៍': 'scholarship_type',
        'scholarship': 'scholarship_type',
        'scholarship_type': 'scholarship_type',
        'fee_type': 'scholarship_type',

        # Father Info
        'ឈ្មោះឪពុក': 'father_name',
        'ឪពុក': 'father_name',
        'father_name': 'father_name',
        'father': 'father_name',
        'លេខទូរស័ព្ទឪពុក': 'father_phone',
        'ទូរស័ព្ទឪពុក': 'father_phone',
        'father_phone': 'father_phone',
        'មុខរបរឪពុក': 'father_job',
        'father_job': 'father_job',
        'father_occupation': 'father_job',

        # Mother Info
        'ឈ្មោះម្តាយ': 'mother_name',
        'ម្តាយ': 'mother_name',
        'mother_name': 'mother_name',
        'mother': 'mother_name',
        'លេខទូរស័ព្ទម្តាយ': 'mother_phone',
        'ទូរស័ព្ទម្តាយ': 'mother_phone',
        'mother_phone': 'mother_phone',
        'មុខរបរម្តាយ': 'mother_job',
        'mother_job': 'mother_job',
        'mother_occupation': 'mother_job',

        # Guardian & Emergency
        'ឈ្មោះអាណាព្យាបាល': 'guardian_name',
        'អាណាព្យាបាល': 'guardian_name',
        'guardian_name': 'guardian_name',
        'guardian': 'guardian_name',
        'លេខទាក់ទងបន្ទាន់': 'emergency_phone',
        'ទូរស័ព្ទបន្ទាន់': 'emergency_phone',
        'emergency_phone': 'emergency_phone',
        'telegram_chat_id': 'telegram_chat_id',
        'telegram': 'telegram_chat_id',
        'telegram_id': 'telegram_chat_id',

        # Student ID
        'student_id': 'student_id',
        'កូដសិស្ស': 'student_id',
        'អត្តលេខ': 'student_id',
        'id': 'student_id',
    }
    return mapping.get(cleaned, cleaned.replace(' ', '_'))


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    
    val_str = str(val).strip()
    if not val_str:
        return None
    
    # Try openpyxl serial float or int
    try:
        if isinstance(val, (int, float)):
            return openpyxl.utils.datetime.from_excel(val).date()
    except Exception:
        pass

    formats = [
        '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
        '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y', '%m-%d-%Y',
        '%d/%m/%y', '%d-%m-%y', '%Y%m%d'
    ]
    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return None


def _parse_gender(val):
    if not val:
        return Student.Gender.MALE
    val_str = str(val).strip().lower()
    if val_str in ['ស្រី', 'f', 'female', 'girl', 'woman', '2']:
        return Student.Gender.FEMALE
    return Student.Gender.MALE


def _parse_scholarship(val):
    if not val:
        return Student.ScholarshipType.FULL_PAY
    val_str = str(val).strip().lower()
    if '50' in val_str:
        return Student.ScholarshipType.SCHOLARSHIP_50
    elif '100' in val_str or 'ឥតគិតថ្លៃ' in val_str or 'free' in val_str or 'ពេញ' in val_str and 'អាហារូបករណ៍' in val_str:
        return Student.ScholarshipType.SCHOLARSHIP_100
    elif 'រំលស់' in val_str or 'installment' in val_str:
        return Student.ScholarshipType.INSTALLMENT
    return Student.ScholarshipType.FULL_PAY


def _find_classroom(val, academic_year=None):
    if not val:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None

    qs = Classroom.objects.all()
    if academic_year:
        qs_year = qs.filter(academic_year=academic_year)
    else:
        qs_year = qs

    # 1. Exact code match (e.g. "7A", "10-SCI")
    c = qs_year.filter(code__iexact=val_str).first()
    if c:
        return c
    # 2. Exact name match (e.g. "ថ្នាក់ទី៧A")
    c = qs_year.filter(name__iexact=val_str).first()
    if c:
        return c
    # 3. Contains code or name
    c = qs_year.filter(code__icontains=val_str).first()
    if c:
        return c
    c = qs_year.filter(name__icontains=val_str).first()
    if c:
        return c

    # Fallback to all years if not found in current year
    if academic_year:
        return qs.filter(Q(code__iexact=val_str) | Q(name__iexact=val_str) | Q(name__icontains=val_str)).first()
    return None


@login_required
@role_required(['ADMIN'])
def student_import(request):
    current_year = AcademicYear.objects.filter(is_current=True).first()
    academic_years = AcademicYear.objects.all()
    classrooms = Classroom.objects.select_related('academic_year').all()

    results = None

    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        selected_year_id = request.POST.get('academic_year')
        target_year = AcademicYear.objects.filter(id=selected_year_id).first() if selected_year_id else current_year

        raw_rows = []
        is_excel = file_name.endswith('.xlsx') or file_name.endswith('.xls')
        is_csv = file_name.endswith('.csv')

        if not (is_excel or is_csv):
            messages.error(request, "សូមជ្រើសរើសឯកសារទម្រង់ Excel (.xlsx, .xls) ឬ CSV (.csv) ប៉ុណ្ណោះ!")
            return redirect('student_import')

        try:
            if is_excel:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                sheet = wb.active
                iter_rows = list(sheet.iter_rows(values_only=True))
                if iter_rows:
                    headers = [_normalize_header(h) for h in iter_rows[0]]
                    for r in iter_rows[1:]:
                        if any(cell is not None and str(cell).strip() != '' for cell in r):
                            row_dict = {}
                            for idx, val in enumerate(r):
                                if idx < len(headers) and headers[idx]:
                                    row_dict[headers[idx]] = _clean_str(val) if not isinstance(val, (datetime, date)) else val
                            raw_rows.append(row_dict)
            else:
                # CSV processing
                decoded_file = uploaded_file.read().decode('utf-8-sig', errors='replace')
                reader = csv.reader(io.StringIO(decoded_file))
                header_row = next(reader, None)
                if header_row:
                    headers = [_normalize_header(h) for h in header_row]
                    for r in reader:
                        if any(cell.strip() != '' for cell in r):
                            row_dict = {}
                            for idx, val in enumerate(r):
                                if idx < len(headers) and headers[idx]:
                                    row_dict[headers[idx]] = val.strip()
                            raw_rows.append(row_dict)

        except Exception as e:
            messages.error(request, f"មានបញ្ហាក្នុងការអានឯកសារ៖ {str(e)}")
            return redirect('student_import')

        if not raw_rows:
            messages.warning(request, "ឯកសារដែលបាន Upload មិនមានទិន្នន័យសិស្សទេ!")
            return redirect('student_import')

        # Process each row
        success_count = 0
        skipped_count = 0
        error_list = []
        imported_students = []
        seen_import_ids = set()

        for idx, row in enumerate(raw_rows, start=2):
            khmer_name = row.get('khmer_name', '')
            if not khmer_name:
                error_list.append({
                    'row': idx,
                    'name': 'មិនស្គាល់',
                    'error': 'ខ្វះឈ្មោះខ្មែរ (Khmer Name is required)'
                })
                skipped_count += 1
                continue

            latin_name = row.get('latin_name', '')
            gender = _parse_gender(row.get('gender'))
            dob = _parse_date(row.get('date_of_birth'))
            if not dob:
                # Default to 15 years ago if missing
                dob = date(datetime.now().year - 15, 1, 1)

            pob = row.get('place_of_birth', '')
            address = row.get('current_address', '')
            phone = row.get('phone', '')
            class_input = row.get('classroom', '')
            classroom = _find_classroom(class_input, target_year)
            scholarship_type = _parse_scholarship(row.get('scholarship_type'))

            father_name = row.get('father_name', '')
            father_phone = row.get('father_phone', '')
            father_job = row.get('father_job', '')
            mother_name = row.get('mother_name', '')
            mother_phone = row.get('mother_phone', '')
            mother_job = row.get('mother_job', '')
            guardian_name = row.get('guardian_name', '')
            emergency_phone = row.get('emergency_phone', '')
            telegram_chat_id = row.get('telegram_chat_id', '')
            student_id_custom = str(row.get('student_id', '')).strip()

            # Check duplicate student_id if provided in Excel
            if student_id_custom:
                if student_id_custom.lower() in seen_import_ids or Student.objects.filter(student_id__iexact=student_id_custom).exists():
                    existing_st = Student.objects.filter(student_id__iexact=student_id_custom).first()
                    err_msg = f"អត្តលេខសិស្ស '{student_id_custom}' ស្ទួនជាមួយសិស្សដែលមានរួចហើយក្នុងប្រព័ន្ធ"
                    if existing_st:
                        err_msg += f" ({existing_st.khmer_name} - {existing_st.classroom.name if existing_st.classroom else 'គ្មានថ្នាក់'})"
                    error_list.append({
                        'row': idx,
                        'name': khmer_name,
                        'error': err_msg
                    })
                    skipped_count += 1
                    continue
                seen_import_ids.add(student_id_custom.lower())

            try:
                with transaction.atomic():
                    student = Student(
                        student_id=student_id_custom if student_id_custom else '',
                        khmer_name=khmer_name,
                        latin_name=latin_name,
                        gender=gender,
                        date_of_birth=dob,
                        place_of_birth=pob,
                        current_address=address,
                        phone=phone,
                        classroom=classroom,
                        academic_year=target_year,
                        scholarship_type=scholarship_type,
                        father_name=father_name,
                        father_phone=father_phone,
                        father_job=father_job,
                        mother_name=mother_name,
                        mother_phone=mother_phone,
                        mother_job=mother_job,
                        guardian_name=guardian_name,
                        emergency_phone=emergency_phone,
                        telegram_chat_id=telegram_chat_id,
                    )
                    student.save()

                    # Create or link user account for login
                    username = student.student_id.lower().replace('-', '_')
                    user = User.objects.filter(username=username).first()
                    if not user:
                        user = User.objects.create_user(
                            username=username,
                            password='p123456',
                            role=User.Role.STUDENT,
                            khmer_name=student.khmer_name,
                            latin_name=student.latin_name,
                            phone=student.phone or student.father_phone or ''
                        )
                    student.user = user
                    student.save(update_fields=['user'])

                    success_count += 1
                    imported_students.append({
                        'id': student.id,
                        'student_id': student.student_id,
                        'khmer_name': student.khmer_name,
                        'latin_name': student.latin_name,
                        'classroom': student.classroom.name if student.classroom else 'គ្មានថ្នាក់',
                        'gender': student.get_gender_display(),
                    })
            except Exception as ex:
                error_list.append({
                    'row': idx,
                    'name': khmer_name,
                    'error': str(ex)
                })
                skipped_count += 1

        results = {
            'total': len(raw_rows),
            'success_count': success_count,
            'skipped_count': skipped_count,
            'errors': error_list,
            'imported_students': imported_students[:50],  # show first 50
        }

        if success_count > 0:
            messages.success(request, f"🎉 ជោគជ័យ! បាន Import សិស្សថ្មីចំនួន {success_count} នាក់ចូលក្នុងប្រព័ន្ធ។")
        if skipped_count > 0:
            messages.warning(request, f"⚠️ មានសិស្សចំនួន {skipped_count} នាក់មិនអាចបញ្ចូលបាន សូមពិនិត្យបញ្ជីកំហុសខាងក្រោម។")

    return render(request, 'students/student_import.html', {
        'academic_years': academic_years,
        'current_year': current_year,
        'classrooms': classrooms,
        'results': results,
    })


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def download_student_template_excel(request):
    """
    Generates a beautifully formatted sample Excel template for importing students
    with reference sheet listing all active classrooms.
    """
    wb = openpyxl.Workbook()
    
    # Sheet 1: Import Template
    ws1 = wb.active
    ws1.title = "Student Import Template"

    headers = [
        'កូដសិស្ស (Student ID)',
        'ឈ្មោះខ្មែរ (Khmer Name)*',
        'ឈ្មោះឡាតាំង (Latin Name)',
        'ភេទ (Gender: ប្រុស/ស្រី)*',
        'ថ្ងៃខែឆ្នាំកំណើត (DOB: DD/MM/YYYY)*',
        'ថ្នាក់រៀន (Class Code: 7A, 8B...)*',
        'លេខទូរស័ព្ទ (Phone)',
        'ទីកន្លែងកំណើត (Place of Birth)',
        'អាសយដ្ឋានបច្ចុប្បន្ន (Address)',
        'ឈ្មោះឪពុក (Father Name)',
        'ទូរស័ព្ទឪពុក (Father Phone)',
        'មុខរបរឪពុក (Father Job)',
        'ឈ្មោះម្តាយ (Mother Name)',
        'ទូរស័ព្ទម្តាយ (Mother Phone)',
        'មុខរបរម្តាយ (Mother Job)',
        'ប្រភេទកម្រៃ (Full Pay / 50% / 100% / Installment)'
    ]
    ws1.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sample rows
    sample_data = [
        ['', 'សេង វណ្ណា', 'SENG VANNA', 'ប្រុស', '15/05/2009', '7A', '012 345 678', 'ភ្នំពេញ', 'ផ្ទះលេខ ១២ ផ្លូវ 2004 សង្កាត់ទឹកថ្លា', 'សេង ចាន់ថន', '012 888 777', 'អាជីវករ', 'អ៊ុក គឹមហុង', '098 777 666', 'មេផ្ទះ', 'Full Pay'],
        ['', 'កែវ មុនីរ័ត្ន', 'KEO MONIROTH', 'ស្រី', '18/04/2008', '8A', '096 900 913', 'កណ្ដាល', 'ភូមិព្រែកតាពៅ សង្កាត់ដើមមៀន', 'កែវ សុខឿន', '012 900 913', 'គ្រូបង្រៀន', 'ឡុង ចរិយា', '012 900 914', 'គណនេយ្យករ', '50%'],
        ['', 'យិន ច័ន្ទរិទ្ធ', 'YIN CHANRITH', 'ប្រុស', '11/03/2007', '10-SCI', '012 900 914', 'សៀមរាប', 'ភូមិមណ្ឌល១ សង្កាត់ស្វាយដង្គំ', 'យិន សំអាត', '012 900 914', 'វិស្វករ', 'ចាន់ ផល្លា', '012 900 915', 'មន្ត្រីរាជការ', '100%'],
        ['', 'សួស ចរិយា', 'SUOS CHORIYA', 'ស្រី', '29/09/2007', '11-SOC', '012 900 915', 'បាត់ដំបង', 'ភូមិរំចេក៤ សង្កាត់រតនៈ', 'សួស ផល', '012 900 915', 'កសិករ', 'ស៊ុន សុភា', '012 900 916', 'កសិករ', 'Full Pay'],
    ]

    for row in sample_data:
        ws1.append(row)

    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 16)
    ws1.row_dimensions[1].height = 28

    # Sheet 2: Available Classrooms Reference
    ws2 = wb.create_sheet(title="Classrooms Reference")
    ws2.append(['កូដថ្នាក់ (Class Code)', 'ឈ្មោះថ្នាក់ (Class Name)', 'កម្រិត (Grade)', 'ជំនាញ (Track)', 'ឆ្នាំសិក្សា (Academic Year)'])
    
    ws2_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    for col_idx in range(1, 6):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = ws2_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for c in Classroom.objects.select_related('academic_year').all():
        ws2.append([
            c.code,
            c.name,
            f"ថ្នាក់ទី{c.grade_level}",
            c.get_track_display(),
            c.academic_year.name if c.academic_year else ''
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 18)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required(['ADMIN', 'ACCOUNTANT'])
def download_student_template_csv(request):
    """
    Generates a UTF-8 with BOM CSV sample template for importing students
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'កូដសិស្ស (Student ID)',
        'ឈ្មោះខ្មែរ (Khmer Name)',
        'ឈ្មោះឡាតាំង (Latin Name)',
        'ភេទ (Gender)',
        'ថ្ងៃខែឆ្នាំកំណើត (DOB: DD/MM/YYYY)',
        'ថ្នាក់រៀន (Class Code)',
        'លេខទូរស័ព្ទ (Phone)',
        'ទីកន្លែងកំណើត (POB)',
        'អាសយដ្ឋានបច្ចុប្បន្ន (Address)',
        'ឈ្មោះឪពុក (Father Name)',
        'ទូរស័ព្ទឪពុក (Father Phone)',
        'មុខរបរឪពុក (Father Job)',
        'ឈ្មោះម្តាយ (Mother Name)',
        'ទូរស័ព្ទម្តាយ (Mother Phone)',
        'មុខរបរម្តាយ (Mother Job)',
        'ប្រភេទកម្រៃ (Scholarship Type)'
    ])
    
    writer.writerow(['', 'សេង វណ្ណា', 'SENG VANNA', 'ប្រុស', '15/05/2009', '7A', '012 345 678', 'ភ្នំពេញ', 'ផ្ទះលេខ ១២ ផ្លូវ 2004 សង្កាត់ទឹកថ្លា', 'សេង ចាន់ថន', '012 888 777', 'អាជីវករ', 'អ៊ុក គឹមហុង', '098 777 666', 'មេផ្ទះ', 'Full Pay'])
    writer.writerow(['', 'កែវ មុនីរ័ត្ន', 'KEO MONIROTH', 'ស្រី', '18/04/2008', '8A', '096 900 913', 'កណ្ដាល', 'ភូមិព្រែកតាពៅ សង្កាត់ដើមមៀន', 'កែវ សុខឿន', '012 900 913', 'គ្រូបង្រៀន', 'ឡុង ចរិយា', '012 900 914', 'គណនេយ្យករ', '50%'])
    writer.writerow(['', 'យិន ច័ន្ទរិទ្ធ', 'YIN CHANRITH', 'ប្រុស', '11/03/2007', '10-SCI', '012 900 914', 'សៀមរាប', 'ភូមិមណ្ឌល១ សង្កាត់ស្វាយដង្គំ', 'យិន សំអាត', '012 900 914', 'វិស្វករ', 'ចាន់ ផល្លា', '012 900 915', 'មន្ត្រីរាជការ', '100%'])
    
    return response


# ==========================================
# EXAM SUSPENSION & EXCLUSION APIS
# ==========================================

@login_required
@role_required(['ADMIN'])
def api_set_student_exam_status(request, pk):
    """
    POST/AJAX endpoint to set or toggle exam exclusion for a student directly from the Student List.
    """
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        is_suspended_raw = request.POST.get('is_exam_suspended')
        if is_suspended_raw is not None:
            is_suspended = is_suspended_raw in ['true', 'True', '1', 'on']
        else:
            is_suspended = not student.is_exam_suspended

        reason = request.POST.get('exam_suspension_reason', Student.ExamExclusionReason.DISCIPLINARY)
        notes = request.POST.get('exam_suspension_notes', '').strip()

        student.is_exam_suspended = is_suspended
        if is_suspended:
            student.exam_suspension_reason = reason
            if notes:
                student.exam_suspension_notes = notes
        else:
            student.exam_suspension_notes = ''
        student.save(update_fields=['is_exam_suspended', 'exam_suspension_reason', 'exam_suspension_notes', 'updated_at'])

        # Sync with ExamStudentExclusion model for unified examinations reporting
        try:
            from apps.examinations.models import ExamStudentExclusion
            if is_suspended:
                ExamStudentExclusion.objects.update_or_create(
                    student=student,
                    academic_year=student.academic_year or (student.classroom.academic_year if student.classroom else None) or AcademicYear.objects.first(),
                    exam_term=None,
                    month=None,
                    defaults={
                        'reason': reason,
                        'notes': notes or f"កំណត់ដកសិទ្ធិពីបញ្ជីសិស្ស (ដោយ {request.user.get_full_name() or request.user.username})",
                        'is_active': True,
                        'excluded_by': request.user
                    }
                )
            else:
                ExamStudentExclusion.objects.filter(student=student).update(is_active=False)
        except Exception:
            pass

        action_label = "🔴 ដកសិទ្ធិប្រឡង (Excluded)" if is_suspended else "🟢 មានសិទ្ធិប្រឡង (Eligible)"
        msg = f"🎉 បានកំណត់សិទ្ធិប្រឡងរបស់សិស្ស «{student.khmer_name}» ទៅជា៖ {action_label} ដោយជោគជ័យ!"
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json':
            return JsonResponse({
                'success': True,
                'is_exam_suspended': student.is_exam_suspended,
                'reason_display': student.get_exam_suspension_reason_display(),
                'notes': student.exam_suspension_notes,
                'message': msg
            })
        messages.success(request, msg)
        return redirect('student_list')

    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)


@login_required
@role_required(['ADMIN'])
def api_batch_set_student_exam_status(request):
    """
    POST endpoint to batch set exam exclusion for multiple selected students from the Student List.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    raw_ids = request.POST.getlist('student_ids')
    if not raw_ids:
        raw_val = request.POST.get('student_ids', '')
        if raw_val:
            raw_ids = [raw_val]

    student_ids = []
    for item in raw_ids:
        for piece in str(item).split(','):
            if piece.strip().isdigit():
                student_ids.append(int(piece.strip()))

    if not student_ids:
        messages.error(request, "សូមជ្រើសរើសសិស្សយ៉ាងហោចណាស់ម្នាក់!")
        return redirect('student_list')

    action = request.POST.get('batch_action', 'suspend')  # 'suspend' or 'allow'
    reason = request.POST.get('exam_suspension_reason', Student.ExamExclusionReason.DISCIPLINARY)
    notes = request.POST.get('exam_suspension_notes', '').strip()

    is_suspended = (action == 'suspend')

    updated_count = Student.objects.filter(id__in=student_ids).update(
        is_exam_suspended=is_suspended,
        exam_suspension_reason=reason if is_suspended else Student.ExamExclusionReason.DISCIPLINARY,
        exam_suspension_notes=notes if is_suspended else ''
    )

    # Sync with ExamStudentExclusion
    try:
        from apps.examinations.models import ExamStudentExclusion
        students = Student.objects.filter(id__in=student_ids)
        for stu in students:
            if is_suspended:
                ExamStudentExclusion.objects.update_or_create(
                    student=stu,
                    academic_year=stu.academic_year or (stu.classroom.academic_year if stu.classroom else None) or AcademicYear.objects.first(),
                    exam_term=None,
                    month=None,
                    defaults={
                        'reason': reason,
                        'notes': notes or f"កំណត់ដកសិទ្ធិជាក្រុមពីបញ្ជីសិស្ស (ដោយ {request.user.get_full_name() or request.user.username})",
                        'is_active': True,
                        'excluded_by': request.user
                    }
                )
            else:
                ExamStudentExclusion.objects.filter(student=stu).update(is_active=False)
    except Exception:
        pass

    action_text = "🔴 ដកសិទ្ធិពីការប្រឡង" if is_suspended else "🟢 អនុញ្ញាតឱ្យចូលរួមការប្រឡងវិញ"
    msg = f"🎉 បានធ្វើបច្ចុប្បន្នភាព {action_text} ចំពោះសិស្សចំនួន {updated_count} នាក់ដោយជោគជ័យ!"

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json':
        return JsonResponse({'success': True, 'updated_count': updated_count, 'message': msg})

    messages.success(request, msg)
    return redirect('student_list')


# =========================================================================
# Academic Year Safe Student Purge & Historical Archival System
# =========================================================================

def _generate_archive_excel_bytes(payload, year_name):
    wb = openpyxl.Workbook()
    header_font = Font(name='Hanuman', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    title_font = Font(name='Hanuman', size=14, bold=True, color='1E3A8A')
    body_font = Font(name='Hanuman', size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')

    # Sheet 1: Students Directory
    ws_students = wb.active
    ws_students.title = "1. បញ្ជីសិស្ស"
    ws_students.views.sheetView[0].showGridLines = True

    ws_students.merge_cells('A1:J1')
    title_cell = ws_students.cell(row=1, column=1, value=f"ប័ណ្ណសារបញ្ជីសិស្ស - ឆ្នាំសិក្សា {year_name}")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_students.row_dimensions[1].height = 35

    headers_students = ['ល.រ', 'អត្តលេខ (ID)', 'ឈ្មោះខ្មែរ', 'ឈ្មោះឡាតាំង', 'ភេទ', 'ថ្ងៃខែឆ្នាំកំណើត', 'ថ្នាក់រៀន', 'លេខទូរស័ព្ទ', 'អាហារូបករណ៍', 'ស្ថានភាព']
    for col_idx, h in enumerate(headers_students, 1):
        c = ws_students.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border
    ws_students.row_dimensions[3].height = 25

    row_idx = 4
    for s_idx, s in enumerate(payload.get('students', []), 1):
        row_vals = [
            s_idx,
            s.get('student_id', ''),
            s.get('khmer_name', ''),
            s.get('latin_name', ''),
            s.get('gender_display', s.get('gender', '')),
            s.get('date_of_birth', ''),
            s.get('classroom_name', ''),
            s.get('phone', ''),
            s.get('scholarship_type', ''),
            s.get('status_display', s.get('status', ''))
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws_students.cell(row=row_idx, column=col_idx, value=val)
            c.font = body_font
            c.border = thin_border
            if col_idx in [1, 2, 5, 6, 7, 10]:
                c.alignment = center_align
            else:
                c.alignment = left_align
        ws_students.row_dimensions[row_idx].height = 22
        row_idx += 1

    for col in ws_students.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_students.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Sheet 2: Grades
    ws_grades = wb.create_sheet(title="2. តារាងពិន្ទុ")
    ws_grades.views.sheetView[0].showGridLines = True
    ws_grades.merge_cells('A1:H1')
    t_cell2 = ws_grades.cell(row=1, column=1, value=f"ប័ណ្ណសារលទ្ធផលពិន្ទុសិស្ស - ឆ្នាំសិក្សា {year_name}")
    t_cell2.font = title_font
    t_cell2.alignment = Alignment(horizontal='center', vertical='center')
    ws_grades.row_dimensions[1].height = 35

    headers_grades = ['ល.រ', 'អត្តលេខ', 'ឈ្មោះសិស្ស', 'ថ្នាក់រៀន', 'មុខវិជ្ជា', 'សម័យប្រឡង', 'ពិន្ទុទទួលបាន', 'និទ្ទេស']
    for col_idx, h in enumerate(headers_grades, 1):
        c = ws_grades.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border
    ws_grades.row_dimensions[3].height = 25

    row_idx = 4
    for g_idx, g in enumerate(payload.get('grades', []), 1):
        row_vals = [
            g_idx,
            g.get('student_id', ''),
            g.get('student_name', ''),
            g.get('classroom_name', ''),
            g.get('subject_name', ''),
            g.get('exam_term_name', ''),
            f"{g.get('score', 0)} / {g.get('max_score', 100)}",
            g.get('grade_letter', '')
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws_grades.cell(row=row_idx, column=col_idx, value=val)
            c.font = body_font
            c.border = thin_border
            if col_idx in [1, 2, 4, 7, 8]:
                c.alignment = center_align
            else:
                c.alignment = left_align
        ws_grades.row_dimensions[row_idx].height = 22
        row_idx += 1

    for col in ws_grades.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_grades.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Sheet 3: Attendances
    ws_att = wb.create_sheet(title="3. វត្តមានសិស្ស")
    ws_att.views.sheetView[0].showGridLines = True
    ws_att.merge_cells('A1:G1')
    t_cell3 = ws_att.cell(row=1, column=1, value=f"ប័ណ្ណសារវត្តមានសិស្ស - ឆ្នាំសិក្សា {year_name}")
    t_cell3.font = title_font
    t_cell3.alignment = Alignment(horizontal='center', vertical='center')
    ws_att.row_dimensions[1].height = 35

    headers_att = ['ល.រ', 'អត្តលេខ', 'ឈ្មោះសិស្ស', 'ថ្នាក់រៀន', 'កាលបរិច្ឆេទ', 'ស្ថានភាពវត្តមាន', 'សម្គាល់/ច្បាប់']
    for col_idx, h in enumerate(headers_att, 1):
        c = ws_att.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border
    ws_att.row_dimensions[3].height = 25

    row_idx = 4
    for a_idx, a in enumerate(payload.get('attendances', []), 1):
        row_vals = [
            a_idx,
            a.get('student_id', ''),
            a.get('student_name', ''),
            a.get('classroom_name', ''),
            a.get('date', ''),
            a.get('status_display', a.get('status', '')),
            a.get('remarks', '')
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws_att.cell(row=row_idx, column=col_idx, value=val)
            c.font = body_font
            c.border = thin_border
            if col_idx in [1, 2, 4, 5, 6]:
                c.alignment = center_align
            else:
                c.alignment = left_align
        ws_att.row_dimensions[row_idx].height = 22
        row_idx += 1

    for col in ws_att.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_att.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Sheet 4: Invoices
    ws_fees = wb.create_sheet(title="4. វិក្កយបត្រ")
    ws_fees.views.sheetView[0].showGridLines = True
    ws_fees.merge_cells('A1:G1')
    t_cell4 = ws_fees.cell(row=1, column=1, value=f"ប័ណ្ណសារវិក្កយបត្រសិស្ស - ឆ្នាំសិក្សា {year_name}")
    t_cell4.font = title_font
    t_cell4.alignment = Alignment(horizontal='center', vertical='center')
    ws_fees.row_dimensions[1].height = 35

    headers_fees = ['ល.រ', 'លេខវិក្កយបត្រ', 'អត្តលេខ', 'ឈ្មោះសិស្ស', 'ទឹកប្រាក់សរុប ($)', 'បានបង់ ($)', 'ស្ថានភាព']
    for col_idx, h in enumerate(headers_fees, 1):
        c = ws_fees.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border
    ws_fees.row_dimensions[3].height = 25

    row_idx = 4
    for f_idx, f in enumerate(payload.get('fees', []), 1):
        row_vals = [
            f_idx,
            f.get('invoice_number', ''),
            f.get('student_id', ''),
            f.get('student_name', ''),
            str(f.get('total_amount', '0.00')),
            str(f.get('paid_amount', '0.00')),
            f.get('status_display', f.get('status', ''))
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws_fees.cell(row=row_idx, column=col_idx, value=val)
            c.font = body_font
            c.border = thin_border
            if col_idx in [1, 2, 3, 7]:
                c.alignment = center_align
            elif col_idx in [5, 6]:
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.alignment = left_align
        ws_fees.row_dimensions[row_idx].height = 22
        row_idx += 1

    for col in ws_fees.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_fees.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@login_required
@role_required(['ADMIN'])
def api_get_academic_year_purge_preview(request):
    """
    Returns real-time counts and challenge code for an Academic Year prior to purging/archiving.
    """
    from .models import AcademicYearStudentArchive
    from django.http import JsonResponse
    import json

    year_id = request.GET.get('academic_year_id')
    if not year_id:
        return JsonResponse({'status': 'error', 'message': 'academic_year_id is required'}, status=400)

    ay = get_object_or_404(AcademicYear, id=year_id)
    students_qs = Student.objects.filter(Q(academic_year=ay) | Q(classroom__academic_year=ay)).distinct()
    students_count = students_qs.count()
    classrooms_count = ay.classrooms.count()
    grades_count = Grade.objects.filter(Q(classroom__academic_year=ay) | Q(exam_term__academic_year=ay)).count()
    attendances_count = StudentAttendance.objects.filter(classroom__academic_year=ay).count()
    fees_count = Invoice.objects.filter(Q(academic_year=ay) | Q(student__in=students_qs)).distinct().count()

    return JsonResponse({
        'status': 'success',
        'academic_year_id': ay.id,
        'academic_year_name': ay.name,
        'is_current': ay.is_current,
        'students_count': students_count,
        'classrooms_count': classrooms_count,
        'grades_count': grades_count,
        'attendances_count': attendances_count,
        'fees_count': fees_count,
        'challenge_text': ay.name,
    })


@login_required
@role_required(['ADMIN'])
def api_execute_academic_year_purge(request):
    """
    Atomically archives all student data, grades, attendances, and fees for a specified
    Academic Year and then executes either Soft Unenroll or Hard Purge with full data preservation.
    """
    from .models import AcademicYearStudentArchive
    from django.core.files.base import ContentFile
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST request required'}, status=400)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        data = request.POST

    year_id = data.get('academic_year_id')
    action_type = data.get('action_type', AcademicYearStudentArchive.ActionType.SOFT_UNENROLL)
    confirmation_text = str(data.get('confirmation_text', '')).strip()
    note = str(data.get('note', '')).strip()

    if not year_id:
        return JsonResponse({'status': 'error', 'message': 'សូមបញ្ជាក់ឆ្នាំសិក្សាដែលត្រូវសម្អាត!'}, status=400)

    ay = get_object_or_404(AcademicYear, id=year_id)

    # Security check: must match exact year name
    if confirmation_text != ay.name.strip():
        return JsonResponse({
            'status': 'error',
            'message': f'⚠️ ពាក្យផ្ទៀងផ្ទាត់មិនត្រឹមត្រូវឡើយ! សូមវាយឈ្មោះឆ្នាំសិក្សា "{ay.name}" ឱ្យបានត្រឹមត្រូវ។'
        }, status=400)

    with transaction.atomic():
        students_qs = Student.objects.filter(Q(academic_year=ay) | Q(classroom__academic_year=ay)).select_related('classroom', 'category').distinct()
        students_count = students_qs.count()
        classrooms_count = ay.classrooms.count()
        
        grades_qs = Grade.objects.filter(Q(classroom__academic_year=ay) | Q(exam_term__academic_year=ay)).select_related('student', 'subject', 'exam_term', 'classroom')
        grades_count = grades_qs.count()

        attendances_qs = StudentAttendance.objects.filter(classroom__academic_year=ay).select_related('student', 'classroom', 'subject')
        attendances_count = attendances_qs.count()

        fees_qs = Invoice.objects.filter(Q(academic_year=ay) | Q(student__in=students_qs)).select_related('student').distinct()
        fees_count = fees_qs.count()

        # 1. Build Comprehensive Snapshot Payload
        students_data = []
        for s in students_qs:
            students_data.append({
                'id': s.id,
                'student_id': s.student_id,
                'khmer_name': s.khmer_name,
                'latin_name': s.latin_name,
                'gender': s.gender,
                'gender_display': s.get_gender_display(),
                'date_of_birth': str(s.date_of_birth) if s.date_of_birth else '',
                'classroom_id': s.classroom.id if s.classroom else None,
                'classroom_name': s.classroom.name if s.classroom else '',
                'phone': s.phone or '',
                'scholarship_type': s.scholarship_type or '',
                'status': s.status,
                'status_display': s.get_status_display(),
                'enrollment_data': s.enrollment_data or {},
                'is_repeating_grade': s.is_repeating_grade,
                'is_exam_suspended': s.is_exam_suspended,
            })

        grades_data = []
        for g in grades_qs:
            grades_data.append({
                'id': g.id,
                'student_id': g.student.student_id if g.student else '',
                'student_name': g.student.khmer_name if g.student else '',
                'classroom_name': g.classroom.name if g.classroom else '',
                'subject_code': g.subject.code if g.subject else '',
                'subject_name': g.subject.name_kh if g.subject else '',
                'exam_term_name': g.exam_term.name if g.exam_term else '',
                'score': float(g.score) if g.score is not None else 0.0,
                'max_score': float(g.max_score) if g.max_score is not None else 100.0,
                'grade_letter': g.grade_letter or '',
                'remarks': g.remarks or '',
            })

        attendances_data = []
        for a in attendances_qs:
            attendances_data.append({
                'id': a.id,
                'student_id': a.student.student_id if a.student else '',
                'student_name': a.student.khmer_name if a.student else '',
                'classroom_name': a.classroom.name if a.classroom else '',
                'date': str(a.date) if a.date else '',
                'session': getattr(a, 'session', 'ALL'),
                'status': a.status,
                'status_display': a.get_status_display() if hasattr(a, 'get_status_display') else str(a.status),
                'remarks': getattr(a, 'notes', '') or getattr(a, 'remarks', '') or '',
            })

        fees_data = []
        for f in fees_qs:
            fees_data.append({
                'id': f.id,
                'invoice_number': f.invoice_number if hasattr(f, 'invoice_number') else f"INV-{f.id}",
                'student_id': f.student.student_id if f.student else '',
                'student_name': f.student.khmer_name if f.student else '',
                'total_amount': float(f.total_amount) if hasattr(f, 'total_amount') and f.total_amount else 0.0,
                'paid_amount': float(f.paid_amount) if hasattr(f, 'paid_amount') and f.paid_amount else 0.0,
                'status': f.status if hasattr(f, 'status') else 'PAID',
                'status_display': f.get_status_display() if hasattr(f, 'get_status_display') else str(getattr(f, 'status', '')),
            })

        archive_payload = {
            'academic_year_id': ay.id,
            'academic_year_name': ay.name,
            'archived_at': datetime.now().isoformat(),
            'archived_by_username': request.user.username,
            'action_type': action_type,
            'students_count': students_count,
            'classrooms_count': classrooms_count,
            'grades_count': grades_count,
            'attendances_count': attendances_count,
            'fees_count': fees_count,
            'students': students_data,
            'grades': grades_data,
            'attendances': attendances_data,
            'fees': fees_data,
            'note': note,
        }

        # 2. Generate Master Excel Archive
        excel_bytes = _generate_archive_excel_bytes(archive_payload, ay.name)
        excel_filename = f"StudentArchive_{ay.name.replace(' ', '_').replace('/', '-')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # 3. Save AcademicYearStudentArchive Record
        archive = AcademicYearStudentArchive.objects.create(
            academic_year=ay,
            academic_year_name=ay.name,
            archived_by=request.user,
            action_type=action_type,
            students_count=students_count,
            classrooms_count=classrooms_count,
            grades_count=grades_count,
            attendances_count=attendances_count,
            fees_count=fees_count,
            archive_payload=archive_payload,
            confirmation_note=note or f"សម្អាតសិស្សដោយ {request.user.get_full_name() or request.user.username}",
        )
        archive.archive_excel.save(excel_filename, ContentFile(excel_bytes), save=True)

        # 4. Perform Execution according to Selected Mode
        if action_type == AcademicYearStudentArchive.ActionType.SOFT_UNENROLL:
            # Soft Unenroll: unassign from classrooms & academic year
            students_qs.update(classroom=None, academic_year=None)
            action_desc = f"បានរក្សាទុកប័ណ្ណសារ និងដកសិស្សចំនួន {students_count} នាក់ចេញពីឆ្នាំសិក្សា {ay.name}"
        else:
            # Full Purge: delete student records belonging to this year
            # Note: Teacher attendance & other years remain 100% untouched
            students_qs.delete()
            action_desc = f"បានរក្សាទុកប័ណ្ណសារ និងលុបសិស្សចំនួន {students_count} នាក់ចេញពីប្រព័ន្ធដោយសុវត្ថិភាព"

    return JsonResponse({
        'status': 'success',
        'message': f"🎉 {action_desc} ដោយជោគជ័យ! ប័ណ្ណសារត្រូវបានរក្សាទុកក្នុងប្រព័ន្ធ។",
        'archive_id': archive.id,
        'students_count': students_count,
        'grades_count': grades_count,
        'attendances_count': attendances_count,
        'download_url': f"/students/archives/{archive.id}/download/",
    })


@login_required
@role_required(['ADMIN'])
def student_archives_list(request):
    """
    Displays list of all historical student archives with details and 1-click Excel download.
    """
    from .models import AcademicYearStudentArchive
    archives = AcademicYearStudentArchive.objects.select_related('academic_year', 'archived_by').all().order_by('-archived_at')
    return render(request, 'students/student_archives_list.html', {
        'archives': archives,
    })


@login_required
@role_required(['ADMIN'])
def download_student_archive_excel(request, pk):
    """
    Download pre-generated archive spreadsheet file.
    """
    from .models import AcademicYearStudentArchive
    from django.http import Http404, HttpResponse

    archive = get_object_or_404(AcademicYearStudentArchive, pk=pk)
    if not archive.archive_excel or not archive.archive_excel.storage.exists(archive.archive_excel.name):
        # Regenerate on the fly if file is missing
        excel_bytes = _generate_archive_excel_bytes(archive.archive_payload, archive.academic_year_name)
        response = HttpResponse(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        clean_name = f"Student_Archive_{archive.academic_year_name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{clean_name}"'
        return response

    response = HttpResponse(archive.archive_excel.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    clean_name = f"Student_Archive_{archive.academic_year_name.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{clean_name}"'
    return response


@login_required
@role_required(['ADMIN'])
def api_get_archive_json_snapshot(request, pk):
    """
    Returns JSON payload of a specific student archive snapshot.
    """
    from .models import AcademicYearStudentArchive
    from django.http import JsonResponse

    archive = get_object_or_404(AcademicYearStudentArchive, pk=pk)
    return JsonResponse({
        'status': 'success',
        'archive_id': archive.id,
        'academic_year_name': archive.academic_year_name,
        'archived_at': archive.archived_at.strftime('%d/%m/%Y %H:%M'),
        'archived_by': archive.archived_by.get_full_name() or archive.archived_by.username if archive.archived_by else 'Admin',
        'action_type': archive.get_action_type_display(),
        'students_count': archive.students_count,
        'grades_count': archive.grades_count,
        'attendances_count': archive.attendances_count,
        'payload': archive.archive_payload,
    })



