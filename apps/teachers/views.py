from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse
import json
import calendar
from datetime import datetime, date, timedelta

from decimal import Decimal
from django.utils import timezone
from apps.accounts.decorators import role_required
from apps.accounts.models import User
from .models import Teacher, TeacherAttendance, TeacherProfileUpdateCampaign
from .forms import TeacherForm
from apps.academics.models import ClassSubject, Timetable, AcademicYear
from apps.academics.utils import get_active_academic_year
from apps.finance.models import Payroll
from .utils import (
    get_teacher_daily_attendance_data,
    get_teacher_range_attendance_report,
    sync_teacher_attendance_from_student_logs
)

@login_required
def teacher_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')

    teachers = Teacher.objects.all()

    if query:
        teachers = teachers.filter(
            Q(teacher_id__icontains=query) |
            Q(khmer_name__icontains=query) |
            Q(latin_name__icontains=query) |
            Q(specialization__icontains=query) |
            Q(phone__icontains=query)
        )

    if status_filter:
        teachers = teachers.filter(status=status_filter)

    return render(request, 'teachers/teacher_list.html', {
        'teachers': teachers,
        'query': query,
        'selected_status': status_filter,
        'total_count': teachers.count()
    })


@login_required
@role_required(['ADMIN'])
def teacher_create(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            teacher = form.save()

            # Create User account for Teacher login
            username = teacher.teacher_id.lower().replace('-', '_')
            if not User.objects.filter(username=username).exists():
                user = User.objects.create_user(
                    username=username,
                    password='password123',
                    role=User.Role.TEACHER,
                    khmer_name=teacher.khmer_name,
                    latin_name=teacher.latin_name,
                    phone=teacher.phone,
                    email=teacher.email
                )
                teacher.user = user
                teacher.save()

            messages.success(request, f"🎉 បានបន្ថែមគ្រូបង្រៀន {teacher.khmer_name} (User: {username}) ដោយជោគជ័យ! Password: 'password123'")
            return redirect('teacher_detail', pk=teacher.pk)
    else:
        form = TeacherForm()

    return render(request, 'teachers/teacher_form.html', {
        'form': form,
        'title': 'បន្ថែមគ្រូបង្រៀនថ្មី / Add Teacher'
    })


@login_required
def teacher_detail(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    
    # Permission check: Teacher can only view their own profile
    if request.user.role == User.Role.TEACHER:
        if not hasattr(request.user, 'teacher_profile') or request.user.teacher_profile.id != teacher.id:
            messages.error(request, "លោកអ្នកអាចមើលបានតែប្រវត្តិរូបរបស់ខ្លួនឯងប៉ុណ្ណោះ!")
            return redirect('teacher_dashboard')

    class_subjects = ClassSubject.objects.filter(teacher=teacher).select_related('classroom', 'subject')
    timetables = Timetable.objects.filter(teacher=teacher).select_related('classroom', 'subject').order_by('day_of_week', 'start_time')
    attendances = TeacherAttendance.objects.filter(teacher=teacher).order_by('-date')[:30]
    payrolls = Payroll.objects.filter(teacher=teacher).order_by('-year', '-month')

    return render(request, 'teachers/teacher_detail.html', {
        'teacher': teacher,
        'class_subjects': class_subjects,
        'timetables': timetables,
        'attendances': attendances,
        'payrolls': payrolls,
    })


@login_required
@role_required(['ADMIN'])
def teacher_edit(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES, instance=teacher)
        if form.is_valid():
            form.save()
            messages.success(request, f"បានកែប្រែព័ត៌មានគ្រូ {teacher.khmer_name} ជោគជ័យ!")
            return redirect('teacher_detail', pk=teacher.pk)
    else:
        form = TeacherForm(instance=teacher)

    return render(request, 'teachers/teacher_form.html', {
        'form': form,
        'title': f'កែប្រែព័ត៌មានគ្រូ {teacher.khmer_name}',
        'teacher': teacher
    })


@login_required
@role_required(['ADMIN'])
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        name = teacher.khmer_name
        with transaction.atomic():
            user = teacher.user
            teacher.delete()
            if user and user.role == User.Role.TEACHER:
                user.delete()
        messages.success(request, f"🗑️ បានលុបគ្រូបង្រៀន {name} ដោយជោគជ័យ!")
        return redirect('teacher_list')
    
    return redirect('teacher_detail', pk=pk)


@login_required
@role_required(['ADMIN'])
def teacher_delete_all(request):
    if request.method == 'POST':
        with transaction.atomic():
            teachers = Teacher.objects.all()
            total_count = teachers.count()
            
            if total_count == 0:
                messages.info(request, "មិនមានទិន្នន័យគ្រូបង្រៀនសម្រាប់លុបឡើយ!")
                return redirect('teacher_list')
            
            # Find associated User accounts with role TEACHER
            user_ids = list(Teacher.objects.filter(user__isnull=False).values_list('user_id', flat=True))
            
            # Delete teachers (cascades to attendances, leave requests, punch logs, etc.)
            teachers.delete()
            
            # Delete associated User accounts
            if user_ids:
                User.objects.filter(id__in=user_ids, role=User.Role.TEACHER).delete()
            
            messages.success(request, f"🗑️ បានលុបទិន្នន័យគ្រូបង្រៀនទាំងអស់ចំនួន {total_count} នាក់ និងគណនី Login ពាក់ព័ន្ធដោយជោគជ័យ!")
            return redirect('teacher_list')
    
    return redirect('teacher_list')


@login_required
@role_required(['ADMIN', 'TEACHER'])
def teacher_attendance_view(request):
    """
    Teacher Daily Attendance Sheet (Admin & Duty Staff)
    Auto-detects whether teachers recorded student attendance for their scheduled classes.
    """
    active_year = get_active_academic_year(request)
    selected_date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = datetime.now().date()

    teachers_qs = Teacher.objects.filter(status='ACTIVE').order_by('teacher_id')
    query = request.GET.get('q', '').strip()
    if query:
        teachers_qs = teachers_qs.filter(
            Q(teacher_id__icontains=query) |
            Q(khmer_name__icontains=query) |
            Q(latin_name__icontains=query) |
            Q(specialization__icontains=query)
        )

    # Handle Auto-Sync action from student attendance logs
    if request.GET.get('action') == 'sync':
        synced_count = sync_teacher_attendance_from_student_logs(selected_date, active_year)
        messages.success(request, f"🔄 បានធ្វើសមកាលកម្ម (Sync) វត្តមានគ្រូចំនួន {synced_count} នាក់ពីការស្រង់វត្តមានសិស្សដោយជោគជ័យ!")
        return redirect(f"/teachers/attendance/?date={selected_date.strftime('%Y-%m-%d')}")

    # Handle Form POST
    if request.method == 'POST':
        saved_count = 0
        all_active_teachers = Teacher.objects.filter(status='ACTIVE')
        for teacher in all_active_teachers:
            status_val = request.POST.get(f'status_{teacher.id}')
            notes_val = request.POST.get(f'notes_{teacher.id}', '')
            if status_val:
                att, created = TeacherAttendance.objects.update_or_create(
                    teacher=teacher,
                    date=selected_date,
                    defaults={'status': status_val, 'notes': notes_val}
                )
                saved_count += 1

        messages.success(request, f"✅ បានកត់ត្រាវត្តមានគ្រូកាលបរិច្ឆេទ {selected_date.strftime('%d/%m/%Y')} ចំនួន {saved_count} នាក់ជោគជ័យ!")
        return redirect(f"/teachers/attendance/?date={selected_date.strftime('%Y-%m-%d')}")

    # Get daily attendance data evaluated from timetable & student attendance
    daily_data = get_teacher_daily_attendance_data(teachers_qs, selected_date, active_year)

    teacher_list_data = []
    for row in daily_data['rows']:
        teacher = row['teacher']
        existing_att = row['existing_att']
        
        # Suggested status based on student attendance:
        if existing_att:
            current_status = existing_att.status
            current_notes = existing_att.notes or ''
        else:
            current_status = row['daily_status'] if row['daily_status'] in ['PRESENT', 'UNEXCUSED_ABSENCE', 'EXCUSED_LEAVE'] else 'PRESENT'
            current_notes = ''
            if row['scheduled_count'] > 0 and row['unrecorded_count'] > 0:
                current_notes = f"មិនបានចុះវត្តមានសិស្ស {row['unrecorded_count']}/{row['scheduled_count']} ម៉ោង"

        teacher_list_data.append({
            'teacher': teacher,
            'status': current_status,
            'notes': current_notes,
            'deduction': row['deduction'],
            'scheduled_count': row['scheduled_count'],
            'recorded_count': row['recorded_count'],
            'unrecorded_count': row['unrecorded_count'],
            'pending_count': row['pending_count'],
            'compliance_rate': row['compliance_rate'],
            'status_label': row['status_label'],
            'badge_class': row['badge_class'],
            'slots_list': row['slots_list'],
            'period_slots': row['period_slots'],
        })

    return render(request, 'teachers/teacher_attendance.html', {
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'selected_date_obj': selected_date,
        'teacher_list_data': teacher_list_data,
        'summary': daily_data['summary'],
        'status_choices': TeacherAttendance.Status.choices,
        'active_year': active_year,
        'query': query,
    })


@login_required
@role_required(['ADMIN', 'TEACHER'])
def teacher_attendance_report(request):
    """
    Teacher Attendance & Compliance Report (Extracts both Recorded and Unrecorded teachers).
    Counts unrecorded periods/absences by Period/Hour (ម៉ោង), Day (ថ្ងៃ), Week (សប្តាហ៍), and Month (ខែ).
    """
    active_year = get_active_academic_year(request)
    filter_type = request.GET.get('filter_type', 'month') # 'period', 'day', 'week', 'month', 'custom'
    status_filter = request.GET.get('status', 'all') # 'all', 'present', 'unrecorded', 'leave', 'no_schedule'
    query = request.GET.get('q', '').strip()
    period_filter = request.GET.get('period', 'all') # 'all', '1'..'8'

    now_dt = datetime.now()
    today = now_dt.date()

    # Determine date range or single date based on filter_type
    if filter_type == 'period' or filter_type == 'day':
        date_str = request.GET.get('date', today.strftime('%Y-%m-%d'))
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = today
        start_date = selected_date
        end_date = selected_date
        week_date_str = selected_date.strftime('%Y-%m-%d')
        month_str = selected_date.strftime('%Y-%m')
        filter_label = f"ថ្ងៃ {selected_date.strftime('%d/%m/%Y')}"
    elif filter_type == 'week':
        week_date_str = request.GET.get('week_date', today.strftime('%Y-%m-%d'))
        try:
            ref_d = datetime.strptime(week_date_str, '%Y-%m-%d').date()
        except ValueError:
            ref_d = today
        # Cambodian school week: Monday (0) to Saturday (5)
        start_date = ref_d - timedelta(days=ref_d.weekday())
        end_date = start_date + timedelta(days=5)
        selected_date = ref_d
        month_str = ref_d.strftime('%Y-%m')
        filter_label = f"សប្តាហ៍ ({start_date.strftime('%d/%m/%Y')} ដល់ {end_date.strftime('%d/%m/%Y')})"
    elif filter_type == 'custom':
        start_str = request.GET.get('start_date', today.strftime('%Y-%m-%d'))
        end_str = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today
            end_date = today
        selected_date = start_date
        week_date_str = today.strftime('%Y-%m-%d')
        month_str = today.strftime('%Y-%m')
        filter_label = f"ចន្លោះថ្ងៃ ({start_date.strftime('%d/%m/%Y')} ដល់ {end_date.strftime('%d/%m/%Y')})"
    else: # Default: 'month'
        filter_type = 'month'
        month_str = request.GET.get('month', today.strftime('%Y-%m'))
        try:
            y, m = map(int, month_str.split('-'))
        except ValueError:
            y, m = today.year, today.month
            month_str = today.strftime('%Y-%m')
        start_date = date(y, m, 1)
        _, last_day = calendar.monthrange(y, m)
        end_date = date(y, m, last_day)
        selected_date = today
        week_date_str = today.strftime('%Y-%m-%d')
        filter_label = f"ប្រចាំខែ {month_str}"

    # Fetch Active Teachers
    teachers_qs = Teacher.objects.filter(status='ACTIVE').order_by('teacher_id')
    if query:
        teachers_qs = teachers_qs.filter(
            Q(teacher_id__icontains=query) |
            Q(khmer_name__icontains=query) |
            Q(latin_name__icontains=query) |
            Q(specialization__icontains=query)
        )

    # Action: Sync teacher attendance from student logs
    if request.GET.get('action') == 'sync':
        synced_count = sync_teacher_attendance_from_student_logs(selected_date, active_year)
        messages.success(request, f"🔄 បានធ្វើសមកាលកម្មវត្តមានគ្រូចំនួន {synced_count} នាក់ពីការស្រង់វត្តមានសិស្សជោគជ័យ!")
        redirect_url = f"/teachers/attendance/report/?filter_type={filter_type}&date={selected_date.strftime('%Y-%m-%d')}&month={month_str}&week_date={week_date_str}"
        return redirect(redirect_url)

    # Process data according to filter type
    daily_report_data = None
    range_report_data = None
    filtered_rows = []

    if filter_type in ['period', 'day']:
        daily_res = get_teacher_daily_attendance_data(teachers_qs, selected_date, active_year, now_dt)
        daily_report_data = daily_res
        rows = daily_res['rows']

        # Apply Status Filter
        for r in rows:
            if status_filter == 'present' and r['daily_status'] != 'PRESENT':
                continue
            if status_filter == 'unrecorded' and r['unrecorded_count'] == 0:
                continue
            if status_filter == 'leave' and r['daily_status'] != 'EXCUSED_LEAVE':
                continue
            if status_filter == 'no_schedule' and r['scheduled_count'] > 0:
                continue
            
            # Period specific filter for 'period' mode
            if filter_type == 'period' and period_filter != 'all' and period_filter.isdigit():
                p_num = int(period_filter)
                slot_info = r['period_slots'].get(p_num)
                # If filtered by period, check if slot exists
                r['focused_period_slot'] = slot_info
            
            filtered_rows.append(r)
        
        summary = daily_res['summary']

    else:
        # Week, Month, or Custom Range
        range_res = get_teacher_range_attendance_report(teachers_qs, start_date, end_date, active_year, now_dt)
        range_report_data = range_res
        rows = range_res['rows']

        # Apply Status Filter
        for r in rows:
            if status_filter == 'present' and (not r['is_perfect']):
                continue
            if status_filter == 'unrecorded' and (not r['has_unrecorded']):
                continue
            if status_filter == 'leave' and r['excused_days_count'] == 0:
                continue
            if status_filter == 'no_schedule' and (not r['no_schedule']):
                continue
            filtered_rows.append(r)

        summary = range_res['summary']

    return render(request, 'teachers/teacher_attendance_report.html', {
        'filter_type': filter_type,
        'filter_label': filter_label,
        'status_filter': status_filter,
        'period_filter': period_filter,
        'query': query,
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'selected_date_obj': selected_date,
        'week_date_str': week_date_str,
        'month_str': month_str,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'daily_report_data': daily_report_data,
        'range_report_data': range_report_data,
        'rows': filtered_rows,
        'summary': summary,
        'active_year': active_year,
        'period_range': range(1, 9),
    })


@login_required
def teacher_leave_list(request):
    """
    Lists teacher leave requests.
    Teachers see their own requests; Admins see all requests.
    Supports filtering by Status (ALL, PENDING, APPROVED, REJECTED) and Category (EMERGENCY, PLANNED).
    """
    from apps.teachers.models import TeacherLeaveRequest
    user = request.user
    status_filter = request.GET.get('status', 'ALL')
    category_filter = request.GET.get('category', 'ALL')
    
    if user.role == 'TEACHER':
        teacher_profile = getattr(user, 'teacher_profile', None)
        if not teacher_profile:
            messages.error(request, "ពុំមានគណនីគ្រូបង្រៀនភ្ជាប់ជាមួយអ្នកប្រើប្រាស់នេះឡើយ។")
            return redirect('dashboard')
        leaves_qs = TeacherLeaveRequest.objects.filter(teacher=teacher_profile)
    else:
        leaves_qs = TeacherLeaveRequest.objects.all()

    if status_filter != 'ALL':
        leaves_qs = leaves_qs.filter(status=status_filter)

    if category_filter != 'ALL':
        leaves_qs = leaves_qs.filter(category=category_filter)

    leaves = leaves_qs.select_related('teacher', 'substitute_teacher', 'applied_by', 'approved_by').order_by('-created_at')

    return render(request, 'teachers/teacher_leave_list.html', {

        'leaves': leaves,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'status_choices': TeacherLeaveRequest.Status.choices,
        'category_choices': TeacherLeaveRequest.Category.choices,
        'leave_type_choices': TeacherLeaveRequest.LeaveType.choices,
    })


@login_required
def teacher_leave_create(request):
    """
    Form for teacher or admin to submit a leave application.
    Supports:
      1. ការសុំច្បាប់ភ្លាមៗ (Emergency Leave): Strictly restricted to TODAY or TOMORROW where the teacher has scheduled classes, prior to cutoff time (e.g. 5:00 PM).
      2. ការសុំច្បាប់ទុកជាមុន (Planned Leave): In advance for formal approval & printable hardcopy.
    """
    from apps.teachers.models import TeacherLeaveRequest, TeacherAttendanceConfig
    from apps.teachers.utils import get_teacher_emergency_leave_schedule
    from apps.attendance.telegram_utils import send_teacher_leave_notification_telegram
    from datetime import time as dtime

    config = TeacherAttendanceConfig.get_settings()
    user = request.user

    teacher_profile = getattr(user, 'teacher_profile', None)
    all_teachers = Teacher.objects.filter(status='ACTIVE').order_by('teacher_id')
    available_substitutes = all_teachers.exclude(id=teacher_profile.id) if teacher_profile else all_teachers

    today = date.today()
    cutoff_time = config.emergency_leave_cutoff_time or dtime(17, 0)
    cutoff_str = cutoff_time.strftime('%H:%M')

    # Initial teacher for emergency schedule evaluation
    initial_teacher = teacher_profile or (all_teachers.first() if all_teachers else None)
    emergency_schedule = get_teacher_emergency_leave_schedule(initial_teacher) if initial_teacher else {'options': [], 'cutoff_str': cutoff_str, 'has_available_dates': False}

    # Precompute emergency schedules for Admin teacher dropdown
    emergency_schedules_map = {}
    if user.role == 'ADMIN':
        for t in all_teachers:
            emergency_schedules_map[str(t.id)] = get_teacher_emergency_leave_schedule(t)

    if request.method == 'POST':
        if user.role == 'ADMIN':
            teacher_id = request.POST.get('teacher_id')
            selected_teacher = get_object_or_404(Teacher, pk=teacher_id)
        else:
            if not teacher_profile:
                messages.error(request, "⚠️ អ្នកគ្មានសិទ្ធិដាក់ពាក្យសុំច្បាប់ឡើយ!")
                return redirect('dashboard')
            selected_teacher = teacher_profile

        category = request.POST.get('category', TeacherLeaveRequest.Category.PLANNED)
        leave_type = request.POST.get('leave_type', 'PERSONAL')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date') or start_date_str
        reason = request.POST.get('reason', '').strip()
        substitute_id = request.POST.get('substitute_teacher_id')
        attachment = request.FILES.get('attachment')

        substitute_teacher = Teacher.objects.filter(id=substitute_id).first() if substitute_id else None

        if start_date_str and reason:
            try:
                s_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                e_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

                if e_date < s_date:
                    messages.error(request, "⚠️ ថ្ងៃបញ្ចប់មិនអាចតូចជាងថ្ងៃចាប់ផ្តើមឡើយ!")
                elif s_date < today:
                    messages.error(request, "⚠️ មិនអាចដាក់ពាក្យសុំច្បាប់សម្រាប់កាលបរិច្ឆេទក្នុងអតីតកាលឡើយ!")
                else:
                    # Validate Category 1: Strict Timetable & Cutoff Rules
                    if category == TeacherLeaveRequest.Category.EMERGENCY:
                        schedule_info = get_teacher_emergency_leave_schedule(selected_teacher)
                        matched_option = next((opt for opt in schedule_info['options'] if opt['date_str'] == start_date_str), None)

                        if not matched_option:
                            messages.error(
                                request,
                                f"⚠️ លោកគ្រូ-អ្នកគ្រូ {selected_teacher.khmer_name} គ្មានកាលវិភាគបង្រៀននៅថ្ងៃ {start_date_str} ឡើយ! ការសុំច្បាប់បន្ទាន់អនុញ្ញាតតែថ្ងៃដែលមានកាលវិភាគបង្រៀន (ថ្ងៃនេះ ឬស្អែក) ប៉ុណ្ណោះ។"
                            )
                            return render(request, 'teachers/teacher_leave_form.html', {
                                'teacher_profile': teacher_profile,
                                'all_teachers': all_teachers if user.role == 'ADMIN' else None,
                                'available_substitutes': available_substitutes,
                                'category_choices': TeacherLeaveRequest.Category.choices,
                                'leave_type_choices': TeacherLeaveRequest.LeaveType.choices,
                                'config': config,
                                'emergency_schedule': emergency_schedule,
                                'emergency_schedules_json': json.dumps(emergency_schedules_map),
                                'today_str': today.strftime('%Y-%m-%d'),
                            })

                        if not matched_option['is_allowed'] and user.role != 'ADMIN':
                            messages.error(
                                request,
                                f"⚠️ ការសុំច្បាប់បន្ទាន់សម្រាប់ {matched_option['label']} បានផុតម៉ោងកំណត់ {schedule_info['cutoff_str']} រសៀលហើយ!"
                            )
                            return render(request, 'teachers/teacher_leave_form.html', {
                                'teacher_profile': teacher_profile,
                                'all_teachers': all_teachers if user.role == 'ADMIN' else None,
                                'available_substitutes': available_substitutes,
                                'category_choices': TeacherLeaveRequest.Category.choices,
                                'leave_type_choices': TeacherLeaveRequest.LeaveType.choices,
                                'config': config,
                                'emergency_schedule': emergency_schedule,
                                'emergency_schedules_json': json.dumps(emergency_schedules_map),
                                'today_str': today.strftime('%Y-%m-%d'),
                            })

                        # Emergency leave is for the specific timetable teaching date
                        e_date = s_date

                    # Generate Unique Leave Code
                    count_today = TeacherLeaveRequest.objects.filter(created_at__date=today).count() + 1
                    leave_code = f"LV-{today.strftime('%Y%m%d')}-{count_today:03d}"
                    proxy_note = request.POST.get('proxy_note', '').strip() if user.role == 'ADMIN' else ''

                    leave_req = TeacherLeaveRequest.objects.create(
                        category=category,
                        leave_code=leave_code,
                        teacher=selected_teacher,
                        applied_by=user,
                        proxy_note=proxy_note,
                        substitute_teacher=substitute_teacher,
                        leave_type=leave_type,
                        start_date=s_date,
                        end_date=e_date,
                        reason=reason,
                        attachment=attachment,
                        status=TeacherLeaveRequest.Status.PENDING
                    )
                    # Notify management via Telegram
                    send_teacher_leave_notification_telegram(leave_req)


                    cat_badge = "ការសុំច្បាប់ភ្លាមៗ (បន្ទាន់)" if category == TeacherLeaveRequest.Category.EMERGENCY else "ការសុំច្បាប់ទុកជាមុន"
                    messages.success(
                        request,
                        f"📨 បានដាក់ពាក្យសុំច្បាប់ [{cat_badge} • លេខកូដ: {leave_code}] ដោយជោគជ័យ! ប្រព័ន្ធបានជូនដំណឹងទៅគណៈគ្រប់គ្រងសាលាហើយ។"
                    )
                    return redirect('teacher_leave_list')
            except Exception as ex:
                messages.error(request, f"⚠️ កំហុសកាលបរិច្ឆេទ ឬប្រព័ន្ធ៖ {str(ex)}")
        else:
            messages.error(request, "⚠️ សូមបំពេញកាលបរិច្ឆេទ និងមូលហេតុសុំច្បាប់ឱ្យបានគ្រប់គ្រាន់!")

    return render(request, 'teachers/teacher_leave_form.html', {
        'teacher_profile': teacher_profile,
        'all_teachers': all_teachers if user.role == 'ADMIN' else None,
        'available_substitutes': available_substitutes,
        'category_choices': TeacherLeaveRequest.Category.choices,
        'leave_type_choices': TeacherLeaveRequest.LeaveType.choices,
        'config': config,
        'emergency_schedule': emergency_schedule,
        'emergency_schedules_json': json.dumps(emergency_schedules_map),
        'today_str': today.strftime('%Y-%m-%d'),
    })



@login_required
def teacher_leave_print_letter(request, pk):
    """
    Formal Khmer A4 Printable Leave Application Letter (លិខិតសុំច្បាប់ផ្លូវការ)
    """
    from apps.teachers.models import TeacherLeaveRequest
    from apps.accounts.models import SchoolProfile

    leave_req = get_object_or_404(TeacherLeaveRequest, pk=pk)
    user = request.user

    # Security check: Teacher can only print their own leave form unless Admin
    if user.role == 'TEACHER':
        if not hasattr(user, 'teacher_profile') or user.teacher_profile.id != leave_req.teacher.id:
            messages.error(request, "លោកអ្នកអាចបោះពុម្ពបានតែលិខិតសុំច្បាប់របស់ខ្លួនឯងប៉ុណ្ណោះ!")
            return redirect('teacher_leave_list')

    school_profile = SchoolProfile.get_settings()

    return render(request, 'teachers/teacher_leave_print.html', {
        'leave': leave_req,
        'school': school_profile,
        'today': date.today(),
        'page_title': f"លិខិតសុំច្បាប់ - {leave_req.teacher.khmer_name} ({leave_req.leave_code or leave_req.id})"
    })



# ----------------- TEACHER EXPORT & IMPORT (EXCEL / CSV) -----------------

@login_required
@role_required(['ADMIN'])
def teacher_export_excel(request):
    """
    Exports all teachers into a beautifully formatted Excel sheet matching official MoEYS standard layout.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    teachers = Teacher.objects.all().order_by('teacher_id')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MoEYS Teachers Directory"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 2-Row Header matching official MoEYS table
    headers_row1 = [
        "ល.រ", "អត្តលេខ", "គោត្តនាម និងនាម", "ភេទ", "ថ្ងៃខែឆ្នាំកំណើត",
        "កម្រិតវប្បធម៌", "ឯកទេស", "កម្រិតបណ្តុះបណ្តាល", "ថ្ងៃចូលបម្រើការងាររដ្ឋ",
        "ថ្ងៃខែឆ្នាំតែងតាំងស៊ប់", "មុខវិជ្ជាឯកទេសទី១", "មុខវិជ្ជាឯកទេសទី២",
        "ភារកិច្ចបច្ចុប្បន្ន", "តាំងស៊ប់តាមប្រកាស", "", "", "លេខទូរស័ព្ទ", "ស្ថានភាព"
    ]

    ws.append(headers_row1)
    # Merge "តាំងស៊ប់តាមប្រកាស" for columns 14, 15, 16
    ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=16)

    headers_row2 = [
        "No.", "Teacher ID", "Khmer & Latin Name", "Gender", "DOB (DD-MM-YYYY)",
        "Qualification", "Specialization", "Training Level", "State Hire Date",
        "Permanent Date", "Subject 1", "Subject 2", "Current Duty",
        "ប្រភេទ / Type", "ឆ្នាំទទួល / Year", "ប្រ.លេខ / Decree No.", "Phone", "Status"
    ]
    ws.append(headers_row2)

    for r_idx in [1, 2]:
        ws.row_dimensions[r_idx].height = 26
        for col_num in range(1, len(headers_row2) + 1):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.fill = header_fill if r_idx == 1 else sub_header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    for idx, t in enumerate(teachers, 1):
        row = [
            idx,
            t.teacher_id,
            f"{t.khmer_name} ({t.latin_name})" if t.latin_name else t.khmer_name,
            t.get_gender_display(),
            t.date_of_birth.strftime('%d-%m-%Y') if t.date_of_birth else '',
            t.qualification or '',
            t.specialization or '',
            t.training_level or '',
            t.state_hire_date.strftime('%d-%m-%Y') if t.state_hire_date else '',
            t.permanent_date.strftime('%d-%m-%Y') if t.permanent_date else '',
            t.primary_subject or '',
            t.secondary_subject or '',
            t.current_duty or 'គ្រូបង្រៀន',
            t.prakas_category or '',
            t.prakas_year or '',
            t.prakas_number or '',
            t.phone or '',
            t.get_status_display()
        ]
        ws.append(row)
        curr_row = idx + 2
        ws.row_dimensions[curr_row].height = 22
        for c_idx in range(1, len(row) + 1):
            cell = ws.cell(row=curr_row, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            if c_idx in [1, 2, 4, 5, 9, 10, 14, 15, 16, 18]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="MoEYS_Teachers_Directory_{date.today().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required(['ADMIN'])
def teacher_import_template_excel(request):
    """Generates an official MoEYS standard Excel template for importing teachers."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teacher Import Template"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "Teacher ID *",
        "Khmer Name *",
        "Latin Name *",
        "Gender (M/F) *",
        "DOB (DD-MM-YYYY)",
        "Qualification",
        "Specialization *",
        "Training Level",
        "State Hire Date (DD-MM-YYYY)",
        "Permanent Date (DD-MM-YYYY)",
        "Primary Subject",
        "Secondary Subject",
        "Current Duty",
        "Prakas Category",
        "Prakas Year",
        "Prakas Number",
        "Phone *",
        "Email",
        "Base Salary ($)",
        "Status (ACTIVE/ON_LEAVE/RESIGNED)"
    ]
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 30

    sample_rows = [
        ["T-001", "សុខ វិបុល", "Sok Vibol", "M", "15-05-1985", "បរិញ្ញាបត្រ", "គណិតវិទ្យា", "គរុកោសល្យឧត្តម (បរិញ្ញាបត្រ+១)", "01-10-2010", "15-12-2012", "គណិតវិទ្យា", "រូបវិទ្យា", "គ្រូបង្រៀន", "ក្របខ័ណ្ឌ ក.១", "2012", "ប្រកាសលេខ ១៤៥", "012345678", "vibol@school.edu.kh", 500, "ACTIVE"],
        ["T-002", "ចាន់ រិទ្ធី", "Chan Rithy", "M", "20-08-1990", "អនុបណ្ឌិត", "រូបវិទ្យា", "គរុកោសល្យឧត្តម", "15-10-2014", "20-11-2016", "រូបវិទ្យា", "គីមីវិទ្យា", "ប្រធានផ្នែកវិទ្យាសាស្ត្រ", "ក្របខ័ណ្ឌ ក.២", "2016", "ប្រកាសលេខ ២២០", "017888999", "rithy@school.edu.kh", 550, "ACTIVE"],
        ["T-003", "ម៉ៅ ស្រីពៅ", "Mao Sreypov", "F", "10-12-1992", "បរិញ្ញាបត្រ", "ភាសាខ្មែរ", "គរុកោសល្យមូលដ្ឋាន (១២+២)", "01-11-2015", "10-01-2018", "តែងសេចក្តី", "អក្សរសិល្ប៍", "គ្រូបង្រៀន", "ក្របខ័ណ្ឌ ខ.១", "2018", "ប្រកាសលេខ ៣០៥", "096123456", "sreypov@school.edu.kh", 500, "ACTIVE"],
    ]
    for r in sample_rows:
        ws.append(r)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 22
        for c_idx in range(1, len(r) + 1):
            cell = ws.cell(row=curr_row, column=c_idx)
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border
            if c_idx in [1, 4, 5, 9, 10, 14, 15, 16, 20]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="teacher_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required(['ADMIN'])
def teacher_import_template_csv(request):
    """Generates a downloadable sample CSV template for importing teachers."""
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="teacher_import_template.csv"'
    writer = csv.writer(response)
    writer.writerow([
        "Teacher ID *",
        "Khmer Name *",
        "Latin Name *",
        "Gender (M/F) *",
        "DOB (DD-MM-YYYY)",
        "Qualification",
        "Specialization *",
        "Training Level",
        "State Hire Date (DD-MM-YYYY)",
        "Permanent Date (DD-MM-YYYY)",
        "Primary Subject",
        "Secondary Subject",
        "Current Duty",
        "Prakas Category",
        "Prakas Year",
        "Prakas Number",
        "Phone *",
        "Email",
        "Base Salary ($)",
        "Status (ACTIVE/ON_LEAVE/RESIGNED)"
    ])
    writer.writerow(["T-001", "សុខ វិបុល", "Sok Vibol", "M", "15-05-1985", "បរិញ្ញាបត្រ", "គណិតវិទ្យា", "គរុកោសល្យឧត្តម (បរិញ្ញាបត្រ+១)", "01-10-2010", "15-12-2012", "គណិតវិទ្យា", "រូបវិទ្យា", "គ្រូបង្រៀន", "ក្របខ័ណ្ឌ ក.១", "2012", "ប្រកាសលេខ ១៤៥", "012345678", "vibol@school.edu.kh", "500", "ACTIVE"])
    writer.writerow(["T-002", "ចាន់ រិទ្ធី", "Chan Rithy", "M", "20-08-1990", "អនុបណ្ឌិត", "រូបវិទ្យា", "គរុកោសល្យឧត្តម", "15-10-2014", "20-11-2016", "រូបវិទ្យា", "គីមីវិទ្យា", "ប្រធានផ្នែកវិទ្យាសាស្ត្រ", "ក្របខ័ណ្ឌ ក.២", "2016", "ប្រកាសលេខ ២២០", "017888999", "rithy@school.edu.kh", "550", "ACTIVE"])
    return response


@login_required
@role_required(['ADMIN'])
def teacher_import(request):
    """Bulk imports teachers from Excel or CSV file supporting both MoEYS 20-col and standard formats."""
    import csv
    import io
    import openpyxl
    from decimal import Decimal
    from datetime import datetime, date
    from apps.accounts.models import User

    def parse_date_flexible(val):
        if not val:
            return None
        if isinstance(val, (datetime, date)):
            return val if isinstance(val, date) else val.date()
        val_clean = str(val).strip()
        for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y'):
            try:
                return datetime.strptime(val_clean, fmt).date()
            except ValueError:
                pass
        return None

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        filename = file.name.lower()
        rows_data = []

        try:
            if filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                # If first row is merged or title, skip to headers
                start_row = 2
                for row in ws.iter_rows(min_row=start_row, values_only=True):
                    if row and any(row):
                        # skip secondary header row if present
                        if str(row[0] or '').strip().lower() in ['no.', 'ល.រ', 'teacher id', 'teacher id *']:
                            continue
                        rows_data.append(list(row))
            elif filename.endswith('.csv'):
                decoded_file = file.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                header = next(reader, None)
                for row in reader:
                    if row and any(row):
                        if str(row[0] or '').strip().lower() in ['no.', 'ល.រ', 'teacher id', 'teacher id *']:
                            continue
                        rows_data.append(row)
            else:
                messages.error(request, "⚠️ ទម្រង់ឯកសារមិនត្រឹមត្រូវ! សូមជ្រើសរើសឯកសារ .xlsx ឬ .csv")
                return redirect('teacher_import')

            success_count = 0
            updated_count = 0
            errors = []

            for idx, r in enumerate(rows_data, 2):
                if len(r) < 3:
                    continue

                t_id = str(r[0] or '').strip()
                k_name = str(r[1] or '').strip()
                l_name = str(r[2] or '').strip()
                gender_raw = str(r[3] or 'M').strip().upper()
                dob = parse_date_flexible(r[4] if len(r) > 4 else None)

                if not t_id or not k_name:
                    errors.append(f"ជួរទី {idx}៖ ខ្វះ Teacher ID ឬឈ្មោះគ្រូ")
                    continue

                gender = Teacher.Gender.FEMALE if gender_raw in ['F', 'ស្រី', 'FEMALE'] else Teacher.Gender.MALE

                # Check format by column length
                if len(r) >= 16:
                    # MoEYS Extended 20-col format
                    qual = str(r[5] or '').strip() if len(r) > 5 else ''
                    spec = str(r[6] or '').strip() if len(r) > 6 else 'ទូទៅ'
                    training_level = str(r[7] or '').strip() if len(r) > 7 else ''
                    state_hire_date = parse_date_flexible(r[8] if len(r) > 8 else None)
                    permanent_date = parse_date_flexible(r[9] if len(r) > 9 else None)
                    primary_subj = str(r[10] or '').strip() if len(r) > 10 else ''
                    secondary_subj = str(r[11] or '').strip() if len(r) > 11 else ''
                    current_duty = str(r[12] or '').strip() if len(r) > 12 else 'គ្រូបង្រៀន'
                    prakas_cat = str(r[13] or '').strip() if len(r) > 13 else ''
                    prakas_yr = str(r[14] or '').strip() if len(r) > 14 else ''
                    prakas_num = str(r[15] or '').strip() if len(r) > 15 else ''
                    phone = str(r[16] or '').strip() if len(r) > 16 else ''
                    email = str(r[17] or '').strip() if len(r) > 17 else ''
                    salary_str = str(r[18] or '500.00').replace('$', '').replace(',', '').strip() if len(r) > 18 else '500.00'
                    salary = Decimal(salary_str) if salary_str else Decimal('500.00')
                    status_raw = str(r[19] or 'ACTIVE').strip().upper() if len(r) > 19 else 'ACTIVE'
                    max_hours = 18
                else:
                    # Legacy 12-col format
                    phone = str(r[5] or '').strip() if len(r) > 5 else ''
                    email = str(r[6] or '').strip() if len(r) > 6 else ''
                    spec = str(r[7] or '').strip() if len(r) > 7 else 'ទូទៅ'
                    qual = str(r[8] or '').strip() if len(r) > 8 else ''
                    max_hours = int(r[9]) if len(r) > 9 and str(r[9]).isdigit() else 18
                    salary_str = str(r[10] or '500.00').replace('$', '').replace(',', '').strip() if len(r) > 10 else '500.00'
                    salary = Decimal(salary_str) if salary_str else Decimal('500.00')
                    status_raw = str(r[11] or 'ACTIVE').strip().upper() if len(r) > 11 else 'ACTIVE'
                    training_level = ''
                    state_hire_date = None
                    permanent_date = None
                    primary_subj = ''
                    secondary_subj = ''
                    current_duty = 'គ្រូបង្រៀន'
                    prakas_cat = ''
                    prakas_yr = ''
                    prakas_num = ''

                status = Teacher.Status.ON_LEAVE if status_raw == 'ON_LEAVE' else (Teacher.Status.RESIGNED if status_raw == 'RESIGNED' else Teacher.Status.ACTIVE)

                teacher, created = Teacher.objects.update_or_create(
                    teacher_id=t_id,
                    defaults={
                        'khmer_name': k_name,
                        'latin_name': l_name or k_name,
                        'gender': gender,
                        'date_of_birth': dob,
                        'phone': phone,
                        'email': email or None,
                        'specialization': spec,
                        'qualification': qual,
                        'training_level': training_level,
                        'state_hire_date': state_hire_date,
                        'permanent_date': permanent_date,
                        'primary_subject': primary_subj,
                        'secondary_subject': secondary_subj,
                        'current_duty': current_duty,
                        'prakas_category': prakas_cat,
                        'prakas_year': prakas_yr,
                        'prakas_number': prakas_num,
                        'max_weekly_hours': max_hours,
                        'base_salary': salary,
                        'status': status,
                    }
                )

                # Auto-create or link User Account
                username = t_id.lower().replace('-', '_').replace(' ', '_')
                user = User.objects.filter(username=username).first()
                if not user:
                    user = User.objects.create_user(
                        username=username,
                        password='password123',
                        role=User.Role.TEACHER,
                        khmer_name=k_name,
                        latin_name=l_name,
                        phone=phone,
                        email=email or ''
                    )
                teacher.user = user
                teacher.save(update_fields=['user'])

                if created:
                    success_count += 1
                else:
                    updated_count += 1

            if success_count > 0 or updated_count > 0:
                messages.success(request, f"🎉 ជោគជ័យ! បានបញ្ចូលគ្រូបង្រៀនថ្មី {success_count} នាក់ និងកែប្រែទិន្នន័យចាស់ {updated_count} នាក់។ គណនី Login ត្រូវបានបង្កើតដោយស្វ័យប្រវត្តិ (Password: password123)")
            if errors:
                for err in errors[:5]:
                    messages.warning(request, f"⚠️ {err}")
            return redirect('teacher_list')

        except Exception as e:
            messages.error(request, f"⚠️ មានបញ្ហាក្នុងការ Import៖ {str(e)}")
            return redirect('teacher_import')

    return render(request, 'teachers/teacher_import.html')


# =========================================================================
# Teacher Information Re-Submission Campaign & Self-Update Portal
# =========================================================================
@login_required
@role_required(['ADMIN'])
def teacher_update_campaign_view(request):
    """
    Allows Admin to configure the Teacher Information Re-Submission Campaign:
    - Activate/Deactivate
    - Tick allowed sections (Identity, Gender/DOB, Contact, Address, Education, Training & Subjects, Civil Service, Photo & Resume)
    - Set Title, Instructions, Target Teachers, Deadline
    - View submission/verification statistics
    """
    campaign = TeacherProfileUpdateCampaign.objects.first()
    if not campaign:
        campaign = TeacherProfileUpdateCampaign.objects.create(
            title="យុទ្ធនាការផ្ទៀងផ្ទាត់ និងបំពេញព័ត៌មានគ្រូបង្រៀន",
            allowed_sections=['identity', 'dob_gender', 'phone_email', 'address', 'education', 'training_subjects', 'civil_service'],
            is_active=True
        )

    if request.method == 'POST':
        title = request.POST.get('title', '').strip() or "យុទ្ធនាការផ្ទៀងផ្ទាត់ និងបំពេញព័ត៌មានគ្រូបង្រៀន"
        instructions = request.POST.get('instructions', '').strip()
        is_active = request.POST.get('is_active') == '1' or 'is_active' in request.POST
        deadline_str = request.POST.get('deadline', '').strip()
        target_all = request.POST.get('target_all') == '1' or 'target_all' in request.POST
        selected_sections = request.POST.getlist('sections')
        
        deadline = None
        if deadline_str:
            try:
                deadline = datetime.strptime(deadline_str, '%Y-%m-%d').date()
            except Exception:
                pass

        campaign.title = title
        campaign.instructions = instructions
        campaign.is_active = is_active
        campaign.allowed_sections = selected_sections
        campaign.target_all = target_all
        campaign.deadline = deadline
        campaign.save()

        # Handle selected teachers if target_all is false
        if not target_all:
            teacher_ids = request.POST.getlist('target_teachers')
            campaign.target_teachers.set(teacher_ids)
        else:
            campaign.target_teachers.clear()

        messages.success(request, "🎉 បានរក្សាទុកការកំណត់យុទ្ធនាការទាមទារព័ត៌មានគ្រូឡើងវិញដោយជោគជ័យ!")
        return redirect('teacher_update_campaign')

    teachers = Teacher.objects.filter(status=Teacher.Status.ACTIVE).order_by('teacher_id')
    total_teachers = teachers.count()
    verified_count = teachers.filter(last_profile_verified_at__isnull=False).count()
    pending_count = total_teachers - verified_count
    progress_pct = int((verified_count / total_teachers * 100)) if total_teachers > 0 else 0

    context = {
        'campaign': campaign,
        'available_sections': TeacherProfileUpdateCampaign.AVAILABLE_SECTIONS,
        'teachers': teachers,
        'total_teachers': total_teachers,
        'verified_count': verified_count,
        'pending_count': pending_count,
        'progress_pct': progress_pct,
    }
    return render(request, 'teachers/teacher_update_campaign.html', context)


@login_required
def teacher_self_update_portal(request):
    """
    Teacher Portal for self-updating/re-submitting profile information.
    Optimized for both Desktop and Mobile Smartphone with Pinch-to-Zoom.
    Only allows editing sections ticked by Admin in the active campaign.
    """
    is_admin = request.user.role == User.Role.ADMIN
    teacher = None

    if hasattr(request.user, 'teacher_profile') and request.user.teacher_profile:
        teacher = request.user.teacher_profile
    elif is_admin:
        teacher_id = request.GET.get('teacher_id')
        if teacher_id:
            teacher = get_object_or_404(Teacher, id=teacher_id)
        else:
            teacher = Teacher.objects.first()
            if not teacher:
                messages.warning(request, "⚠️ មិនទាន់មានគ្រូបង្រៀននៅក្នុងប្រព័ន្ធនៅឡើយទេ។")
                return redirect('teacher_list')

    if not teacher:
        messages.error(request, "⚠️ គណនីរបស់អ្នកមិនមានភ្ជាប់ជាមួយប្រវត្តិរូបគ្រូបង្រៀនឡើយ។")
        return redirect('dashboard_redirect')

    campaign = TeacherProfileUpdateCampaign.get_current_active(teacher)
    if not campaign and not is_admin:
        messages.info(request, "ℹ️ បច្ចុប្បន្នគ្មានយុទ្ធនាការទាមទារឱ្យបំពេញព័ត៌មានគ្រូឡើងវិញឡើយ។")
        return redirect('teacher_detail', pk=teacher.id)

    # Allowed sections list
    allowed_sections = campaign.allowed_sections if campaign else [s[0] for s in TeacherProfileUpdateCampaign.AVAILABLE_SECTIONS]

    if request.method == 'POST':
        # Process updates for allowed fields
        if 'identity' in allowed_sections:
            if request.POST.get('khmer_name'):
                teacher.khmer_name = request.POST.get('khmer_name').strip()
            if request.POST.get('latin_name'):
                teacher.latin_name = request.POST.get('latin_name').strip()

        if 'dob_gender' in allowed_sections:
            if request.POST.get('gender'):
                teacher.gender = request.POST.get('gender')
            dob_str = request.POST.get('date_of_birth')
            if dob_str:
                try:
                    teacher.date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date()
                except Exception:
                    pass

        if 'phone_email' in allowed_sections:
            if request.POST.get('phone'):
                teacher.phone = request.POST.get('phone').strip()
            teacher.email = request.POST.get('email', '').strip() or None

        if 'address' in allowed_sections:
            addr = request.POST.get('address', '').strip()
            if addr:
                teacher.address = addr

        if 'education' in allowed_sections:
            teacher.qualification = request.POST.get('qualification', '').strip()
            if request.POST.get('specialization'):
                teacher.specialization = request.POST.get('specialization').strip()

        if 'training_subjects' in allowed_sections:
            teacher.training_level = request.POST.get('training_level', '').strip()
            teacher.primary_subject = request.POST.get('primary_subject', '').strip()
            teacher.secondary_subject = request.POST.get('secondary_subject', '').strip()

        if 'civil_service' in allowed_sections:
            teacher.current_duty = request.POST.get('current_duty', '').strip() or 'គ្រូបង្រៀន'
            teacher.prakas_category = request.POST.get('prakas_category', '').strip()
            teacher.prakas_year = request.POST.get('prakas_year', '').strip()
            teacher.prakas_number = request.POST.get('prakas_number', '').strip()
            state_hire_str = request.POST.get('state_hire_date')
            if state_hire_str:
                try:
                    teacher.state_hire_date = datetime.strptime(state_hire_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            permanent_str = request.POST.get('permanent_date')
            if permanent_str:
                try:
                    teacher.permanent_date = datetime.strptime(permanent_str, '%Y-%m-%d').date()
                except Exception:
                    pass

        if 'photo_resume' in allowed_sections:
            if 'photo' in request.FILES:
                teacher.photo = request.FILES['photo']
            if 'resume' in request.FILES:
                teacher.resume = request.FILES['resume']

        teacher.last_profile_verified_at = timezone.now()
        teacher.save()

        # Update User object name/phone if changed
        if teacher.user:
            teacher.user.khmer_name = teacher.khmer_name
            teacher.user.latin_name = teacher.latin_name
            teacher.user.phone = teacher.phone
            teacher.user.email = teacher.email or ''
            teacher.user.save(update_fields=['khmer_name', 'latin_name', 'phone', 'email'])

        # Notify Admin via Telegram if configured
        try:
            from apps.accounts.models import SchoolProfile
            profile = SchoolProfile.objects.first()
            if profile and profile.telegram_bot_token and profile.telegram_channel:
                import urllib.request
                import urllib.parse
                bot_token = profile.telegram_bot_token.strip()
                chat_id = profile.telegram_channel.strip()
                if not chat_id.startswith('@') and not chat_id.startswith('-'):
                    chat_id = f"@{chat_id}"
                msg_text = (
                    f"✅ <b>គ្រូបង្រៀនបានធ្វើបច្ចុប្បន្នភាពព័ត៌មាន (Teacher Portal Update)</b>\n\n"
                    f"👤 ឈ្មោះគ្រូ៖ <b>{teacher.khmer_name}</b> ({teacher.teacher_id})\n"
                    f"📞 លេខទូរស័ព្ទ៖ <code>{teacher.phone}</code>\n"
                    f"📚 ឯកទេស៖ <b>{teacher.specialization}</b>\n"
                    f"⏰ ពេលវេលា៖ {timezone.now().strftime('%d-%m-%Y %H:%M')}\n"
                )
                url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                payload = urllib.parse.urlencode({
                    'chat_id': chat_id,
                    'text': msg_text,
                    'parse_mode': 'HTML'
                }).encode('utf-8')
                req = urllib.request.Request(url, data=payload, headers={'User-Agent': 'SchoolSM-Bot'})
                urllib.request.urlopen(req, timeout=4)
        except Exception:
            pass

        messages.success(request, f"🎉 អរគុណលោកគ្រូ/អ្នកគ្រូ {teacher.khmer_name}! ព័ត៌មានត្រូវបានផ្ទៀងផ្ទាត់ និងធ្វើបច្ចុប្បន្នភាពរួចរាល់ដោយជោគជ័យ។")
        if is_admin:
            return redirect('teacher_detail', pk=teacher.id)
        return redirect('dashboard_redirect')

    context = {
        'teacher': teacher,
        'campaign': campaign,
        'allowed_sections': allowed_sections,
        'is_admin': is_admin,
    }
    return render(request, 'teachers/teacher_self_update_portal.html', context)



