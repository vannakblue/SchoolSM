import os
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

import openpyxl
from datetime import datetime, date
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth.hashers import make_password
from apps.accounts.models import User
from apps.teachers.models import Teacher


KHMER_LATIN_DICT = {
    'កង': 'Kang', 'កញ្ញា': 'Kanya', 'កន្យា': 'Kanya', 'កាន': 'Kan', 'កុសល': 'Kosal',
    'កឿន': 'Koeun', 'កែវ': 'Keo', 'ក្រឹង': 'Kreung', 'ខឹម': 'Khim', 'ខៀវ': 'Khiev',
    'ខេមរិន្ទ': 'Khemrinth', 'គង់': 'Kong', 'គន្ធា': 'Kunthea', 'គាន': 'Kean', 'គ្រីន': 'Krin',
    'ឃាង': 'Kheang', 'ឃឹម': 'Khim', 'ឃុត': 'Khut', 'ឃុន': 'Khun', 'ងួន': 'Nguon',
    'ចន្ថា': 'Chantha', 'ចន្ទ្រា': 'Chantrea', 'ចរិយា': 'Chariya', 'ចាន់': 'Chan',
    'ចាន់ណាក់': 'Channak', 'ចាន់ថា': 'Chantha', 'ចាន់នី': 'Channy', 'ចាន់រ៉ា': 'Chanra',
    'ចាន់សុផាន់ណា': 'Chansophanna', 'ចិន្តា': 'Chinda', 'ចេង': 'Cheng', 'ចំរើន': 'Chamroeun',
    'ច័ន្ទសុធី': 'Chansothey', 'ឆាយ': 'Chhay', 'ឆេង': 'Chheng', 'ជឹង': 'Cheung', 'ជឹម': 'Chim',
    'ជុំ': 'Chum', 'ជួ': 'Chou', 'ជួង': 'Chhoung', 'ជៀស': 'Chieas', 'ជៃ': 'Chay',
    'ជំនិត': 'Chomnit', 'ឈាង': 'Chheang', 'ឈឿន': 'Chhoeun', 'ដាវណ្ណ': 'Davann', 'ដាវី': 'Davy',
    'ដុក': 'Dok', 'ដួង': 'Duong', 'ឌីណា': 'Dina', 'ឌីនីន': 'Dinin', 'ឌីម៉ង់': 'Dimang',
    'ឌុច': 'Duch', 'ណារី': 'Nary', 'ណារ៉ា': 'Nara', 'ណាសួន': 'Nasoun', 'ណុប': 'Nop',
    'ណុំ': 'Nom', 'ថោង': 'Thaong', 'ទិត': 'Tith', 'ទិន': 'Tin', 'ទឹម': 'Tim',
    'ទុន': 'Tun', 'ទូច': 'Touch', 'ទ្រី': 'Try', 'ធី': 'Thy', 'ធីតា': 'Thida',
    'នាង': 'Neang', 'និមល': 'Nimol', 'និស្សិត': 'Nissith', 'នី': 'Ny', 'បុណ្ណវេទ': 'Bonnveth',
    'បូ': 'Bo', 'បូរាមី': 'Boramy', 'ប៉ន': 'Porn', 'ប៊ុន': 'Bun', 'ប៊ុនណារិទ្ធ': 'Bunnarith',
    'ប៊ុនថន': 'Bunthon', 'ប៊ុនធន': 'Bunthon', 'ប្រាក់': 'Prak', 'ផន': 'Phorn', 'ផល': 'Phal',
    'ផល្លី': 'Phally', 'ផាត់': 'Phat', 'ផេង': 'Pheng', 'ពិដោរ': 'Pidor', 'ពិសាល': 'Pisal',
    'ពិសី': 'Pisey', 'ពិសេស': 'Pises', 'ពឺន': 'Poeun', 'ពុទ្ធាវី': 'Putheavy', 'ពូន': 'Poun',
    'ពេជ្រ': 'Pech', 'ពៅ': 'Pov', 'ភារុន': 'Phearun', 'ភ័ស': 'Phorn', 'មករា': 'Makara',
    'មាស': 'Meas', 'មូល': 'Moul', 'មៀច': 'Miech', 'ម៉ង់': 'Mang', 'ម៉ានិន': 'Manin',
    'ម៉ាលីស': 'Malis', 'ម៉ូនីដា': 'Monida', 'ម៉ែន': 'Men', 'យូ': 'You', 'យូណៃ': 'Younai',
    'យ៉ន': 'Yorn', 'យ៉ាង': 'Yang', 'យ៉េន': 'Yen', 'រក្សា': 'Raksa', 'រចនា': 'Rachana',
    'រតនា': 'Rattana', 'រិទ្ធីយ៉ា': 'Rithiya', 'រុនស្រី': 'Ronsrey', 'រ៉ន': 'Rorn',
    'លក្ខិណា': 'Leakhena', 'លន': 'Lon', 'លាងឃន': 'Leangkhon', 'លី': 'Ly', 'លីឆាយ': 'Lychhay',
    'លឿង': 'Loeung', 'វណ្ណៈ': 'Vannak', 'វាសនា': 'Veasna', 'វិន': 'Vin', 'វិសាល': 'Visal',
    'វ៉ាង': 'Vang', 'វ៉ាន់': 'Van', 'វ៉េង': 'Veng', 'សម្បត្តិ': 'Sambath', 'សា': 'Sa',
    'សានម៉ូណាវី': 'Sanmonavy', 'សានសុផានី': 'Sansophanith', 'សាន់': 'San', 'សាមឌី': 'Samdy',
    'សារិន': 'Sarin', 'សាវិន': 'Savin', 'សីហា': 'Seyha', 'សុខ': 'Sok', 'សុខឃៀង': 'Sokkheang',
    'សុខចាន់': 'Sokchan', 'សុខម៉េត': 'Sokmet', 'សុខា': 'Sokha', 'សុខុម': 'Sokhom',
    'សុគង់': 'Sokong', 'សុគន្ធារី': 'Sokuntheary', 'សុង': 'Song', 'សុជាតា': 'Socheata',
    'សុដានី': 'Sodany', 'សុទ្ធ': 'Soth', 'សុន': 'Son', 'សុផន': 'Sophon', 'សុផា': 'Sopha',
    'សុផាន': 'Sophan', 'សុផារិទ្ធ': 'Sopharith', 'សុភារៈ': 'Sophearak', 'សុភី': 'Sophea',
    'សុភ័ក្រ': 'Sopheak', 'សុមនី': 'Somony', 'សុម៉ាឡា': 'Somala', 'សុសៅគន្ធ': 'Sosaokunth',
    'សូ': 'So', 'សូកាន': 'Sokan', 'សូរីយា': 'Soriya', 'សួន': 'Suon', 'សួរ': 'Sour',
    'សួស': 'Suos', 'សឿន': 'Soeun', 'សេង': 'Seng', 'សេងហៃ': 'Senghai', 'សេត': 'Seth',
    'សេរីពង្ស': 'Sereypong', 'សេស': 'Ses', 'សែត': 'Set', 'សោម៉នវីរៈ': 'Somonvirak',
    'សំ': 'Sam', 'សំអុល': 'Sam Ol', 'សំអឿន': 'Samoeun', 'ស៊ិន': 'Sin', 'ស៊ីដារ៉ា': 'Sidara',
    'ស៊ីនាង': 'Siniang', 'ស៊ុយ': 'Suy', 'ស៊ុំ': 'Sum', 'ស្រស់': 'Sros', 'ស្រីណែត': 'Sreynet',
    'ស្រីនាង': 'Sreyniang', 'ស្រីពៅ': 'Sreypov', 'ស្រីរ័ត្ន': 'Sreyroth', 'ស្រីលក្ខ័': 'Sreyleak',
    'ស្រ៊ុន': 'Srun', 'ហន': 'Horn', 'ហុង': 'Hong', 'ហួត': 'Huot', 'ហៀង': 'Heang',
    'ហេង': 'Heng', 'ហ៊ឺ': 'Heu', 'ហ៊ូ': 'Hou', 'ឡាង': 'Lang', 'ឡុច': 'Loch',
    'ឡេង': 'Leng', 'អមរា': 'Amara', 'អាន': 'An', 'អៀ': 'Iea', 'អេង': 'Eng',
    'អោម': 'Aom', 'អ៊ិន': 'In', 'អ៊ឹម': 'Im', 'អ៊ុយ': 'Uy', 'ឯក': 'Ek'
}


def transliterate_khmer_name(kh_name):
    if not kh_name:
        return ''
    cleaned = kh_name.replace('\u200b', ' ').strip()
    words = cleaned.split()
    latin_words = []
    for w in words:
        w_clean = w.strip()
        if w_clean in KHMER_LATIN_DICT:
            latin_words.append(KHMER_LATIN_DICT[w_clean])
        else:
            latin_words.append(w_clean)
    return ' '.join(latin_words)


def format_phone_number(raw_phone):
    if not raw_phone:
        return ''
    p_str = str(raw_phone).strip()
    digits = ''.join(c for c in p_str if c.isdigit())
    if not digits:
        return ''
    if not digits.startswith('0'):
        digits = '0' + digits
    return digits


def parse_flexible_date(val):
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    val_str = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d.%m.%Y'):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    return None


class Command(BaseCommand):
    help = "Imports official MoEYS Teacher / Staff records from Excel sheet into Database (SQLite / Postgres / Supabase / Render)."

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default=r'E:\បញ្ជីគ្រប់គ្រងបុគ្គលិក.xlsx',
            help='Absolute path to the Excel file containing teacher records'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Clear dummy or existing teachers before importing'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        clear_first = options.get('clear', False)

        if not os.path.exists(file_path):
            alt_path = os.path.join(os.getcwd(), os.path.basename(file_path))
            if os.path.exists(alt_path):
                file_path = alt_path
            else:
                self.stderr.write(self.style.ERROR(f"❌ រកមិនឃើញឯកសារ Excel នៅទីតាំង៖ {file_path}"))
                return

        self.stdout.write(self.style.SUCCESS(f"📂 កំពុងអានឯកសារ Excel: {file_path}"))

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_name = '2026' if '2026' in wb.sheetnames else wb.sheetnames[0]
            sheet = wb[sheet_name]
        except Exception as ex:
            self.stderr.write(self.style.ERROR(f"❌ មិនអាចបើកឯកសារ Excel បានទេ៖ {str(ex)}"))
            return

        default_password_hash = make_password('password123')

        with transaction.atomic():
            if clear_first:
                self.stdout.write("🧹 កំពុងសម្អាតទិន្នន័យគ្រូបង្រៀនចាស់ៗ...")
                from apps.academics.models import ClassSubject, Timetable, Classroom
                Timetable.objects.all().delete()
                ClassSubject.objects.filter(teacher__isnull=False).update(teacher=None)
                Classroom.objects.filter(homeroom_teacher__isnull=False).update(homeroom_teacher=None)
                
                old_user_ids = list(Teacher.objects.filter(user__isnull=False).values_list('user_id', flat=True))
                Teacher.objects.all().delete()
                if old_user_ids:
                    User.objects.filter(id__in=old_user_ids, role=User.Role.TEACHER).delete()
                self.stdout.write(self.style.SUCCESS("✅ បានសម្អាតទិន្នន័យចាស់ៗរួចរាល់!"))

            created_count = 0
            updated_count = 0

            for r in range(8, sheet.max_row + 1):
                col1 = sheet.cell(row=r, column=1).value
                col2 = sheet.cell(row=r, column=2).value
                col3 = sheet.cell(row=r, column=3).value

                if col1 is None and col2 is None:
                    continue
                if isinstance(col1, str) and any(kw in col1 for kw in ['បញ្ឈប់', 'ក្នុងនោះ', 'ថ្នាក់ទី']):
                    break
                if isinstance(col2, str) and any(kw in col2 for kw in ['បញ្ឈប់', 'ក្នុងនោះ']):
                    break
                if not col2 or not col3:
                    continue

                t_id = str(col2).strip()
                k_name = str(col3).replace('\u200b', ' ').strip()
                gender_raw = str(sheet.cell(row=r, column=4).value or '').strip()
                gender = Teacher.Gender.FEMALE if gender_raw in ['ស', 'F', 'ស្រី', 'Female'] else Teacher.Gender.MALE

                dob = parse_flexible_date(sheet.cell(row=r, column=5).value)
                qual = str(sheet.cell(row=r, column=6).value or '').strip()
                spec_val = str(sheet.cell(row=r, column=7).value or '').strip()
                train_level = str(sheet.cell(row=r, column=8).value or '').strip()

                state_hire = parse_flexible_date(sheet.cell(row=r, column=9).value)
                perm_date = parse_flexible_date(sheet.cell(row=r, column=10).value)

                subj1 = str(sheet.cell(row=r, column=11).value or '').strip()
                subj2 = str(sheet.cell(row=r, column=12).value or '').strip()
                duty = str(sheet.cell(row=r, column=13).value or '').strip() or 'គ្រូបង្រៀន'

                cat = str(sheet.cell(row=r, column=14).value or '').strip()
                class_num = sheet.cell(row=r, column=15).value
                step_num = sheet.cell(row=r, column=16).value
                prakas_yr_raw = sheet.cell(row=r, column=17).value
                prakas_num_raw = str(sheet.cell(row=r, column=18).value or '').strip()
                order_num = sheet.cell(row=r, column=19).value
                raw_phone = sheet.cell(row=r, column=20).value
                col30_cat = str(sheet.cell(row=r, column=30).value or '').strip()

                if col30_cat:
                    prakas_cat = f"ក្របខ័ណ្ឌ {col30_cat}"
                elif cat and class_num and step_num:
                    prakas_cat = f"ក្របខ័ណ្ឌ {cat}.{class_num}.{step_num}"
                elif cat:
                    prakas_cat = f"ក្របខ័ណ្ឌ {cat}"
                else:
                    prakas_cat = ''

                prakas_yr = ''
                if isinstance(prakas_yr_raw, (datetime, date)):
                    prakas_yr = str(prakas_yr_raw.year)
                elif prakas_yr_raw:
                    prakas_yr = str(prakas_yr_raw).strip()[:4]

                prakas_number = prakas_num_raw
                if order_num:
                    prakas_number = f"{prakas_num_raw} (ល.រ {order_num})".strip()

                specialization = spec_val or subj1 or 'ទូទៅ'
                phone = format_phone_number(raw_phone)
                latin_name = transliterate_khmer_name(k_name)
                is_fee_collector = duty in ['បេឡា', 'គណនេយ្យ', 'លេខា']

                teacher, created = Teacher.objects.update_or_create(
                    teacher_id=t_id,
                    defaults={
                        'khmer_name': k_name,
                        'latin_name': latin_name or k_name,
                        'gender': gender,
                        'date_of_birth': dob,
                        'phone': phone,
                        'qualification': qual,
                        'specialization': specialization,
                        'training_level': train_level,
                        'state_hire_date': state_hire,
                        'permanent_date': perm_date,
                        'primary_subject': subj1,
                        'secondary_subject': subj2,
                        'current_duty': duty,
                        'prakas_category': prakas_cat,
                        'prakas_year': prakas_yr,
                        'prakas_number': prakas_number,
                        'base_salary': Decimal('500.00'),
                        'max_weekly_hours': 18,
                        'is_fee_collector': is_fee_collector,
                        'status': Teacher.Status.ACTIVE,
                    }
                )

                username = t_id
                user = User.objects.filter(username=username).first()
                if not user:
                    user = User.objects.create(
                        username=username,
                        password=default_password_hash,
                        role=User.Role.TEACHER,
                        khmer_name=k_name,
                        latin_name=latin_name,
                        phone=phone,
                        email=f"{username}@hunsenkkt.edu.kh",
                        is_active=True
                    )
                else:
                    user.khmer_name = k_name
                    user.latin_name = latin_name
                    user.phone = phone
                    user.role = User.Role.TEACHER
                    user.save(update_fields=['khmer_name', 'latin_name', 'phone', 'role'])

                teacher.user = user
                teacher.save(update_fields=['user'])

                if created:
                    created_count += 1
                else:
                    updated_count += 1

                self.stdout.write(f"  [{'NEW' if created else 'UPDATED'}] #{col1:03d} | ID: {t_id} | {k_name} ({latin_name}) | {duty} | {specialization} | 📞 {phone}")

        self.stdout.write(self.style.SUCCESS(
            f"\n=======================================================\n"
            f"🎉 ជោគជ័យពេញលេញក្នុងការបញ្ចូលទិន្នន័យគ្រូបង្រៀន!\n"
            f"   - បញ្ចូលថ្មី (Created): {created_count} នាក់\n"
            f"   - កែប្រែទិន្នន័យ (Updated): {updated_count} នាក់\n"
            f"   - សរុបទាំងអស់ (Total): {Teacher.objects.count()} នាក់\n"
            f"   - គណនី Login (Users): បង្កើតរួចជាស្រេច (Username: អត្តលេខ, Password: password123)\n"
            f"======================================================="
        ))
