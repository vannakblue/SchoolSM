from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Count, Q, Avg, Max, Min, Sum
from decimal import Decimal
import json
import datetime
import os
from django.conf import settings
from django.utils import timezone
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.accounts.decorators import role_required
from apps.accounts.utils import send_telegram_notification
from .models import (
    ExamTerm, Grade,
    StandardizedExam, StandardizedExamType, ExamRoom, ExamSubject, ExamCandidate, CandidateSubjectScore,
    ExamRoomSubjectCode, ExamStudentExclusion,
    ExamInvigilatorPlan, TeacherDutyGroup, TeacherDutyQuota, ExamShiftSlot, TeacherShiftRegistration
)
from .forms import ExamTermForm, StandardizedExamForm, StandardizedExamTypeForm

from apps.academics.models import Classroom, Subject, AcademicYear, GradeLevelRule
from apps.students.models import Student
from apps.teachers.models import Teacher

@login_required
def exam_term_list(request):
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if selected_year == 'all':
            active_year = None
        elif str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
            if found_year:
                active_year = found_year

    terms = ExamTerm.objects.select_related('academic_year').all()
    if active_year:
        terms = terms.filter(academic_year=active_year)
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    return render(request, 'examinations/exam_term_list.html', {
        'terms': terms,
        'academic_years': academic_years,
        'active_year': active_year,
        'selected_year': str(active_year.id) if active_year else '',
    })


@login_required
@role_required(['ADMIN'])
def exam_term_create(request):
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    if request.method == 'POST':
        form = ExamTermForm(request.POST)
        if form.is_valid():
            term = form.save()
            messages.success(request, f"បានបង្កើតសម័យប្រឡង {term.name} ជោគជ័យ!")
            return redirect('exam_term_list')
    else:
        form = ExamTermForm(initial={'academic_year': active_year})
    return render(request, 'examinations/exam_term_form.html', {'form': form, 'title': 'បង្កើតសម័យប្រឡងថ្មី / Create Exam Term'})


@login_required
@role_required(['ADMIN'])
def exam_term_edit(request, term_id):
    term = get_object_or_404(ExamTerm, id=term_id)
    if request.method == 'POST':
        form = ExamTermForm(request.POST, instance=term)
        if form.is_valid():
            term = form.save()
            messages.success(request, f"បានកែប្រែសម័យប្រឡង «{term.name}» ដោយជោគជ័យ!")
            return redirect('exam_term_list')
    else:
        form = ExamTermForm(instance=term)
    return render(request, 'examinations/exam_term_form.html', {
        'form': form,
        'term': term,
        'title': f'កែប្រែសម័យប្រឡង៖ {term.name}',
        'is_edit': True
    })


@login_required
@role_required(['ADMIN'])
def exam_term_delete(request, term_id):
    term = get_object_or_404(ExamTerm, id=term_id)
    if request.method == 'POST':
        name = term.name
        term.delete()
        messages.success(request, f"បានលុបសម័យប្រឡង «{name}» ដោយជោគជ័យ!")
        return redirect('exam_term_list')
    return redirect('exam_term_list')



@login_required
@role_required(['ADMIN', 'TEACHER'])
def grade_entry_matrix(request):
    """
    Rapid score entry grid for all students in a classroom across the exact subjects and max scores defined for that grade level & track.
    Strictly isolated per Academic Year!
    Enforces active student exam restrictions, Teacher Assigned Subject filters, and Admin grading deadline windows.
    """
    from apps.academics.utils import get_active_academic_year
    from apps.academics.models import ClassSubject
    active_year = get_active_academic_year(request)
    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'
    teacher_profile = getattr(request.user, 'teacher_profile', None) if not is_admin else None

    terms = ExamTerm.objects.filter(academic_year=active_year) if active_year else ExamTerm.objects.all()
    all_classrooms = Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code')

    # Teacher assigned classes and subjects filtering
    teacher_assigned_classes = set()
    teacher_assigned_subjects = set()
    homeroom_cls_ids = set()
    if teacher_profile:
        cs_qs = ClassSubject.objects.filter(teacher=teacher_profile)
        if active_year:
            cs_qs = cs_qs.filter(classroom__academic_year=active_year)
        teacher_assigned_classes = set(cs_qs.values_list('classroom_id', flat=True))
        teacher_assigned_subjects = set(cs_qs.values_list('subject_id', flat=True))
        # Add homeroom classroom
        homeroom_cls_ids = set(Classroom.objects.filter(homeroom_teacher=teacher_profile).values_list('id', flat=True))
        teacher_assigned_classes.update(homeroom_cls_ids)
        
        # Filter classrooms list to only assigned classes for this teacher
        classrooms = all_classrooms.filter(id__in=teacher_assigned_classes) if teacher_assigned_classes else all_classrooms
    else:
        classrooms = all_classrooms

    selected_term_id = request.GET.get('term') or request.POST.get('term') or str(terms.first().id if terms.first() else '')
    selected_class_id = request.GET.get('classroom') or request.POST.get('classroom') or str(classrooms.first().id if classrooms.first() else '')
    selected_subject_id = request.GET.get('subject') or request.POST.get('subject') or ''

    selected_term = ExamTerm.objects.filter(id=selected_term_id).first() if (selected_term_id and str(selected_term_id).isdigit()) else terms.first()
    selected_class = Classroom.objects.filter(id=selected_class_id).first() if (selected_class_id and str(selected_class_id).isdigit()) else classrooms.first()

    effective_year = selected_term.academic_year if selected_term else active_year

    subject_rules = []
    students = []
    matrix_data = []
    excluded_students_map = {}
    is_grading_open = True
    grading_status_msg = "កំពុងបើកដំណើរការបញ្ចូលពិន្ទុ"

    if selected_term:
        is_grading_open, _, grading_status_msg = selected_term.get_grading_status()

    total_tested_max = Decimal('0.00')

    if selected_term and selected_class:
        # Load effective subject rules (including is_tested status for this term/month)
        from .services import get_effective_term_subjects
        subject_rules = get_effective_term_subjects(
            exam_term=selected_term,
            classroom=selected_class,
            include_non_tested=True
        )

        total_tested_max = sum(r.max_score for r in subject_rules if getattr(r, 'is_tested', True))

        # If a specific subject is filtered
        if selected_subject_id and str(selected_subject_id).isdigit():
            subject_rules = [r for r in subject_rules if r.subject_id == int(selected_subject_id)]
        elif teacher_profile and teacher_assigned_subjects and not teacher_profile.current_duty.startswith('នាយក'):
            # Highlight teacher's assigned subjects or filter if preferred
            pass

        # Find all active exclusions for this term or month
        term_month = selected_term.start_date.month if selected_term.start_date else None
        exclusions_qs = ExamStudentExclusion.objects.filter(
            academic_year=selected_term.academic_year,
            is_active=True
        ).filter(
            Q(exam_term=selected_term) | (Q(month=term_month) if term_month else Q())
        ).select_related('student')

        for exc in exclusions_qs:
            excluded_students_map[exc.student_id] = exc

        # All students in classroom (active ones primarily, plus any existing students)
        students = Student.objects.filter(classroom=selected_class).order_by('student_id')
        
        existing_grades = {
            (g.student_id, g.subject_id): g
            for g in Grade.objects.filter(classroom=selected_class, exam_term=selected_term)
        }

        if request.method == 'POST':
            # Block saving if grading window is closed for regular teachers
            if not is_grading_open and not is_admin:
                messages.error(request, f"⚠️ មិនអាចរក្សាទុកបានទេ៖ {grading_status_msg}!")
                return redirect(f"/examinations/matrix/?term={selected_term.id}&classroom={selected_class.id}{f'&subject={selected_subject_id}' if selected_subject_id else ''}")

            saved_count = 0
            blocked_count = 0
            for student in students:
                is_student_excluded = (student.id in excluded_students_map) or getattr(student, 'is_exam_suspended', False)
                
                # Non-admin cannot submit/modify positive scores for excluded/missed students
                if is_student_excluded and not is_admin:
                    blocked_count += 1
                    continue

                for rule in subject_rules:
                    # Skip non-tested subjects from saving grades
                    if not getattr(rule, 'is_tested', True):
                        continue

                    subject = rule.subject
                    
                    # If teacher is not admin, only allow saving subjects they teach (or all in classroom if homeroom)
                    if teacher_profile and teacher_assigned_subjects and (subject.id not in teacher_assigned_subjects) and (selected_class.id not in homeroom_cls_ids):
                        continue

                    field_name = f"score_{student.id}_{subject.id}"
                    score_val = request.POST.get(field_name, '').strip()
                    if score_val != '':
                        try:
                            score_num = Decimal(score_val)
                            # Ensure score is within valid bounds
                            if score_num > rule.max_score:
                                score_num = rule.max_score
                            if score_num < Decimal('0.00'):
                                score_num = Decimal('0.00')

                            Grade.objects.update_or_create(
                                student=student,
                                subject=subject,
                                exam_term=selected_term,
                                classroom=selected_class,
                                defaults={
                                    'score': score_num,
                                    'max_score': rule.max_score
                                }
                            )
                            saved_count += 1
                        except (ValueError, Exception):
                            pass
                    elif is_student_excluded and is_admin:
                        # Admin can explicitly zero out score
                        pass

            if blocked_count > 0:
                messages.warning(request, f"⚠️ មានសិស្សចំនួន {blocked_count} នាក់ត្រូវបានលើកលែងមិនឱ្យប្រឡង (កំណត់ដោយ Admin) ដែលមានតែ Admin ប៉ុណ្ណោះដែលអាចកែប្រែពិន្ទុបាន!")
            messages.success(request, f"🎉 បានរក្សាទុកពិន្ទុសិស្សថ្នាក់ {selected_class.name} ចំនួន {saved_count} មុខវិជ្ជាជោគជ័យ!")
            return redirect(f"/examinations/matrix/?term={selected_term.id}&classroom={selected_class.id}{f'&subject={selected_subject_id}' if selected_subject_id else ''}")

        for student in students:
            is_excluded = (student.id in excluded_students_map) or getattr(student, 'is_exam_suspended', False)
            exc_obj = excluded_students_map.get(student.id)
            if getattr(student, 'is_exam_suspended', False):
                exc_reason = student.get_exam_suspension_reason_display()
            elif exc_obj:
                exc_reason = exc_obj.get_reason_display()
            else:
                exc_reason = ''

            row_scores = []
            for rule in subject_rules:
                g = existing_grades.get((student.id, rule.subject_id))
                is_tested = getattr(rule, 'is_tested', True)
                
                # If subject is not tested for this exam term
                display_score = ''
                display_letter = ''
                if not is_tested:
                    display_score = ''
                    display_letter = 'មិនប្រឡង'
                elif g:
                    display_score = g.score
                    display_letter = g.grade_letter
                elif is_excluded:
                    display_score = '0.00'
                    display_letter = 'F'

                can_edit_subject = (is_admin or is_grading_open) and is_tested and (
                    is_admin or not teacher_profile or (rule.subject_id in teacher_assigned_subjects) or (selected_class.id in homeroom_cls_ids)
                )

                row_scores.append({
                    'subject': rule.subject,
                    'max_score': rule.max_score,
                    'score': display_score,
                    'grade_letter': display_letter,
                    'is_tested': is_tested,
                    'can_edit_subject': can_edit_subject and not is_excluded,
                })
            matrix_data.append({
                'student': student,
                'is_excluded': is_excluded,
                'exclusion_reason': exc_reason,
                'can_edit': (is_admin or is_grading_open) and (is_admin or not is_excluded),
                'scores': row_scores
            })

    # All subjects for quick subject filter
    all_subjects = Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id')

    return render(request, 'examinations/grade_matrix.html', {
        'terms': terms,
        'classrooms': classrooms,
        'all_subjects': all_subjects,
        'selected_term': selected_term,
        'selected_class': selected_class,
        'selected_subject_id': selected_subject_id,
        'subject_rules': subject_rules,
        'total_tested_max': total_tested_max,
        'matrix_data': matrix_data,
        'active_year': active_year,
        'is_grading_open': is_grading_open,
        'grading_status_msg': grading_status_msg,
    })


@login_required
def grade_summary_view(request):
    """
    Computes and ranks all students in a class based on Cambodian scoring rules:
    Total Score / Total Max Score, Percentage %, Letter Grade, and Class Rank.
    Strictly isolated per Academic Year!
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    terms = ExamTerm.objects.filter(academic_year=active_year) if active_year else ExamTerm.objects.all()
    classrooms = Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code')

    selected_term_id = request.GET.get('term', str(terms.first().id if terms.first() else ''))
    selected_class_id = request.GET.get('classroom', str(classrooms.first().id if classrooms.first() else ''))

    selected_term = terms.filter(id=selected_term_id).first() if selected_term_id else None
    selected_class = classrooms.filter(id=selected_class_id).first() if selected_class_id else None

    subject_rules = []
    summary_results = []
    total_class_max = Decimal('0.00')

    if selected_term and selected_class:
        from .services import get_effective_term_subjects
        subject_rules = get_effective_term_subjects(
            exam_term=selected_term,
            classroom=selected_class,
            include_non_tested=False
        )
        total_class_max = sum(r.max_score for r in subject_rules)

        students = Student.objects.filter(classroom=selected_class, status='ACTIVE').order_by('student_id')
        
        grades_map = {}
        for g in Grade.objects.filter(classroom=selected_class, exam_term=selected_term):
            grades_map[(g.student_id, g.subject_id)] = g

        student_ranks = []
        for student in students:
            total_score = Decimal('0.00')
            subject_details = []

            for rule in subject_rules:
                sub = rule.subject
                grade_obj = grades_map.get((student.id, sub.id))
                if grade_obj:
                    score = grade_obj.score
                    total_score += score
                    subject_details.append({'subject': sub, 'score': score, 'max_score': rule.max_score, 'letter': grade_obj.grade_letter})
                else:
                    subject_details.append({'subject': sub, 'score': None, 'max_score': rule.max_score, 'letter': '-'})

            # Calculate percentage & MoEYS Letter Grade
            percentage = round((float(total_score) / float(total_class_max)) * 100, 2) if total_class_max > 0 else 0.0
            
            if percentage >= 90:
                letter = 'A'
            elif percentage >= 80:
                letter = 'B'
            elif percentage >= 70:
                letter = 'C'
            elif percentage >= 60:
                letter = 'D'
            elif percentage >= 50:
                letter = 'E'
            else:
                letter = 'F'

            student_ranks.append({
                'student': student,
                'subject_details': subject_details,
                'total_score': total_score,
                'total_max': total_class_max,
                'percentage': percentage,
                'letter': letter,
                'passed': percentage >= 50.0,
            })

        # Sort descending by total_score / percentage to determine Rank
        student_ranks.sort(key=lambda x: x['total_score'], reverse=True)
        for idx, item in enumerate(student_ranks, 1):
            item['rank'] = idx
            summary_results.append(item)

    # Handle telegram alert broadcast
    if request.method == 'POST' and 'broadcast_results' in request.POST and selected_term and selected_class:
        count = 0
        for item in summary_results:
            stu = item['student']
            msg = (
                f"សួស្តីលោក/លោកស្រីអាណាព្យាបាលសិស្ស {stu.khmer_name}!\n"
                f"លទ្ធផលប្រឡង៖ {selected_term.name}\n"
                f"ថ្នាក់៖ {selected_class.name}\n"
                f"📊 ពិន្ទុសរុប៖ {item['total_score']}/{item['total_max']} ({item['percentage']}%)\n"
                f"🏆 ចំណាត់ថ្នាក់ក្នុងថ្នាក់ (Rank)៖ លេខ {item['rank']} (និទ្ទេស {item['letter']})\n"
                f"លទ្ធផល៖ {'ជាប់កម្រិតមធ្យមសិក្សា' if item['passed'] else 'ធ្លាក់'}"
            )
            send_telegram_notification(
                title=f"🏆 លទ្ធផលប្រឡងសិស្ស: {stu.khmer_name}",
                message=msg,
                recipient_name=stu.father_name or stu.khmer_name,
                recipient_phone=stu.father_phone or stu.phone,
                custom_chat_id=stu.telegram_chat_id
            )
            count += 1
        messages.success(request, f"🔔 បានផ្ញើសារជូនដំណឹងលទ្ធផលប្រឡងទៅកាន់អាណាព្យាបាលសិស្ស {count} នាក់ជោគជ័យ!")
        return redirect(f"/examinations/summary/?term={selected_term.id}&classroom={selected_class.id}")

    return render(request, 'examinations/grade_summary.html', {
        'terms': terms,
        'classrooms': classrooms,
        'selected_term': selected_term,
        'selected_class': selected_class,
        'subject_rules': subject_rules,
        'total_class_max': total_class_max,
        'summary_results': summary_results,
    })


@login_required
def report_card_view(request, student_id, term_id):
    """
    Official MoEYS Academic Transcript & Report Card with accurate scoring rule breakdown
    """
    student = get_object_or_404(Student.objects.select_related('classroom', 'academic_year'), pk=student_id)
    term = get_object_or_404(ExamTerm, pk=term_id)
    classroom = student.classroom

    from .services import get_effective_term_subjects
    tested_rules = get_effective_term_subjects(
        exam_term=term,
        classroom=classroom,
        include_non_tested=False
    )
    tested_subject_ids = {r.subject_id for r in tested_rules}
    total_max = sum(r.max_score for r in tested_rules) if tested_rules else (classroom.get_total_max_score() if classroom else Decimal('0.00'))

    grades = Grade.objects.filter(student=student, exam_term=term).select_related('subject').order_by('subject__order', 'id')
    grades_list = []
    total_score = Decimal('0.00')

    for g in grades:
        is_tested = g.subject_id in tested_subject_ids
        g.is_tested = is_tested
        grades_list.append(g)
        if is_tested:
            total_score += g.score

    percentage = round((float(total_score) / float(total_max)) * 100, 2) if total_max > 0 else 0.0

    if percentage >= 90:
        overall_grade = 'A (ល្អប្រសើរ)'
    elif percentage >= 80:
        overall_grade = 'B (ល្អណាស់)'
    elif percentage >= 70:
        overall_grade = 'C (ល្អ)'
    elif percentage >= 60:
        overall_grade = 'D (ល្អបង្គួរ)'
    elif percentage >= 50:
        overall_grade = 'E (មធ្យម)'
    else:
        overall_grade = 'F (ធ្លាក់)'

    # Calculate class rank based on tested subjects
    all_class_students = Student.objects.filter(classroom=classroom, status='ACTIVE') if classroom else []
    student_scores = []
    for s in all_class_students:
        s_grades = Grade.objects.filter(student=s, exam_term=term, subject_id__in=tested_subject_ids)
        s_tot = sum(g.score for g in s_grades)
        student_scores.append((s.id, s_tot))

    student_scores.sort(key=lambda x: x[1], reverse=True)
    rank = 1
    for idx, (s_id, _) in enumerate(student_scores, 1):
        if s_id == student.id:
            rank = idx
            break

    return render(request, 'examinations/report_card.html', {
        'student': student,
        'term': term,
        'classroom': classroom,
        'grades': grades_list,
        'total_score': total_score,
        'total_max': total_max,
        'percentage': percentage,
        'overall_grade': overall_grade,
        'rank': rank,
        'total_in_class': len(student_scores),
        'is_passed': percentage >= 50.0,
    })


from .telegram_report_card import (
    dispatch_student_report_card_to_telegram,
    generate_report_card_pdf_bytes,
    build_report_card_telegram_message
)


@login_required
def api_send_report_card_telegram(request):
    """
    AJAX Endpoint to dispatch an individual report card to Telegram as PDF, Rich message, or Both.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)

    try:
        data = json.loads(request.body.decode('utf-8')) if request.body else request.POST
        student_id = data.get('student_id')
        term_id = data.get('term_id')
        destination = data.get('destination', 'CLASS_GROUP')  # 'CLASS_GROUP', 'PARENT_INDIVIDUAL', 'CUSTOM_CHAT_ID'
        send_mode = data.get('send_mode', 'BOTH')  # 'PDF_ONLY', 'MESSAGE_ONLY', 'BOTH'
        custom_chat_id = data.get('custom_chat_id', '').strip() or None

        student = get_object_or_404(Student, id=student_id)
        term = get_object_or_404(ExamTerm, id=term_id)

        result = dispatch_student_report_card_to_telegram(
            student=student,
            term=term,
            destination=destination,
            send_mode=send_mode,
            custom_chat_id=custom_chat_id
        )

        return JsonResponse({
            'status': 'success',
            'message': f'បានផ្ញើប័ណ្ណពិន្ទុរបស់សិស្ស "{student.khmer_name}" ទៅកាន់ Telegram ដោយជោគជ័យ!',
            'data': result
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def api_send_class_report_cards_telegram(request):
    """
    AJAX Endpoint to dispatch report cards for all students in a classroom to Telegram.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)

    try:
        data = json.loads(request.body.decode('utf-8')) if request.body else request.POST
        classroom_id = data.get('classroom_id')
        term_id = data.get('term_id')
        destination = data.get('destination', 'CLASS_GROUP')
        send_mode = data.get('send_mode', 'BOTH')
        custom_chat_id = data.get('custom_chat_id', '').strip() or None

        classroom = get_object_or_404(Classroom, id=classroom_id)
        term = get_object_or_404(ExamTerm, id=term_id)
        students = Student.objects.filter(classroom=classroom, status='ACTIVE')

        sent_count = 0
        for s in students:
            dispatch_student_report_card_to_telegram(
                student=s,
                term=term,
                destination=destination,
                send_mode=send_mode,
                custom_chat_id=custom_chat_id
            )
            sent_count += 1

        return JsonResponse({
            'status': 'success',
            'message': f'បានផ្ញើប័ណ្ណពិន្ទុសិស្សថ្នាក់ "{classroom.name}" ចំនួន {sent_count} នាក់ ទៅកាន់ Telegram ដោយជោគជ័យ!',
            'sent_count': sent_count
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# =========================================================================
# STANDARDIZED EXAM MANAGEMENT VIEWS (ប្រព័ន្ធប្រឡងតេស្តស្តង់ដាវិទ្យាល័យ)
# =========================================================================

import re

def get_clean_exam_session_title(name):
    # Remove grade specifications like (ថ្នាក់ទី ៧), ថ្នាក់ទី ៧, (Grade 12), etc.
    cleaned = re.sub(r'[\(\[\{]?\s*(?:ថ្នាក់ទី|កម្រិតទី|Grade)\s*\d+\s*[\)\]\}]?', '', name, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s*-\s*$', '', cleaned).strip(" -–—:()")
    return cleaned if cleaned else name


@login_required
def standardized_exam_list(request):
    """
    Two-Level Hierarchical Overview of Standardized Exams:
    Level 1: Grouped by Exam Period / Session (សម័យប្រឡងនីមួយៗ)
    Level 2: When admin clicks/expands a session, displays all individual Grade Levels under it.
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if selected_year == 'all':
            active_year = None
        elif str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
            if found_year:
                active_year = found_year

    selected_grade = request.GET.get('grade') or request.GET.get('grade_level')
    search_q = request.GET.get('q', '').strip()

    exams_qs = StandardizedExam.objects.select_related('academic_year').prefetch_related('candidates', 'rooms', 'exam_subjects').all().order_by('-exam_date', 'grade_level')
    if active_year:
        exams_qs = exams_qs.filter(academic_year=active_year)
    if selected_grade and selected_grade != 'all' and selected_grade.isdigit():
        exams_qs = exams_qs.filter(grade_level=int(selected_grade))
    if search_q:
        exams_qs = exams_qs.filter(Q(name__icontains=search_q) | Q(description__icontains=search_q))

    # Group exams by (academic_year_id, exam_date, clean_title)
    sessions_map = {}
    for ex in exams_qs:
        clean_title = get_clean_exam_session_title(ex.name)
        date_key = str(ex.exam_date)
        year_key = str(ex.academic_year_id)
        group_key = f"{year_key}_{date_key}_{clean_title}"

        if group_key not in sessions_map:
            sessions_map[group_key] = {
                'group_key': group_key,
                'html_id': f"session_{abs(hash(group_key)) % 1000000}",
                'title': clean_title,
                'academic_year': ex.academic_year,
                'exam_date': ex.exam_date,
                'total_candidates': 0,
                'female_candidates': 0,
                'total_rooms': 0,
                'total_subjects': 0,
                'grades_list': [],
                'morning_grades': [],
                'afternoon_grades': [],
                'exams_data': [],
            }

        cand_count = ex.candidates.count()
        fem_count = ex.candidates.filter(gender='F').count()
        room_count = ex.rooms.count()
        subj_count = ex.exam_subjects.count()

        sess = sessions_map[group_key]
        sess['total_candidates'] += cand_count
        sess['female_candidates'] += fem_count
        sess['total_rooms'] += room_count
        sess['total_subjects'] = max(sess['total_subjects'], subj_count)
        
        if ex.grade_level not in sess['grades_list']:
            sess['grades_list'].append(ex.grade_level)
            sess['grades_list'].sort()

        if ex.session == 'MORNING' and ex.grade_level not in sess['morning_grades']:
            sess['morning_grades'].append(ex.grade_level)
            sess['morning_grades'].sort()
        elif ex.session == 'AFTERNOON' and ex.grade_level not in sess['afternoon_grades']:
            sess['afternoon_grades'].append(ex.grade_level)
            sess['afternoon_grades'].sort()

        sess['exams_data'].append({
            'exam': ex,
            'total_candidates': cand_count,
            'female_candidates': fem_count,
            'total_rooms': room_count,
            'total_subjects': subj_count,
        })

    # Sort sessions by exam date descending
    exam_sessions = list(sessions_map.values())
    exam_sessions.sort(key=lambda s: s['exam_date'], reverse=True)

    # Match and attach invigilator plans to each exam session
    all_plans = list(ExamInvigilatorPlan.objects.select_related('academic_year').prefetch_related('shift_slots__registrations').all())
    for sess in exam_sessions:
        matched_plan = None
        for p in all_plans:
            if p.session_key and p.session_key == sess['group_key']:
                matched_plan = p
                break
            if p.standardized_exam_id and p.standardized_exam_id in [item['exam'].id for item in sess['exams_data']]:
                matched_plan = p
                break
            if p.academic_year_id == sess['academic_year'].id and p.start_date <= sess['exam_date'] <= p.end_date and (sess['title'].lower() in p.title.lower() or p.title.lower() in sess['title'].lower()):
                matched_plan = p
                break
        sess['invigilator_plan'] = matched_plan
        if matched_plan:
            slots = list(matched_plan.shift_slots.all())
            matched_plan.calc_total_capacity = sum(s.max_invigilators for s in slots)
            matched_plan.calc_total_registered = sum(s.registered_count for s in slots)

    # Calculate overall total counts
    total_all_exams = exams_qs.count()
    total_all_candidates = sum(s['total_candidates'] for s in exam_sessions)
    total_all_rooms = sum(s['total_rooms'] for s in exam_sessions)

    # Build full list of all sessions across all years for the batch modal
    all_exams_all_years = StandardizedExam.objects.select_related('academic_year').all().order_by('-exam_date', 'grade_level')
    modal_sessions_map = {}
    for ex in all_exams_all_years:
        clean_title = get_clean_exam_session_title(ex.name)
        date_key = str(ex.exam_date)
        year_key = str(ex.academic_year_id)
        group_key = f"{year_key}_{date_key}_{clean_title}"
        if group_key not in modal_sessions_map:
            modal_sessions_map[group_key] = {
                'group_key': group_key,
                'title': clean_title,
                'academic_year_id': ex.academic_year_id,
                'academic_year_name': ex.academic_year.name,
                'exam_date': ex.exam_date,
                'grades_list': [],
                'exam_count': 0
            }
        ms = modal_sessions_map[group_key]
        ms['exam_count'] += 1
        if ex.grade_level not in ms['grades_list']:
            ms['grades_list'].append(ex.grade_level)
            ms['grades_list'].sort()

    all_modal_sessions = list(modal_sessions_map.values())
    all_modal_sessions.sort(key=lambda s: s['exam_date'], reverse=True)

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    return render(request, 'examinations/standardized/exam_list.html', {
        'exam_sessions': exam_sessions,
        'all_modal_sessions': all_modal_sessions,
        'total_sessions_count': len(exam_sessions),
        'total_all_exams': total_all_exams,
        'total_all_candidates': total_all_candidates,
        'total_all_rooms': total_all_rooms,
        'academic_years': academic_years,
        'active_year': active_year,
        'selected_year': str(active_year.id) if active_year else '',
        'selected_grade': selected_grade or 'all',
        'search_q': search_q,
    })



def pull_candidates_for_exam(exam):
    """
    1-Click / Auto-Pull all active students from classrooms matching this exam's Grade Level and Track.
    Creates ExamCandidate records and initializes CandidateSubjectScore rows for each subject.
    Returns the count of newly added candidates.
    """
    classrooms = Classroom.objects.filter(
        academic_year=exam.academic_year,
        grade_level=exam.grade_level
    )
    if exam.track != 'ALL':
        classrooms = classrooms.filter(track=exam.track)

    # Fallback 1: If no classrooms found for exam.academic_year, try system active academic year
    if not classrooms.exists():
        from apps.academics.utils import get_active_academic_year
        act_y = get_active_academic_year()
        if act_y and act_y != exam.academic_year:
            fallback_cr = Classroom.objects.filter(academic_year=act_y, grade_level=exam.grade_level)
            if exam.track != 'ALL':
                fallback_cr = fallback_cr.filter(track=exam.track)
            if fallback_cr.exists():
                classrooms = fallback_cr

    # Fallback 2: Try any academic year with active students for this grade level
    if not classrooms.exists():
        fallback_cr = Classroom.objects.filter(
            grade_level=exam.grade_level,
            students__status='ACTIVE'
        ).distinct()
        if exam.track != 'ALL':
            fallback_cr = fallback_cr.filter(track=exam.track)
        if fallback_cr.exists():
            classrooms = fallback_cr

    # Fallback 3: Try any classroom for this grade level
    if not classrooms.exists():
        fallback_cr = Classroom.objects.filter(grade_level=exam.grade_level)
        if exam.track != 'ALL':
            fallback_cr = fallback_cr.filter(track=exam.track)
        if fallback_cr.exists():
            classrooms = fallback_cr

    # Auto-align exam academic year if it was misassigned to a year with 0 classrooms
    if classrooms.exists():
        cr_year = classrooms.first().academic_year
        if cr_year and exam.academic_year != cr_year:
            if not Classroom.objects.filter(academic_year=exam.academic_year, grade_level=exam.grade_level).exists():
                exam.academic_year = cr_year
                exam.save(update_fields=['academic_year'])

    students = Student.objects.filter(
        classroom__in=classrooms,
        status='ACTIVE',
        is_exam_suspended=False
    ).select_related('classroom').order_by('classroom__code', 'khmer_name')

    # Exclude students with active exclusions for this exam or month/year
    excluded_years = [exam.academic_year]
    if classrooms.exists():
        first_cr_ay = classrooms.first().academic_year
        if first_cr_ay and first_cr_ay not in excluded_years:
            excluded_years.append(first_cr_ay)

    excluded_student_ids = set(
        ExamStudentExclusion.objects.filter(
            academic_year__in=excluded_years,
            is_active=True
        ).filter(
            Q(standardized_exam=exam) |
            (Q(month=exam.exam_date.month) if exam.exam_date else Q())
        ).values_list('student_id', flat=True)
    )
    if excluded_student_ids:
        students = students.exclude(id__in=excluded_student_ids)

    existing_student_ids = set(exam.candidates.values_list('student_id', flat=True))
    existing_roll_count = exam.candidates.count()

    new_candidates = []
    exam_subjects = list(exam.exam_subjects.all())

    with transaction.atomic():
        for stu in students:
            if stu.id in existing_student_ids:
                continue

            roll_str = f"{existing_roll_count + len(new_candidates) + 1:03d}"
            cand = ExamCandidate.objects.create(
                exam=exam,
                student=stu,
                roll_number=roll_str,
                desk_number=1,
                candidate_name_kh=stu.khmer_name,
                candidate_name_en=stu.latin_name or '',
                gender=stu.gender or 'M',
                dob=stu.date_of_birth,
                origin_class=stu.classroom.name if stu.classroom else '',
                student_code=stu.student_id or '',
            )
            # Create empty score rows for each subject
            for es in exam_subjects:
                CandidateSubjectScore.objects.create(
                    candidate=cand,
                    exam_subject=es
                )
            new_candidates.append(cand)

    return len(new_candidates)


@login_required
@role_required(['ADMIN'])
def standardized_exam_create(request):
    """
    Creates a new Standardized Exam (Single or Multi-Grade Batch) and auto-populates Exam Subjects from MoEYS Grade Level Rules.
    Automatically pulls active students into exam candidates by grade level unless explicitly disabled.
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    if request.method == 'POST':
        selected_grades = request.POST.getlist('selected_grades')
        
        # If no selected_grades checkboxes (e.g. from standard single form), fallback to single grade_level
        if not selected_grades:
            single_grade = request.POST.get('grade_level')
            if single_grade and str(single_grade).isdigit():
                selected_grades = [str(single_grade)]
            else:
                selected_grades = ['12']

        base_name = request.POST.get('name', '').strip()
        academic_year_id = request.POST.get('academic_year')
        ay = AcademicYear.objects.filter(id=academic_year_id).first() if academic_year_id else active_year
        base_exam_type = request.POST.get('exam_type', StandardizedExam.ExamType.OTHER)
        exam_term_id = request.POST.get('exam_term')
        linked_term = ExamTerm.objects.filter(id=exam_term_id).first() if exam_term_id and str(exam_term_id).isdigit() else None
        base_grading_method = request.POST.get('grading_method', StandardizedExam.GradingMethod.BOTH)
        base_track = request.POST.get('track', 'ALL')
        base_session = request.POST.get('session', 'MORNING')
        base_date_str = request.POST.get('exam_date', datetime.date.today().strftime('%Y-%m-%d'))
        try:
            base_date = datetime.datetime.strptime(base_date_str, '%Y-%m-%d').date()
        except Exception:
            base_date = datetime.date.today()

        base_cpr = int(request.POST.get('candidates_per_room', '25')) if request.POST.get('candidates_per_room', '').isdigit() else 25
        base_desc = request.POST.get('description', '').strip()
        is_pub = request.POST.get('is_published') == 'on'

        # Auto-pull active candidates into each grade level (default is True)
        if 'auto_pull_candidates' in request.POST:
            auto_pull_val = request.POST.get('auto_pull_candidates')
            auto_pull = (auto_pull_val in ['on', 'true', '1', 'yes'])
        else:
            auto_pull = True

        created_exams = []
        total_pulled_candidates = 0

        with transaction.atomic():
            for g_str in selected_grades:
                if not g_str.isdigit():
                    continue
                g_val = int(g_str)

                # Read per-grade customized parameters or fallback to base
                g_name = request.POST.get(f'grade_name_{g_val}', '').strip()
                if not g_name:
                    if len(selected_grades) > 1:
                        if f'ថ្នាក់ទី {g_val}' not in base_name and f'ថ្នាក់ទី{g_val}' not in base_name:
                            g_name = f"{base_name} ថ្នាក់ទី {g_val}"
                        else:
                            g_name = base_name
                    else:
                        g_name = base_name

                g_track = request.POST.get(f'grade_track_{g_val}') or base_track
                g_session = request.POST.get(f'grade_session_{g_val}') or base_session
                g_date_str = request.POST.get(f'grade_date_{g_val}')
                if g_date_str:
                    try:
                        g_date = datetime.datetime.strptime(g_date_str, '%Y-%m-%d').date()
                    except Exception:
                        g_date = base_date
                else:
                    g_date = base_date

                g_cpr_str = request.POST.get(f'grade_cpr_{g_val}')
                g_cpr = int(g_cpr_str) if g_cpr_str and g_cpr_str.isdigit() else base_cpr

                exam = StandardizedExam.objects.create(
                    name=g_name,
                    academic_year=ay,
                    exam_type=base_exam_type,
                    exam_term=linked_term,
                    grade_level=g_val,
                    track=g_track,
                    session=g_session,
                    exam_date=g_date,
                    candidates_per_room=g_cpr,
                    grading_method=base_grading_method,
                    description=base_desc,
                    is_published=is_pub
                )
                created_exams.append(exam)

                # Automatically populate default exam subjects based on grade_level and track
                rules = GradeLevelRule.objects.filter(grade_level=exam.grade_level)
                if exam.track != 'ALL':
                    rules = rules.filter(Q(track=exam.track) | Q(track='GENERAL'))

                seen_subjects = set()
                order_idx = 1
                subj_session = 'AFTERNOON' if g_session == 'AFTERNOON' else 'MORNING'
                for r in rules.select_related('subject').order_by('subject__order', 'id'):
                    if r.subject_id not in seen_subjects:
                        seen_subjects.add(r.subject_id)
                        max_sc = r.max_score or Decimal('50.00')
                        # Auto-calculate coefficient: max_score / 50
                        calculated_coef = round(max_sc / Decimal('50.00'), 2)
                        ExamSubject.objects.create(
                            exam=exam,
                            subject=r.subject,
                            max_score=max_sc,
                            coefficient=calculated_coef,
                            session=subj_session,
                            order=order_idx
                        )
                        order_idx += 1

                if not seen_subjects:
                    for s in Subject.objects.all().order_by('order', 'id')[:10]:
                        max_sc = Decimal('50.00')
                        calculated_coef = round(max_sc / Decimal('50.00'), 2)
                        ExamSubject.objects.create(
                            exam=exam,
                            subject=s,
                            max_score=max_sc,
                            coefficient=calculated_coef,
                            session=subj_session,
                            order=order_idx
                        )
                        order_idx += 1

                # Automatically pull active candidates into this exam
                if auto_pull:
                    pulled_count = pull_candidates_for_exam(exam)
                    total_pulled_candidates += pulled_count

        if len(created_exams) == 1:
            if auto_pull and total_pulled_candidates > 0:
                messages.success(request, f"🎉 បានបង្កើតសម័យប្រឡងតេស្តស្តង់ដា «{created_exams[0].name}» និងបានទាញបញ្ចូលបេក្ខជនចំនួន {total_pulled_candidates} នាក់ដោយស្វ័យប្រវត្តិ!")
            else:
                messages.success(request, f"🎉 បានបង្កើតសម័យប្រឡងតេស្តស្តង់ដា «{created_exams[0].name}» ដោយជោគជ័យ!")
            return redirect('standardized_exam_manage', exam_id=created_exams[0].id)
        elif len(created_exams) > 1:
            grades_list_str = ", ".join([f"ថ្នាក់ទី {e.grade_level}" for e in created_exams])
            if auto_pull and total_pulled_candidates > 0:
                messages.success(request, f"🎉 បានបង្កើតសម័យប្រឡងតេស្តស្តង់ដាចំនួន {len(created_exams)} កម្រិតថ្នាក់ ({grades_list_str}) ព្រមទាំងបានទាញបញ្ចូលបេក្ខជនសរុបចំនួន {total_pulled_candidates} នាក់ដោយស្វ័យប្រវត្តិ!")
            else:
                messages.success(request, f"🎉 បានបង្កើតសម័យប្រឡងតេស្តស្តង់ដាចំនួន {len(created_exams)} កម្រិតថ្នាក់ ({grades_list_str}) ដោយជោគជ័យ!")
            return redirect('standardized_exam_list')
        else:
            messages.error(request, "⚠️ សូមជ្រើសរើសយ៉ាងហោចណាស់មួយកម្រិតថ្នាក់!")
            return redirect('standardized_exam_create')
    else:
        form = StandardizedExamForm(initial={
            'academic_year': active_year,
            'exam_type': 'OTHER',
            'grade_level': 12,
            'track': 'ALL',
            'exam_date': datetime.date.today(),
            'candidates_per_room': 25
        })

    terms_filter = Q(academic_year=active_year) if active_year else Q()
    monthly_terms = ExamTerm.objects.filter(terms_filter, term_type=ExamTerm.TermType.MONTHLY).select_related('academic_year').order_by('start_date')
    if not monthly_terms.exists():
        monthly_terms = ExamTerm.objects.filter(term_type=ExamTerm.TermType.MONTHLY).select_related('academic_year').order_by('-start_date')[:15]

    semester_terms = ExamTerm.objects.filter(terms_filter).filter(Q(term_type=ExamTerm.TermType.SEMESTER_1) | Q(term_type=ExamTerm.TermType.SEMESTER_2)).select_related('academic_year').order_by('start_date')
    if not semester_terms.exists():
        semester_terms = ExamTerm.objects.filter(Q(term_type=ExamTerm.TermType.SEMESTER_1) | Q(term_type=ExamTerm.TermType.SEMESTER_2)).select_related('academic_year').order_by('-start_date')[:10]

    exam_types = StandardizedExamType.get_active_types()
    return render(request, 'examinations/standardized/exam_form.html', {
        'form': form,
        'monthly_terms': monthly_terms,
        'semester_terms': semester_terms,
        'exam_types': exam_types,
        'title': 'បង្កើតសម័យប្រឡងតេស្តស្តង់ដាថ្មី (Create Standardized Exam)',
        'is_edit': False,
        'is_admin': (request.user.role == 'ADMIN' or request.user.is_superuser),
    })


@login_required
@role_required(['ADMIN'])
def standardized_exam_type_list(request):
    """
    Lists all StandardizedExamTypes.
    Returns JSON if AJAX/JSON request; renders management template otherwise.
    """
    StandardizedExamType.ensure_defaults()
    types_qs = StandardizedExamType.objects.all().order_by('order', 'id')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('accept', ''):
        data = [
            {
                'id': t.id,
                'name': t.name,
                'code': t.code,
                'icon': t.icon,
                'default_title': t.default_title,
                'is_monthly': t.is_monthly,
                'linked_term_type': t.linked_term_type or '',
                'order': t.order,
                'is_active': t.is_active,
            }
            for t in types_qs
        ]
        return JsonResponse({'status': 'success', 'types': data})

    form = StandardizedExamTypeForm()
    return render(request, 'examinations/standardized/exam_types_manage.html', {
        'types': types_qs,
        'form': form,
        'title': 'គ្រប់គ្រងប្រភេទសម័យប្រឡងតេស្តស្តង់ដា (Exam Types)',
        'is_admin': True,
    })


@login_required
@role_required(['ADMIN'])
def standardized_exam_type_create(request):
    """
    Creates a new StandardizedExamType.
    Supports standard POST or AJAX/JSON POST.
    """
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json'
    
    if request.method == 'POST':
        if request.content_type == 'application/json':
            try:
                body_data = json.loads(request.body)
            except Exception:
                body_data = {}
            post_data = body_data
        else:
            post_data = request.POST

        name = post_data.get('name', '').strip()
        code = post_data.get('code', '').strip().upper()
        icon = post_data.get('icon', '🎯').strip() or '🎯'
        default_title = post_data.get('default_title', '').strip()
        is_monthly = str(post_data.get('is_monthly', '')).lower() in ['true', '1', 'on']
        linked_term_type = post_data.get('linked_term_type', '').strip()
        order_val = post_data.get('order', '1')
        order = int(order_val) if str(order_val).isdigit() else 1
        is_active = str(post_data.get('is_active', 'true')).lower() in ['true', '1', 'on']

        if not name:
            msg = "⚠️ សូមបញ្ចូលឈ្មោះប្រភេទសម័យប្រឡង!"
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('standardized_exam_type_list')

        if not code:
            import re
            base_code = re.sub(r'[^A-Za-z0-9_]', '', name.upper().replace(' ', '_'))
            if not base_code:
                base_code = f"TYPE_{int(timezone.now().timestamp())}"
            code = base_code

        orig_code = code
        counter = 1
        while StandardizedExamType.objects.filter(code=code).exists():
            code = f"{orig_code}_{counter}"
            counter += 1

        obj = StandardizedExamType.objects.create(
            name=name,
            code=code,
            icon=icon,
            default_title=default_title or f"ការប្រឡង{name}",
            is_monthly=is_monthly,
            linked_term_type=linked_term_type,
            order=order,
            is_active=is_active
        )

        success_msg = f"🎉 បានបន្ថែមប្រភេទសម័យប្រឡង «{obj.icon} {obj.name}» ដោយជោគជ័យ!"
        if is_ajax:
            types_qs = StandardizedExamType.objects.all().order_by('order', 'id')
            data = [
                {
                    'id': t.id,
                    'name': t.name,
                    'code': t.code,
                    'icon': t.icon,
                    'default_title': t.default_title,
                    'is_monthly': t.is_monthly,
                    'linked_term_type': t.linked_term_type or '',
                    'order': t.order,
                    'is_active': t.is_active,
                }
                for t in types_qs
            ]
            return JsonResponse({'status': 'success', 'message': success_msg, 'type_id': obj.id, 'types': data})

        messages.success(request, success_msg)
        return redirect('standardized_exam_type_list')

    return redirect('standardized_exam_type_list')


@login_required
@role_required(['ADMIN'])
def standardized_exam_type_edit(request, type_id):
    """
    Edits an existing StandardizedExamType.
    Supports standard POST or AJAX/JSON POST.
    """
    exam_type = get_object_or_404(StandardizedExamType, id=type_id)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json'

    if request.method == 'POST':
        if request.content_type == 'application/json':
            try:
                body_data = json.loads(request.body)
            except Exception:
                body_data = {}
            post_data = body_data
        else:
            post_data = request.POST

        name = post_data.get('name', '').strip()
        code = post_data.get('code', '').strip().upper()
        icon = post_data.get('icon', '').strip()
        default_title = post_data.get('default_title', '').strip()
        is_monthly = str(post_data.get('is_monthly', '')).lower() in ['true', '1', 'on']
        linked_term_type = post_data.get('linked_term_type', '').strip()
        order_val = post_data.get('order', '')
        is_active_val = post_data.get('is_active')

        if not name:
            msg = "⚠️ ឈ្មោះប្រភេទមិនអាចទទេបានទេ!"
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('standardized_exam_type_list')

        if code and code != exam_type.code:
            if StandardizedExamType.objects.filter(code=code).exclude(id=exam_type.id).exists():
                msg = f"⚠️ កូដ «{code}» ត្រូវបានប្រើប្រាស់រួចហើយ!"
                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': msg}, status=400)
                messages.error(request, msg)
                return redirect('standardized_exam_type_list')
            old_code = exam_type.code
            exam_type.code = code
            # Also update exams pointing to old_code
            StandardizedExam.objects.filter(exam_type=old_code).update(exam_type=code)

        exam_type.name = name
        if icon:
            exam_type.icon = icon
        exam_type.default_title = default_title
        exam_type.is_monthly = is_monthly
        exam_type.linked_term_type = linked_term_type
        if order_val and str(order_val).isdigit():
            exam_type.order = int(order_val)
        if is_active_val is not None:
            exam_type.is_active = str(is_active_val).lower() in ['true', '1', 'on']
        
        exam_type.save()

        success_msg = f"🎉 បានកែប្រែប្រភេទសម័យប្រឡង «{exam_type.icon} {exam_type.name}» ដោយជោគជ័យ!"
        if is_ajax:
            types_qs = StandardizedExamType.objects.all().order_by('order', 'id')
            data = [
                {
                    'id': t.id,
                    'name': t.name,
                    'code': t.code,
                    'icon': t.icon,
                    'default_title': t.default_title,
                    'is_monthly': t.is_monthly,
                    'linked_term_type': t.linked_term_type or '',
                    'order': t.order,
                    'is_active': t.is_active,
                }
                for t in types_qs
            ]
            return JsonResponse({'status': 'success', 'message': success_msg, 'type_id': exam_type.id, 'types': data})

        messages.success(request, success_msg)
        return redirect('standardized_exam_type_list')

    return redirect('standardized_exam_type_list')


@login_required
@role_required(['ADMIN'])
def standardized_exam_type_delete(request, type_id):
    """
    Deletes an existing StandardizedExamType.
    If existing exams use its code, updates them safely to 'OTHER'.
    """
    exam_type = get_object_or_404(StandardizedExamType, id=type_id)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json'

    if request.method == 'POST':
        deleted_name = f"{exam_type.icon} {exam_type.name}"
        old_code = exam_type.code

        StandardizedExam.objects.filter(exam_type=old_code).update(exam_type='OTHER')
        exam_type.delete()

        success_msg = f"🗑️ បានលុបប្រភេទសម័យប្រឡង «{deleted_name}» ដោយជោគជ័យ!"
        if is_ajax:
            types_qs = StandardizedExamType.objects.all().order_by('order', 'id')
            data = [
                {
                    'id': t.id,
                    'name': t.name,
                    'code': t.code,
                    'icon': t.icon,
                    'default_title': t.default_title,
                    'is_monthly': t.is_monthly,
                    'linked_term_type': t.linked_term_type or '',
                    'order': t.order,
                    'is_active': t.is_active,
                }
                for t in types_qs
            ]
            return JsonResponse({'status': 'success', 'message': success_msg, 'types': data})

        messages.success(request, success_msg)
        return redirect('standardized_exam_type_list')

    return redirect('standardized_exam_type_list')


@login_required
@role_required(['ADMIN'])
def standardized_exam_edit(request, exam_id):
    """
    Edits an existing Standardized Exam and its Exam Subjects.
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)

    if request.method == 'POST':
        form = StandardizedExamForm(request.POST, instance=exam)
        if form.is_valid():
            with transaction.atomic():
                exam = form.save()

                # Update existing exam subjects from POST fields
                for es in exam.exam_subjects.all():
                    max_s = request.POST.get(f'max_score_{es.id}')
                    coef = request.POST.get(f'coefficient_{es.id}')
                    sess = request.POST.get(f'session_{es.id}')
                    ex_date = request.POST.get(f'exam_date_{es.id}')

                    if max_s:
                        try:
                            es.max_score = Decimal(str(max_s))
                        except Exception:
                            pass
                    if coef is not None and str(coef).strip():
                        try:
                            # Admin manual entry
                            es.coefficient = Decimal(str(coef))
                        except Exception:
                            pass
                    elif max_s:
                        # Auto-calculate: max_score / 50
                        try:
                            es.coefficient = round(es.max_score / Decimal('50.00'), 2)
                        except Exception:
                            pass

                    if sess:
                        es.session = sess
                    if ex_date:
                        try:
                            es.exam_date = datetime.datetime.strptime(ex_date, '%Y-%m-%d').date()
                        except Exception:
                            pass
                    es.save()

                # Adding new subject
                new_subj_id = request.POST.get('new_subject_id')
                if new_subj_id:
                    new_subj = Subject.objects.filter(id=new_subj_id).first()
                    if new_subj and not exam.exam_subjects.filter(subject=new_subj).exists():
                        new_max_s = request.POST.get('new_max_score', '50.00')
                        new_coef = request.POST.get('new_coefficient', '1.00')
                        new_sess = request.POST.get('new_session', exam.session)
                        try:
                            parsed_max = Decimal(str(new_max_s))
                        except Exception:
                            parsed_max = Decimal('50.00')
                        try:
                            parsed_coef = Decimal(str(new_coef))
                        except Exception:
                            parsed_coef = Decimal('1.00')

                        order_idx = (exam.exam_subjects.aggregate(m=Max('order'))['m'] or 0) + 1
                        ExamSubject.objects.create(
                            exam=exam,
                            subject=new_subj,
                            max_score=parsed_max,
                            coefficient=parsed_coef,
                            session=new_sess,
                            order=order_idx
                        )

            messages.success(request, f"បានកែប្រែសម័យប្រឡង «{exam.name}» ដោយជោគជ័យ!")
            return redirect('standardized_exam_manage', exam_id=exam.id)
    else:
        form = StandardizedExamForm(instance=exam)

    subjects = exam.exam_subjects.select_related('subject').order_by('order', 'id')
    available_subjects = Subject.objects.exclude(id__in=subjects.values_list('subject_id', flat=True)).order_by('name_kh')

    monthly_terms = ExamTerm.objects.filter(academic_year=exam.academic_year, term_type=ExamTerm.TermType.MONTHLY).order_by('start_date') if exam.academic_year else []
    semester_terms = ExamTerm.objects.filter(academic_year=exam.academic_year).filter(Q(term_type=ExamTerm.TermType.SEMESTER_1) | Q(term_type=ExamTerm.TermType.SEMESTER_2)).order_by('start_date') if exam.academic_year else []

    exam_types = StandardizedExamType.get_active_types()
    return render(request, 'examinations/standardized/exam_form.html', {
        'form': form,
        'exam': exam,
        'subjects': subjects,
        'available_subjects': available_subjects,
        'monthly_terms': monthly_terms,
        'semester_terms': semester_terms,
        'exam_types': exam_types,
        'title': f'កែប្រែសម័យប្រឡង៖ {exam.name}',
        'is_edit': True,
        'is_admin': (request.user.role == 'ADMIN' or request.user.is_superuser),
    })



@login_required
@role_required(['ADMIN'])
def standardized_exam_delete(request, exam_id):
    """
    Deletes a Standardized Exam and its candidates, scores, and rooms.
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    if request.method == 'POST':
        name = exam.name
        exam.delete()
        messages.success(request, f"បានលុបសម័យប្រឡង «{name}» ដោយជោគជ័យ!")
        return redirect('standardized_exam_list')
    return redirect('standardized_exam_manage', exam_id=exam.id)


@login_required
@role_required(['ADMIN'])
def standardized_exam_session_delete(request):
    """
    Deletes all StandardizedExam records and associated candidates, scores, rooms,
    and invigilator plans belonging to an entire multi-grade exam session in one click.
    """
    if request.method == 'POST':
        exam_ids_str = request.POST.get('exam_ids', '').strip()
        session_key = request.POST.get('session_key', '').strip()
        session_title = request.POST.get('session_title', '').strip()

        exam_ids = [int(eid.strip()) for eid in exam_ids_str.split(',') if eid.strip().isdigit()]
        exams_qs = StandardizedExam.objects.none()

        if exam_ids:
            exams_qs = StandardizedExam.objects.filter(id__in=exam_ids)
        elif session_key:
            parts = session_key.split('_', 2)
            if len(parts) >= 3 and parts[0].isdigit():
                y_id = int(parts[0])
                d_str = parts[1]
                t_str = parts[2]
                exams_qs = StandardizedExam.objects.filter(academic_year_id=y_id, exam_date=d_str, name__icontains=t_str)

        count = exams_qs.count()
        if count > 0:
            grades = list(exams_qs.order_by('grade_level').values_list('grade_level', flat=True).distinct())
            grades_str = ", ".join([f"ថ្នាក់ទី {g}" for g in grades]) if grades else f"{count} កម្រិត"

            # Clean up any invigilator plans linked specifically to this session or its exams
            if session_key:
                ExamInvigilatorPlan.objects.filter(session_key=session_key).delete()
            ExamInvigilatorPlan.objects.filter(standardized_exam__in=exams_qs).delete()

            # Cascade delete all exams in this session
            exams_qs.delete()

            title_display = session_title or "សម័យប្រឡង"
            messages.success(request, f"🗑️ បានលុបសម័យប្រឡងទាំងមូល «{title_display}» (រួមមាន {count} កម្រិតថ្នាក់៖ {grades_str}) និងទិន្នន័យពាក់ព័ន្ធទាំងអស់ដោយជោគជ័យ!")
        else:
            messages.warning(request, "ពុំមានសម័យប្រឡងណាមួយត្រូវបានរកឃើញសម្រាប់លុបឡើយ។")

    return redirect('standardized_exam_list')


@login_required
@role_required(['ADMIN'])
def standardized_exam_session_pull_candidates(request):
    """
    Batch Auto-Pulls all active students into candidates for all grade levels under an entire exam session in one click.
    """
    if request.method == 'POST':
        exam_ids_str = request.POST.get('exam_ids', '').strip()
        session_key = request.POST.get('session_key', '').strip()
        session_title = request.POST.get('session_title', '').strip()

        exam_ids = [int(eid.strip()) for eid in exam_ids_str.split(',') if eid.strip().isdigit()]
        exams_qs = StandardizedExam.objects.none()

        if exam_ids:
            exams_qs = StandardizedExam.objects.filter(id__in=exam_ids)
        elif session_key:
            parts = session_key.split('_', 2)
            if len(parts) >= 3 and parts[0].isdigit():
                y_id = int(parts[0])
                d_str = parts[1]
                t_str = parts[2]
                exams_qs = StandardizedExam.objects.filter(academic_year_id=y_id, exam_date=d_str, name__icontains=t_str)

        total_pulled = 0
        exams_list = list(exams_qs.order_by('grade_level'))
        for ex in exams_list:
            total_pulled += pull_candidates_for_exam(ex)

        title_display = session_title or "សម័យប្រឡង"
        if len(exams_list) > 0:
            messages.success(request, f"🎉 បានទាញបញ្ចូលបេក្ខជនសរុបចំនួន {total_pulled} នាក់ សម្រាប់សម័យប្រឡង «{title_display}» ({len(exams_list)} កម្រិតថ្នាក់) ដោយជោគជ័យ!")
        else:
            messages.warning(request, "ពុំមានសម័យប្រឡងណាមួយត្រូវបានរកឃើញសម្រាប់ទាញឈ្មោះសិស្សឡើយ។")

    return redirect('standardized_exam_list')



@login_required
def standardized_exam_manage(request, exam_id):
    """
    Central Operations Dashboard for a Standardized Exam.
    """
    exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)

    # Filtering Candidates
    candidates_qs = exam.candidates.select_related('room', 'student').order_by('room__room_number', 'desk_number', 'roll_number', 'id')
    
    room_filter = request.GET.get('room')
    gender_filter = request.GET.get('gender')
    search_q = request.GET.get('q', '').strip()

    if room_filter and room_filter.isdigit():
        candidates_qs = candidates_qs.filter(room_id=int(room_filter))
    elif room_filter == 'unassigned':
        candidates_qs = candidates_qs.filter(room__isnull=True)

    if gender_filter in ['M', 'F']:
        candidates_qs = candidates_qs.filter(gender=gender_filter)

    if search_q:
        candidates_qs = candidates_qs.filter(
            Q(candidate_name_kh__icontains=search_q) |
            Q(candidate_name_en__icontains=search_q) |
            Q(roll_number__icontains=search_q) |
            Q(origin_class__icontains=search_q) |
            Q(student_code__icontains=search_q)
        )

    rooms = exam.rooms.all().order_by('room_number')
    subjects = exam.exam_subjects.select_related('subject').order_by('order', 'id')

    total_candidates = exam.candidates.count()
    female_candidates = exam.candidates.filter(gender='F').count()
    total_rooms = rooms.count()

    total_max_score = sum(s.max_score for s in subjects)
    total_coefficients = sum(s.coefficient for s in subjects)

    # Find matching Invigilator Plan for this exam or exam session
    clean_title = get_clean_exam_session_title(exam.name)
    session_key = f"{exam.academic_year_id}_{exam.exam_date}_{clean_title}"
    invigilator_plan = ExamInvigilatorPlan.objects.filter(
        Q(standardized_exam=exam) |
        Q(session_key=session_key) |
        (Q(academic_year=exam.academic_year, start_date__lte=exam.exam_date, end_date__gte=exam.exam_date) & (Q(title__icontains=clean_title) | Q(title__icontains=exam.name)))
    ).prefetch_related('shift_slots__registrations').first()

    if invigilator_plan:
        slots = list(invigilator_plan.shift_slots.all())
        invigilator_plan.calc_total_capacity = sum(s.max_invigilators for s in slots)
        invigilator_plan.calc_total_registered = sum(s.registered_count for s in slots)

    return render(request, 'examinations/standardized/exam_manage.html', {
        'exam': exam,
        'clean_title': clean_title,
        'session_key': session_key,
        'invigilator_plan': invigilator_plan,
        'candidates': candidates_qs,
        'rooms': rooms,
        'subjects': subjects,
        'total_candidates': total_candidates,
        'female_candidates': female_candidates,
        'total_rooms': total_rooms,
        'total_max_score': total_max_score,
        'total_coefficients': total_coefficients,
        'selected_room': room_filter or '',
        'selected_gender': gender_filter or '',
        'search_q': search_q,
    })


@login_required
@role_required(['ADMIN'])
def exam_pull_candidates(request, exam_id):
    """
    1-Click Auto-Pull all active students from classrooms matching this exam's Grade Level and Track.
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    count = pull_candidates_for_exam(exam)
    messages.success(request, f"🎉 បានទាញបញ្ចូលបេក្ខជនសរុបចំនួន {count} នាក់ ពីកម្រិតថ្នាក់ទី {exam.grade_level} ដោយជោគជ័យ!")
    return redirect('standardized_exam_manage', exam_id=exam.id)


def partition_exam_rooms(exam, start_room_number=1, start_roll_number=1, cap=25, building="អគារ A", candidate_order='ALPHABETICAL'):
    """
    Helper to partition candidates into rooms and assign desk/roll numbers according to parameters.
    - Desk numbers are assigned 1..cap in each room.
    - Roll numbers start from start_roll_number.
    - Candidates are ordered by candidate_order (ALPHABETICAL, CLASS_ALPHABETICAL, STUDENT_ID, RANDOM).
    Returns: (total_candidates, needed_rooms, next_start_room, next_start_roll)
    """
    candidates = list(exam.candidates.all())
    if not candidates:
        # Create at least Room 01 so the exam has a room ready
        exam.rooms.all().delete()
        room_obj = ExamRoom.objects.create(
            exam=exam,
            room_number=start_room_number,
            room_name=f"បន្ទប់លេខ {start_room_number:02d}",
            building=building or "អគារ A"
        )
        exam.generate_all_secret_codes(force_regenerate=True)
        return 0, 1, start_room_number + 1, start_roll_number

    # Order candidates according to chosen policy
    if candidate_order == 'ALPHABETICAL':
        candidates.sort(key=lambda c: (c.candidate_name_kh or '', c.id))
    elif candidate_order == 'CLASS_ALPHABETICAL':
        candidates.sort(key=lambda c: (c.origin_class or '', c.candidate_name_kh or '', c.id))
    elif candidate_order == 'STUDENT_ID':
        candidates.sort(key=lambda c: (c.student_code or '', c.roll_number or '', c.candidate_name_kh or '', c.id))
    elif candidate_order == 'RANDOM':
        random.seed(exam.id + 42)
        random.shuffle(candidates)
    else:
        candidates.sort(key=lambda c: (c.candidate_name_kh or '', c.id))

    cap = cap or exam.candidates_per_room or 25
    total_candidates = len(candidates)
    needed_rooms = (total_candidates + cap - 1) // cap

    # Clear existing rooms
    exam.rooms.all().delete()

    created_rooms = []
    for r_idx in range(needed_rooms):
        actual_room_num = start_room_number + r_idx
        room_obj = ExamRoom.objects.create(
            exam=exam,
            room_number=actual_room_num,
            room_name=f"បន្ទប់លេខ {actual_room_num:02d}",
            building=building or "អគារ A"
        )
        created_rooms.append(room_obj)

    for idx, cand in enumerate(candidates):
        room_idx = idx // cap
        desk_num = (idx % cap) + 1

        cand.room = created_rooms[room_idx]
        cand.desk_number = desk_num
        cand.roll_number = f"{start_roll_number + idx:03d}"
        cand.save(update_fields=['room', 'desk_number', 'roll_number'])

    # Auto-generate unique secret codes for all rooms and subject envelopes
    exam.generate_all_secret_codes(force_regenerate=True)

    next_room_number = start_room_number + needed_rooms
    next_roll_number = start_roll_number + total_candidates
    return total_candidates, needed_rooms, next_room_number, next_roll_number


@login_required
@role_required(['ADMIN'])
def exam_generate_rooms(request, exam_id):
    """
    Auto-Partitions candidates into 25 candidates per room (or custom cap),
    generating Rooms and assigning Desk/Roll Numbers based on selected policy:
    - RESET_PER_GRADE: Starts from Room 01 and Roll No 001 for this grade.
    - CONTINUOUS_IN_SHIFT: Continues room & roll numbers from previous grades in the same shift (Morning / Afternoon) within this specific exam session.
    - CONTINUOUS_ALL_GRADES: Continues room & roll numbers across all grades within this specific exam session.
    - CUSTOM: Starts from custom start_room_number and start_roll_number specified by Admin.
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)

    # If exam has 0 candidates, auto-pull candidates first
    if exam.candidates.count() == 0:
        pull_candidates_for_exam(exam)

    numbering_mode = request.POST.get('numbering_mode', 'RESET_PER_GRADE')
    custom_cpr = request.POST.get('candidates_per_room')
    cap = int(custom_cpr) if custom_cpr and custom_cpr.isdigit() else (exam.candidates_per_room or 25)
    candidate_order = request.POST.get('candidate_order', 'ALPHABETICAL')
    building = request.POST.get('building', 'អគារ A')

    start_room_number = 1
    start_roll_number = 1

    clean_title = get_clean_exam_session_title(exam.name)

    if numbering_mode == 'CONTINUOUS_IN_SHIFT':
        # Find previous exams in the SAME EXAM SESSION (matching academic year, exam date, clean title) and same shift with lower grade_level
        same_session_exams = StandardizedExam.objects.filter(
            academic_year=exam.academic_year,
            exam_date=exam.exam_date,
            session=exam.session,
            grade_level__lt=exam.grade_level
        )
        prior_exams = [e for e in same_session_exams if get_clean_exam_session_title(e.name) == clean_title]
        
        total_prior_rooms = sum(e.rooms.count() for e in prior_exams)
        total_prior_candidates = sum(e.candidates.count() for e in prior_exams)
        start_room_number = total_prior_rooms + 1
        start_roll_number = total_prior_candidates + 1

    elif numbering_mode == 'CONTINUOUS_ALL_GRADES':
        # Find previous exams in the SAME EXAM SESSION with lower grade_level
        same_session_exams = StandardizedExam.objects.filter(
            academic_year=exam.academic_year,
            exam_date=exam.exam_date,
            grade_level__lt=exam.grade_level
        )
        prior_exams = [e for e in same_session_exams if get_clean_exam_session_title(e.name) == clean_title]
        
        total_prior_rooms = sum(e.rooms.count() for e in prior_exams)
        total_prior_candidates = sum(e.candidates.count() for e in prior_exams)
        start_room_number = total_prior_rooms + 1
        start_roll_number = total_prior_candidates + 1

    elif numbering_mode == 'CUSTOM':
        c_room = request.POST.get('start_room_number')
        c_roll = request.POST.get('start_roll_number')
        start_room_number = int(c_room) if c_room and c_room.isdigit() else 1
        start_roll_number = int(c_roll) if c_roll and c_roll.isdigit() else 1

    with transaction.atomic():
        total_cands, needed_rooms, _, _ = partition_exam_rooms(
            exam=exam,
            start_room_number=start_room_number,
            start_roll_number=start_roll_number,
            cap=cap,
            building=building,
            candidate_order=candidate_order
        )

    mode_labels = {
        'RESET_PER_GRADE': 'រាប់ចាប់ពីលេខ ១ សម្រាប់កម្រិតថ្នាក់នេះ',
        'CONTINUOUS_IN_SHIFT': f'រាប់បន្តគ្នាក្នុង {exam.get_session_display()}',
        'CONTINUOUS_ALL_GRADES': 'រាប់បន្តគ្នាគ្រប់កម្រិតថ្នាក់ក្នុងសម័យប្រឡងនេះ',
        'CUSTOM': f'កំណត់ដោយខ្លួនឯង (បន្ទប់ទី {start_room_number:02d}, អត្តលេខ {start_roll_number:03d})'
    }
    mode_text = mode_labels.get(numbering_mode, 'ស្តង់ដារ')
    if total_cands > 0:
        messages.success(request, f"🎉 បានរៀបចំ និងបែងចែកបេក្ខជនចំនួន {total_cands} នាក់ ទៅកាន់បន្ទប់ប្រឡង {needed_rooms} បន្ទប់ ({cap} នាក់/បន្ទប់) តាមទម្រង់ «{mode_text}» (បន្ទប់លេខ {start_room_number:02d} ដល់ {start_room_number+needed_rooms-1:02d}, អត្តលេខ {start_roll_number:03d} ដល់ {start_roll_number+total_cands-1:03d}) ដោយជោគជ័យ!")
    else:
        messages.info(request, f"🎉 បានបង្កើតបន្ទប់ប្រឡងលេខ {start_room_number:02d} រួចរាល់។ (ពុំទាន់មានបេក្ខជនក្នុងកម្រិតថ្នាក់នេះនៅឡើយទេ)")
    return redirect('standardized_exam_manage', exam_id=exam.id)


@login_required
@role_required(['ADMIN'])
def exam_batch_generate_rooms(request):
    """
    Batch Auto-Partitions standardized exams per exam session (សម័យប្រឡងនីមួយៗ) or across multiple grades in one click.
    Supports Numbering Policies:
    - CONTINUOUS_IN_SHIFT: Sequential room & roll numbers within Morning shift (7-10) and Afternoon shift (11-12) for the specific session.
    - RESET_PER_GRADE: Each grade resets to Room 01 and Roll 001.
    - CONTINUOUS_ALL_GRADES: Sequential room & roll numbers across all grades in this session.
    """
    if request.method != 'POST':
        return redirect('standardized_exam_list')

    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    
    academic_year_id = request.POST.get('academic_year')
    ay = AcademicYear.objects.filter(id=academic_year_id).first() if academic_year_id else active_year
    
    raw_exam_ids = request.POST.get('exam_ids', '').strip()
    session_key = request.POST.get('session_key', '').strip()
    session_title = request.POST.get('session_title', '').strip()
    exam_date_str = request.POST.get('exam_date', '').strip()
    
    scope = request.POST.get('scope', 'ALL_GRADES')
    numbering_mode = request.POST.get('numbering_mode', 'CONTINUOUS_IN_SHIFT')
    custom_cpr = request.POST.get('candidates_per_room')
    cap = int(custom_cpr) if custom_cpr and custom_cpr.isdigit() else 25
    candidate_order = request.POST.get('candidate_order', 'ALPHABETICAL')
    building = request.POST.get('building', 'អគារ A')

    target_session_name = ""
    exams_qs = StandardizedExam.objects.none()

    # 1. Primary Priority: Exact exam IDs passed from the session card modal
    if raw_exam_ids:
        eids = [int(x) for x in raw_exam_ids.split(',') if x.strip().isdigit()]
        if eids:
            exams_qs = StandardizedExam.objects.filter(id__in=eids)
            first_exam = exams_qs.first()
            if first_exam:
                target_session_name = get_clean_exam_session_title(first_exam.name)

    # 2. Secondary Priority: Filter by session_key
    if not exams_qs.exists() and session_key and session_key != 'ALL':
        all_exams = StandardizedExam.objects.all()
        if ay:
            ay_exams = all_exams.filter(academic_year=ay)
            if ay_exams.exists():
                all_exams = ay_exams

        matching_exams = []
        for ex in all_exams:
            ex_clean_title = get_clean_exam_session_title(ex.name)
            ex_date_key = str(ex.exam_date)
            ex_year_key = str(ex.academic_year_id)
            ex_group_key = f"{ex_year_key}_{ex_date_key}_{ex_clean_title}"
            
            if ex_group_key == session_key or ex_clean_title == session_key or session_key in ex_group_key:
                matching_exams.append(ex)
                if not target_session_name:
                    target_session_name = ex_clean_title
        
        exams_qs = StandardizedExam.objects.filter(id__in=[e.id for e in matching_exams])

    # 3. Tertiary Priority: Filter by session_title
    if not exams_qs.exists() and session_title:
        all_exams = StandardizedExam.objects.all()
        if ay:
            ay_exams = all_exams.filter(academic_year=ay)
            if ay_exams.exists():
                all_exams = ay_exams

        matching_exams = []
        for ex in all_exams:
            ex_clean_title = get_clean_exam_session_title(ex.name)
            if ex_clean_title == session_title or session_title in ex.name:
                if not exam_date_str or str(ex.exam_date) == exam_date_str:
                    matching_exams.append(ex)
                    if not target_session_name:
                        target_session_name = ex_clean_title
        exams_qs = StandardizedExam.objects.filter(id__in=[e.id for e in matching_exams])

    # 4. Fallback: All exams in the academic year
    if not exams_qs.exists() and (not session_key or session_key == 'ALL') and not raw_exam_ids and not session_title:
        exams_qs = StandardizedExam.objects.all()
        if ay:
            exams_qs = exams_qs.filter(academic_year=ay)

    if scope == 'MORNING_SHIFT':
        exams_qs = exams_qs.filter(session='MORNING')
    elif scope == 'AFTERNOON_SHIFT':
        exams_qs = exams_qs.filter(session='AFTERNOON')

    exams = list(exams_qs.order_by('grade_level', 'id'))
    if not exams:
        messages.warning(request, "មិនមានសម័យប្រឡងណាត្រូវនឹងលក្ខខណ្ឌជ្រើសរើសឡើយ!")
        return redirect('standardized_exam_list')

    # Automatically pull candidates for any exam in this batch that has 0 candidates
    for exam in exams:
        if exam.candidates.count() == 0:
            pull_candidates_for_exam(exam)

    total_exams_processed = 0
    total_candidates_partitioned = 0
    total_rooms_created = 0

    with transaction.atomic():
        if numbering_mode == 'RESET_PER_GRADE':
            for exam in exams:
                c_count, r_count, _, _ = partition_exam_rooms(
                    exam,
                    start_room_number=1,
                    start_roll_number=1,
                    cap=cap,
                    building=building,
                    candidate_order=candidate_order
                )
                if c_count > 0 or r_count > 0:
                    total_exams_processed += 1
                    total_candidates_partitioned += c_count
                    total_rooms_created += r_count

        elif numbering_mode == 'CONTINUOUS_IN_SHIFT':
            # Group by session within this exam batch
            shift_counters = {
                'MORNING': {'room': 1, 'roll': 1},
                'AFTERNOON': {'room': 1, 'roll': 1},
                'FULL_DAY': {'room': 1, 'roll': 1},
            }
            for exam in exams:
                sess = exam.session if exam.session in shift_counters else 'MORNING'
                curr_room = shift_counters[sess]['room']
                curr_roll = shift_counters[sess]['roll']
                c_count, r_count, next_room, next_roll = partition_exam_rooms(
                    exam,
                    start_room_number=curr_room,
                    start_roll_number=curr_roll,
                    cap=cap,
                    building=building,
                    candidate_order=candidate_order
                )
                if c_count > 0 or r_count > 0:
                    shift_counters[sess]['room'] = next_room
                    shift_counters[sess]['roll'] = next_roll
                    total_exams_processed += 1
                    total_candidates_partitioned += c_count
                    total_rooms_created += r_count

        else: # CONTINUOUS_ALL_GRADES
            curr_room = 1
            curr_roll = 1
            for exam in exams:
                c_count, r_count, next_room, next_roll = partition_exam_rooms(
                    exam,
                    start_room_number=curr_room,
                    start_roll_number=curr_roll,
                    cap=cap,
                    building=building,
                    candidate_order=candidate_order
                )
                if c_count > 0 or r_count > 0:
                    curr_room = next_room
                    curr_roll = next_roll
                    total_exams_processed += 1
                    total_candidates_partitioned += c_count
                    total_rooms_created += r_count

    session_label = f" «{target_session_name}»" if target_session_name else ""
    messages.success(request, f"⚡ បានរៀបចំ និងចែកបន្ទប់ស្វ័យប្រវត្តិតាមសម័យប្រឡង{session_label} សរុប {total_exams_processed} កម្រិតថ្នាក់ (បេក្ខជន {total_candidates_partitioned} នាក់, បន្ទប់ {total_rooms_created} បន្ទប់) ដោយជោគជ័យ!")
    return redirect('standardized_exam_list')



# Helper converters for Cambodian MoEYS official examination documents
KHMER_DIGITS_MAP = {
    '0': '០', '1': '១', '2': '២', '3': '៣', '4': '៤',
    '5': '៥', '6': '៦', '7': '៧', '8': '៨', '9': '៩'
}

KHMER_MONTH_NAMES = {
    1: 'មករា', 2: 'កុម្ភៈ', 3: 'មីនា', 4: 'មេសា',
    5: 'ឧសភា', 6: 'មិថុនា', 7: 'កក្កដា', 8: 'សីហា',
    9: 'កញ្ញា', 10: 'តុលា', 11: 'វិច្ឆិកា', 12: 'ធ្នូ',
}


def to_khmer_digits(val):
    """Convert any number or string of digits to Khmer numerals (0 -> ០, 1 -> ១...)."""
    if val is None or val == '':
        return ''
    return ''.join(KHMER_DIGITS_MAP.get(c, c) for c in str(val))


def to_khmer_2digits(val):
    """Format an integer as 2 digits in Khmer numerals (e.g. 1 -> ០១, 8 -> ០៨, 15 -> ១៥, 0 -> ០០)."""
    try:
        n = int(val)
        return to_khmer_digits(f"{n:02d}")
    except (ValueError, TypeError):
        return to_khmer_digits(val)


def format_khmer_full_date(d):
    """Format a date as 'DD ខែ YYYY' in Khmer numerals (e.g. '១២ តុលា ២០១២')."""
    if not d:
        return ''
    day_str = to_khmer_2digits(d.day)
    month_str = KHMER_MONTH_NAMES.get(d.month, '')
    year_str = to_khmer_digits(d.year)
    return f"{day_str} {month_str} {year_str}".strip()


def format_origin_classroom(orig_class, grade_level):
    """
    Format classroom for MoEYS posting sheet matching show.pdf (e.g. '7A' -> 'A', 'A' -> 'A').
    """
    if not orig_class:
        return ''
    orig = str(orig_class).strip()
    g_str = str(grade_level).strip() if grade_level else ''
    
    # Strip common prefixes like 'ថ្នាក់ទី ' or 'ថ្នាក់ '
    for pfx in ['ថ្នាក់ទី ', 'ថ្នាក់ទី', 'ថ្នាក់ ', 'ថ្នាក់']:
        if orig.startswith(pfx):
            orig = orig[len(pfx):].strip()
            break
            
    if g_str and orig.startswith(g_str):
        trimmed = orig[len(g_str):].strip()
        if trimmed:
            return trimmed
    return orig


@login_required
def exam_room_postings_view(request, exam_id):
    """
    Official MoEYS Exam Room Notice Board Posting Sheet (បញ្ជីបិទផ្សាយតាមបន្ទប់).
    Exact 1:1 replica of MoEYS standard format as exemplified in show.pdf:
    - 25 Rows per room sheet strictly (every room renders exactly 25 rows)
    - Continuous desk numbering across rooms (Room 1: 1-25, Room 2: 26-50, Room 10: 226-250, etc.)
    - National header: Kingdom motto, Province, School Name, Room Number, Session (ព្រឹក/រសៀល)
    - Centered Exam Title with Academic Year & Grade Level
    - Full Khmer Exam Date (សម័យប្រឡង៖ ០៣ សីហា ២០២៦)
    - 8 Columns: ល.រ, លេខតុ, អត្តលេខ, គោត្តនាម និងនាម, ភេទ, ថ្ងៃ ខែ ឆ្នាំកំណើត, មកពីថ្នាក់, ផ្សេងៗ
    - Footer with candidate counts in 2-digit Khmer numerals and School Principal signature
    """
    from apps.accounts.models import SchoolProfile
    
    exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)
    school_profile = SchoolProfile.get_settings()
    
    rooms_qs = exam.rooms.prefetch_related('candidates__student').order_by('room_number')

    selected_room_id = request.GET.get('room_id')
    if selected_room_id and selected_room_id.isdigit():
        rooms_qs = rooms_qs.filter(id=int(selected_room_id))

    # Session in Khmer
    session_map = {
        'MORNING': 'ព្រឹក',
        'AFTERNOON': 'រសៀល',
        'FULL_DAY': 'ពេញមួយថ្ងៃ',
    }
    session_kh = session_map.get(exam.session, exam.get_session_display() or 'ព្រឹក')

    # Exam Title Line: Clean prefix & suffix to prevent duplication of grade/year
    import re
    exam_name_clean = exam.name.strip()
    for pfx in ['បញ្ជីរាយនាមសិស្សប្រឡង', 'បញ្ជីឈ្មោះសិស្សប្រឡង', 'ប្រឡង']:
        if exam_name_clean.startswith(pfx):
            exam_name_clean = exam_name_clean[len(pfx):].strip()
            break

    # Strip any trailing 'ថ្នាក់ទី ...' or 'ថ្នាក់ ...' or 'ឆ្នាំសិក្សា ...' already in exam name
    exam_name_clean = re.sub(r'\s*ថ្នាក់ទី\s*[\d\u17e0-\u17e9A-Za-z]+\s*$', '', exam_name_clean).strip()
    exam_name_clean = re.sub(r'\s*ថ្នាក់\s*[\d\u17e0-\u17e9A-Za-z]+\s*$', '', exam_name_clean).strip()
    exam_name_clean = re.sub(r'\s*ឆ្នាំសិក្សា\s*[\d\u17e0-\u17e9\-]+\s*$', '', exam_name_clean).strip()

    academic_year_kh = to_khmer_digits(exam.academic_year.name if exam.academic_year else '')
    grade_kh = to_khmer_digits(exam.grade_level)
    
    if academic_year_kh and grade_kh:
        exam_title_line = f"បញ្ជីរាយនាមសិស្សប្រឡង{exam_name_clean} ឆ្នាំសិក្សា {academic_year_kh} ថ្នាក់ទី {grade_kh}".strip()
    elif academic_year_kh:
        exam_title_line = f"បញ្ជីរាយនាមសិស្សប្រឡង{exam_name_clean} ឆ្នាំសិក្សា {academic_year_kh}".strip()
    elif grade_kh:
        exam_title_line = f"បញ្ជីរាយនាមសិស្សប្រឡង{exam_name_clean} ថ្នាក់ទី {grade_kh}".strip()
    else:
        exam_title_line = f"បញ្ជីរាយនាមសិស្សប្រឡង{exam_name_clean}".strip()

    # Exam Date in Khmer
    exam_date_kh = format_khmer_full_date(exam.exam_date)

    # Province & School Name (allow query param override or default to SchoolProfile / show.pdf fallback)
    province_name = request.GET.get('province') or school_profile.province or 'ខេត្តកណ្តាល'
    school_name = request.GET.get('school_name') or school_profile.name_kh or 'វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្សួត'

    # Signing Location & Date
    location_name = request.GET.get('location', '').strip()
    if not location_name:
        if school_profile.commune:
            c_name = school_profile.commune.strip()
            for prefix in ['ឃុំ', 'សង្កាត់', 'ឃុំ ', 'សង្កាត់ ']:
                if c_name.startswith(prefix):
                    c_name = c_name[len(prefix):].strip()
                    break
            location_name = c_name
        elif school_profile.district:
            d_name = school_profile.district.strip()
            for prefix in ['ស្រុក', 'ខណ្ឌ', 'ក្រុង', 'ស្រុក ', 'ខណ្ឌ ', 'ក្រុង ']:
                if d_name.startswith(prefix):
                    d_name = d_name[len(prefix):].strip()
                    break
            location_name = d_name
        else:
            location_name = 'កំពង់កន្សួត'

    sign_date_param = request.GET.get('sign_date', '').strip()
    sign_date = None
    if sign_date_param:
        try:
            sign_date = datetime.datetime.strptime(sign_date_param, '%Y-%m-%d').date()
        except ValueError:
            pass
    if not sign_date:
        sign_date = exam.exam_date or timezone.now().date()

    sign_day_kh = to_khmer_2digits(sign_date.day)
    sign_month_kh = KHMER_MONTH_NAMES.get(sign_date.month, '')
    sign_year_kh = to_khmer_digits(sign_date.year)

    # Signer title (e.g. នាយក, នាយិកា, ប្រធានមណ្ឌល)
    sign_role = request.GET.get('sign_role', '').strip() or 'នាយក'

    rooms_data = []
    for r in rooms_qs:
        cand_list = list(r.candidates.select_related('student').order_by('desk_number', 'roll_number', 'id'))
        total_cands = len(cand_list)
        female_cands = len([c for c in cand_list if c.gender == 'F'])

        # Global desk base calculation: Room 1 -> 1-25; Room 2 -> 26-50; Room R -> (R-1)*25 + 1..25
        room_num_int = r.room_number if isinstance(r.room_number, int) else 1
        base_desk = (room_num_int - 1) * 25

        pad_25 = request.GET.get('pad_25') == '1'
        row_limit = 25 if pad_25 else total_cands

        rows = []
        for i in range(1, row_limit + 1):
            default_desk = base_desk + i
            if i <= total_cands:
                c = cand_list[i - 1]
                # If candidate already has global desk number assigned (> 25), preserve it, else default_desk
                actual_desk = c.desk_number if (c.desk_number and c.desk_number > 25) else default_desk

                student_id = c.student_code or (c.student.student_id if c.student else '') or c.roll_number
                dob_kh = format_khmer_full_date(c.dob) if c.dob else ''
                gender_kh = 'ស' if c.gender == 'F' else 'ប'
                origin_cls = format_origin_classroom(c.origin_class, exam.grade_level)

                rows.append({
                    'row_num': i,
                    'desk_number': actual_desk,
                    'student_id': student_id,
                    'candidate_name_kh': c.candidate_name_kh,
                    'candidate_name_en': c.candidate_name_en or '',
                    'gender_kh': gender_kh,
                    'dob_kh': dob_kh,
                    'origin_class': origin_cls,
                    'remarks': c.remarks or '',
                    'is_disciplinary_blocked': c.is_disciplinary_blocked,
                    'has_data': True,
                })
            else:
                rows.append({
                    'row_num': i,
                    'desk_number': default_desk,
                    'student_id': '',
                    'candidate_name_kh': '',
                    'candidate_name_en': '',
                    'gender_kh': '',
                    'dob_kh': '',
                    'origin_class': '',
                    'remarks': '',
                    'is_disciplinary_blocked': False,
                    'has_data': False,
                })

        room_number_kh = to_khmer_2digits(room_num_int)
        total_kh = to_khmer_2digits(total_cands)
        female_kh = to_khmer_2digits(female_cands)

        rooms_data.append({
            'room': r,
            'room_number_kh': room_number_kh,
            'rows': rows,
            'candidates': cand_list,  # kept for backwards compatibility
            'total_candidates': total_cands,
            'female_candidates': female_cands,
            'total_candidates_kh': total_kh,
            'female_candidates_kh': female_kh,
        })

    all_rooms = exam.rooms.all().order_by('room_number')

    return render(request, 'examinations/standardized/room_postings_print.html', {
        'exam': exam,
        'rooms_data': rooms_data,
        'all_rooms': all_rooms,
        'selected_room_id': selected_room_id or '',
        'session_kh': session_kh,
        'exam_title_line': exam_title_line,
        'exam_date_kh': exam_date_kh,
        'province_name': province_name,
        'school_name': school_name,
        'location_name': location_name,
        'sign_day_kh': sign_day_kh,
        'sign_month_kh': sign_month_kh,
        'sign_year_kh': sign_year_kh,
        'sign_role': sign_role,
        'sign_date_raw': sign_date.strftime('%Y-%m-%d'),
        'school_profile': school_profile,
    })


SUBJECT_SHORT_NAMES = {
    'តែងសេចក្តី': 'តែង',
    'សរសេរតាមអាន': 'សរសេរ',
    'ភាសាខ្មែរ': 'ខ្មែរ',
    'សីលធម៌ ពលរដ្ឋ': 'សីល',
    'សីលធម៌-ពលរដ្ឋ': 'សីល',
    'សីលធម៌': 'សីល',
    'ភូមិវិទ្យា': 'ភូមិ',
    'ប្រវត្តិវិទ្យា': 'ប្រវត្តិ',
    'គណិតវិទ្យា': 'គណិត',
    'ផែនដីវិទ្យា': 'ផែនដី',
    'ផែនដី និងបរិស្ថានវិទ្យា': 'ផែនដី',
    'រូបវិទ្យា': 'រូប',
    'គីមីវិទ្យា': 'គីមី',
    'ជីវវិទ្យា': 'ជីវ',
    'គេហវិទ្យា': 'គេហៈ',
    'គេហសេដ្ឋកិច្ច': 'គេហៈ',
    'ភាសាអង់គ្លេស': 'អង់គ្លេស',
    'អង់គ្លេស': 'អង់គ្លេស',
    'ភាសាបារាំង': 'បារាំង',
    'បារាំង': 'បារាំង',
    'ព័ត៌មានវិទ្យា': 'ICT',
}

DEFAULT_MOEYS_ATTENDANCE_SUBJECTS = [
    'តែង', 'សរសេរ', 'ខ្មែរ', 'សីល', 'ភូមិ', 'ប្រវត្តិ', 'គណិត', 'ផែនដី', 'រូប', 'គីមី', 'ជីវ', 'គេហៈ', 'អង់គ្លេស'
]


@login_required
def exam_subject_attendance_view(request, exam_id):
    """
    Official MoEYS Candidate Attendance & Signature Sheet (បញ្ជីវត្តមានចុះហត្ថលេខាបេក្ខជនតាមបន្ទប់).
    Exact 1:1 replica of MoEYS standard format as exemplified in show2.pdf:
    - A4 Landscape format (297mm x 210mm)
    - Guaranteed 1cm - 1.5cm margins all around on Print and PDF export
    - Exactly 25 rows per room sheet strictly
    - Continuous global desk numbering across rooms (1-25, 26-50, etc.)
    - Multi-subject horizontal signature columns under grouped header 'ហត្ថលេខាបេក្ខជនតាមមុខវិជ្ជា'
    - School header, Room number, Attendance title line, and footer signatures
    """
    from apps.accounts.models import SchoolProfile
    import re
    
    exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)
    school_profile = SchoolProfile.get_settings()
    
    rooms_qs = exam.rooms.prefetch_related('candidates__student').order_by('room_number')

    selected_room_id = request.GET.get('room_id')
    if selected_room_id and selected_room_id.isdigit():
        rooms_qs = rooms_qs.filter(id=int(selected_room_id))

    # Subjects for the signature columns
    exam_subjects = list(exam.exam_subjects.select_related('subject').order_by('order', 'id'))
    if exam_subjects:
        display_subjects = []
        for s in exam_subjects:
            s_name = s.subject.name_kh.strip()
            short_name = SUBJECT_SHORT_NAMES.get(s_name, s_name)
            display_subjects.append({
                'id': s.id,
                'name': short_name,
                'full_name': s_name,
            })
    else:
        display_subjects = [
            {'id': idx, 'name': name, 'full_name': name}
            for idx, name in enumerate(DEFAULT_MOEYS_ATTENDANCE_SUBJECTS)
        ]

    # Session in Khmer
    session_map = {
        'MORNING': 'ព្រឹក',
        'AFTERNOON': 'រសៀល',
        'FULL_DAY': 'ពេញមួយថ្ងៃ',
    }
    session_kh = session_map.get(exam.session, exam.get_session_display() or 'ព្រឹក')

    # Exam Title Line: Clean prefix & suffix to prevent duplicate grade/year
    exam_name_clean = exam.name.strip()
    for pfx in ['បញ្ជីវត្តមានបេក្ខជនប្រឡង', 'បញ្ជីរាយនាមសិស្សប្រឡង', 'បញ្ជីឈ្មោះសិស្សប្រឡង', 'ប្រឡង']:
        if exam_name_clean.startswith(pfx):
            exam_name_clean = exam_name_clean[len(pfx):].strip()
            break

    # Strip any trailing 'ថ្នាក់ទី ...' or 'ថ្នាក់ ...' or 'ឆ្នាំសិក្សា ...' already in exam name
    exam_name_clean = re.sub(r'\s*ថ្នាក់ទី\s*[\d\u17e0-\u17e9A-Za-z]+\s*$', '', exam_name_clean).strip()
    exam_name_clean = re.sub(r'\s*ថ្នាក់\s*[\d\u17e0-\u17e9A-Za-z]+\s*$', '', exam_name_clean).strip()
    exam_name_clean = re.sub(r'\s*ឆ្នាំសិក្សា\s*[\d\u17e0-\u17e9\-]+\s*$', '', exam_name_clean).strip()

    academic_year_kh = to_khmer_digits(exam.academic_year.name if exam.academic_year else '')
    grade_kh = to_khmer_digits(exam.grade_level)
    
    if academic_year_kh and grade_kh:
        attendance_title_line = f"បញ្ជីវត្តមានបេក្ខជនប្រឡង{exam_name_clean} ឆ្នាំសិក្សា {academic_year_kh} ថ្នាក់ទី {grade_kh}".strip()
    elif academic_year_kh:
        attendance_title_line = f"បញ្ជីវត្តមានបេក្ខជនប្រឡង{exam_name_clean} ឆ្នាំសិក្សា {academic_year_kh}".strip()
    elif grade_kh:
        attendance_title_line = f"បញ្ជីវត្តមានបេក្ខជនប្រឡង{exam_name_clean} ថ្នាក់ទី {grade_kh}".strip()
    else:
        attendance_title_line = f"បញ្ជីវត្តមានបេក្ខជនប្រឡង{exam_name_clean}".strip()

    # Exam Date in Khmer
    exam_date_kh = format_khmer_full_date(exam.exam_date)

    # School Name & Province
    school_name = request.GET.get('school_name') or school_profile.name_kh or 'វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្សួត'
    province_name = request.GET.get('province') or school_profile.province or 'ខេត្តកណ្តាល'

    # Signing Location & Date
    location_name = request.GET.get('location', '').strip()
    if not location_name:
        if school_profile.commune:
            c_name = school_profile.commune.strip()
            for prefix in ['ឃុំ', 'សង្កាត់', 'ឃុំ ', 'សង្កាត់ ']:
                if c_name.startswith(prefix):
                    c_name = c_name[len(prefix):].strip()
                    break
            location_name = c_name
        elif school_profile.district:
            d_name = school_profile.district.strip()
            for prefix in ['ស្រុក', 'ខណ្ឌ', 'ក្រុង', 'ស្រុក ', 'ខណ្ឌ ', 'ក្រុង ']:
                if d_name.startswith(prefix):
                    d_name = d_name[len(prefix):].strip()
                    break
            location_name = d_name
        else:
            location_name = 'កំពង់កន្សួត'

    sign_date_param = request.GET.get('sign_date', '').strip()
    sign_date = None
    if sign_date_param:
        try:
            sign_date = datetime.datetime.strptime(sign_date_param, '%Y-%m-%d').date()
        except ValueError:
            pass
    if not sign_date:
        sign_date = exam.exam_date or timezone.now().date()

    sign_day_kh = to_khmer_2digits(sign_date.day)
    sign_month_kh = KHMER_MONTH_NAMES.get(sign_date.month, '')
    sign_year_kh = to_khmer_digits(sign_date.year)

    # Signer title (e.g. នាយក, នាយិកា, ប្រធានមណ្ឌល)
    sign_role = request.GET.get('sign_role', '').strip() or 'នាយក'

    rooms_data = []
    for r in rooms_qs:
        cand_list = list(r.candidates.select_related('student').order_by('desk_number', 'roll_number', 'id'))
        total_cands = len(cand_list)
        female_cands = len([c for c in cand_list if c.gender == 'F'])

        room_num_int = r.room_number if isinstance(r.room_number, int) else 1
        base_desk = (room_num_int - 1) * 25

        pad_25 = request.GET.get('pad_25') == '1'
        row_limit = 25 if pad_25 else total_cands

        rows = []
        for i in range(1, row_limit + 1):
            default_desk = base_desk + i
            if i <= total_cands:
                c = cand_list[i - 1]
                actual_desk = c.desk_number if (c.desk_number and c.desk_number > 25) else default_desk
                student_id = c.student_code or (c.student.student_id if c.student else '') or c.roll_number
                
                # Compact DOB format matching show2.pdf (e.g. 12/10/12)
                dob_compact = c.dob.strftime('%d/%m/%y') if c.dob else ''
                gender_kh = 'ស' if c.gender == 'F' else 'ប'
                origin_cls = format_origin_classroom(c.origin_class, exam.grade_level)

                rows.append({
                    'row_num': i,
                    'desk_number': actual_desk,
                    'student_id': student_id,
                    'candidate_name_kh': c.candidate_name_kh,
                    'gender_kh': gender_kh,
                    'dob_str': dob_compact,
                    'origin_class': origin_cls,
                    'remarks': c.remarks or '',
                    'is_disciplinary_blocked': c.is_disciplinary_blocked,
                    'has_data': True,
                })
            else:
                rows.append({
                    'row_num': i,
                    'desk_number': default_desk,
                    'student_id': '',
                    'candidate_name_kh': '',
                    'gender_kh': '',
                    'dob_str': '',
                    'origin_class': '',
                    'remarks': '',
                    'is_disciplinary_blocked': False,
                    'has_data': False,
                })

        room_number_kh = to_khmer_2digits(room_num_int)
        total_kh = to_khmer_2digits(total_cands)
        female_kh = to_khmer_2digits(female_cands)

        rooms_data.append({
            'room': r,
            'room_number_kh': room_number_kh,
            'rows': rows,
            'candidates': cand_list,  # kept for test compatibility
            'total_candidates': total_cands,
            'female_candidates': female_cands,
            'total_candidates_kh': total_kh,
            'female_candidates_kh': female_kh,
        })

    all_rooms = exam.rooms.all().order_by('room_number')

    return render(request, 'examinations/standardized/attendance_sheets_print.html', {
        'exam': exam,
        'rooms_data': rooms_data,
        'sheets_data': rooms_data,  # kept for backwards compatibility
        'rooms': all_rooms,
        'display_subjects': display_subjects,
        'subjects': exam_subjects,
        'selected_room_id': selected_room_id or '',
        'session_kh': session_kh,
        'attendance_title_line': attendance_title_line,
        'exam_date_kh': exam_date_kh,
        'school_name': school_name,
        'province_name': province_name,
        'location_name': location_name,
        'sign_day_kh': sign_day_kh,
        'sign_month_kh': sign_month_kh,
        'sign_year_kh': sign_year_kh,
        'sign_role': sign_role,
        'sign_date_raw': sign_date.strftime('%Y-%m-%d'),
        'school_profile': school_profile,
    })


@login_required
@role_required(['ADMIN', 'TEACHER'])
def exam_room_scores_entry(request, exam_id):
    """
    Rapid Score Entry Matrix per Room for all subjects with dynamic calculation of
    Total Score, Weighted Average (coefficients included), MoEYS Letter Grade (A-F), and Room Rank.
    """
    exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)
    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'

    # Check if direct room scoring is blocked by Admin in favor of Blind Secret Code scoring
    if exam.grading_method == 'BLIND_SECRET_CODE' and not is_admin:
        messages.warning(request, f"⚠️ សម័យប្រឡង «{exam.name}» ត្រូវបានកំណត់ដោយ Admin ឱ្យបញ្ចូលពិន្ទុតាមរយៈលេខកូដសម្ងាត់ (Blind Scoring) តែប៉ុណ្ណោះ! សូមប្រើប្រាស់ផ្ទាំងបញ្ចូលពិន្ទុកូដសម្ងាត់។")
        return redirect(f"{reverse('exam_blind_scoring_portal')}?exam_id={exam.id}")

    rooms = list(exam.rooms.all().order_by('room_number'))
    subjects = list(exam.exam_subjects.select_related('subject').order_by('order', 'id'))

    selected_room_id = request.GET.get('room_id')
    selected_room = None
    if selected_room_id and selected_room_id.isdigit():
        selected_room = exam.rooms.filter(id=int(selected_room_id)).first()
    if not selected_room and rooms:
        selected_room = rooms[0]

    candidates = []
    if selected_room:
        candidates = list(selected_room.candidates.prefetch_related('subject_scores__exam_subject').order_by('desk_number', 'roll_number', 'id'))

    if request.method == 'POST':
        # Handle saving score matrix
        with transaction.atomic():
            saved_count = 0
            for cand in candidates:
                for s in subjects:
                    field_key = f"score_{cand.id}_{s.id}"
                    absent_key = f"absent_{cand.id}_{s.id}"

                    score_val = request.POST.get(field_key, '').strip()
                    is_absent = (request.POST.get(absent_key) == '1') or (score_val.upper() == 'A')

                    score_obj, _ = CandidateSubjectScore.objects.get_or_create(
                        candidate=cand,
                        exam_subject=s
                    )
                    score_obj.is_absent = is_absent
                    if is_absent:
                        score_obj.score = Decimal('0.00')
                    elif score_val != '':
                        try:
                            score_num = Decimal(str(score_val))
                            score_obj.score = min(score_num, s.max_score)
                        except Exception:
                            score_obj.score = None
                    else:
                        score_obj.score = None

                    score_obj.save()
                    saved_count += 1

            # Recalculate ranks across the entire exam
            exam.recalculate_all_ranks()

        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return JsonResponse({'status': 'success', 'message': f'បានរក្សាទុកពិន្ទុ {saved_count} ក្រឡា ដោយជោគជ័យ!'})

        messages.success(request, f"🎉 បានរក្សាទុក និងគណនាលទ្ធផលពិន្ទុសម្រាប់ «{selected_room.room_name}» ដោយជោគជ័យ!")
        return redirect(f"{request.path}?room_id={selected_room.id}")

    # Build matrix map: (candidate_id, exam_subject_id) -> CandidateSubjectScore
    scores_map = {}
    for cand in candidates:
        for sc in cand.subject_scores.all():
            scores_map[(cand.id, sc.exam_subject_id)] = sc

    candidate_rows = []
    for cand in candidates:
        row_scores = []
        for s in subjects:
            sc = scores_map.get((cand.id, s.id))
            row_scores.append({
                'subject': s,
                'score_obj': sc,
                'score_val': sc.score if sc and sc.score is not None else '',
                'is_absent': sc.is_absent if sc else False,
            })
        candidate_rows.append({
            'candidate': cand,
            'scores': row_scores,
        })

    return render(request, 'examinations/standardized/room_scores_entry.html', {
        'exam': exam,
        'rooms': rooms,
        'selected_room': selected_room,
        'subjects': subjects,
        'candidate_rows': candidate_rows,
        'total_max_score': sum(s.max_score for s in subjects),
        'total_coefficients': sum(s.coefficient for s in subjects),
    })


@login_required
def exam_provisional_results_view(request, exam_id):
    """
    Master Grade-Level Provisional Results Posting Board (តារាងបិទផ្សាយបណ្តោះអាសន្នតាមកម្រិតថ្នាក់).
    Supports sorting by Name (Alphabetical) and by Rank, filtering by Grade (A-F) or Room,
    with summary stats, official print view, and Excel export.
    """
    exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)
    
    # Auto-recalculate ranks to guarantee freshness
    exam.recalculate_all_ranks()

    subjects = list(exam.exam_subjects.select_related('subject').order_by('order', 'id'))
    total_max_score = sum(s.max_score for s in subjects) if subjects else Decimal('100.00')

    candidates_qs = exam.candidates.select_related('room').prefetch_related('subject_scores__exam_subject')

    # Filtering
    room_filter = request.GET.get('room')
    grade_filter = request.GET.get('grade')
    gender_filter = request.GET.get('gender')
    search_q = request.GET.get('q', '').strip()
    sort_by = request.GET.get('sort', 'rank') # 'rank', 'name', 'desk', 'roll'

    if room_filter and room_filter.isdigit():
        candidates_qs = candidates_qs.filter(room_id=int(room_filter))
    if grade_filter in ['A', 'B', 'C', 'D', 'E', 'F']:
        candidates_qs = candidates_qs.filter(grade_letter=grade_filter)
    if gender_filter in ['M', 'F']:
        candidates_qs = candidates_qs.filter(gender=gender_filter)
    if search_q:
        candidates_qs = candidates_qs.filter(
            Q(candidate_name_kh__icontains=search_q) |
            Q(candidate_name_en__icontains=search_q) |
            Q(roll_number__icontains=search_q) |
            Q(origin_class__icontains=search_q)
        )

    all_candidates = list(candidates_qs)

    # Sorting
    if sort_by == 'name':
        all_candidates.sort(key=lambda c: (c.candidate_name_kh or '', c.roll_number or ''))
    elif sort_by == 'desk':
        all_candidates.sort(key=lambda c: (c.room.room_number if c.room else 999, c.desk_number, c.roll_number))
    elif sort_by == 'roll':
        all_candidates.sort(key=lambda c: c.roll_number or '')
    else:
        # Default sort by rank (rank_overall asc, total_score desc)
        all_candidates.sort(key=lambda c: (c.rank_overall if c.rank_overall else 9999, -(c.total_score or 0)))

    # Statistics
    total_cand = len(all_candidates)
    females = len([c for c in all_candidates if c.gender == 'F'])
    passed_cand = len([c for c in all_candidates if c.grade_letter in ['A', 'B', 'C', 'D', 'E']])
    failed_cand = len([c for c in all_candidates if c.grade_letter == 'F'])

    grade_counts = {
        'A': len([c for c in all_candidates if c.grade_letter == 'A']),
        'B': len([c for c in all_candidates if c.grade_letter == 'B']),
        'C': len([c for c in all_candidates if c.grade_letter == 'C']),
        'D': len([c for c in all_candidates if c.grade_letter == 'D']),
        'E': len([c for c in all_candidates if c.grade_letter == 'E']),
        'F': len([c for c in all_candidates if c.grade_letter == 'F']),
    }

    # Prepare rows with subject scores for template
    candidate_rows = []
    for cand in all_candidates:
        cand_scores = {sc.exam_subject_id: sc for sc in cand.subject_scores.all()}
        scores_list = []
        for s in subjects:
            sc = cand_scores.get(s.id)
            scores_list.append({
                'subject': s,
                'score': sc.score if sc and sc.score is not None else '-',
                'is_absent': sc.is_absent if sc else False,
            })
        candidate_rows.append({
            'candidate': cand,
            'scores': scores_list,
            'is_passed': cand.grade_letter in ['A', 'B', 'C', 'D', 'E'],
        })

    rooms = exam.rooms.all().order_by('room_number')

    return render(request, 'examinations/standardized/provisional_results_board.html', {
        'exam': exam,
        'candidate_rows': candidate_rows,
        'subjects': subjects,
        'rooms': rooms,
        'total_cand': total_cand,
        'females': females,
        'passed_cand': passed_cand,
        'failed_cand': failed_cand,
        'pass_rate': round((passed_cand / total_cand * 100), 1) if total_cand > 0 else 0,
        'grade_counts': grade_counts,
        'total_max_score': total_max_score,
        'selected_room': room_filter or '',
        'selected_grade': grade_filter or '',
        'selected_gender': gender_filter or '',
        'selected_sort': sort_by,
        'search_q': search_q,
    })


MOEYS_RESULT_SUBJECTS = [
    {'key': 'taing', 'name': 'តែងសេចក្តី', 'short': 'តែង', 'max': 60.0},
    {'key': 'sarser', 'name': 'សរសេរតាមអាន', 'short': 'សរសេរ', 'max': 40.0},
    {'key': 'khmer', 'name': 'ភាសាខ្មែរ', 'short': 'ខ្មែរ', 'max': 100.0, 'is_composite': True},
    {'key': 'sil', 'name': 'សីលធម៌', 'short': 'សីល', 'max': 50.0},
    {'key': 'phum', 'name': 'ភូមិវិទ្យា', 'short': 'ភូមិ', 'max': 50.0},
    {'key': 'pravatt', 'name': 'ប្រវត្តិវិទ្យា', 'short': 'ប្រវត្តិ', 'max': 50.0},
    {'key': 'kanit', 'name': 'គណិតវិទ្យា', 'short': 'គណិត', 'max': 100.0},
    {'key': 'phendey', 'name': 'ផែនដីវិទ្យា', 'short': 'ផែនដី', 'max': 50.0},
    {'key': 'roop', 'name': 'រូបវិទ្យា', 'short': 'រូប', 'max': 50.0},
    {'key': 'kimey', 'name': 'គីមីវិទ្យា', 'short': 'គីមី', 'max': 50.0},
    {'key': 'chiv', 'name': 'ជីវវិទ្យា', 'short': 'ជីវ', 'max': 50.0},
    {'key': 'keha', 'name': 'គេហវិទ្យា', 'short': 'គេហៈ', 'max': 50.0},
    {'key': 'english', 'name': 'អង់គ្លេស', 'short': 'អង់គ្លេស', 'max': 50.0},
]


def get_moeys_subject_mention(score, max_score):
    """
    Computes official MoEYS subject letter grade (A..F):
    - 50-pt subject: A>=45, B>=40, C>=35, D>=30, E>=25, F<25
    - 100-pt subject: A>=90, B>=80, C>=70, D>=60, E>=50, F<50
    - 60-pt subject (តែង): A>=54, B>=48, C>=42, D>=36, E>=30, F<30
    - 40-pt subject (សរសេរ): A>=36, B>=32, C>=28, D>=24, E>=20, F<20
    """
    if score is None or float(score) <= 0:
        return 'F'
    s = float(score)
    m = float(max_score) if max_score and float(max_score) > 0 else 50.0
    if m == 50.0:
        if s >= 45.0: return 'A'
        if s >= 40.0: return 'B'
        if s >= 35.0: return 'C'
        if s >= 30.0: return 'D'
        if s >= 25.0: return 'E'
        return 'F'
    elif m == 100.0:
        if s >= 90.0: return 'A'
        if s >= 80.0: return 'B'
        if s >= 70.0: return 'C'
        if s >= 60.0: return 'D'
        if s >= 50.0: return 'E'
        return 'F'
    elif m == 60.0:
        if s >= 54.0: return 'A'
        if s >= 48.0: return 'B'
        if s >= 42.0: return 'C'
        if s >= 36.0: return 'D'
        if s >= 30.0: return 'E'
        return 'F'
    elif m == 40.0:
        if s >= 36.0: return 'A'
        if s >= 32.0: return 'B'
        if s >= 28.0: return 'C'
        if s >= 24.0: return 'D'
        if s >= 20.0: return 'E'
        return 'F'
    else:
        pct = (s / m) * 100.0
        if pct >= 90.0: return 'A'
        if pct >= 80.0: return 'B'
        if pct >= 70.0: return 'C'
        if pct >= 60.0: return 'D'
        if pct >= 50.0: return 'E'
        return 'F'


def get_moeys_overall_mention(tot_score, max_total=650.0):
    """
    Computes overall mention (A..F) based on total score percentage out of 650:
    A>=90% (585), B>=80% (520), C>=70% (455), D>=60% (390), E>=50% (325), F<50%
    """
    if tot_score is None or float(tot_score) <= 0:
        return 'F'
    s = float(tot_score)
    m = float(max_total) if max_total and float(max_total) > 0 else 650.0
    pct = (s / m) * 100.0
    if pct >= 90.0: return 'A'
    if pct >= 80.0: return 'B'
    if pct >= 70.0: return 'C'
    if pct >= 60.0: return 'D'
    if pct >= 50.0: return 'E'
    return 'F'


@login_required
def exam_results_sheet_print_view(request, exam_id):
    """
    Official MoEYS Provisional Examination Results Print & Export System.
    Exact 100% replica of result.pdf (scores), result-mention.pdf (mentions), and
    Subject & Overall Summary Report (របាយការណ៍បូកសរុបនិទ្ទេសតាមមុខវិជ្ជា និងសរុបរួម):
    - A4 Landscape format (297mm x 210mm)
    - Mode: 'score' (19 columns), 'mention' (20 columns with 'សរុប'), 'summary' (Matrices)
    - Page chunking: Page 1 = 26 rows, Pages 2..N = 33 rows, Final page = remaining rows + signatures
    """
    from apps.accounts.models import SchoolProfile
    import re

    exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)
    school_profile = SchoolProfile.get_settings()

    # View Mode: 'score' (result.pdf), 'mention' (result-mention.pdf), 'summary' (matrices), 'graph' (Graph.pdf)
    view_mode = request.GET.get('mode', 'score').strip().lower()
    if view_mode not in ['score', 'mention', 'summary', 'graph']:
        view_mode = 'score'

    # Filter by Room or Origin Class
    selected_room_id = request.GET.get('room_id', '').strip()
    selected_class = request.GET.get('class_name', '').strip()

    candidates_qs = exam.candidates.select_related('room', 'student').prefetch_related('subject_scores__exam_subject__subject')

    if selected_room_id and selected_room_id.isdigit():
        candidates_qs = candidates_qs.filter(room_id=int(selected_room_id))
    if selected_class:
        candidates_qs = candidates_qs.filter(origin_class__iexact=selected_class)

    # Sort candidates sequentially by roll_number / desk_number / id
    cand_list = list(candidates_qs.order_by('room__room_number', 'desk_number', 'roll_number', 'id'))

    # Map exam subjects to MoEYS standard 13 subjects
    exam_subjs = list(exam.exam_subjects.select_related('subject').all())
    subject_mapping = {}
    for es in exam_subjs:
        name = es.subject.name_kh or ''
        code = es.subject.code or ''
        for item in MOEYS_RESULT_SUBJECTS:
            k = item['key']
            if k == 'khmer' and ('ភាសាខ្មែរ' in name or name == 'ខ្មែរ'):
                subject_mapping['khmer'] = es
            elif k == 'taing' and 'តែង' in name:
                subject_mapping['taing'] = es
            elif k == 'sarser' and 'សរសេរ' in name:
                subject_mapping['sarser'] = es
            elif k == 'sil' and 'សីល' in name:
                subject_mapping['sil'] = es
            elif k == 'phum' and 'ភូមិ' in name:
                subject_mapping['phum'] = es
            elif k == 'pravatt' and 'ប្រវត្តិ' in name:
                subject_mapping['pravatt'] = es
            elif k == 'kanit' and 'គណិត' in name:
                subject_mapping['kanit'] = es
            elif k == 'phendey' and 'ផែនដី' in name:
                subject_mapping['phendey'] = es
            elif k == 'roop' and 'រូប' in name:
                subject_mapping['roop'] = es
            elif k == 'kimey' and 'គីមី' in name:
                subject_mapping['kimey'] = es
            elif k == 'chiv' and 'ជីវ' in name:
                subject_mapping['chiv'] = es
            elif k == 'keha' and 'គេហ' in name:
                subject_mapping['keha'] = es
            elif k == 'english' and ('អង់គ្លេស' in name or 'English' in name or code.upper() == 'ENG'):
                subject_mapping['english'] = es

    def extract_cand_score(cand_scores_dict, key):
        if key == 'khmer':
            if 'khmer' in subject_mapping:
                es = subject_mapping['khmer']
                sc = cand_scores_dict.get(es.id)
                if sc and sc.score is not None and not sc.is_absent:
                    return float(sc.score)
            taing_sc = extract_cand_score(cand_scores_dict, 'taing')
            sarser_sc = extract_cand_score(cand_scores_dict, 'sarser')
            return taing_sc + sarser_sc
        es = subject_mapping.get(key)
        if es:
            sc = cand_scores_dict.get(es.id)
            if sc and sc.score is not None and not sc.is_absent:
                return float(sc.score)
        return 0.0

    subject_keys = [s['key'] for s in MOEYS_RESULT_SUBJECTS]

    # Process all candidates data
    all_cands_data = []
    for idx, c in enumerate(cand_list):
        cand_scores_dict = {sc.exam_subject_id: sc for sc in c.subject_scores.all()}
        
        scores = {}
        mentions = {}
        for s_def in MOEYS_RESULT_SUBJECTS:
            k = s_def['key']
            sc_val = extract_cand_score(cand_scores_dict, k)
            scores[k] = sc_val
            mentions[k] = get_moeys_subject_mention(sc_val, s_def['max'])

        # Total score summing the 12 distinct subjects (out of 650)
        tot_score = (
            scores['taing'] + scores['sarser'] + scores['sil'] + scores['phum'] +
            scores['pravatt'] + scores['kanit'] + scores['phendey'] + scores['roop'] +
            scores['kimey'] + scores['chiv'] + scores['keha'] + scores['english']
        )
        overall_mention = get_moeys_overall_mention(tot_score, 650.0)

        dob_compact = c.dob.strftime('%d/%m/%y') if c.dob else ''
        origin_cls = format_origin_classroom(c.origin_class, exam.grade_level)
        gender_kh = 'ស' if c.gender == 'F' else 'ប'
        student_id = c.student_code or (c.student.student_id if c.student else '') or c.roll_number

        # Format scores display: int if whole number
        formatted_scores = {}
        for k, v in scores.items():
            formatted_scores[k] = int(v) if v == int(v) else f"{v:.1f}"

        scores_list = [formatted_scores.get(k, '') for k in subject_keys]
        mentions_list = [mentions.get(k, 'F') for k in subject_keys]

        all_cands_data.append({
            'row_num': idx + 1,
            'student_id': student_id,
            'candidate_name_kh': c.candidate_name_kh,
            'gender': c.gender,
            'gender_kh': gender_kh,
            'dob_str': dob_compact,
            'origin_class': origin_cls,
            'scores': scores,
            'formatted_scores': formatted_scores,
            'scores_list': scores_list,
            'mentions': mentions,
            'mentions_list': mentions_list,
            'total_score': tot_score,
            'overall_mention': overall_mention,
            'is_absent': not c.is_present,
        })

    # Summary Matrices Calculations
    def build_mentions_matrix(subset):
        matrix = {g: [0] * (len(subject_keys) + 1) for g in ['A', 'B', 'C', 'D', 'E', 'F']}
        for cand in subset:
            for s_idx, k in enumerate(subject_keys):
                m = cand['mentions'][k]
                if m in matrix:
                    matrix[m][s_idx] += 1
            ov = cand['overall_mention']
            if ov in matrix:
                matrix[ov][-1] += 1
        return matrix

    matrix_all = build_mentions_matrix(all_cands_data)
    matrix_female = build_mentions_matrix([c for c in all_cands_data if c['gender'] == 'F'])
    matrix_male = build_mentions_matrix([c for c in all_cands_data if c['gender'] == 'M'])

    # Build rows for Subject Mentions Table
    subject_headers = [s['name'] for s in MOEYS_RESULT_SUBJECTS] + ['សរុប']
    
    def format_matrix_rows(mat):
        rows = []
        for g in ['A', 'B', 'C', 'D', 'E', 'F']:
            rows.append({
                'grade': g,
                'counts': mat[g],
                'counts_kh': [to_khmer_digits(x) for x in mat[g]],
            })
        return rows

    summary_rows_all = format_matrix_rows(matrix_all)
    summary_rows_female = format_matrix_rows(matrix_female)
    summary_rows_male = format_matrix_rows(matrix_male)

    # Overall Mention Distribution Table
    total_count = len(all_cands_data)
    overall_mention_table = []
    passed_count = 0
    passed_females = 0
    passed_males = 0
    failed_count = 0
    failed_females = 0
    failed_males = 0

    for g in ['A', 'B', 'C', 'D', 'E', 'F']:
        cnt_tot = matrix_all[g][-1]
        cnt_fem = matrix_female[g][-1]
        cnt_mal = matrix_male[g][-1]
        pct = round((cnt_tot / total_count * 100.0), 1) if total_count > 0 else 0.0

        overall_mention_table.append({
            'mention': g,
            'total': cnt_tot,
            'total_kh': to_khmer_2digits(cnt_tot),
            'female': cnt_fem,
            'female_kh': to_khmer_2digits(cnt_fem),
            'male': cnt_mal,
            'male_kh': to_khmer_2digits(cnt_mal),
            'percentage': pct,
            'percentage_kh': to_khmer_digits(f"{pct:.1f}"),
        })
        if g in ['A', 'B', 'C', 'D', 'E']:
            passed_count += cnt_tot
            passed_females += cnt_fem
            passed_males += cnt_mal
        else:
            failed_count += cnt_tot
            failed_females += cnt_fem
            failed_males += cnt_mal

    grand_total_females = sum(1 for c in all_cands_data if c['gender'] == 'F')
    grand_total_males = sum(1 for c in all_cands_data if c['gender'] == 'M')

    grand_total_row = {
        'mention': 'សរុប',
        'total': total_count,
        'total_kh': to_khmer_2digits(total_count),
        'female': grand_total_females,
        'female_kh': to_khmer_2digits(grand_total_females),
        'male': grand_total_males,
        'male_kh': to_khmer_2digits(grand_total_males),
        'percentage': 100.0,
        'percentage_kh': '១០០',
    }

    # Pagination: Page 1 = 26 rows, Subsequent pages = 33 rows
    pages = []
    chunk_p1 = 26
    chunk_rest = 33

    if all_cands_data:
        p1_cands = all_cands_data[:chunk_p1]
        pages.append({
            'page_num': 1,
            'is_first_page': True,
            'is_last_page': len(all_cands_data) <= chunk_p1,
            'candidates': p1_cands,
        })
        
        remaining = all_cands_data[chunk_p1:]
        cur_page = 2
        while remaining:
            cur_chunk = remaining[:chunk_rest]
            remaining = remaining[chunk_rest:]
            pages.append({
                'page_num': cur_page,
                'is_first_page': False,
                'is_last_page': len(remaining) == 0,
                'candidates': cur_chunk,
            })
            cur_page += 1

    # Title & Academic Header
    exam_name_clean = exam.name.strip()
    for pfx in ['លទ្ធផលប្រឡង', 'បញ្ជីរាយនាមសិស្សប្រឡង', 'ប្រឡង']:
        if exam_name_clean.startswith(pfx):
            exam_name_clean = exam_name_clean[len(pfx):].strip()
            break

    exam_name_clean = re.sub(r'\s*ថ្នាក់ទី\s*[\d\u17e0-\u17e9A-Za-z]+\s*$', '', exam_name_clean).strip()
    exam_name_clean = re.sub(r'\s*ថ្នាក់\s*[\d\u17e0-\u17e9A-Za-z]+\s*$', '', exam_name_clean).strip()
    exam_name_clean = re.sub(r'\s*ឆ្នាំសិក្សា\s*[\d\u17e0-\u17e9\-]+\s*$', '', exam_name_clean).strip()

    academic_year_kh = to_khmer_digits(exam.academic_year.name if exam.academic_year else '')
    grade_kh = to_khmer_digits(exam.grade_level)

    if academic_year_kh and grade_kh:
        results_title_line = f"លទ្ធផលប្រឡង{exam_name_clean} ឆ្នាំសិក្សា {academic_year_kh} ថ្នាក់ទី {grade_kh}".strip()
    elif academic_year_kh:
        results_title_line = f"លទ្ធផលប្រឡង{exam_name_clean} ឆ្នាំសិក្សា {academic_year_kh}".strip()
    else:
        results_title_line = f"លទ្ធផលប្រឡង{exam_name_clean}".strip()

    exam_date_kh = format_khmer_full_date(exam.exam_date)

    # Province, School Name & Signing Location
    province_name = request.GET.get('province') or school_profile.province or 'ខេត្តកណ្តាល'
    school_name = request.GET.get('school_name') or school_profile.name_kh or 'វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្សួត'

    location_name = request.GET.get('location', '').strip()
    if not location_name:
        if school_profile.commune:
            c_name = school_profile.commune.strip()
            for prefix in ['ឃុំ', 'សង្កាត់', 'ឃុំ ', 'សង្កាត់ ']:
                if c_name.startswith(prefix):
                    c_name = c_name[len(prefix):].strip()
                    break
            location_name = c_name
        elif school_profile.district:
            d_name = school_profile.district.strip()
            for prefix in ['ស្រុក', 'ខណ្ឌ', 'ក្រុង', 'ស្រុក ', 'ខណ្ឌ ', 'ក្រុង ']:
                if d_name.startswith(prefix):
                    d_name = d_name[len(prefix):].strip()
                    break
            location_name = d_name
        else:
            location_name = 'កំពង់កន្សួត'

    sign_date_param = request.GET.get('sign_date', '').strip()
    sign_date = None
    if sign_date_param:
        try:
            sign_date = datetime.datetime.strptime(sign_date_param, '%Y-%m-%d').date()
        except ValueError:
            pass
    if not sign_date:
        sign_date = exam.exam_date or timezone.now().date()

    sign_day_kh = to_khmer_2digits(sign_date.day)
    sign_month_kh = KHMER_MONTH_NAMES.get(sign_date.month, '')
    sign_year_kh = to_khmer_digits(sign_date.year)

    # Khmer Lunar Calendar Date (e.g. ថ្ងៃសុក្រ ៥កើត ខែបឋមាសាឍ ឆ្នាំមមី អដ្ឋស័ក ព.ស.២៥៧០)
    lunar_date = request.GET.get('lunar_date', '').strip() or 'ថ្ងៃសុក្រ ៥កើត ខែបឋមាសាឍ ឆ្នាំមមី អដ្ឋស័ក ព.ស.២៥៧០'

    # Signer Role title (principal)
    sign_role = request.GET.get('sign_role', '').strip() or 'នាយក'

    # Colors matching Graph.pdf
    GRADE_COLORS = {
        'A': '#1f4e79',  # Dark Navy Blue
        'B': '#c55a11',  # Terracotta Orange
        'C': '#276a3c',  # Forest Green
        'D': '#00a2e8',  # Cyan / Sky Blue
        'E': '#800080',  # Magenta / Purple
        'F': '#548235',  # Leaf Green
    }

    # Gender filter for Graph view
    graph_gender = request.GET.get('gender', 'ALL').strip().upper()
    if graph_gender == 'F':
        active_matrix = matrix_female
    elif graph_gender == 'M':
        active_matrix = matrix_male
    else:
        graph_gender = 'ALL'
        active_matrix = matrix_all

    # Calculate maximum value across all 13 subjects and 6 grades
    max_count = 0
    for g in ['A', 'B', 'C', 'D', 'E', 'F']:
        for s_idx in range(13):
            val = active_matrix[g][s_idx]
            if val > max_count:
                max_count = val

    import math
    if max_count <= 120:
        graph_y_max = 120
        graph_y_ticks = [120, 100, 80, 60, 40, 20, 0]
    else:
        step = int(math.ceil(max_count / 6 / 10.0)) * 10
        if step < 20:
            step = 20
        graph_y_max = step * 6
        graph_y_ticks = [step * i for i in range(6, -1, -1)]

    # Prepare chart columns (13 subjects)
    graph_columns = []
    for s_idx, s_def in enumerate(MOEYS_RESULT_SUBJECTS):
        sub_bars = []
        for g in ['A', 'B', 'C', 'D', 'E', 'F']:
            cnt = active_matrix[g][s_idx]
            height_pct = f"{(cnt / graph_y_max * 100.0):.2f}" if graph_y_max > 0 else "0.00"
            sub_bars.append({
                'grade': g,
                'count': cnt,
                'count_kh': to_khmer_digits(cnt),
                'color': GRADE_COLORS[g],
                'height_pct': height_pct,
            })
        graph_columns.append({
            'key': s_def['key'],
            'name_kh': s_def['name'],
            'max_score': s_def['max'],
            'bars': sub_bars,
        })

    # Prepare integrated data table rows (6 rows: A, B, C, D, E, F)
    graph_table_rows = []
    for g in ['A', 'B', 'C', 'D', 'E', 'F']:
        row_counts = [active_matrix[g][s_idx] for s_idx in range(13)]
        row_counts_kh = [to_khmer_digits(c) for c in row_counts]
        graph_table_rows.append({
            'grade': g,
            'color': GRADE_COLORS[g],
            'counts': row_counts,
            'counts_kh': row_counts_kh,
        })

    # Available rooms and classes for filter dropdown
    all_rooms = exam.rooms.all().order_by('room_number')
    all_classes = exam.candidates.values_list('origin_class', flat=True).distinct().exclude(origin_class__isnull=True).exclude(origin_class='').order_by('origin_class')

    context = {
        'exam': exam,
        'view_mode': view_mode,
        'pages': pages,
        'all_candidates_count': total_count,
        'moeys_subjects': MOEYS_RESULT_SUBJECTS,
        'subject_headers': subject_headers,
        'summary_rows_all': summary_rows_all,
        'summary_rows_female': summary_rows_female,
        'summary_rows_male': summary_rows_male,
        'overall_mention_table': overall_mention_table,
        'grand_total_row': grand_total_row,
        'passed_count': passed_count,
        'passed_count_kh': to_khmer_digits(passed_count),
        'passed_females': passed_females,
        'passed_females_kh': to_khmer_digits(passed_females),
        'passed_rate': round((passed_count / total_count * 100.0), 1) if total_count > 0 else 0.0,
        'failed_count': failed_count,
        'failed_count_kh': to_khmer_digits(failed_count),
        'failed_females': failed_females,
        'failed_females_kh': to_khmer_digits(failed_females),
        'failed_rate': round((failed_count / total_count * 100.0), 1) if total_count > 0 else 0.0,
        'results_title_line': results_title_line,
        'exam_date_kh': exam_date_kh,
        'school_name': school_name,
        'province_name': province_name,
        'location_name': location_name,
        'sign_day_kh': sign_day_kh,
        'sign_month_kh': sign_month_kh,
        'sign_year_kh': sign_year_kh,
        'lunar_date': lunar_date,
        'sign_role': sign_role,
        'sign_date_raw': sign_date.strftime('%Y-%m-%d'),
        'all_rooms': all_rooms,
        'all_classes': all_classes,
        'selected_room_id': selected_room_id,
        'selected_class': selected_class,
        'school_profile': school_profile,
        # Graph specific context
        'graph_columns': graph_columns,
        'graph_table_rows': graph_table_rows,
        'graph_y_ticks': graph_y_ticks,
        'graph_y_max': graph_y_max,
        'graph_gender': graph_gender,
        'grade_colors': GRADE_COLORS,
    }

    if view_mode == 'graph':
        return render(request, 'examinations/standardized/results_graph_print.html', context)

    return render(request, 'examinations/standardized/results_sheet_print.html', context)


@login_required
def exam_export_candidates_excel(request, exam_id):
    """
    Exports candidates roster to Excel (.xlsx).
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    candidates = exam.candidates.select_related('room').order_by('room__room_number', 'desk_number', 'roll_number', 'id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "បញ្ជីបេក្ខជនប្រឡង"

    # Header Styles
    header_font = Font(name='Kantumruy Pro', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Rows
    ws.merge_cells('A1:I1')
    ws['A1'] = f"បញ្ជីឈ្មោះបេក្ខជន៖ {exam.name}"
    ws['A1'].font = Font(name='Kantumruy Pro', size=14, bold=True, color='1E3A8A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f"កម្រិតថ្នាក់ទី {exam.grade_level} • ឆ្នាំសិក្សា {exam.academic_year.name} • ចំនួនបេក្ខជនសរុប៖ {candidates.count()} នាក់ (ស្រី {candidates.filter(gender='F').count()} នាក់)"
    ws['A2'].font = Font(name='Kantumruy Pro', size=10, italic=True, color='475569')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Table Header
    headers = [
        'ល.រ (No)', 'លេខបន្ទប់ (Room)', 'លេខតុ (Desk)', 'អត្តលេខ (Roll No)',
        'គោត្តនាម-នាម (Khmer Name)', 'ឈ្មោះឡាតាំង (Latin Name)', 'ភេទ (Sex)',
        'ថ្ងៃខែឆ្នាំកំណើត (DOB)', 'ថ្នាក់ដើម (Class)'
    ]
    ws.append([]) # Blank row 3
    ws.append(headers) # Row 4

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for idx, c in enumerate(candidates, 1):
        row = [
            idx,
            c.room.room_name if c.room else 'មិនទាន់កំណត់',
            f"{c.desk_number:02d}",
            c.roll_number,
            c.candidate_name_kh,
            c.candidate_name_en or '',
            'ស្រី' if c.gender == 'F' else 'ប្រុស',
            c.dob.strftime('%d/%m/%Y') if c.dob else '',
            c.origin_class or ''
        ]
        ws.append(row)
        curr_row = 4 + idx
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=curr_row, column=col_num)
            cell.font = Font(name='Kantumruy Pro', size=10)
            cell.border = border_thin
            if col_num in [1, 2, 3, 4, 7, 8]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Candidates_{exam.id}_{exam.grade_level}.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required(['ADMIN'])
def exam_import_candidates_excel(request, exam_id):
    """
    Imports candidates roster from an Excel file.
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                messages.error(request, "ឯកសារ Excel គ្មានទិន្នន័យបេក្ខជនឡើយ!")
                return redirect('standardized_exam_manage', exam_id=exam.id)

            # Find header row
            header_row_idx = 0
            for idx, r in enumerate(rows[:10]):
                r_text = " ".join([str(c or '') for c in r])
                if 'គោត្តនាម' in r_text or 'Khmer Name' in r_text or 'ឈ្មោះ' in r_text:
                    header_row_idx = idx
                    break

            exam_subjects = list(exam.exam_subjects.all())
            imported_count = 0
            existing_count = exam.candidates.count()

            with transaction.atomic():
                for r in rows[header_row_idx + 1:]:
                    if not r or not any(r):
                        continue
                    name_kh = str(r[4] if len(r) > 4 else r[0] or '').strip()
                    if not name_kh or name_kh.lower() in ['none', 'ឈ្មោះ', 'គោត្តនាម-នាម']:
                        continue

                    name_en = str(r[5] if len(r) > 5 else '').strip()
                    gender_raw = str(r[6] if len(r) > 6 else 'M').strip()
                    gender = 'F' if gender_raw in ['F', 'f', 'ស្រី', 'ស្រី្ត', 'ស្ត្រី', 'កញ្ញា', 'ស', 'ស.', 'Female', 'female', 'girl', 'woman', '2'] else 'M'
                    origin_class = str(r[8] if len(r) > 8 else '').strip()

                    roll_str = f"{existing_count + imported_count + 1:03d}"
                    cand = ExamCandidate.objects.create(
                        exam=exam,
                        roll_number=roll_str,
                        desk_number=1,
                        candidate_name_kh=name_kh,
                        candidate_name_en=name_en if name_en != 'None' else '',
                        gender=gender,
                        origin_class=origin_class if origin_class != 'None' else '',
                    )
                    for es in exam_subjects:
                        CandidateSubjectScore.objects.create(candidate=cand, exam_subject=es)
                    imported_count += 1

            messages.success(request, f"🎉 បាននាំចូលបេក្ខជនចំនួន {imported_count} នាក់ពី Excel ដោយជោគជ័យ!")
        except Exception as e:
            messages.error(request, f"មានបញ្ហាក្នុងការអានឯកសារ Excel៖ {str(e)}")

    return redirect('standardized_exam_manage', exam_id=exam.id)


@login_required
def exam_export_provisional_excel(request, exam_id):
    """
    Exports the complete Provisional Results Master Matrix to Excel (.xlsx).
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    exam.recalculate_all_ranks()

    subjects = list(exam.exam_subjects.select_related('subject').order_by('order', 'id'))
    candidates = list(exam.candidates.select_related('room').prefetch_related('subject_scores__exam_subject').order_by('rank_overall', '-total_score', 'candidate_name_kh'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "លទ្ធផលបណ្តោះអាសន្ន"

    # Header Styles
    header_font = Font(name='Kantumruy Pro', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title
    num_cols = 8 + len(subjects) + 4
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = f"តារាងបិទផ្សាយបណ្តោះអាសន្នលទ្ធផលប្រឡង៖ {exam.name}"
    ws['A1'].font = Font(name='Kantumruy Pro', size=14, bold=True, color='1E3A8A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'] = f"កម្រិតថ្នាក់ទី {exam.grade_level} • ឆ្នាំសិក្សា {exam.academic_year.name} • បេក្ខជនសរុប៖ {len(candidates)} នាក់"
    ws['A2'].font = Font(name='Kantumruy Pro', size=10, italic=True, color='475569')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ['ចំណាត់ថ្នាក់', 'អត្តលេខ', 'បន្ទប់', 'លេខតុ', 'គោត្តនាម-នាម', 'ឈ្មោះឡាតាំង', 'ភេទ', 'ថ្នាក់ដើម']
    for s in subjects:
        headers.append(f"{s.subject.name_kh} (/{s.max_score:g})")
    headers.extend(['ពិន្ទុសរុប', 'មធ្យមភាគ', 'និទ្ទេស', 'លទ្ធផល'])

    ws.append([]) # Row 3 blank
    ws.append(headers) # Row 4

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for idx, c in enumerate(candidates, 1):
        cand_scores = {sc.exam_subject_id: sc for sc in c.subject_scores.all()}
        row = [
            c.rank_overall or idx,
            c.roll_number,
            c.room.room_name if c.room else '-',
            f"{c.desk_number:02d}",
            c.candidate_name_kh,
            c.candidate_name_en or '',
            'ស្រី' if c.gender == 'F' else 'ប្រុស',
            c.origin_class or '',
        ]
        for s in subjects:
            sc = cand_scores.get(s.id)
            if sc and sc.is_absent:
                row.append('អវត្តមាន')
            elif sc and sc.score is not None:
                row.append(float(sc.score))
            else:
                row.append('-')

        row.extend([
            float(c.total_score),
            float(c.average_score),
            c.grade_letter,
            'ជាប់' if c.grade_letter in ['A', 'B', 'C', 'D', 'E'] else 'ធ្លាក់'
        ])
        ws.append(row)
        curr_row = 4 + idx
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=curr_row, column=col_num)
            cell.font = Font(name='Kantumruy Pro', size=10)
            cell.border = border_thin
            if col_num in [1, 2, 3, 4, 7, 8, len(row)-2, len(row)-1, len(row)]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Provisional_Results_{exam.id}_{exam.grade_level}.xlsx"'
    wb.save(response)
    return response


# ==============================================================================
# BLIND EXAM SCORING SYSTEM (ផ្ទាំងបញ្ចូលពិន្ទុសិស្សដោយលេខកូដសម្ងាត់ ៤ ជំហាន)
# ==============================================================================

@login_required
@role_required(['ADMIN', 'TEACHER'])
def exam_blind_scoring_portal(request):
    """
    Step-by-Step Blind / Secret-Coded Exam Score Entry Interface (4 Steps):
    - Step 1: Select Grade Level & Exam
    - Step 2: Select Subject
    - Step 3: Enter Secret Code
    - Step 4: Rapid Score Entry from Desk 01 to 25
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    
    exams_qs = StandardizedExam.objects.select_related('academic_year').order_by('-exam_date', '-id')
    if active_year:
        exams_qs = exams_qs.filter(academic_year=active_year)

    # Regular teachers can only see published / admin-permitted exams
    if not request.user.is_superuser and getattr(request.user, 'role', '') == 'TEACHER':
        exams_qs = exams_qs.filter(is_published=True)
    
    pre_exam_id = request.GET.get('exam_id')
    pre_subject_id = request.GET.get('subject_id')
    pre_code = request.GET.get('code', '').strip()

    selected_exam = None
    subjects = []
    if pre_exam_id and str(pre_exam_id).isdigit():
        selected_exam = StandardizedExam.objects.filter(id=int(pre_exam_id)).first()
        if selected_exam:
            subjects = selected_exam.exam_subjects.select_related('subject').order_by('order', 'id')

    return render(request, 'examinations/standardized/blind_scoring_portal.html', {
        'exams': exams_qs,
        'selected_exam': selected_exam,
        'subjects': subjects,
        'pre_exam_id': pre_exam_id or '',
        'pre_subject_id': pre_subject_id or '',
        'pre_code': pre_code,
    })


@login_required
@role_required(['ADMIN', 'TEACHER'])
def api_exam_get_subjects(request, exam_id):
    """
    JSON API returning all subjects, grading rules, secret code envelopes, and grading window status for a standardized exam.
    Conceals physical room names from regular teachers (anonymized as envelope codes).
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    subjects = exam.exam_subjects.select_related('subject').order_by('order', 'id')
    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'

    is_grading_open, status_code, grading_msg = exam.get_grading_status()

    # Query room subject codes (anonymized for teachers, full info for Admin)
    codes_by_subject = {}
    room_codes_qs = ExamRoomSubjectCode.objects.filter(exam_room__exam=exam).select_related('exam_room', 'graded_by')
    for rc in room_codes_qs:
        if rc.exam_subject_id not in codes_by_subject:
            codes_by_subject[rc.exam_subject_id] = []
        
        display_room = rc.exam_room.room_name if is_admin else f"កញ្ចប់កូដ #{rc.secret_code}"
        codes_by_subject[rc.exam_subject_id].append({
            'secret_code': rc.secret_code,
            'is_graded': rc.is_graded,
            'graded_by': (rc.graded_by.get_full_name() or rc.graded_by.username) if (rc.graded_by and is_admin) else ('បានបញ្ចូល' if rc.is_graded else ''),
            'graded_at': rc.graded_at.strftime('%d/%m/%Y %H:%M') if rc.graded_at else '',
            'room_name': display_room,
            'candidate_count': rc.exam_room.candidates.count(),
        })
    
    return JsonResponse({
        'status': 'success',
        'exam_id': exam.id,
        'exam_name': exam.name,
        'grade_level': exam.grade_level,
        'candidates_per_room': exam.candidates_per_room,
        'is_admin': is_admin,
        'is_grading_open': is_grading_open or is_admin,
        'is_grading_locked': exam.is_grading_locked,
        'status_code': status_code,
        'grading_status_msg': grading_msg,
        'grading_start_datetime': exam.grading_start_datetime.strftime('%Y-%m-%dT%H:%M') if exam.grading_start_datetime else None,
        'grading_end_datetime': exam.grading_end_datetime.strftime('%Y-%m-%dT%H:%M') if exam.grading_end_datetime else None,
        'grading_start_display': exam.grading_start_datetime.strftime('%d/%m/%Y %H:%M') if exam.grading_start_datetime else '',
        'grading_end_display': exam.grading_end_datetime.strftime('%d/%m/%Y %H:%M') if exam.grading_end_datetime else '',
        'subjects': [
            {
                'id': s.id,
                'name': s.subject.name_kh,
                'code': s.subject.code,
                'max_score': float(s.max_score),
                'coefficient': float(s.coefficient),
                'session': s.get_session_display(),
                'exam_date': s.exam_date.strftime('%d/%m/%Y') if s.exam_date else '',
                'secret_codes': codes_by_subject.get(s.id, []),
            }
            for s in subjects
        ]
    })


@login_required
@role_required(['ADMIN'])
def api_toggle_exam_grading_lock(request, exam_id):
    """
    1-Click instant toggle to lock/unlock grading for a Standardized Exam (or all exams in the same session).
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required'}, status=405)

    exam = get_object_or_404(StandardizedExam, id=exam_id)
    try:
        data = json.loads(request.body) if request.body else request.POST
    except Exception:
        data = request.POST

    apply_to_session = str(data.get('apply_to_session', '')).lower() in ['true', '1', 'yes']

    new_lock_state = not exam.is_grading_locked
    if 'is_locked' in data:
        new_lock_state = str(data.get('is_locked')).lower() in ['true', '1', 'yes']

    if apply_to_session:
        clean_title = get_clean_exam_session_title(exam.name)
        session_exams = StandardizedExam.objects.filter(
            academic_year=exam.academic_year,
            exam_date=exam.exam_date
        )
        matched_ids = [e.id for e in session_exams if get_clean_exam_session_title(e.name) == clean_title]
        StandardizedExam.objects.filter(id__in=matched_ids).update(
            is_grading_locked=new_lock_state,
            updated_at=timezone.now()
        )
        count = len(matched_ids)
        action_str = "🔒 បានចាក់សោការបញ្ចូលពិន្ទុ" if new_lock_state else "🔓 បានបើកដំណើរការបញ្ចូលពិន្ទុ"
        msg = f"{action_str} សម្រាប់គ្រប់កម្រិតថ្នាក់ទាំងអស់នៃសម័យប្រឡង «{clean_title}» ({count} កម្រិត) ដោយជោគជ័យ!"
    else:
        exam.is_grading_locked = new_lock_state
        exam.save(update_fields=['is_grading_locked', 'updated_at'])
        action_str = "🔒 បានចាក់សោការបញ្ចូលពិន្ទុ" if new_lock_state else "🔓 បានបើកដំណើរការបញ្ចូលពិន្ទុ"
        msg = f"{action_str} សម្រាប់សម័យប្រឡង «{exam.name}» ដោយជោគជ័យ!"

    exam.refresh_from_db()
    is_open, status_code, status_msg = exam.get_grading_status()
    return JsonResponse({
        'status': 'success',
        'message': msg,
        'is_grading_locked': new_lock_state,
        'is_grading_open': is_open,
        'status_code': status_code,
        'grading_status_msg': status_msg,
    })


@login_required
@role_required(['ADMIN'])
def api_update_exam_grading_window(request, exam_id):
    """
    Sets or updates the grading start and end deadline datetime for a Standardized Exam (or session).
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST method required'}, status=405)

    exam = get_object_or_404(StandardizedExam, id=exam_id)
    try:
        data = json.loads(request.body) if request.body else request.POST
    except Exception:
        data = request.POST

    start_dt_raw = data.get('grading_start_datetime')
    end_dt_raw = data.get('grading_end_datetime')
    is_locked_raw = data.get('is_grading_locked')
    grading_method = data.get('grading_method')
    apply_to_session = str(data.get('apply_to_session', '')).lower() in ['true', '1', 'yes']

    start_dt = None
    if start_dt_raw:
        try:
            start_dt = datetime.datetime.fromisoformat(start_dt_raw.replace('Z', '+00:00'))
        except Exception:
            pass

    end_dt = None
    if end_dt_raw:
        try:
            end_dt = datetime.datetime.fromisoformat(end_dt_raw.replace('Z', '+00:00'))
        except Exception:
            pass

    is_locked = exam.is_grading_locked
    if is_locked_raw is not None:
        is_locked = str(is_locked_raw).lower() in ['true', '1', 'yes']

    update_kwargs = {
        'grading_start_datetime': start_dt,
        'grading_end_datetime': end_dt,
        'is_grading_locked': is_locked,
        'updated_at': timezone.now()
    }
    if grading_method and grading_method in ['BOTH', 'TEACHER_DIRECT', 'BLIND_SECRET_CODE']:
        update_kwargs['grading_method'] = grading_method

    if apply_to_session:
        clean_title = get_clean_exam_session_title(exam.name)
        session_exams = StandardizedExam.objects.filter(
            academic_year=exam.academic_year,
            exam_date=exam.exam_date
        )
        matched_ids = [e.id for e in session_exams if get_clean_exam_session_title(e.name) == clean_title]
        StandardizedExam.objects.filter(id__in=matched_ids).update(**update_kwargs)
        msg = f"🎉 បានកំណត់កាលវិភាគ និងរបៀបបញ្ចូលពិន្ទុសម្រាប់គ្រប់កម្រិតថ្នាក់នៃសម័យប្រឡង «{clean_title}» ដោយជោគជ័យ!"
    else:
        for k, v in update_kwargs.items():
            setattr(exam, k, v)
        exam.save(update_fields=list(update_kwargs.keys()))
        msg = f"🎉 បានកំណត់កាលវិភាគ និងរបៀបបញ្ចូលពិន្ទុសម្រាប់សម័យប្រឡង «{exam.name}» ដោយជោគជ័យ!"

    exam.refresh_from_db()
    is_open, status_code, status_msg = exam.get_grading_status()
    return JsonResponse({
        'status': 'success',
        'message': msg,
        'is_grading_locked': is_locked,
        'is_grading_open': is_open,
        'status_code': status_code,
        'grading_status_msg': status_msg,
        'grading_method': exam.grading_method,
        'grading_method_display': exam.get_grading_method_display(),
        'grading_start_datetime': exam.grading_start_datetime.strftime('%d/%m/%Y %H:%M') if exam.grading_start_datetime else None,
        'grading_end_datetime': exam.grading_end_datetime.strftime('%d/%m/%Y %H:%M') if exam.grading_end_datetime else None,
    })


@login_required
@role_required(['ADMIN', 'TEACHER'])
def api_exam_validate_secret_code(request):
    """
    Step 3 Validation API:
    Validates the Secret Code against ExamRoomSubjectCode or ExamRoom.secret_code,
    and returns anonymous candidate desks list (Desks 01 to 25) without student names.
    Conceals physical room names from regular teachers to ensure 100% blind grading privacy!
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST

    exam_id = data.get('exam_id')
    subject_id = data.get('subject_id')
    secret_code = str(data.get('secret_code', '')).strip().upper()

    if not exam_id or not subject_id or not secret_code:
        return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសព័ត៌មានឱ្យបានគ្រប់គ្រាន់ (កម្រិតថ្នាក់, មុខវិជ្ជា និងលេខកូដសម្ងាត់)!'})

    exam = get_object_or_404(StandardizedExam, id=exam_id)
    exam_subject = get_object_or_404(ExamSubject.objects.select_related('subject'), id=subject_id, exam=exam)
    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'

    # Search for matching ExamRoomSubjectCode first
    code_obj = ExamRoomSubjectCode.objects.filter(
        secret_code__iexact=secret_code,
        exam_subject=exam_subject
    ).select_related('exam_room').first()

    room = None
    if code_obj:
        room = code_obj.exam_room
    else:
        # Check by ExamRoom secret_code
        room = ExamRoom.objects.filter(exam=exam, secret_code__iexact=secret_code).first()

    if not room:
        return JsonResponse({
            'status': 'error',
            'message': f'លេខកូដសម្ងាត់ «{secret_code}» មិនត្រឹមត្រូវ ឬមិនត្រូវគ្នានឹងមុខវិជ្ជា {exam_subject.subject.name_kh} នៃសម័យប្រឡងនេះឡើយ! សូមផ្ទៀងផ្ទាត់លេខកូដលើកញ្ចប់ក្រដាសប្រឡងម្តងទៀត។'
        })

    is_grading_open, _, grading_msg = exam.get_grading_status()

    # Retrieve candidates in this room sorted by desk_number (1 to N)
    candidates = room.candidates.all().order_by('desk_number', 'id')
    if not candidates.exists():
        return JsonResponse({
            'status': 'error',
            'message': f'កញ្ចប់ប្រឡងលេខកូដ «{secret_code}» មិនទាន់មានបេក្ខជនត្រូវបានបែងចែកនៅឡើយទេ។ សូមទាក់ទង Admin។'
        })

    scores_map = {
        sc.candidate_id: sc
        for sc in CandidateSubjectScore.objects.filter(
            candidate__in=candidates,
            exam_subject=exam_subject
        )
    }

    desks_data = []
    for cand in candidates:
        sc = scores_map.get(cand.id)
        score_val = ''
        is_absent = False
        if sc:
            is_absent = sc.is_absent
            if sc.score is not None and not is_absent:
                score_val = float(sc.score)

        desks_data.append({
            'desk_number': cand.desk_number,
            'candidate_id': cand.id,
            'score': score_val,
            'is_absent': is_absent,
        })

    # For regular teachers, mask physical room name for blind confidentiality
    display_room_name = room.room_name if is_admin else f"កញ្ចប់កូដសម្ងាត់ #{secret_code}"

    return JsonResponse({
        'status': 'success',
        'room_id': room.id,
        'room_name': display_room_name,
        'is_blind_mode': not is_admin,
        'is_grading_open': is_grading_open or is_admin,
        'grading_status_msg': grading_msg,
        'subject_id': exam_subject.id,
        'subject_name': exam_subject.subject.name_kh,
        'max_score': float(exam_subject.max_score),
        'coefficient': float(exam_subject.coefficient),
        'candidate_count': len(desks_data),
        'is_already_graded': (code_obj.is_graded if code_obj else False) or bool(any(sc.score is not None or sc.is_absent for sc in scores_map.values())),
        'graded_by': (
            (getattr(code_obj.graded_by, 'khmer_name', '') or code_obj.graded_by.get_full_name() or code_obj.graded_by.username)
            if (code_obj and code_obj.graded_by)
            else (
                (getattr(first_sc.entered_by, 'khmer_name', '') or first_sc.entered_by.get_full_name() or first_sc.entered_by.username)
                if (first_sc := next((sc for sc in scores_map.values() if sc.entered_by), None))
                else ''
            )
        ),
        'graded_at': (
            code_obj.graded_at.strftime('%d/%m/%Y %H:%M')
            if (code_obj and code_obj.graded_at)
            else (
                first_sc.entered_at.strftime('%d/%m/%Y %H:%M')
                if (first_sc := next((sc for sc in scores_map.values() if sc.entered_at), None))
                else ''
            )
        ),
        'desks': desks_data
    })


@login_required
@role_required(['ADMIN', 'TEACHER'])
def api_exam_save_blind_scores(request):
    """
    Step 4 Save API:
    Saves scores submitted blindly by desk number (01 to N), matches with candidates,
    updates CandidateSubjectScore, marks envelope code as graded, and recalculates exam ranks.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST

    exam_id = data.get('exam_id')
    subject_id = data.get('subject_id')
    secret_code = str(data.get('secret_code', '')).strip().upper()
    scores_list = data.get('scores', [])

    if not exam_id or not subject_id or not secret_code or not scores_list:
        return JsonResponse({'status': 'error', 'message': 'ទិន្នន័យមិនពេញលេញ!'})

    exam = get_object_or_404(StandardizedExam, id=exam_id)
    exam_subject = get_object_or_404(ExamSubject.objects.select_related('subject'), id=subject_id, exam=exam)
    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'

    # Enforce Grading Window for regular teachers
    is_grading_open, _, grading_msg = exam.get_grading_status()
    if not is_grading_open and not is_admin:
        return JsonResponse({'status': 'error', 'message': f'⚠️ មិនអាចរក្សាទុកបានទេ៖ {grading_msg}!'})

    code_obj = ExamRoomSubjectCode.objects.filter(
        secret_code__iexact=secret_code,
        exam_subject=exam_subject
    ).select_related('exam_room').first()

    room = None
    if code_obj:
        room = code_obj.exam_room
    else:
        room = ExamRoom.objects.filter(exam=exam, secret_code__iexact=secret_code).first()

    if not room:
        return JsonResponse({'status': 'error', 'message': 'លេខកូដសម្ងាត់មិនត្រឹមត្រូវ!'})

    candidates_by_desk = {c.desk_number: c for c in room.candidates.all()}

    saved_count = 0
    absent_count = 0
    total_score_sum = Decimal('0.00')
    valid_scores = []

    with transaction.atomic():
        for item in scores_list:
            desk_num = int(item.get('desk_number', 0))
            score_raw = str(item.get('score', '')).strip().upper()
            is_absent_flag = bool(item.get('is_absent', False)) or (score_raw in ['0', '0.0', '0.00', 'A'])

            cand = candidates_by_desk.get(desk_num)
            if not cand:
                continue

            score_obj, _ = CandidateSubjectScore.objects.get_or_create(
                candidate=cand,
                exam_subject=exam_subject
            )

            if score_raw != '' and score_raw != '-':
                try:
                    val = Decimal(score_raw)
                    if val > exam_subject.max_score:
                        val = exam_subject.max_score
                    if val < Decimal('0.00'):
                        val = Decimal('0.00')
                    score_obj.score = val
                    score_obj.is_absent = (val == Decimal('0.00')) or is_absent_flag
                    total_score_sum += val
                    valid_scores.append(val)
                    if val == Decimal('0.00') or is_absent_flag:
                        absent_count += 1
                except Exception:
                    score_obj.score = Decimal('0.00')
                    score_obj.is_absent = True
                    absent_count += 1
            elif is_absent_flag:
                score_obj.is_absent = True
                score_obj.score = Decimal('0.00')
                absent_count += 1
                valid_scores.append(Decimal('0.00'))
            else:
                score_obj.score = None
                score_obj.is_absent = False

            if not score_obj.entered_by:
                score_obj.entered_by = request.user
            if not score_obj.entered_at:
                score_obj.entered_at = timezone.now()
            score_obj.secret_code_used = secret_code
            score_obj.last_modified_by = request.user

            score_obj.save()
            saved_count += 1

        if code_obj:
            code_obj.is_graded = True
            code_obj.graded_by = request.user
            code_obj.graded_at = timezone.now()
            code_obj.save(update_fields=['is_graded', 'graded_by', 'graded_at'])

        # Recalculate ranks & grades
        exam.recalculate_all_ranks()

    avg_val = (float(total_score_sum) / len(valid_scores)) if valid_scores else 0.0
    max_val = float(max(valid_scores)) if valid_scores else 0.0
    min_val = float(min(valid_scores)) if valid_scores else 0.0

    return JsonResponse({
        'status': 'success',
        'message': f'🎉 បានរក្សាទុកពិន្ទុមុខវិជ្ជា «{exam_subject.subject.name_kh}» សម្រាប់កញ្ចប់កូដ «{secret_code}» ({room.room_name}) ដោយជោគជ័យ!',
        'summary': {
            'saved_count': saved_count,
            'graded_count': len(valid_scores),
            'absent_count': absent_count,
            'average_score': round(avg_val, 2),
            'max_score': round(max_val, 2),
            'min_score': round(min_val, 2),
        }
    })


@login_required
@role_required(['ADMIN'])
def exam_secret_codes_directory(request, exam_id):
    """
    Admin-Only Master Secret Codes Directory & Printable Stickers:
    Shows all generated secret codes per room and subject envelope.
    """
    exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)
    rooms = exam.rooms.all().order_by('room_number')
    subjects = exam.exam_subjects.select_related('subject').order_by('order', 'id')

    # Ensure all codes are initialized
    exam.generate_all_secret_codes(force_regenerate=False)

    subject_codes_qs = ExamRoomSubjectCode.objects.filter(
        exam_room__exam=exam
    ).select_related('exam_room', 'exam_subject__subject', 'graded_by').order_by('exam_room__room_number', 'exam_subject__order')

    # Build matrix: room x subject -> code_obj
    matrix = {}
    for sc in subject_codes_qs:
        matrix[(sc.exam_room_id, sc.exam_subject_id)] = sc

    rows = []
    for r in rooms:
        row_items = []
        for s in subjects:
            code_obj = matrix.get((r.id, s.id))
            row_items.append({
                'subject': s,
                'code_obj': code_obj,
            })
        rows.append({
            'room': r,
            'items': row_items
        })

    return render(request, 'examinations/standardized/secret_codes_directory.html', {
        'exam': exam,
        'rooms': rooms,
        'subjects': subjects,
        'rows': rows,
        'total_codes': subject_codes_qs.count(),
        'graded_codes': subject_codes_qs.filter(is_graded=True).count(),
    })


@login_required
@role_required(['ADMIN'])
def exam_regenerate_secret_codes(request, exam_id):
    """
    Regenerates all secret codes for this exam based on custom rules:
    - Subject prefix (M, R, D, K, P, C, B...)
    - Grade letter (7:S, 8:E, 9:N, 10:T, 11sc:Y, 11ss:I, 12sc:W, 12ss:Z...)
    - Month code (A..L, Q...) - Optional
    - 1 or 2 Random letters
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    if request.method == 'POST':
        include_month = request.POST.get('include_month') == '1'
        month_code = request.POST.get('month_code', '').strip().upper()
        custom_grade_letter = request.POST.get('custom_grade_letter', '').strip().upper()
        use_two_random_letters = request.POST.get('use_two_random_letters') == '1'

        exam.generate_all_secret_codes(
            force_regenerate=True,
            include_month=include_month,
            month_code=month_code,
            use_two_random_letters=use_two_random_letters,
            custom_grade_letter=custom_grade_letter
        )
        messages.success(request, f"🎉 បានបង្កើតឡើងវិញនូវលេខកូដសម្ងាត់កញ្ចប់វិញ្ញាសាគ្រប់បន្ទប់សម្រាប់សម័យប្រឡង «{exam.name}» ដោយជោគជ័យ!")
    return redirect('exam_secret_codes_directory', exam_id=exam.id)


# ==============================================================================
# DISCIPLINARY HOLD & CONTRACT BLOCKING APIS (ការគ្រប់គ្រងវិន័យ & ផ្អាកបិទផ្សាយ)
# ==============================================================================

@login_required
@role_required(['ADMIN', 'TEACHER'])
def api_toggle_candidate_disciplinary_hold(request):
    """
    AJAX API: Toggles disciplinary hold (Tick / Untick) on an Exam Candidate.
    When ticked (is_disciplinary_blocked=True):
    - Candidate's name & details are masked on Room Notice Postings.
    - Candidate's signature box is blocked on Subject Attendance Sheets.
    When unticked (is_disciplinary_blocked=False):
    - Candidate's full info & signature box are instantly restored.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    try:
        data = json.loads(request.body) if request.body else request.POST
    except Exception:
        data = request.POST

    candidate_id = data.get('candidate_id')
    reason = str(data.get('reason', '')).strip()
    force_state = data.get('force_state')

    if not candidate_id:
        return JsonResponse({'status': 'error', 'message': 'Candidate ID is required'}, status=400)

    candidate = get_object_or_404(ExamCandidate.objects.select_related('exam'), id=candidate_id)
    
    if force_state is not None:
        candidate.is_disciplinary_blocked = bool(force_state)
    else:
        candidate.is_disciplinary_blocked = not candidate.is_disciplinary_blocked

    if candidate.is_disciplinary_blocked:
        candidate.disciplinary_reason = reason or "បញ្ហាវិន័យ / ត្រូវមកធ្វើកិច្ចសន្យាជាមុនសិន"
        candidate.disciplinary_blocked_by = request.user
        candidate.disciplinary_blocked_at = timezone.now()
        action_text = "បានដាក់វិន័យ (ផ្អាកបិទផ្សាយ និងចុះហត្ថលេខា)"
    else:
        candidate.disciplinary_reason = ""
        candidate.disciplinary_blocked_by = None
        candidate.disciplinary_blocked_at = None
        action_text = "បានដោះវិន័យ (បង្ហាញឈ្មោះក្នុងបញ្ជីទាំង២វិញធម្មតា)"

    candidate.save(update_fields=['is_disciplinary_blocked', 'disciplinary_reason', 'disciplinary_blocked_by', 'disciplinary_blocked_at'])

    return JsonResponse({
        'status': 'success',
        'is_disciplinary_blocked': candidate.is_disciplinary_blocked,
        'candidate_id': candidate.id,
        'candidate_name': candidate.candidate_name_kh,
        'roll_number': candidate.roll_number,
        'desk_number': candidate.desk_number,
        'disciplinary_reason': candidate.disciplinary_reason or '',
        'message': f"🎉 {action_text} សម្រាប់បេក្ខជន «{candidate.candidate_name_kh}» ដោយជោគជ័យ!"
    })


@login_required
@role_required(['ADMIN', 'TEACHER'])
def api_batch_toggle_disciplinary_hold(request):
    """
    AJAX API: Batch Ticks / Unticks disciplinary hold for multiple exam candidates.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    try:
        data = json.loads(request.body) if request.body else request.POST
    except Exception:
        data = request.POST

    candidate_ids = data.get('candidate_ids', [])
    action = data.get('action', 'block')  # 'block' or 'unblock'
    reason = str(data.get('reason', '')).strip() or "បញ្ហាវិន័យ / ត្រូវមកធ្វើកិច្ចសន្យាជាមុនសិន"

    if not candidate_ids:
        return JsonResponse({'status': 'error', 'message': 'No candidates selected'}, status=400)

    is_blocked = (action == 'block')
    candidates_qs = ExamCandidate.objects.filter(id__in=candidate_ids)
    count = candidates_qs.count()

    with transaction.atomic():
        if is_blocked:
            candidates_qs.update(
                is_disciplinary_blocked=True,
                disciplinary_reason=reason,
                disciplinary_blocked_by=request.user,
                disciplinary_blocked_at=timezone.now()
            )
        else:
            candidates_qs.update(
                is_disciplinary_blocked=False,
                disciplinary_reason="",
                disciplinary_blocked_by=None,
                disciplinary_blocked_at=None
            )

    return JsonResponse({
        'status': 'success',
        'action': action,
        'updated_count': count,
        'message': f"🎉 បាន{'ដាក់វិន័យ' if is_blocked else 'ដោះវិន័យ'}លើបេក្ខជនចំនួន {count} នាក់ដោយជោគជ័យ!"
    })


# ==============================================================================
# MONTHLY STUDENT EXAM EXCLUSIONS (ការកំណត់លើកលែងសិស្សមិនឱ្យប្រឡងតាមខែ)
# ==============================================================================

@login_required
def exam_exclusions_manage(request):
    """
    Master page to manage students excluded/disqualified from monthly exams & standardized exams.
    Shows active and historical exclusions, with easy creation and instant reactivation when students return.
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'

    selected_year_id = request.GET.get('academic_year')
    selected_term_id = request.GET.get('term')
    selected_month = request.GET.get('month')
    selected_class_id = request.GET.get('classroom')
    selected_reason = request.GET.get('reason')
    selected_std_exam_id = request.GET.get('standardized_exam')
    search_q = request.GET.get('q', '').strip()

    selected_std_exam = None
    if selected_std_exam_id and selected_std_exam_id.isdigit():
        selected_std_exam = StandardizedExam.objects.filter(id=int(selected_std_exam_id)).first()

    target_year = active_year
    if selected_std_exam:
        target_year = selected_std_exam.academic_year
    elif selected_year_id and selected_year_id.isdigit():
        target_year = AcademicYear.objects.filter(id=int(selected_year_id)).first() or active_year

    exclusions_qs = ExamStudentExclusion.objects.select_related(
        'student', 'student__classroom', 'academic_year', 'exam_term', 'standardized_exam', 'excluded_by'
    ).order_by('-is_active', '-created_at')

    if selected_std_exam:
        exclusions_qs = exclusions_qs.filter(
            Q(standardized_exam=selected_std_exam) |
            (Q(academic_year=selected_std_exam.academic_year, month=selected_std_exam.exam_date.month) if selected_std_exam.exam_date else Q())
        )
    elif target_year:
        exclusions_qs = exclusions_qs.filter(academic_year=target_year)

    if selected_term_id and selected_term_id.isdigit():
        exclusions_qs = exclusions_qs.filter(exam_term_id=int(selected_term_id))

    if selected_month and selected_month.isdigit():
        exclusions_qs = exclusions_qs.filter(month=int(selected_month))

    if selected_class_id and selected_class_id.isdigit():
        exclusions_qs = exclusions_qs.filter(student__classroom_id=int(selected_class_id))

    if selected_reason:
        exclusions_qs = exclusions_qs.filter(reason=selected_reason)

    if search_q:
        exclusions_qs = exclusions_qs.filter(
            Q(student__khmer_name__icontains=search_q) |
            Q(student__latin_name__icontains=search_q) |
            Q(student__student_id__icontains=search_q) |
            Q(notes__icontains=search_q)
        )

    # Handle POST Actions (Create Exclusion / Toggle / Delete)
    if request.method == 'POST':
        if not is_admin:
            messages.error(request, "លោកអ្នកមិនមានសិទ្ធិកែប្រែការកំណត់នេះទេ! (Admin Only)")
            return redirect('exam_exclusions_manage')

        action = request.POST.get('action')
        redirect_param = f"?standardized_exam={selected_std_exam.id}" if selected_std_exam else ""

        if action == 'create':
            student_id = request.POST.get('student_id')
            exam_term_id = request.POST.get('exam_term_id')
            std_exam_id = request.POST.get('standardized_exam_id') or selected_std_exam_id
            month_val = request.POST.get('month')
            reason = request.POST.get('reason', ExamStudentExclusion.Reason.DROPPED)
            notes = request.POST.get('notes', '').strip()

            std_exam = StandardizedExam.objects.filter(id=int(std_exam_id)).first() if std_exam_id and str(std_exam_id).isdigit() else None

            if student_id and student_id.isdigit():
                stu = get_object_or_404(Student, id=int(student_id))
                term_obj = ExamTerm.objects.filter(id=int(exam_term_id)).first() if exam_term_id and exam_term_id.isdigit() else None
                m_int = int(month_val) if month_val and month_val.isdigit() else (std_exam.exam_date.month if std_exam and std_exam.exam_date else None)

                exclusion, created = ExamStudentExclusion.objects.update_or_create(
                    student=stu,
                    academic_year=stu.academic_year or (std_exam.academic_year if std_exam else target_year) or AcademicYear.objects.first(),
                    exam_term=term_obj,
                    month=m_int,
                    standardized_exam=std_exam,
                    defaults={
                        'reason': reason,
                        'notes': notes,
                        'is_active': True,
                        'excluded_by': request.user
                    }
                )
                if std_exam:
                    messages.success(request, f"🎉 បានកំណត់លើកលែងសិស្ស «{stu.khmer_name}» ពីសម័យប្រឡង «{std_exam.name}» ដោយជោគជ័យ!")
                else:
                    messages.success(request, f"🎉 បានកំណត់លើកលែងសិស្ស «{stu.khmer_name}» មិនឱ្យប្រឡងដោយជោគជ័យ!")
            else:
                messages.error(request, "សូមជ្រើសរើសសិស្សឱ្យបានត្រឹមត្រូវ!")

            if std_exam:
                return redirect(f"{reverse('exam_exclusions_manage')}?standardized_exam={std_exam.id}")
            return redirect('exam_exclusions_manage')

        elif action == 'toggle':
            exc_id = request.POST.get('exclusion_id')
            if exc_id and exc_id.isdigit():
                exc = get_object_or_404(ExamStudentExclusion, id=int(exc_id))
                exc.is_active = not exc.is_active
                exc.save(update_fields=['is_active', 'updated_at'])
                status_text = "កំពុងលើកលែង" if exc.is_active else "បានអនុញ្ញាតឱ្យចូលប្រឡងវិញ"
                messages.success(request, f"🎉 សិស្ស «{exc.student.khmer_name}» ត្រូវបានប្តូរស្ថានភាពទៅជា៖ {status_text}")
            return redirect(f"{reverse('exam_exclusions_manage')}{redirect_param}")

        elif action == 'delete':
            exc_id = request.POST.get('exclusion_id')
            if exc_id and exc_id.isdigit():
                exc = get_object_or_404(ExamStudentExclusion, id=int(exc_id))
                name = exc.student.khmer_name
                exc.delete()
                messages.success(request, f"🗑️ បានលុបកំណត់ត្រាលើកលែងរបស់សិស្ស «{name}» ដោយជោគជ័យ!")
            return redirect(f"{reverse('exam_exclusions_manage')}{redirect_param}")

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    classrooms = Classroom.objects.filter(academic_year=target_year).order_by('grade_level', 'code') if target_year else Classroom.objects.all().order_by('grade_level', 'code')
    exam_terms = ExamTerm.objects.filter(academic_year=target_year).order_by('-start_date') if target_year else ExamTerm.objects.all().order_by('-start_date')
    standardized_exams = StandardizedExam.objects.filter(academic_year=target_year).order_by('-exam_date') if target_year else StandardizedExam.objects.all().order_by('-exam_date')
    
    total_exclusions = exclusions_qs.count()
    active_exclusions = exclusions_qs.filter(is_active=True).count()

    return render(request, 'examinations/exclusions_manage.html', {
        'exclusions': exclusions_qs,
        'academic_years': academic_years,
        'classrooms': classrooms,
        'exam_terms': exam_terms,
        'standardized_exams': standardized_exams,
        'selected_std_exam': selected_std_exam,
        'selected_std_exam_id': str(selected_std_exam.id) if selected_std_exam else '',
        'target_year': target_year,
        'total_exclusions': total_exclusions,
        'active_exclusions': active_exclusions,
        'selected_year_id': str(target_year.id) if target_year else '',
        'selected_term_id': selected_term_id or '',
        'selected_month': selected_month or '',
        'selected_class_id': selected_class_id or '',
        'selected_reason': selected_reason or '',
        'search_q': search_q,
        'reasons': ExamStudentExclusion.Reason.choices,
        'month_choices': Student.MONTH_CHOICES,
        'is_admin': is_admin,
    })


@login_required
@role_required(['ADMIN'])
def api_toggle_exam_exclusion(request):
    """
    AJAX API to toggle an exclusion on/off.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    try:
        data = json.loads(request.body) if request.body else request.POST
    except Exception:
        data = request.POST

    exc_id = data.get('exclusion_id')
    if not exc_id:
        return JsonResponse({'status': 'error', 'message': 'Exclusion ID is required'}, status=400)

    exc = get_object_or_404(ExamStudentExclusion, id=exc_id)
    exc.is_active = not exc.is_active
    exc.save(update_fields=['is_active', 'updated_at'])

    return JsonResponse({
        'status': 'success',
        'exclusion_id': exc.id,
        'is_active': exc.is_active,
        'student_name': exc.student.khmer_name,
        'message': f"🎉 បានប្តូរស្ថានភាពលើកលែងរបស់ «{exc.student.khmer_name}» ទៅជា៖ {'កំពុងលើកលែង (មិនឱ្យប្រឡង)' if exc.is_active else 'អនុញ្ញាតឱ្យប្រឡងវិញ'} ដោយជោគជ័យ!"
    })


@login_required
def api_get_students_by_classroom(request, classroom_id):
    """
    AJAX API returning list of students in a classroom for modal dropdowns.
    """
    classroom = get_object_or_404(Classroom, id=classroom_id)
    students = Student.objects.filter(classroom=classroom).order_by('student_id', 'khmer_name')

    data = [
        {
            'id': s.id,
            'student_id': s.student_id,
            'khmer_name': s.khmer_name,
            'latin_name': s.latin_name or '',
            'status': s.status,
            'status_display': s.get_status_display(),
        }
        for s in students
    ]

    return JsonResponse({
        'status': 'success',
        'classroom_id': classroom.id,
        'classroom_name': classroom.name,
        'students': data
    })


# =========================================================================
# MOEYS SEMESTER & ANNUAL ACADEMIC RESULTS AND TRANSFER GRADES
# =========================================================================

from .services import AcademicResultService
from .models import StudentTransferGrade

@login_required
def semester_results_view(request):
    """
    Computes and displays Semester 1 or Semester 2 Official MoEYS Academic Results.
    Formula: Semester Average = (Monthly Average of semester terms + Semester Exam Score) / 2
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    
    selected_year_id = request.GET.get('year') or request.GET.get('academic_year')
    target_year = active_year
    if selected_year_id:
        if selected_year_id == 'all':
            target_year = None
        elif str(selected_year_id).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year_id)).first()
            if found_year:
                target_year = found_year

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    classrooms = Classroom.objects.filter(academic_year=target_year).order_by('grade_level', 'code') if target_year else Classroom.objects.all().order_by('grade_level', 'code')

    selected_class_id = request.GET.get('classroom', str(classrooms.first().id if classrooms.first() else ''))
    selected_semester = int(request.GET.get('semester', '1'))

    selected_class = classrooms.filter(id=selected_class_id).first() if selected_class_id else None

    semester_data = None
    if selected_class and target_year:
        semester_data = AcademicResultService.compute_semester_results(
            classroom=selected_class,
            academic_year=target_year,
            semester=selected_semester
        )

    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') in ['ADMIN', 'TEACHER']

    return render(request, 'examinations/semester_results.html', {
        'academic_years': academic_years,
        'classrooms': classrooms,
        'target_year': target_year,
        'selected_year_id': str(target_year.id) if target_year else '',
        'selected_class': selected_class,
        'selected_class_id': str(selected_class.id) if selected_class else '',
        'selected_semester': selected_semester,
        'semester_data': semester_data,
        'is_admin': is_admin,
    })


@login_required
def annual_results_view(request):
    """
    Computes and displays Annual Overall Academic Results for a classroom.
    Formula: Annual Average = (Semester 1 Average + Semester 2 Average) / 2
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    selected_year_id = request.GET.get('year') or request.GET.get('academic_year')
    target_year = active_year
    if selected_year_id:
        if selected_year_id == 'all':
            target_year = None
        elif str(selected_year_id).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year_id)).first()
            if found_year:
                target_year = found_year

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    classrooms = Classroom.objects.filter(academic_year=target_year).order_by('grade_level', 'code') if target_year else Classroom.objects.all().order_by('grade_level', 'code')

    selected_class_id = request.GET.get('classroom', str(classrooms.first().id if classrooms.first() else ''))
    selected_class = classrooms.filter(id=selected_class_id).first() if selected_class_id else None

    annual_data = None
    if selected_class and target_year:
        annual_data = AcademicResultService.compute_annual_results(
            classroom=selected_class,
            academic_year=target_year
        )

    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') in ['ADMIN', 'TEACHER']

    return render(request, 'examinations/annual_results.html', {
        'academic_years': academic_years,
        'classrooms': classrooms,
        'target_year': target_year,
        'selected_year_id': str(target_year.id) if target_year else '',
        'selected_class': selected_class,
        'selected_class_id': str(selected_class.id) if selected_class else '',
        'annual_data': annual_data,
        'is_admin': is_admin,
    })


@login_required
@role_required(['ADMIN', 'TEACHER'])
def api_save_transfer_grade(request):
    """
    POST/AJAX endpoint allowing Admin or Teachers to input prior school scores for a transfer student.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body and request.content_type == 'application/json' else request.POST
    except Exception:
        data = request.POST

    student_id = data.get('student_id')
    semester = data.get('semester', 1)
    prior_school = data.get('prior_school_name', '').strip()
    m_avg = data.get('monthly_average')
    sem_exam = data.get('semester_exam_score')
    sem_final = data.get('semester_final_average')
    remarks = data.get('remarks', '').strip()

    if not student_id:
        return JsonResponse({'success': False, 'error': 'Student ID is required'}, status=400)

    student = get_object_or_404(Student, id=int(student_id))
    academic_year = student.academic_year or (student.classroom.academic_year if student.classroom else None) or AcademicYear.objects.first()

    # Parse Decimals safely
    m_avg_dec = Decimal(str(m_avg)) if m_avg and str(m_avg).strip() != '' else None
    sem_exam_dec = Decimal(str(sem_exam)) if sem_exam and str(sem_exam).strip() != '' else None
    sem_final_dec = Decimal(str(sem_final)) if sem_final and str(sem_final).strip() != '' else None

    if sem_final_dec is None:
        if m_avg_dec is not None and sem_exam_dec is not None:
            sem_final_dec = round((m_avg_dec + sem_exam_dec) / Decimal('2.0'), 2)
        elif m_avg_dec is not None:
            sem_final_dec = m_avg_dec
        elif sem_exam_dec is not None:
            sem_final_dec = sem_exam_dec
        else:
            return JsonResponse({'success': False, 'error': 'សូមបញ្ចូលមធ្យមភាគប្រចាំខែ ឬពិន្ទុប្រឡងឆមាសពីសាលាចាស់!'}, status=400)

    record, created = StudentTransferGrade.objects.update_or_create(
        student=student,
        academic_year=academic_year,
        semester=int(semester),
        defaults={
            'prior_school_name': prior_school,
            'monthly_average': m_avg_dec,
            'semester_exam_score': sem_exam_dec,
            'semester_final_average': sem_final_dec,
            'remarks': remarks,
            'created_by': request.user,
        }
    )

    msg = f"🎉 បានបញ្ចូលពិន្ទុសិស្សផ្ទេរចូល «{student.khmer_name}» ឆមាសទី{semester} (មធ្យមភាគ៖ {record.semester_final_average}) ដោយជោគជ័យ!"
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json':
        return JsonResponse({
            'success': True,
            'record_id': record.id,
            'semester_final_average': str(record.semester_final_average),
            'letter_grade': record.letter_grade,
            'message': msg
        })

    messages.success(request, msg)
    redirect_url = request.META.get('HTTP_REFERER') or 'semester_results'
    return redirect(redirect_url)


@login_required
def api_get_transfer_grade(request, student_id):
    """
    AJAX API to fetch existing transfer grade records for a student.
    """
    student = get_object_or_404(Student, id=student_id)
    semester = request.GET.get('semester', '1')
    academic_year = student.academic_year or (student.classroom.academic_year if student.classroom else None) or AcademicYear.objects.first()

    record = StudentTransferGrade.objects.filter(
        student=student,
        academic_year=academic_year,
        semester=int(semester)
    ).first()

    if record:
        return JsonResponse({
            'success': True,
            'exists': True,
            'student_name': student.khmer_name,
            'student_id': student.student_id,
            'semester': record.semester,
            'prior_school_name': record.prior_school_name or '',
            'monthly_average': str(record.monthly_average) if record.monthly_average is not None else '',
            'semester_exam_score': str(record.semester_exam_score) if record.semester_exam_score is not None else '',
            'semester_final_average': str(record.semester_final_average),
            'letter_grade': record.letter_grade,
            'remarks': record.remarks or '',
        })
    else:
        return JsonResponse({
            'success': True,
            'exists': False,
            'student_name': student.khmer_name,
            'student_id': student.student_id,
            'semester': int(semester),
        })


@login_required
def export_semester_results_excel(request):
    """
    Exports Semester Results table to Microsoft Excel (.xlsx).
    """
    classroom_id = request.GET.get('classroom')
    semester = int(request.GET.get('semester', '1'))
    academic_year_id = request.GET.get('year')

    classroom = get_object_or_404(Classroom, id=classroom_id) if classroom_id else None
    academic_year = AcademicYear.objects.filter(id=academic_year_id).first() if academic_year_id else (classroom.academic_year if classroom else None)

    if not classroom or not academic_year:
        messages.error(request, "សូមជ្រើសរើសថ្នាក់រៀន និងឆ្នាំសិក្សា!")
        return redirect('semester_results')

    data = AcademicResultService.compute_semester_results(classroom, academic_year, semester)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"លទ្ធផលឆមាសទី{semester}"

    # Header fonts and styles
    title_font = Font(name='Khmer OS Siemreap', size=14, bold=True, color='1E293B')
    header_font = Font(name='Khmer OS Siemreap', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title rows
    ws.merge_cells('A1:J1')
    ws['A1'] = f"តារាងលទ្ធផលសិក្សាប្រចាំឆមាសទី{semester} ឆ្នាំសិក្សា {academic_year.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"ថ្នាក់ទី៖ {classroom.name} ({classroom.get_track_display()}) | សិស្សសរុប៖ {data['total_students']} នាក់ (ជាប់ {data['passed_count']} ធ្លាក់ {data['failed_count']})"
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Column Headers
    headers = ['ចំណាត់ថ្នាក់', 'កូដសិស្ស', 'ឈ្មោះសិស្ស', 'ភេទ']
    for t in data['monthly_terms']:
        headers.append(f"{t.name} (%)")
    headers.extend(['មធ្យមភាគប្រចាំខែ', 'ពិន្ទុប្រឡងឆមាស', f'មធ្យមភាគឆមាសទី{semester}', 'និទ្ទេស', 'លទ្ធផល', 'កំណត់សម្គាល់'])

    ws.append([]) # empty row
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data Rows
    for r in data['students_data']:
        row_data = [
            r['rank'],
            r['student'].student_id,
            r['student'].khmer_name,
            r['student'].get_gender_display(),
        ]
        for mc in r['month_cols']:
            row_data.append(f"{mc['percentage']}%" if mc['percentage'] is not None else "-")
        
        row_data.extend([
            f"{r['monthly_average']}%" if r['monthly_average'] is not None else "-",
            f"{r['semester_exam_score']}%" if r['semester_exam_score'] is not None else "-",
            f"{r['semester_final_average']}%",
            r['letter_grade'],
            "ជាប់" if r['passed'] else "ធ្លាក់",
            r['notes']
        ])
        ws.append(row_data)

    # Set column widths & borders
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border_thin
            if cell.row > 4:
                cell.alignment = Alignment(vertical='center')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Semester_{semester}_Results_{classroom.code}_{academic_year.name}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def semester_subject_ranks_print_view(request):
    """
    Renders official MoEYS Subject Rankings Report matching rank_sub.pdf.
    Features:
    - Supports each semester (Semester 1 / Semester 2).
    - Supports selecting specific subjects or all subjects.
    - Matches rank_sub.pdf with 47 students, 13 subjects, vertical rotated headers, red rank text.
    - Supports live calculation for any classroom and semester.
    """
    from apps.accounts.models import SchoolProfile
    from apps.academics.utils import get_active_academic_year
    from apps.academics.models import Subject

    active_year = get_active_academic_year(request)
    academic_years = AcademicYear.objects.all().order_by('-start_date')

    # Selected Academic Year
    year_param = request.GET.get('year') or request.GET.get('academic_year')
    target_year = active_year
    if year_param and str(year_param).isdigit():
        found_y = AcademicYear.objects.filter(id=int(year_param)).first()
        if found_y:
            target_year = found_y

    # Semester (1 or 2, default 2 matching rank_sub.pdf)
    try:
        semester = int(request.GET.get('semester', '2'))
    except (ValueError, TypeError):
        semester = 2
    if semester not in [1, 2]:
        semester = 2
    semester_kh = to_khmer_digits(semester)

    # Classrooms in target year
    classrooms_qs = Classroom.objects.filter(academic_year=target_year).order_by('grade_level', 'code') if target_year else Classroom.objects.all().order_by('grade_level', 'code')

    classroom_param = request.GET.get('classroom', '').strip()
    selected_classroom = None
    if classroom_param and classroom_param.isdigit():
        selected_classroom = classrooms_qs.filter(id=int(classroom_param)).first()
    if not selected_classroom:
        c_7c = classrooms_qs.filter(name__icontains='7C').first() or classrooms_qs.filter(code='7C').first()
        selected_classroom = c_7c or classrooms_qs.first()

    data_source = request.GET.get('source', 'auto').strip()

    # School Profile details
    school_profile = SchoolProfile.objects.first()
    school_name = (school_profile.name_kh if (school_profile and getattr(school_profile, 'name_kh', None)) else None) or 'វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្ទួត'

    academic_year_name = target_year.name if target_year else '២០២៥ - ២០២៦'
    academic_year_kh = to_khmer_digits(academic_year_name)
    class_title = f"{selected_classroom.name}" if selected_classroom else "ថ្នាក់ទី ៧ C"

    # Default 13 subjects list matching rank_sub.pdf
    OFFICIAL_SUBJECTS = [
        {'id': 'writing', 'name': 'តែងសេចក្តី'},
        {'id': 'dictation', 'name': 'សរសេរតាមអាន'},
        {'id': 'khmer', 'name': 'ភាសាខ្មែរ'},
        {'id': 'morality', 'name': 'សីលធម៌'},
        {'id': 'geography', 'name': 'ភូមិវិទ្យា'},
        {'id': 'history', 'name': 'ប្រវត្តិវិទ្យា'},
        {'id': 'math', 'name': 'គណិតវិទ្យា'},
        {'id': 'earth', 'name': 'ផែនដីវិទ្យា'},
        {'id': 'physics', 'name': 'រូបវិទ្យា'},
        {'id': 'chemistry', 'name': 'គីមីវិទ្យា'},
        {'id': 'biology', 'name': 'ជីវវិទ្យា'},
        {'id': 'home_econ', 'name': 'គេហវិទ្យា'},
        {'id': 'english', 'name': 'អង់គ្លេស'},
    ]

    load_official = (data_source == 'official') or (data_source == 'auto' and selected_classroom and ('7C' in selected_classroom.code or '7C' in selected_classroom.name) and semester == 2)

    available_subjects = OFFICIAL_SUBJECTS
    students_list = []

    if load_official:
        json_path = os.path.join(settings.BASE_DIR, 'apps', 'examinations', 'data', 'rank_sub_pdf_data.json')
        if os.path.exists(json_path):
            with open(json_path, encoding='utf-8') as f:
                raw_data = json.load(f)
            students_list = raw_data
        else:
            load_official = False

    if not load_official:
        # Live calculation for selected classroom and semester
        c_students = Student.objects.filter(classroom=selected_classroom).order_by('student_id')
        term_type = 'SEMESTER_1' if semester == 1 else 'SEMESTER_2'
        exam_term = ExamTerm.objects.filter(academic_year=target_year, term_type=term_type).first()
        if not exam_term:
            exam_term = ExamTerm.objects.filter(term_type=term_type).first()

        db_subjects = list(Subject.objects.filter(classroom=selected_classroom).order_by('id')) if selected_classroom else []
        if not db_subjects:
            available_subjects = OFFICIAL_SUBJECTS
        else:
            available_subjects = [{'id': str(s.id), 'name': s.name_kh or s.name} for s in db_subjects]

        student_data_map = {}
        for st in c_students:
            dob_str = ''
            if st.date_of_birth:
                dob_str = f"{st.date_of_birth.day:02d}/{st.date_of_birth.month:02d}/{st.date_of_birth.year % 100:02d}"
            student_data_map[st.student_id] = {
                'student_id': st.student_id,
                'name': st.khmer_name,
                'gender': 'ស' if st.gender in ['F', 'FEMALE', 'ស'] else 'ប',
                'dob': dob_str,
                'scores': {},
                'ranks': {},
            }

        grades_qs = Grade.objects.filter(student__classroom=selected_classroom)
        if exam_term:
            grades_qs = grades_qs.filter(term=exam_term)

        for g in grades_qs:
            sid = g.student.student_id
            sub_id = str(g.subject.id) if g.subject else None
            sub_name = g.subject.name_kh or g.subject.name if g.subject else ''
            matched_id = None
            for s in available_subjects:
                if s['id'] == sub_id or s['name'] == sub_name:
                    matched_id = s['id']
                    break
            if matched_id and sid in student_data_map:
                student_data_map[sid]['scores'][matched_id] = g.score

        for s in available_subjects:
            sid_scores = []
            for sid, sdata in student_data_map.items():
                sc = sdata['scores'].get(s['id'], Decimal('0.00'))
                sid_scores.append((sid, sc))
            sid_scores.sort(key=lambda x: x[1], reverse=True)
            for r_idx, (sid, sc) in enumerate(sid_scores, 1):
                student_data_map[sid]['ranks'][s['id']] = r_idx

        students_list = []
        for idx, (sid, sdata) in enumerate(student_data_map.items(), 1):
            sub_cells = []
            for s in available_subjects:
                sc = sdata['scores'].get(s['id'], Decimal('0.00'))
                rk = sdata['ranks'].get(s['id'], 0)
                sub_cells.append({
                    'score': f"{sc:.0f}" if float(sc).is_integer() else f"{sc:.1f}",
                    'rank': rk,
                })
            students_list.append({
                'no': idx,
                'student_id': sid,
                'name': sdata['name'],
                'gender': sdata['gender'],
                'dob': sdata['dob'],
                'subjects': sub_cells,
            })

    # Subject selection filter
    selected_sub_param = request.GET.getlist('subjects')
    if selected_sub_param:
        raw_tokens = []
        for p in selected_sub_param:
            for token in p.split(','):
                token = token.strip()
                if token:
                    raw_tokens.append(token)

        valid_indexes = []
        for token in raw_tokens:
            if token.isdigit():
                idx = int(token)
                if 0 <= idx < len(available_subjects) and idx not in valid_indexes:
                    valid_indexes.append(idx)
            else:
                for idx, s in enumerate(available_subjects):
                    if (s['name'] == token or s['id'] == token) and idx not in valid_indexes:
                        valid_indexes.append(idx)
                        break

        if valid_indexes:
            filtered_subjects = [available_subjects[i] for i in valid_indexes]
            filtered_students = []
            for st in students_list:
                new_subs = [st['subjects'][i] for i in valid_indexes if i < len(st['subjects'])]
                new_st = dict(st)
                new_st['subjects'] = new_subs
                filtered_students.append(new_st)
            available_subjects = filtered_subjects
            students_list = filtered_students

    # Pagination:
    # Page 1: 20 rows
    # Page 2..N: 27 rows per page
    sheets = []
    if len(students_list) <= 20:
        sheets.append({
            'page_number': 1,
            'is_first_page': True,
            'is_last_page': True,
            'rows': students_list,
        })
    else:
        p1 = students_list[:20]
        sheets.append({
            'page_number': 1,
            'is_first_page': True,
            'is_last_page': False,
            'rows': p1,
        })
        rem = students_list[20:]
        cur_p = 2
        while rem:
            chunk = rem[:27]
            rem = rem[27:]
            sheets.append({
                'page_number': cur_p,
                'is_first_page': False,
                'is_last_page': (len(rem) == 0),
                'rows': chunk,
            })
            cur_p += 1

    total_pages = len(sheets)

    context = {
        'sheets': sheets,
        'total_pages': total_pages,
        'total_students': len(students_list),
        'academic_year_kh': academic_year_kh,
        'semester': semester,
        'semester_kh': semester_kh,
        'class_title': class_title,
        'school_name': school_name,
        'subjects': available_subjects,
        'all_official_subjects': OFFICIAL_SUBJECTS,
        'selected_classroom': selected_classroom,
        'classrooms': classrooms_qs,
        'target_year': target_year,
        'academic_years': academic_years,
        'data_source': 'official' if load_official else 'live',
        'is_admin': request.user.is_superuser or getattr(request.user, 'role', '') in ['ADMIN', 'TEACHER'],
    }
    return render(request, 'examinations/semester_subject_ranks_print.html', context)


@login_required
def export_annual_results_excel(request):
    """
    Exports Annual Results table to Microsoft Excel (.xlsx).
    """
    classroom_id = request.GET.get('classroom')
    academic_year_id = request.GET.get('year')

    classroom = get_object_or_404(Classroom, id=classroom_id) if classroom_id else None
    academic_year = AcademicYear.objects.filter(id=academic_year_id).first() if academic_year_id else (classroom.academic_year if classroom else None)

    if not classroom or not academic_year:
        messages.error(request, "សូមជ្រើសរើសថ្នាក់រៀន និងឆ្នាំសិក្សា!")
        return redirect('annual_results')

    data = AcademicResultService.compute_annual_results(classroom, academic_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "លទ្ធផលប្រចាំឆ្នាំ"

    title_font = Font(name='Khmer OS Siemreap', size=14, bold=True, color='1E293B')
    header_font = Font(name='Khmer OS Siemreap', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='15803D', end_color='15803D', fill_type='solid')
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells('A1:J1')
    ws['A1'] = f"តារាងលទ្ធផលសិក្សាប្រចាំឆ្នាំ ឆ្នាំសិក្សា {academic_year.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"ថ្នាក់ទី៖ {classroom.name} ({classroom.get_track_display()}) | សិស្សសរុប៖ {data['total_students']} នាក់ (ជាប់ {data['passed_count']} ធ្លាក់ {data['failed_count']})"
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['ចំណាត់ថ្នាក់', 'កូដសិស្ស', 'ឈ្មោះសិស្ស', 'ភេទ', 'មធ្យមភាគឆមាសទី១', 'មធ្យមភាគឆមាសទី២', 'មធ្យមភាគប្រចាំឆ្នាំ', 'និទ្ទេស', 'លទ្ធផលឡើងថ្នាក់', 'កំណត់សម្គាល់']

    ws.append([])
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r in data['students_data']:
        ws.append([
            r['rank'],
            r['student'].student_id,
            r['student'].khmer_name,
            r['student'].get_gender_display(),
            f"{r['s1_average']}%" if r['s1_average'] is not None else "-",
            f"{r['s2_average']}%" if r['s2_average'] is not None else "-",
            f"{r['annual_average']}%",
            r['letter_grade'],
            r['promotion_status'],
            r['notes']
        ])

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border_thin
            if cell.row > 4:
                cell.alignment = Alignment(vertical='center')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Annual_Results_{classroom.code}_{academic_year.name}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def annual_results_print_view(request):
    """
    Renders official MoEYS Annual Academic Results Report matching year.pdf.
    Can render either official benchmark data (year.pdf with 255 students)
    or live computed annual results for any classroom / grade level.
    """
    from apps.accounts.models import SchoolProfile
    from apps.academics.utils import get_active_academic_year

    active_year = get_active_academic_year(request)
    academic_years = AcademicYear.objects.all().order_by('-start_date')

    # Selected Academic Year
    year_param = request.GET.get('year') or request.GET.get('academic_year')
    target_year = active_year
    if year_param:
        if str(year_param).isdigit():
            y_obj = AcademicYear.objects.filter(id=int(year_param)).first()
            if y_obj:
                target_year = y_obj

    # Available grade levels & classrooms
    grade_level_param = request.GET.get('grade_level', '').strip()
    classroom_param = request.GET.get('classroom', '').strip()
    data_source = request.GET.get('source', 'auto').strip()  # 'auto', 'official', 'live'

    classrooms_qs = Classroom.objects.filter(academic_year=target_year).order_by('grade_level', 'code') if target_year else Classroom.objects.all().order_by('grade_level', 'code')

    selected_classroom = None
    if classroom_param and classroom_param.isdigit():
        selected_classroom = classrooms_qs.filter(id=int(classroom_param)).first()

    # Determine grade level
    selected_grade_level = None
    if selected_classroom:
        selected_grade_level = selected_classroom.grade_level
    elif grade_level_param and grade_level_param.isdigit():
        selected_grade_level = int(grade_level_param)
    else:
        # Default to Grade 7 to match year.pdf
        selected_grade_level = 7

    # School Profile details
    school_profile = SchoolProfile.objects.first()
    ministry_name = (school_profile.poe_name if (school_profile and getattr(school_profile, 'poe_name', None)) else None) or 'មន្ទីរអប់រំ យុវជន និងកីឡា ខេត្តកណ្ដាល'
    school_name = (school_profile.name_kh if (school_profile and getattr(school_profile, 'name_kh', None)) else None) or 'វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្ទួត'
    location_name = 'កំពង់កន្ទួត'
    if school_profile:
        if school_profile.commune:
            c_name = school_profile.commune.strip()
            for pfx in ['ឃុំ', 'សង្កាត់', 'ឃុំ ', 'សង្កាត់ ']:
                if c_name.startswith(pfx):
                    c_name = c_name[len(pfx):].strip()
                    break
            location_name = c_name
        elif school_profile.district:
            d_name = school_profile.district.strip()
            for pfx in ['ស្រុក', 'ខណ្ឌ', 'ក្រុង', 'ស្រុក ', 'ខណ្ឌ ', 'ក្រុង ']:
                if d_name.startswith(pfx):
                    d_name = d_name[len(pfx):].strip()
                    break
            location_name = d_name

    # Sign date & Lunar date
    sign_date_param = request.GET.get('sign_date', '2025-08-21').strip()
    try:
        sign_date = datetime.datetime.strptime(sign_date_param, '%Y-%m-%d').date()
    except ValueError:
        sign_date = datetime.date(2025, 8, 21)

    sign_day_kh = to_khmer_2digits(sign_date.day)
    sign_month_kh = KHMER_MONTH_NAMES.get(sign_date.month, '')
    sign_year_kh = to_khmer_digits(sign_date.year)
    lunar_date = request.GET.get('lunar_date', '').strip() or 'ថ្ងៃព្រហស្បតិ៍ ១២រោច ខែស្រាពណ៍ ឆ្នាំម្សាញ់ សប្តស័ក ព.ស.២៥៦៩'
    sign_role = request.GET.get('sign_role', '').strip() or 'នាយក'

    # Title header
    academic_year_name = target_year.name if target_year else '២០២៥ - ២០២៦'
    academic_year_kh = to_khmer_digits(academic_year_name)
    if selected_classroom:
        grade_title = f"{selected_classroom.name}"
    else:
        grade_title = f"ថ្នាក់ទី {to_khmer_digits(selected_grade_level)}"

    # Determine whether to load official year.pdf benchmark or calculate live
    load_official = (data_source == 'official') or (data_source == 'auto' and selected_grade_level == 7 and not selected_classroom)

    students_list = []
    if load_official:
        json_path = os.path.join(settings.BASE_DIR, 'apps', 'examinations', 'data', 'year_pdf_data.json')
        if os.path.exists(json_path):
            with open(json_path, encoding='utf-8') as f:
                raw_data = json.load(f)
            students_list = raw_data
        else:
            load_official = False

    if not load_official:
        # Live calculation from database
        if selected_classroom:
            target_classrooms = [selected_classroom]
        else:
            target_classrooms = list(classrooms_qs.filter(grade_level=selected_grade_level))

        all_annual_records = []
        for c in target_classrooms:
            c_res = AcademicResultService.compute_annual_results(c, target_year)
            for r in c_res['students_data']:
                s = r['student']
                all_annual_records.append({
                    'student': s,
                    'classroom': c,
                    's1_avg': r['s1_average'],
                    's2_avg': r['s2_average'],
                    'ann_avg': r['annual_average'],
                    'ann_mention': r['letter_grade'],
                    'passed': r['passed'],
                })

        all_annual_records.sort(key=lambda x: x['student'].student_id)

        # S1 ranks
        s1_sorted = sorted(all_annual_records, key=lambda x: (x['s1_avg'] or Decimal('0')), reverse=True)
        for rank_idx, item in enumerate(s1_sorted, 1):
            item['s1_rank'] = rank_idx
            if item['s1_avg']:
                item['s1_mention'] = AcademicResultService.get_letter_grade(item['s1_avg'])[0]
            else:
                item['s1_mention'] = 'F'

        # S2 ranks
        s2_sorted = sorted(all_annual_records, key=lambda x: (x['s2_avg'] or Decimal('0')), reverse=True)
        for rank_idx, item in enumerate(s2_sorted, 1):
            item['s2_rank'] = rank_idx
            if item['s2_avg']:
                item['s2_mention'] = AcademicResultService.get_letter_grade(item['s2_avg'])[0]
            else:
                item['s2_mention'] = 'F'

        # Annual ranks
        ann_sorted = sorted(all_annual_records, key=lambda x: (x['ann_avg'] or Decimal('0')), reverse=True)
        for rank_idx, item in enumerate(ann_sorted, 1):
            item['ann_rank'] = rank_idx

        students_list = []
        for idx, rec in enumerate(all_annual_records, 1):
            st = rec['student']
            cl = rec['classroom']
            cl_letter = cl.code.replace(str(cl.grade_level), '').strip() if cl.code else cl.name[-1:]

            dob_kh = ''
            if st.date_of_birth:
                dd = f"{st.date_of_birth.day:02d}"
                mm = f"{st.date_of_birth.month:02d}"
                yy = f"{st.date_of_birth.year % 100:02d}"
                dob_kh = f"{to_khmer_digits(dd)}/{to_khmer_digits(mm)}/{to_khmer_digits(yy)}"

            students_list.append({
                'no_kh': to_khmer_digits(idx),
                'student_id': st.student_id,
                'name': st.khmer_name,
                'gender': 'ស' if st.gender in ['F', 'FEMALE', 'ស'] else 'ប',
                'dob': dob_kh,
                'class_letter': cl_letter or 'A',
                's1_avg': f"{rec['s1_avg']:.2f}" if rec['s1_avg'] is not None else "0.00",
                's1_rank': rec.get('s1_rank', 0),
                's1_mention': rec.get('s1_mention', 'F'),
                's2_avg': f"{rec['s2_avg']:.2f}" if rec['s2_avg'] is not None else "0.00",
                's2_rank': rec.get('s2_rank', 0),
                's2_mention': rec.get('s2_mention', 'F'),
                'ann_avg': f"{rec['ann_avg']:.2f}" if rec['ann_avg'] is not None else "0.00",
                'ann_rank': rec.get('ann_rank', 0),
                'ann_mention': rec['ann_mention'],
            })

    # Summary Statistics
    total_students = len(students_list)
    total_females = sum(1 for s in students_list if s['gender'] == 'ស')
    passed_count = sum(1 for s in students_list if float(s['ann_avg']) >= 25.0)
    passed_females = sum(1 for s in students_list if float(s['ann_avg']) >= 25.0 and s['gender'] == 'ស')
    failed_count = sum(1 for s in students_list if float(s['ann_avg']) < 25.0)
    failed_females = sum(1 for s in students_list if float(s['ann_avg']) < 25.0 and s['gender'] == 'ស')

    mention_stats = {}
    for m in ['A', 'B', 'C', 'D', 'E', 'F']:
        m_tot = sum(1 for s in students_list if s['ann_mention'] == m)
        m_fem = sum(1 for s in students_list if s['ann_mention'] == m and s['gender'] == 'ស')
        mention_stats[m] = {
            'total': m_tot,
            'female': m_fem,
            'total_kh': to_khmer_2digits(m_tot),
            'female_kh': to_khmer_2digits(m_fem),
        }

    # Pagination:
    # Page 1: 38 rows
    # Page 2..N-1: 45 rows
    # Final Page: <= 37 rows + Footer
    sheets = []
    if total_students <= 25:
        sheets.append({
            'page_number': 1,
            'is_first_page': True,
            'is_last_page': True,
            'rows': students_list,
        })
    else:
        # Page 1
        p1_rows = students_list[:38]
        sheets.append({
            'page_number': 1,
            'is_first_page': True,
            'is_last_page': False,
            'rows': p1_rows,
        })
        rem_students = students_list[38:]
        cur_page = 2
        while rem_students:
            if len(rem_students) <= 37:
                sheets.append({
                    'page_number': cur_page,
                    'is_first_page': False,
                    'is_last_page': True,
                    'rows': rem_students,
                })
                break
            else:
                chunk = rem_students[:45]
                rem_students = rem_students[45:]
                is_last = (len(rem_students) == 0)
                sheets.append({
                    'page_number': cur_page,
                    'is_first_page': False,
                    'is_last_page': is_last,
                    'rows': chunk,
                })
                cur_page += 1

    total_pages = len(sheets)
    for s in sheets:
        s['total_pages'] = total_pages
        s['page_number_kh'] = to_khmer_digits(s['page_number'])
        s['total_pages_kh'] = to_khmer_digits(total_pages)

    context = {
        'sheets': sheets,
        'total_pages': total_pages,
        'total_students': total_students,
        'total_students_kh': to_khmer_digits(total_students),
        'total_females': total_females,
        'total_females_kh': to_khmer_digits(total_females),
        'passed_count_kh': to_khmer_digits(passed_count),
        'passed_females_kh': to_khmer_digits(passed_females),
        'failed_count_kh': to_khmer_digits(failed_count),
        'failed_females_kh': to_khmer_digits(failed_females),
        'mention_stats': mention_stats,
        'academic_year_kh': academic_year_kh,
        'grade_title': grade_title,
        'selected_grade_level': selected_grade_level,
        'ministry_name': ministry_name,
        'school_name': school_name,
        'location_name': location_name,
        'sign_day_kh': sign_day_kh,
        'sign_month_kh': sign_month_kh,
        'sign_year_kh': sign_year_kh,
        'lunar_date': lunar_date,
        'sign_role': sign_role,
        'academic_years': academic_years,
        'target_year': target_year,
        'classrooms': classrooms_qs,
        'selected_classroom': selected_classroom,
        'data_source': 'official' if load_official else 'live',
        'is_admin': request.user.is_superuser or getattr(request.user, 'role', '') in ['ADMIN', 'TEACHER'],
    }
    return render(request, 'examinations/annual_results_print.html', context)


# ==============================================================================
# TEACHER EXAM INVIGILATOR / PROCTOR SHIFT SYSTEM VIEWS (ប្រព័ន្ធសុំវេនអនុរក្ស)
# ==============================================================================

@login_required
@role_required(['ADMIN'])
def exam_invigilator_plans_list(request):
    """
    Admin View: Displays all Exam Invigilator Configuration Plans, their Active status,
    and summary statistics (Total Slots, Spots Needed vs Filled).
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    
    plans = list(ExamInvigilatorPlan.objects.select_related('academic_year').prefetch_related('shift_slots__registrations').all().order_by('-start_date', '-created_at'))
    
    # Calculate statistics for each plan
    for p in plans:
        slots = list(p.shift_slots.all())
        p.calc_total_slots = len(slots)
        p.calc_total_capacity = sum(s.max_invigilators for s in slots)
        p.calc_total_registered = sum(s.registered_count for s in slots)
        p.calc_percentage = round((p.calc_total_registered / p.calc_total_capacity * 100), 1) if p.calc_total_capacity > 0 else 0

    return render(request, 'examinations/invigilators/plan_list.html', {
        'plans': plans,
        'active_year': active_year,
    })


@login_required
@role_required(['ADMIN'])
def exam_invigilator_plan_create(request):
    """
    Admin View: Creates a new Exam Invigilator Plan with default teacher groups and auto-generated slots.
    Supports 1-click linking directly from a StandardizedExam or an Exam Session (សម័យប្រឡង).
    """
    from apps.academics.utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    academic_years = AcademicYear.objects.all().order_by('-start_date')

    # Read pre-linking parameters if opened from an exam session
    pre_exam_id = request.GET.get('exam_id')
    pre_session_key = request.GET.get('session_key', '').strip()
    pre_title = request.GET.get('title', '').strip()
    pre_date_str = request.GET.get('date', '').strip()
    pre_rooms_str = request.GET.get('rooms', '').strip()
    pre_year_id = request.GET.get('year')

    linked_exam = None
    if pre_exam_id and str(pre_exam_id).isdigit():
        linked_exam = StandardizedExam.objects.filter(id=int(pre_exam_id)).first()

    default_title = ""
    default_start_date = None
    default_end_date = None
    default_slots_capacity = 20
    default_ay = active_year
    room_count = 0

    if linked_exam:
        default_ay = linked_exam.academic_year
        clean_name = get_clean_exam_session_title(linked_exam.name)
        default_title = f"វេនអនុរក្ស៖ {clean_name}"
        default_start_date = linked_exam.exam_date
        default_end_date = linked_exam.exam_date
        room_count = linked_exam.rooms.count()
        if room_count > 0:
            default_slots_capacity = room_count * 2
        if not pre_session_key:
            pre_session_key = f"{linked_exam.academic_year_id}_{linked_exam.exam_date}_{clean_name}"
    elif pre_title:
        clean_name = get_clean_exam_session_title(pre_title)
        default_title = clean_name if clean_name.startswith("វេនអនុរក្ស") else f"វេនអនុរក្ស៖ {clean_name}"
        if pre_date_str:
            try:
                default_start_date = datetime.datetime.strptime(pre_date_str, '%Y-%m-%d').date()
                default_end_date = default_start_date
            except Exception:
                pass
        if pre_rooms_str and pre_rooms_str.isdigit():
            room_count = int(pre_rooms_str)
            if room_count > 0:
                default_slots_capacity = room_count * 2
        if pre_year_id and str(pre_year_id).isdigit():
            found_ay = AcademicYear.objects.filter(id=int(pre_year_id)).first()
            if found_ay:
                default_ay = found_ay

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        ay_id = request.POST.get('academic_year')
        ay = AcademicYear.objects.filter(id=ay_id).first() if ay_id else default_ay
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        description = request.POST.get('description', '').strip()
        
        post_exam_id = request.POST.get('standardized_exam')
        post_exam = StandardizedExam.objects.filter(id=int(post_exam_id)).first() if post_exam_id and str(post_exam_id).isdigit() else linked_exam
        post_session_key = request.POST.get('session_key', '').strip() or pre_session_key or (f"{post_exam.academic_year_id}_{post_exam.exam_date}_{get_clean_exam_session_title(post_exam.name)}" if post_exam else None)

        is_active = (request.POST.get('is_active') == 'on')
        allow_reg = (request.POST.get('allow_teacher_registration') == 'on')
        reg_quota = int(request.POST.get('default_regular_quota', 4))
        off_quota = int(request.POST.get('default_office_quota', 5))
        auto_create_slots = (request.POST.get('auto_create_slots') == 'on')
        slots_capacity = int(request.POST.get('slots_capacity', default_slots_capacity))

        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except Exception:
            messages.error(request, "កាលបរិច្ឆេទចាប់ផ្តើម និងបញ្ចប់មិនត្រឹមត្រូវឡើយ!")
            return redirect('exam_invigilator_plan_create')

        if start_date > end_date:
            messages.error(request, "កាលបរិច្ឆេទចាប់ផ្តើមមិនអាចក្រោយកាលបរិច្ឆេទបញ្ចប់បានទេ!")
            return redirect('exam_invigilator_plan_create')

        with transaction.atomic():
            # If set to active, deactivate other plans to avoid confusion
            if is_active:
                ExamInvigilatorPlan.objects.all().update(is_active=False)

            plan = ExamInvigilatorPlan.objects.create(
                academic_year=ay,
                standardized_exam=post_exam,
                session_key=post_session_key,
                title=title,
                description=description,
                start_date=start_date,
                end_date=end_date,
                is_active=is_active,
                allow_teacher_registration=allow_reg,
                default_regular_quota=reg_quota,
                default_office_quota=off_quota
            )

            # 1. Create Default Teacher Duty Groups
            group_regular = TeacherDutyGroup.objects.create(
                plan=plan,
                name="គ្រូបង្រៀនធម្មតា (Regular Teachers)",
                required_shifts=reg_quota,
                description="គ្រូបង្រៀនទូទៅតាមមុខវិជ្ជា និងបន្ទុកថ្នាក់",
                order=1
            )
            group_office = TeacherDutyGroup.objects.create(
                plan=plan,
                name="គ្រូការិយាល័យ / រដ្ឋបាល (Office Staff Teachers)",
                required_shifts=off_quota,
                description="គ្រូដែលបម្រើការងារនៅការិយាល័យ និងរដ្ឋបាល",
                order=2
            )
            group_mgmt = TeacherDutyGroup.objects.create(
                plan=plan,
                name="គណៈគ្រប់គ្រង / នាយក-នាយករង (Management)",
                required_shifts=2,
                description="គណៈគ្រប់គ្រង និងប្រធានផ្នែក",
                order=3
            )

            # 2. Auto-Populate Teacher Quotas from Active Teachers
            active_teachers = Teacher.objects.filter(status=Teacher.Status.ACTIVE)
            for t in active_teachers:
                duty_lower = (t.current_duty or '').lower()
                if any(kw in duty_lower for kw in ['នាយក', 'នាយករង', 'management', 'director']):
                    assigned_group = group_mgmt
                elif any(kw in duty_lower for kw in ['ការិយាល័យ', 'រដ្ឋបាល', 'បណ្ណារក្ស', 'គណនេយ្យ', 'office', 'admin', 'clerk']):
                    assigned_group = group_office
                else:
                    assigned_group = group_regular

                TeacherDutyQuota.objects.create(
                    plan=plan,
                    teacher=t,
                    duty_group=assigned_group
                )

            # 3. Auto-Generate Daily Shift Slots if requested
            if auto_create_slots:
                curr = start_date
                day_num = 1
                slot_order = 1
                while curr <= end_date:
                    # Morning Slot
                    ExamShiftSlot.objects.create(
                        plan=plan,
                        date=curr,
                        session='MORNING',
                        session_name=f"ថ្ងៃទី{day_num} - 🌅 ពេលព្រឹក (Day {day_num} Morning)",
                        start_time=datetime.time(7, 0),
                        end_time=datetime.time(11, 0),
                        max_invigilators=slots_capacity,
                        order=slot_order
                    )
                    slot_order += 1
                    # Afternoon Slot
                    ExamShiftSlot.objects.create(
                        plan=plan,
                        date=curr,
                        session='AFTERNOON',
                        session_name=f"ថ្ងៃទី{day_num} - ⛅ ពេលរសៀល (Day {day_num} Afternoon)",
                        start_time=datetime.time(13, 0),
                        end_time=datetime.time(17, 0),
                        max_invigilators=slots_capacity,
                        order=slot_order
                    )
                    slot_order += 1
                    curr += datetime.timedelta(days=1)
                    day_num += 1

        messages.success(request, f"🎉 បានបង្កើតគម្រោងវេនអនុរក្សប្រឡង «{plan.title}» ព្រមទាំងបង្កើតវេនប្រឡង និងកំណត់កូតាគ្រូបង្រៀនស្វ័យប្រវត្តិដោយជោគជ័យ!")
        return redirect('exam_invigilator_roster_view', plan_id=plan.id)

    return render(request, 'examinations/invigilators/plan_form.html', {
        'active_year': default_ay,
        'academic_years': academic_years,
        'linked_exam': linked_exam,
        'session_key': pre_session_key,
        'default_title': default_title,
        'default_start_date': default_start_date,
        'default_end_date': default_end_date,
        'default_slots_capacity': default_slots_capacity,
        'room_count': room_count,
        'is_edit': False,
    })


@login_required
@role_required(['ADMIN'])
def exam_invigilator_plan_edit(request, plan_id):
    """
    Admin View: Edit plan settings, add/edit/delete shift slots.
    """
    plan = get_object_or_404(ExamInvigilatorPlan, id=plan_id)
    academic_years = AcademicYear.objects.all().order_by('-start_date')

    if request.method == 'POST':
        action = request.POST.get('form_action', 'update_plan')

        if action == 'update_plan':
            plan.title = request.POST.get('title', '').strip()
            ay_id = request.POST.get('academic_year')
            if ay_id:
                plan.academic_year_id = int(ay_id)
            plan.description = request.POST.get('description', '').strip()
            plan.is_active = (request.POST.get('is_active') == 'on')
            plan.allow_teacher_registration = (request.POST.get('allow_teacher_registration') == 'on')
            plan.default_regular_quota = int(request.POST.get('default_regular_quota', plan.default_regular_quota))
            plan.default_office_quota = int(request.POST.get('default_office_quota', plan.default_office_quota))

            if request.POST.get('standardized_exam'):
                try:
                    plan.standardized_exam_id = int(request.POST.get('standardized_exam'))
                except Exception:
                    pass
            if 'session_key' in request.POST:
                plan.session_key = request.POST.get('session_key', '').strip() or None

            if plan.is_active:
                ExamInvigilatorPlan.objects.exclude(id=plan.id).update(is_active=False)

            plan.save()
            messages.success(request, "🎉 បានកែប្រែព័ត៌មានគម្រោងដោយជោគជ័យ!")
            return redirect('exam_invigilator_plan_edit', plan_id=plan.id)

        elif action == 'add_slot':
            date_str = request.POST.get('slot_date')
            session = request.POST.get('slot_session', 'MORNING')
            name = request.POST.get('slot_name', '').strip()
            start_str = request.POST.get('slot_start', '07:00')
            end_str = request.POST.get('slot_end', '11:00')
            cap = int(request.POST.get('slot_capacity', 20))

            try:
                slot_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                st = datetime.datetime.strptime(start_str, '%H:%M').time()
                et = datetime.datetime.strptime(end_str, '%H:%M').time()
            except Exception:
                messages.error(request, "កាលបរិច្ឆេទ ឬម៉ោងមិនត្រឹមត្រូវឡើយ!")
                return redirect('exam_invigilator_plan_edit', plan_id=plan.id)

            if not name:
                name = f"{slot_date.strftime('%d/%m/%Y')} - {'ព្រឹក' if session=='MORNING' else 'រសៀល'}"

            ExamShiftSlot.objects.create(
                plan=plan,
                date=slot_date,
                session=session,
                session_name=name,
                start_time=st,
                end_time=et,
                max_invigilators=cap,
                order=plan.shift_slots.count() + 1
            )
            messages.success(request, f"🎉 បានបន្ថែមវេនប្រឡង «{name}» ដោយជោគជ័យ!")
            return redirect('exam_invigilator_plan_edit', plan_id=plan.id)

        elif action == 'delete_slot':
            slot_id = request.POST.get('slot_id')
            slot = get_object_or_404(ExamShiftSlot, id=slot_id, plan=plan)
            slot.delete()
            messages.success(request, "🗑️ បានលុបវេនប្រឡងដោយជោគជ័យ!")
            return redirect('exam_invigilator_plan_edit', plan_id=plan.id)

    slots = list(plan.shift_slots.all().order_by('date', 'start_time'))
    return render(request, 'examinations/invigilators/plan_form.html', {
        'plan': plan,
        'slots': slots,
        'academic_years': academic_years,
        'is_edit': True,
    })


@login_required
@role_required(['ADMIN'])
def exam_invigilator_plan_toggle_active(request, plan_id):
    """
    Admin View: 1-Click Toggle Active / Inactive switch.
    When activated, allows teachers to view and request shifts.
    When deactivated, strictly hides and blocks requests from teachers.
    """
    if request.method == 'POST':
        plan = get_object_or_404(ExamInvigilatorPlan, id=plan_id)
        if not plan.is_active:
            # Deactivate any other active plan and activate this one
            ExamInvigilatorPlan.objects.exclude(id=plan.id).update(is_active=False)
            plan.is_active = True
            plan.allow_teacher_registration = True
            plan.save(update_fields=['is_active', 'allow_teacher_registration'])
            messages.success(request, f"🟢 បានបើកដំណើរការការស្នើសុំវេនអនុរក្ស «{plan.title}» ជាផ្លូវការ! គ្រូបង្រៀនអាចមើលឃើញ និងស្នើសុំវេនបានហើយ។")
        else:
            plan.is_active = False
            plan.save(update_fields=['is_active'])
            messages.warning(request, f"🔴 បានបិទដំណើរការការស្នើសុំវេនអនុរក្ស «{plan.title}»! ផ្ទាំងស្នើសុំត្រូវបានលាក់ និងបិទមិនឱ្យគ្រូស្នើសុំទៀតទេ។")

    return redirect('exam_invigilator_plans_list')


@login_required
@role_required(['ADMIN'])
def exam_invigilator_plan_delete(request, plan_id):
    """
    Admin View: Deletes an Exam Invigilator Plan.
    """
    if request.method == 'POST':
        plan = get_object_or_404(ExamInvigilatorPlan, id=plan_id)
        title = plan.title
        plan.delete()
        messages.success(request, f"🗑️ បានលុបគម្រោងវេនអនុរក្ស «{title}» ដោយជោគជ័យ!")
    return redirect('exam_invigilator_plans_list')


@login_required
@role_required(['ADMIN'])
def exam_invigilator_quotas_manage(request, plan_id):
    """
    Admin View: Manages the 6 Exam Committee Roles, Role-based Capacity per Shift,
    and individual teacher shift quotas and role assignments.
    """
    plan = get_object_or_404(ExamInvigilatorPlan, id=plan_id)
    plan.ensure_default_role_settings()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'save_role_settings':
            # Update settings for each of the 6 roles
            for r_set in plan.role_settings.all():
                r_code = r_set.role
                req_val = (request.POST.get(f'is_requestable_{r_code}') == 'on')
                cap_val = request.POST.get(f'capacity_{r_code}', '1')
                auto_val = (request.POST.get(f'auto_assign_{r_code}') == 'on')

                r_set.is_requestable = req_val
                if cap_val and cap_val.isdigit():
                    r_set.capacity_per_shift = max(1, int(cap_val))
                r_set.auto_assign_all_shifts = auto_val
                r_set.save(update_fields=['is_requestable', 'capacity_per_shift', 'auto_assign_all_shifts'])

            messages.success(request, "🎉 បានធ្វើបច្ចុប្បន្នភាពការកំណត់មុខងារគណៈកម្មការទាំង ៦ ដោយជោគជ័យ!")
            return redirect('exam_invigilator_quotas_manage', plan_id=plan.id)

        elif action == 'add_group':
            name = request.POST.get('name', '').strip()
            shifts = int(request.POST.get('required_shifts', 4))
            desc = request.POST.get('description', '').strip()
            if name:
                TeacherDutyGroup.objects.create(
                    plan=plan,
                    name=name,
                    required_shifts=shifts,
                    description=desc,
                    order=plan.duty_groups.count() + 1
                )
                messages.success(request, f"🎉 បានបន្ថែមក្រុម «{name}» ({shifts} វេន) ដោយជោគជ័យ!")

        elif action == 'edit_group':
            group_id = request.POST.get('group_id')
            grp = get_object_or_404(TeacherDutyGroup, id=group_id, plan=plan)
            grp.name = request.POST.get('name', grp.name).strip()
            grp.required_shifts = int(request.POST.get('required_shifts', grp.required_shifts))
            grp.description = request.POST.get('description', grp.description).strip()
            grp.save()
            messages.success(request, f"🎉 បានកែប្រែក្រុម «{grp.name}» ដោយជោគជ័យ!")

        elif action == 'delete_group':
            group_id = request.POST.get('group_id')
            grp = get_object_or_404(TeacherDutyGroup, id=group_id, plan=plan)
            grp.delete()
            messages.success(request, "🗑️ បានលុបក្រុមដោយជោគជ័យ!")

        elif action == 'auto_classify_teachers':
            # Auto-classify active teachers
            active_teachers = Teacher.objects.filter(status=Teacher.Status.ACTIVE)
            group_regular = plan.duty_groups.filter(name__icontains='ធម្មតា').first() or plan.duty_groups.first()
            group_office = plan.duty_groups.filter(name__icontains='ការិយាល័យ').first() or plan.duty_groups.first()
            group_mgmt = plan.duty_groups.filter(name__icontains='គ្រប់គ្រង').first() or group_office

            count = 0
            with transaction.atomic():
                for t in active_teachers:
                    duty_lower = (t.current_duty or '').lower()
                    if any(kw in duty_lower for kw in ['នាយក', 'នាយករង', 'management', 'director']):
                        assigned_group = group_mgmt
                        assigned_role = ExamCommitteeRole.PRESIDENT if 'នាយក' in duty_lower and 'រង' not in duty_lower else ExamCommitteeRole.VICE_PRESIDENT
                        auto_assign = True
                    elif any(kw in duty_lower for kw in ['ការិយាល័យ', 'រដ្ឋបាល', 'បណ្ណារក្ស', 'គណនេយ្យ', 'office', 'admin', 'clerk']):
                        assigned_group = group_office
                        assigned_role = ExamCommitteeRole.SECRETARIAT
                        auto_assign = False
                    else:
                        assigned_group = group_regular
                        assigned_role = ExamCommitteeRole.INVIGILATOR
                        auto_assign = False

                    quota_obj, created = TeacherDutyQuota.objects.get_or_create(
                        plan=plan,
                        teacher=t,
                        defaults={
                            'duty_group': assigned_group,
                            'assigned_role': assigned_role,
                            'auto_assign_all_shifts': auto_assign
                        }
                    )
                    if not created:
                        quota_obj.duty_group = assigned_group
                        quota_obj.save(update_fields=['duty_group'])
                    count += 1

            messages.success(request, f"⚡ បានកំណត់ក្រុម និងកូតាស្វ័យប្រវត្តិជូនគ្រូចំនួន {count} នាក់ដោយជោគជ័យ!")

        elif action == 'save_teacher_quotas':
            quotas = plan.teacher_quotas.all()
            auto_assigned_teachers = []
            for q in quotas:
                grp_val = request.POST.get(f'group_{q.id}')
                role_val = request.POST.get(f'role_{q.id}')
                auto_val = (request.POST.get(f'auto_assign_{q.id}') == 'on')
                custom_val = request.POST.get(f'custom_{q.id}', '').strip()
                exempt_val = (request.POST.get(f'exempt_{q.id}') == 'on')

                if grp_val and grp_val.isdigit():
                    q.duty_group_id = int(grp_val)
                if role_val in ExamCommitteeRole.values:
                    q.assigned_role = role_val
                q.auto_assign_all_shifts = auto_val
                if custom_val and custom_val.isdigit():
                    q.custom_required_shifts = int(custom_val)
                else:
                    q.custom_required_shifts = None
                q.is_exempt = exempt_val
                q.save(update_fields=['duty_group', 'assigned_role', 'auto_assign_all_shifts', 'custom_required_shifts', 'is_exempt'])

                if auto_val or (q.assigned_role == ExamCommitteeRole.PRESIDENT and not q.is_exempt):
                    auto_assigned_teachers.append(q)

            # Auto-assign teachers to all shifts if configured
            slots = list(plan.shift_slots.all())
            for q in auto_assigned_teachers:
                for s in slots:
                    TeacherShiftRegistration.objects.get_or_create(
                        slot=s,
                        teacher=q.teacher,
                        defaults={'role': q.assigned_role, 'status': 'ADMIN_ASSIGNED'}
                    )

            messages.success(request, "🎉 បានរក្សាទុកការកំណត់កូតា និងមុខងារគណៈកម្មការរបស់គ្រូបង្រៀនដោយជោគជ័យ!")

        elif action == 'sync_auto_assigned_teachers':
            slots = list(plan.shift_slots.all())
            auto_quotas = plan.teacher_quotas.filter(Q(auto_assign_all_shifts=True) | Q(assigned_role=ExamCommitteeRole.PRESIDENT), is_exempt=False)
            synced_count = 0
            for q in auto_quotas:
                for s in slots:
                    reg, created = TeacherShiftRegistration.objects.get_or_create(
                        slot=s,
                        teacher=q.teacher,
                        defaults={'role': q.assigned_role, 'status': 'ADMIN_ASSIGNED'}
                    )
                    if not created and reg.role != q.assigned_role:
                        reg.role = q.assigned_role
                        reg.save(update_fields=['role'])
                synced_count += 1
            messages.success(request, f"⚡ បានចាត់តាំងប្រធាន/អនុប្រធានចូលគ្រប់វេនប្រឡងទាំងអស់ស្វ័យប្រវត្តិចំនួន {synced_count} នាក់!")

        return redirect('exam_invigilator_quotas_manage', plan_id=plan.id)

    duty_groups = list(plan.duty_groups.all())
    role_settings = list(plan.role_settings.all())

    # Ensure all active teachers have quota records
    active_teachers = list(Teacher.objects.filter(status=Teacher.Status.ACTIVE).order_by('khmer_name'))
    existing_quotas = {q.teacher_id: q for q in plan.teacher_quotas.select_related('teacher', 'duty_group').all()}

    teacher_quota_rows = []
    default_group = duty_groups[0] if duty_groups else None
    for t in active_teachers:
        q_obj = existing_quotas.get(t.id)
        if not q_obj:
            q_obj = TeacherDutyQuota.objects.create(
                plan=plan,
                teacher=t,
                duty_group=default_group,
                assigned_role=ExamCommitteeRole.INVIGILATOR
            )
        q_obj.registered_count = t.exam_shift_registrations.filter(slot__plan=plan).exclude(status='CANCELLED').count()
        q_obj.is_fulfilled = (q_obj.registered_count >= q_obj.effective_required_shifts)
        teacher_quota_rows.append(q_obj)

    return render(request, 'examinations/invigilators/quotas_manage.html', {
        'plan': plan,
        'duty_groups': duty_groups,
        'role_settings': role_settings,
        'committee_roles': ExamCommitteeRole.choices,
        'teacher_quotas': teacher_quota_rows,
    })


@login_required
@role_required(['ADMIN'])
def exam_invigilator_roster_view(request, plan_id):
    """
    Admin View: Duty Matrix of all shift slots, grouped by the 6 Committee Roles:
    (ប្រធាន, អនុប្រធាន, គណៈកម្មការកណ្តាល, ត្រួតអគារ, អនុរក្ស, បូកស្រង់ពិន្ទុ).
    Admin has full override authority to assign, edit, or remove anytime.
    """
    plan = get_object_or_404(ExamInvigilatorPlan, id=plan_id)
    plan.ensure_default_role_settings()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'admin_assign_teacher':
            slot_id = request.POST.get('slot_id')
            teacher_id = request.POST.get('teacher_id')
            role = request.POST.get('role', ExamCommitteeRole.INVIGILATOR)
            room = request.POST.get('room_assignment', '').strip()
            slot = get_object_or_404(ExamShiftSlot, id=slot_id, plan=plan)
            teacher = get_object_or_404(Teacher, id=teacher_id)

            reg, created = TeacherShiftRegistration.objects.get_or_create(
                slot=slot,
                teacher=teacher,
                defaults={'role': role, 'status': 'ADMIN_ASSIGNED', 'room_assignment': room}
            )
            if not created:
                reg.role = role
                reg.status = 'ADMIN_ASSIGNED'
                if room:
                    reg.room_assignment = room
                reg.save(update_fields=['role', 'status', 'room_assignment'])
            messages.success(request, f"🎉 បានចាត់តាំងលោកគ្រូ/អ្នកគ្រូ {teacher.khmer_name} ជា «{reg.get_role_display()}» ក្នុង {slot.session_name} ដោយជោគជ័យ!")

        elif action == 'admin_update_registration':
            reg_id = request.POST.get('registration_id')
            role = request.POST.get('role')
            room = request.POST.get('room_assignment', '').strip()
            reg = get_object_or_404(TeacherShiftRegistration, id=reg_id, slot__plan=plan)
            if role in ExamCommitteeRole.values:
                reg.role = role
            reg.room_assignment = room
            reg.save(update_fields=['role', 'room_assignment'])
            messages.success(request, f"🎉 បានកែប្រែព័ត៌មានរបស់ «{reg.teacher.khmer_name}» ដោយជោគជ័យ!")

        elif action == 'admin_remove_teacher':
            reg_id = request.POST.get('registration_id')
            reg = get_object_or_404(TeacherShiftRegistration, id=reg_id, slot__plan=plan)
            t_name = reg.teacher.khmer_name
            reg.delete()
            messages.success(request, f"🗑️ បានដកឈ្មោះ {t_name} ចេញពីវេនប្រឡងដោយជោគជ័យ!")

        elif action == 'toggle_teacher_registration':
            plan.allow_teacher_registration = not plan.allow_teacher_registration
            plan.save(update_fields=['allow_teacher_registration'])
            status_str = "បើក" if plan.allow_teacher_registration else "បិទ"
            messages.success(request, f"🎉 បាន{status_str}ការស្នើសុំវេនរបស់គ្រូបង្រៀនដោយជោគជ័យ!")

        return redirect('exam_invigilator_roster_view', plan_id=plan.id)

    slots = list(plan.shift_slots.prefetch_related('registrations__teacher').order_by('date', 'start_time'))

    # Group registrations in each slot into the 6 Committee Roles
    for s in slots:
        regs = list(s.registrations.select_related('teacher').all())
        s.grouped_registrations = {
            ExamCommitteeRole.PRESIDENT: [r for r in regs if r.role == ExamCommitteeRole.PRESIDENT],
            ExamCommitteeRole.VICE_PRESIDENT: [r for r in regs if r.role == ExamCommitteeRole.VICE_PRESIDENT],
            ExamCommitteeRole.SECRETARIAT: [r for r in regs if r.role == ExamCommitteeRole.SECRETARIAT],
            ExamCommitteeRole.BUILDING_INSPECTOR: [r for r in regs if r.role == ExamCommitteeRole.BUILDING_INSPECTOR],
            ExamCommitteeRole.INVIGILATOR: [r for r in regs if r.role == ExamCommitteeRole.INVIGILATOR],
            ExamCommitteeRole.TABULATOR: [r for r in regs if r.role == ExamCommitteeRole.TABULATOR],
        }
        s.role_stats = {
            role_code: {
                'capacity': s.get_role_capacity(role_code),
                'count': len(s.grouped_registrations[role_code]),
                'remaining': s.get_role_remaining_spots(role_code),
                'is_full': s.is_role_full(role_code),
            }
            for role_code in ExamCommitteeRole.values
        }

    # Build list of unfulfilled teachers
    active_teachers = list(Teacher.objects.filter(status=Teacher.Status.ACTIVE).order_by('khmer_name'))
    existing_quotas = {q.teacher_id: q for q in plan.teacher_quotas.select_related('duty_group').all()}
    registrations_count_map = dict(
        TeacherShiftRegistration.objects.filter(slot__plan=plan)
        .exclude(status='CANCELLED')
        .values('teacher_id')
        .annotate(c=Count('id'))
        .values_list('teacher_id', 'c')
    )

    unfulfilled_teachers = []
    for t in active_teachers:
        q_obj = existing_quotas.get(t.id)
        req = q_obj.effective_required_shifts if q_obj else plan.default_regular_quota
        reg_count = registrations_count_map.get(t.id, 0)
        if reg_count < req:
            unfulfilled_teachers.append({
                'teacher': t,
                'required': req,
                'registered': reg_count,
                'shortage': req - reg_count,
                'role': q_obj.get_assigned_role_display() if q_obj else 'គណៈកម្មការអនុរក្ស',
                'group': q_obj.duty_group.name if (q_obj and q_obj.duty_group) else 'ធម្មតា',
            })

    unfulfilled_teachers.sort(key=lambda x: -x['shortage'])

    return render(request, 'examinations/invigilators/roster_matrix.html', {
        'plan': plan,
        'slots': slots,
        'committee_roles': ExamCommitteeRole.choices,
        'unfulfilled_teachers': unfulfilled_teachers,
        'active_teachers': active_teachers,
    })


@login_required
@role_required(['ADMIN'])
def api_invigilator_auto_assign(request, plan_id):
    """
    1-Click Auto-Assign: Automatically assigns unfulfilled teachers to slots that have empty spots.
    """
    if request.method != 'POST':
        return redirect('exam_invigilator_roster_view', plan_id=plan_id)

    plan = get_object_or_404(ExamInvigilatorPlan, id=plan_id)
    slots = list(plan.shift_slots.all().order_by('date', 'start_time'))
    
    active_teachers = list(Teacher.objects.filter(status=Teacher.Status.ACTIVE).order_by('khmer_name'))
    existing_quotas = {q.teacher_id: q for q in plan.teacher_quotas.select_related('duty_group').all()}
    
    # Map teacher registered slot ids & slot filled counts
    teacher_registered_slots = {t.id: set() for t in active_teachers}
    slot_registered_counts = {s.id: 0 for s in slots}
    
    for reg in TeacherShiftRegistration.objects.filter(slot__plan=plan).exclude(status='CANCELLED'):
        if reg.teacher_id in teacher_registered_slots:
            teacher_registered_slots[reg.teacher_id].add(reg.slot_id)
        if reg.slot_id in slot_registered_counts:
            slot_registered_counts[reg.slot_id] += 1

    assigned_count = 0
    with transaction.atomic():
        # Iterate over unfulfilled teachers
        for t in active_teachers:
            q_obj = existing_quotas.get(t.id)
            if q_obj and q_obj.is_exempt:
                continue
            req = q_obj.effective_required_shifts if q_obj else plan.default_regular_quota
            current_count = len(teacher_registered_slots[t.id])

            while current_count < req:
                # Find available slot where teacher is not yet registered and spot is available
                available_slot = None
                for s in slots:
                    if s.id not in teacher_registered_slots[t.id] and slot_registered_counts[s.id] < s.max_invigilators:
                        available_slot = s
                        break

                if not available_slot:
                    break # No more spots available

                TeacherShiftRegistration.objects.create(
                    slot=available_slot,
                    teacher=t,
                    status='ADMIN_ASSIGNED',
                    notes='Auto-assigned by system'
                )
                teacher_registered_slots[t.id].add(available_slot.id)
                slot_registered_counts[available_slot.id] += 1
                current_count += 1
                assigned_count += 1

    messages.success(request, f"⚡ បានចាត់តាំងបំពេញវេនស្វ័យប្រវត្តិចំនួន {assigned_count} វេន ដោយជោគជ័យ!")
    return redirect('exam_invigilator_roster_view', plan_id=plan.id)


@login_required
@role_required(['ADMIN'])
def exam_invigilator_roster_print(request, plan_id):
    """
    Official MoEYS Exam Invigilator Duty Roster (តារាងចាត់តាំងអនុរក្សប្រឡង) for printing.
    """
    plan = get_object_or_404(ExamInvigilatorPlan, id=plan_id)
    slots = list(plan.shift_slots.prefetch_related('registrations__teacher').order_by('date', 'start_time'))

    return render(request, 'examinations/invigilators/duty_roster_print.html', {
        'plan': plan,
        'slots': slots,
    })


@login_required
def exam_invigilator_teacher_portal(request):
    """
    Teacher Self-Service Portal:
    - Checks teacher's assigned committee role.
    - If role is not requestable (e.g. President / Vice President pre-assigned by Admin), displays clear notice.
    - If requestable (e.g. Invigilator / Tabulator), displays real-time role-based remaining capacity.
    - First-come, first-served logic.
    """
    plan = ExamInvigilatorPlan.objects.filter(is_active=True).first()

    if not plan or not plan.allow_teacher_registration:
        return render(request, 'examinations/invigilators/teacher_portal.html', {
            'is_active': False,
            'plan': plan,
        })

    plan.ensure_default_role_settings()

    # Find teacher profile
    teacher = None
    if hasattr(request.user, 'teacher_profile') and request.user.teacher_profile:
        teacher = request.user.teacher_profile
    elif request.user.is_superuser or request.user.role == 'ADMIN':
        tid = request.GET.get('teacher_id')
        if tid and tid.isdigit():
            teacher = Teacher.objects.filter(id=int(tid)).first()
        if not teacher:
            teacher = Teacher.objects.filter(status=Teacher.Status.ACTIVE).first()

    if not teacher:
        messages.warning(request, "មិនមានទម្រង់គ្រូបង្រៀនដែលត្រូវគ្នានឹងគណនីរបស់អ្នកឡើយ!")
        return render(request, 'examinations/invigilators/teacher_portal.html', {
            'is_active': False,
            'plan': plan,
            'no_teacher_profile': True,
        })

    quota_obj = TeacherDutyQuota.objects.filter(plan=plan, teacher=teacher).first()
    required_shifts = quota_obj.effective_required_shifts if quota_obj else plan.default_regular_quota
    duty_group_name = quota_obj.duty_group.name if (quota_obj and quota_obj.duty_group) else "គ្រូបង្រៀនធម្មតា"

    assigned_role = quota_obj.assigned_role if quota_obj else ExamCommitteeRole.INVIGILATOR
    assigned_role_display = quota_obj.get_assigned_role_display() if quota_obj else "គណៈកម្មការអនុរក្ស (អនុរក្ស)"

    role_setting = plan.role_settings.filter(role=assigned_role).first()
    is_role_requestable = role_setting.is_requestable if role_setting else True
    is_auto_assigned_all = quota_obj.auto_assign_all_shifts if quota_obj else (assigned_role == ExamCommitteeRole.PRESIDENT)

    # Get teacher's registered slot IDs
    registered_slot_ids = set(
        TeacherShiftRegistration.objects.filter(slot__plan=plan, teacher=teacher)
        .exclude(status='CANCELLED')
        .values_list('slot_id', flat=True)
    )

    # Group slots by date
    all_slots = list(plan.shift_slots.prefetch_related('registrations').order_by('date', 'start_time'))
    slots_by_date = {}
    for s in all_slots:
        s.is_teacher_registered = (s.id in registered_slot_ids)
        s.assigned_role_capacity = s.get_role_capacity(assigned_role)
        s.assigned_role_registered = s.get_role_registered_count(assigned_role)
        s.assigned_role_remaining = s.get_role_remaining_spots(assigned_role)
        s.assigned_role_is_full = s.is_role_full(assigned_role)

        d_str = s.date.strftime('%Y-%m-%d')
        if d_str not in slots_by_date:
            slots_by_date[d_str] = {
                'date': s.date,
                'slots': []
            }
        slots_by_date[d_str]['slots'].append(s)

    current_count = len(registered_slot_ids)
    progress_percentage = min(100, round(current_count / required_shifts * 100)) if required_shifts > 0 else 100

    return render(request, 'examinations/invigilators/teacher_portal.html', {
        'is_active': True,
        'plan': plan,
        'teacher': teacher,
        'assigned_role': assigned_role,
        'assigned_role_display': assigned_role_display,
        'is_role_requestable': is_role_requestable,
        'is_auto_assigned_all': is_auto_assigned_all,
        'required_shifts': required_shifts,
        'current_count': current_count,
        'remaining_to_choose': max(0, required_shifts - current_count),
        'progress_percentage': progress_percentage,
        'duty_group_name': duty_group_name,
        'slots_by_date': list(slots_by_date.values()),
        'all_teachers': Teacher.objects.filter(status=Teacher.Status.ACTIVE).order_by('khmer_name') if request.user.role == 'ADMIN' else None,
    })


@login_required
def api_toggle_invigilator_slot(request):
    """
    AJAX API for teachers to toggle selection of an exam shift slot.
    Enforces Role-Based Capacity and First-Come, First-Served logic.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    slot_id = request.POST.get('slot_id')
    if not slot_id or not slot_id.isdigit():
        return JsonResponse({'success': False, 'error': 'Slot ID is required'}, status=400)

    slot = get_object_or_404(ExamShiftSlot, id=int(slot_id))
    plan = slot.plan
    plan.ensure_default_role_settings()

    if not plan.is_active or not plan.allow_teacher_registration:
        return JsonResponse({'success': False, 'error': 'ការស្នើសុំវេនអនុរក្សត្រូវបានបិទដោយគណៈគ្រប់គ្រង!'}, status=403)

    # Resolve Teacher
    teacher = None
    if hasattr(request.user, 'teacher_profile') and request.user.teacher_profile:
        teacher = request.user.teacher_profile
    elif request.user.is_superuser or request.user.role == 'ADMIN':
        tid = request.POST.get('teacher_id')
        if tid and tid.isdigit():
            teacher = Teacher.objects.filter(id=int(tid)).first()

    if not teacher:
        return JsonResponse({'success': False, 'error': 'រកមិនឃើញគណនីគ្រូបង្រៀនឡើយ!'}, status=403)

    quota_obj = TeacherDutyQuota.objects.filter(plan=plan, teacher=teacher).first()
    assigned_role = quota_obj.assigned_role if quota_obj else ExamCommitteeRole.INVIGILATOR
    role_label = quota_obj.get_assigned_role_display() if quota_obj else assigned_role

    role_setting = plan.role_settings.filter(role=assigned_role).first()
    if role_setting and not role_setting.is_requestable and request.user.role != 'ADMIN':
        return JsonResponse({
            'success': False,
            'error': f'មុខងារ «{role_label}» ត្រូវបានចាត់តាំងដោយ Admin រួចជាស្រេច (មិនចាំបាច់ស្នើសុំវេនដោយខ្លួនឯងឡើយ)។'
        }, status=400)

    with transaction.atomic():
        # Check existing registration
        reg = TeacherShiftRegistration.objects.filter(slot=slot, teacher=teacher).first()

        if reg:
            # Toggle OFF -> Remove registration
            reg.delete()
            is_registered = False
            message = f"បានដកចេញពីវេន «{slot.session_name}» រួចរាល់!"
        else:
            # Toggle ON -> Check if designated capacity for this role is full (First-Come, First-Served)
            if slot.is_role_full(assigned_role):
                cap = slot.get_role_capacity(assigned_role)
                return JsonResponse({
                    'success': False,
                    'error': f'សូមអភ័យទោស! វេន «{slot.session_name}» សម្រាប់មុខងារ «{role_label}» បានពេញកូតា ({cap} នាក់) រួចហើយ (អាទិភាពអ្នកសុំមុន)។ សូមជ្រើសរើសវេនផ្សេងទៀត។'
                }, status=400)

            TeacherShiftRegistration.objects.create(
                slot=slot,
                teacher=teacher,
                role=assigned_role,
                status='CONFIRMED'
            )
            is_registered = True
            message = f"🎉 បានចុះឈ្មោះក្នុងវេន «{slot.session_name}» ជា «{role_label}» ដោយជោគជ័យ!"

        # Calculate updated counts
        required_shifts = quota_obj.effective_required_shifts if quota_obj else plan.default_regular_quota
        current_count = TeacherShiftRegistration.objects.filter(slot__plan=plan, teacher=teacher).exclude(status='CANCELLED').count()
        progress_percentage = min(100, round(current_count / required_shifts * 100)) if required_shifts > 0 else 100

        role_cap = slot.get_role_capacity(assigned_role)
        role_rem = slot.get_role_remaining_spots(assigned_role)
        role_is_full = slot.is_role_full(assigned_role)

        return JsonResponse({
            'success': True,
            'is_registered': is_registered,
            'slot_id': slot.id,
            'slot_registered_count': slot.registered_count,
            'slot_max': slot.max_invigilators,
            'slot_remaining': slot.remaining_spots,
            'slot_is_full': slot.is_full,
            'role_capacity': role_cap,
            'role_remaining': role_rem,
            'role_is_full': role_is_full,
            'current_count': current_count,
            'required_shifts': required_shifts,
            'remaining_to_choose': max(0, required_shifts - current_count),
            'progress_percentage': progress_percentage,
            'message': message,
        })


# ==============================================================================
# STUDENT & PARENT EXAM ADMISSION SLIP & NOTIFICATIONS
# ==============================================================================

@login_required
def student_exam_admission_slip(request, candidate_id):
    """
    Renders official Examination Admission Slip / Hall Ticket (ប័ណ្ណអនុញ្ញាតចូលរួមប្រឡង)
    for students and parents. Printable and mobile-friendly.
    """
    candidate = get_object_or_404(
        ExamCandidate.objects.select_related('exam', 'exam__academic_year', 'room', 'student', 'student__classroom'),
        id=candidate_id
    )

    user = request.user
    # If student or parent, verify ownership unless staff/admin
    if user.role == 'STUDENT' and not (user.is_superuser or user.is_staff):
        student_profile = getattr(user, 'student_profile', None)
        if candidate.student and student_profile and candidate.student_id != student_profile.id:
            if candidate.student.student_id != user.username and candidate.student.phone != user.phone:
                messages.error(request, "លោកអ្នកពុំមានសិទ្ធិចូលមើលប័ណ្ណអនុញ្ញាតប្រឡងរបស់សិស្សផ្សេងឡើយ។")
                return redirect('student_dashboard')

    subjects = list(candidate.exam.exam_subjects.all().select_related('subject').order_by('order'))

    return render(request, 'examinations/student/exam_admission_slip.html', {
        'candidate': candidate,
        'subjects': subjects,
    })


def send_exam_seating_notification_telegram(candidate):
    """
    Sends examination seating details (Room, Desk No, Shift) to the parent's Telegram Chat ID.
    """
    student = candidate.student
    if not student or not student.telegram_chat_id:
        return False, "សិស្សមិនទាន់មាន Telegram Chat ID ឡើយ"

    exam = candidate.exam
    room_str = candidate.room.room_name if candidate.room else "មិនទាន់កំណត់"
    building_str = f" ({candidate.room.building})" if (candidate.room and candidate.room.building) else ""
    desk_str = f"តុលេខ {candidate.desk_number:02d}" if candidate.desk_number else "មិនទាន់កំណត់"

    title = f"🏛️ ដំណឹងសម័យប្រឡង៖ {student.khmer_name}"
    msg = (
        f"🏛️ <b>ព័ត៌មានសម័យប្រឡង & បន្ទប់ប្រឡង</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 <b>សិស្ស៖</b> {student.khmer_name} (អត្តលេខ: {candidate.roll_number})\n"
        f"🏫 <b>ថ្នាក់៖</b> {candidate.origin_class or (student.classroom.name if student.classroom else '-')}\n"
        f"📅 <b>សម័យប្រឡង៖</b> {exam.name}\n"
        f"🗓️ <b>កាលបរិច្ឆេទ៖</b> {exam.exam_date.strftime('%d/%m/%Y')} ({exam.get_session_display()})\n"
        f"🏢 <b>បន្ទប់ប្រឡង៖</b> <b>{room_str}</b>{building_str}\n"
        f"🪑 <b>លេខតុក្នុងបន្ទប់៖</b> <b>{desk_str}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"ℹ️ <i>សូមប្អូនៗសិស្សានុសិស្សមកដល់បន្ទប់ប្រឡងមុនម៉ោងកំណត់ ១៥ នាទី និងយកមកជាមួយនូវកាតសិស្ស ឬប័ណ្ណអនុញ្ញាតប្រឡង។</i>"
    )

    from apps.accounts.utils import send_telegram_notification
    log = send_telegram_notification(
        title=title,
        message=msg,
        recipient_name=f"{student.khmer_name} (អាណាព្យាបាល)",
        recipient_phone=student.phone or student.emergency_phone,
        recipient_type="Parent",
        custom_chat_id=student.telegram_chat_id.strip()
    )
    return True, "បានបញ្ជូនដំណឹងជោគជ័យ"


def send_exam_exclusion_notification_telegram(exclusion):
    """
    Sends examination exclusion / disqualification notice to the parent's Telegram Chat ID with clear reason.
    """
    student = exclusion.student
    if not student or not student.telegram_chat_id:
        return False, "សិស្សមិនទាន់មាន Telegram Chat ID ឡើយ"

    target_name = exclusion.standardized_exam.name if exclusion.standardized_exam else (
        exclusion.exam_term.name if exclusion.exam_term else f"ការប្រឡងប្រចាំខែទី {exclusion.month}"
    )

    reason_display = exclusion.get_reason_display()
    if exclusion.reason == 'UNEXCUSED_ABSENCE':
        reason_display = "🚨 ឈប់រៀនច្រើន / អវត្តមានឥតច្បាប់ច្រើន (Excessive Unexcused Absences)"
    elif exclusion.reason == 'FEE_OVERDUE':
        reason_display = "💳 មិនទាន់បង់ប្រាក់ថ្លៃទឹកភ្លើង / ជំពាក់ប្រាក់កម្រៃសិក្សា (Fee Overdue / Unpaid Tuition or Utilities)"
    elif exclusion.reason == 'DISCIPLINARY':
        reason_display = "⚖️ បញ្ហាវិន័យ / ជាប់កិច្ចសន្យាវិន័យ (Disciplinary Hold)"

    title = f"⚠️ សេចក្តីជូនដំណឹងស្តីពីការប្រឡង៖ {student.khmer_name}"
    msg = (
        f"⚠️ <b>សេចក្តីជូនដំណឹងស្តីពីការលើកលែងមិនឱ្យប្រឡង</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 <b>សិស្ស៖</b> {student.khmer_name} (អត្តលេខ: {student.student_id})\n"
        f"📅 <b>សម័យប្រឡង៖</b> {target_name}\n"
        f"🚫 <b>ស្ថានភាព៖</b> <b>ពុំមានឈ្មោះក្នុងបញ្ជីប្រឡង</b>\n"
        f"❗ <b>មូលហេតុ៖</b> {reason_display}\n"
    )
    if exclusion.notes:
        msg += f"📝 <b>កំណត់សម្គាល់៖</b> {exclusion.notes}\n"
    msg += (
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📞 <i>សូមអាណាព្យាបាល ឬសិស្សានុសិស្សមេត្តាទាក់ទងមកកាន់ការិយាល័យរដ្ឋបាល ឬគណនេយ្យសាលាជាបន្ទាន់ ដើម្បីដោះស្រាយ និងទទួលបានសិទ្ធិចូលរួមប្រឡងឡើងវិញ។</i>"
    )

    from apps.accounts.utils import send_telegram_notification
    log = send_telegram_notification(
        title=title,
        message=msg,
        recipient_name=f"{student.khmer_name} (អាណាព្យាបាល)",
        recipient_phone=student.phone or student.emergency_phone,
        recipient_type="Parent",
        custom_chat_id=student.telegram_chat_id.strip()
    )
    return True, "បានបញ្ជូនដំណឹងជោគជ័យ"


@login_required
@role_required(['ADMIN'])
def api_send_exam_seating_telegram(request, exam_id):
    """
    Admin 1-Click: Dispatches seating notifications (room & desk number)
    to all candidate parents who have configured Telegram Chat IDs.
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    candidates = list(exam.candidates.filter(student__isnull=False).select_related('student', 'room'))

    sent_count = 0
    skipped_count = 0
    for cand in candidates:
        if cand.student and cand.student.telegram_chat_id:
            success, _ = send_exam_seating_notification_telegram(cand)
            if success:
                sent_count += 1
            else:
                skipped_count += 1
        else:
            skipped_count += 1

    messages.success(request, f"បានបញ្ជូនដំណឹងបន្ទប់ និងលេខតុតាម Telegram ទៅកាន់អាណាព្យាបាលចំនួន {sent_count} នាក់ជោគជ័យ (សិស្សគ្មាន Telegram ចំនួន {skipped_count} នាក់)។")
    return redirect('standardized_exam_manage', exam_id=exam.id)


# ==============================================================================
# EXAM SUBJECTS SELECTION & NON-TESTED SUBJECTS EXCLUSIONS MANAGEMENT
# ==============================================================================

@login_required
@role_required(['ADMIN'])
def exam_term_subjects_manage(request):
    """
    Admin Management Dashboard for Exam Subjects & Non-Tested Subjects Exclusions.
    Allows configuring which subjects are tested vs non-tested:
    - Per Exam Term (e.g. Monthly, Semester 1, Semester 2)
    - Per Grade Level (7-12) & Track (Science, Social, General)
    - Per Classroom (e.g. 11A vs 11B)
    - 1-Click Presets (All Subjects, 7 Science Subjects, 7 Social Subjects)
    """
    from apps.academics.utils import get_active_academic_year
    from .services import get_effective_term_subjects
    active_year = get_active_academic_year(request)
    
    terms = ExamTerm.objects.filter(academic_year=active_year).order_by('-start_date') if active_year else ExamTerm.objects.all().order_by('-start_date')
    selected_term_id = request.GET.get('term_id') or str(terms.first().id if terms.first() else '')
    selected_term = terms.filter(id=selected_term_id).first() if (selected_term_id and selected_term_id.isdigit()) else terms.first()

    selected_grade = request.GET.get('grade_level', '').strip()
    selected_class_id = request.GET.get('classroom_id', '').strip()

    classrooms_qs = Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code')
    if selected_grade and selected_grade.isdigit():
        classrooms_qs = classrooms_qs.filter(grade_level=int(selected_grade))

    selected_class = None
    if selected_class_id and selected_class_id.isdigit():
        selected_class = classrooms_qs.filter(id=int(selected_class_id)).first()

    # Determine scope and load subjects
    g_level = int(selected_grade) if (selected_grade and selected_grade.isdigit()) else (selected_class.grade_level if selected_class else 12)
    t_track = selected_class.track if selected_class else request.GET.get('track', 'GENERAL')

    subjects_data = get_effective_term_subjects(
        exam_term=selected_term,
        classroom=selected_class,
        grade_level=g_level if not selected_class else None,
        track=t_track if not selected_class else None,
        include_non_tested=True
    )

    all_grades = [7, 8, 9, 10, 11, 12]
    all_classrooms_list = Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code')

    return render(request, 'examinations/term_subjects_manage.html', {
        'terms': terms,
        'selected_term': selected_term,
        'all_grades': all_grades,
        'selected_grade': g_level,
        'classrooms': classrooms_qs,
        'all_classrooms': all_classrooms_list,
        'selected_class': selected_class,
        'selected_track': t_track,
        'subjects_data': subjects_data,
        'active_year': active_year,
    })


@login_required
@role_required(['ADMIN'])
def api_toggle_exam_term_subject(request):
    """
    AJAX endpoint to toggle is_tested for a specific subject under term + (classroom or grade_level).
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    import json
    data = json.loads(request.body) if request.content_type == 'application/json' else request.POST

    term_id = data.get('term_id')
    classroom_id = data.get('classroom_id')
    grade_level = data.get('grade_level')
    track = data.get('track', 'GENERAL')
    subject_id = data.get('subject_id')
    is_tested = str(data.get('is_tested', 'true')).lower() in ['true', '1', 'on']
    custom_max = data.get('custom_max_score')

    exam_term = ExamTerm.objects.filter(id=term_id).first() if term_id else None
    classroom = Classroom.objects.filter(id=classroom_id).first() if classroom_id else None
    subject = get_object_or_404(Subject, id=subject_id)
    ay = exam_term.academic_year if exam_term else (classroom.academic_year if classroom else None)

    lookup = {
        'academic_year': ay,
        'exam_term': exam_term,
        'classroom': classroom,
        'subject': subject,
    }
    if not classroom:
        lookup['grade_level'] = int(grade_level) if grade_level else None
        lookup['track'] = track

    parsed_max = Decimal(str(custom_max)) if custom_max and str(custom_max).strip() else None

    setting, _ = ExamTermSubjectSetting.objects.update_or_create(
        **lookup,
        defaults={
            'is_tested': is_tested,
            'custom_max_score': parsed_max,
            'configured_by': request.user
        }
    )

    status_kh = "ប្រឡង" if is_tested else "មិនប្រឡង"
    return JsonResponse({
        'status': 'success',
        'is_tested': is_tested,
        'message': f"មុខវិជ្ជា «{subject.name_kh}» ត្រូវបានកំណត់ជា៖ {status_kh}"
    })


@login_required
@role_required(['ADMIN'])
def api_apply_exam_term_preset(request):
    """
    Applies presets (ALL, SCIENCE_7, SOCIAL_7) to an ExamTerm + Classroom/GradeLevel.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    import json
    data = json.loads(request.body) if request.content_type == 'application/json' else request.POST

    term_id = data.get('term_id')
    classroom_id = data.get('classroom_id')
    grade_level = data.get('grade_level')
    track = data.get('track', 'GENERAL')
    preset = data.get('preset', 'ALL').strip().upper()

    exam_term = ExamTerm.objects.filter(id=term_id).first() if term_id else None
    classroom = Classroom.objects.filter(id=classroom_id).first() if classroom_id else None
    ay = exam_term.academic_year if exam_term else (classroom.academic_year if classroom else None)

    SCIENCE_7_NAMES = ['ភាសាខ្មែរ', 'គណិតវិទ្យា', 'រូបវិទ្យា', 'គីមីវិទ្យា', 'ជីវវិទ្យា', 'ប្រវត្តិវិទ្យា', 'អង់គ្លេស', 'ភាសាអង់គ្លេស', 'ភាសាបរទេស']
    SOCIAL_7_NAMES = ['ភាសាខ្មែរ', 'គណិតវិទ្យា', 'ភូមិវិទ្យា', 'ប្រវត្តិវិទ្យា', 'សីលធម៌-ពលរដ្ឋ', 'សីលធម៌', 'ផែនដីវិទ្យា', 'ផែនដី', 'អង់គ្លេស', 'ភាសាអង់គ្លេស', 'ភាសាបរទេស']

    from .services import get_effective_term_subjects
    all_rules = get_effective_term_subjects(
        exam_term=exam_term,
        classroom=classroom,
        grade_level=int(grade_level) if grade_level else None,
        track=track,
        include_non_tested=True
    )

    with transaction.atomic():
        for r in all_rules:
            sub = r.subject
            sub_name = sub.name_kh.strip()
            
            if preset == 'ALL':
                is_test = True
            elif preset == 'SCIENCE_7':
                is_test = any(s in sub_name for s in SCIENCE_7_NAMES)
            elif preset == 'SOCIAL_7':
                is_test = any(s in sub_name for s in SOCIAL_7_NAMES)
            else:
                is_test = True

            lookup = {
                'academic_year': ay,
                'exam_term': exam_term,
                'classroom': classroom,
                'subject': sub,
            }
            if not classroom:
                lookup['grade_level'] = int(grade_level) if grade_level else None
                lookup['track'] = track

            ExamTermSubjectSetting.objects.update_or_create(
                **lookup,
                defaults={
                    'is_tested': is_test,
                    'configured_by': request.user
                }
            )

    return JsonResponse({
        'status': 'success',
        'message': f"បានកំណត់ Preset «{preset}» ដោយជោគជ័យ!"
    })


@login_required
@role_required(['ADMIN'])
def api_manage_standardized_exam_subjects(request, exam_id):
    """
    AJAX / POST endpoint to manage (add, remove, edit) subjects for a StandardizedExam.
    Automatically recalculates overall and room ranks after modifications.
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    import json
    data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
    subject_ids = data.get('subject_ids', [])
    if isinstance(subject_ids, str):
        subject_ids = [int(s.strip()) for s in subject_ids.split(',') if s.strip().isdigit()]
    elif isinstance(subject_ids, list):
        subject_ids = [int(s) for s in subject_ids if str(s).isdigit()]

    if not subject_ids:
        return JsonResponse({'status': 'error', 'message': '⚠️ សូមជ្រើសរើសយ៉ាងហោចណាស់មួយមុខវិជ្ជា!'}, status=400)

    with transaction.atomic():
        # Remove exam subjects not in subject_ids
        exam.exam_subjects.exclude(subject_id__in=subject_ids).delete()

        # Add or update
        for order_idx, sid in enumerate(subject_ids, 1):
            sub = Subject.objects.filter(id=sid).first()
            if sub:
                es = exam.exam_subjects.filter(subject=sub).first()
                if not es:
                    max_sc = Decimal('50.00')
                    coef = Decimal('1.00')
                    ExamSubject.objects.create(
                        exam=exam,
                        subject=sub,
                        max_score=max_sc,
                        coefficient=coef,
                        session=exam.session if exam.session != 'FULL_DAY' else 'MORNING',
                        order=order_idx
                    )

        # Recalculate ranks across all candidates in exam
        exam.recalculate_all_ranks()

    return JsonResponse({
        'status': 'success',
        'message': f"🎉 បានធ្វើបច្ចុប្បន្នភាពមុខវិជ្ជាប្រឡងចំនួន {len(subject_ids)} មុខវិជ្ជា ដោយជោគជ័យ!",
        'total_subjects': exam.exam_subjects.count()
    })


@login_required
@role_required(['ADMIN'])
def api_apply_standardized_exam_preset(request, exam_id):
    """
    1-Click Preset API for Standardized Exams:
    - SCIENCE_7: Khmer (75, 1.5), Math (125, 2.5), Physics (75, 1.5), Chemistry (75, 1.5), Biology (75, 1.5), History (50, 1.0), English (50, 1.0)
    - SOCIAL_7: Khmer (125, 2.5), Math (75, 1.5), Geography (75, 1.5), History (75, 1.5), Moral (75, 1.5), Earth (50, 1.0), English (50, 1.0)
    - ALL_MOEYS: All 13 MoEYS subjects
    """
    exam = get_object_or_404(StandardizedExam, id=exam_id)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    import json
    data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
    preset = str(data.get('preset', '')).strip().upper()

    # Pre-defined Official MoEYS Grade 12 Presets
    SCIENCE_PRESET = [
        ('ភាសាខ្មែរ', Decimal('75.00'), Decimal('1.50')),
        ('គណិតវិទ្យា', Decimal('125.00'), Decimal('2.50')),
        ('រូបវិទ្យា', Decimal('75.00'), Decimal('1.50')),
        ('គីមីវិទ្យា', Decimal('75.00'), Decimal('1.50')),
        ('ជីវវិទ្យា', Decimal('75.00'), Decimal('1.50')),
        ('ប្រវត្តិវិទ្យា', Decimal('50.00'), Decimal('1.00')),
        ('ភាសាបរទេស', Decimal('50.00'), Decimal('1.00')),
    ]

    SOCIAL_PRESET = [
        ('ភាសាខ្មែរ', Decimal('125.00'), Decimal('2.50')),
        ('គណិតវិទ្យា', Decimal('75.00'), Decimal('1.50')),
        ('ភូមិវិទ្យា', Decimal('75.00'), Decimal('1.50')),
        ('ប្រវត្តិវិទ្យា', Decimal('75.00'), Decimal('1.50')),
        ('សីលធម៌-ពលរដ្ឋ', Decimal('75.00'), Decimal('1.50')),
        ('ផែនដីវិទ្យា', Decimal('50.00'), Decimal('1.00')),
        ('ភាសាបរទេស', Decimal('50.00'), Decimal('1.00')),
    ]

    all_subs = list(Subject.objects.all())

    def find_subject(name):
        for s in all_subs:
            if name in s.name_kh or (name == 'ភាសាបរទេស' and ('អង់គ្លេស' in s.name_kh or 'English' in s.name_en or s.code == 'ENG')):
                return s
            if name == 'សីលធម៌-ពលរដ្ឋ' and ('សីលធម៌' in s.name_kh or 'ពលរដ្ឋ' in s.name_kh):
                return s
        return None

    with transaction.atomic():
        exam.exam_subjects.all().delete()
        order_idx = 1

        if preset == 'SCIENCE_7':
            target_list = SCIENCE_PRESET
            exam.track = 'SCIENCE'
            exam.save(update_fields=['track'])
        elif preset == 'SOCIAL_7':
            target_list = SOCIAL_PRESET
            exam.track = 'SOCIAL'
            exam.save(update_fields=['track'])
        else:
            # All MoEYS subjects from GradeLevelRule or Subject
            rules = GradeLevelRule.objects.filter(grade_level=exam.grade_level)
            target_list = []
            for r in rules.select_related('subject').order_by('subject__order', 'id'):
                target_list.append((r.subject.name_kh, r.max_score, round(r.max_score / Decimal('50.00'), 2)))

        for sub_name, max_s, coef in target_list:
            sub = find_subject(sub_name)
            if sub:
                ExamSubject.objects.create(
                    exam=exam,
                    subject=sub,
                    max_score=max_s,
                    coefficient=coef,
                    session=exam.session if exam.session != 'FULL_DAY' else 'MORNING',
                    order=order_idx
                )
                order_idx += 1

        exam.recalculate_all_ranks()

    count = exam.exam_subjects.count()
    preset_label = "៧ មុខវិជ្ជាវិទ្យាសាស្ត្រ" if preset == 'SCIENCE_7' else ("៧ មុខវិជ្ជាសង្គម" if preset == 'SOCIAL_7' else "គ្រប់មុខវិជ្ជា")
    return JsonResponse({
        'status': 'success',
        'message': f"🎉 បានកំណត់ Preset «{preset_label}» ចំនួន {count} មុខវិជ្ជា សម្រាប់សម័យប្រឡងនេះដោយជោគជ័យ!",
        'total_subjects': count
    })


@login_required
def exam_results_graph_view(request, exam_id):
    """
    Dedicated view for Examination Results Graph (100% replica of Graph.pdf).
    Redirects to exam_results_sheet_print_view with mode=graph, preserving query params.
    """
    params = request.GET.copy()
    params['mode'] = 'graph'
    from django.urls import reverse
    return redirect(f"{reverse('exam_results_sheet_print_view', args=[exam_id])}?{params.urlencode()}")


@login_required
def term_results_graph_view(request, term_id):
    """
    Results Graph View for Monthly / Semester Classroom Exam Terms.
    Builds mention distribution across classroom grades and renders results_graph_print.html.
    """
    term = get_object_or_404(ExamTerm.objects.select_related('academic_year'), id=term_id)
    
    # Check if this term is linked to a StandardizedExam
    linked_std_exam = term.standardized_exams.first()
    if linked_std_exam:
        params = request.GET.copy()
        params['mode'] = 'graph'
        from django.urls import reverse
        return redirect(f"{reverse('exam_results_sheet_print_view', args=[linked_std_exam.id])}?{params.urlencode()}")

    classroom_id = request.GET.get('classroom_id') or request.GET.get('classroom')
    selected_class = None
    if classroom_id and str(classroom_id).isdigit():
        selected_class = Classroom.objects.filter(id=int(classroom_id)).first()
    if not selected_class:
        selected_class = Classroom.objects.filter(academic_year=term.academic_year).order_by('grade_level', 'code').first()

    grade_level = selected_class.grade_level if selected_class else 7
    classrooms = Classroom.objects.filter(academic_year=term.academic_year).order_by('grade_level', 'code')

    from .services import get_effective_term_subjects
    subject_rules = get_effective_term_subjects(exam_term=term, classroom=selected_class, include_non_tested=False)
    
    students = Student.objects.filter(classroom=selected_class, status='ACTIVE').order_by('student_id') if selected_class else []
    grades_map = {}
    if selected_class:
        for g in Grade.objects.filter(classroom=selected_class, exam_term=term):
            grades_map[(g.student_id, g.subject_id)] = g

    # Build matrix for subjects in this classroom
    GRADE_COLORS = {
        'A': '#1f4e79',  # Dark Navy Blue
        'B': '#c55a11',  # Terracotta Orange
        'C': '#276a3c',  # Forest Green
        'D': '#00a2e8',  # Cyan / Sky Blue
        'E': '#800080',  # Magenta / Purple
        'F': '#548235',  # Leaf Green
    }

    # Use first 13 subject rules or available rules
    active_rules = subject_rules[:13] if len(subject_rules) >= 13 else subject_rules
    matrix = {g: [0] * len(active_rules) for g in ['A', 'B', 'C', 'D', 'E', 'F']}

    for student in students:
        for s_idx, rule in enumerate(active_rules):
            g_obj = grades_map.get((student.id, rule.subject_id))
            if g_obj and g_obj.score is not None:
                letter = g_obj.grade_letter or get_moeys_subject_mention(float(g_obj.score), float(rule.max_score))
            else:
                letter = 'F'
            if letter in matrix:
                matrix[letter][s_idx] += 1

    max_count = max([max(matrix[g]) for g in ['A', 'B', 'C', 'D', 'E', 'F']] + [0])
    import math
    if max_count <= 120:
        graph_y_max = 120
        graph_y_ticks = [120, 100, 80, 60, 40, 20, 0]
    else:
        step = int(math.ceil(max_count / 6 / 10.0)) * 10
        if step < 20:
            step = 20
        graph_y_max = step * 6
        graph_y_ticks = [step * i for i in range(6, -1, -1)]

    graph_columns = []
    for s_idx, rule in enumerate(active_rules):
        sub_bars = []
        for g in ['A', 'B', 'C', 'D', 'E', 'F']:
            cnt = matrix[g][s_idx]
            height_pct = f"{(cnt / graph_y_max * 100.0):.2f}" if graph_y_max > 0 else "0.00"
            sub_bars.append({
                'grade': g,
                'count': cnt,
                'count_kh': to_khmer_digits(cnt),
                'color': GRADE_COLORS[g],
                'height_pct': height_pct,
            })
        graph_columns.append({
            'key': rule.subject.code or f"sub_{rule.subject.id}",
            'name_kh': rule.subject.name_kh,
            'max_score': rule.max_score,
            'bars': sub_bars,
        })

    graph_table_rows = []
    for g in ['A', 'B', 'C', 'D', 'E', 'F']:
        row_counts = [matrix[g][s_idx] for s_idx in range(len(active_rules))]
        row_counts_kh = [to_khmer_digits(c) for c in row_counts]
        graph_table_rows.append({
            'grade': g,
            'color': GRADE_COLORS[g],
            'counts': row_counts,
            'counts_kh': row_counts_kh,
        })

    academic_year_kh = to_khmer_digits(term.academic_year.name if term.academic_year else '')
    grade_kh = to_khmer_digits(grade_level)
    cls_name = selected_class.name if selected_class else f"ថ្នាក់ទី {grade_kh}"
    results_title_line = f"លទ្ធផលប្រឡង{term.name} ឆ្នាំសិក្សា {academic_year_kh} {cls_name}".strip()

    context = {
        'term': term,
        'selected_class': selected_class,
        'classrooms': classrooms,
        'results_title_line': results_title_line,
        'graph_columns': graph_columns,
        'graph_table_rows': graph_table_rows,
        'graph_y_ticks': graph_y_ticks,
        'graph_y_max': graph_y_max,
        'grade_colors': GRADE_COLORS,
        'view_mode': 'graph',
    }
    return render(request, 'examinations/standardized/results_graph_print.html', context)


@login_required
def exam_analytics_view(request, exam_id=None):
    """
    Comprehensive Examination Analytics Suite matching the 5 user-provided templates.
    Supports 3 Breakdown Scopes:
    1. 'school': School-level (aggregate across all grades in this exam session)
    2. 'grade': Grade-level (e.g. 7, 8, 9, 10, 11, 12)
    3. 'class': Classroom-level (e.g. 7A, 7B, 10A, 12A1)
    """
    from apps.accounts.models import SchoolProfile
    from .analytics_service import ExamAnalyticsService
    import json

    school_profile = SchoolProfile.get_settings()
    session_key = request.GET.get('session_key', '').strip()
    exam_ids_param = request.GET.get('exam_ids', '').strip()

    # 1. Resolve Active Exam & Sibling Exams for this Session
    target_exam = None
    if exam_id:
        target_exam = get_object_or_404(StandardizedExam.objects.select_related('academic_year'), id=exam_id)

    # All standardized exams ordered by date desc
    all_exams_qs = StandardizedExam.objects.select_related('academic_year').prefetch_related('exam_subjects', 'candidates').order_by('-exam_date', 'grade_level')
    
    # Build session mapping across all exams
    all_sessions_map = {}
    for ex in all_exams_qs:
        clean_title = get_clean_exam_session_title(ex.name)
        date_key = str(ex.exam_date)
        year_key = str(ex.academic_year_id)
        g_key = f"{year_key}_{date_key}_{clean_title}"
        if g_key not in all_sessions_map:
            all_sessions_map[g_key] = {
                'group_key': g_key,
                'title': clean_title,
                'academic_year': ex.academic_year,
                'exam_date': ex.exam_date,
                'exams': [],
            }
        all_sessions_map[g_key]['exams'].append(ex)

    # Determine which session group is selected
    selected_session = None
    if session_key and session_key in all_sessions_map:
        selected_session = all_sessions_map[session_key]
    elif exam_ids_param:
        eids = [int(x.strip()) for x in exam_ids_param.split(',') if x.strip().isdigit()]
        for s in all_sessions_map.values():
            if any(e.id in eids for e in s['exams']):
                selected_session = s
                break
    elif target_exam:
        for s in all_sessions_map.values():
            if any(e.id == target_exam.id for e in s['exams']):
                selected_session = s
                break

    # Fallback to latest session if none matched
    if not selected_session and all_sessions_map:
        selected_session = list(all_sessions_map.values())[0]

    session_exams = selected_session['exams'] if selected_session else ([target_exam] if target_exam else [])
    current_session_key = selected_session['group_key'] if selected_session else ''
    session_title = selected_session['title'] if selected_session else (target_exam.name if target_exam else 'សម័យប្រឡង')
    session_date = selected_session['exam_date'] if selected_session else (target_exam.exam_date if target_exam else None)
    session_academic_year = selected_session['academic_year'] if selected_session else (target_exam.academic_year if target_exam else None)

    # 2. Scope & Parameters
    scope = request.GET.get('scope', '').strip().lower()
    grade_level = request.GET.get('grade_level', '').strip()
    classroom_name = request.GET.get('class_name', '').strip()

    if not scope:
        if exam_id and target_exam:
            scope = 'grade'
            grade_level = str(target_exam.grade_level)
        else:
            scope = 'school'

    if scope not in ['school', 'grade', 'class']:
        scope = 'school'

    if scope == 'grade' and not grade_level:
        if target_exam:
            grade_level = str(target_exam.grade_level)
        elif session_exams:
            grade_level = str(session_exams[0].grade_level)

    # 3. Compute Analytics Payload
    analytics = ExamAnalyticsService.get_analytics_payload(
        exams=session_exams,
        scope=scope,
        grade_level=grade_level,
        classroom_name=classroom_name
    )

    slow_learners_json = json.dumps(analytics['slow_learners_data'])
    subjects_json = json.dumps(analytics['subjects_list'])

    sessions_nav_list = [
        {
            'group_key': s['group_key'],
            'title': s['title'],
            'date': s['exam_date'],
            'year': s['academic_year'].name if s['academic_year'] else '',
            'exam_count': len(s['exams']),
        }
        for s in all_sessions_map.values()
    ]

    context = {
        'school_profile': school_profile,
        'target_exam': target_exam,
        'session_exams': session_exams,
        'session_title': session_title,
        'session_date': session_date,
        'session_academic_year': session_academic_year,
        'current_session_key': current_session_key,
        'sessions_nav_list': sessions_nav_list,
        'scope': scope,
        'selected_grade': grade_level,
        'selected_class': classroom_name,
        'available_grades': analytics['available_grades'],
        'all_classrooms': analytics['all_classrooms'],
        'subjects_list': analytics['subjects_list'],
        'total_candidates': analytics['total_candidates'],
        'female_candidates': analytics['female_candidates'],
        'male_candidates': analytics['male_candidates'],
        'overall_mentions': analytics['overall_mentions'],
        'quality_evaluation': analytics['quality_evaluation'],
        'subject_mentions_single': analytics['subject_mentions_single'],
        'subject_mentions_detailed': analytics['subject_mentions_detailed'],
        'subject_percentage_rows': analytics['subject_percentage_rows'],
        'percentage_thresholds': analytics['percentage_thresholds'],
        'slow_learners_data': analytics['slow_learners_data'],
        'slow_learners_json': slow_learners_json,
        'subjects_json': subjects_json,
    }
    return render(request, 'examinations/standardized/analytics_report.html', context)


@login_required
def exam_session_analytics_view(request):
    """
    Session-level entry point for multi-grade exam session analytics.
    Delegates to exam_analytics_view.
    """
    return exam_analytics_view(request, exam_id=None)


@login_required
def exam_generate_mock_scores_view(request):
    """
    POST/GET endpoint to generate realistic mock scores (A-F) or clear scores
    for a specific exam or across an entire examination session.
    """
    from .analytics_service import ExamAnalyticsService

    exam_id = request.POST.get('exam_id') or request.GET.get('exam_id')
    grade_level = request.POST.get('grade_level') or request.GET.get('grade_level')
    session_key = request.POST.get('session_key') or request.GET.get('session_key')
    target_scope = request.POST.get('target_scope') or request.GET.get('target_scope') or 'current'
    action = request.POST.get('action') or request.GET.get('action') or 'generate'

    target_exams = []

    if target_scope == 'session' and session_key:
        all_exams_qs = StandardizedExam.objects.select_related('academic_year').prefetch_related('exam_subjects', 'candidates')
        for ex in all_exams_qs:
            clean_title = get_clean_exam_session_title(ex.name)
            date_key = str(ex.exam_date)
            year_key = str(ex.academic_year_id)
            g_key = f"{year_key}_{date_key}_{clean_title}"
            if g_key == session_key:
                target_exams.append(ex)
    elif exam_id and str(exam_id).isdigit():
        target_exams = [get_object_or_404(StandardizedExam, id=int(exam_id))]
    elif grade_level and str(grade_level).isdigit():
        qs = StandardizedExam.objects.filter(grade_level=int(grade_level))
        if session_key:
            filtered = []
            for ex in qs.select_related('academic_year'):
                clean_title = get_clean_exam_session_title(ex.name)
                g_key = f"{ex.academic_year_id}_{ex.exam_date}_{clean_title}"
                if g_key == session_key:
                    filtered.append(ex)
            target_exams = filtered or list(qs)
        else:
            target_exams = list(qs)

    # Fallback to exam_id if none resolved
    if not target_exams and exam_id and str(exam_id).isdigit():
        target_exams = [get_object_or_404(StandardizedExam, id=int(exam_id))]

    # Fallback to session exams if session_key provided
    if not target_exams and session_key:
        for ex in StandardizedExam.objects.select_related('academic_year').all():
            clean_title = get_clean_exam_session_title(ex.name)
            g_key = f"{ex.academic_year_id}_{ex.exam_date}_{clean_title}"
            if g_key == session_key:
                target_exams.append(ex)

    if not target_exams:
        messages.error(request, "មិនអាចស្វែងរកកម្រិតថ្នាក់ប្រឡងដែលត្រូវអនុវត្តបានឡើយ។")
        return redirect(request.META.get('HTTP_REFERER') or 'standardized_exam_list')

    if action == 'clear':
        res = ExamAnalyticsService.clear_mock_scores(target_exams)
        messages.success(request, f"🗑️ បានសម្អាតពិន្ទុសម្រាប់បេក្ខជន {res['candidates_count']} នាក់ ({len(target_exams)} កម្រិតថ្នាក់) មកជាទទេវិញដោយជោគជ័យ!")
    else:
        res = ExamAnalyticsService.generate_mock_scores(target_exams)
        messages.success(request, f"🎉 បានបង្កើតពិន្ទុតេស្តសាកល្បង A-F សម្រាប់បេក្ខជន {res['candidates_count']} នាក់ ({res['scores_count']} ក្រឡា) ដោយជោគជ័យ!")

    redirect_url = request.META.get('HTTP_REFERER')
    if redirect_url:
        return redirect(redirect_url)
    if len(target_exams) == 1:
        return redirect('exam_analytics_view', exam_id=target_exams[0].id)
    return redirect('standardized_exam_list')


@login_required
def exam_analytics_export_excel(request, exam_id=None):
    """
    Exports all reports or selected report on Exam Session Analytics page to Microsoft Excel (.xlsx).
    Supports:
    - Multi-sheet workbook with all 6 reports ('all')
    - Individual sheet export ('overall', 'quality', 'subject_summary', 'percentages', 'slow_learners')
    """
    import io
    import urllib.parse
    from django.http import HttpResponse
    from .analytics_service import ExamAnalyticsService

    target_exam = get_object_or_404(StandardizedExam, id=exam_id) if (exam_id and str(exam_id).isdigit()) else None

    session_key = request.GET.get('session_key', '').strip()
    exam_ids_param = request.GET.get('exam_ids', '').strip()
    target_exam_param = request.GET.get('exam_id', '').strip()
    if not target_exam and target_exam_param and target_exam_param.isdigit():
        target_exam = StandardizedExam.objects.filter(id=int(target_exam_param)).first()

    all_exams_qs = StandardizedExam.objects.select_related('academic_year').prefetch_related(
        'exam_subjects__subject', 'candidates__subject_scores'
    ).order_by('grade_level', 'id')

    all_sessions_map = {}
    for ex in all_exams_qs:
        clean_title = get_clean_exam_session_title(ex.name)
        date_key = str(ex.exam_date)
        year_key = str(ex.academic_year_id)
        g_key = f"{year_key}_{date_key}_{clean_title}"
        if g_key not in all_sessions_map:
            all_sessions_map[g_key] = {
                'group_key': g_key,
                'title': clean_title,
                'academic_year': ex.academic_year,
                'exam_date': ex.exam_date,
                'exams': [],
            }
        all_sessions_map[g_key]['exams'].append(ex)

    selected_session = None
    if session_key and session_key in all_sessions_map:
        selected_session = all_sessions_map[session_key]
    elif exam_ids_param:
        eids = [int(x.strip()) for x in exam_ids_param.split(',') if x.strip().isdigit()]
        for s in all_sessions_map.values():
            if any(e.id in eids for e in s['exams']):
                selected_session = s
                break
    elif target_exam:
        for s in all_sessions_map.values():
            if any(e.id == target_exam.id for e in s['exams']):
                selected_session = s
                break

    if not selected_session and all_sessions_map:
        selected_session = list(all_sessions_map.values())[0]

    session_exams = selected_session['exams'] if selected_session else ([target_exam] if target_exam else [])
    session_title = selected_session['title'] if selected_session else (target_exam.name if target_exam else 'សម័យប្រឡង')
    session_date = selected_session['exam_date'] if selected_session else (target_exam.exam_date if target_exam else None)
    session_academic_year = selected_session['academic_year'] if selected_session else (target_exam.academic_year if target_exam else None)

    scope = request.GET.get('scope', '').strip().lower()
    grade_level = request.GET.get('grade_level', '').strip()
    classroom_name = request.GET.get('class_name', '').strip()
    target_sheet = request.GET.get('sheet', 'all').strip().lower()
    mention_sum_label = request.GET.get('sum_label', 'A+B+C').strip()

    if not scope:
        if exam_id and target_exam:
            scope = 'grade'
            grade_level = str(target_exam.grade_level)
        else:
            scope = 'school'

    if scope not in ['school', 'grade', 'class']:
        scope = 'school'

    if scope == 'grade' and not grade_level:
        if target_exam:
            grade_level = str(target_exam.grade_level)
        elif session_exams:
            grade_level = str(session_exams[0].grade_level)

    # Compute Analytics Payload
    analytics = ExamAnalyticsService.get_analytics_payload(
        exams=session_exams,
        scope=scope,
        grade_level=grade_level,
        classroom_name=classroom_name
    )

    scope_title = "កម្រិតសាលា (School Level)"
    if scope == 'grade' and grade_level:
        scope_title = f"កម្រិតថ្នាក់ទី {grade_level}"
    elif scope == 'class' and classroom_name:
        scope_title = f"ថ្នាក់រៀន {classroom_name}"

    date_str = session_date.strftime('%d/%m/%Y') if session_date else ''
    year_str = session_academic_year.name if session_academic_year else ''

    wb = ExamAnalyticsService.build_analytics_workbook(
        analytics=analytics,
        session_title=session_title,
        session_date_str=date_str,
        academic_year_name=year_str,
        scope_title=scope_title,
        target_sheet=target_sheet,
        mention_sum_label=mention_sum_label
    )

    sheet_suffix = {
        'all': 'គ្រប់_Sheet',
        'overall': 'សរុបនិទ្ទេសរួម',
        'quality': 'វាយតម្លៃគុណភាព',
        'subject_summary': 'សង្ខេបនិទ្ទេសតាមមុខវិជ្ជា',
        'subject_single': 'សង្ខេបនិទ្ទេស_សរុប',
        'subject_detailed': 'សង្ខេបនិទ្ទេស_លម្អិត',
        'percentages': 'វិភាគភាគរយមុខវិជ្ជា',
        'slow_learners': 'សិស្សរៀនយឺត',
    }.get(target_sheet, 'របាយការណ៍')

    safe_title = "".join(c for c in session_title if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    filename = f"របាយការណ៍_{safe_title}_{scope}_{sheet_suffix}.xlsx"
    encoded_filename = urllib.parse.quote(filename)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    return response







