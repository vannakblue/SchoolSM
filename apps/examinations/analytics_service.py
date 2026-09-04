from decimal import Decimal
from django.db.models import Q
from .models import StandardizedExam, ExamCandidate, ExamSubject, CandidateSubjectScore


def get_subject_mention(score, max_score):
    """
    Computes official MoEYS subject letter grade (A..F):
    - A: >= 90%
    - B: >= 80% and < 90%
    - C: >= 70% and < 80%
    - D: >= 60% and < 70%
    - E: >= 50% and < 60%
    - F: < 50%
    """
    if score is None:
        return '-'
    try:
        s = float(score)
    except (ValueError, TypeError):
        return '-'
    m = float(max_score) if max_score and float(max_score) > 0 else 50.0
    pct = (s / m) * 100.0
    if pct >= 90.0:
        return 'A'
    if pct >= 80.0:
        return 'B'
    if pct >= 70.0:
        return 'C'
    if pct >= 60.0:
        return 'D'
    if pct >= 50.0:
        return 'E'
    return 'F'


def get_overall_mention(total_score, max_total_score):
    """
    Computes overall exam letter grade (A..F) based on percentage of total score.
    """
    if total_score is None:
        return '-'
    try:
        s = float(total_score)
    except (ValueError, TypeError):
        return '-'
    m = float(max_total_score) if max_total_score and float(max_total_score) > 0 else 100.0
    if m <= 0:
        return '-'
    pct = (s / m) * 100.0
    if pct >= 90.0:
        return 'A'
    if pct >= 80.0:
        return 'B'
    if pct >= 70.0:
        return 'C'
    if pct >= 60.0:
        return 'D'
    if pct >= 50.0:
        return 'E'
    return 'F'


class ExamAnalyticsService:
    """
    Comprehensive Analytics Service for Standardized Examination Sessions
    Reproduces reports across 3 breakdown scopes:
      1. 'school': Across all grades in the examination session
      2. 'grade': Filtered by specific grade level (e.g. 7, 8, 9, 10, 11, 12)
      3. 'class': Filtered by specific classroom (e.g. 7A, 10A, 12A1)
    """

    @classmethod
    def get_analytics_payload(cls, exams, scope='school', grade_level=None, classroom_name=None):
        """
        Main calculation engine returning full data structures for all 4 report components:
        - overall_mentions (Image 2 - Left)
        - quality_evaluation (Image 2 - Right)
        - subject_mentions (Images 3 & 4 - Single and Detailed views)
        - subject_percentages (Image 5)
        - slow_learners (Image 1 - Candidates & scores table)
        """
        if not hasattr(exams, '__iter__'):
            exams = [exams]
        else:
            exams = list(exams)

        if not exams:
            return cls._empty_payload()

        # Apply Scope Filtering to Exams
        if scope == 'grade' and grade_level:
            try:
                g_int = int(grade_level)
                active_exams = [e for e in exams if e.grade_level == g_int]
                if not active_exams:
                    active_exams = exams
            except (ValueError, TypeError):
                active_exams = exams
        else:
            active_exams = exams

        exam_ids = [e.id for e in active_exams]

        # Fetch Candidates with scores preloaded
        candidates_qs = ExamCandidate.objects.filter(
            exam_id__in=exam_ids
        ).select_related('exam', 'student', 'room').prefetch_related(
            'subject_scores__exam_subject__subject'
        ).order_by('exam__grade_level', 'origin_class', 'roll_number', 'id')

        if scope == 'class' and classroom_name:
            candidates_qs = candidates_qs.filter(origin_class__iexact=str(classroom_name).strip())

        candidates = list(candidates_qs)

        # Collect and deduplicate subjects in logical display order
        # Map: subject_id -> {'name': subject_name, 'max_score': max_score, 'order': order}
        subjects_map = {}
        for exam in active_exams:
            for es in exam.exam_subjects.select_related('subject').order_by('order', 'id'):
                s_id = es.subject_id
                if s_id not in subjects_map:
                    subjects_map[s_id] = {
                        'id': s_id,
                        'code': es.subject.code or f"sub_{s_id}",
                        'name': es.subject.name_kh or es.subject.name_en,
                        'short': es.subject.name_kh[:4] if es.subject.name_kh else es.subject.name_en[:4],
                        'max_score': float(es.max_score),
                        'order': es.order,
                    }
                else:
                    # Update max score if larger
                    if float(es.max_score) > subjects_map[s_id]['max_score']:
                        subjects_map[s_id]['max_score'] = float(es.max_score)

        subjects_list = sorted(subjects_map.values(), key=lambda s: (s['order'], s['id']))

        # Process Candidates: Compute individual candidate subject scores, mentions, and overall total
        processed_candidates = []
        for cand in candidates:
            # Build quick score lookup for candidate: subject_id -> score_obj
            c_scores = {}
            for sc in cand.subject_scores.all():
                if sc.exam_subject and sc.exam_subject.subject_id:
                    c_scores[sc.exam_subject.subject_id] = sc

            cand_subject_data = {}
            cand_total_score = 0.0
            cand_max_total = 0.0
            has_any_score = False

            for s in subjects_list:
                s_id = s['id']
                sc_obj = c_scores.get(s_id)
                score_val = None
                is_absent = False
                if sc_obj:
                    is_absent = sc_obj.is_absent
                    if sc_obj.score is not None and not is_absent:
                        score_val = float(sc_obj.score)
                        cand_total_score += score_val
                        has_any_score = True

                mention = get_subject_mention(score_val, s['max_score'])
                cand_max_total += s['max_score']

                cand_subject_data[s_id] = {
                    'score': score_val,
                    'is_absent': is_absent,
                    'mention': mention,
                    'is_weak': mention in ['E', 'F'] or (score_val is not None and score_val < (s['max_score'] * 0.5)),
                    'max_score': s['max_score'],
                }

            # If candidate total_score was already stored on model, use it or computed sum
            final_tot = float(cand.total_score) if (cand.total_score and float(cand.total_score) > 0) else cand_total_score
            overall_m = cand.grade_letter if (cand.grade_letter and cand.grade_letter in ['A', 'B', 'C', 'D', 'E', 'F']) else get_overall_mention(final_tot, cand_max_total)

            # Overall percentage
            pct = (final_tot / cand_max_total * 100.0) if cand_max_total > 0 and has_any_score else 0.0

            processed_candidates.append({
                'id': cand.id,
                'student_code': cand.student_code or (cand.student.student_id if cand.student else '') or cand.roll_number,
                'candidate_name_kh': cand.candidate_name_kh,
                'gender': cand.gender,
                'gender_kh': 'ស្រី' if cand.gender == 'F' else 'ប្រុស',
                'gender_short': 'ស' if cand.gender == 'F' else 'ប',
                'grade_level': cand.exam.grade_level,
                'origin_class': cand.origin_class or f"ថ្នាក់ទី{cand.exam.grade_level}",
                'total_score': round(final_tot, 2),
                'overall_mention': overall_m,
                'percentage': round(pct, 1),
                'subjects': cand_subject_data,
            })

        total_candidates_count = len(processed_candidates)
        female_candidates_count = sum(1 for c in processed_candidates if c['gender'] == 'F')
        male_candidates_count = total_candidates_count - female_candidates_count

        # -------------------------------------------------------------
        # 1. Overall Mentions Matrix (Image 2 - Left)
        # -------------------------------------------------------------
        # Mentions: A, B, C, D, E, F, A+B+C, Total
        grades_list = ['A', 'B', 'C', 'D', 'E', 'F']
        
        counts_total = {g: 0 for g in grades_list}
        counts_female = {g: 0 for g in grades_list}
        counts_male = {g: 0 for g in grades_list}

        for c in processed_candidates:
            m = c['overall_mention']
            if m in counts_total:
                counts_total[m] += 1
                if c['gender'] == 'F':
                    counts_female[m] += 1
                else:
                    counts_male[m] += 1

        sum_abc_total = counts_total['A'] + counts_total['B'] + counts_total['C']
        sum_abc_female = counts_female['A'] + counts_female['B'] + counts_female['C']
        sum_abc_male = counts_male['A'] + counts_male['B'] + counts_male['C']

        overall_mentions = {
            'grades': grades_list,
            'total_row': {
                'label': 'សរុប',
                'counts': [counts_total[g] for g in grades_list],
                'sum_abc': sum_abc_total,
                'grand_total': total_candidates_count,
            },
            'female_row': {
                'label': 'ស្រី',
                'counts': [counts_female[g] for g in grades_list],
                'sum_abc': sum_abc_female,
                'grand_total': female_candidates_count,
            },
            'male_row': {
                'label': 'ប្រុស',
                'counts': [counts_male[g] for g in grades_list],
                'sum_abc': sum_abc_male,
                'grand_total': male_candidates_count,
            }
        }

        # -------------------------------------------------------------
        # 2. Quality Evaluation Matrix (Image 2 - Right)
        # -------------------------------------------------------------
        # Tiers:
        # - ល្អ (>=40 or >=80%): matches A + B
        # - ល្អបង្គួរ (>=32.5 or 65%..79.9%): matches C + high D
        # - មធ្យម (>=25 or 50%..64.9%): matches low D + E
        # - ខ្សោយ (<25 or <50%): matches F
        quality_counts = {
            'good': {'total': 0, 'female': 0, 'male': 0},         # >=80%
            'fairly_good': {'total': 0, 'female': 0, 'male': 0},  # 65% - 79.9%
            'average': {'total': 0, 'female': 0, 'male': 0},      # 50% - 64.9%
            'weak': {'total': 0, 'female': 0, 'male': 0},         # < 50%
        }

        for c in processed_candidates:
            pct = c['percentage']
            gen = 'female' if c['gender'] == 'F' else 'male'
            
            if pct >= 80.0:
                tier = 'good'
            elif pct >= 65.0:
                tier = 'fairly_good'
            elif pct >= 50.0:
                tier = 'average'
            else:
                tier = 'weak'

            quality_counts[tier]['total'] += 1
            quality_counts[tier][gen] += 1

        quality_evaluation = {
            'headers': [
                {'key': 'good', 'label': 'ល្អ (>=40)', 'sub': '>=80%', 'class': 'text-primary'},
                {'key': 'fairly_good', 'label': 'ល្អបង្គួរ (>=32.5)', 'sub': '>=65%', 'class': 'text-success'},
                {'key': 'average', 'label': 'មធ្យម (>=25)', 'sub': '>=50%', 'class': 'text-warning text-dark'},
                {'key': 'weak', 'label': 'ខ្សោយ (<25)', 'sub': '<50%', 'class': 'text-danger'},
                {'key': 'total', 'label': 'សរុប', 'sub': '', 'class': 'text-dark'},
            ],
            'total_row': {
                'label': 'សរុប',
                'good': quality_counts['good']['total'],
                'fairly_good': quality_counts['fairly_good']['total'],
                'average': quality_counts['average']['total'],
                'weak': quality_counts['weak']['total'],
                'grand_total': total_candidates_count,
            },
            'female_row': {
                'label': 'ស្រី',
                'good': quality_counts['good']['female'],
                'fairly_good': quality_counts['fairly_good']['female'],
                'average': quality_counts['average']['female'],
                'weak': quality_counts['weak']['female'],
                'grand_total': female_candidates_count,
            },
            'male_row': {
                'label': 'ប្រុស',
                'good': quality_counts['good']['male'],
                'fairly_good': quality_counts['fairly_good']['male'],
                'average': quality_counts['average']['male'],
                'weak': quality_counts['weak']['male'],
                'grand_total': male_candidates_count,
            },
        }

        # -------------------------------------------------------------
        # 3. Subject Mentions Matrix (Images 3 & 4)
        # -------------------------------------------------------------
        # For each subject:
        # Counts of A, B, C, D, E, F, A+B+C, Total
        # Single mode: Total counts per subject
        # Detailed mode: Total, Female, Male counts for each grade
        subject_mentions_single = []
        subject_mentions_detailed = []

        for s in subjects_list:
            s_id = s['id']
            max_s = int(s['max_score']) if s['max_score'] == int(s['max_score']) else s['max_score']
            subj_display_title = f"{s['name']} ({max_s})"

            # Count per grade
            s_counts = {g: {'total': 0, 'female': 0, 'male': 0} for g in grades_list}
            
            for c in processed_candidates:
                s_data = c['subjects'].get(s_id)
                if s_data and s_data['mention'] in s_counts:
                    m = s_data['mention']
                    s_counts[m]['total'] += 1
                    if c['gender'] == 'F':
                        s_counts[m]['female'] += 1
                    else:
                        s_counts[m]['male'] += 1

            # Calculations
            abc_tot = s_counts['A']['total'] + s_counts['B']['total'] + s_counts['C']['total']
            abc_fem = s_counts['A']['female'] + s_counts['B']['female'] + s_counts['C']['female']
            abc_mal = s_counts['A']['male'] + s_counts['B']['male'] + s_counts['C']['male']

            total_tot = sum(s_counts[g]['total'] for g in grades_list)
            total_fem = sum(s_counts[g]['female'] for g in grades_list)
            total_mal = sum(s_counts[g]['male'] for g in grades_list)

            # Single row structure (Image 3)
            subject_mentions_single.append({
                'subject_id': s_id,
                'title': subj_display_title,
                'name': s['name'],
                'max_score': s['max_score'],
                'a': s_counts['A']['total'],
                'b': s_counts['B']['total'],
                'c': s_counts['C']['total'],
                'd': s_counts['D']['total'],
                'e': s_counts['E']['total'],
                'f': s_counts['F']['total'],
                'sum_abc': abc_tot,
                'total': total_tot,
            })

            # Detailed row structure (Image 4)
            subject_mentions_detailed.append({
                'subject_id': s_id,
                'title': subj_display_title,
                'name': s['name'],
                'max_score': s['max_score'],
                'grades': {
                    g: {
                        'tot': s_counts[g]['total'],
                        'fem': s_counts[g]['female'],
                        'mal': s_counts[g]['male']
                    } for g in grades_list
                },
                'sum_abc': {
                    'tot': abc_tot,
                    'fem': abc_fem,
                    'mal': abc_mal
                },
                'total': {
                    'tot': total_tot,
                    'fem': total_fem,
                    'mal': total_mal
                }
            })

        # -------------------------------------------------------------
        # 4. Subject Percentage Matrix (Image 5)
        # -------------------------------------------------------------
        # Thresholds:
        # Group 1 (>=): 95%, 90%, 85%, 80%, 75%, 70%, 65%, 60%, 55%, 50%
        # Group 2 (Between): 60-80%, 50-<80%
        # Group 3 Weak (<): 60%, 50% (highlighted in red)
        percentage_thresholds = [95, 90, 85, 80, 75, 70, 65, 60, 55, 50]
        subject_percentage_rows = []

        for s in subjects_list:
            s_id = s['id']
            max_s = int(s['max_score']) if s['max_score'] == int(s['max_score']) else s['max_score']
            subj_display_title = f"{s['name']} ({max_s})"

            # Collect percentages for all candidates who took this subject
            candidate_pcts = []
            for c in processed_candidates:
                s_data = c['subjects'].get(s_id)
                if s_data and s_data['score'] is not None and not s_data['is_absent']:
                    pct = (s_data['score'] / s['max_score']) * 100.0
                    candidate_pcts.append(pct)

            # Calculate >= thresholds
            gte_counts = []
            for th in percentage_thresholds:
                cnt = sum(1 for p in candidate_pcts if p >= th)
                gte_counts.append(cnt if cnt > 0 else '-')

            # Calculate Ranges: 60-80%, 50-<80%
            between_60_80 = sum(1 for p in candidate_pcts if 60.0 <= p < 80.0)
            between_50_80 = sum(1 for p in candidate_pcts if 50.0 <= p < 80.0)

            # Calculate Weak: < 60%, < 50%
            lt_60 = sum(1 for p in candidate_pcts if p < 60.0)
            lt_50 = sum(1 for p in candidate_pcts if p < 50.0)

            subject_percentage_rows.append({
                'subject_id': s_id,
                'title': subj_display_title,
                'name': s['name'],
                'max_score': s['max_score'],
                'gte_counts': gte_counts,
                'between_60_80': between_60_80,
                'between_50_80': between_50_80,
                'lt_60': lt_60,
                'lt_50': lt_50,
            })

        # -------------------------------------------------------------
        # 5. Slow Learners / Candidates Matrix (Image 1)
        # -------------------------------------------------------------
        # In Image 1, rows display student records with scores or mentions
        # Default view shows students who scored E or F in any subject
        slow_learners_data = []
        for idx, c in enumerate(processed_candidates, 1):
            cand_row = {
                'index': idx,
                'id': c['id'],
                'student_code': c['student_code'],
                'candidate_name_kh': c['candidate_name_kh'],
                'gender': c['gender'],
                'gender_short': c['gender_short'],
                'origin_class': c['origin_class'],
                'overall_mention': c['overall_mention'],
                'has_weak_grade': any(c['subjects'][s['id']]['is_weak'] for s in subjects_list if s['id'] in c['subjects']),
                'subject_scores': {},
                'subject_mentions': {},
                'subject_weak_flags': {},
            }
            for s in subjects_list:
                s_id = s['id']
                s_data = c['subjects'].get(s_id, {})
                cand_row['subject_scores'][s_id] = s_data.get('score')
                cand_row['subject_mentions'][s_id] = s_data.get('mention', '-')
                cand_row['subject_weak_flags'][s_id] = s_data.get('is_weak', False)

            slow_learners_data.append(cand_row)

        # Collect unique classroom names across active exams for classroom scope filter
        all_classrooms = sorted(list(set(
            c['origin_class'] for c in processed_candidates if c['origin_class']
        )))

        # Available grade levels across the active exams
        available_grades = sorted(list(set(e.grade_level for e in active_exams)))

        return {
            'scope': scope,
            'selected_grade': grade_level,
            'selected_class': classroom_name,
            'available_grades': available_grades,
            'all_classrooms': all_classrooms,
            'subjects_list': subjects_list,
            'total_candidates': total_candidates_count,
            'female_candidates': female_candidates_count,
            'male_candidates': male_candidates_count,
            'overall_mentions': overall_mentions,
            'quality_evaluation': quality_evaluation,
            'subject_mentions_single': subject_mentions_single,
            'subject_mentions_detailed': subject_mentions_detailed,
            'subject_percentage_rows': subject_percentage_rows,
            'percentage_thresholds': percentage_thresholds,
            'slow_learners_data': slow_learners_data,
        }

    @classmethod
    def generate_mock_scores(cls, exams):
        """
        Generates realistic, natural mock examination scores with grades A to F
        for all candidates across all subjects for the given list of exams.
        """
        import random
        from decimal import Decimal

        if not hasattr(exams, '__iter__'):
            exams = [exams]
        else:
            exams = list(exams)

        # Realistic tier definitions: (tier_name, target_weight, pct_min, pct_max)
        tiers = [
            ('A', 0.10, 90.0, 98.0),   # 10% A
            ('B', 0.18, 80.0, 89.0),   # 18% B
            ('C', 0.30, 70.0, 79.0),   # 30% C
            ('D', 0.22, 60.0, 69.0),   # 22% D
            ('E', 0.12, 50.0, 59.0),   # 12% E
            ('F', 0.08, 28.0, 48.0),   # 8% F (slow learners)
        ]
        tier_choices = [t[0] for t in tiers]
        tier_weights = [t[1] for t in tiers]
        tier_ranges = {t[0]: (t[2], t[3]) for t in tiers}

        total_candidates_scored = 0
        total_scores_saved = 0

        for exam in exams:
            subjects = list(exam.exam_subjects.all().select_related('subject'))
            if not subjects:
                continue

            candidates = list(exam.candidates.all())
            if not candidates:
                continue

            # Preload existing CandidateSubjectScore objects
            existing_scores = {
                (sc.candidate_id, sc.exam_subject_id): sc
                for sc in CandidateSubjectScore.objects.filter(candidate__exam=exam)
            }

            scores_to_create = []
            scores_to_update = []

            for cand in candidates:
                chosen_tier = random.choices(tier_choices, weights=tier_weights, k=1)[0]
                t_min, t_max = tier_ranges[chosen_tier]
                student_absent = (random.random() < 0.015)  # 1.5% chance absent

                for s in subjects:
                    max_s = float(s.max_score) if s.max_score and float(s.max_score) > 0 else 50.0

                    if student_absent:
                        is_absent = True
                        score_val = Decimal('0.00')
                    else:
                        is_absent = False
                        base_pct = random.uniform(t_min, t_max)
                        subj_jitter = random.uniform(-6.0, 6.0)
                        final_pct = max(15.0, min(100.0, base_pct + subj_jitter))
                        calculated_score = round((final_pct / 100.0) * max_s, 1)
                        score_val = Decimal(str(calculated_score))

                    key = (cand.id, s.id)
                    if key in existing_scores:
                        sc = existing_scores[key]
                        sc.score = score_val
                        sc.is_absent = is_absent
                        sc.signature_present = not is_absent
                        scores_to_update.append(sc)
                    else:
                        scores_to_create.append(CandidateSubjectScore(
                            candidate=cand,
                            exam_subject=s,
                            score=score_val,
                            is_absent=is_absent,
                            signature_present=not is_absent
                        ))

                total_candidates_scored += 1

            if scores_to_create:
                CandidateSubjectScore.objects.bulk_create(scores_to_create, batch_size=1000)
                total_scores_saved += len(scores_to_create)
            if scores_to_update:
                CandidateSubjectScore.objects.bulk_update(scores_to_update, ['score', 'is_absent', 'signature_present'], batch_size=1000)
                total_scores_saved += len(scores_to_update)

            # Recalculate candidate averages, ranks, and grade letters
            exam.recalculate_all_ranks()

        return {
            'exams_count': len(exams),
            'candidates_count': total_candidates_scored,
            'scores_count': total_scores_saved,
        }

    @classmethod
    def clear_mock_scores(cls, exams):
        """
        Clears all scores and resets grades back to '-' for the given exams.
        """
        from decimal import Decimal

        if not hasattr(exams, '__iter__'):
            exams = [exams]
        else:
            exams = list(exams)

        total_candidates_cleared = 0
        for exam in exams:
            cands = exam.candidates.all()
            CandidateSubjectScore.objects.filter(candidate__in=cands).update(score=None, is_absent=False)
            cands.update(
                total_score=Decimal('0.00'),
                average_score=Decimal('0.00'),
                grade_letter='-',
                rank_overall=None,
                rank_in_room=None,
            )
            total_candidates_cleared += cands.count()

        return {
            'exams_count': len(exams),
            'candidates_count': total_candidates_cleared,
        }

    @classmethod
    def _empty_payload(cls):
        return {
            'scope': 'school',
            'selected_grade': None,
            'selected_class': None,
            'available_grades': [],
            'all_classrooms': [],
            'subjects_list': [],
            'total_candidates': 0,
            'female_candidates': 0,
            'male_candidates': 0,
            'overall_mentions': {'grades': ['A', 'B', 'C', 'D', 'E', 'F'], 'total_row': {'counts': [0]*6, 'sum_abc': 0, 'grand_total': 0}, 'female_row': {'counts': [0]*6, 'sum_abc': 0, 'grand_total': 0}, 'male_row': {'counts': [0]*6, 'sum_abc': 0, 'grand_total': 0}},
            'quality_evaluation': {'headers': [], 'total_row': {'good': 0, 'fairly_good': 0, 'average': 0, 'weak': 0, 'grand_total': 0}, 'female_row': {'good': 0, 'fairly_good': 0, 'average': 0, 'weak': 0, 'grand_total': 0}, 'male_row': {'good': 0, 'fairly_good': 0, 'average': 0, 'weak': 0, 'grand_total': 0}},
            'subject_mentions_single': [],
            'subject_mentions_detailed': [],
            'subject_percentage_rows': [],
            'percentage_thresholds': [95, 90, 85, 80, 75, 70, 65, 60, 55, 50],
            'slow_learners_data': [],
        }

    @classmethod
    def build_analytics_workbook(cls, analytics, session_title="សម័យប្រឡង", session_date_str="", academic_year_name="", scope_title="", target_sheet='all', mention_sum_label="A+B+C"):
        """
        Builds a professionally formatted multi-sheet or single-sheet Excel (.xlsx) workbook
        representing all reports on the Exam Session Analytics page.
        Supported target_sheet:
          - 'all': Multi-sheet workbook containing all 6 reports
          - 'overall': Sheet 1 (សរុបនិទ្ទេសរួម)
          - 'quality': Sheet 2 (វាយតម្លៃគុណភាព)
          - 'subject_summary': Sheet 3 & 4 (សង្ខេបនិទ្ទេស សរុប និង លម្អិត)
          - 'subject_single': Sheet 3 (សង្ខេបនិទ្ទេស សរុប)
          - 'subject_detailed': Sheet 4 (សង្ខេបនិទ្ទេស លម្អិត)
          - 'percentages': Sheet 5 (វិភាគភាគរយមុខវិជ្ជា)
          - 'slow_learners': Sheet 6 (របាយការណ៍សិស្សរៀនយឺត)
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        default_sheet = wb.active

        font_family = 'Khmer OS Siemreap'

        title_font = Font(name=font_family, size=13, bold=True, color='0F172A')
        subtitle_font = Font(name=font_family, size=9.5, bold=False, color='475569')
        header_font_white = Font(name=font_family, size=9.5, bold=True, color='FFFFFF')
        header_font_dark = Font(name=font_family, size=9.5, bold=True, color='1E293B')
        bold_font = Font(name=font_family, size=9.5, bold=True, color='0F172A')
        regular_font = Font(name=font_family, size=9.5, bold=False, color='1E293B')
        danger_font = Font(name=font_family, size=9.5, bold=True, color='DC2626')
        female_font = Font(name=font_family, size=9.5, bold=True, color='DB2777')
        male_font = Font(name=font_family, size=9.5, bold=True, color='2563EB')

        fill_teal = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
        fill_emerald = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        fill_cyan = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
        fill_coral = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
        fill_light_slate = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        fill_yellow_highlight = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid')
        fill_green_highlight = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

        border_thin = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        def write_header_block(ws, title_text, col_span=9):
            col_letter = get_column_letter(max(col_span, 6))
            ws.merge_cells(f'A1:{col_letter}1')
            ws['A1'] = f"របាយការណ៍វិភាគលទ្ធផលប្រឡង៖ {session_title}"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(f'A2:{col_letter}2')
            ws['A2'] = f"{title_text} | វិសាលភាព៖ {scope_title} | ឆ្នាំសិក្សា៖ {academic_year_name} | កាលបរិច្ឆេទ៖ {session_date_str}"
            ws['A2'].font = subtitle_font
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(f'A3:{col_letter}3')
            ws['A3'] = f"បេក្ខជនសរុប៖ {analytics.get('total_candidates', 0)} នាក់ (ស្រី៖ {analytics.get('female_candidates', 0)} នាក់, ប្រុស៖ {analytics.get('male_candidates', 0)} នាក់)"
            ws['A3'].font = subtitle_font
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[1].height = 25
            ws.row_dimensions[2].height = 18
            ws.row_dimensions[3].height = 18
            ws.append([])

        sum_label = mention_sum_label or "A+B+C"

        # -------------------------------------------------------------
        # SHEET 1: សរុបនិទ្ទេសរួម
        # -------------------------------------------------------------
        if target_sheet in ['all', 'overall']:
            ws1 = wb.create_sheet(title="សរុបនិទ្ទេសរួម")
            write_header_block(ws1, "១. តារាងសរុបនិទ្ទេសរួម", col_span=9)
            headers = ['ភេទ', 'A', 'B', 'C', 'D', 'E', 'F', sum_label, 'សរុបរួម']
            ws1.append(headers)
            ws1.row_dimensions[5].height = 24
            for c_idx in range(1, 10):
                cell = ws1.cell(row=5, column=c_idx)
                cell.font = header_font_white
                cell.fill = fill_teal
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin

            om = analytics.get('overall_mentions', {})
            t_row = om.get('total_row', {'label': 'សរុប', 'counts': [0]*6, 'sum_abc': 0, 'grand_total': 0})
            f_row = om.get('female_row', {'label': 'ស្រី', 'counts': [0]*6, 'sum_abc': 0, 'grand_total': 0})
            m_row = om.get('male_row', {'label': 'ប្រុស', 'counts': [0]*6, 'sum_abc': 0, 'grand_total': 0})

            rows_data = [
                (t_row['label'], t_row['counts'], t_row['sum_abc'], t_row['grand_total'], bold_font, fill_light_slate),
                (f_row['label'], f_row['counts'], f_row['sum_abc'], f_row['grand_total'], female_font, None),
                (m_row['label'], m_row['counts'], m_row['sum_abc'], m_row['grand_total'], male_font, None),
            ]
            for r_idx, (lbl, counts, sum_abc, grand_tot, f_style, row_fill) in enumerate(rows_data, 6):
                ws1.row_dimensions[r_idx].height = 22
                row_vals = [lbl] + list(counts) + [sum_abc, grand_tot]
                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws1.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = f_style
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    if row_fill:
                        cell.fill = row_fill
                    if c_idx == 8:
                        cell.fill = fill_yellow_highlight
                        cell.font = bold_font
                    elif c_idx == 9:
                        cell.fill = fill_green_highlight
                        cell.font = bold_font

        # -------------------------------------------------------------
        # SHEET 2: វាយតម្លៃគុណភាព
        # -------------------------------------------------------------
        if target_sheet in ['all', 'quality']:
            ws2 = wb.create_sheet(title="វាយតម្លៃគុណភាព")
            write_header_block(ws2, "២. តារាងរបាយការណ៍វាយតម្លៃគុណភាព", col_span=6)
            headers2 = ['ភេទ', 'ល្អ (>=40)', 'ល្អបង្គួរ (>=32.5)', 'មធ្យម (>=25)', 'ខ្សោយ (<25)', 'សរុប']
            ws2.append(headers2)
            ws2.row_dimensions[5].height = 24
            for c_idx in range(1, 7):
                cell = ws2.cell(row=5, column=c_idx)
                cell.font = header_font_white
                cell.fill = fill_teal
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin

            qe = analytics.get('quality_evaluation', {})
            t_qe = qe.get('total_row', {'label': 'សរុប', 'good': 0, 'fairly_good': 0, 'average': 0, 'weak': 0, 'grand_total': 0})
            f_qe = qe.get('female_row', {'label': 'ស្រី', 'good': 0, 'fairly_good': 0, 'average': 0, 'weak': 0, 'grand_total': 0})
            m_qe = qe.get('male_row', {'label': 'ប្រុស', 'good': 0, 'fairly_good': 0, 'average': 0, 'weak': 0, 'grand_total': 0})

            rows_data2 = [
                (t_qe['label'], t_qe['good'], t_qe['fairly_good'], t_qe['average'], t_qe['weak'], t_qe['grand_total'], bold_font, fill_light_slate),
                (f_qe['label'], f_qe['good'], f_qe['fairly_good'], f_qe['average'], f_qe['weak'], f_qe['grand_total'], female_font, None),
                (m_qe['label'], m_qe['good'], m_qe['fairly_good'], m_qe['average'], m_qe['weak'], m_qe['grand_total'], male_font, None),
            ]
            for r_idx, (lbl, gd, fg, av, wk, gt, f_style, row_fill) in enumerate(rows_data2, 6):
                ws2.row_dimensions[r_idx].height = 22
                row_vals = [lbl, gd, fg, av, wk, gt]
                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = f_style
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    if row_fill:
                        cell.fill = row_fill
                    if c_idx == 5:
                        cell.font = danger_font
                    elif c_idx == 6:
                        cell.fill = fill_green_highlight
                        cell.font = bold_font

        # -------------------------------------------------------------
        # SHEET 3: សង្ខេបនិទ្ទេស (សរុប)
        # -------------------------------------------------------------
        if target_sheet in ['all', 'subject_summary', 'subject_single']:
            ws3 = wb.create_sheet(title="សង្ខេបនិទ្ទេស (សរុប)")
            write_header_block(ws3, "៣. សង្ខេបនិទ្ទេសតាមមុខវិជ្ជា (បង្ហាញតែមួយ - សរុប)", col_span=9)
            headers3 = ['មុខវិជ្ជា (ពិន្ទុពេញ)', 'A', 'B', 'C', 'D', 'E', 'F', sum_label, 'សរុបរួម']
            ws3.append(headers3)
            ws3.row_dimensions[5].height = 24
            for c_idx in range(1, 10):
                cell = ws3.cell(row=5, column=c_idx)
                cell.font = header_font_white
                cell.fill = fill_emerald
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
            ws3.cell(row=5, column=1).alignment = Alignment(horizontal='left', vertical='center')

            for r_idx, s_row in enumerate(analytics.get('subject_mentions_single', []), 6):
                ws3.row_dimensions[r_idx].height = 20
                row_vals = [s_row['title'], s_row['a'], s_row['b'], s_row['c'], s_row['d'], s_row['e'], s_row['f'], s_row['sum_abc'], s_row['total']]
                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = regular_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    if c_idx == 1:
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif c_idx == 7:
                        cell.font = danger_font
                    elif c_idx == 8:
                        cell.fill = fill_yellow_highlight
                        cell.font = bold_font
                    elif c_idx == 9:
                        cell.fill = fill_green_highlight
                        cell.font = bold_font

        # -------------------------------------------------------------
        # SHEET 4: សង្ខេបនិទ្ទេស (លម្អិត)
        # -------------------------------------------------------------
        if target_sheet in ['all', 'subject_summary', 'subject_detailed']:
            ws4 = wb.create_sheet(title="សង្ខេបនិទ្ទេស (លម្អិត)")
            write_header_block(ws4, "៤. សង្ខេបនិទ្ទេសតាមមុខវិជ្ជា (បង្ហាញលម្អិត - សរុប, ស្រី, ប្រុស)", col_span=25)

            ws4.merge_cells('A5:A6')
            ws4['A5'] = 'មុខវិជ្ជា (ពិន្ទុពេញ)'
            top_groups = [
                ('B5:D5', 'A'), ('E5:G5', 'B'), ('H5:J5', 'C'),
                ('K5:M5', 'D'), ('N5:P5', 'E'), ('Q5:S5', 'F'),
                ('T5:V5', sum_label), ('W5:Y5', 'សរុបរួម')
            ]
            for rng, txt in top_groups:
                ws4.merge_cells(rng)
                start_cell = rng.split(':')[0]
                ws4[start_cell] = txt

            ws4.row_dimensions[5].height = 22
            ws4.row_dimensions[6].height = 20

            sub_headers = [''] + ['ស', 'ស្រ', 'ប'] * 8
            ws4.append(sub_headers)

            for col_idx in range(1, 26):
                cell5 = ws4.cell(row=5, column=col_idx)
                cell5.font = header_font_white
                cell5.fill = fill_emerald if col_idx < 20 else (fill_yellow_highlight if col_idx < 23 else fill_teal)
                if col_idx in [20, 21, 22]:
                    cell5.font = header_font_dark
                cell5.alignment = Alignment(horizontal='center', vertical='center')
                cell5.border = border_thin

                cell6 = ws4.cell(row=6, column=col_idx)
                cell6.font = header_font_white
                cell6.fill = fill_teal if col_idx < 20 else (fill_yellow_highlight if col_idx < 23 else fill_teal)
                if col_idx in [20, 21, 22]:
                    cell6.font = header_font_dark
                cell6.alignment = Alignment(horizontal='center', vertical='center')
                cell6.border = border_thin

            for r_idx, d_row in enumerate(analytics.get('subject_mentions_detailed', []), 7):
                ws4.row_dimensions[r_idx].height = 20
                row_vals = [d_row['title']]
                for g in ['A', 'B', 'C', 'D', 'E', 'F']:
                    g_data = d_row.get('grades', {}).get(g, {'tot': 0, 'fem': 0, 'mal': 0})
                    row_vals.extend([g_data['tot'], g_data['fem'], g_data['mal']])
                s_abc = d_row.get('sum_abc', {'tot': 0, 'fem': 0, 'mal': 0})
                row_vals.extend([s_abc['tot'], s_abc['fem'], s_abc['mal']])
                s_tot = d_row.get('total', {'tot': 0, 'fem': 0, 'mal': 0})
                row_vals.extend([s_tot['tot'], s_tot['fem'], s_tot['mal']])

                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws4.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = regular_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    if c_idx == 1:
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif c_idx in [17, 18, 19]:
                        cell.font = danger_font
                    elif c_idx in [20, 21, 22]:
                        cell.fill = fill_yellow_highlight
                        cell.font = bold_font
                    elif c_idx in [23, 24, 25]:
                        cell.fill = fill_green_highlight
                        cell.font = bold_font

        # -------------------------------------------------------------
        # SHEET 5: វិភាគភាគរយមុខវិជ្ជា
        # -------------------------------------------------------------
        if target_sheet in ['all', 'percentages']:
            ws5 = wb.create_sheet(title="វិភាគភាគរយមុខវិជ្ជា")
            write_header_block(ws5, "៥. វិភាគភាគរយមុខវិជ្ជា (ផ្អែកលើពិន្ទុអតិបរមា)", col_span=15)

            ws5.merge_cells('A5:A6')
            ws5['A5'] = 'មុខវិជ្ជា (ពិន្ទុពេញ)'
            ws5.merge_cells('B5:K5')
            ws5['B5'] = 'កម្រិតភាគរយ (>=)'
            ws5.merge_cells('L5:M5')
            ws5['L5'] = 'ចន្លោះភាគរយ'
            ws5.merge_cells('N5:O5')
            ws5['N5'] = 'កម្រិតខ្សោយ (<)'

            ws5.row_dimensions[5].height = 22
            ws5.row_dimensions[6].height = 20

            pct_subs = [''] + ['95%', '90%', '85%', '80%', '75%', '70%', '65%', '60%', '55%', '50%', '60-80%', '50-<80%', '60%', '50%']
            ws5.append(pct_subs)

            for col_idx in range(1, 16):
                cell5 = ws5.cell(row=5, column=col_idx)
                cell5.font = header_font_white
                cell5.fill = fill_teal if col_idx <= 11 else (fill_cyan if col_idx <= 13 else fill_coral)
                cell5.alignment = Alignment(horizontal='center', vertical='center')
                cell5.border = border_thin

                cell6 = ws5.cell(row=6, column=col_idx)
                cell6.font = header_font_white
                cell6.fill = fill_teal if col_idx <= 11 else (fill_cyan if col_idx <= 13 else fill_coral)
                cell6.alignment = Alignment(horizontal='center', vertical='center')
                cell6.border = border_thin

            for r_idx, p_row in enumerate(analytics.get('subject_percentage_rows', []), 7):
                ws5.row_dimensions[r_idx].height = 20
                row_vals = [p_row['title']] + p_row['gte_counts'] + [p_row['between_60_80'], p_row['between_50_80'], p_row['lt_60'], p_row['lt_50']]
                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws5.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = regular_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    if c_idx == 1:
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif c_idx in [12, 13]:
                        cell.font = male_font
                    elif c_idx in [14, 15]:
                        cell.font = danger_font

        # -------------------------------------------------------------
        # SHEET 6: របាយការណ៍សិស្សរៀនយឺត
        # -------------------------------------------------------------
        if target_sheet in ['all', 'slow_learners']:
            ws6 = wb.create_sheet(title="សិស្សរៀនយឺត")
            subjects_list = analytics.get('subjects_list', [])
            total_cols = 6 + (len(subjects_list) * 2)
            write_header_block(ws6, "៦. របាយការណ៍សិស្សរៀនយឺត (តាមមុខវិជ្ជា)", col_span=total_cols)

            headers6 = ['ល.រ', 'អត្តលេខ', 'ឈ្មោះសិស្ស', 'ភេទ', 'ថ្នាក់រៀន', 'និទ្ទេសរួម']
            for s in subjects_list:
                headers6.extend([f"{s['name']} (ពិន្ទុ)", f"{s['name']} (និទ្ទេស)"])

            ws6.append(headers6)
            ws6.row_dimensions[5].height = 24
            for col_idx in range(1, len(headers6) + 1):
                cell = ws6.cell(row=5, column=col_idx)
                cell.font = header_font_white
                cell.fill = fill_teal if col_idx <= 6 else fill_emerald
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
            ws6.cell(row=5, column=3).alignment = Alignment(horizontal='left', vertical='center')

            candidates = analytics.get('slow_learners_data', [])
            slow_candidates = [c for c in candidates if c.get('overall_mention') in ['E', 'F'] or c.get('has_weak_grade')]
            if not slow_candidates:
                slow_candidates = candidates

            for r_idx, cand in enumerate(slow_candidates, 6):
                ws6.row_dimensions[r_idx].height = 20
                row_vals = [
                    r_idx - 5,
                    cand.get('student_code', ''),
                    cand.get('candidate_name_kh', ''),
                    cand.get('gender_short', ''),
                    cand.get('origin_class', ''),
                    cand.get('overall_mention', ''),
                ]
                for s in subjects_list:
                    s_id = s['id']
                    sc = cand.get('subject_scores', {}).get(s_id)
                    sc_str = f"{sc:.1f}" if sc is not None else "-"
                    m_str = cand.get('subject_mentions', {}).get(s_id, '-')
                    row_vals.extend([sc_str, m_str])

                for c_idx, val in enumerate(row_vals, 1):
                    cell = ws6.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = regular_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    if c_idx == 3:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif c_idx == 6 and val in ['E', 'F']:
                        cell.font = danger_font
                    elif c_idx > 6 and (c_idx % 2 == 0) and val in ['E', 'F']:
                        cell.font = danger_font

        # Auto-fit Column Widths for all created sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row >= 5 and cell.value is not None:
                        val_str = str(cell.value)
                        val_len = len(val_str)
                        if val_len > max_len:
                            max_len = val_len
                adjusted_width = max(max_len + 4, 10)
                if col_letter == 'A':
                    adjusted_width = max(adjusted_width, 22)
                elif col_letter in ['B', 'C']:
                    adjusted_width = max(adjusted_width, 16)
                sheet.column_dimensions[col_letter].width = min(adjusted_width, 42)

        # Remove the default empty sheet if custom sheets were created
        if len(wb.worksheets) > 1 and default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        return wb

