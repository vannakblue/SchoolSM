import os
import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
from apps.academics.models import Province, District, Commune, Village


class Command(BaseCommand):
    help = "Seeds Cambodia administrative locations (Provinces, Districts, Communes, Villages) from Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            '--force',
            action='store_true',
            help='Force re-seed even if locations already exist in the database',
        )

    def handle(self, *args, **options):
        force = options.get('force', False)
        existing_provinces = Province.objects.count()

        if existing_provinces > 0 and not force:
            self.stdout.write(self.style.SUCCESS(
                f"[SKIP] Cambodia locations already seeded ({existing_provinces} provinces, "
                f"{District.objects.count()} districts, {Commune.objects.count()} communes, {Village.objects.count()} villages). "
                "Use --force to reload."
            ))
            return

        excel_file = os.path.join(settings.BASE_DIR, 'Cambodia All List2025.xlsx')
        if not os.path.exists(excel_file):
            self.stdout.write(self.style.ERROR(f"[ERROR] Excel file not found at: {excel_file}"))
            return

        self.stdout.write(f"Loading Excel file: {excel_file} ...")
        wb = openpyxl.load_workbook(excel_file, data_only=True)

        # 1. Provinces
        self.stdout.write("Seeding Provinces...")
        ws_prov = wb['CambodiaProvinceList2025']
        provinces_to_create = []
        for r in range(2, ws_prov.max_row + 1):
            p_code = str(ws_prov.cell(row=r, column=1).value or '').strip()
            p_kh = str(ws_prov.cell(row=r, column=2).value or '').strip()
            p_en = str(ws_prov.cell(row=r, column=3).value or '').strip()
            if p_code and p_kh:
                provinces_to_create.append(Province(code=p_code, name_kh=p_kh, name_en=p_en))

        with transaction.atomic():
            Province.objects.all().delete()
            Province.objects.bulk_create(provinces_to_create)

        province_map = {p.code: p.id for p in Province.objects.all()}

        # 2. Districts
        self.stdout.write("Seeding Districts...")
        ws_dist = wb['CambodiaDistrictList2025']
        districts_to_create = []
        for r in range(2, ws_dist.max_row + 1):
            p_code = str(ws_dist.cell(row=r, column=1).value or '').strip()
            d_code = str(ws_dist.cell(row=r, column=2).value or '').strip()
            d_kh = str(ws_dist.cell(row=r, column=3).value or '').strip()
            d_en = str(ws_dist.cell(row=r, column=4).value or '').strip()
            if d_code and d_kh and p_code in province_map:
                districts_to_create.append(District(province_id=province_map[p_code], code=d_code, name_kh=d_kh, name_en=d_en))

        with transaction.atomic():
            District.objects.all().delete()
            District.objects.bulk_create(districts_to_create)

        district_map = {d.code: d.id for d in District.objects.all()}

        # 3. Communes
        self.stdout.write("Seeding Communes...")
        ws_comm = wb['CambodiaCommuneList2025']
        communes_to_create = []
        for r in range(2, ws_comm.max_row + 1):
            d_code = str(ws_comm.cell(row=r, column=2).value or '').strip()
            c_code = str(ws_comm.cell(row=r, column=3).value or '').strip()
            c_kh = str(ws_comm.cell(row=r, column=4).value or '').strip()
            c_en = str(ws_comm.cell(row=r, column=5).value or '').strip()
            if c_code and c_kh and d_code in district_map:
                communes_to_create.append(Commune(district_id=district_map[d_code], code=c_code, name_kh=c_kh, name_en=c_en))

        with transaction.atomic():
            Commune.objects.all().delete()
            Commune.objects.bulk_create(communes_to_create, batch_size=1000)

        commune_map = {c.code: c.id for c in Commune.objects.all()}

        # 4. Villages
        self.stdout.write("Seeding Villages...")
        ws_vill = wb['CambodiaVillagesList2025']
        villages_to_create = []
        for r in range(2, ws_vill.max_row + 1):
            c_code = str(ws_vill.cell(row=r, column=3).value or '').strip()
            v_code = str(ws_vill.cell(row=r, column=4).value or '').strip()
            v_kh = str(ws_vill.cell(row=r, column=5).value or '').strip()
            v_en = str(ws_vill.cell(row=r, column=6).value or '').strip()
            if v_kh and c_code in commune_map:
                villages_to_create.append(Village(commune_id=commune_map[c_code], code=v_code, name_kh=v_kh, name_en=v_en))

        with transaction.atomic():
            Village.objects.all().delete()
            Village.objects.bulk_create(villages_to_create, batch_size=2000)

        self.stdout.write(self.style.SUCCESS(
            f"[OK] Successfully seeded Cambodia locations! "
            f"Provinces: {Province.objects.count()}, Districts: {District.objects.count()}, "
            f"Communes: {Commune.objects.count()}, Villages: {Village.objects.count()}"
        ))
