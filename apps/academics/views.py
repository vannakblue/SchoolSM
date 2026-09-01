from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import ValidationError
from django.db.models import Q, Count
from decimal import Decimal
from collections import defaultdict
import datetime
from django.utils import timezone
import json
import csv
import os
import re
import string
from apps.accounts.decorators import role_required
from .models import AcademicYear, Classroom, Subject, ClassSubject, Timetable, GradeLevelRule, SavedDefaultConfig, GradeLevel, Province, District, Commune, Village, GradeEnrollmentOption, AcademicTrack, TeacherDutySchedule, TeacherDutyType
from .forms import ClassroomForm, SubjectForm, TimetableForm, GradeLevelForm, AcademicYearForm, GradeEnrollmentOptionForm, AcademicTrackForm
from apps.students.models import Student
from apps.teachers.models import Teacher
from apps.examinations.models import Grade, ExamTerm

# Official 14 MoEYS Subjects in EXACT specified order
DEFAULT_MOEYS_SUBJECTS = [
    ('តែងសេចក្តី', 'Composition / Essay', 'R', 2, '#4f46e5', 1),
    ('សរសេរតាមអាន', 'Dictation', 'D', 2, '#6366f1', 2),
    ('ភាសាខ្មែរ', 'Khmer Language', 'K', 4, '#0ea5e9', 3),
    ('សីលធម៌', 'Civics & Moral / Ethics', 'I', 2, '#06b6d4', 4),
    ('ភូមិវិទ្យា', 'Geography', 'G', 2, '#d97706', 5),
    ('ប្រវត្តិវិទ្យា', 'History', 'H', 2, '#f59e0b', 6),
    ('គណិតវិទ្យា', 'Mathematics', 'M', 4, '#dc2626', 7),
    ('ផែនដីវិទ្យា', 'Earth Science', 'Es', 2, '#84cc16', 8),
    ('រូបវិទ្យា', 'Physics', 'P', 3, '#8b5cf6', 9),
    ('គីមីវិទ្យា', 'Chemistry', 'C', 3, '#10b981', 10),
    ('ជីវវិទ្យា', 'Biology', 'B', 3, '#14b8a6', 11),
    ('គេហវិទ្យា', 'Home Economics', 'He', 2, '#ec4899', 12),
    ('សេដ្ឋកិច្ច', 'Economics', 'Ec', 2, '#f97316', 13),
    ('អង់គ្លេស', 'English Language', 'E', 3, '#3b82f6', 14),
]

OFFICIAL_CODES = [s[2] for s in DEFAULT_MOEYS_SUBJECTS]

DEFAULT_MOEYS_STREAMS = [
    ('ថ្នាក់ទី ៧', 7, 'GENERAL', 1),
    ('ថ្នាក់ទី ៨', 8, 'GENERAL', 2),
    ('ថ្នាក់ទី ៩', 9, 'GENERAL', 3),
    ('ថ្នាក់ទី ១០', 10, 'GENERAL', 4),
    ('ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រសង្គម', 11, 'SOCIAL', 5),
    ('ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រ', 11, 'SCIENCE', 6),
    ('ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រសង្គម', 12, 'SOCIAL', 7),
    ('ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រ', 12, 'SCIENCE', 8),
]

# Official MoEYS Standard Scoring Rules Matrix for the 8 Streams
MOEYS_SCORING_RULES = {
    (7, 'GENERAL'): {
        'តែងសេចក្តី': 60, 'សរសេរតាមអាន': 40, 'សីលធម៌': 50, 'ភូមិវិទ្យា': 50, 'ប្រវត្តិវិទ្យា': 50,
        'គណិតវិទ្យា': 100, 'ផែនដីវិទ្យា': 50, 'រូបវិទ្យា': 50, 'គីមីវិទ្យា': 50, 'ជីវវិទ្យា': 50,
        'គេហវិទ្យា': 50, 'អង់គ្លេស': 50
    },
    (8, 'GENERAL'): {
        'តែងសេចក្តី': 60, 'សរសេរតាមអាន': 40, 'សីលធម៌': 50, 'ភូមិវិទ្យា': 50, 'ប្រវត្តិវិទ្យា': 50,
        'គណិតវិទ្យា': 100, 'ផែនដីវិទ្យា': 50, 'រូបវិទ្យា': 50, 'គីមីវិទ្យា': 50, 'ជីវវិទ្យា': 50,
        'គេហវិទ្យា': 50, 'អង់គ្លេស': 50
    },
    (9, 'GENERAL'): {
        'តែងសេចក្តី': 60, 'សរសេរតាមអាន': 40, 'សីលធម៌': 35, 'ភូមិវិទ្យា': 32, 'ប្រវត្តិវិទ្យា': 33,
        'គណិតវិទ្យា': 100, 'ផែនដីវិទ្យា': 25, 'រូបវិទ្យា': 35, 'គីមីវិទ្យា': 25, 'ជីវវិទ្យា': 35,
        'គេហវិទ្យា': 50, 'អង់គ្លេស': 50
    },
    (10, 'GENERAL'): {
        'ភាសាខ្មែរ': 150, 'សីលធម៌': 38, 'ភូមិវិទ្យា': 38, 'ប្រវត្តិវិទ្យា': 37,
        'គណិតវិទ្យា': 150, 'ផែនដីវិទ្យា': 25, 'រូបវិទ្យា': 50, 'គីមីវិទ្យា': 37,
        'ជីវវិទ្យា': 38, 'គេហវិទ្យា': 37, 'អង់គ្លេស': 100
    },
    (11, 'SOCIAL'): {
        'ភាសាខ្មែរ': 125, 'សីលធម៌': 75, 'ភូមិវិទ្យា': 75, 'ប្រវត្តិវិទ្យា': 75,
        'គណិតវិទ្យា': 75, 'ផែនដីវិទ្យា': 50, 'រូបវិទ្យា': 50, 'គីមីវិទ្យា': 50,
        'ជីវវិទ្យា': 50, 'សេដ្ឋកិច្ច': 50, 'អង់គ្លេស': 50
    },
    (11, 'SCIENCE'): {
        'ភាសាខ្មែរ': 75, 'សីលធម៌': 50, 'ភូមិវិទ្យា': 50, 'ប្រវត្តិវិទ្យា': 50,
        'គណិតវិទ្យា': 125, 'ផែនដីវិទ្យា': 50, 'រូបវិទ្យា': 75, 'គីមីវិទ្យា': 75,
        'ជីវវិទ្យា': 75, 'សេដ្ឋកិច្ច': 50, 'អង់គ្លេស': 50
    },
    (12, 'SOCIAL'): {
        'ភាសាខ្មែរ': 125, 'សីលធម៌': 75, 'ភូមិវិទ្យា': 75, 'ប្រវត្តិវិទ្យា': 75,
        'គណិតវិទ្យា': 75, 'ផែនដីវិទ្យា': 50, 'រូបវិទ្យា': 50, 'គីមីវិទ្យា': 50,
        'ជីវវិទ្យា': 50, 'សេដ្ឋកិច្ច': 50, 'អង់គ្លេស': 50
    },
    (12, 'SCIENCE'): {
        'ភាសាខ្មែរ': 75, 'សីលធម៌': 50, 'ភូមិវិទ្យា': 50, 'ប្រវត្តិវិទ្យា': 50,
        'គណិតវិទ្យា': 125, 'ផែនដីវិទ្យា': 50, 'រូបវិទ្យា': 75, 'គីមីវិទ្យា': 75,
        'ជីវវិទ្យា': 75, 'សេដ្ឋកិច្ច': 50, 'អង់គ្លេស': 50
    },
}

DEFAULT_MOEYS_CLASSROOMS = [
    ('7A', 'ថ្នាក់ទី ៧A', 7, Classroom.Track.GENERAL, 'បន្ទប់ 001'),
    ('8A', 'ថ្នាក់ទី ៨A', 8, Classroom.Track.GENERAL, 'បន្ទប់ 002'),
    ('9A', 'ថ្នាក់ទី ៩A', 9, Classroom.Track.GENERAL, 'បន្ទប់ 003'),
    ('10A', 'ថ្នាក់ទី ១០A', 10, Classroom.Track.GENERAL, 'បន្ទប់ 101'),
    ('11-SOC', 'ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រសង្គម', 11, Classroom.Track.SOCIAL, 'បន្ទប់ 201'),
    ('11-SCI', 'ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រ', 11, Classroom.Track.SCIENCE, 'បន្ទប់ 202'),
    ('12-SOC', 'ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រសង្គម', 12, Classroom.Track.SOCIAL, 'បន្ទប់ 301'),
    ('12-SCI', 'ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រ', 12, Classroom.Track.SCIENCE, 'បន្ទប់ 302'),
]

# ----------------- GRADE LEVELS CRUD (កម្រិតថ្នាក់) -----------------

@login_required
@role_required(['ADMIN'])
def grade_level_list(request):
    """Lists all configurable Grade Levels / Streams -> Managed in Scoring Rules"""
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def grade_level_create(request):
    """Create a new Grade Level (e.g. ថ្នាក់ទី ១០ វិទ្យាសាស្ត្រ, ថ្នាក់ទី ១០ វិទ្យាសាស្ត្រសង្គម)"""
    if request.method == 'POST':
        form = GradeLevelForm(request.POST)
        if form.is_valid():
            gl = form.save()
            # Initialize default rules for all subjects if none exist
            subjects = Subject.objects.all()
            for s in subjects:
                GradeLevelRule.objects.get_or_create(
                    grade_level=gl.grade_number,
                    track=gl.track,
                    subject=s,
                    defaults={'max_score': Decimal('50.00'), 'order': s.order}
                )
            messages.success(request, f"🎉 បានបង្កើតកម្រិតថ្នាក់ថ្មី '{gl.name}' ជោគជ័យ! លោកអ្នកអាចកំណត់ពិន្ទុអតិបរមានៅលើតារាងច្បាប់ពិន្ទុ។")
            return redirect('grade_rules_manager')
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"⚠️ កំហុស [{field}]: {err}")
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def grade_level_edit(request, pk):
    """Edit Grade Level name or order"""
    gl = get_object_or_404(GradeLevel, pk=pk)
    if request.method == 'POST':
        form = GradeLevelForm(request.POST, instance=gl)
        if form.is_valid():
            form.save()
            messages.success(request, f"បានកែប្រែកម្រិតថ្នាក់ '{gl.name}' ជោគជ័យ!")
            return redirect('grade_rules_manager')
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def grade_level_delete(request, pk):
    """Delete a Grade Level and its scoring rules"""
    gl = get_object_or_404(GradeLevel, pk=pk)
    if request.method == 'POST':
        name = gl.name
        # Delete associated scoring rules
        GradeLevelRule.objects.filter(grade_level=gl.grade_number, track=gl.track).delete()
        gl.delete()
        messages.success(request, f"🗑️ បានលុបកម្រិតថ្នាក់ '{name}' និងច្បាប់ពិន្ទុពាក់ព័ន្ធជោគជ័យ!")
    return redirect('grade_rules_manager')


# ----------------- ACADEMIC TRACKS (PROGRAMS / STREAMS) CRUD -----------------

@login_required
@role_required(['ADMIN'])
def academic_track_list(request):
    """List & manage all academic tracks"""
    return redirect('grade_options_manager')


@login_required
@role_required(['ADMIN'])
def academic_track_create(request):
    """Create a new academic track (Supports AJAX or regular form)"""
    if request.method == 'POST':
        form = AcademicTrackForm(request.POST)
        if form.is_valid():
            track = form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('is_ajax') == '1':
                return JsonResponse({
                    'success': True,
                    'id': track.id,
                    'code': track.code,
                    'name_kh': track.name_kh,
                    'name_en': track.name_en or '',
                    'message': f"បានបង្កើតជំនាញ '{track.name_kh}' ជោគជ័យ!"
                })
            messages.success(request, f"🎉 បានបង្កើតជំនាញសិក្សាថ្មី '{track.name_kh}' ជោគជ័យ!")
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('is_ajax') == '1':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"⚠️ [{field}]: {err}")
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def academic_track_edit(request, pk):
    """Edit an academic track"""
    track = get_object_or_404(AcademicTrack, pk=pk)
    if request.method == 'POST':
        form = AcademicTrackForm(request.POST, instance=track)
        if form.is_valid():
            form.save()
            messages.success(request, f"បានកែប្រែជំនាញសិក្សា '{track.name_kh}' ជោគជ័យ!")
            return redirect('grade_rules_manager')
    else:
        form = AcademicTrackForm(instance=track)
    return render(request, 'academics/track_form.html', {'form': form, 'track': track})


@login_required
@role_required(['ADMIN'])
def academic_track_delete(request, pk):
    """Delete an academic track"""
    track = get_object_or_404(AcademicTrack, pk=pk)
    if request.method == 'POST':
        name = track.name_kh
        code = track.code
        # Check if classrooms or grade levels are using this track
        cl_count = Classroom.objects.filter(track=code).count()
        gl_count = GradeLevel.objects.filter(track=code).count()
        if cl_count > 0 or gl_count > 0:
            messages.warning(request, f"⚠️ មិនអាចលុបជំនាញ '{name}' បានទេ ដោយសារមាន {gl_count} កម្រិតថ្នាក់ និង {cl_count} ថ្នាក់រៀនកំពុងប្រើប្រាស់!")
        else:
            track.delete()
            messages.success(request, f"🗑️ បានលុបជំនាញសិក្សា '{name}' ជោគជ័យ!")
    return redirect('grade_rules_manager')


@login_required
def api_academic_tracks(request):
    """API endpoint returning list of active tracks"""
    tracks = AcademicTrack.objects.all().order_by('order', 'id')
    data = [{'id': t.id, 'code': t.code, 'name_kh': t.name_kh, 'name_en': t.name_en or '', 'is_default': t.is_default} for t in tracks]
    return JsonResponse({'success': True, 'tracks': data})


# ----------------- GRADE ENROLLMENT OPTIONS (CUSTOM FIELDS) -----------------

@login_required
@role_required(['ADMIN'])
def grade_options_manager(request):
    """Admin Manager for Grade-Level Specific Enrollment Options"""
    grade_levels = GradeLevel.objects.prefetch_related('enrollment_options').all().order_by('order', 'grade_number', 'track')
    form = GradeEnrollmentOptionForm()
    
    selected_gl_id = request.GET.get('grade_level')
    selected_gl = None
    if selected_gl_id and str(selected_gl_id).isdigit():
        selected_gl = GradeLevel.objects.filter(id=int(selected_gl_id)).first()

    return render(request, 'academics/grade_options_manager.html', {
        'grade_levels': grade_levels,
        'selected_gl': selected_gl,
        'form': form,
    })


@login_required
@role_required(['ADMIN'])
def grade_option_save(request, pk=None):
    opt = get_object_or_404(GradeEnrollmentOption, pk=pk) if pk else None
    if request.method == 'POST':
        form = GradeEnrollmentOptionForm(request.POST, instance=opt)
        if form.is_valid():
            saved_opt = form.save()
            messages.success(request, f"🎉 បានរក្សាទុកជម្រើស '{saved_opt.label}' សម្រាប់ {saved_opt.grade_level.name} ជោគជ័យ!")
        else:
            for f, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"កំហុស [{f}]: {e}")
    return redirect('grade_options_manager')


@login_required
@role_required(['ADMIN'])
def grade_option_delete(request, pk):
    opt = get_object_or_404(GradeEnrollmentOption, pk=pk)
    if request.method == 'POST':
        label = opt.label
        gl_name = opt.grade_level.name
        opt.delete()
        messages.success(request, f"🗑️ បានលុបជម្រើស '{label}' ពី {gl_name} ជោគជ័យ!")
    return redirect('grade_options_manager')


@login_required
@role_required(['ADMIN'])
def grade_options_reorder(request):
    """AJAX POST to reorder options via drag and drop"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ordered_ids = data.get('ordered_ids', [])
            for idx, opt_id in enumerate(ordered_ids, start=1):
                GradeEnrollmentOption.objects.filter(id=opt_id).update(order=idx)
            return JsonResponse({'status': 'success', 'message': 'បានតម្រៀបលំដាប់លំដោយជោគជ័យ!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
@role_required(['ADMIN'])
def grade_option_update_width(request, pk):
    """AJAX POST to update column width (col-12, col-6, col-4, col-3)"""
    opt = get_object_or_404(GradeEnrollmentOption, pk=pk)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_width = int(data.get('col_width', 6))
            if new_width in [12, 6, 4, 3]:
                opt.col_width = new_width
                opt.save(update_fields=['col_width'])
                return JsonResponse({'status': 'success', 'col_width': opt.col_width})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)



# ----------------- CLASSROOM CRUD & BULK ACTIONS -----------------

@login_required
def classroom_list(request):
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
        else:
            found_year = AcademicYear.objects.filter(name=str(selected_year).strip()).first()
        if found_year:
            active_year = found_year
            try:
                request.session['active_academic_year_id'] = active_year.id
            except Exception:
                pass

    classrooms = Classroom.objects.select_related(
        'academic_year', 'homeroom_teacher', 'assembly_duty_teacher', 'class_monitor', 'vice_monitor'
    ).prefetch_related(
        'assigned_subjects__subject', 'assigned_subjects__teacher', 'students'
    ).all()
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    all_subjects = Subject.objects.all().order_by('order', 'id')
    
    if active_year:
        classrooms = classrooms.filter(academic_year=active_year)

    # Build a lookup of all GradeLevelRule (grade_level, track, subject_id) -> max_score
    rules_dict = {}
    for r in GradeLevelRule.objects.all():
        rules_dict[(r.grade_level, r.track, r.subject_id)] = r.max_score

    classroom_items = []
    total_students_school = 0
    total_female_school = 0
    total_capacity_school = 0
    classes_with_homeroom = 0
    
    grade_grouped = defaultdict(list)

    for c in classrooms:
        tot_stu = c.total_students
        fem_stu = c.female_students
        male_stu = max(0, tot_stu - fem_stu)
        
        total_students_school += tot_stu
        total_female_school += fem_stu
        if c.capacity:
            total_capacity_school += c.capacity
        if c.homeroom_teacher:
            classes_with_homeroom += 1
            
        grade_grouped[c.grade_level].append(c)

        # Build map of assigned subjects to their teacher and weekly hours
        cs_map = {}
        for cs in c.assigned_subjects.all():
            cs_map[cs.subject_id] = {
                'teacher': cs.teacher,
                'weekly_hours': cs.weekly_hours,
            }

        assigned_ids = set(cs_map.keys())
        
        # If none explicitly assigned yet, fallback to subjects in rules for this grade/track
        if not assigned_ids:
            assigned_ids = {sub.id for sub in all_subjects if (c.grade_level, c.track, sub.id) in rules_dict}

        subjects_with_meta = []
        tot_max = Decimal('0.00')
        active_cnt = 0
        for sub in all_subjects:
            is_active = sub.id in assigned_ids
            sc = rules_dict.get((c.grade_level, c.track, sub.id), Decimal('50.00'))
            cs_info = cs_map.get(sub.id, {})
            assigned_teacher = cs_info.get('teacher')
            weekly_hours = cs_info.get('weekly_hours') or (4 if sub.code in ['M', 'K'] else 2)
            
            if is_active:
                tot_max += sc
                active_cnt += 1
                
            subjects_with_meta.append({
                'subject': sub,
                'is_active': is_active,
                'max_score': sc,
                'assigned_teacher': assigned_teacher,
                'weekly_hours': weekly_hours,
            })

        occupancy_pct = round((tot_stu / c.capacity) * 100, 1) if (c.capacity and c.capacity > 0) else None

        classroom_items.append({
            'classroom': c,
            'assigned_ids': assigned_ids,
            'subjects_with_meta': subjects_with_meta,
            'active_subjects_count': active_cnt,
            'total_max_score': tot_max,
            'male_students': male_stu,
            'female_students': fem_stu,
            'total_students': tot_stu,
            'occupancy_pct': occupancy_pct,
        })

    # Grade Level Breakdown for Summary Modal
    grade_breakdown = []
    for g_num in sorted(grade_grouped.keys()):
        g_classes = grade_grouped[g_num]
        g_tot_stu = sum(c.total_students for c in g_classes)
        g_fem_stu = sum(c.female_students for c in g_classes)
        g_male_stu = max(0, g_tot_stu - g_fem_stu)
        g_cap = sum(c.capacity or 0 for c in g_classes)
        g_homeroom_cnt = sum(1 for c in g_classes if c.homeroom_teacher)
        
        grade_breakdown.append({
            'grade_number': g_num,
            'grade_name': f"ថ្នាក់ទី {g_num}",
            'classes_count': len(g_classes),
            'classes_list': g_classes,
            'students_count': g_tot_stu,
            'female_count': g_fem_stu,
            'male_count': g_male_stu,
            'total_capacity': g_cap,
            'homeroom_assigned_count': g_homeroom_cnt,
        })

    total_male_school = max(0, total_students_school - total_female_school)
    classes_without_homeroom = max(0, len(classroom_items) - classes_with_homeroom)

    grade_levels = list(GradeLevel.objects.all().order_by('order', 'grade_number', 'track', 'id'))
    if not grade_levels:
        for idx, (name, g_num, trk, ord_idx) in enumerate(DEFAULT_MOEYS_STREAMS):
            GradeLevel.objects.get_or_create(grade_number=g_num, track=trk, defaults={'name': name, 'order': ord_idx})
        grade_levels = list(GradeLevel.objects.all().order_by('order', 'grade_number', 'track', 'id'))

    return render(request, 'academics/classroom_list.html', {
        'classroom_items': classroom_items,
        'all_subjects': all_subjects,
        'academic_years': academic_years,
        'selected_year': str(active_year.id) if active_year else '',
        'active_year': active_year,
        'grade_levels': grade_levels,
        'total_classrooms_count': len(classroom_items),
        'total_students_school': total_students_school,
        'total_female_school': total_female_school,
        'total_male_school': total_male_school,
        'total_capacity_school': total_capacity_school,
        'classes_with_homeroom': classes_with_homeroom,
        'classes_without_homeroom': classes_without_homeroom,
        'grade_breakdown': grade_breakdown,
    })


KHMER_NUMERALS_MAP = {'0': '០', '1': '១', '2': '២', '3': '៣', '4': '៤', '5': '៥', '6': '៦', '7': '៧', '8': '៨', '9': '៩'}

def to_khmer_digits(n):
    return ''.join(KHMER_NUMERALS_MAP.get(c, c) for c in str(n))

def get_next_available_letters(academic_year, grade_number, count=1):
    existing_classes = Classroom.objects.filter(
        academic_year=academic_year,
        grade_level=grade_number
    )
    
    used_letters = set()
    for c in existing_classes:
        code = (c.code or '').strip().upper()
        # Extract trailing letters or letter after grade_number
        match = re.search(r'([A-Z]+)$', code)
        if match:
            used_letters.add(match.group(1))
        else:
            match2 = re.search(rf'^{grade_number}\s*[-_]?\s*([A-Z]+)', code)
            if match2:
                used_letters.add(match2.group(1))

    all_letters = list(string.ascii_uppercase)
    for first_char in string.ascii_uppercase:
        for second_char in string.ascii_uppercase:
            all_letters.append(f"{first_char}{second_char}")

    available = [L for L in all_letters if L not in used_letters]
    return available[:count]

def generate_classroom_name(grade_number, track, letter, grade_level_obj=None):
    kh_grade = to_khmer_digits(grade_number)
    if track == 'GENERAL' or track == Classroom.Track.GENERAL:
        return f"ថ្នាក់ទី {kh_grade}{letter}"
    elif track == 'SCIENCE' or track == Classroom.Track.SCIENCE:
        return f"ថ្នាក់ទី {kh_grade}{letter} វិទ្យាសាស្ត្រ"
    elif track == 'SOCIAL' or track == Classroom.Track.SOCIAL:
        return f"ថ្នាក់ទី {kh_grade}{letter} វិទ្យាសាស្ត្រសង្គម"
    else:
        track_name = grade_level_obj.name if grade_level_obj else track
        return f"ថ្នាក់ទី {kh_grade}{letter} ({track_name})"

@login_required
@role_required(['ADMIN'])
def classroom_auto_generate_preview(request):
    """
    AJAX endpoint to live-preview the classrooms that will be generated.
    Prevents letter clashes across streams of the same grade level (e.g. 11 Science & 11 Social).
    """
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    
    year_id = request.GET.get('academic_year_id') or request.GET.get('year')
    if year_id:
        found_year = AcademicYear.objects.filter(id=year_id).first()
        if found_year:
            active_year = found_year

    if not active_year:
        return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសឆ្នាំសិក្សា!'}, status=400)

    grade_level_id = request.GET.get('grade_level_id')
    grade_number = request.GET.get('grade_number')
    track = request.GET.get('track', 'GENERAL')
    grade_level_obj = None

    if grade_level_id and str(grade_level_id).isdigit():
        grade_level_obj = GradeLevel.objects.filter(id=int(grade_level_id)).first()
        if grade_level_obj:
            grade_number = grade_level_obj.grade_number
            track = grade_level_obj.track

    if not grade_number:
        return JsonResponse({'status': 'error', 'message': 'សូមជ្រើសរើសកម្រិតថ្នាក់!'}, status=400)

    try:
        grade_number = int(grade_number)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'កម្រិតថ្នាក់មិនត្រឹមត្រូវ!'}, status=400)

    try:
        count = int(request.GET.get('count', 1))
        if count < 1:
            count = 1
        elif count > 26:
            count = 26
    except ValueError:
        count = 1

    room_prefix = (request.GET.get('room_prefix') or '').strip()

    available_letters = get_next_available_letters(active_year, grade_number, count)
    
    existing_classes = Classroom.objects.filter(
        academic_year=active_year,
        grade_level=grade_number
    ).values('code', 'name', 'track')

    preview_items = []
    for letter in available_letters:
        code = f"{grade_number}{letter}"
        name = generate_classroom_name(grade_number, track, letter, grade_level_obj)
        room = f"{room_prefix} {letter}".strip() if room_prefix else f"បន្ទប់ {code}"
        preview_items.append({
            'code': code,
            'name': name,
            'grade_level': grade_number,
            'track': track,
            'room_number': room,
            'letter': letter,
        })

    return JsonResponse({
        'status': 'success',
        'academic_year': active_year.name,
        'grade_number': grade_number,
        'track': track,
        'count': len(preview_items),
        'preview': preview_items,
        'existing_classes': list(existing_classes),
    })

@login_required
@role_required(['ADMIN'])
def classroom_auto_generate(request):
    """
    Auto batch creates classrooms for a selected Grade Level and Count.
    Ensures conflict-free sequential letters across tracks for the same grade (e.g. 11A-11E Science -> 11F-11H Social).
    Auto-assigns standard subjects from GradeLevelRule.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid HTTP Method'}, status=405)

    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    year_id = request.POST.get('academic_year_id') or request.POST.get('year')
    if year_id:
        found_year = AcademicYear.objects.filter(id=year_id).first()
        if found_year:
            active_year = found_year

    if not active_year:
        messages.error(request, "⚠️ សូមជ្រើសរើសឆ្នាំសិក្សាជាមុនសិន!")
        return redirect('classroom_list')

    grade_level_id = request.POST.get('grade_level_id')
    grade_number = request.POST.get('grade_number')
    track = request.POST.get('track', 'GENERAL')
    grade_level_obj = None

    if grade_level_id and str(grade_level_id).isdigit():
        grade_level_obj = GradeLevel.objects.filter(id=int(grade_level_id)).first()
        if grade_level_obj:
            grade_number = grade_level_obj.grade_number
            track = grade_level_obj.track

    if not grade_number:
        messages.error(request, "⚠️ សូមជ្រើសរើសកម្រិតថ្នាក់!")
        return redirect('classroom_list')

    try:
        grade_number = int(grade_number)
    except ValueError:
        messages.error(request, "⚠️ កម្រិតថ្នាក់មិនត្រឹមត្រូវ!")
        return redirect('classroom_list')

    try:
        count = int(request.POST.get('count', 1))
        if count < 1:
            count = 1
        elif count > 26:
            count = 26
    except ValueError:
        count = 1

    capacity = None
    capacity_raw = (request.POST.get('capacity') or '').strip()
    if capacity_raw:
        try:
            capacity = max(1, int(capacity_raw))
        except ValueError:
            capacity = None

    room_prefix = (request.POST.get('room_prefix') or '').strip()

    available_letters = get_next_available_letters(active_year, grade_number, count)
    if not available_letters:
        messages.error(request, f"⚠️ មិនមានអក្សរកូដថ្នាក់ទំនេរសម្រាប់កម្រិតថ្នាក់ទី {grade_number} ឡើយ!")
        return redirect('classroom_list')

    created_classrooms = []
    with transaction.atomic():
        for letter in available_letters:
            code = f"{grade_number}{letter}"
            name = generate_classroom_name(grade_number, track, letter, grade_level_obj)
            room = f"{room_prefix} {letter}".strip() if room_prefix else f"បន្ទប់ {code}"
            
            classroom, created = Classroom.objects.get_or_create(
                code=code,
                academic_year=active_year,
                defaults={
                    'name': name,
                    'grade_level': grade_number,
                    'track': track,
                    'capacity': capacity,
                    'room_number': room,
                }
            )
            if not created:
                classroom.name = name
                classroom.grade_level = grade_number
                classroom.track = track
                classroom.capacity = capacity
                if room_prefix:
                    classroom.room_number = room
                classroom.save()

            # Auto-assign standard subjects from GradeLevelRule
            sub_ids = list(GradeLevelRule.objects.filter(
                grade_level=grade_number,
                track=track,
                weekly_hours__gt=0
            ).values_list('subject_id', flat=True))
            if not sub_ids:
                sub_ids = list(GradeLevelRule.objects.filter(
                    grade_level=grade_number,
                    track='GENERAL',
                    weekly_hours__gt=0
                ).values_list('subject_id', flat=True))

            if sub_ids:
                classroom.sync_assigned_subjects(sub_ids)

            created_classrooms.append(classroom)

    codes_str = ", ".join(c.code for c in created_classrooms)
    success_msg = f"🎉 បានបង្កើតថ្នាក់រៀនដោយស្វ័យប្រវត្តិចំនួន {len(created_classrooms)} ថ្នាក់ ({codes_str}) សម្រាប់ឆ្នាំសិក្សា {active_year.name} ដោយជោគជ័យ!"
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
        return JsonResponse({
            'status': 'success',
            'message': success_msg,
            'count': len(created_classrooms),
            'classrooms': [{'id': c.id, 'code': c.code, 'name': c.name} for c in created_classrooms],
        })

    messages.success(request, success_msg)
    return redirect('classroom_list')


@login_required
@role_required(['ADMIN'])
def classroom_manage_subjects(request, pk):
    """Admin ticks/unticks subjects for a specific classroom"""
    classroom = get_object_or_404(Classroom, pk=pk)
    if request.method == 'POST':
        subject_ids = request.POST.getlist('subject_ids')
        if not subject_ids:
            messages.warning(request, f"⚠️ សូមជ្រើសរើសមុខវិជ្ជាយ៉ាងហោចណាស់មួយសម្រាប់ថ្នាក់ {classroom.name}!")
            return redirect('classroom_list')
        
        classroom.sync_assigned_subjects(subject_ids)
        count = len(subject_ids)
        tot_max = classroom.get_total_max_score()
        messages.success(request, f"🎉 បានកំណត់មុខវិជ្ជាសម្រាប់ថ្នាក់ '{classroom.name}' ចំនួន {count} មុខវិជ្ជា (ពិន្ទុសរុបពេញ៖ {tot_max:g} ពិន្ទុ) ដោយជោគជ័យ!")
    return redirect('classroom_list')


@login_required
@role_required(['ADMIN'])
def classroom_create(request):
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    if request.method == 'POST':
        form = ClassroomForm(request.POST, academic_year=active_year)
        if form.is_valid():
            classroom = form.save()
            subject_ids = request.POST.getlist('subject_ids')
            if subject_ids:
                classroom.sync_assigned_subjects(subject_ids)
            else:
                # Default: assign standard subjects from GradeLevelRule for this grade/track
                default_sub_ids = list(GradeLevelRule.objects.filter(
                    grade_level=classroom.grade_level,
                    track=classroom.track
                ).values_list('subject_id', flat=True))
                if default_sub_ids:
                    classroom.sync_assigned_subjects(default_sub_ids)
            messages.success(request, f"បានបង្កើតថ្នាក់ {classroom.name} ជោគជ័យ!")
            return redirect('classroom_list')
    else:
        form = ClassroomForm(initial={'academic_year': active_year}, academic_year=active_year)
    
    all_subjects = Subject.objects.all().order_by('order', 'id')
    return render(request, 'academics/classroom_form.html', {
        'form': form,
        'all_subjects': all_subjects,
        'title': 'បង្កើតថ្នាក់រៀនថ្មី / Create Classroom'
    })


@login_required
@role_required(['ADMIN'])
def classroom_edit(request, pk):
    classroom = get_object_or_404(Classroom, pk=pk)
    if request.method == 'POST':
        form = ClassroomForm(request.POST, instance=classroom, academic_year=classroom.academic_year)
        if form.is_valid():
            form.save()
            subject_ids = request.POST.getlist('subject_ids')
            if subject_ids:
                classroom.sync_assigned_subjects(subject_ids)
            messages.success(request, f"បានកែប្រែថ្នាក់ {classroom.name} ជោគជ័យ!")
            return redirect('classroom_list')
    else:
        form = ClassroomForm(instance=classroom, academic_year=classroom.academic_year)
    
    all_subjects = Subject.objects.all().order_by('order', 'id')
    assigned_subject_ids = classroom.get_assigned_subject_ids()
    if not assigned_subject_ids:
        assigned_subject_ids = list(GradeLevelRule.objects.filter(
            grade_level=classroom.grade_level,
            track=classroom.track
        ).values_list('subject_id', flat=True))

    return render(request, 'academics/classroom_form.html', {
        'form': form,
        'classroom': classroom,
        'all_subjects': all_subjects,
        'assigned_subject_ids': assigned_subject_ids,
        'title': f'កែប្រែថ្នាក់ {classroom.name}'
    })


@login_required
@role_required(['ADMIN'])
def classroom_delete(request, pk):
    classroom = get_object_or_404(Classroom, pk=pk)
    if request.method == 'POST':
        name = classroom.name
        classroom.delete()
        messages.success(request, f"បានលុបថ្នាក់ {name} ដោយជោគជ័យ!")
    return redirect('classroom_list')


@login_required
@role_required(['ADMIN'])
def classroom_bulk_delete(request):
    """Bulk delete selected classrooms"""
    if request.method == 'POST':
        classroom_ids = request.POST.getlist('classroom_ids')
        if classroom_ids:
            deleted_count, _ = Classroom.objects.filter(id__in=classroom_ids).delete()
            messages.success(request, f"🗑️ បានលុបថ្នាក់រៀនដែលបានជ្រើសរើសចំនួន {deleted_count} ថ្នាក់ជោគជ័យ!")
        else:
            messages.warning(request, "សូមជ្រើសរើសថ្នាក់រៀនយ៉ាងតិចមួយដើម្បីលុប!")
    return redirect('classroom_list')


@login_required
@role_required(['ADMIN'])
def classroom_delete_all(request):
    """Delete all classrooms in the active academic year"""
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    if request.method == 'POST':
        if active_year:
            count = Classroom.objects.filter(academic_year=active_year).count()
            Classroom.objects.filter(academic_year=active_year).delete()
            messages.success(request, f"⚠️ បានលុបថ្នាក់រៀនទាំងអស់ ({count} ថ្នាក់) ក្នុងឆ្នាំសិក្សា {active_year.name} ជោគជ័យ!")
        else:
            count = Classroom.objects.count()
            Classroom.objects.all().delete()
            messages.success(request, f"⚠️ បានលុបថ្នាក់រៀនទាំងអស់ ({count} ថ្នាក់) ចេញពីប្រព័ន្ធជោគជ័យ!")
    return redirect('classroom_list')


@login_required
@role_required(['ADMIN'])
def classroom_restore_default(request):
    """Restore default 8 MoEYS classrooms (Grade 7 to 12) for active academic year"""
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    if not active_year:
        active_year, _ = AcademicYear.objects.get_or_create(
            name='2025-2026',
            defaults={'start_date': '2025-09-01', 'end_date': '2026-07-15', 'is_current': True}
        )

    teachers = list(Teacher.objects.filter(status='ACTIVE'))
    restored_cnt = 0

    with transaction.atomic():
        for idx, (code, name, grade, track, room) in enumerate(DEFAULT_MOEYS_CLASSROOMS):
            homeroom = teachers[idx % len(teachers)] if teachers else None
            cls_obj, _ = Classroom.objects.update_or_create(
                code=code,
                academic_year=active_year,
                defaults={
                    'name': name,
                    'grade_level': grade,
                    'track': track,
                    'room_number': room,
                    'capacity': 40,
                    'homeroom_teacher': homeroom
                }
            )
            # Assign standard subjects from rules
            sub_ids = list(GradeLevelRule.objects.filter(
                grade_level=cls_obj.grade_level,
                track=cls_obj.track
            ).values_list('subject_id', flat=True))
            if sub_ids:
                cls_obj.sync_assigned_subjects(sub_ids)
            restored_cnt += 1

    messages.success(request, f"🎉 បានស្តារឡើងវិញនូវថ្នាក់រៀនលំនាំដើម MoEYS ទាំង {restored_cnt} ថ្នាក់ (ថ្នាក់ទី៧ ដល់ទី១២) សម្រាប់ឆ្នាំសិក្សា {active_year.name} ដោយជោគជ័យ!")
    return redirect('classroom_list')



# ----------------- SUBJECTS MANAGEMENT & RESTORE -----------------

@login_required
def subject_list(request):
    subjects = Subject.objects.all().order_by('order', 'id')
    return render(request, 'academics/subject_list.html', {'subjects': subjects})


@login_required
@role_required(['ADMIN'])
def subject_create(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            subject = form.save()
            
            # Automatically initialize GradeLevelRule for the new subject across all existing GradeLevels
            for gl in GradeLevel.objects.all():
                GradeLevelRule.objects.get_or_create(
                    grade_level=gl.grade_number,
                    track=gl.track,
                    subject=subject,
                    defaults={'max_score': Decimal('50.00'), 'order': subject.order}
                )

            messages.success(request, f"🎉 បានបង្កើតមុខវិជ្ជាថ្មី '{subject.name_kh}' ({subject.code}) ជោគជ័យ! និងបានបន្ថែមទៅក្នុងតារាងច្បាប់ពិន្ទុ (Scoring Matrix) គ្រប់កម្រិតថ្នាក់រួចរាល់។")
            return redirect('subject_list')
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"⚠️ កំហុស [{field}]: {err}")
    else:
        form = SubjectForm()
    return render(request, 'academics/subject_form.html', {'form': form, 'title': 'បង្កើតមុខវិជ្ជាថ្មី / Create Subject'})


@login_required
@role_required(['ADMIN'])
def subject_edit(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, f"បានកែប្រែមុខវិជ្ជា {subject.name_kh} ជោគជ័យ!")
            return redirect('subject_list')
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"⚠️ កំហុស [{field}]: {err}")
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'academics/subject_form.html', {'form': form, 'title': f'កែប្រែមុខវិជ្ជា {subject.name_kh}', 'subject': subject})


@login_required
@role_required(['ADMIN'])
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        name = subject.name_kh
        subject.delete()
        messages.success(request, f"បានលុបមុខវិជ្ជា {name} ដោយជោគជ័យ!")
    return redirect('subject_list')


@login_required
@role_required(['ADMIN'])
def subject_restore_default(request):
    """Restore/Upsert the official 14 MoEYS subjects with exact short codes."""
    with transaction.atomic():
        restored_cnt = 0
        for name_kh, name_en, short_code, credit, color, sort_order in DEFAULT_MOEYS_SUBJECTS:
            sub = Subject.objects.filter(code=short_code).first() or Subject.objects.filter(name_kh=name_kh).first()
            if sub:
                sub.name_kh = name_kh
                sub.name_en = name_en
                sub.code = short_code
                sub.credit = credit
                sub.color_code = color
                sub.order = sort_order
                sub.save()
            else:
                Subject.objects.create(
                    name_kh=name_kh,
                    name_en=name_en,
                    code=short_code,
                    credit=credit,
                    color_code=color,
                    order=sort_order
                )
            restored_cnt += 1

    messages.success(request, f"🎉 បានស្តារឡើងវិញនូវមុខវិជ្ជា និងអក្សរកាត់លំនាំដើម MoEYS ទាំង {restored_cnt} មុខវិជ្ជា ដោយជោគជ័យ!")
    return redirect('subject_list')


# ----------------- SCORING RULES MANAGER & RESTORE -----------------

@login_required
@role_required(['ADMIN'])
def grade_rules_manager(request):
    """
    Admin Scoring Rules Matrix View
    Displays all dynamic Grade Levels across ALL subjects in exact order.
    """
    # Load all subjects in exact ordered sequence
    subjects = Subject.objects.all().order_by('order', 'id')
    if not subjects.exists():
        subject_restore_default(request)
        subjects = Subject.objects.all().order_by('order', 'id')

    # Ensure default streams exist in GradeLevel model if empty
    if not GradeLevel.objects.exists():
        for name, g_num, trk, ord_idx in DEFAULT_MOEYS_STREAMS:
            GradeLevel.objects.create(
                name=name,
                grade_number=g_num,
                track=trk,
                order=ord_idx
            )

    grade_levels = GradeLevel.objects.all().order_by('order', 'grade_number', 'track', 'id')

    if request.method == 'POST' and 'update_rules' in request.POST:
        updated_cnt = 0
        for gl in grade_levels:
            g = gl.grade_number
            t = gl.track
            for sub in subjects:
                input_key = f"score_{g}_{t}_{sub.id}"
                if input_key in request.POST:
                    val_str = request.POST.get(input_key, '').strip()
                    if val_str:
                        try:
                            score_val = Decimal(val_str)
                            if score_val > 0:
                                GradeLevelRule.objects.update_or_create(
                                    grade_level=g,
                                    track=t,
                                    subject=sub,
                                    defaults={'max_score': score_val, 'order': sub.order}
                                )
                                updated_cnt += 1
                            else:
                                GradeLevelRule.objects.filter(grade_level=g, track=t, subject=sub).delete()
                        except Exception:
                            pass
                    else:
                        GradeLevelRule.objects.filter(grade_level=g, track=t, subject=sub).delete()

        messages.success(request, f"✅ បានធ្វើបច្ចុប្បន្នភាពច្បាប់ពិន្ទុអតិបរមា {updated_cnt} មុខវិជ្ជាជោគជ័យ!")
        return redirect('grade_rules_manager')

    # Load matrix data
    rules_dict = {}
    for r in GradeLevelRule.objects.select_related('subject').all():
        rules_dict[(r.grade_level, r.track, r.subject_id)] = r.max_score

    streams_data = []
    for gl in grade_levels:
        g = gl.grade_number
        t = gl.track
        sub_scores = []
        total_max = Decimal('0.00')
        for sub in subjects:
            sc = rules_dict.get((g, t, sub.id))
            if sc:
                total_max += sc
            sub_scores.append({
                'subject': sub,
                'max_score': sc,
            })
        streams_data.append({
            'id': gl.id,
            'grade': g,
            'track': t,
            'name': gl.name,
            'total_max': total_max,
            'sub_scores': sub_scores,
        })

    has_saved_custom = SavedDefaultConfig.objects.filter(key='custom_scoring_rules').exists()
    gl_form = GradeLevelForm()
    all_tracks = AcademicTrack.objects.all().order_by('order', 'id')
    track_form = AcademicTrackForm()

    return render(request, 'academics/grade_rules_manager.html', {
        'subjects': subjects,
        'streams_data': streams_data,
        'has_saved_custom': has_saved_custom,
        'gl_form': gl_form,
        'all_tracks': all_tracks,
        'track_form': track_form,
    })


@login_required
@role_required(['ADMIN'])
def save_current_as_default(request):
    """
    Saves the current scoring rules configuration as the new custom default preset.
    """
    if request.method == 'POST':
        rules_data = []
        for r in GradeLevelRule.objects.select_related('subject').all():
            rules_data.append({
                'grade_level': r.grade_level,
                'track': r.track,
                'subject_code': r.subject.code,
                'subject_name_kh': r.subject.name_kh,
                'max_score': str(r.max_score),
                'order': r.order,
            })
        
        streams_data = []
        for gl in GradeLevel.objects.all():
            streams_data.append({
                'name': gl.name,
                'grade_number': gl.grade_number,
                'track': gl.track,
                'order': gl.order
            })

        SavedDefaultConfig.objects.update_or_create(
            key='custom_scoring_rules',
            defaults={'data': {'rules': rules_data, 'streams': streams_data}}
        )
        messages.success(request, "💾 ជោគជ័យ! បានរក្សាទុកការកំណត់ច្បាប់ពិន្ទុបច្ចុប្បន្នជា Default សម្រាប់ប្រើប្រាស់ឡើងវិញគ្រប់ពេល។")
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def restore_saved_custom_default(request):
    """
    Restores the custom default preset saved earlier by the Admin.
    """
    cfg = SavedDefaultConfig.objects.filter(key='custom_scoring_rules').first()
    if not cfg or not cfg.data or 'rules' not in cfg.data:
        messages.warning(request, "មិនទាន់មានទិន្នន័យ Custom Default ត្រូវបានរក្សាទុកនៅឡើយទេ។ ប្រព័ន្ធនឹងស្តារតាម MoEYS Default!")
        return redirect('reset_grade_rules_to_moeys')

    with transaction.atomic():
        # Restore custom streams if present
        if 'streams' in cfg.data:
            GradeLevel.objects.all().delete()
            for s in cfg.data['streams']:
                GradeLevel.objects.create(
                    name=s['name'],
                    grade_number=s['grade_number'],
                    track=s['track'],
                    order=s['order']
                )

        GradeLevelRule.objects.all().delete()
        count = 0
        for item in cfg.data['rules']:
            sub = Subject.objects.filter(code=item['subject_code']).first() or Subject.objects.filter(name_kh=item['subject_name_kh']).first()
            if sub:
                GradeLevelRule.objects.create(
                    grade_level=item['grade_level'],
                    track=item['track'],
                    subject=sub,
                    max_score=Decimal(item['max_score']),
                    order=item.get('order', sub.order)
                )
                count += 1

    messages.success(request, f"🎉 បានស្តារឡើងវិញនូវច្បាប់ពិន្ទុលំនាំដើមដែលលោកអ្នកបាន Save ទុក ({count} មុខវិជ្ជា) ដោយជោគជ័យ!")
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def grade_rules_delete_all(request):
    """Delete all grade scoring rules"""
    if request.method == 'POST':
        count = GradeLevelRule.objects.count()
        GradeLevelRule.objects.all().delete()
        messages.success(request, f"⚠️ បានលុបច្បាប់ពិន្ទុទាំងអស់ ({count} ច្បាប់) ចេញពីប្រព័ន្ធជោគជ័យ!")
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def reset_grade_rules_to_moeys(request):
    """Resets all GradeLevel and GradeLevelRule records to the standard MoEYS matrix."""
    with transaction.atomic():
        # 1. First ensure only the clean 14 subjects exist
        Subject.objects.exclude(code__in=OFFICIAL_CODES).delete()
        for name_kh, name_en, short_code, credit, color, sort_order in DEFAULT_MOEYS_SUBJECTS:
            sub = Subject.objects.filter(code=short_code).first() or Subject.objects.filter(name_kh=name_kh).first()
            if sub:
                sub.name_kh = name_kh
                sub.name_en = name_en
                sub.code = short_code
                sub.credit = credit
                sub.color_code = color
                sub.order = sort_order
                sub.save()
            else:
                Subject.objects.create(
                    name_kh=name_kh,
                    name_en=name_en,
                    code=short_code,
                    credit=credit,
                    color_code=color,
                    order=sort_order
                )

        # 3. Reset Scoring Rules
        GradeLevelRule.objects.all().delete()
        created_count = 0
        for (g, track), sub_map in MOEYS_SCORING_RULES.items():
            for sub_name, max_sc in sub_map.items():
                sub = Subject.objects.filter(name_kh=sub_name).first()
                if sub:
                    GradeLevelRule.objects.create(
                        grade_level=g,
                        track=track,
                        subject=sub,
                        max_score=Decimal(str(max_sc)),
                        order=sub.order
                    )
                    created_count += 1

    messages.success(request, f"🎉 បានកំណត់ឡើងវិញនូវច្បាប់ពិន្ទុស្តង់ដារ MoEYS ទាំង ៨ កម្រិតថ្នាក់ ({created_count} មុខវិជ្ជា) ដោយជោគជ័យ!")
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def grade_rules_export_excel(request):
    """
    Exports the MoEYS Scoring Rules Matrix to a styled Excel (.xlsx) file.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scoring Rules Matrix"

    # Fonts & Fills
    title_font = Font(name="Khmer OS Siemreap", size=14, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Khmer OS Siemreap", size=10, italic=True, color="475569")
    header_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Khmer OS Siemreap", size=10)
    bold_font = Font(name="Khmer OS Siemreap", size=10, bold=True)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    total_header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    total_cell_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Block
    ws.merge_cells('A1:Q1')
    ws['A1'] = "តារាងច្បាប់ពិន្ទុអតិបរមាតាមកម្រិតថ្នាក់ (MoEYS Scoring Rules Matrix)"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:Q2')
    export_time = timezone.now().strftime("%d/%m/%Y %H:%M")
    ws['A2'] = f"កាលបរិច្ឆេទ Export: {export_time} | ប្រព័ន្ធគ្រប់គ្រងសាលារៀន SchoolSM"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    subjects = list(Subject.objects.all().order_by('order', 'id'))
    headers = ["ល.រ", "កម្រិតថ្នាក់ (Grade Level)", "លេខកម្រិត", "ជំនាញ (Track)"]
    for sub in subjects:
        headers.append(f"{sub.name_kh} ({sub.code})")
    headers.append("ពិន្ទុសរុបពេញ (Total)")

    ws.append([]) # Row 3 empty
    ws.append(headers) # Row 4
    ws.row_dimensions[4].height = 28

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        if col_num == len(headers):
            cell.fill = total_header_fill
        else:
            cell.fill = header_fill

    # Data Rows
    grade_levels = list(GradeLevel.objects.all().order_by('order', 'grade_number', 'track', 'id'))
    rules_dict = {}
    for r in GradeLevelRule.objects.select_related('subject').all():
        rules_dict[(r.grade_level, r.track, r.subject_id)] = r.max_score

    for idx, gl in enumerate(grade_levels, start=1):
        row_num = 4 + idx
        g = gl.grade_number
        t = gl.track
        row_data = [idx, gl.name, g, t]
        total_max = Decimal('0.00')

        for sub in subjects:
            sc = rules_dict.get((g, t, sub.id))
            if sc and sc > 0:
                total_max += sc
                row_data.append(float(sc))
            else:
                row_data.append("-")
        row_data.append(float(total_max))

        ws.append(row_data)
        ws.row_dimensions[row_num].height = 22

        # Style data cells
        is_alt = (idx % 2 == 0)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.font = bold_font if col_num in [1, 2, len(headers)] else data_font

            if col_num == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if col_num == len(headers):
                cell.fill = total_cell_fill
            elif is_alt:
                cell.fill = alt_row_fill

    # Auto Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row in [1, 2]:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"moeys_scoring_rules_matrix_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@role_required(['ADMIN'])
def grade_rules_export_csv(request):
    """
    Exports the MoEYS Scoring Rules Matrix as a UTF-8 BOM CSV file for Excel compatibility.
    """
    import csv

    subjects = list(Subject.objects.all().order_by('order', 'id'))
    grade_levels = list(GradeLevel.objects.all().order_by('order', 'grade_number', 'track', 'id'))
    rules_dict = {}
    for r in GradeLevelRule.objects.select_related('subject').all():
        rules_dict[(r.grade_level, r.track, r.subject_id)] = r.max_score

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f"moeys_scoring_rules_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Write UTF-8 BOM
    response.write('\ufeff'.encode('utf-8'))

    writer = csv.writer(response)
    header = ["No", "Grade_Level_Name", "Grade_Number", "Track"]
    for sub in subjects:
        header.append(f"{sub.name_kh} ({sub.code})")
    header.append("Total_Max_Score")
    writer.writerow(header)

    for idx, gl in enumerate(grade_levels, start=1):
        g = gl.grade_number
        t = gl.track
        row = [idx, gl.name, g, t]
        total_max = Decimal('0.00')
        for sub in subjects:
            sc = rules_dict.get((g, t, sub.id))
            if sc and sc > 0:
                total_max += sc
                row.append(str(sc))
            else:
                row.append("0")
        row.append(str(total_max))
        writer.writerow(row)

    return response


@login_required
@role_required(['ADMIN'])
def grade_rules_import(request):
    """
    Bulk imports/updates Scoring Rules from an uploaded Excel (.xlsx, .xls) or CSV (.csv) file.
    """
    import csv
    import io
    if request.method != 'POST':
        return redirect('grade_rules_manager')

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        messages.error(request, "⚠️ សូមជ្រើសរើសឯកសារ Excel (.xlsx) ឬ CSV (.csv) ដើម្បី Upload!")
        return redirect('grade_rules_manager')

    fname = uploaded_file.name.lower()
    subjects = list(Subject.objects.all().order_by('order', 'id'))
    sub_by_code = {s.code.upper(): s for s in subjects if s.code}
    sub_by_name = {s.name_kh.strip(): s for s in subjects if s.name_kh}

    rows_data = []

    try:
        if fname.endswith('.xlsx') or fname.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if any(row):
                    rows_data.append([str(c).strip() if c is not None else '' for c in row])
        elif fname.endswith('.csv'):
            content = uploaded_file.read().decode('utf-8-sig', errors='ignore')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                if any(row):
                    rows_data.append([str(c).strip() for c in row])
        else:
            messages.error(request, "⚠️ ប្រព័ន្ធគាំទ្រតែឯកសារប្រភេទ .xlsx, .xls ឬ .csv ប៉ុណ្ណោះ!")
            return redirect('grade_rules_manager')
    except Exception as e:
        messages.error(request, f"❌ កំហុសក្នុងការអានឯកសារ៖ {str(e)}")
        return redirect('grade_rules_manager')

    if not rows_data:
        messages.error(request, "⚠️ ឯកសារទទេ មិនមានទិន្នន័យឡើយ!")
        return redirect('grade_rules_manager')

    # Locate Header Row
    header_idx = -1
    for idx, r in enumerate(rows_data[:10]):
        row_str = " ".join(r).lower()
        if "កម្រិតថ្នាក់" in row_str or "grade" in row_str or "track" in row_str or "ជំនាញ" in row_str:
            header_idx = idx
            break

    if header_idx == -1:
        header_idx = 0

    header_row = rows_data[header_idx]
    
    # Map column indexes to subjects
    col_subject_map = {}
    grade_col = -1
    grade_num_col = -1
    track_col = -1

    for c_idx, h_text in enumerate(header_row):
        h_clean = h_text.strip()
        h_lower = h_clean.lower()
        if "កម្រិតថ្នាក់" in h_lower or "grade_level_name" in h_lower:
            grade_col = c_idx
        elif "លេខកម្រិត" in h_lower or "grade_number" in h_lower or h_lower == "grade":
            grade_num_col = c_idx
        elif "ជំនាញ" in h_lower or "track" in h_lower:
            track_col = c_idx
        else:
            # Check subject match by code or name
            matched_sub = None
            for code, sub_obj in sub_by_code.items():
                if f"({code})" in h_clean.upper() or h_clean.upper() == code:
                    matched_sub = sub_obj
                    break
            if not matched_sub:
                for name, sub_obj in sub_by_name.items():
                    if name in h_clean:
                        matched_sub = sub_obj
                        break
            if matched_sub:
                col_subject_map[c_idx] = matched_sub

    updated_count = 0
    with transaction.atomic():
        for r_idx in range(header_idx + 1, len(rows_data)):
            r = rows_data[r_idx]
            if len(r) <= max(grade_col, track_col, 1):
                continue

            grade_name = r[grade_col] if grade_col >= 0 and grade_col < len(r) else ''
            g_num_str = r[grade_num_col] if grade_num_col >= 0 and grade_num_col < len(r) else ''
            track_str = r[track_col].upper() if track_col >= 0 and track_col < len(r) else 'GENERAL'

            # Parse Grade Number
            g_num = None
            if g_num_str and g_num_str.isdigit():
                g_num = int(g_num_str)
            else:
                for digit in ['12', '11', '10', '9', '8', '7']:
                    if digit in grade_name:
                        g_num = int(digit)
                        break

            if not g_num:
                continue

            # Standardize Track
            if 'SCIENCE' in track_str or 'វិទ្យាសាស្ត្រ' in track_str:
                if 'សង្គម' in track_str or 'SOCIAL' in track_str:
                    track_val = 'SOCIAL'
                else:
                    track_val = 'SCIENCE'
            elif 'SOCIAL' in track_str or 'សង្គម' in track_str:
                track_val = 'SOCIAL'
            else:
                track_val = 'GENERAL'

            # Update or create GradeLevel
            if grade_name:
                GradeLevel.objects.get_or_create(
                    grade_number=g_num,
                    track=track_val,
                    defaults={'name': grade_name}
                )

            # Update Subject Max Scores
            for c_idx, sub_obj in col_subject_map.items():
                if c_idx < len(r):
                    val_str = r[c_idx].strip()
                    if val_str and val_str not in ['-', 'N/A', 'n/a']:
                        try:
                            score_val = Decimal(val_str)
                            if score_val > 0:
                                GradeLevelRule.objects.update_or_create(
                                    grade_level=g_num,
                                    track=track_val,
                                    subject=sub_obj,
                                    defaults={'max_score': score_val, 'order': sub_obj.order}
                                )
                                updated_count += 1
                            else:
                                GradeLevelRule.objects.filter(grade_level=g_num, track=track_val, subject=sub_obj).delete()
                        except Exception:
                            pass

    messages.success(request, f"🎉 ជោគជ័យ! បាន Import និងធ្វើបច្ចុប្បន្នភាពច្បាប់ពិន្ទុចំនួន {updated_count} មុខវិជ្ជាពីឯកសារ {uploaded_file.name}!")
    return redirect('grade_rules_manager')


@login_required
@role_required(['ADMIN'])
def master_restore_defaults(request):
    """
    Master 1-Click Restore:
    1. Purges obsolete subjects, keeps exactly 14 official subjects (R, D, K, I, G, H, M, Es, P, C, B, He, Ec, E)
    2. Restores standard 8 GradeLevel records & scoring rules matrix
    3. Restores 8 default classrooms for active academic year
    """
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    if not active_year:
        active_year, _ = AcademicYear.objects.get_or_create(
            name='2025-2026',
            defaults={'start_date': '2025-09-01', 'end_date': '2026-07-15', 'is_current': True}
        )

    with transaction.atomic():
        # 1. Clean Subjects
        Subject.objects.exclude(code__in=OFFICIAL_CODES).delete()
        for name_kh, name_en, short_code, credit, color, sort_order in DEFAULT_MOEYS_SUBJECTS:
            sub = Subject.objects.filter(code=short_code).first() or Subject.objects.filter(name_kh=name_kh).first()
            if sub:
                sub.name_kh = name_kh
                sub.name_en = name_en
                sub.code = short_code
                sub.credit = credit
                sub.color_code = color
                sub.order = sort_order
                sub.save()
            else:
                Subject.objects.create(
                    name_kh=name_kh,
                    name_en=name_en,
                    code=short_code,
                    credit=credit,
                    color_code=color,
                    order=sort_order
                )

        # 2. Grade Levels
        GradeLevel.objects.all().delete()
        for name, g_num, trk, ord_idx in DEFAULT_MOEYS_STREAMS:
            GradeLevel.objects.create(
                name=name,
                grade_number=g_num,
                track=trk,
                order=ord_idx
            )

        # 3. Scoring Rules
        GradeLevelRule.objects.all().delete()
        for (g, track), sub_map in MOEYS_SCORING_RULES.items():
            for sub_name, max_sc in sub_map.items():
                sub = Subject.objects.filter(name_kh=sub_name).first()
                if sub:
                    GradeLevelRule.objects.create(
                        grade_level=g,
                        track=track,
                        subject=sub,
                        max_score=Decimal(str(max_sc)),
                        order=sub.order
                    )

        # 4. Classrooms
        teachers = list(Teacher.objects.filter(status='ACTIVE'))
        for idx, (code, name, grade, track, room) in enumerate(DEFAULT_MOEYS_CLASSROOMS):
            homeroom = teachers[idx % len(teachers)] if teachers else None
            cls_obj, _ = Classroom.objects.update_or_create(
                code=code,
                academic_year=active_year,
                defaults={
                    'name': name,
                    'grade_level': grade,
                    'track': track,
                    'room_number': room,
                    'capacity': 40,
                    'homeroom_teacher': homeroom
                }
            )
            # Assign standard subjects from rules
            sub_ids = list(GradeLevelRule.objects.filter(
                grade_level=cls_obj.grade_level,
                track=cls_obj.track
            ).values_list('subject_id', flat=True))
            if sub_ids:
                cls_obj.sync_assigned_subjects(sub_ids)

    messages.success(request, f"🎉 ជោគជ័យ! បានស្តារប្រព័ន្ធទាំងមូលឡើងវិញទៅតាមស្តង់ដារលំនាំដើម MoEYS សម្រាប់ឆ្នាំសិក្សា {active_year.name}!")
    return redirect('classroom_list')


# ----------------- MASTER TIMETABLE MATRIX & GENERATION -----------------

STANDARD_PERIOD_TIMES = {
    1: (datetime.time(7, 0), datetime.time(7, 50)),
    2: (datetime.time(7, 55), datetime.time(8, 45)),
    3: (datetime.time(9, 5), datetime.time(9, 55)),
    4: (datetime.time(10, 0), datetime.time(10, 50)),
    5: (datetime.time(13, 0), datetime.time(13, 50)),
    6: (datetime.time(13, 55), datetime.time(14, 45)),
    7: (datetime.time(15, 5), datetime.time(15, 55)),
    8: (datetime.time(16, 0), datetime.time(16, 50)),
}

DAYS_OF_WEEK = [
    {'num': 1, 'name_kh': 'ច័ន្ទ', 'name_en': 'Monday'},
    {'num': 2, 'name_kh': 'អង្គារ', 'name_en': 'Tuesday'},
    {'num': 3, 'name_kh': 'ពុធ', 'name_en': 'Wednesday'},
    {'num': 4, 'name_kh': 'ព្រហស្បតិ៍', 'name_en': 'Thursday'},
    {'num': 5, 'name_kh': 'សុក្រ', 'name_en': 'Friday'},
    {'num': 6, 'name_kh': 'សៅរ៍', 'name_en': 'Saturday'},
]

PERIODS_LIST = [1, 2, 3, 4, 5, 6, 7, 8]


@login_required
def timetable_view(request):
    """
    Master Timetable Matrix (កាលវិភាគរួម) View.
    Displays school-wide timetable matrix with rows as classrooms and columns as Day x Periods 1-8.
    Supports Undo/Redo, filtered teacher/subject assignment selection, real-time clash detection,
    and conditional formatting for required vs scheduled weekly hours.
    Strictly isolated per Academic Year!
    """
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
        else:
            found_year = AcademicYear.objects.filter(name=str(selected_year).strip()).first()
        if found_year:
            active_year = found_year
            try:
                request.session['active_academic_year_id'] = active_year.id
            except Exception:
                pass

    academic_years = list(AcademicYear.objects.all().order_by('-start_date'))
    classrooms = list(Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code'))
    teachers = list(Teacher.objects.filter(status='ACTIVE').order_by('khmer_name'))
    subjects = list(Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id'))
    subjects_by_id = {s.id: s for s in subjects}
    teachers_by_id = {t.id: t for t in teachers}
    
    # Existing timetable entries for active academic year
    timetables_qs = Timetable.objects.filter(classroom__academic_year=active_year).select_related('classroom', 'subject', 'teacher') if active_year else Timetable.objects.select_related('classroom', 'subject', 'teacher').all()
    timetables = list(timetables_qs)

    timetables_by_classroom = defaultdict(list)
    timetables_by_teacher = defaultdict(int)
    for entry in timetables:
        timetables_by_classroom[entry.classroom_id].append(entry)
        if entry.teacher_id:
            timetables_by_teacher[entry.teacher_id] += 1
    
    # Pre-fetch ClassSubject assignments (Only teachers assigned to this class and subject in this academic year)
    class_subject_assignments_qs = ClassSubject.objects.filter(
        classroom__academic_year=active_year,
        teacher__isnull=False
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('classroom', 'subject', 'teacher') if active_year else ClassSubject.objects.filter(
        teacher__isnull=False
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('classroom', 'subject', 'teacher')
    class_subject_assignments = list(class_subject_assignments_qs)

    cs_by_teacher = defaultdict(list)
    cs_pairs_set = set()
    for cs in class_subject_assignments:
        if cs.teacher_id:
            cs_by_teacher[cs.teacher_id].append(cs)
            cs_pairs_set.add((cs.subject_id, cs.teacher_id))

    # Build unique sequential teacher-subject codes (e.g. K1, K2, M1, M2, P1, P2...)
    distinct_assignments = sorted(list(cs_pairs_set), key=lambda x: (x[0] or 0, x[1] or 0))

    teacher_subject_code_map = {}
    subject_teacher_counters = {}

    for (s_id, t_id) in distinct_assignments:
        sub = subjects_by_id.get(s_id)
        sub_code = sub.code if sub and sub.code else 'S'
        
        if s_id not in subject_teacher_counters:
            subject_teacher_counters[s_id] = 1
        else:
            subject_teacher_counters[s_id] += 1
            
        code = f"{sub_code}{subject_teacher_counters[s_id]}"
        teacher_subject_code_map[(s_id, t_id)] = code

    # Fallback assignment for any active teachers/subjects
    for s in subjects:
        for t in teachers:
            if (s.id, t.id) not in teacher_subject_code_map:
                if s.id not in subject_teacher_counters:
                    subject_teacher_counters[s.id] = 1
                else:
                    subject_teacher_counters[s.id] += 1
                teacher_subject_code_map[(s.id, t.id)] = f"{s.code or 'S'}{subject_teacher_counters[s.id]}"

    # Build options per classroom with slot_code (e.g. K1, M2...)
    class_options_map = {}
    for cs in class_subject_assignments:
        c_id = cs.classroom_id
        if c_id not in class_options_map:
            class_options_map[c_id] = []
        sub_code = cs.subject.code if cs.subject and cs.subject.code else 'S'
        slot_code = teacher_subject_code_map.get((cs.subject_id, cs.teacher_id), f"{sub_code}1")
        tch_name = (cs.teacher.khmer_name or cs.teacher.name or '') if cs.teacher else ''
        class_options_map[c_id].append({
            'subject_id': cs.subject.id if cs.subject else None,
            'subject_code': sub_code,
            'slot_code': slot_code,
            'subject_name': cs.subject.name_kh if cs.subject else '',
            'subject_color': cs.subject.color_code if cs.subject and cs.subject.color_code else '#4f46e5',
            'category': cs.subject.category if cs.subject else 'GENERAL',
            'teacher_id': cs.teacher.id if cs.teacher else None,
            'teacher_name': tch_name,
            'teacher_short': tch_name[:6],
        })

    # Pre-fetch required hours from GradeLevelRule
    grade_rules = list(GradeLevelRule.objects.filter(
        weekly_hours__gt=0
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('subject'))

    requirements_map = {}
    for r in grade_rules:
        k = (r.grade_level, r.track)
        if k not in requirements_map:
            requirements_map[k] = {}
        requirements_map[k][r.subject_id] = r.weekly_hours

    # Build master grid records for each classroom
    classrooms_data = []
    matrix_state = {}

    for cls in classrooms:
        cls_reqs = requirements_map.get((cls.grade_level, cls.track), {})
        if not cls_reqs:
            cls_reqs = requirements_map.get((cls.grade_level, 'GENERAL'), {})
        
        total_req_hours = sum(cls_reqs.values())
        
        # Classroom slot map: (day_of_week, period_number) -> slot info
        cls_slots = {}
        cls_entries = timetables_by_classroom.get(cls.id, [])
        
        for entry in cls_entries:
            sub_name = entry.subject.name_kh if entry.subject else ''
            sub_code = entry.subject.code if entry.subject and entry.subject.code else ''
            sub_color = entry.subject.color_code if entry.subject and entry.subject.color_code else '#4f46e5'
            sub_category = entry.subject.category if entry.subject else 'GENERAL'
            tch_name = (entry.teacher.khmer_name or entry.teacher.name or '') if entry.teacher else ''
            tch_short = tch_name[:6]
            slot_code = teacher_subject_code_map.get(
                (entry.subject_id, entry.teacher_id), 
                sub_code
            )
            cls_slots[(entry.day_of_week, entry.period_number)] = {
                'id': entry.id,
                'subject_id': entry.subject_id,
                'subject_name': sub_name,
                'subject_code': sub_code,
                'slot_code': slot_code,
                'subject_color': sub_color,
                'category': sub_category,
                'teacher_id': entry.teacher_id,
                'teacher_name': tch_name,
                'teacher_short': tch_short,
            }
            matrix_state[f"{cls.id}_{entry.day_of_week}_{entry.period_number}"] = {
                'subject_id': entry.subject_id,
                'teacher_id': entry.teacher_id,
                'slot_code': slot_code,
            }

        # Build grid cells per day and period
        days_grid = []
        for day in DAYS_OF_WEEK:
            day_periods = []
            for p in PERIODS_LIST:
                slot_info = cls_slots.get((day['num'], p))
                day_periods.append({
                    'period': p,
                    'slot': slot_info,
                    'cell_id': f"cell_{cls.id}_{day['num']}_{p}",
                })
            days_grid.append({
                'day': day,
                'periods': day_periods,
            })

        classrooms_data.append({
            'classroom': cls,
            'total_req_hours': total_req_hours,
            'scheduled_hours': len(cls_entries),
            'days_grid': days_grid,
        })

    # Merge session-stored blocked slots for active academic year into matrix_state
    session_key = f"blocked_slots_{active_year.id if active_year else 'all'}"
    saved_blocked = request.session.get(session_key, [])
    if isinstance(saved_blocked, list):
        for blk in saved_blocked:
            if isinstance(blk, dict):
                c_id = blk.get('classroom_id')
                d_num = blk.get('day_of_week')
                p_num = blk.get('period_number')
                if c_id and d_num and p_num:
                    k = f"{c_id}_{d_num}_{p_num}"
                    if k not in matrix_state:
                        matrix_state[k] = {'is_blocked': True, 'is_locked': True}

    # Pre-calculate classroom requirements mapping for frontend validation
    class_requirements_map = {}
    for cls in classrooms:
        reqs = requirements_map.get((cls.grade_level, cls.track), {})
        if not reqs:
            reqs = requirements_map.get((cls.grade_level, 'GENERAL'), {})
        class_requirements_map[cls.id] = {str(k): v for k, v in reqs.items()}

    # Calculate teacher statistics (Weekly load vs Max hours) within active academic year
    teacher_hours_report = []
    teacher_assigned_hours_map = {}
    teacher_max_hours_map = {}

    for t in teachers:
        assigned_cs = cs_by_teacher.get(t.id, [])
        total_assigned_h = 0
        for cs in assigned_cs:
            cls = cs.classroom
            if cls:
                h = requirements_map.get((cls.grade_level, cls.track), {}).get(cs.subject_id)
                if h is None:
                    h = requirements_map.get((cls.grade_level, 'GENERAL'), {}).get(cs.subject_id, 0)
                total_assigned_h += h

        scheduled_h = timetables_by_teacher.get(t.id, 0)
        t_max = t.max_weekly_hours or 18
        teacher_assigned_hours_map[t.id] = total_assigned_h
        teacher_max_hours_map[t.id] = t_max

        teacher_hours_report.append({
            'teacher': t,
            'assigned_hours': total_assigned_h,
            'scheduled_hours': scheduled_h,
            'max_hours': t_max,
            'diff': scheduled_h - total_assigned_h,
            'is_over': scheduled_h > t_max or (total_assigned_h > 0 and scheduled_h > total_assigned_h),
            'is_complete': scheduled_h == total_assigned_h and total_assigned_h > 0,
        })

    teacher_code_directory = []
    for (s_id, t_id), code in sorted(teacher_subject_code_map.items(), key=lambda x: str(x[1])):
        if (s_id, t_id) in cs_pairs_set:
            sub = subjects_by_id.get(s_id)
            tch = teachers_by_id.get(t_id)
            if sub and tch:
                teacher_code_directory.append({
                    'code': code,
                    'subject': sub,
                    'teacher': tch,
                })

    subjects_list = [{'id': s.id, 'name_kh': s.name_kh, 'code': s.code, 'category': s.category} for s in subjects]
    classrooms_list = [{'id': c.id, 'name': c.name, 'grade_level': c.grade_level, 'track': c.track} for c in classrooms]

    context = {
        'active_year': active_year,
        'academic_years': academic_years,
        'selected_year': str(active_year.id) if active_year else '',
        'classrooms': classrooms,
        'teachers': teachers,
        'subjects': subjects,
        'days': DAYS_OF_WEEK,
        'periods': PERIODS_LIST,
        'classrooms_data': classrooms_data,
        'class_options_json': json.dumps(class_options_map),
        'matrix_state_json': json.dumps(matrix_state),
        'class_requirements_json': json.dumps(class_requirements_map),
        'teacher_requirements_json': json.dumps(teacher_assigned_hours_map),
        'teacher_max_hours_json': json.dumps(teacher_max_hours_map),
        'teacher_assigned_hours_json': json.dumps(teacher_assigned_hours_map),
        'subjects_json': json.dumps(subjects_list),
        'classrooms_json': json.dumps(classrooms_list),
        'teacher_hours_report': teacher_hours_report,
        'teacher_code_directory': teacher_code_directory,
    }
    return render(request, 'academics/timetable.html', context)


@login_required
@role_required(['ADMIN'])
def timetable_save_matrix(request):
    """
    Saves the entire Master Timetable Matrix via AJAX JSON payload in a single atomic transaction.
    Scoped strictly to the active academic year!
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid HTTP method'}, status=405)

    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    try:
        data = json.loads(request.body.decode('utf-8'))
        year_param = data.get('academic_year_id') or data.get('year')
        if year_param:
            if str(year_param).isdigit():
                found_year = AcademicYear.objects.filter(id=int(year_param)).first()
            else:
                found_year = AcademicYear.objects.filter(name=str(year_param).strip()).first()
            if found_year:
                active_year = found_year

        matrix_items = data.get('matrix', [])
        blocked_items = data.get('blocked_slots')
        
        session_key = f"blocked_slots_{active_year.id if active_year else 'all'}"
        if blocked_items is not None and hasattr(request, 'session'):
            request.session[session_key] = blocked_items
            request.session.modified = True
        
        with transaction.atomic():
            if active_year:
                Timetable.objects.filter(classroom__academic_year=active_year).delete()
            else:
                Timetable.objects.all().delete()
            
            created_entries = []
            for item in matrix_items:
                cls_id = item.get('classroom_id')
                day_num = int(item.get('day_of_week'))
                p_num = int(item.get('period_number'))
                sub_id = item.get('subject_id')
                tch_id = item.get('teacher_id')
                
                if cls_id and sub_id and tch_id:
                    st_time, et_time = STANDARD_PERIOD_TIMES.get(
                        p_num, 
                        (datetime.time(7, 0), datetime.time(7, 50))
                    )
                    created_entries.append(Timetable(
                        classroom_id=cls_id,
                        subject_id=sub_id,
                        teacher_id=tch_id,
                        day_of_week=day_num,
                        period_number=p_num,
                        start_time=st_time,
                        end_time=et_time,
                    ))
            
            # Synchronize ClassSubject teacher assignments if provided during backup restore
            class_subjects_data = data.get('class_subject_assignments') or data.get('class_subjects')
            if class_subjects_data:
                if isinstance(class_subjects_data, dict):
                    for cid_str, items in class_subjects_data.items():
                        try:
                            c_id = int(cid_str)
                            for itm in items:
                                s_id = itm.get('subject_id')
                                t_id = itm.get('teacher_id')
                                if s_id and t_id:
                                    ClassSubject.objects.update_or_create(
                                        classroom_id=c_id,
                                        subject_id=s_id,
                                        defaults={'teacher_id': t_id}
                                    )
                        except Exception:
                            pass
                elif isinstance(class_subjects_data, list):
                    for itm in class_subjects_data:
                        c_id = itm.get('classroom_id')
                        s_id = itm.get('subject_id')
                        t_id = itm.get('teacher_id')
                        if c_id and s_id and t_id:
                            ClassSubject.objects.update_or_create(
                                classroom_id=c_id,
                                subject_id=s_id,
                                defaults={'teacher_id': t_id}
                            )

            if created_entries:
                Timetable.objects.bulk_create(created_entries)

        return JsonResponse({
            'status': 'success',
            'message': f'បានរក្សាទុកកាលវិភាគរួម ({len(created_entries)} ម៉ោង) សម្រាប់ឆ្នាំសិក្សា {active_year.name if active_year else ""} ដោយជោគជ័យ!',
            'count': len(created_entries),
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@role_required(['ADMIN'])
def timetable_transfer_class(request):
    """
    Transfers, Clones, or Swaps timetable entries between two classrooms (e.g. 7A -> 7B).
    Modes:
      - 'copy_with_teachers': Copies schedule from source to target, assigning the same teachers to target (and syncing target ClassSubject).
      - 'copy_with_target_teachers': Copies subject periods from source to target, but using target class's designated teachers.
      - 'move': Moves schedule from source to target (clearing source class schedule).
      - 'swap': Swaps schedule between source and target classes.
    Recalculates teacher hours and ensures real-time sync with ClassSubject!
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid HTTP method'}, status=405)

    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    try:
        data = json.loads(request.body.decode('utf-8'))
        source_id = int(data.get('source_class_id'))
        target_id = int(data.get('target_class_id'))
        mode = data.get('mode', 'copy_with_teachers')
        overwrite = data.get('overwrite', True)

        if source_id == target_id:
            return JsonResponse({'status': 'error', 'message': 'ថ្នាក់ប្រភព និងថ្នាក់គោលដៅមិនអាចដូចគ្នាឡើយ! (Source and target classes cannot be the same)'}, status=400)

        source_class = Classroom.objects.filter(id=source_id).first()
        target_class = Classroom.objects.filter(id=target_id).first()

        if not source_class or not target_class:
            return JsonResponse({'status': 'error', 'message': 'រកមិនឃើញថ្នាក់រៀនដែលបានជ្រើសរើសឡើយ!'}, status=404)

        source_entries = list(Timetable.objects.filter(classroom=source_class))
        target_entries = list(Timetable.objects.filter(classroom=target_class))

        if not source_entries and mode in ['copy_with_teachers', 'copy_with_target_teachers', 'move']:
            return JsonResponse({'status': 'error', 'message': f'ថ្នាក់ {source_class.name} មិនទាន់មានកាលវិភាគដើម្បីផ្ទេរឡើយ!'}, status=400)

        # Target class assigned teachers lookup from ClassSubject
        target_cs_map = {cs.subject_id: cs.teacher_id for cs in ClassSubject.objects.filter(classroom=target_class) if cs.teacher_id}

        with transaction.atomic():
            if mode == 'swap':
                # Swap slots between source and target
                Timetable.objects.filter(classroom=source_class).delete()
                Timetable.objects.filter(classroom=target_class).delete()

                new_source_entries = []
                for te in target_entries:
                    new_source_entries.append(Timetable(
                        classroom=source_class,
                        subject=te.subject,
                        teacher=te.teacher,
                        day_of_week=te.day_of_week,
                        period_number=te.period_number,
                        start_time=te.start_time,
                        end_time=te.end_time,
                    ))

                new_target_entries = []
                for se in source_entries:
                    new_target_entries.append(Timetable(
                        classroom=target_class,
                        subject=se.subject,
                        teacher=se.teacher,
                        day_of_week=se.day_of_week,
                        period_number=se.period_number,
                        start_time=se.start_time,
                        end_time=se.end_time,
                    ))

                if new_source_entries:
                    Timetable.objects.bulk_create(new_source_entries)
                if new_target_entries:
                    Timetable.objects.bulk_create(new_target_entries)

                msg = f"បានប្តូរកាលវិភាគរវាងថ្នាក់ {source_class.name} និង {target_class.name} ទៅវិញទៅមកដោយជោគជ័យ!"

            elif mode == 'move':
                # Move: Delete target if overwrite, delete source entries
                if overwrite:
                    Timetable.objects.filter(classroom=target_class).delete()

                new_target_entries = []
                for se in source_entries:
                    new_target_entries.append(Timetable(
                        classroom=target_class,
                        subject=se.subject,
                        teacher=se.teacher,
                        day_of_week=se.day_of_week,
                        period_number=se.period_number,
                        start_time=se.start_time,
                        end_time=se.end_time,
                    ))
                    # Sync target ClassSubject
                    ClassSubject.objects.update_or_create(
                        classroom=target_class,
                        subject=se.subject,
                        defaults={'teacher': se.teacher}
                    )

                Timetable.objects.filter(classroom=source_class).delete()
                if new_target_entries:
                    Timetable.objects.bulk_create(new_target_entries)

                msg = f"បានផ្ទេរកាលវិភាគទាំងអស់ពី {source_class.name} ទៅកាន់ {target_class.name} ដោយជោគជ័យ!"

            elif mode == 'copy_with_target_teachers':
                # Copy subject periods but use target class's assigned teachers
                if overwrite:
                    Timetable.objects.filter(classroom=target_class).delete()

                new_target_entries = []
                for se in source_entries:
                    t_id = target_cs_map.get(se.subject_id) or se.teacher_id
                    t_obj = Teacher.objects.filter(id=t_id).first() if t_id else se.teacher
                    new_target_entries.append(Timetable(
                        classroom=target_class,
                        subject=se.subject,
                        teacher=t_obj,
                        day_of_week=se.day_of_week,
                        period_number=se.period_number,
                        start_time=se.start_time,
                        end_time=se.end_time,
                    ))

                if new_target_entries:
                    Timetable.objects.bulk_create(new_target_entries)

                msg = f"បានចម្លងគ្រោងកាលវិភាគពី {source_class.name} ទៅកាន់ {target_class.name} (ដោយប្រើគ្រូចាត់តាំងនៅ {target_class.name}) ដោយជោគជ័យ!"

            else:  # 'copy_with_teachers'
                if overwrite:
                    Timetable.objects.filter(classroom=target_class).delete()

                new_target_entries = []
                for se in source_entries:
                    new_target_entries.append(Timetable(
                        classroom=target_class,
                        subject=se.subject,
                        teacher=se.teacher,
                        day_of_week=se.day_of_week,
                        period_number=se.period_number,
                        start_time=se.start_time,
                        end_time=se.end_time,
                    ))
                    # Sync target ClassSubject with copied teachers
                    ClassSubject.objects.update_or_create(
                        classroom=target_class,
                        subject=se.subject,
                        defaults={'teacher': se.teacher}
                    )

                if new_target_entries:
                    Timetable.objects.bulk_create(new_target_entries)

                msg = f"បានចម្លងកាលវិភាគ និងគ្រូបង្រៀនពី {source_class.name} ទៅកាន់ {target_class.name} ដោយជោគជ័យ!"

        # Fetch updated entries for real-time frontend matrix sync
        updated_source = [
            {
                'classroom_id': e.classroom_id,
                'subject_id': e.subject_id,
                'teacher_id': e.teacher_id,
                'day_of_week': e.day_of_week,
                'period_number': e.period_number,
            } for e in Timetable.objects.filter(classroom=source_class)
        ]
        updated_target = [
            {
                'classroom_id': e.classroom_id,
                'subject_id': e.subject_id,
                'teacher_id': e.teacher_id,
                'day_of_week': e.day_of_week,
                'period_number': e.period_number,
            } for e in Timetable.objects.filter(classroom=target_class)
        ]

        return JsonResponse({
            'status': 'success',
            'message': msg,
            'source_entries': updated_source,
            'target_entries': updated_target,
            'source_class_id': source_id,
            'target_class_id': target_id,
            'mode': mode,
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@role_required(['ADMIN'])
def timetable_auto_generate(request):
    """
    Intelligent Timetable Auto-Generator (Constraint-Satisfaction Solver).
    Strict Priorities:
      1. Alternating Days (រំលង ១ ថ្ងៃ ទៅ ១ ថ្ងៃទៀត): Spaced on alternating days (Mon/Wed/Fri or Tue/Thu/Sat) in 2-hour consecutive blocks.
      2. Remainder Placement (បើសិននៅសល់ម៉ោង): Places remaining single hours or overflow into intermediate available days/periods.
      3. Science vs Social Science: Morning priority (Periods 1-4) for Science, balanced complementary placement for Social Science.
      4. Zero teacher conflict, preservation of locked & blocked slots, and exact requirement quota match.
      Scoped strictly to active academic year!
    """
    import random
    from collections import defaultdict
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    # Parse potential locked / blocked slots & academic year from request
    locked_slots_input = []
    if request.body:
        try:
            req_data = json.loads(request.body.decode('utf-8'))
            locked_slots_input = req_data.get('locked_slots', [])
            year_param = req_data.get('academic_year_id') or req_data.get('year')
            if year_param:
                if str(year_param).isdigit():
                    found_year = AcademicYear.objects.filter(id=int(year_param)).first()
                else:
                    found_year = AcademicYear.objects.filter(name=str(year_param).strip()).first()
                if found_year:
                    active_year = found_year
        except Exception:
            pass

    if not active_year:
        active_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.first()

    session_key = f"blocked_slots_{active_year.id if active_year else 'all'}"
    blocked_from_input = [ls for ls in locked_slots_input if ls.get('is_blocked')]
    if blocked_from_input and hasattr(request, 'session'):
        request.session[session_key] = blocked_from_input
        request.session.modified = True

    classrooms = list(Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code'))
    subjects = list(Subject.objects.exclude(code__in=['R', 'D']))
    teachers = list(Teacher.objects.filter(status='ACTIVE'))
    teacher_dict = {t.id: t for t in teachers}
    subject_dict = {s.id: s for s in subjects}

    # Pre-fetch assignments for this academic year
    class_subject_assignments = ClassSubject.objects.filter(
        classroom__academic_year=active_year,
        teacher__isnull=False
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('classroom', 'subject', 'teacher') if active_year else ClassSubject.objects.filter(
        teacher__isnull=False
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('classroom', 'subject', 'teacher')

    class_teacher_map = {}
    for cs in class_subject_assignments:
        class_teacher_map[(cs.classroom_id, cs.subject_id)] = cs.teacher

    # Pre-fetch requirements
    grade_rules = GradeLevelRule.objects.filter(
        weekly_hours__gt=0
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('subject')

    requirements_map = {}
    for r in grade_rules:
        k = (r.grade_level, r.track)
        if k not in requirements_map:
            requirements_map[k] = []
        requirements_map[k].append(r)

    # Map pre-locked and blocked slots per classroom: cls_id -> (d, p) -> slot_dict
    locked_by_class = defaultdict(dict)
    for ls in locked_slots_input:
        c_id = int(ls.get('classroom_id', 0))
        d_num = int(ls.get('day_of_week', 0))
        p_num = int(ls.get('period_number', 0))
        if c_id and d_num and p_num:
            locked_by_class[c_id][(d_num, p_num)] = ls

    DAYS = [1, 2, 3, 4, 5, 6]
    ALL_PERIODS = [1, 2, 3, 4, 5, 6, 7, 8]
    PERIOD_PAIRS_MORNING = [(1, 2), (3, 4)]
    PERIOD_PAIRS_AFTERNOON = [(5, 6), (7, 8)]
    SINGLE_PERIODS_MORNING = [1, 2, 3, 4]
    SINGLE_PERIODS_AFTERNOON = [5, 6, 7, 8]

    # Alternating day pairs & triplets (រំលង ១ ថ្ងៃ)
    ALT_DAY_PAIRS = [(1, 3), (2, 4), (3, 5), (4, 6), (1, 5), (2, 6), (1, 4), (3, 6)]
    ALT_DAY_TRIPLETS = [(1, 3, 5), (2, 4, 6), (1, 3, 6), (1, 4, 6), (2, 4, 5)]

    best_solution = None
    best_score = -999999

    # Run multi-pass optimization solver to find the optimal clash-free, alternating-spaced schedule
    for attempt in range(15):
        teacher_occupancy = set()
        class_slots = defaultdict(dict)
        day_sub_hours = defaultdict(lambda: defaultdict(int))
        total_placed = 0
        total_demanded = 0
        alternating_matches = 0

        # Step 0: Register all locked and blocked slots first across all classes
        for cls in classrooms:
            cls_locked_map = locked_by_class.get(cls.id, {})
            for (d, p), ls in cls_locked_map.items():
                is_blocked = ls.get('is_blocked', False)
                sub_id = ls.get('subject_id')
                tch_id = ls.get('teacher_id')

                if is_blocked or not sub_id:
                    class_slots[cls.id][(d, p)] = {'is_blocked': True}
                else:
                    sub = subject_dict.get(int(sub_id))
                    tch = teacher_dict.get(int(tch_id))
                    if sub and tch:
                        class_slots[cls.id][(d, p)] = {'subject': sub, 'teacher': tch, 'is_locked': True}
                        teacher_occupancy.add((tch.id, d, p))
                        day_sub_hours[(cls.id, d)][sub.id] += 1

        cls_list = list(classrooms)
        random.shuffle(cls_list)

        for cls in cls_list:
            rules = requirements_map.get((cls.grade_level, cls.track)) or requirements_map.get((cls.grade_level, 'GENERAL'), [])
            cls_locked_map = locked_by_class.get(cls.id, {})
            locked_sub_counts = defaultdict(int)
            for (d, p), ls in cls_locked_map.items():
                if not ls.get('is_blocked') and ls.get('subject_id'):
                    locked_sub_counts[int(ls.get('subject_id'))] += 1

            has_any_assignments = class_subject_assignments.exists()
            demands = []
            for r in rules:
                tch = class_teacher_map.get((cls.id, r.subject_id))
                if not tch:
                    if has_any_assignments:
                        # STRICT ISOLATION & ASSIGNMENT: If assignments exist for this year, only schedule assigned classes!
                        continue
                    else:
                        sub_teachers = [t for t in teachers if r.subject.name_kh in (t.specialization or '')]
                        tch = sub_teachers[0] if sub_teachers else (teachers[0] if teachers else None)
                
                if not tch:
                    continue
                needed_h = max(0, r.weekly_hours - locked_sub_counts[r.subject_id])
                if needed_h <= 0:
                    continue


                total_demanded += needed_h
                cat = r.subject.category
                # Priority: Science=1, Khmer=2, Social=3, Other=4
                if cat == Subject.SubjectCategory.SCIENCE:
                    prio = 1
                elif r.subject.code == 'K':
                    prio = 2
                elif cat == Subject.SubjectCategory.SOCIAL:
                    prio = 3
                else:
                    prio = 4

                demands.append({
                    'subject': r.subject,
                    'teacher': tch,
                    'hours': needed_h,
                    'priority': prio,
                })

            # Sort demands by priority then hours desc
            demands.sort(key=lambda x: (x['priority'], -x['hours']))

            for dem in demands:
                sub = dem['subject']
                tch = dem['teacher']
                rem_h = dem['hours']

                # Decompose into teaching units: e.g. 4 -> 2+2, 3 -> 2+1, 2 -> 2, 6 -> 2+2+2
                blocks = []
                while rem_h >= 2:
                    blocks.append(2)
                    rem_h -= 2
                if rem_h == 1:
                    blocks.append(1)

                # ==========================================
                # PRIORITY 1: Alternating Days (រំលង ១ ថ្ងៃ ទៅ ១ ថ្ងៃទៀត)
                # ==========================================
                if len(blocks) == 2 and blocks == [2, 2]:
                    pair_options = list(ALT_DAY_PAIRS)
                    random.shuffle(pair_options)
                    pair_placed = False

                    for d1, d2 in pair_options:
                        period_options = (PERIOD_PAIRS_MORNING + PERIOD_PAIRS_AFTERNOON) if dem['priority'] <= 2 else (PERIOD_PAIRS_MORNING + PERIOD_PAIRS_AFTERNOON)
                        
                        found_p1 = None
                        found_p2 = None

                        for p_a, p_b in period_options:
                            if (d1, p_a) not in class_slots[cls.id] and (d1, p_b) not in class_slots[cls.id]:
                                if (tch.id, d1, p_a) not in teacher_occupancy and (tch.id, d1, p_b) not in teacher_occupancy:
                                    if day_sub_hours[(cls.id, d1)][sub.id] + 2 <= 2:
                                        found_p1 = (p_a, p_b)
                                        break

                        if found_p1:
                            for p_a, p_b in period_options:
                                if (d2, p_a) not in class_slots[cls.id] and (d2, p_b) not in class_slots[cls.id]:
                                    if (tch.id, d2, p_a) not in teacher_occupancy and (tch.id, d2, p_b) not in teacher_occupancy:
                                        if day_sub_hours[(cls.id, d2)][sub.id] + 2 <= 2:
                                            found_p2 = (p_a, p_b)
                                            break

                        if found_p1 and found_p2:
                            for p in found_p1:
                                class_slots[cls.id][(d1, p)] = {'subject': sub, 'teacher': tch}
                                teacher_occupancy.add((tch.id, d1, p))
                            day_sub_hours[(cls.id, d1)][sub.id] += 2

                            for p in found_p2:
                                class_slots[cls.id][(d2, p)] = {'subject': sub, 'teacher': tch}
                                teacher_occupancy.add((tch.id, d2, p))
                            day_sub_hours[(cls.id, d2)][sub.id] += 2

                            total_placed += 4
                            alternating_matches += 1
                            pair_placed = True
                            blocks = [] # Successfully placed in alternating days
                            break

                elif len(blocks) == 3 and blocks == [2, 2, 2]:
                    triplet_options = list(ALT_DAY_TRIPLETS)
                    random.shuffle(triplet_options)
                    trip_placed = False

                    for d1, d2, d3 in triplet_options:
                        p_opts = PERIOD_PAIRS_MORNING + PERIOD_PAIRS_AFTERNOON
                        f_p1, f_p2, f_p3 = None, None, None
                        
                        for p_a, p_b in p_opts:
                            if (d1, p_a) not in class_slots[cls.id] and (d1, p_b) not in class_slots[cls.id] and (tch.id, d1, p_a) not in teacher_occupancy and (tch.id, d1, p_b) not in teacher_occupancy and day_sub_hours[(cls.id, d1)][sub.id] + 2 <= 2:
                                f_p1 = (p_a, p_b)
                                break
                        if f_p1:
                            for p_a, p_b in p_opts:
                                if (d2, p_a) not in class_slots[cls.id] and (d2, p_b) not in class_slots[cls.id] and (tch.id, d2, p_a) not in teacher_occupancy and (tch.id, d2, p_b) not in teacher_occupancy and day_sub_hours[(cls.id, d2)][sub.id] + 2 <= 2:
                                    f_p2 = (p_a, p_b)
                                    break
                        if f_p1 and f_p2:
                            for p_a, p_b in p_opts:
                                if (d3, p_a) not in class_slots[cls.id] and (d3, p_b) not in class_slots[cls.id] and (tch.id, d3, p_a) not in teacher_occupancy and (tch.id, d3, p_b) not in teacher_occupancy and day_sub_hours[(cls.id, d3)][sub.id] + 2 <= 2:
                                    f_p3 = (p_a, p_b)
                                    break

                        if f_p1 and f_p2 and f_p3:
                            for d_cur, f_p in [(d1, f_p1), (d2, f_p2), (d3, f_p3)]:
                                for p in f_p:
                                    class_slots[cls.id][(d_cur, p)] = {'subject': sub, 'teacher': tch}
                                    teacher_occupancy.add((tch.id, d_cur, p))
                                day_sub_hours[(cls.id, d_cur)][sub.id] += 2
                            total_placed += 6
                            alternating_matches += 2
                            trip_placed = True
                            blocks = []
                            break

                # ==========================================
                # PRIORITY 2: Remainder Placement (បើសិននៅសល់ម៉ោង គឺត្រូវបន្ថែមចន្លោះពេល ឬថ្ងៃណាមួយ)
                # ==========================================
                for blk in blocks:
                    placed = False
                    avail_days = list(DAYS)
                    # Sort days by least hours of this subject in that day, then least class total periods
                    avail_days.sort(key=lambda d: (
                        day_sub_hours[(cls.id, d)][sub.id],
                        len([p for p in ALL_PERIODS if (d, p) in class_slots[cls.id]])
                    ))

                    if blk == 2:
                        p_opts = PERIOD_PAIRS_MORNING + PERIOD_PAIRS_AFTERNOON
                        for d in avail_days:
                            if day_sub_hours[(cls.id, d)][sub.id] + 2 > 2:
                                continue
                            for p_a, p_b in p_opts:
                                if (d, p_a) not in class_slots[cls.id] and (d, p_b) not in class_slots[cls.id]:
                                    if (tch.id, d, p_a) not in teacher_occupancy and (tch.id, d, p_b) not in teacher_occupancy:
                                        class_slots[cls.id][(d, p_a)] = {'subject': sub, 'teacher': tch}
                                        class_slots[cls.id][(d, p_b)] = {'subject': sub, 'teacher': tch}
                                        teacher_occupancy.add((tch.id, d, p_a))
                                        teacher_occupancy.add((tch.id, d, p_b))
                                        day_sub_hours[(cls.id, d)][sub.id] += 2
                                        total_placed += 2
                                        placed = True
                                        break
                            if placed:
                                break

                    if not placed:
                        # 1h single block or fallback
                        needed_single = blk
                        for d in avail_days:
                            if needed_single == 0:
                                break
                            if day_sub_hours[(cls.id, d)][sub.id] >= 2:
                                continue
                            p_list = (SINGLE_PERIODS_MORNING + SINGLE_PERIODS_AFTERNOON) if dem['priority'] <= 2 else ALL_PERIODS
                            for p in p_list:
                                if (d, p) not in class_slots[cls.id] and (tch.id, d, p) not in teacher_occupancy:
                                    class_slots[cls.id][(d, p)] = {'subject': sub, 'teacher': tch}
                                    teacher_occupancy.add((tch.id, d, p))
                                    day_sub_hours[(cls.id, d)][sub.id] += 1
                                    total_placed += 1
                                    needed_single -= 1
                                    if needed_single == 0:
                                        placed = True
                                        break

        score = (total_placed * 100) + (alternating_matches * 20)
        if score > best_score:
            best_score = score
            best_solution = class_slots
            if total_placed == total_demanded:
                # 100% placed with maximum alternating efficiency!
                break

    # Persist the winning schedule to database
    generated_timetable_entries = []
    if best_solution:
        with transaction.atomic():
            if active_year:
                Timetable.objects.filter(classroom__academic_year=active_year).delete()
            else:
                Timetable.objects.all().delete()
            for cls in classrooms:
                cls_slots = best_solution.get(cls.id, {})
                for (d, p), item in cls_slots.items():
                    if item and not item.get('is_blocked') and item.get('subject') and item.get('teacher'):
                        st_time, et_time = STANDARD_PERIOD_TIMES.get(p, (datetime.time(7, 0), datetime.time(7, 50)))
                        generated_timetable_entries.append(Timetable(
                            classroom=cls,
                            subject=item['subject'],
                            teacher=item['teacher'],
                            day_of_week=d,
                            period_number=p,
                            start_time=st_time,
                            end_time=et_time,
                            room=cls.room_number or f"បន្ទប់ {cls.code}"
                        ))

            if generated_timetable_entries:
                Timetable.objects.bulk_create(generated_timetable_entries)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
        return JsonResponse({
            'status': 'success',
            'message': f'បានរៀបចំកាលវិភាគស្វ័យប្រវត្តិចំនួន {len(generated_timetable_entries)} ម៉ោងដោយជោគជ័យ (គោរពតាមគោលការណ៍រំលងថ្ងៃ រៀបចំម៉ោងនៅសល់ និងគ្មានការជាន់ម៉ោងគ្នា)!',
            'count': len(generated_timetable_entries),
        })
    messages.success(request, f"បានរៀបចំកាលវិភាគស្វ័យប្រវត្តិចំនួន {len(generated_timetable_entries)} ម៉ោងដោយជោគជ័យ!")
    return redirect('timetable_view')


@login_required
def student_teacher_timetable_view(request):
    """
    Individual & Batch Printable Timetable View for Students (Classrooms) and Teachers.
    Matches official Ministry of Education, Youth and Sport (MoEYS) timetable layout.
    Strictly isolated per Academic Year!
    """
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
        else:
            found_year = AcademicYear.objects.filter(name=str(selected_year).strip()).first()
        if found_year:
            active_year = found_year

    academic_years = list(AcademicYear.objects.all().order_by('-start_date'))
    classrooms = list(Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code'))
    teachers = list(Teacher.objects.filter(status='ACTIVE').order_by('khmer_name'))
    timetables_qs = Timetable.objects.filter(classroom__academic_year=active_year).select_related('classroom', 'subject', 'teacher') if active_year else Timetable.objects.select_related('classroom', 'subject', 'teacher').all()
    timetables = list(timetables_qs)

    timetables_by_classroom = defaultdict(list)
    timetables_by_teacher = defaultdict(list)
    for entry in timetables:
        timetables_by_classroom[entry.classroom_id].append(entry)
        if entry.teacher_id:
            timetables_by_teacher[entry.teacher_id].append(entry)

    # Pre-fetch ClassSubject for teachers' subjects within active academic year
    class_subjects_qs = ClassSubject.objects.filter(classroom__academic_year=active_year, teacher__isnull=False).select_related('subject', 'teacher', 'classroom') if active_year else ClassSubject.objects.filter(teacher__isnull=False).select_related('subject', 'teacher', 'classroom')
    class_subjects = list(class_subjects_qs)
    teacher_subjects_map = defaultdict(set)
    for cs in class_subjects:
        if cs.teacher_id and cs.subject:
            teacher_subjects_map[cs.teacher_id].add(cs.subject.name_kh)

    # Pre-fetch Teacher Duty Schedules for active academic year
    duty_entries_qs = TeacherDutySchedule.objects.filter(academic_year=active_year).select_related('teacher') if active_year else TeacherDutySchedule.objects.all().select_related('teacher')
    duty_entries = list(duty_entries_qs)
    duty_by_teacher = defaultdict(list)
    for d in duty_entries:
        if d.teacher_id:
            duty_by_teacher[d.teacher_id].append(d)

    raw_duty_types = TeacherDutyType.get_all_duty_types()
    duty_types_dict = {dt.code: dt.name for dt in raw_duty_types}

    # 1. Build Classroom Timetables Data
    classrooms_timetables = []
    for cls in classrooms:
        cls_entries = timetables_by_classroom.get(cls.id, [])
        slots_map = {}
        for entry in cls_entries:
            t_name = (entry.teacher.khmer_name or entry.teacher.name or '') if entry.teacher else ""
            t_gender = getattr(entry.teacher, 'gender', 'M') if entry.teacher else 'M'
            title = "អ្នកគ្រូ" if t_gender == 'F' else "លោកគ្រូ"
            if t_name.startswith('លោកគ្រូ') or t_name.startswith('អ្នកគ្រូ'):
                display_teacher = t_name
            else:
                display_teacher = f"{title} {t_name}" if t_name else ""

            slots_map[(entry.day_of_week, entry.period_number)] = {
                'subject_name': entry.subject.name_kh if entry.subject else '',
                'subject_code': entry.subject.code if entry.subject else '',
                'teacher_name': t_name,
                'teacher_title': title,
                'teacher_display': display_teacher,
                'teacher_gender': t_gender,
            }

        # Build 4 morning periods (1-4) and 4 afternoon periods (5-8)
        morning_rows = []
        for p in [1, 2, 3, 4]:
            p_slots = [slots_map.get((d['num'], p)) for d in DAYS_OF_WEEK]
            morning_rows.append({'period': p, 'slots': p_slots})

        afternoon_rows = []
        for p in [5, 6, 7, 8]:
            p_slots = [slots_map.get((d['num'], p)) for d in DAYS_OF_WEEK]
            afternoon_rows.append({'period': p, 'slots': p_slots})

        classrooms_timetables.append({
            'classroom': cls,
            'homeroom_teacher': cls.homeroom_teacher,
            'academic_year': cls.academic_year.name if cls.academic_year else (active_year.name if active_year else "២០២៥-២០២៦"),
            'morning_rows': morning_rows,
            'afternoon_rows': afternoon_rows,
            'total_hours': len(slots_map),
        })

    # 2. Build Teacher Timetables Data (Teaching slots + Duty schedule slots in active academic year)
    teachers_timetables = []
    for tch in teachers:
        tch_entries = timetables_by_teacher.get(tch.id, [])
        slots_map = {}
        teaching_hours = 0
        for entry in tch_entries:
            slots_map[(entry.day_of_week, entry.period_number)] = {
                'subject_name': entry.subject.name_kh if entry.subject else '',
                'subject_code': entry.subject.code if entry.subject else '',
                'classroom_name': entry.classroom.name if entry.classroom else '',
                'classroom_code': entry.classroom.code if entry.classroom else '',
                'is_duty': False,
            }
            teaching_hours += 1

        # Fill in duty slots if no teaching class in that slot
        tch_duties = duty_by_teacher.get(tch.id, [])
        duty_hours = 0
        for d in tch_duties:
            k = (d.day_of_week, d.period_number)
            if k not in slots_map:
                duty_name = duty_types_dict.get(d.duty_type, d.duty_type)
                slots_map[k] = {
                    'is_duty': True,
                    'duty_name': duty_name,
                    'duty_code': d.duty_type,
                    'duty_notes': d.notes or '',
                    'is_auto': d.is_auto_assigned,
                }
                duty_hours += 1

        morning_rows = []
        for p in [1, 2, 3, 4]:
            p_slots = [slots_map.get((d['num'], p)) for d in DAYS_OF_WEEK]
            morning_rows.append({'period': p, 'slots': p_slots})

        afternoon_rows = []
        for p in [5, 6, 7, 8]:
            p_slots = [slots_map.get((d['num'], p)) for d in DAYS_OF_WEEK]
            afternoon_rows.append({'period': p, 'slots': p_slots})

        assigned_subs = sorted(list(teacher_subjects_map.get(tch.id, set())))
        subjects_display = ", ".join(assigned_subs) if assigned_subs else (tch.specialization or "-")
        title = "អ្នកគ្រូ" if tch.gender == 'F' else "លោកគ្រូ"

        teachers_timetables.append({
            'teacher': tch,
            'title': title,
            'subjects_display': subjects_display,
            'academic_year': active_year.name if active_year else "២០២៥-២០២៦",
            'morning_rows': morning_rows,
            'afternoon_rows': afternoon_rows,
            'teaching_hours': teaching_hours,
            'duty_hours': duty_hours,
            'total_hours': teaching_hours + duty_hours,
        })

    # Date formatting in Khmer
    now = datetime.datetime.now()
    khmer_digits = {'0': '០', '1': '១', '2': '២', '3': '៣', '4': '៤', '5': '៥', '6': '៦', '7': '៧', '8': '៨', '9': '៩'}
    def to_khmer_num(n):
        return ''.join(khmer_digits.get(c, c) for c in str(n))

    kh_months = ['', 'មករា', 'កុម្ភៈ', 'មីនា', 'មេសា', 'ឧសភា', 'មិថុនា', 'កក្កដា', 'សីហា', 'កញ្ញា', 'តុលា', 'វិច្ឆិកា', 'ធ្នូ']
    kh_days_name = ['ច័ន្ទ', 'អង្គារ', 'ពុធ', 'ព្រហស្បតិ៍', 'សុក្រ', 'សៅរ៍', 'អាទិត្យ']

    today_kh_day = to_khmer_num(now.day)
    today_kh_month = kh_months[now.month] if 1 <= now.month <= 12 else ''
    today_kh_year = to_khmer_num(now.year)
    today_kh_dow = kh_days_name[now.weekday()]

    context = {
        'classrooms': classrooms,
        'teachers': teachers,
        'classrooms_timetables': classrooms_timetables,
        'teachers_timetables': teachers_timetables,
        'academic_year': active_year,
        'academic_years': academic_years,
        'selected_year': str(active_year.id) if active_year else '',
        'today_kh_day': today_kh_day,
        'today_kh_month': today_kh_month,
        'today_kh_year': today_kh_year,
        'today_kh_dow': today_kh_dow,
        'days': DAYS_OF_WEEK,
    }
    return render(request, 'academics/student_teacher_timetable.html', context)


@login_required
def student_teacher_timetable_export_excel(request):
    """
    Exports Student (Classroom) and Teacher Timetables to Excel (.xlsx) using openpyxl.
    Supports:
      - mode='class' & id='all'   : All classrooms in separate sheets (1 sheet per classroom in 1 workbook)
      - mode='class' & id=<int>   : Single classroom timetable sheet
      - mode='teacher' & id='all' : All teachers in separate sheets (1 sheet per teacher in 1 workbook)
      - mode='teacher' & id=<int> : Single teacher timetable sheet
    Strictly isolated per Academic Year!
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from .utils import get_active_academic_year

    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
        else:
            found_year = AcademicYear.objects.filter(name=str(selected_year).strip()).first()
        if found_year:
            active_year = found_year

    year_name = active_year.name if active_year else '២០២៦-២០២៧'
    mode = request.GET.get('mode', 'class') # 'class' or 'teacher'
    target_id = request.GET.get('id', 'all')

    classrooms = list(Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code'))
    teachers = list(Teacher.objects.filter(status='ACTIVE').order_by('khmer_name'))
    timetables_qs = Timetable.objects.filter(classroom__academic_year=active_year).select_related('classroom', 'subject', 'teacher') if active_year else Timetable.objects.select_related('classroom', 'subject', 'teacher').all()
    timetables = list(timetables_qs)

    timetables_by_classroom = defaultdict(list)
    timetables_by_teacher = defaultdict(list)
    for entry in timetables:
        timetables_by_classroom[entry.classroom_id].append(entry)
        if entry.teacher_id:
            timetables_by_teacher[entry.teacher_id].append(entry)

    # ClassSubject mapping
    class_subjects_qs = ClassSubject.objects.filter(classroom__academic_year=active_year, teacher__isnull=False).select_related('subject', 'teacher', 'classroom') if active_year else ClassSubject.objects.filter(teacher__isnull=False).select_related('subject', 'teacher', 'classroom')
    teacher_subjects_map = defaultdict(set)
    for cs in class_subjects_qs:
        if cs.teacher_id and cs.subject:
            teacher_subjects_map[cs.teacher_id].add(cs.subject.name_kh)

    # Teacher Duties
    duty_entries_qs = TeacherDutySchedule.objects.filter(academic_year=active_year).select_related('teacher') if active_year else TeacherDutySchedule.objects.all().select_related('teacher')
    duty_by_teacher = defaultdict(list)
    for d in duty_entries_qs:
        if d.teacher_id:
            duty_by_teacher[d.teacher_id].append(d)

    raw_duty_types = TeacherDutyType.get_all_duty_types()
    duty_types_dict = {dt.code: dt.name for dt in raw_duty_types}

    # Styling Tokens
    font_title = Font(name='Khmer OS Muol Light', size=12, bold=True)
    font_sub = Font(name='Khmer OS Muol Light', size=11, bold=True)
    font_meta = Font(name='Khmer OS Siemreap', size=10, italic=True)
    font_head = Font(name='Khmer OS Siemreap', size=10, bold=True)
    font_session = Font(name='Khmer OS Siemreap', size=10, bold=True)
    font_slot = Font(name='Khmer OS Siemreap', size=9.5)
    font_slot_bold = Font(name='Khmer OS Siemreap', size=9.5, bold=True)
    font_duty = Font(name='Khmer OS Siemreap', size=9.5, bold=True, color='B91C1C')
    font_sig_title = Font(name='Khmer OS Siemreap', size=10, bold=True)
    font_sig_name = Font(name='Khmer OS Siemreap', size=10, bold=True)

    fill_header = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    fill_session = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    fill_duty = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='1E293B'),
        right=Side(style='thin', color='1E293B'),
        top=Side(style='thin', color='1E293B'),
        bottom=Side(style='thin', color='1E293B')
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    def sanitize_title(title):
        invalid_chars = r':\/?*[]'
        for char in invalid_chars:
            title = title.replace(char, '')
        return title.strip()[:31]

    # Helper: Build 1 Classroom Timetable Sheet
    def render_classroom_sheet(workbook, cls):
        sheet_name = sanitize_title(f"ថ្នាក់ {cls.name}")
        # Ensure sheet title uniqueness
        counter = 1
        base_name = sheet_name
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{base_name[:28]}_{counter}"
            counter += 1

        ws = workbook.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        # Header Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្តែត"
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:G2')
        ws['A2'] = "កាលវិភាគបង្រៀន និងរៀនប្រចាំសប្តាហ៍"
        ws['A2'].font = font_sub
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        homeroom_str = f"{cls.homeroom_teacher.khmer_name}" if (cls.homeroom_teacher and hasattr(cls, 'homeroom_teacher')) else "...................."
        ws.merge_cells('A3:G3')
        ws['A3'] = f"ថ្នាក់ទី៖ {cls.name} | គ្រូបន្ទុកថ្នាក់៖ {homeroom_str} | ឆ្នាំសិក្សា៖ {year_name}"
        ws['A3'].font = font_meta
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

        # Headers Row 5
        headers = ['ម៉ោងទី', 'ច័ន្ទ', 'អង្គារ', 'ពុធ', 'ព្រហស្បតិ៍', 'សុក្រ', 'សៅរ៍']
        ws.append([]) # Row 4
        ws.append(headers) # Row 5
        for col_idx in range(1, 8):
            cell = ws.cell(row=5, column=col_idx)
            cell.font = font_head
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_header
            cell.border = thin_border
        ws.row_dimensions[5].height = 26

        cls_entries = timetables_by_classroom.get(cls.id, [])
        slots_map = {}
        for entry in cls_entries:
            t_name = entry.teacher.khmer_name if entry.teacher else ""
            t_gender = getattr(entry.teacher, 'gender', 'M') if entry.teacher else 'M'
            title = "អ្នកគ្រូ" if t_gender == 'F' else "លោកគ្រូ"
            disp_tch = f"{title} {t_name}" if t_name else ""
            sub_name = entry.subject.name_kh if entry.subject else ""
            slots_map[(entry.day_of_week, entry.period_number)] = f"{sub_name}\n({disp_tch})" if (sub_name and disp_tch) else (sub_name or "-")

        # Session 1: ពេលព្រឹក (Row 6)
        ws.append(['ពេលព្រឹក', '', '', '', '', '', ''])
        ws.merge_cells('A6:G6')
        for col_idx in range(1, 8):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_session
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_session
            cell.border = thin_border
        ws.row_dimensions[6].height = 22

        # Periods 1 to 4 (Rows 7 to 10)
        curr_row = 7
        for p in [1, 2, 3, 4]:
            row_data = [p] + [slots_map.get((d['num'], p), '-') for d in DAYS_OF_WEEK]
            ws.append(row_data)
            ws.row_dimensions[curr_row].height = 36
            for col_idx in range(1, 8):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = font_slot_bold if col_idx > 1 else font_head
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            curr_row += 1

        # Session 2: ពេលរសៀល (Row 11)
        ws.append(['ពេលរសៀល', '', '', '', '', '', ''])
        ws.merge_cells('A11:G11')
        for col_idx in range(1, 8):
            cell = ws.cell(row=11, column=col_idx)
            cell.font = font_session
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_session
            cell.border = thin_border
        ws.row_dimensions[11].height = 22

        # Periods 5 to 8 (Rows 12 to 15)
        curr_row = 12
        for p in [5, 6, 7, 8]:
            row_data = [p] + [slots_map.get((d['num'], p), '-') for d in DAYS_OF_WEEK]
            ws.append(row_data)
            ws.row_dimensions[curr_row].height = 36
            for col_idx in range(1, 8):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = font_slot_bold if col_idx > 1 else font_head
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            curr_row += 1

        # Signatures (Rows 17-20)
        curr_row += 1
        ws.cell(row=curr_row + 1, column=2, value="បានឃើញ និងឯកភាព").font = font_sig_title
        ws.cell(row=curr_row + 2, column=2, value="នាយកសាលា").font = font_sig_title
        ws.cell(row=curr_row + 1, column=6, value="ហត្ថលេខាគ្រូបន្ទុកថ្នាក់").font = font_sig_title
        ws.cell(row=curr_row + 4, column=6, value=homeroom_str).font = font_sig_name

        # Column widths
        ws.column_dimensions['A'].width = 10
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 23

    # Helper: Build 1 Teacher Timetable Sheet
    def render_teacher_sheet(workbook, tch):
        sheet_name = sanitize_title(f"{tch.khmer_name}")
        counter = 1
        base_name = sheet_name
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{base_name[:28]}_{counter}"
            counter += 1

        ws = workbook.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        tch_gender_title = "អ្នកគ្រូ" if tch.gender == 'F' else "លោកគ្រូ"
        full_name = f"{tch_gender_title} {tch.khmer_name}"

        # Header Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្តែត"
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:G2')
        ws['A2'] = "កាលវិភាគបង្រៀនប្រចាំសប្តាហ៍"
        ws['A2'].font = font_sub
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        assigned_subs = sorted(list(teacher_subjects_map.get(tch.id, set())))
        subs_str = ", ".join(assigned_subs) if assigned_subs else (tch.specialization or "-")
        
        tch_entries = timetables_by_teacher.get(tch.id, [])
        teaching_hours = len(tch_entries)
        tch_duties = duty_by_teacher.get(tch.id, [])
        duty_hours = len(tch_duties)
        total_hours = teaching_hours + duty_hours

        ws.merge_cells('A3:G3')
        ws['A3'] = f"ឈ្មោះគ្រូបង្រៀន៖ {full_name} | មុខវិជ្ជា៖ {subs_str} | ម៉ោងបង្រៀន៖ {teaching_hours} | ម៉ោងប្រចាំការ៖ {duty_hours} | សរុប៖ {total_hours} ម៉ោង | ឆ្នាំសិក្សា៖ {year_name}"
        ws['A3'].font = font_meta
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

        # Headers Row 5
        headers = ['ម៉ោងទី', 'ច័ន្ទ', 'អង្គារ', 'ពុធ', 'ព្រហស្បតិ៍', 'សុក្រ', 'សៅរ៍']
        ws.append([]) # Row 4
        ws.append(headers) # Row 5
        for col_idx in range(1, 8):
            cell = ws.cell(row=5, column=col_idx)
            cell.font = font_head
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_header
            cell.border = thin_border
        ws.row_dimensions[5].height = 26

        slots_map = {}
        is_duty_map = {}
        for entry in tch_entries:
            sub_name = entry.subject.name_kh if entry.subject else ""
            cls_name = entry.classroom.name if entry.classroom else ""
            slots_map[(entry.day_of_week, entry.period_number)] = f"{sub_name}\n(ថ្នាក់ {cls_name})" if (sub_name and cls_name) else (sub_name or "-")
            is_duty_map[(entry.day_of_week, entry.period_number)] = False

        for d in tch_duties:
            k = (d.day_of_week, d.period_number)
            if k not in slots_map:
                duty_name = duty_types_dict.get(d.duty_type, 'ប្រចាំការ')
                if duty_name in ['ប្រចាំការទូទៅ', 'GENERAL']:
                    duty_name = 'ប្រចាំការ'
                slots_map[k] = duty_name
                is_duty_map[k] = True

        # Session 1: ពេលព្រឹក (Row 6)
        ws.append(['ពេលព្រឹក', '', '', '', '', '', ''])
        ws.merge_cells('A6:G6')
        for col_idx in range(1, 8):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_session
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_session
            cell.border = thin_border
        ws.row_dimensions[6].height = 22

        # Periods 1 to 4 (Rows 7 to 10)
        curr_row = 7
        for p in [1, 2, 3, 4]:
            row_data = [p] + [slots_map.get((d['num'], p), '-') for d in DAYS_OF_WEEK]
            ws.append(row_data)
            ws.row_dimensions[curr_row].height = 36
            for col_idx in range(1, 8):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if col_idx == 1:
                    cell.font = font_head
                else:
                    is_duty = is_duty_map.get((DAYS_OF_WEEK[col_idx-2]['num'], p), False)
                    if is_duty:
                        cell.font = font_duty
                        cell.fill = fill_duty
                    else:
                        cell.font = font_slot_bold
            curr_row += 1

        # Session 2: ពេលរសៀល (Row 11)
        ws.append(['ពេលរសៀល', '', '', '', '', '', ''])
        ws.merge_cells('A11:G11')
        for col_idx in range(1, 8):
            cell = ws.cell(row=11, column=col_idx)
            cell.font = font_session
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_session
            cell.border = thin_border
        ws.row_dimensions[11].height = 22

        # Periods 5 to 8 (Rows 12 to 15)
        curr_row = 12
        for p in [5, 6, 7, 8]:
            row_data = [p] + [slots_map.get((d['num'], p), '-') for d in DAYS_OF_WEEK]
            ws.append(row_data)
            ws.row_dimensions[curr_row].height = 36
            for col_idx in range(1, 8):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if col_idx == 1:
                    cell.font = font_head
                else:
                    is_duty = is_duty_map.get((DAYS_OF_WEEK[col_idx-2]['num'], p), False)
                    if is_duty:
                        cell.font = font_duty
                        cell.fill = fill_duty
                    else:
                        cell.font = font_slot_bold
            curr_row += 1

        # Signatures (Rows 17-20)
        curr_row += 1
        ws.cell(row=curr_row + 1, column=2, value="បានឃើញ និងឯកភាព").font = font_sig_title
        ws.cell(row=curr_row + 2, column=2, value="នាយកសាលា").font = font_sig_title
        ws.cell(row=curr_row + 1, column=6, value="ហត្ថលេខាសាមីខ្លួន").font = font_sig_title
        ws.cell(row=curr_row + 4, column=6, value=f"{tch.khmer_name}").font = font_sig_name

        # Column widths
        ws.column_dimensions['A'].width = 10
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 23

    # EXECUTE REQUESTED EXPORT
    if mode == 'teacher':
        if target_id != 'all' and str(target_id).isdigit():
            single_teacher = next((t for t in teachers if t.id == int(target_id)), None)
            if single_teacher:
                render_teacher_sheet(wb, single_teacher)
                filename = f"timetable_teacher_{single_teacher.khmer_name}.xlsx"
            else:
                for tch in teachers:
                    render_teacher_sheet(wb, tch)
                filename = "all_teachers_timetables.xlsx"
        else:
            # All Teachers into separate sheets
            for tch in teachers:
                render_teacher_sheet(wb, tch)
            filename = "all_teachers_timetables.xlsx"
    else:
        # mode == 'class'
        if target_id != 'all' and str(target_id).isdigit():
            single_class = next((c for c in classrooms if c.id == int(target_id)), None)
            if single_class:
                render_classroom_sheet(wb, single_class)
                filename = f"timetable_class_{single_class.name}.xlsx"
            else:
                for cls in classrooms:
                    render_classroom_sheet(wb, cls)
                filename = "all_classrooms_timetables.xlsx"
        else:
            # All Classrooms into separate sheets
            for cls in classrooms:
                render_classroom_sheet(wb, cls)
            filename = "all_classrooms_timetables.xlsx"

    # If no sheets were created, create a blank placeholder
    if not wb.sheetnames:
        ws = wb.create_sheet(title="គ្មានទិន្នន័យ")
        ws['A1'] = "គ្មានទិន្នន័យកាលវិភាគសម្រាប់ឆ្នាំសិក្សាដែលបានជ្រើសរើសឡើយ"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def timetable_daily_reports_view(request):
    """
    Daily Duty Sign-In Sheets Generator & Related Timetable Reports.
    Matches MoEYS standard layout: No, Teacher ID, Name, Period 1-4 / 5-8, Sign In, Sign Out, Remarks.
    Strictly isolated per Academic Year!
    """
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
        else:
            found_year = AcademicYear.objects.filter(name=str(selected_year).strip()).first()
        if found_year:
            active_year = found_year

    teachers = list(Teacher.objects.filter(status='ACTIVE').order_by('teacher_id', 'khmer_name'))
    classrooms = list(Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code'))
    subjects = list(Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id'))
    subjects_by_id = {s.id: s for s in subjects}
    timetables_qs = Timetable.objects.filter(classroom__academic_year=active_year).select_related('classroom', 'subject', 'teacher') if active_year else Timetable.objects.select_related('classroom', 'subject', 'teacher').all()
    timetables = list(timetables_qs)

    timetables_by_day = defaultdict(list)
    timetables_by_teacher = defaultdict(int)
    timetables_by_classroom = defaultdict(int)
    for entry in timetables:
        timetables_by_day[entry.day_of_week].append(entry)
        if entry.teacher_id:
            timetables_by_teacher[entry.teacher_id] += 1
        if entry.classroom_id:
            timetables_by_classroom[entry.classroom_id] += 1

    # Pre-fetch rules & assignments
    grade_rules = list(GradeLevelRule.objects.filter(weekly_hours__gt=0).exclude(subject__code__in=['R', 'D']))
    requirements_map = {}
    for r in grade_rules:
        k = (r.grade_level, r.track)
        if k not in requirements_map:
            requirements_map[k] = {}
        requirements_map[k][r.subject_id] = r.weekly_hours

    # Teacher subject code map
    class_subject_assignments_qs = ClassSubject.objects.filter(
        classroom__academic_year=active_year,
        teacher__isnull=False
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('classroom', 'subject', 'teacher') if active_year else ClassSubject.objects.filter(
        teacher__isnull=False
    ).exclude(
        subject__code__in=['R', 'D']
    ).select_related('classroom', 'subject', 'teacher')
    class_subject_assignments = list(class_subject_assignments_qs)

    cs_by_teacher = defaultdict(list)
    cs_pairs_set = set()
    teacher_subject_classes_map = defaultdict(list)
    for cs in class_subject_assignments:
        if cs.teacher_id:
            cs_by_teacher[cs.teacher_id].append(cs)
            cs_pairs_set.add((cs.subject_id, cs.teacher_id))
            if cs.classroom:
                teacher_subject_classes_map[(cs.subject_id, cs.teacher_id)].append(cs.classroom.name)

    distinct_assignments = sorted(list(cs_pairs_set), key=lambda x: (x[0] or 0, x[1] or 0))

    teacher_subject_code_map = {}
    subject_teacher_counters = {}
    for (s_id, t_id) in distinct_assignments:
        sub = subjects_by_id.get(s_id)
        sub_code = sub.code if sub and sub.code else 'S'
        if s_id not in subject_teacher_counters:
            subject_teacher_counters[s_id] = 1
        else:
            subject_teacher_counters[s_id] += 1
        teacher_subject_code_map[(s_id, t_id)] = f"{sub_code}{subject_teacher_counters[s_id]}"

    # Selected filters from request
    selected_day = request.GET.get('day', 'all')
    selected_session = request.GET.get('session', 'all')
    selected_tab = request.GET.get('tab', 'duty_sheets')

    days_to_render = DAYS_OF_WEEK if selected_day == 'all' else [d for d in DAYS_OF_WEEK if str(d['num']) == str(selected_day)]

    # 1. Query Duty Schedules for this academic year
    duty_entries_qs = TeacherDutySchedule.objects.filter(academic_year=active_year).select_related('teacher') if active_year else TeacherDutySchedule.objects.all().select_related('teacher')
    duty_entries = list(duty_entries_qs)
    duty_by_day = defaultdict(list)
    duty_by_teacher = defaultdict(int)
    for duty_s in duty_entries:
        duty_by_day[duty_s.day_of_week].append(duty_s)
        if duty_s.teacher_id:
            duty_by_teacher[duty_s.teacher_id] += 1

    # Build Duty Sign-In Sheets (Includes Classroom Teaching & On-Duty Staff/Teachers)
    duty_sheets = []
    for d in days_to_render:
        d_entries = timetables_by_day.get(d['num'], [])
        d_duties = duty_by_day.get(d['num'], [])

        day_teacher_slots = {}
        # 1. Populate Classroom Teaching Slots
        for entry in d_entries:
            if entry.teacher_id:
                if entry.teacher_id not in day_teacher_slots:
                    day_teacher_slots[entry.teacher_id] = {}
                p_num = entry.period_number or 1
                sub_code = entry.subject.code if entry.subject and entry.subject.code else 'S'
                slot_code = teacher_subject_code_map.get((entry.subject_id, entry.teacher_id), sub_code)
                cls_name = entry.classroom.code if (entry.classroom and entry.classroom.code) else (entry.classroom.name if entry.classroom else '')
                day_teacher_slots[entry.teacher_id][p_num] = f"{cls_name}({slot_code})"

        # 2. Populate On-Duty Shifts (for office staff and teachers on duty)
        for duty_s in d_duties:
            if duty_s.teacher_id:
                if duty_s.teacher_id not in day_teacher_slots:
                    day_teacher_slots[duty_s.teacher_id] = {}
                p_num = duty_s.period_number or 1
                if p_num not in day_teacher_slots[duty_s.teacher_id]:
                    day_teacher_slots[duty_s.teacher_id][p_num] = "ប្រចាំការ"

        # Morning session (Periods 1, 2, 3, 4)
        if selected_session in ['all', 'morning']:
            morning_rows = []
            no_idx = 1
            for tch in teachers:
                tch_slots = day_teacher_slots.get(tch.id, {})
                p1 = tch_slots.get(1, '-')
                p2 = tch_slots.get(2, '-')
                p3 = tch_slots.get(3, '-')
                p4 = tch_slots.get(4, '-')
                has_classes = any(p != '-' for p in [p1, p2, p3, p4])
                
                if has_classes:
                    gender_title = "អ្នកគ្រូ" if tch.gender == 'F' else "លោកគ្រូ"
                    tch_name = (tch.khmer_name or tch.name or '')
                    morning_rows.append({
                        'no': no_idx,
                        'teacher_id': tch.teacher_id or '',
                        'teacher_name': f"{gender_title} {tch_name}",
                        'specialization': tch.specialization or tch.current_duty or 'បុគ្គលិក',
                        'p1': p1,
                        'p2': p2,
                        'p3': p3,
                        'p4': p4,
                    })
                    no_idx += 1

            duty_sheets.append({
                'day_num': d['num'],
                'day_name': d['name_kh'],
                'session': 'morning',
                'session_name': 'ពេលព្រឹក',
                'session_badge': 'ព្រឹក (ម៉ោង ១-៤)',
                'period_labels': ['ម៉ោងទី១', 'ម៉ោងទី២', 'ម៉ោងទី៣', 'ម៉ោងទី៤'],
                'rows': morning_rows,
                'total_teachers': len(morning_rows),
            })

        # Afternoon session (Periods 5, 6, 7, 8)
        if selected_session in ['all', 'afternoon']:
            afternoon_rows = []
            no_idx = 1
            for tch in teachers:
                tch_slots = day_teacher_slots.get(tch.id, {})
                p5 = tch_slots.get(5, '-')
                p6 = tch_slots.get(6, '-')
                p7 = tch_slots.get(7, '-')
                p8 = tch_slots.get(8, '-')
                has_classes = any(p != '-' for p in [p5, p6, p7, p8])
                
                if has_classes:
                    gender_title = "អ្នកគ្រូ" if tch.gender == 'F' else "លោកគ្រូ"
                    tch_name = (tch.khmer_name or tch.name or '')
                    afternoon_rows.append({
                        'no': no_idx,
                        'teacher_id': tch.teacher_id or '',
                        'teacher_name': f"{gender_title} {tch_name}",
                        'specialization': tch.specialization or tch.current_duty or 'បុគ្គលិក',
                        'p1': p5,
                        'p2': p6,
                        'p3': p7,
                        'p4': p8,
                    })
                    no_idx += 1

            duty_sheets.append({
                'day_num': d['num'],
                'day_name': d['name_kh'],
                'session': 'afternoon',
                'session_name': 'ពេលរសៀល',
                'session_badge': 'រសៀល (ម៉ោង ៥-៨)',
                'period_labels': ['ម៉ោងទី៥', 'ម៉ោងទី៦', 'ម៉ោងទី៧', 'ម៉ោងទី៨'],
                'rows': afternoon_rows,
                'total_teachers': len(afternoon_rows),
            })

    # 2. Teacher Teaching & Duty Hours Load Report
    teacher_load_report = []
    for t in teachers:
        t_slots_count = timetables_by_teacher.get(t.id, 0)
        t_duty_count = duty_by_teacher.get(t.id, 0)
        t_total_actual = t_slots_count + t_duty_count

        t_assigned_cs = cs_by_teacher.get(t.id, [])
        t_assigned_sum = 0
        for cs in t_assigned_cs:
            cls = cs.classroom
            if cls:
                cls_reqs = requirements_map.get((cls.grade_level, cls.track), {})
                if not cls_reqs:
                    cls_reqs = requirements_map.get((cls.grade_level, 'GENERAL'), {})
                t_assigned_sum += cls_reqs.get(cs.subject_id, 0)

        t_max = t.max_weekly_hours or 18
        t_codes = [
            code for (s_id, t_id), code in teacher_subject_code_map.items() 
            if t_id == t.id and (s_id, t_id) in cs_pairs_set
        ]

        if t_total_actual > t_max:
            status_text = f"លើសម៉ោងកំណត់ ({t_total_actual - t_max} ម៉ោង)"
            status_badge = "bg-warning text-dark border border-warning"
        elif t_total_actual == t_max:
            status_text = "គ្រប់ម៉ោងតាមការកំណត់ (100%)"
            status_badge = "bg-success text-white"
        elif t_total_actual < t_max and t_total_actual > 0:
            status_text = f"ខ្វះ {t_max - t_total_actual} ម៉ោង"
            status_badge = "bg-info text-dark"
        else:
            status_text = "មិនទាន់មានម៉ោង"
            status_badge = "bg-secondary text-white"

        teacher_load_report.append({
            'teacher': t,
            'codes': ", ".join(t_codes) or "-",
            'max_hours': t_max,
            'assigned_hours': t_assigned_sum,
            'scheduled_hours': t_slots_count,
            'duty_hours': t_duty_count,
            'total_actual_hours': t_total_actual,
            'diff': t_total_actual - t_max,
            'status_text': status_text,
            'status_badge': status_badge,
        })

    # 3. Teacher Subject Code Directory
    teacher_code_directory = []
    for (s_id, t_id), code in sorted(teacher_subject_code_map.items(), key=lambda x: str(x[1])):
        if (s_id, t_id) in cs_pairs_set:
            sub = subjects_by_id.get(s_id)
            tch = next((t for t in teachers if t.id == t_id), None)
            if sub and tch:
                assigned_classes = sorted(list(set(teacher_subject_classes_map.get((s_id, t_id), []))))
                cls_names = ", ".join(assigned_classes) or "-"
                teacher_code_directory.append({
                    'code': code,
                    'subject': sub,
                    'teacher': tch,
                    'classes': cls_names,
                })

    # 4. Classrooms Summary
    classrooms_summary = []
    for cls in classrooms:
        cls_reqs = requirements_map.get((cls.grade_level, cls.track), {})
        if not cls_reqs:
            cls_reqs = requirements_map.get((cls.grade_level, 'GENERAL'), {})
        req_total = sum(cls_reqs.values())
        sched_total = timetables_by_classroom.get(cls.id, 0)
        diff = sched_total - req_total
        classrooms_summary.append({
            'classroom': cls,
            'required_hours': req_total,
            'scheduled_hours': sched_total,
            'diff': diff,
            'status_class': 'success' if diff == 0 else ('warning text-dark' if diff > 0 else 'info text-dark'),
            'status_text': 'ពេញលេញ' if diff == 0 else (f'លើស {diff} ម៉ោង' if diff > 0 else f'ខ្វះ {-diff} ម៉ោង'),
        })

    # Date in Khmer
    now = datetime.datetime.now()
    khmer_digits = {'0': '០', '1': '១', '2': '២', '3': '៣', '4': '៤', '5': '៥', '6': '៦', '7': '៧', '8': '៨', '9': '៩'}
    def to_khmer_num(n):
        return ''.join(khmer_digits.get(c, c) for c in str(n))

    kh_months = ['', 'មករា', 'កុម្ភៈ', 'មីនា', 'មេសា', 'ឧសភា', 'មិថុនា', 'កក្កដា', 'សីហា', 'កញ្ញា', 'តុលា', 'វិច្ឆិកា', 'ធ្នូ']
    kh_days_name = ['ច័ន្ទ', 'អង្គារ', 'ពុធ', 'ព្រហស្បតិ៍', 'សុក្រ', 'សៅរ៍', 'អាទិត្យ']

    today_kh_day = to_khmer_num(now.day)
    today_kh_month = kh_months[now.month] if 1 <= now.month <= 12 else ''
    today_kh_year = to_khmer_num(now.year)
    today_kh_dow = kh_days_name[now.weekday()]

    context = {
        'days': DAYS_OF_WEEK,
        'selected_day': selected_day,
        'selected_session': selected_session,
        'selected_tab': selected_tab,
        'duty_sheets': duty_sheets,
        'teacher_load_report': teacher_load_report,
        'teacher_code_directory': teacher_code_directory,
        'classrooms_summary': classrooms_summary,
        'academic_year': active_year,
        'today_kh_day': today_kh_day,
        'today_kh_month': today_kh_month,
        'today_kh_year': today_kh_year,
        'today_kh_dow': today_kh_dow,
    }
    return render(request, 'academics/daily_reports.html', context)


@login_required
def timetable_daily_reports_export_excel(request):
    """
    Exports formatted Timetable Reports to Excel (.xlsx) using openpyxl.
    Supports 4 Report Types:
      1. 'duty_sheets'  : Daily Duty Sign-In Sheets (១២ ថស / តាមថ្ងៃ)
      2. 'teacher_load'  : Teacher Teaching Load Report (បន្ទុកបង្រៀនរបស់គ្រូ)
      3. 'subject_codes' : Teacher Subject Codes Directory (កូដគ្រូ-មុខវិជ្ជា)
      4. 'class_summary' : Classrooms Timetable Summary (ម៉ោងតាមថ្នាក់រៀន)
      5. 'all_reports'   : All 4 reports combined into a single workbook!
    Strictly for the active academic year!
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from .utils import get_active_academic_year
    academic_year = get_active_academic_year(request)
    year_name = academic_year.name if academic_year else '២០២៦-២០២៧'

    report_type = request.GET.get('report_type') or request.GET.get('tab') or 'duty_sheets'
    selected_day = request.GET.get('day', 'all')
    selected_session = request.GET.get('session', 'all')

    teachers = list(Teacher.objects.filter(status='ACTIVE').order_by('teacher_id', 'khmer_name'))
    subjects = list(Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id'))
    subjects_by_id = {s.id: s for s in subjects}
    classrooms = list(Classroom.objects.filter(academic_year=academic_year).order_by('grade_level', 'code')) if academic_year else list(Classroom.objects.all().order_by('grade_level', 'code'))
    timetables = list(Timetable.objects.filter(classroom__academic_year=academic_year).select_related('classroom', 'subject', 'teacher')) if academic_year else list(Timetable.objects.select_related('classroom', 'subject', 'teacher').all())

    # Build ClassSubject relations
    cs_query = ClassSubject.objects.filter(classroom__academic_year=academic_year, teacher__isnull=False).select_related('classroom', 'teacher', 'subject') if academic_year else ClassSubject.objects.filter(teacher__isnull=False).select_related('classroom', 'teacher', 'subject')
    cs_pairs_set = set(cs_query.values_list('subject_id', 'teacher_id'))
    cs_by_teacher = defaultdict(list)
    teacher_subject_classes_map = defaultdict(list)
    for cs in cs_query:
        if cs.teacher_id:
            cs_by_teacher[cs.teacher_id].append(cs)
            if cs.subject_id and cs.classroom:
                teacher_subject_classes_map[(cs.subject_id, cs.teacher_id)].append(cs.classroom.code or cs.classroom.name)

    # Teacher subject code map
    distinct_assignments = sorted(list(cs_pairs_set), key=lambda x: (x[0], x[1]))
    teacher_subject_code_map = {}
    subject_teacher_counters = {}
    for s_id, t_id in distinct_assignments:
        sub = subjects_by_id.get(s_id)
        sub_code = sub.code if sub else 'S'
        if s_id not in subject_teacher_counters:
            subject_teacher_counters[s_id] = 1
        else:
            subject_teacher_counters[s_id] += 1
        teacher_subject_code_map[(s_id, t_id)] = f"{sub_code}{subject_teacher_counters[s_id]}"

    # Requirements map
    requirements_map = {}
    for r in GradeLevelRule.objects.all():
        key = (r.grade_level, r.track)
        if key not in requirements_map:
            requirements_map[key] = {}
        requirements_map[key][r.subject_id] = r.weekly_hours or 0

    # Scheduled hours map
    timetables_by_teacher = defaultdict(int)
    timetables_by_classroom = defaultdict(int)
    for e in timetables:
        if e.teacher_id:
            timetables_by_teacher[e.teacher_id] += 1
        if e.classroom_id:
            timetables_by_classroom[e.classroom_id] += 1

    all_duties = list(TeacherDutySchedule.objects.filter(academic_year=academic_year).select_related('teacher')) if academic_year else list(TeacherDutySchedule.objects.select_related('teacher').all())
    duty_hours_by_teacher = defaultdict(int)
    for duty in all_duties:
        if duty.teacher_id:
            duty_hours_by_teacher[duty.teacher_id] += 1

    # Styling Tokens
    font_title = Font(name='Khmer OS Muol Light', size=13, bold=True)
    font_sub = Font(name='Khmer OS Muol Light', size=13, bold=True)
    font_meta = Font(name='Khmer OS Siemreap', size=10, italic=True)
    font_head = Font(name='Khmer OS Siemreap', size=10, bold=True)
    font_body = Font(name='Khmer OS Siemreap', size=10)
    font_bold = Font(name='Khmer OS Siemreap', size=10, bold=True)

    fill_header = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    fill_accent = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    fill_success = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    fill_warning = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='A0AEC0'),
        right=Side(style='thin', color='A0AEC0'),
        top=Side(style='thin', color='A0AEC0'),
        bottom=Side(style='thin', color='A0AEC0')
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default blank sheet

    def auto_fit_columns(ws, min_widths=None):
        min_widths = min_widths or {}
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val:
                    # Account for unicode/khmer length approx
                    length = len(val) + sum(1 for c in val if ord(c) > 127) * 0.4
                    if length > max_len:
                        max_len = length
            min_w = min_widths.get(col[0].column, 10)
            ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, 50))

    # ---------------- 1. TEACHER TEACHING LOAD SHEET BUILDER ----------------
    def build_teacher_load_sheet(workbook):
        ws = workbook.create_sheet(title="បន្ទុកបង្រៀនរបស់គ្រូ"[:31])
        ws.views.sheetView[0].showGridLines = True

        # Header Title
        ws.merge_cells('A1:L1')
        ws['A1'] = "វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្តែត"
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:L2')
        ws['A2'] = "របាយការណ៍បន្ទុកម៉ោងបង្រៀនរបស់គ្រូ (Teacher Teaching Hours Load Report)"
        ws['A2'].font = font_sub
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A3:L3')
        ws['A3'] = f"ឆ្នាំសិក្សា៖ {year_name} | សរុបគ្រូ៖ {len(teachers)} នាក់"
        ws['A3'].font = font_meta
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

        headers = [
            'ល.រ', 'អត្តលេខ', 'ឈ្មោះគ្រូបង្រៀន', 'ភេទ', 'ឯកទេស', 'កូដមុខវិជ្ជា',
            'ម៉ោងកំណត់គោល', 'ម៉ោងចាត់តាំង', 'រៀបចំក្នុងកាលវិភាគ', 'ម៉ោងប្រចាំការ', 'ម៉ោងសរុបជាក់ស្តែង', 'ខុសគ្នា', 'ស្ថានភាព'
        ]
        ws.append([]) # Row 4
        ws.append(headers) # Row 5

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.font = font_head
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_header
            cell.border = thin_border

        curr_row = 6
        tot_max = 0
        tot_assigned = 0
        tot_scheduled = 0
        tot_duty = 0
        tot_actual = 0

        for idx, t in enumerate(teachers, start=1):
            t_slots = timetables_by_teacher.get(t.id, 0)
            t_duty = duty_hours_by_teacher.get(t.id, 0)
            t_total_actual = t_slots + t_duty

            t_assigned_cs = cs_by_teacher.get(t.id, [])
            t_assigned_sum = 0
            for cs in t_assigned_cs:
                cls = cs.classroom
                if cls:
                    cls_reqs = requirements_map.get((cls.grade_level, cls.track), {})
                    if not cls_reqs:
                        cls_reqs = requirements_map.get((cls.grade_level, 'GENERAL'), {})
                    t_assigned_sum += cls_reqs.get(cs.subject_id, 0)

            t_max = t.max_weekly_hours or 18
            t_codes = [
                code for (s_id, t_id), code in teacher_subject_code_map.items() 
                if t_id == t.id and (s_id, t_id) in cs_pairs_set
            ]
            codes_str = ", ".join(t_codes) or "-"
            diff = t_total_actual - t_max

            if t_total_actual > t_max:
                status_text = f"លើស {diff} ម៉ោង"
            elif t_total_actual == t_max:
                status_text = "គ្រប់ម៉ោង (100%)"
            elif t_total_actual < t_max and t_total_actual > 0:
                status_text = f"ខ្វះ {-diff} ម៉ោង"
            else:
                status_text = "មិនទាន់មានម៉ោង"

            gender_str = "ស្រី" if t.gender == 'F' else "ប្រុស"
            name_str = f"{t.khmer_name} ({t.latin_name})" if t.latin_name else t.khmer_name

            row_vals = [
                idx, t.teacher_id, name_str, gender_str, t.specialization or '-', codes_str,
                t_max, t_assigned_sum, t_slots, t_duty, t_total_actual, diff, status_text
            ]
            ws.append(row_vals)

            tot_max += t_max
            tot_assigned += t_assigned_sum
            tot_scheduled += t_slots
            tot_duty += t_duty
            tot_actual += t_total_actual

            for col_idx in range(1, len(row_vals) + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = font_body
                cell.border = thin_border
                if col_idx in [1, 2, 4, 6, 7, 8, 9, 10, 11, 12, 13]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            curr_row += 1

        # Total summary row
        tot_diff = tot_actual - tot_max
        summary_row = [
            'សរុប', '', '', '', '', '',
            tot_max, tot_assigned, tot_scheduled, tot_duty, tot_actual, tot_diff, ''
        ]
        ws.append(summary_row)
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws.cell(row=curr_row, column=1, value="សរុបម៉ោងទូទាំងសាលា").alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(1, len(summary_row) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = font_bold
            cell.border = thin_border
            cell.fill = fill_accent
            if col_idx > 6:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        auto_fit_columns(ws, {1: 6, 2: 14, 3: 26, 4: 8, 5: 18, 6: 14, 7: 12, 8: 12, 9: 14, 10: 12, 11: 14, 12: 10, 13: 16})

    # ---------------- 2. TEACHER SUBJECT CODES DIRECTORY BUILDER ----------------
    def build_subject_codes_sheet(workbook):
        ws = workbook.create_sheet(title="កូដគ្រូ-មុខវិជ្ជា"[:31])
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells('A1:F1')
        ws['A1'] = "វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្តែត"
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = "បញ្ជីកូដគ្រូបង្រៀន និងមុខវិជ្ជា (Teacher Subject Codes Directory)"
        ws['A2'].font = font_sub
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A3:F3')
        ws['A3'] = f"ឆ្នាំសិក្សា៖ {year_name}"
        ws['A3'].font = font_meta
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

        headers = ['ល.រ', 'កូដសម្គាល់', 'មុខវិជ្ជាបង្រៀន', 'ឈ្មោះគ្រូទទួលបន្ទុក', 'អត្តលេខគ្រូ', 'ថ្នាក់ដែលត្រូវបង្រៀន']
        ws.append([])
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.font = font_head
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_header
            cell.border = thin_border

        curr_row = 6
        no_idx = 1
        for (s_id, t_id), code in sorted(teacher_subject_code_map.items(), key=lambda x: str(x[1])):
            if (s_id, t_id) in cs_pairs_set:
                sub = subjects_by_id.get(s_id)
                tch = next((t for t in teachers if t.id == t_id), None)
                if sub and tch:
                    assigned_classes = sorted(list(set(teacher_subject_classes_map.get((s_id, t_id), []))))
                    cls_names = ", ".join(assigned_classes) or "-"
                    tch_name = f"{tch.khmer_name} ({tch.latin_name})" if tch.latin_name else tch.khmer_name

                    row_vals = [no_idx, code, sub.name_kh, tch_name, tch.teacher_id, cls_names]
                    ws.append(row_vals)

                    for col_idx in range(1, len(row_vals) + 1):
                        cell = ws.cell(row=curr_row, column=col_idx)
                        cell.font = font_body
                        cell.border = thin_border
                        if col_idx in [1, 2, 3, 5]:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')

                    curr_row += 1
                    no_idx += 1

        auto_fit_columns(ws, {1: 6, 2: 14, 3: 20, 4: 24, 5: 14, 6: 30})

    # ---------------- 3. CLASSROOMS SUMMARY SHEET BUILDER ----------------
    def build_class_summary_sheet(workbook):
        ws = workbook.create_sheet(title="ម៉ោងតាមថ្នាក់រៀន"[:31])
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells('A1:I1')
        ws['A1'] = "វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្តែត"
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:I2')
        ws['A2'] = "របាយការណ៍ម៉ោងតាមថ្នាក់រៀន (Classrooms Timetable Allocation Summary)"
        ws['A2'].font = font_sub
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A3:I3')
        ws['A3'] = f"ឆ្នាំសិក្សា៖ {year_name} | សរុបថ្នាក់៖ {len(classrooms)} ថ្នាក់"
        ws['A3'].font = font_meta
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

        headers = ['ល.រ', 'ឈ្មោះថ្នាក់', 'កម្រិតថ្នាក់', 'ផ្នែក/Track', 'គ្រូបន្ទុកថ្នាក់', 'ម៉ោងកំណត់ក្នុងថ្នាក់', 'ម៉ោងរៀបចំជាក់ស្តែង', 'ខុសគ្នា', 'ស្ថានភាព']
        ws.append([])
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.font = font_head
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_header
            cell.border = thin_border

        curr_row = 6
        tot_req = 0
        tot_sched = 0

        for idx, cls in enumerate(classrooms, start=1):
            cls_reqs = requirements_map.get((cls.grade_level, cls.track), {})
            if not cls_reqs:
                cls_reqs = requirements_map.get((cls.grade_level, 'GENERAL'), {})
            req_total = sum(cls_reqs.values())
            sched_total = timetables_by_classroom.get(cls.id, 0)
            diff = sched_total - req_total
            status_text = 'ពេញលេញ (100%)' if diff == 0 else (f'លើស {diff} ម៉ោង' if diff > 0 else f'ខ្វះ {-diff} ម៉ោង')

            homeroom_name = cls.homeroom_teacher.khmer_name if (cls.homeroom_teacher and hasattr(cls, 'homeroom_teacher')) else '-'

            row_vals = [
                idx, cls.name, f"ថ្នាក់ទី {cls.grade_level}", cls.get_track_display() if hasattr(cls, 'get_track_display') else cls.track,
                homeroom_name, req_total, sched_total, diff, status_text
            ]
            ws.append(row_vals)

            tot_req += req_total
            tot_sched += sched_total

            for col_idx in range(1, len(row_vals) + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = font_body
                cell.border = thin_border
                if col_idx in [1, 2, 3, 4, 6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            curr_row += 1

        # Summary Row
        tot_diff = tot_sched - tot_req
        summary_row = ['សរុប', '', '', '', '', tot_req, tot_sched, tot_diff, '']
        ws.append(summary_row)
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
        ws.cell(row=curr_row, column=1, value="សរុបម៉ោងថ្នាក់រៀនទាំងអស់").alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(1, len(summary_row) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = font_bold
            cell.border = thin_border
            cell.fill = fill_accent
            if col_idx > 5:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        auto_fit_columns(ws, {1: 6, 2: 16, 3: 14, 4: 16, 5: 20, 6: 16, 7: 16, 8: 12, 9: 18})

    # ---------------- 4. DAILY DUTY SIGN-IN SHEETS BUILDER ----------------
    def build_duty_sheets(workbook):
        days_to_render = DAYS_OF_WEEK if selected_day == 'all' else [d for d in DAYS_OF_WEEK if str(d['num']) == str(selected_day)]

        for d in days_to_render:
            d_entries = [e for e in timetables if e.day_of_week == d['num']]
            d_duties = [duty for duty in all_duties if duty.day_of_week == d['num']]

            day_teacher_slots = {}
            for entry in d_entries:
                if entry.teacher_id:
                    if entry.teacher_id not in day_teacher_slots:
                        day_teacher_slots[entry.teacher_id] = {}
                    p_num = entry.period_number or 1
                    slot_code = teacher_subject_code_map.get((entry.subject_id, entry.teacher_id), entry.subject.code if entry.subject else 'S')
                    cls_name = entry.classroom.code or entry.classroom.name if entry.classroom else '-'
                    day_teacher_slots[entry.teacher_id][p_num] = f"{cls_name}({slot_code})"

            # Populate on-duty shifts
            for duty_s in d_duties:
                if duty_s.teacher_id:
                    if duty_s.teacher_id not in day_teacher_slots:
                        day_teacher_slots[duty_s.teacher_id] = {}
                    p_num = duty_s.period_number or 1
                    if p_num not in day_teacher_slots[duty_s.teacher_id]:
                        day_teacher_slots[duty_s.teacher_id][p_num] = "ប្រចាំការ"

            sessions = []
            if selected_session in ['all', 'morning']:
                sessions.append(('morning', 'ពេលព្រឹក', [1, 2, 3, 4], ['ម៉ោងទី១', 'ម៉ោងទី២', 'ម៉ោងទី៣', 'ម៉ោងទី៤']))
            if selected_session in ['all', 'afternoon']:
                sessions.append(('afternoon', 'ពេលរសៀល', [5, 6, 7, 8], ['ម៉ោងទី៥', 'ម៉ោងទី៦', 'ម៉ោងទី៧', 'ម៉ោងទី៨']))

            for sess_code, sess_name, p_nums, p_labels in sessions:
                sheet_title = f"{d['name_kh']}_{sess_name}"[:31]
                ws = workbook.create_sheet(title=sheet_title)
                ws.views.sheetView[0].showGridLines = True

                # Title
                ws.merge_cells('A1:D1')
                ws['A1'] = "វិទ្យាល័យ ហ៊ុន សែន កំពង់កន្តែត"
                ws['A1'].font = font_title
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

                ws.merge_cells('A2:J2')
                ws['A2'] = "បញ្ជីចុះហត្ថលេខាវត្តមានគ្រូបង្រៀនប្រចាំថ្ងៃ"
                ws['A2'].font = font_sub
                ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

                ws.merge_cells('A3:J3')
                ws['A3'] = f"ឆ្នាំសិក្សា៖ {year_name} ថ្ងៃ {d['name_kh']} {sess_name}"
                ws['A3'].font = font_meta
                ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

                # Table Header Row
                headers = ['ល.រ', 'អត្តលេខ', 'ឈ្មោះគ្រូបង្រៀន'] + p_labels + ['ហត្ថលេខាចូល', 'ហត្ថលេខាចេញ', 'ផ្សេងៗ']
                ws.append([])
                ws.append(headers)

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=5, column=col_idx)
                    cell.font = font_head
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = fill_header
                    cell.border = thin_border

                current_row = 6
                no_idx = 1
                for tch in teachers:
                    tch_slots = day_teacher_slots.get(tch.id, {})
                    p_vals = [tch_slots.get(p, '-') for p in p_nums]
                    if any(v != '-' for v in p_vals):
                        gender_title = "អ្នកគ្រូ" if tch.gender == 'F' else "លោកគ្រូ"
                        row_data = [no_idx, tch.teacher_id, f"{gender_title} {tch.khmer_name}"] + p_vals + ['', '', '']
                        ws.append(row_data)

                        for col_idx in range(1, len(headers) + 1):
                            cell = ws.cell(row=current_row, column=col_idx)
                            cell.font = font_body
                            cell.border = thin_border
                            if col_idx in [1, 2, 4, 5, 6, 7]:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='left', vertical='center')

                        current_row += 1
                        no_idx += 1

                # Footer Signatures
                current_row += 2
                ws.cell(row=current_row, column=7, value="បានឃើញ និងឯកភាព").font = font_bold
                ws.cell(row=current_row + 1, column=7, value="នាយកសាលា").font = font_bold
                ws.cell(row=current_row, column=2, value="អ្នករៀបចំរបាយការណ៍").font = font_bold

                auto_fit_columns(ws, {1: 5, 2: 12, 3: 18, 4: 10, 5: 10, 6: 10, 7: 10, 8: 15, 9: 15, 10: 12})

    # ROUTING TO THE REQUESTED REPORT BUILDER
    if report_type == 'teacher_load':
        build_teacher_load_sheet(wb)
        filename = "teacher_teaching_load_report.xlsx"
    elif report_type == 'subject_codes':
        build_subject_codes_sheet(wb)
        filename = "teacher_subject_codes_directory.xlsx"
    elif report_type == 'class_summary':
        build_class_summary_sheet(wb)
        filename = "classrooms_timetable_summary.xlsx"
    elif report_type in ['all_reports', 'all']:
        build_teacher_load_sheet(wb)
        build_subject_codes_sheet(wb)
        build_class_summary_sheet(wb)
        build_duty_sheets(wb)
        filename = "all_timetable_reports.xlsx"
    else:
        # Default: duty_sheets
        build_duty_sheets(wb)
        filename = "teacher_daily_duty_sheets.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def timetable_export_excel(request):
    """
    Export Master Timetable Matrix to Excel-compatible CSV format with UTF-8 BOM for Khmer text.
    Strictly for the active academic year!
    """
    import csv
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="master_timetable.csv"'
    response.write('\ufeff') # UTF-8 BOM for Microsoft Excel Khmer font rendering

    writer = csv.writer(response)
    
    # Headers
    header_row_1 = ['ថ្នាក់រៀន']
    for d in DAYS_OF_WEEK:
        for p in PERIODS_LIST:
            header_row_1.append(f"{d['name_kh']} - ម៉ោង {p}")
    writer.writerow(header_row_1)

    classrooms = Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code') if active_year else Classroom.objects.all().order_by('grade_level', 'code')
    timetables = Timetable.objects.filter(classroom__academic_year=active_year).select_related('classroom', 'subject', 'teacher') if active_year else Timetable.objects.select_related('classroom', 'subject', 'teacher').all()

    for cls in classrooms:
        row = [cls.name]
        cls_entries = {(e.day_of_week, e.period_number): e for e in timetables.filter(classroom=cls)}
        for d in DAYS_OF_WEEK:
            for p in PERIODS_LIST:
                entry = cls_entries.get((d['num'], p))
                if entry:
                    row.append(f"{entry.subject.name_kh} ({entry.teacher.khmer_name})")
                else:
                    row.append("")
        writer.writerow(row)

    return response


@login_required
@role_required(['ADMIN'])
def timetable_clear_all(request):
    """
    Clear all timetable entries for the active academic year.
    """
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    if active_year:
        count = Timetable.objects.filter(classroom__academic_year=active_year).count()
        Timetable.objects.filter(classroom__academic_year=active_year).delete()
        msg = f"បានលុបទិន្នន័យកាលវិភាគទាំងអស់ ({count} ម៉ោង) នៃឆ្នាំសិក្សា {active_year.name} រួចរាល់!"
    else:
        count = Timetable.objects.count()
        Timetable.objects.all().delete()
        msg = f"បានលុបទិន្នន័យកាលវិភាគទាំងអស់ ({count} ម៉ោង) រួចរាល់!"

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success', 'message': msg})
    messages.success(request, msg)
    return redirect('timetable_view')


@login_required
@role_required(['ADMIN'])
def timetable_create(request):
    if request.method == 'POST':
        form = TimetableForm(request.POST)
        if form.is_valid():
            try:
                timetable = form.save()
                messages.success(request, f"បានបញ្ចូលកាលវិភាគ {timetable.subject.name_kh} ជោគជ័យ!")
                return redirect('timetable_view')
            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for err in errors:
                        messages.error(request, f"⚠️ បរាជ័យ: {err}")
        else:
            for field, errors in form.errors.items():
                for err in errors:
                    messages.error(request, f"⚠️ បញ្ហាទិន្នន័យ [{field}]: {err}")
    return redirect('timetable_view')


@login_required
@role_required(['ADMIN'])
def timetable_edit(request, pk):
    entry = get_object_or_404(Timetable, pk=pk)
    if request.method == 'POST':
        form = TimetableForm(request.POST, instance=entry)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f"បានកែប្រែម៉ោងបង្រៀន {entry.subject.name_kh} ជោគជ័យ!")
                return redirect('timetable_view')
            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for err in errors:
                        messages.error(request, f"⚠️ បរាជ័យ: {err}")
        else:
            for field, errors in form.errors.items():
                for err in errors:
                    messages.error(request, f"⚠️ កំហុស [{field}]: {err}")
    return redirect('timetable_view')


@login_required
@role_required(['ADMIN'])
def timetable_delete(request, pk):
    entry = get_object_or_404(Timetable, pk=pk)
    entry.delete()
    messages.success(request, "បានលុបម៉ោងបង្រៀនចេញពីកាលវិភាគជោគជ័យ!")
    return redirect('timetable_view')


@login_required
@role_required(['ADMIN'])
def timetable_clear_class(request, class_id):
    cls = get_object_or_404(Classroom, pk=class_id)
    count = Timetable.objects.filter(classroom=cls).count()
    Timetable.objects.filter(classroom=cls).delete()
    messages.success(request, f"បានលុបកាលវិភាគទាំងអស់ ({count} ម៉ោង) របស់ថ្នាក់ {cls.name} រួចរាល់!")
    return redirect('timetable_view')



# ----------------- SUBJECT REQUIREMENTS MATRIX -----------------

@login_required
@role_required(['ADMIN'])
def subject_requirements_manager(request):
    """
    Manage Subjects & Weekly Hours Requirements (Matrix matching user's design)
    Columns: Code, Subject Name, 7, 8, 9, 10, 11SC, 11SS, 12SC, 12SS, Actions
    Excludes R (Essay) and D (Dictation) since Khmer language (K) covers all Khmer teaching hours.
    """
    subjects = Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id')
    grade_levels = GradeLevel.objects.all().order_by('order', 'grade_number', 'track', 'id')
    
    if not grade_levels.exists():
        default_levels = [
            (7, 'GENERAL', 'ថ្នាក់ទី ៧', 1),
            (8, 'GENERAL', 'ថ្នាក់ទី ៨', 2),
            (9, 'GENERAL', 'ថ្នាក់ទី ៩', 3),
            (10, 'GENERAL', 'ថ្នាក់ទី ១០', 4),
            (11, 'SCIENCE', 'ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រ', 5),
            (11, 'SOCIAL', 'ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រសង្គម', 6),
            (12, 'SCIENCE', 'ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រ', 7),
            (12, 'SOCIAL', 'ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រសង្គម', 8),
        ]
        for g_num, trk, nm, ord_num in default_levels:
            GradeLevel.objects.get_or_create(grade_number=g_num, track=trk, defaults={'name': nm, 'order': ord_num})
        grade_levels = GradeLevel.objects.all().order_by('order', 'grade_number', 'track', 'id')

    streams_meta = []
    for gl in grade_levels:
        if gl.track == 'SCIENCE':
            lbl = f"{gl.grade_number}SC"
        elif gl.track == 'SOCIAL':
            lbl = f"{gl.grade_number}SS"
        else:
            lbl = f"{gl.grade_number}"
        streams_meta.append({
            'gl': gl,
            'label': lbl,
            'grade_number': gl.grade_number,
            'track': gl.track,
        })

    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('hours_'):
                parts = key.split('_')
                if len(parts) == 4:
                    _, g_num, trk, sub_id = parts
                    try:
                        hours_val = int(value) if value and value.isdigit() else 0
                        hours_val = max(0, hours_val)
                        rule, _ = GradeLevelRule.objects.get_or_create(
                            grade_level=int(g_num),
                            track=trk,
                            subject_id=int(sub_id),
                            defaults={'weekly_hours': hours_val}
                        )
                        if rule.weekly_hours != hours_val:
                            rule.weekly_hours = hours_val
                            rule.save(update_fields=['weekly_hours'])
                    except Exception:
                        pass
        messages.success(request, "បានរក្សាទុក និងធ្វើបច្ចុប្បន្នភាពម៉ោងសិក្សាជោគជ័យ!")
        return redirect('subject_requirements_manager')

    rules_dict = {}
    for r in GradeLevelRule.objects.all():
        rules_dict[(r.subject_id, r.grade_level, r.track)] = r.weekly_hours

    matrix_rows = []
    stream_totals = [0] * len(streams_meta)
    
    for sub in subjects:
        cells = []
        total_hours_for_sub = 0
        for idx, sm in enumerate(streams_meta):
            hrs = rules_dict.get((sub.id, sm['grade_number'], sm['track']), 0)
            total_hours_for_sub += hrs
            stream_totals[idx] += hrs
            cells.append({
                'grade_number': sm['grade_number'],
                'track': sm['track'],
                'hours': hrs,
                'input_name': f"hours_{sm['grade_number']}_{sm['track']}_{sub.id}",
            })
        matrix_rows.append({
            'subject': sub,
            'cells': cells,
            'total_hours': total_hours_for_sub,
        })

    has_custom_default = SavedDefaultConfig.objects.filter(key='custom_subject_requirements').exists()

    return render(request, 'academics/subject_requirements.html', {
        'streams_meta': streams_meta,
        'matrix_rows': matrix_rows,
        'stream_totals': stream_totals,
        'subjects_count': subjects.count(),
        'has_custom_default': has_custom_default,
    })


@login_required
@role_required(['ADMIN'])
def subject_requirements_reset(request):
    """Reset all weekly teaching hours to 0."""
    GradeLevelRule.objects.all().update(weekly_hours=0)
    messages.success(request, "បានសម្អាតទិន្នន័យម៉ោងសិក្សាទាំងអស់ទៅជា 0 ជោគជ័យ!")
    return redirect('subject_requirements_manager')


@login_required
@role_required(['ADMIN'])
def subject_requirements_restore_moeys(request):
    """
    Restore official Ministry of Education (MoEYS) standard weekly teaching hours per grade level.
    """
    moeys_hours_matrix = {
        # ថ្នាក់ទី ៧ (Grade 7 - Junior High)
        (7, 'GENERAL'): {
            'K': 5, 'M': 5, 'P': 2, 'C': 1, 'B': 2, 'Es': 1,
            'I': 2, 'G': 2, 'H': 2, 'He': 2, 'Ec': 0, 'E': 2, 'Ed': 2, 'Ag': 2, 'IT': 2,
        },
        # ថ្នាក់ទី ៨ (Grade 8 - Junior High)
        (8, 'GENERAL'): {
            'K': 5, 'M': 5, 'P': 2, 'C': 2, 'B': 2, 'Es': 1,
            'I': 2, 'G': 2, 'H': 2, 'He': 2, 'Ec': 0, 'E': 2, 'Ed': 2, 'Ag': 2, 'IT': 2,
        },
        # ថ្នាក់ទី ៩ (Grade 9 - Junior High Diploma)
        (9, 'GENERAL'): {
            'K': 5, 'M': 5, 'P': 2, 'C': 2, 'B': 2, 'Es': 1,
            'I': 2, 'G': 2, 'H': 2, 'He': 2, 'Ec': 0, 'E': 2, 'Ed': 2, 'Ag': 2, 'IT': 2,
        },
        # ថ្នាក់ទី ១០ (Grade 10 - High School Foundation)
        (10, 'GENERAL'): {
            'K': 5, 'M': 5, 'P': 3, 'C': 3, 'B': 3, 'Es': 2,
            'I': 2, 'G': 2, 'H': 2, 'He': 2, 'Ec': 2, 'E': 3, 'Ed': 2, 'Ag': 0, 'IT': 2,
        },
        # ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រ (Grade 11 - Science Stream)
        (11, 'SCIENCE'): {
            'K': 4, 'M': 6, 'P': 4, 'C': 4, 'B': 4, 'Es': 2,
            'I': 2, 'G': 2, 'H': 2, 'He': 0, 'Ec': 2, 'E': 2, 'Ed': 2, 'Ag': 0, 'IT': 2,
        },
        # ថ្នាក់ទី ១១ វិទ្យាសាស្ត្រសង្គម (Grade 11 - Social Science Stream)
        (11, 'SOCIAL'): {
            'K': 6, 'M': 4, 'P': 2, 'C': 1, 'B': 2, 'Es': 2,
            'I': 4, 'G': 4, 'H': 4, 'He': 0, 'Ec': 3, 'E': 2, 'Ed': 2, 'Ag': 0, 'IT': 2,
        },
        # ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រ (Grade 12 - Science Stream BacII)
        (12, 'SCIENCE'): {
            'K': 4, 'M': 6, 'P': 4, 'C': 4, 'B': 4, 'Es': 2,
            'I': 2, 'G': 2, 'H': 2, 'He': 0, 'Ec': 2, 'E': 2, 'Ed': 2, 'Ag': 0, 'IT': 2,
        },
        # ថ្នាក់ទី ១២ វិទ្យាសាស្ត្រសង្គម (Grade 12 - Social Science Stream BacII)
        (12, 'SOCIAL'): {
            'K': 6, 'M': 4, 'P': 2, 'C': 1, 'B': 2, 'Es': 2,
            'I': 4, 'G': 4, 'H': 4, 'He': 0, 'Ec': 3, 'E': 2, 'Ed': 2, 'Ag': 0, 'IT': 2,
        },
    }

    count = 0
    with transaction.atomic():
        for (g_num, trk), sub_map in moeys_hours_matrix.items():
            for sub_code, hrs in sub_map.items():
                sub = Subject.objects.filter(code=sub_code).first()
                if sub:
                    rule, _ = GradeLevelRule.objects.get_or_create(
                        grade_level=g_num,
                        track=trk,
                        subject=sub,
                        defaults={'weekly_hours': hrs}
                    )
                    if rule.weekly_hours != hrs:
                        rule.weekly_hours = hrs
                        rule.save(update_fields=['weekly_hours'])
                    count += 1

    messages.success(request, f"បានទាញយក និងកំណត់ម៉ោងសិក្សាតាមស្តង់ដារផ្លូវការក្រសួងអប់រំ (MoEYS) ជោគជ័យ!")
    return redirect('subject_requirements_manager')


@login_required
@role_required(['ADMIN'])
def subject_requirements_save_custom_default(request):
    """
    Save the current weekly hours configuration as the institution's custom default preset.
    """
    rules = GradeLevelRule.objects.filter(weekly_hours__gt=0).select_related('subject')
    saved_dict = {}
    for r in rules:
        saved_dict[f"{r.grade_level}_{r.track}_{r.subject.code}"] = r.weekly_hours

    SavedDefaultConfig.objects.update_or_create(
        key='custom_subject_requirements',
        defaults={'data': saved_dict}
    )
    messages.success(request, "បានរក្សាទុកទម្រង់ម៉ោងសិក្សាបច្ចុប្បន្នជា «លំនាំដើមផ្ទាល់ខ្លួន» របស់សាលាជោគជ័យ!")
    return redirect('subject_requirements_manager')


@login_required
@role_required(['ADMIN'])
def subject_requirements_restore_custom_default(request):
    """
    Restore the institution's custom default preset.
    """
    preset = SavedDefaultConfig.objects.filter(key='custom_subject_requirements').first()
    if not preset or not preset.data:
        messages.warning(request, "មិនទាន់មានទម្រង់លំនាំដើមផ្ទាល់ខ្លួនដែលបានរក្សាទុកនៅឡើយទេ!")
        return redirect('subject_requirements_manager')

    with transaction.atomic():
        GradeLevelRule.objects.all().update(weekly_hours=0)
        for k, hrs in preset.data.items():
            parts = k.split('_')
            if len(parts) == 3:
                g_num, trk, sub_code = parts
                sub = Subject.objects.filter(code=sub_code).first()
                if sub:
                    rule, _ = GradeLevelRule.objects.get_or_create(
                        grade_level=int(g_num),
                        track=trk,
                        subject=sub,
                        defaults={'weekly_hours': int(hrs)}
                    )
                    if rule.weekly_hours != int(hrs):
                        rule.weekly_hours = int(hrs)
                        rule.save(update_fields=['weekly_hours'])

    messages.success(request, "បានទាញយកទម្រង់ម៉ោងសិក្សាលំនាំដើមផ្ទាល់ខ្លួនរបស់សាលាជោគជ័យ!")
    return redirect('subject_requirements_manager')


@login_required
@role_required(['ADMIN'])
def subject_requirement_row_delete(request, subject_id):
    sub = get_object_or_404(Subject, pk=subject_id)
    GradeLevelRule.objects.filter(subject=sub).update(weekly_hours=0)
    messages.success(request, f"បានលុបម៉ោងសិក្សាសម្រាប់មុខវិជ្ជា {sub.name_kh} ជោគជ័យ!")
    return redirect('subject_requirements_manager')


# ----------------- TEACHER CLASS & SUBJECT ASSIGNMENTS -----------------

DEFAULT_TRAINING_LEVEL_QUOTAS = {
    'គ្រូទុតិយភូមិ': 16,
    'គ្រូបឋមភូមិ': 18,
    'គ្រូកម្រិតបឋម': 18,
    'default': 18,
}

def get_training_level_quotas():
    try:
        config = SavedDefaultConfig.objects.filter(key='training_level_quotas').first()
        if config and config.data and isinstance(config.data, dict):
            res = dict(DEFAULT_TRAINING_LEVEL_QUOTAS)
            res.update(config.data)
            return res
    except Exception:
        pass
    return dict(DEFAULT_TRAINING_LEVEL_QUOTAS)


@login_required
@role_required(['ADMIN'])
def teacher_assignments_manager(request):
    """
    Teacher Class & Subject Assignments Manager.
    Admin can select any teacher and tick multiple classrooms and multiple subjects assigned to that teacher.
    Strictly isolated per Academic Year!
    """
    try:
        from .utils import get_active_academic_year
        active_year = get_active_academic_year(request)
        selected_year = request.GET.get('year') or request.GET.get('academic_year')
        if selected_year:
            if str(selected_year).strip().isdigit():
                found_year = AcademicYear.objects.filter(id=int(str(selected_year).strip())).first()
            else:
                found_year = AcademicYear.objects.filter(name=str(selected_year).strip()).first()
            if found_year:
                active_year = found_year

        teachers = list(Teacher.objects.filter(status='ACTIVE').order_by('khmer_name'))
        classrooms = list(Classroom.objects.filter(academic_year=active_year).order_by('grade_level', 'code')) if active_year else list(Classroom.objects.all().order_by('grade_level', 'code'))
        subjects = list(Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id'))

        selected_teacher_id = request.GET.get('teacher')
        selected_teacher = None
        if selected_teacher_id:
            s_tid = str(selected_teacher_id).strip()
            if s_tid.isdigit():
                selected_teacher = Teacher.objects.filter(id=int(s_tid)).first()
            else:
                selected_teacher = Teacher.objects.filter(khmer_name=s_tid).first()
        if not selected_teacher and teachers:
            selected_teacher = teachers[0]

        if request.method == 'POST' and selected_teacher:
            # Handle Max Hours update
            max_h_str = request.POST.get('max_weekly_hours', '').strip()
            apply_all = request.POST.get('apply_to_all_teachers') in ['true', 'on', '1']
            if max_h_str and max_h_str.isdigit():
                val = int(max_h_str)
                if apply_all:
                    Teacher.objects.all().update(max_weekly_hours=val)
                    selected_teacher.max_weekly_hours = val
                else:
                    selected_teacher.max_weekly_hours = val
                    selected_teacher.save(update_fields=['max_weekly_hours'])

            checked_pairs = set()
            for key in request.POST.keys():
                if key.startswith('assign_'):
                    parts = key.split('_')
                    if len(parts) == 3 and parts[1].isdigit() and parts[2].isdigit():
                        cls_id, sub_id = int(parts[1]), int(parts[2])
                        checked_pairs.add((cls_id, sub_id))

            assigned_count = 0
            transferred_changes = []
            removed_changes = []
            new_additions = []
            total_tt_synced = 0

            with transaction.atomic():
                # 1. Unassign unchecked pairs previously belonging to this teacher in this academic year
                existing_assigned = list(
                    ClassSubject.objects.filter(teacher=selected_teacher, classroom__academic_year=active_year).select_related('classroom', 'subject')
                    if active_year else ClassSubject.objects.filter(teacher=selected_teacher).select_related('classroom', 'subject')
                )
                unassign_cs_list = [cs for cs in existing_assigned if (cs.classroom_id, cs.subject_id) not in checked_pairs]
                
                for cs in unassign_cs_list:
                    tt_qs = Timetable.objects.filter(classroom_id=cs.classroom_id, subject_id=cs.subject_id, teacher=selected_teacher)
                    if active_year:
                        tt_qs = tt_qs.filter(classroom__academic_year=active_year)
                    tt_count = tt_qs.count()
                    if tt_count > 0:
                        tt_qs.update(teacher=None)
                        total_tt_synced += tt_count
                    removed_changes.append({
                        'classroom_name': cs.classroom.name if cs.classroom else f'ថ្នាក់ #{cs.classroom_id}',
                        'subject_name': cs.subject.name_kh if cs.subject else f'មុខវិជ្ជា #{cs.subject_id}',
                        'tt_count': tt_count,
                    })

                if unassign_cs_list:
                    ClassSubject.objects.filter(id__in=[cs.id for cs in unassign_cs_list]).update(teacher=None)

                # 2. Assign checked pairs to this teacher & detect transfers from other teachers
                for cls_id, sub_id in checked_pairs:
                    cs = ClassSubject.objects.filter(classroom_id=cls_id, subject_id=sub_id).select_related('classroom', 'subject', 'teacher').first()
                    if not cs:
                        cls_obj = Classroom.objects.filter(id=cls_id).first()
                        sub_obj = Subject.objects.filter(id=sub_id).first()
                        cs = ClassSubject.objects.create(
                            classroom_id=cls_id,
                            subject_id=sub_id,
                            teacher=selected_teacher,
                        )
                        tt_qs = Timetable.objects.filter(classroom_id=cls_id, subject_id=sub_id)
                        if active_year:
                            tt_qs = tt_qs.filter(classroom__academic_year=active_year)
                        tt_synced = tt_qs.exclude(teacher=selected_teacher).update(teacher=selected_teacher)
                        total_tt_synced += tt_synced
                        new_additions.append({
                            'classroom_name': cls_obj.name if cls_obj else f'ថ្នាក់ #{cls_id}',
                            'subject_name': sub_obj.name_kh if sub_obj else f'មុខវិជ្ជា #{sub_id}',
                            'tt_updated': tt_synced,
                        })
                    else:
                        old_teacher = cs.teacher
                        old_teacher_id = getattr(old_teacher, 'id', None)
                        if old_teacher_id and old_teacher_id != selected_teacher.id:
                            # Transferred from another teacher!
                            cs.teacher = selected_teacher
                            cs.save(update_fields=['teacher'])
                            tt_qs = Timetable.objects.filter(classroom_id=cls_id, subject_id=sub_id)
                            if active_year:
                                tt_qs = tt_qs.filter(classroom__academic_year=active_year)
                            tt_synced = tt_qs.exclude(teacher=selected_teacher).update(teacher=selected_teacher)
                            total_tt_synced += tt_synced
                            transferred_changes.append({
                                'classroom_name': cs.classroom.name if cs.classroom else f'ថ្នាក់ #{cls_id}',
                                'subject_name': cs.subject.name_kh if cs.subject else f'មុខវិជ្ជា #{sub_id}',
                                'old_teacher_name': old_teacher.khmer_name or old_teacher.latin_name or 'គ្រូផ្សេង',
                                'tt_updated': tt_synced,
                            })
                        else:
                            if cs.teacher_id != selected_teacher.id:
                                cs.teacher = selected_teacher
                                cs.save(update_fields=['teacher'])
                                tt_qs = Timetable.objects.filter(classroom_id=cls_id, subject_id=sub_id)
                                if active_year:
                                    tt_qs = tt_qs.filter(classroom__academic_year=active_year)
                                tt_synced = tt_qs.exclude(teacher=selected_teacher).update(teacher=selected_teacher)
                                total_tt_synced += tt_synced
                                new_additions.append({
                                    'classroom_name': cs.classroom.name if cs.classroom else f'ថ្នាក់ #{cls_id}',
                                    'subject_name': cs.subject.name_kh if cs.subject else f'មុខវិជ្ជា #{sub_id}',
                                    'tt_updated': tt_synced,
                                })
                    assigned_count += 1

                # 3. Check for timetable slot clashes/conflicts for selected_teacher in Master Timetable
                clash_details = []
                tt_all = list(
                    Timetable.objects.filter(teacher=selected_teacher, classroom__academic_year=active_year).select_related('classroom', 'subject')
                    if active_year else Timetable.objects.filter(teacher=selected_teacher).select_related('classroom', 'subject')
                )
                day_period_map = defaultdict(list)
                for entry in tt_all:
                    day_period_map[(entry.day_of_week, entry.period_number)].append(entry)
                
                kh_days_map = {1: 'ចន្ទ', 2: 'អង្គារ', 3: 'ពុធ', 4: 'ព្រហស្បតិ៍', 5: 'សុក្រ', 6: 'សៅរ៍', 7: 'អាទិត្យ'}
                for (d_num, p_num), entries_in_slot in day_period_map.items():
                    if len(entries_in_slot) > 1:
                        classes_str = ", ".join([f"{e.classroom.name} ({e.subject.name_kh})" for e in entries_in_slot])
                        clash_details.append(f"ថ្ងៃ {kh_days_map.get(d_num, d_num)} ម៉ោងទី {p_num} ៖ ជាន់គ្នារវាង {classes_str}")

            # Compose informative notification messages
            msg_parts = [f"✅ បានរក្សាទុកការចាត់តាំងមុខវិជ្ជា និងថ្នាក់បង្រៀន ({assigned_count} ថ្នាក់-មុខវិជ្ជា) សម្រាប់គ្រូ <strong>{selected_teacher.khmer_name}</strong> ជោគជ័យ!"]
            if total_tt_synced > 0:
                msg_parts.append(f"🔄 <strong>ធ្វើបច្ចុប្បន្នភាពកាលវិភាគរួម (Master Timetable)៖</strong> បាន Sync ម៉ោងបង្រៀនចំនួន <strong>{total_tt_synced} ម៉ោង</strong> ក្នុងកាលវិភាគរួមស្វ័យប្រវត្តិ។")
            if transferred_changes:
                transfers_str = "<br>• " + "<br>• ".join([
                    f"ផ្ទេរថ្នាក់ <strong>{t['classroom_name']}</strong> ({t['subject_name']}) ពី <strong>{t['old_teacher_name']}</strong> មកកាន់ <strong>{selected_teacher.khmer_name}</strong>" + (f" (Sync កាលវិភាគ {t['tt_updated']} ម៉ោង)" if t['tt_updated'] > 0 else "")
                    for t in transferred_changes
                ])
                msg_parts.append(f"👥 <strong>បម្រែបម្រួលផ្ទេរគ្រូបង្រៀន ({len(transferred_changes)} មុខវិជ្ជា)៖</strong>{transfers_str}")
            if removed_changes:
                removals_str = "<br>• " + "<br>• ".join([
                    f"ដកចេញពីថ្នាក់ <strong>{r['classroom_name']}</strong> ({r['subject_name']})" + (f" (ម៉ោងក្នុងកាលវិភាគ {r['tt_count']} ម៉ោងត្រូវទំនេរ)" if r['tt_count'] > 0 else "")
                    for r in removed_changes
                ])
                msg_parts.append(f"🗑️ <strong>មុខវិជ្ជាដែលបានដកចេញ ({len(removed_changes)} មុខវិជ្ជា)៖</strong>{removals_str}")

            messages.success(request, "<br><br>".join(msg_parts))

            if clash_details:
                clashes_str = "<br>• " + "<br>• ".join(clash_details)
                messages.warning(request, f"⚠️ <strong>ការព្រមានជាន់ម៉ោងក្នុងកាលវិភាគរួម (Timetable Conflicts)!</strong><br>លោកគ្រូ/អ្នកគ្រូ {selected_teacher.khmer_name} មានម៉ោងជាន់គ្នា៖{clashes_str}<br><em>សូមចូលទៅកាន់ «កាលវិភាគរួម (Master Timetable)» ដើម្បីសម្រួលរៀបចំម៉ោងឡើងវិញ!</em>")

            return redirect(f"/academics/teacher-assignments/?teacher={selected_teacher.id}{f'&year={active_year.id}' if active_year else ''}")

        # Build matrix for display with zero N+1 queries
        selected_teacher_pairs = set()
        if selected_teacher:
            selected_teacher_pairs = set(
                ClassSubject.objects.filter(teacher=selected_teacher, classroom__academic_year=active_year).values_list('classroom_id', 'subject_id') if active_year else ClassSubject.objects.filter(teacher=selected_teacher).values_list('classroom_id', 'subject_id')
            )

        all_assignments = {}
        cs_query = ClassSubject.objects.filter(
            classroom__academic_year=active_year, 
            teacher__isnull=False
        ).select_related('teacher', 'classroom', 'subject') if active_year else ClassSubject.objects.filter(
            teacher__isnull=False
        ).select_related('teacher', 'classroom', 'subject')
        
        teacher_assigned_map = defaultdict(list)
        for cs in cs_query:
            all_assignments[(cs.classroom_id, cs.subject_id)] = cs.teacher
            if cs.teacher_id:
                teacher_assigned_map[cs.teacher_id].append(cs)

        all_rules = list(GradeLevelRule.objects.all())
        rules_dict = {}
        rules_by_grade_track = defaultdict(set)
        for r in all_rules:
            rules_dict[(r.subject_id, r.grade_level, r.track)] = r.weekly_hours
            if r.weekly_hours and r.weekly_hours > 0:
                rules_by_grade_track[(r.grade_level, r.track)].add(r.subject_id)

        cls_assigned_subs_map = defaultdict(set)
        if classrooms:
            for cls_id, sub_id in ClassSubject.objects.filter(classroom__in=classrooms).values_list('classroom_id', 'subject_id'):
                cls_assigned_subs_map[cls_id].add(sub_id)

        teacher_stats = []
        for t in teachers:
            assigned_cs = teacher_assigned_map.get(t.id, [])
            t_hours = 0
            for cs in assigned_cs:
                cls_grade = cs.classroom.grade_level if cs.classroom else 10
                cls_track = cs.classroom.track if cs.classroom else 'GENERAL'
                h = rules_dict.get((cs.subject_id, cls_grade, cls_track))
                if h is None:
                    h = rules_dict.get((cs.subject_id, cls_grade, 'GENERAL'), 0)
                t_hours += (h or 0)

            t_max = t.max_weekly_hours or 18
            teacher_stats.append({
                'teacher': t,
                'assigned_count': len(assigned_cs),
                'assigned_hours': t_hours,
                'max_weekly_hours': t_max,
                'is_selected': bool(selected_teacher and t.id == selected_teacher.id),
                'is_over': t_hours > t_max,
            })

        timetables_active = list(
            Timetable.objects.filter(classroom__academic_year=active_year).select_related('teacher').only('id', 'classroom_id', 'subject_id', 'teacher__id', 'teacher__khmer_name')
            if active_year else Timetable.objects.select_related('teacher').only('id', 'classroom_id', 'subject_id', 'teacher__id', 'teacher__khmer_name').all()
        )
        tt_slot_count_map = defaultdict(int)
        tt_teacher_map = {}
        for t_entry in timetables_active:
            tt_slot_count_map[(t_entry.classroom_id, t_entry.subject_id)] += 1
            if t_entry.teacher:
                tt_teacher_map[(t_entry.classroom_id, t_entry.subject_id)] = t_entry.teacher.khmer_name

        matrix_grid = []
        selected_subject_hours = {sub.id: 0 for sub in subjects}
        selected_total_assigned_hours = 0

        for cls in classrooms:
            cells = []
            cls_assigned_subs = cls_assigned_subs_map.get(cls.id)
            if not cls_assigned_subs:
                cls_grade = cls.grade_level if cls.grade_level is not None else 10
                cls_track = cls.track or 'GENERAL'
                cls_assigned_subs = rules_by_grade_track.get((cls_grade, cls_track)) or rules_by_grade_track.get((cls_grade, 'GENERAL')) or set()

            for sub in subjects:
                is_checked = (cls.id, sub.id) in selected_teacher_pairs
                other_teacher = all_assignments.get((cls.id, sub.id))
                is_valid_for_class = sub.id in cls_assigned_subs
                scheduled_slots_count = tt_slot_count_map.get((cls.id, sub.id), 0)
                scheduled_teacher_name = tt_teacher_map.get((cls.id, sub.id))

                cls_grade = cls.grade_level if cls.grade_level is not None else 10
                cls_track = cls.track or 'GENERAL'
                h_req = rules_dict.get((sub.id, cls_grade, cls_track))
                if h_req is None:
                    h_req = rules_dict.get((sub.id, cls_grade, 'GENERAL'), 0)
                h_req = h_req or 0

                if is_checked:
                    selected_subject_hours[sub.id] += h_req
                    selected_total_assigned_hours += h_req

                cells.append({
                    'subject': sub,
                    'classroom': cls,
                    'input_name': f"assign_{cls.id}_{sub.id}",
                    'is_checked': is_checked,
                    'hours_required': h_req,
                    'other_teacher': other_teacher if (other_teacher and other_teacher != selected_teacher) else None,
                    'is_valid_for_class': is_valid_for_class,
                    'scheduled_slots_count': scheduled_slots_count,
                    'scheduled_teacher_name': scheduled_teacher_name,
                })
            matrix_grid.append({
                'classroom': cls,
                'cells': cells,
            })

        # Build teacher-subject code mapping (e.g. K1, K2, M1...) using in-memory cs_query
        distinct_assignments = []
        seen_pairs = set()
        for cs in cs_query:
            if cs.subject_id and cs.teacher_id:
                sub = next((s for s in subjects if s.id == cs.subject_id), None)
                if sub and sub.code in ['R', 'D']:
                    continue
                pair = (cs.subject_id, cs.teacher_id)
                if pair not in seen_pairs:
                    seen_pairs.add(pair)
                    distinct_assignments.append({'subject_id': cs.subject_id, 'teacher_id': cs.teacher_id})
        distinct_assignments.sort(key=lambda x: (x['subject_id'], x['teacher_id']))

        teacher_subject_code_map = {}
        subject_teacher_counters = {}
        for item in distinct_assignments:
            s_id = item['subject_id']
            t_id = item['teacher_id']
            sub = next((s for s in subjects if s.id == s_id), None)
            sub_code = sub.code if (sub and sub.code) else 'S'
            if s_id not in subject_teacher_counters:
                subject_teacher_counters[s_id] = 1
            else:
                subject_teacher_counters[s_id] += 1
            teacher_subject_code_map[(s_id, t_id)] = f"{sub_code}{subject_teacher_counters[s_id]}"

        selected_teacher_codes = []
        if selected_teacher:
            for s in subjects:
                if (s.id, selected_teacher.id) in teacher_subject_code_map:
                    selected_teacher_codes.append(teacher_subject_code_map[(s.id, selected_teacher.id)])

        # Retrieve dynamic training level quotas for modal customization from in-memory teachers
        training_quotas = get_training_level_quotas()
        raw_levels = [t.training_level for t in teachers if t.training_level]
        distinct_levels = sorted(list(set([str(lvl).strip() for lvl in raw_levels if lvl and str(lvl).strip()])))
        if 'គ្រូទុតិយភូមិ' not in distinct_levels:
            distinct_levels.insert(0, 'គ្រូទុតិយភូមិ')
        if 'គ្រូបឋមភូមិ' not in distinct_levels:
            distinct_levels.append('គ្រូបឋមភូមិ')

        training_level_settings = []
        teacher_levels_count = defaultdict(int)
        for t in teachers:
            if t.training_level:
                teacher_levels_count[t.training_level] += 1

        for lvl in distinct_levels:
            training_level_settings.append({
                'name': lvl,
                'hours': training_quotas.get(lvl, 16 if 'ទុតិយភូមិ' in lvl else 18),
                'count': teacher_levels_count.get(lvl, 0),
            })

        return render(request, 'academics/teacher_assignments.html', {
            'teachers': teachers,
            'teacher_stats': teacher_stats,
            'selected_teacher': selected_teacher,
            'selected_teacher_codes': ", ".join(selected_teacher_codes) or None,
            'classrooms': classrooms,
            'subjects': subjects,
            'matrix_grid': matrix_grid,
            'selected_assigned_count': len(selected_teacher_pairs),
            'selected_assigned_hours': selected_total_assigned_hours,
            'selected_subject_hours': selected_subject_hours,
            'selected_max_hours': selected_teacher.max_weekly_hours if (selected_teacher and selected_teacher.max_weekly_hours) else 18,
            'training_level_settings': training_level_settings,
            'training_quotas': training_quotas,
        })
    except Exception as e:
        messages.error(request, f"កំហុសក្នុងការទាញយកទិន្នន័យចាត់តាំងគ្រូ៖ {str(e)}")
        try:
            return render(request, 'academics/teacher_assignments.html', {
                'teachers': list(Teacher.objects.all().order_by('khmer_name')),
                'teacher_stats': [],
                'selected_teacher': None,
                'selected_teacher_codes': None,
                'classrooms': [],
                'subjects': list(Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id')),
                'matrix_grid': [],
                'selected_assigned_count': 0,
                'selected_assigned_hours': 0,
                'selected_subject_hours': {},
                'selected_max_hours': 18,
                'training_level_settings': [],
                'training_quotas': DEFAULT_TRAINING_LEVEL_QUOTAS,
            })
        except Exception as inner_e:
            return HttpResponse(
                f"<div style='padding:30px;font-family:sans-serif;'><h3>កំហុសប្រព័ន្ធ (System Error)</h3><p>{str(e)}</p><p><small>{str(inner_e)}</small></p><a href='/academics/timetable/'>ត្រឡប់ទៅកាន់កាលវិភាគ</a></div>",
                status=200,
                content_type="text/html; charset=utf-8"
            )


@login_required
@role_required(['ADMIN'])
def teacher_assignments_training_quotas_save(request):
    """
    Save custom training level teaching hour quotas (e.g. គ្រូទុតិយភូមិ=16h, គ្រូបឋមភូមិ=18h)
    and sync/apply them to all active teachers in the database.
    """
    if request.method == 'POST':
        quotas = get_training_level_quotas()
        for key in request.POST.keys():
            if key.startswith('quota_'):
                level_name = key[6:].strip()
                val_str = request.POST.get(key, '').strip()
                if val_str.isdigit():
                    quotas[level_name] = max(1, min(60, int(val_str)))

        # Handle adding new training level dynamically
        new_name = request.POST.get('new_level_name', '').strip()
        new_hours_str = request.POST.get('new_level_hours', '').strip()
        if new_name and new_hours_str.isdigit():
            quotas[new_name] = max(1, min(60, int(new_hours_str)))

        SavedDefaultConfig.objects.update_or_create(
            key='training_level_quotas',
            defaults={'data': quotas}
        )

        # Apply to all teachers in bulk
        teachers_to_update = []
        teachers = Teacher.objects.all()
        for t in teachers:
            t_level = (t.training_level or '').strip()
            if t_level in quotas:
                new_max = quotas[t_level]
            elif 'ទុតិយភូមិ' in t_level:
                new_max = quotas.get('គ្រូទុតិយភូមិ', 16)
            elif 'បឋមភូមិ' in t_level:
                new_max = quotas.get('គ្រូបឋមភូមិ', 18)
            else:
                new_max = quotas.get('default', 18)

            if t.max_weekly_hours != new_max:
                t.max_weekly_hours = new_max
                teachers_to_update.append(t)

        if teachers_to_update:
            try:
                Teacher.objects.bulk_update(teachers_to_update, ['max_weekly_hours'])
            except Exception:
                for t in teachers_to_update:
                    try:
                        t.save(update_fields=['max_weekly_hours'])
                    except Exception:
                        pass

        messages.success(request, f"បានរក្សាទុក និងកំណត់កូតាម៉ោងបង្រៀន (គ្រូទុតិយភូមិ = {quotas.get('គ្រូទុតិយភូមិ', 16)} ម៉ោង, ផ្សេងៗ = {quotas.get('default', 18)} ម៉ោង) ទៅគ្រូទាំងអស់ជោគជ័យ!")
    
    return redirect('teacher_assignments_manager')


@login_required
@role_required(['ADMIN'])
def teacher_assignments_reset_teacher(request, teacher_id):
    """Reset/Clear assignments for a single teacher."""
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    
    cs_query = ClassSubject.objects.filter(teacher=teacher)
    if active_year:
        cs_query = cs_query.filter(classroom__academic_year=active_year)
    
    count = cs_query.count()
    cs_query.update(teacher=None)
    messages.success(request, f"បានសម្អាតការចាត់តាំងថ្នាក់ ({count} ថ្នាក់-មុខវិជ្ជា) សម្រាប់គ្រូ {teacher.khmer_name} ជោគជ័យ!")
    return redirect(f"/academics/teacher-assignments/?teacher={teacher.id}{f'&year={active_year.id}' if active_year else ''}")


@login_required
@role_required(['ADMIN'])
def teacher_assignments_reset_all(request):
    """Reset/Clear all teacher class assignments for the active academic year."""
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    
    cs_query = ClassSubject.objects.filter(teacher__isnull=False)
    if active_year:
        cs_query = cs_query.filter(classroom__academic_year=active_year)
    
    count = cs_query.count()
    cs_query.update(teacher=None)
    messages.success(request, f"បានសម្អាតការចាត់តាំងគ្រូទាំងអស់ទូទាំងសាលា ({count} ការចាត់តាំង) ជោគជ័យ! លោកអ្នកអាចចាប់ផ្តើមចាត់តាំងថ្មីបាន។")
    return redirect(f"/academics/teacher-assignments/{f'?year={active_year.id}' if active_year else ''}")


@login_required
@role_required(['ADMIN'])
def teacher_assignments_auto_assign(request):
    """
    Intelligently auto-assign active teachers to classes based on their subject specialization
    and pedagogical training level quota (គ្រូទុតិយភូមិ = 16h, ផ្សេងៗ = 18h).
    """
    active_year = None
    try:
        from .utils import get_active_academic_year
        active_year = get_active_academic_year(request)
        
        quotas = get_training_level_quotas()
        teachers = list(Teacher.objects.filter(status='ACTIVE').order_by('id'))
        
        if not teachers:
            messages.warning(request, "មិនទាន់មានគ្រូបង្រៀនសកម្ម (ACTIVE) ក្នុងប្រព័ន្ធដើម្បីចាត់តាំងស្វ័យប្រវត្តិឡើយ!")
            return redirect(f"/academics/teacher-assignments/{f'?year={active_year.id}' if active_year else ''}")

        # 1. Sync / Update teacher max hours based on training level in bulk
        teachers_to_update = []
        for t in teachers:
            t_level = (t.training_level or '').strip()
            if t_level in quotas:
                quota_val = quotas[t_level]
            elif 'ទុតិយភូមិ' in t_level:
                quota_val = quotas.get('គ្រូទុតិយភូមិ', 16)
            elif 'បឋមភូមិ' in t_level:
                quota_val = quotas.get('គ្រូបឋមភូមិ', 18)
            else:
                quota_val = quotas.get('default', 18)
            
            if not t.max_weekly_hours or t.max_weekly_hours in [16, 18]:
                if t.max_weekly_hours != quota_val:
                    t.max_weekly_hours = quota_val
                    teachers_to_update.append(t)

        if teachers_to_update:
            try:
                Teacher.objects.bulk_update(teachers_to_update, ['max_weekly_hours'])
            except Exception:
                for t in teachers_to_update:
                    try:
                        t.save(update_fields=['max_weekly_hours'])
                    except Exception:
                        pass

        # Classrooms ordered by High School (12, 11, 10) down to Middle School (9, 8, 7)
        classrooms = list(Classroom.objects.filter(academic_year=active_year).order_by('-grade_level', 'code') if active_year else Classroom.objects.all().order_by('-grade_level', 'code'))
        subjects = list(Subject.objects.exclude(code__in=['R', 'D']).order_by('order', 'id'))
        
        rules_dict = {}
        for r in GradeLevelRule.objects.all():
            rules_dict[(r.subject_id, r.grade_level, r.track)] = r.weekly_hours
        
        teacher_loads = {t.id: 0 for t in teachers}
        teacher_max = {t.id: (t.max_weekly_hours or 18) for t in teachers}
        assigned_count = 0

        # Pre-fetch existing ClassSubject mapping to avoid N+1 DB calls
        existing_cs_map = {}
        if classrooms:
            for cs in ClassSubject.objects.filter(classroom__in=classrooms):
                existing_cs_map[(cs.classroom_id, cs.subject_id)] = cs

        cs_to_update = []
        cs_to_create = []

        for cls in classrooms:
            cls_grade = cls.grade_level if cls.grade_level is not None else 10
            cls_track = cls.track or 'GENERAL'
            is_high_school = cls_grade >= 10
            
            for sub in subjects:
                h_req = rules_dict.get((sub.id, cls_grade, cls_track))
                if h_req is None:
                    h_req = rules_dict.get((sub.id, cls_grade, 'GENERAL'), 0)
                
                if not h_req or h_req <= 0:
                    continue

                sub_kh = (sub.name_kh or '').strip()
                sub_en = (sub.name_en or '').strip().lower()

                # Find candidate teachers specializing in this subject
                candidates = []
                for t in teachers:
                    spec = (t.specialization or '').strip().lower()
                    if sub_kh and sub_kh.lower() in spec:
                        candidates.append(t)
                    elif sub_en and sub_en in spec:
                        candidates.append(t)
                
                if not candidates:
                    candidates = list(teachers)  # Fallback to any teacher

                # Score candidates: Prioritize High School Teachers (គ្រូទុតិយភូមិ) for Grades 10-12, Middle School (គ្រូបឋមភូមិ) for Grades 7-9
                def score_candidate(cand):
                    is_tutiya = 'ទុតិយភូមិ' in (cand.training_level or '')
                    level_match_bonus = 0
                    if is_high_school and is_tutiya:
                        level_match_bonus = -50  # Lower score is prioritized
                    elif not is_high_school and not is_tutiya:
                        level_match_bonus = -50
                    return (level_match_bonus, teacher_loads.get(cand.id, 0), cand.id)

                candidates.sort(key=score_candidate)
                
                chosen = None
                for cand in candidates:
                    if teacher_loads.get(cand.id, 0) + h_req <= teacher_max.get(cand.id, 18):
                        chosen = cand
                        break
                if not chosen and candidates:
                    # Over-quota fallback candidate with minimum load
                    candidates.sort(key=lambda t: (teacher_loads.get(t.id, 0), t.id))
                    chosen = candidates[0]

                if chosen:
                    cs = existing_cs_map.get((cls.id, sub.id))
                    if not cs:
                        new_cs = ClassSubject(
                            classroom=cls,
                            subject=sub,
                            teacher=chosen,
                            weekly_hours=h_req or 4,
                        )
                        cs_to_create.append(new_cs)
                        existing_cs_map[(cls.id, sub.id)] = new_cs
                    else:
                        cs.teacher = chosen
                        if h_req:
                            cs.weekly_hours = h_req
                        cs_to_update.append(cs)
                    
                    teacher_loads[chosen.id] = teacher_loads.get(chosen.id, 0) + h_req
                    assigned_count += 1

        with transaction.atomic():
            if cs_to_create:
                ClassSubject.objects.bulk_create(cs_to_create, ignore_conflicts=True)
            if cs_to_update:
                ClassSubject.objects.bulk_update(cs_to_update, ['teacher', 'weekly_hours'])

        messages.success(request, f"បានចាត់តាំងគ្រូបង្រៀនតាមឯកទេស និងកម្រិតបណ្តុះបណ្តាលស្វ័យប្រវត្តិ ({assigned_count} ការចាត់តាំង, គ្រូទុតិយភូមិ=16h, ផ្សេងៗ=18h) ជោគជ័យ!")
    except Exception as e:
        messages.error(request, f"កំហុសក្នុងការចាត់តាំងស្វ័យប្រវត្តិ៖ {str(e)}")
        
    return redirect(f"/academics/teacher-assignments/{f'?year={active_year.id}' if active_year else ''}")

# ----------------- TEACHER & STAFF ON-DUTY ALLOCATION (ម៉ោងប្រចាំការ) -----------------

@login_required
@role_required(['ADMIN', 'TEACHER'])
def teacher_duty_manager(request):
    """
    Teacher & Office Staff On-Duty (ម៉ោងប្រចាំការ) Allocation Manager.
    - Tracks teaching hours vs target weekly quota (e.g. 18 hours).
    - Under-quota teachers get deficit duty hours assigned (manually or auto/random).
    - Office/Administrative staff have 100% duty hours assigned manually with priority.
    - Fully scoped per Academic Year!
    """
    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)
    selected_year = request.GET.get('year') or request.GET.get('academic_year')
    if selected_year:
        if str(selected_year).isdigit():
            found_year = AcademicYear.objects.filter(id=int(selected_year)).first()
        else:
            found_year = AcademicYear.objects.filter(name=str(selected_year).strip()).first()
        if found_year:
            active_year = found_year

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    all_teachers = Teacher.objects.filter(status='ACTIVE').order_by('khmer_name')

    # Query all timetable teaching slots for this active year
    timetables = Timetable.objects.filter(
        classroom__academic_year=active_year
    ).select_related('classroom', 'subject', 'teacher') if active_year else Timetable.objects.all().select_related('classroom', 'subject', 'teacher')

    # Query all duty schedules for this active year
    duty_entries = TeacherDutySchedule.objects.filter(
        academic_year=active_year
    ).select_related('teacher') if active_year else TeacherDutySchedule.objects.all().select_related('teacher')

    # Map teaching hours per teacher
    teacher_teaching_map = {}
    teacher_teaching_slots_map = {}
    for entry in timetables:
        tid = entry.teacher_id
        if not tid:
            continue
        teacher_teaching_map[tid] = teacher_teaching_map.get(tid, 0) + 1
        if tid not in teacher_teaching_slots_map:
            teacher_teaching_slots_map[tid] = {}
        slot_k = f"{entry.day_of_week}_{entry.period_number}"
        teacher_teaching_slots_map[tid][slot_k] = {
            'classroom_name': entry.classroom.name,
            'subject_name': entry.subject.name_kh,
            'code': entry.subject.code
        }

    raw_duty_types = TeacherDutyType.get_all_duty_types()
    duty_types_dict = {dt.code: dt.name for dt in raw_duty_types}

    # Map duty hours per teacher
    teacher_duty_map = {}
    teacher_duty_slots_map = {}
    for d in duty_entries:
        tid = d.teacher_id
        teacher_duty_map[tid] = teacher_duty_map.get(tid, 0) + 1
        if tid not in teacher_duty_slots_map:
            teacher_duty_slots_map[tid] = {}
        slot_k = f"{d.day_of_week}_{d.period_number}"
        teacher_duty_slots_map[tid][slot_k] = {
            'id': d.id,
            'duty_type': d.duty_type,
            'duty_label': duty_types_dict.get(d.duty_type, d.duty_type),
            'is_auto': d.is_auto_assigned,
            'notes': d.notes or ''
        }

    # Build teacher and office staff statistics
    staff_list = []
    office_keywords = ['ការិយាល័យ', 'រដ្ឋបាល', 'បណ្ណារក្ស', 'គណនេយ្យ', 'នាយក', 'នាយករង', 'បុគ្គលិក', 'សន្តិសុខ', 'អនាម័យ']
    
    total_teachers_count = 0
    total_office_count = 0
    total_deficit_count = 0
    total_exact_count = 0
    total_duty_hours_school = sum(teacher_duty_map.values())
    total_teaching_hours_school = sum(teacher_teaching_map.values())

    selected_teacher_id = request.GET.get('teacher')
    selected_teacher = None
    if selected_teacher_id and str(selected_teacher_id).isdigit():
        selected_teacher = all_teachers.filter(id=int(selected_teacher_id)).first()

    for t in all_teachers:
        t_duty_str = (t.current_duty or '').strip()
        t_spec_str = (t.specialization or '').strip()
        is_office = any(kw in t_duty_str for kw in office_keywords) or any(kw in t_spec_str for kw in office_keywords)
        teaching_h = teacher_teaching_map.get(t.id, 0)
        if teaching_h == 0 and not t_spec_str:
            is_office = True

        duty_h = teacher_duty_map.get(t.id, 0)
        target_max = t.max_weekly_hours or 18
        total_h = teaching_h + duty_h
        deficit_h = max(0, target_max - teaching_h) if not is_office else target_max
        remaining_duty_needed = max(0, target_max - total_h)

        if is_office:
            total_office_count += 1
        else:
            total_teachers_count += 1
            if total_h >= target_max:
                total_exact_count += 1
            else:
                total_deficit_count += 1

        staff_list.append({
            'id': t.id,
            'teacher_id': t.teacher_id,
            'khmer_name': t.khmer_name,
            'latin_name': t.latin_name,
            'gender': t.gender,
            'gender_kh': 'ស្រី' if t.gender == 'F' else 'ប្រុស',
            'phone': t.phone or '',
            'current_duty': t.current_duty or ('បុគ្គលិកការិយាល័យ' if is_office else 'គ្រូបង្រៀន'),
            'specialization': t.specialization or '',
            'is_office': is_office,
            'teaching_hours': teaching_h,
            'duty_hours': duty_h,
            'total_hours': total_h,
            'target_max': target_max,
            'deficit_hours': deficit_h,
            'remaining_needed': remaining_duty_needed,
            'status': 'EXACT' if total_h == target_max else ('OVER' if total_h > target_max else 'DEFICIT'),
        })

    if not selected_teacher and staff_list:
        selected_teacher = all_teachers.first()

    active_staff_duty_slots = teacher_duty_slots_map.get(selected_teacher.id, {}) if selected_teacher else {}
    active_staff_teaching_slots = teacher_teaching_slots_map.get(selected_teacher.id, {}) if selected_teacher else {}

    duty_types = [
        {
            'id': dt.id,
            'code': dt.code,
            'name': dt.name,
            'icon': dt.icon,
            'color': dt.color,
            'order': dt.order
        }
        for dt in raw_duty_types
    ]

    context = {
        'active_year': active_year,
        'academic_years': academic_years,
        'staff_list': staff_list,
        'staff_list_json': json.dumps(staff_list, ensure_ascii=False),
        'selected_teacher': selected_teacher,
        'active_staff_duty_slots': active_staff_duty_slots,
        'active_staff_duty_slots_json': json.dumps(active_staff_duty_slots, ensure_ascii=False),
        'active_staff_teaching_slots': active_staff_teaching_slots,
        'active_staff_teaching_slots_json': json.dumps(active_staff_teaching_slots, ensure_ascii=False),
        'all_duty_slots_json': json.dumps(teacher_duty_slots_map, ensure_ascii=False),
        'all_teaching_slots_json': json.dumps(teacher_teaching_slots_map, ensure_ascii=False),
        'duty_types': duty_types,
        'duty_types_json': json.dumps(duty_types, ensure_ascii=False),
        'total_teachers_count': total_teachers_count,
        'total_office_count': total_office_count,
        'total_deficit_count': total_deficit_count,
        'total_exact_count': total_exact_count,
        'total_duty_hours_school': total_duty_hours_school,
        'total_teaching_hours_school': total_teaching_hours_school,
    }
    return render(request, 'academics/teacher_duty_manager.html', context)


@login_required
@role_required(['ADMIN'])
def teacher_duty_save_manual(request):
    """
    Saves manually assigned duty slots for a teacher/staff.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid HTTP method'}, status=405)

    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    try:
        data = json.loads(request.body.decode('utf-8'))
        teacher_id = data.get('teacher_id')
        year_id = data.get('academic_year_id')
        if year_id:
            found_year = AcademicYear.objects.filter(id=int(year_id)).first()
            if found_year:
                active_year = found_year

        if not active_year:
            return JsonResponse({'status': 'error', 'message': 'មិនមានឆ្នាំសិក្សាសកម្មឡើយ'}, status=400)

        teacher = Teacher.objects.filter(id=teacher_id).first()
        if not teacher:
            return JsonResponse({'status': 'error', 'message': 'រកមិនឃើញគ្រូ ឬបុគ្គលិកឡើយ'}, status=404)

        # Optional update max_weekly_hours
        max_h = data.get('max_weekly_hours')
        if max_h is not None and str(max_h).isdigit():
            teacher.max_weekly_hours = int(max_h)
            teacher.save(update_fields=['max_weekly_hours'])

        slots = data.get('slots', [])

        with transaction.atomic():
            TeacherDutySchedule.objects.filter(
                academic_year=active_year,
                teacher=teacher
            ).delete()

            new_entries = []
            for s in slots:
                d_num = int(s.get('day_of_week'))
                p_num = int(s.get('period_number'))
                dtype = s.get('duty_type') or 'OFFICE'
                notes = s.get('notes') or ''
                new_entries.append(TeacherDutySchedule(
                    academic_year=active_year,
                    teacher=teacher,
                    day_of_week=d_num,
                    period_number=p_num,
                    duty_type=dtype,
                    is_auto_assigned=False,
                    notes=notes,
                ))
            if new_entries:
                TeacherDutySchedule.objects.bulk_create(new_entries)

        return JsonResponse({
            'status': 'success',
            'message': f'បានរក្សាទុកម៉ោងប្រចាំការ ({len(new_entries)} ម៉ោង) ជូន {teacher.khmer_name} ដោយជោគជ័យ!',
            'count': len(new_entries),
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@role_required(['ADMIN'])
def teacher_duty_auto_assign(request):
    """
    Intelligent Auto / Random Duty Allocation Solver.
    Calculates deficit teaching hours for under-quota teachers,
    finds completely free timetable slots (no teaching & no existing manual duty),
    and distributes deficit duty hours so they reach 100% full quota!
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid HTTP method'}, status=405)

    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    try:
        data = json.loads(request.body.decode('utf-8')) if request.body else {}
        target_teacher_id = data.get('teacher_id')
        year_id = data.get('academic_year_id')
        default_duty_type = data.get('duty_type') or 'OFFICE'
        reset_existing_auto = data.get('reset_existing_auto', True)

        if year_id:
            found_year = AcademicYear.objects.filter(id=int(year_id)).first()
            if found_year:
                active_year = found_year

        if not active_year:
            return JsonResponse({'status': 'error', 'message': 'មិនមានឆ្នាំសិក្សាសកម្មឡើយ'}, status=400)

        teachers_query = Teacher.objects.filter(status='ACTIVE')
        if target_teacher_id:
            teachers_query = teachers_query.filter(id=target_teacher_id)

        # Get existing timetable teaching slots
        timetables = Timetable.objects.filter(classroom__academic_year=active_year)
        teaching_slots_by_teacher = {}
        for entry in timetables:
            if entry.teacher_id not in teaching_slots_by_teacher:
                teaching_slots_by_teacher[entry.teacher_id] = set()
            teaching_slots_by_teacher[entry.teacher_id].add((entry.day_of_week, entry.period_number))

        # Get existing manual duty slots
        existing_duties = TeacherDutySchedule.objects.filter(academic_year=active_year)
        manual_duties_by_teacher = {}
        for d in existing_duties:
            if not d.is_auto_assigned:
                if d.teacher_id not in manual_duties_by_teacher:
                    manual_duties_by_teacher[d.teacher_id] = set()
                manual_duties_by_teacher[d.teacher_id].add((d.day_of_week, d.period_number))

        import random
        assigned_total_slots = 0
        assigned_teachers_count = 0

        with transaction.atomic():
            if target_teacher_id:
                if reset_existing_auto:
                    TeacherDutySchedule.objects.filter(
                        academic_year=active_year,
                        teacher_id=target_teacher_id,
                        is_auto_assigned=True
                    ).delete()
            else:
                if reset_existing_auto:
                    TeacherDutySchedule.objects.filter(
                        academic_year=active_year,
                        is_auto_assigned=True
                    ).delete()

            new_duty_entries = []

            for teacher in teachers_query:
                t_teach_slots = teaching_slots_by_teacher.get(teacher.id, set())
                t_manual_slots = manual_duties_by_teacher.get(teacher.id, set())
                
                teaching_count = len(t_teach_slots)
                manual_duty_count = len(t_manual_slots)
                target_max = teacher.max_weekly_hours or 18

                needed_duty = target_max - (teaching_count + manual_duty_count)
                if needed_duty <= 0:
                    continue

                all_possible_slots = []
                for day in range(1, 7):
                    for period in range(1, 9):
                        slot_tuple = (day, period)
                        if slot_tuple not in t_teach_slots and slot_tuple not in t_manual_slots:
                            all_possible_slots.append(slot_tuple)

                if not all_possible_slots:
                    continue

                random.shuffle(all_possible_slots)
                teaching_days = {d for (d, p) in t_teach_slots}
                def slot_priority(s):
                    day, period = s
                    same_day_score = 0 if day in teaching_days else 10
                    return (same_day_score, day, period)

                all_possible_slots.sort(key=slot_priority)
                
                chosen_slots = all_possible_slots[:needed_duty]
                for (day, period) in chosen_slots:
                    new_duty_entries.append(TeacherDutySchedule(
                        academic_year=active_year,
                        teacher=teacher,
                        day_of_week=day,
                        period_number=period,
                        duty_type=default_duty_type,
                        is_auto_assigned=True,
                        notes='បែងចែកម៉ោងខ្វះស្វ័យប្រវត្តិ'
                    ))
                    assigned_total_slots += 1

                assigned_teachers_count += 1

            if new_duty_entries:
                TeacherDutySchedule.objects.bulk_create(new_duty_entries)

        msg = f"បានបែងចែកម៉ោងប្រចាំការស្វ័យប្រវត្តិចំនួន {assigned_total_slots} ម៉ោង ជូនគ្រូ/បុគ្គលិក {assigned_teachers_count} នាក់ ដោយជោគជ័យ!"
        return JsonResponse({
            'status': 'success',
            'message': msg,
            'total_slots': assigned_total_slots,
            'teachers_count': assigned_teachers_count
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@role_required(['ADMIN'])
def teacher_duty_clear(request):
    """
    Clears duty schedules for a specific teacher or all teachers in active year.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid HTTP method'}, status=405)

    from .utils import get_active_academic_year
    active_year = get_active_academic_year(request)

    try:
        data = json.loads(request.body.decode('utf-8')) if request.body else {}
        teacher_id = data.get('teacher_id')
        auto_only = data.get('auto_only', False)

        query = TeacherDutySchedule.objects.filter(academic_year=active_year) if active_year else TeacherDutySchedule.objects.all()
        if teacher_id:
            query = query.filter(teacher_id=teacher_id)
        if auto_only:
            query = query.filter(is_auto_assigned=True)

        deleted_count, _ = query.delete()
        return JsonResponse({
            'status': 'success',
            'message': f'បានសម្អាតម៉ោងប្រចាំការចំនួន {deleted_count} ម៉ោងរួចរាល់!',
            'deleted_count': deleted_count
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@role_required(['ADMIN'])
def api_duty_types_list(request):
    """
    Returns JSON list of all configured duty types.
    """
    duty_types = TeacherDutyType.get_all_duty_types()
    data = [
        {
            'id': dt.id,
            'code': dt.code,
            'name': dt.name,
            'icon': dt.icon,
            'color': dt.color,
            'order': dt.order
        }
        for dt in duty_types
    ]
    return JsonResponse({'status': 'success', 'duty_types': data})


@login_required
@role_required(['ADMIN'])
def api_duty_type_create(request):
    """
    Creates a new duty type.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST request required'}, status=405)

    try:
        data = json.loads(request.body.decode('utf-8')) if request.content_type == 'application/json' else request.POST
        name = data.get('name', '').strip()
        icon = data.get('icon', 'fa-clock').strip()
        color = data.get('color', '#4f46e5').strip()
        code = data.get('code', '').strip().upper()

        if not name:
            return JsonResponse({'status': 'error', 'message': 'សូមបញ្ចូលឈ្មោះប្រភេទប្រចាំការ!'}, status=400)

        # Generate unique code if not provided
        if not code:
            import re
            base_code = re.sub(r'[^A-Z0-9_]', '', name.upper().replace(' ', '_')).strip('_')[:20]
            if not base_code:
                base_code = f"DUTY_{TeacherDutyType.objects.count() + 1}"
            code = base_code
            counter = 1
            while TeacherDutyType.objects.filter(code=code).exists():
                code = f"{base_code}_{counter}"
                counter += 1
        elif TeacherDutyType.objects.filter(code=code).exists():
            return JsonResponse({'status': 'error', 'message': f'កូដ "{code}" មានរួចហើយ!'}, status=400)

        max_order = TeacherDutyType.objects.count() + 1
        dt = TeacherDutyType.objects.create(
            code=code,
            name=name,
            icon=icon or 'fa-clock',
            color=color or '#4f46e5',
            order=max_order
        )

        return JsonResponse({
            'status': 'success',
            'message': f'🎉 បានបន្ថែមប្រភេទប្រចាំការ "{dt.name}" ដោយជោគជ័យ!',
            'duty_type': {
                'id': dt.id,
                'code': dt.code,
                'name': dt.name,
                'icon': dt.icon,
                'color': dt.color,
                'order': dt.order
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
@role_required(['ADMIN'])
def api_duty_type_edit(request, type_id):
    """
    Updates an existing duty type.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST request required'}, status=405)

    try:
        dt = get_object_or_404(TeacherDutyType, id=type_id)
        data = json.loads(request.body.decode('utf-8')) if request.content_type == 'application/json' else request.POST
        name = data.get('name', '').strip()
        icon = data.get('icon', '').strip()
        color = data.get('color', '').strip()

        if not name:
            return JsonResponse({'status': 'error', 'message': 'ឈ្មោះប្រភេទប្រចាំការមិនអាចទទេបានឡើយ!'}, status=400)

        dt.name = name
        if icon:
            dt.icon = icon
        if color:
            dt.color = color
        dt.save()

        return JsonResponse({
            'status': 'success',
            'message': f'🎉 បានកែប្រែប្រភេទប្រចាំការ "{dt.name}" ដោយជោគជ័យ!',
            'duty_type': {
                'id': dt.id,
                'code': dt.code,
                'name': dt.name,
                'icon': dt.icon,
                'color': dt.color,
                'order': dt.order
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
@role_required(['ADMIN'])
def api_duty_type_delete(request, type_id):
    """
    Deletes an existing duty type safely.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST request required'}, status=405)

    try:
        dt = get_object_or_404(TeacherDutyType, id=type_id)
        if TeacherDutyType.objects.count() <= 1:
            return JsonResponse({'status': 'error', 'message': 'មិនអាចលុបបានទេ ត្រូវមានយ៉ាងហោចណាស់ប្រភេទប្រចាំការមួយក្នុងប្រព័ន្ធ!'}, status=400)

        name = dt.name
        code = dt.code
        dt.delete()

        return JsonResponse({
            'status': 'success',
            'message': f'🗑️ បានលុបប្រភេទប្រចាំការ "{name}" ដោយជោគជ័យ!',
            'deleted_code': code
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)




@login_required
@role_required(['ADMIN', 'TEACHER'])
def student_promotion_view(request):
    """
    Student Promotion & Grade Retention Matrix (ឧបករណ៍ផ្ទេរ ឡើងថ្នាក់ និងត្រួតថ្នាក់សិស្ស).
    Allows Admin & Authorized Teachers to specify individual promotion / retention decisions per student
    with MoEYS standard reasons, target classrooms, and full audit logs.
    """
    from apps.students.models import StudentPromotionRecord

    is_admin = request.user.is_superuser or getattr(request.user, 'role', '') == 'ADMIN'
    teacher_profile = None

    if not is_admin:
        teacher_profile = Teacher.objects.filter(user=request.user).first()
        if not teacher_profile:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនមានសិទ្ធិចាត់ចែងការឡើងថ្នាក់/ត្រួតថ្នាក់សិស្សឡើយ!")
            return redirect('dashboard')
        # Allowed classrooms where teacher is class head or teaches
        taught_class_ids = set(ClassSubject.objects.filter(teacher=teacher_profile).values_list('classroom_id', flat=True))
        if hasattr(Classroom, 'teacher'):
            taught_class_ids.update(Classroom.objects.filter(teacher=teacher_profile).values_list('id', flat=True))
        classrooms = Classroom.objects.filter(id__in=taught_class_ids).select_related('academic_year')
    else:
        classrooms = Classroom.objects.all().select_related('academic_year')

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    all_target_classrooms = Classroom.objects.all().select_related('academic_year').order_by('grade_level', 'name')

    source_class_id = request.GET.get('source_class')
    students = []
    source_class = None

    if source_class_id and str(source_class_id).strip().isdigit():
        cls_int_id = int(str(source_class_id).strip())
        source_class = classrooms.filter(id=cls_int_id).first() if not is_admin else Classroom.objects.filter(id=cls_int_id).first()
        if source_class:
            students = Student.objects.filter(classroom=source_class, status='ACTIVE').order_by('student_id')

    if request.method == 'POST':
        source_class_post = request.POST.get('source_class')
        target_year_id = request.POST.get('target_year')
        global_action = request.POST.get('global_promotion_action', 'PROMOTE')
        global_target_class_id = request.POST.get('global_target_class')
        selected_student_ids = request.POST.getlist('student_ids')

        redirect_url = f"/academics/promotion/?source_class={source_class_post}" if (source_class_post and str(source_class_post).strip().isdigit()) else "/academics/promotion/"

        if not selected_student_ids:
            messages.error(request, "⚠️ សូមជ្រើសរើសសិស្សយ៉ាងតិចម្នាក់ដើម្បីដំណើរការ!")
            return redirect(redirect_url)

        target_year = AcademicYear.objects.filter(pk=target_year_id).first() if (target_year_id and str(target_year_id).strip().isdigit()) else None
        global_target_class = Classroom.objects.filter(pk=global_target_class_id).first() if (global_target_class_id and str(global_target_class_id).strip().isdigit()) else None

        promoted_count = 0
        retained_count = 0
        other_count = 0

        with transaction.atomic():
            for student_id in selected_student_ids:
                if not str(student_id).strip().isdigit():
                    continue
                student = Student.objects.filter(id=int(str(student_id).strip())).first()
                if not student:
                    continue

                # Individual decision or fallback to global
                action = request.POST.get(f'action_{student_id}', global_action)
                target_cid = request.POST.get(f'target_class_{student_id}') or global_target_class_id
                target_cls = Classroom.objects.filter(id=int(str(target_cid).strip())).first() if (target_cid and str(target_cid).strip().isdigit()) else global_target_class
                standard_reason = request.POST.get(f'reason_{student_id}', 'PASSED_YEAR')
                custom_notes = request.POST.get(f'notes_{student_id}', '').strip()

                old_class = student.classroom
                old_year = student.academic_year

                # Map action to Cambodian label for reason
                reason_display = dict(StudentPromotionRecord.StandardReason.choices).get(standard_reason, standard_reason)
                full_reason = f"{reason_display}" + (f" ({custom_notes})" if custom_notes else "")

                if action == 'PROMOTE':
                    student.academic_year = target_year or old_year
                    if target_cls:
                        student.classroom = target_cls
                    student.status = 'ACTIVE'
                    student.is_repeating_grade = False
                    student.last_promotion_status = 'ឡើងថ្នាក់'
                    student.last_promotion_reason = full_reason
                    student.save()
                    promoted_count += 1

                elif action == 'RETAIN':
                    student.academic_year = target_year or old_year
                    if target_cls:
                        student.classroom = target_cls
                    student.status = 'ACTIVE'
                    student.is_repeating_grade = True
                    student.last_promotion_status = 'ត្រួតថ្នាក់'
                    student.last_promotion_reason = full_reason
                    student.save()
                    retained_count += 1

                elif action == 'GRADUATE':
                    student.status = 'GRADUATED'
                    student.is_repeating_grade = False
                    student.last_promotion_status = 'បញ្ចប់ការសិក្សា'
                    student.last_promotion_reason = full_reason
                    student.save()
                    other_count += 1

                elif action == 'TRANSFER':
                    student.status = 'TRANSFERRED'
                    student.last_promotion_status = 'ផ្ទេរចេញ'
                    student.last_promotion_reason = full_reason
                    student.save()
                    other_count += 1

                elif action == 'DROP':
                    student.status = 'DROPPED'
                    student.last_promotion_status = 'ឈប់រៀន'
                    student.last_promotion_reason = full_reason
                    student.save()
                    other_count += 1

                # Record Audit History
                StudentPromotionRecord.objects.create(
                    student=student,
                    from_academic_year=old_year,
                    to_academic_year=target_year or old_year,
                    from_classroom=old_class,
                    to_classroom=target_cls,
                    action=action,
                    standard_reason=standard_reason,
                    custom_notes=custom_notes,
                    processed_by=request.user
                )

        total_done = promoted_count + retained_count + other_count
        messages.success(
            request,
            f"🎉 ជោគជ័យ! បានដំណើរការឡើងថ្នាក់/ត្រួតថ្នាក់សិស្សសរុប {total_done} នាក់ "
            f"(ឡើងថ្នាក់: {promoted_count} នាក់, ត្រួតថ្នាក់: {retained_count} នាក់, ផ្សេងៗ: {other_count} នាក់)។"
        )
        return redirect(redirect_url)

    # Recent promotion audit logs
    recent_promotions = StudentPromotionRecord.objects.select_related(
        'student', 'from_classroom', 'to_classroom', 'from_academic_year', 'to_academic_year', 'processed_by'
    ).all()[:60]

    return render(request, 'academics/promotion.html', {
        'academic_years': academic_years,
        'classrooms': classrooms,
        'all_target_classrooms': all_target_classrooms,
        'source_class_id': source_class_id,
        'source_class': source_class,
        'students': students,
        'recent_promotions': recent_promotions,
        'standard_reasons': StudentPromotionRecord.StandardReason.choices,
        'is_admin': is_admin,
    })


# =========================================================================
# Academic Years Manager (គ្រប់គ្រងឆ្នាំសិក្សា)
# =========================================================================
@login_required
@role_required(['ADMIN'])
def academic_year_list(request):
    """
    Manage Academic Years (ឆ្នាំសិក្សា) - List, Add, Set Current, Edit, Delete.
    """
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    form = AcademicYearForm()
    
    years_data = []
    for ay in academic_years:
        years_data.append({
            'year': ay,
            'students_count': ay.enrolled_students.count(),
            'classrooms_count': ay.classrooms.count(),
        })

    return render(request, 'academics/academic_year_list.html', {
        'years_data': years_data,
        'form': form,
    })


@login_required
@role_required(['ADMIN'])
def academic_year_create(request):
    if request.method == 'POST':
        form = AcademicYearForm(request.POST)
        if form.is_valid():
            ay = form.save()
            messages.success(request, f"🎉 បានបន្ថែមឆ្នាំសិក្សា {ay.name} ដោយជោគជ័យ!")
            return redirect('academic_year_list')
        else:
            messages.error(request, "សូមពិនិត្យទិន្នន័យឆ្នាំសិក្សាឡើងវិញ!")
    return redirect('academic_year_list')


@login_required
@role_required(['ADMIN'])
def academic_year_edit(request, pk):
    ay = get_object_or_404(AcademicYear, pk=pk)
    if request.method == 'POST':
        form = AcademicYearForm(request.POST, instance=ay)
        if form.is_valid():
            form.save()
            messages.success(request, f"បានកែប្រែឆ្នាំសិក្សា {ay.name} ដោយជោគជ័យ!")
            return redirect('academic_year_list')
    return redirect('academic_year_list')


@login_required
@role_required(['ADMIN'])
def academic_year_delete(request, pk):
    ay = get_object_or_404(AcademicYear, pk=pk)
    if ay.is_current:
        messages.error(request, f"មិនអាចលុបឆ្នាំសិក្សាបច្ចុប្បន្ន ({ay.name}) បានទេ! សូមកំណត់ឆ្នាំសិក្សាផ្សេងជាបច្ចុប្បន្នសិន។")
        return redirect('academic_year_list')
    
    if ay.enrolled_students.exists() or ay.classrooms.exists():
        messages.error(request, f"មិនអាចលុបឆ្នាំសិក្សា {ay.name} បានទេ ព្រោះមានទិន្នន័យសិស្ស ឬថ្នាក់រៀនភ្ជាប់ជាមួយ!")
        return redirect('academic_year_list')

    name = ay.name
    ay.delete()
    messages.success(request, f"បានលុបឆ្នាំសិក្សា {name} ដោយជោគជ័យ!")
    return redirect('academic_year_list')


@login_required
def academic_year_switch(request, pk):
    """
    Session-based Academic Year Switcher for the active session.
    """
    ay = get_object_or_404(AcademicYear, pk=pk)
    request.session['active_academic_year_id'] = ay.id
    messages.info(request, f"🔄 បានប្តូរទៅកាន់ឆ្នាំសិក្សា៖ {ay.name}")
    next_url = request.META.get('HTTP_REFERER') or '/'
    return redirect(next_url)


@login_required
@role_required(['ADMIN'])
def academic_year_set_current(request, pk):
    ay = get_object_or_404(AcademicYear, pk=pk)
    AcademicYear.objects.filter(is_current=True).update(is_current=False)
    ay.is_current = True
    ay.save()
    request.session['active_academic_year_id'] = ay.id
    messages.success(request, f"🎉 បានកំណត់ឆ្នាំសិក្សា {ay.name} ជាឆ្នាំសិក្សាបច្ចុប្បន្ន (Current Active Year)!")
    return redirect('academic_year_list')



# =========================================================================
# Location AJAX APIs (For Cascading Dropdowns in Forms)
# =========================================================================
def _location_sort_key(x):
    code_str = str(x.get('code') or '').strip()
    try:
        return (0, int(code_str), code_str)
    except ValueError:
        return (1, 0, code_str)


def api_locations_provinces(request):
    provinces = list(Province.objects.all().values('id', 'code', 'name_kh', 'name_en'))
    provinces.sort(key=_location_sort_key)
    return JsonResponse({'status': 'success', 'data': provinces})


def api_locations_districts(request):
    province_id = request.GET.get('province_id')
    districts = District.objects.all()
    if province_id:
        districts = districts.filter(province_id=province_id)
    data = list(districts.values('id', 'code', 'name_kh', 'name_en', 'province_id'))
    data.sort(key=_location_sort_key)
    return JsonResponse({'status': 'success', 'data': data})


def api_locations_communes(request):
    district_id = request.GET.get('district_id')
    communes = Commune.objects.all()
    if district_id:
        communes = communes.filter(district_id=district_id)
    data = list(communes.values('id', 'code', 'name_kh', 'name_en', 'district_id'))
    data.sort(key=_location_sort_key)
    return JsonResponse({'status': 'success', 'data': data})


def api_locations_villages(request):
    commune_id = request.GET.get('commune_id')
    villages = Village.objects.all()
    if commune_id:
        villages = villages.filter(commune_id=commune_id)
    data = list(villages.values('id', 'code', 'name_kh', 'name_en', 'commune_id'))
    data.sort(key=_location_sort_key)
    return JsonResponse({'status': 'success', 'data': data})


# =========================================================================
# Admin Location Manager (គ្រប់គ្រង និងកែតម្រូវឈ្មោះខេត្ត ស្រុក ឃុំ ភូមិ)
# =========================================================================
@login_required
@role_required(['ADMIN'])
def location_manager_view(request):
    """
    Administrative Divisions Manager (Admin only).
    Allows Admin to browse, search, sort, and edit Province, District, Commune, and Village names.
    """
    level = request.GET.get('level', 'province') # province, district, commune, village
    search = request.GET.get('q', '').strip()
    province_id = request.GET.get('province_id', '')
    district_id = request.GET.get('district_id', '')
    commune_id = request.GET.get('commune_id', '')
    sort_by = request.GET.get('sort', 'code') # code, name_kh, name_en, parent
    order = request.GET.get('order', 'asc') # asc, desc

    provinces = Province.objects.all().order_by('code')
    districts = District.objects.filter(province_id=province_id).order_by('code') if province_id else District.objects.none()
    communes = Commune.objects.filter(district_id=district_id).order_by('code') if district_id else Commune.objects.none()

    if level == 'province':
        qs = Province.objects.all()
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search))
    elif level == 'district':
        qs = District.objects.select_related('province').all()
        if province_id:
            qs = qs.filter(province_id=province_id)
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search))
    elif level == 'commune':
        qs = Commune.objects.select_related('district__province').all()
        if district_id:
            qs = qs.filter(district_id=district_id)
        elif province_id:
            qs = qs.filter(district__province_id=province_id)
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search))
    elif level == 'village':
        qs = Village.objects.select_related('commune__district__province').all()
        if commune_id:
            qs = qs.filter(commune_id=commune_id)
        elif district_id:
            qs = qs.filter(commune__district_id=district_id)
        elif province_id:
            qs = qs.filter(commune__district__province_id=province_id)
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search))
    else:
        qs = Province.objects.all()

    items = list(qs)

    # Natural sorting helper
    def natural_sort_key(item):
        if sort_by == 'code':
            c_str = str(getattr(item, 'code', '') or '').strip()
            try:
                return (0, int(c_str), c_str)
            except ValueError:
                return (1, 0, c_str)
        elif sort_by == 'name_kh':
            return (0, 0, getattr(item, 'name_kh', '') or '')
        elif sort_by == 'name_en':
            return (0, 0, (getattr(item, 'name_en', '') or '').lower())
        elif sort_by == 'parent':
            if hasattr(item, 'province'):
                return (0, 0, item.province.name_kh)
            elif hasattr(item, 'district'):
                return (0, 0, item.district.name_kh)
            elif hasattr(item, 'commune'):
                return (0, 0, item.commune.name_kh)
            return (0, 0, '')
        return (0, 0, '')

    is_reverse = (order == 'desc')
    items.sort(key=natural_sort_key, reverse=is_reverse)
    items = items[:250]

    level_names = {
        'province': 'ខេត្ត/រាជធានី',
        'district': 'ស្រុក/ខណ្ឌ/ក្រុង',
        'commune': 'ឃុំ/សង្កាត់',
        'village': 'ភូមិ',
    }
    level_kh = level_names.get(level, 'ខេត្ត/រាជធានី')

    context = {
        'level': level,
        'level_kh': level_kh,
        'search': search,
        'province_id': province_id,
        'district_id': district_id,
        'commune_id': commune_id,
        'sort_by': sort_by,
        'order': order,
        'provinces': provinces,
        'districts': districts,
        'communes': communes,
        'items': items,
        'total_provinces': Province.objects.count(),
        'total_districts': District.objects.count(),
        'total_communes': Commune.objects.count(),
        'total_villages': Village.objects.count(),
    }
    return render(request, 'academics/location_manager.html', context)


@login_required
@role_required(['ADMIN'])
def location_item_create(request):
    """
    POST create new location item (Province, District, Commune, Village).
    """
    if request.method == 'POST':
        level = request.POST.get('level')
        code = request.POST.get('code', '').strip()
        name_kh = request.POST.get('name_kh', '').strip()
        name_en = request.POST.get('name_en', '').strip()

        if not name_kh:
            messages.error(request, "សូមបំពេញឈ្មោះជាភាសាខ្មែរ!")
            return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))

        if level == 'province':
            if not code:
                last_code = Province.objects.order_by('-id').values_list('code', flat=True).first()
                code = str(int(last_code) + 1) if last_code and last_code.isdigit() else str(Province.objects.count() + 1)
            Province.objects.create(code=code, name_kh=name_kh, name_en=name_en)
            messages.success(request, f"🎉 បានបន្ថែមខេត្ត/រាជធានី {name_kh} ដោយជោគជ័យ!")

        elif level == 'district':
            province_id = request.POST.get('province_id')
            if not province_id:
                messages.error(request, "សូមជ្រើសរើសខេត្ត/រាជធានី!")
                return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))
            province = get_object_or_404(Province, id=province_id)
            if not code:
                code = f"{province.code}{District.objects.filter(province=province).count() + 1:02d}"
            District.objects.create(province=province, code=code, name_kh=name_kh, name_en=name_en)
            messages.success(request, f"🎉 បានបន្ថែមស្រុក/ខណ្ឌ/ក្រុង {name_kh} ក្នុង {province.name_kh} ដោយជោគជ័យ!")

        elif level == 'commune':
            district_id = request.POST.get('district_id')
            if not district_id:
                messages.error(request, "សូមជ្រើសរើសស្រុក/ខណ្ឌ!")
                return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))
            district = get_object_or_404(District, id=district_id)
            if not code:
                code = f"{district.code}{Commune.objects.filter(district=district).count() + 1:02d}"
            Commune.objects.create(district=district, code=code, name_kh=name_kh, name_en=name_en)
            messages.success(request, f"🎉 បានបន្ថែមឃុំ/សង្កាត់ {name_kh} ក្នុង {district.name_kh} ដោយជោគជ័យ!")

        elif level == 'village':
            commune_id = request.POST.get('commune_id')
            if not commune_id:
                messages.error(request, "សូមជ្រើសរើសឃុំ/សង្កាត់!")
                return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))
            commune = get_object_or_404(Commune, id=commune_id)
            if not code:
                code = f"{commune.code}{Village.objects.filter(commune=commune).count() + 1:02d}"
            Village.objects.create(commune=commune, code=code, name_kh=name_kh, name_en=name_en)
            messages.success(request, f"🎉 បានបន្ថែមភូមិ {name_kh} ក្នុង {commune.name_kh} ដោយជោគជ័យ!")

    return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))


@login_required
@role_required(['ADMIN'])
def location_item_delete(request):
    """
    POST delete a location item (Province, District, Commune, Village).
    """
    if request.method == 'POST':
        level = request.POST.get('level')
        item_id = request.POST.get('item_id')

        if level == 'province':
            item = get_object_or_404(Province, id=item_id)
            name = item.name_kh
            item.delete()
            messages.success(request, f"បានលុបខេត្ត/រាជធានី {name} ដោយជោគជ័យ!")
        elif level == 'district':
            item = get_object_or_404(District, id=item_id)
            name = item.name_kh
            item.delete()
            messages.success(request, f"បានលុបស្រុក/ខណ្ឌ {name} ដោយជោគជ័យ!")
        elif level == 'commune':
            item = get_object_or_404(Commune, id=item_id)
            name = item.name_kh
            item.delete()
            messages.success(request, f"បានលុបឃុំ/សង្កាត់ {name} ដោយជោគជ័យ!")
        elif level == 'village':
            item = get_object_or_404(Village, id=item_id)
            name = item.name_kh
            item.delete()
            messages.success(request, f"បានលុបភូមិ {name} ដោយជោគជ័យ!")

    return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))


@login_required
@role_required(['ADMIN'])
def location_item_edit(request):
    """
    POST edit location item name (Province, District, Commune, Village).
    """
    if request.method == 'POST':
        level = request.POST.get('level')
        item_id = request.POST.get('item_id')
        code = request.POST.get('code', '').strip()
        name_kh = request.POST.get('name_kh', '').strip()
        name_en = request.POST.get('name_en', '').strip()

        if not name_kh:
            messages.error(request, "ឈ្មោះមិនអាចទទេបានទេ!")
            return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))

        if level == 'province':
            item = get_object_or_404(Province, id=item_id)
        elif level == 'district':
            item = get_object_or_404(District, id=item_id)
        elif level == 'commune':
            item = get_object_or_404(Commune, id=item_id)
        elif level == 'village':
            item = get_object_or_404(Village, id=item_id)
        else:
            messages.error(request, "ប្រភេទមិនត្រឹមត្រូវ!")
            return redirect('location_manager_view')

        if code:
            item.code = code
        item.name_kh = name_kh
        item.name_en = name_en
        item.save()
        messages.success(request, f"🎉 បានកែសម្រួលឈ្មោះ {name_kh} ដោយជោគជ័យ!")

    return redirect(request.META.get('HTTP_REFERER', 'location_manager_view'))


@login_required
@role_required(['ADMIN'])
def location_sync_excel(request):
    """
    Re-sync all locations from E:/SchoolSM/Cambodia All List2025.xlsx.
    """
    import openpyxl
    excel_file = os.path.join(settings.BASE_DIR, 'Cambodia All List2025.xlsx')
    if not os.path.exists(excel_file):
        messages.error(request, f"រកមិនឃើញឯកសារ Excel នៅ {excel_file} ទេ!")
        return redirect('location_manager_view')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        # 1. Provinces
        ws_prov = wb['CambodiaProvinceList2025']
        provinces_to_create = []
        for r in range(2, ws_prov.max_row + 1):
            p_code = str(ws_prov.cell(row=r, column=1).value or '').strip()
            p_kh = str(ws_prov.cell(row=r, column=2).value or '').strip()
            p_en = str(ws_prov.cell(row=r, column=3).value or '').strip()
            if p_code and p_kh:
                provinces_to_create.append(Province(code=p_code, name_kh=p_kh, name_en=p_en))

        with transaction.atomic():
            Province.objects.all().delete()
            Province.objects.bulk_create(provinces_to_create)

        province_map = {p.code: p.id for p in Province.objects.all()}

        # 2. Districts
        ws_dist = wb['CambodiaDistrictList2025']
        districts_to_create = []
        for r in range(2, ws_dist.max_row + 1):
            p_code = str(ws_dist.cell(row=r, column=1).value or '').strip()
            d_code = str(ws_dist.cell(row=r, column=2).value or '').strip()
            d_kh = str(ws_dist.cell(row=r, column=3).value or '').strip()
            d_en = str(ws_dist.cell(row=r, column=4).value or '').strip()
            if d_code and d_kh and p_code in province_map:
                districts_to_create.append(District(province_id=province_map[p_code], code=d_code, name_kh=d_kh, name_en=d_en))

        with transaction.atomic():
            District.objects.all().delete()
            District.objects.bulk_create(districts_to_create)

        district_map = {d.code: d.id for d in District.objects.all()}

        # 3. Communes
        ws_comm = wb['CambodiaCommuneList2025']
        communes_to_create = []
        for r in range(2, ws_comm.max_row + 1):
            d_code = str(ws_comm.cell(row=r, column=2).value or '').strip()
            c_code = str(ws_comm.cell(row=r, column=3).value or '').strip()
            c_kh = str(ws_comm.cell(row=r, column=4).value or '').strip()
            c_en = str(ws_comm.cell(row=r, column=5).value or '').strip()
            if c_code and c_kh and d_code in district_map:
                communes_to_create.append(Commune(district_id=district_map[d_code], code=c_code, name_kh=c_kh, name_en=c_en))

        with transaction.atomic():
            Commune.objects.all().delete()
            Commune.objects.bulk_create(communes_to_create, batch_size=1000)

        commune_map = {c.code: c.id for c in Commune.objects.all()}

        # 4. Villages
        ws_vill = wb['CambodiaVillagesList2025']
        villages_to_create = []
        for r in range(2, ws_vill.max_row + 1):
            c_code = str(ws_vill.cell(row=r, column=3).value or '').strip()
            v_code = str(ws_vill.cell(row=r, column=4).value or '').strip()
            v_kh = str(ws_vill.cell(row=r, column=5).value or '').strip()
            v_en = str(ws_vill.cell(row=r, column=6).value or '').strip()
            if v_kh and c_code in commune_map:
                villages_to_create.append(Village(commune_id=commune_map[c_code], code=v_code, name_kh=v_kh, name_en=v_en))

        with transaction.atomic():
            Village.objects.all().delete()
            Village.objects.bulk_create(villages_to_create, batch_size=2000)

        messages.success(request, f"🎉 ជោគជ័យ! បានទាញយក និងធ្វើបច្ចុប្បន្នភាពទិន្នន័យ (ខេត្ត: {Province.objects.count()}, ស្រុក: {District.objects.count()}, ឃុំ: {Commune.objects.count()}, ភូមិ: {Village.objects.count()}) ពី Excel រួចរាល់!")
    except Exception as e:
        messages.error(request, f"កំហុសក្នុងការ Sync ទិន្នន័យ៖ {str(e)}")

    return redirect('location_manager_view')


@login_required
@role_required(['ADMIN'])
def location_export_excel(request):
    """
    Export Cambodia Administrative Divisions (Province, District, Commune, Village) to styled Excel (.xlsx).
    Supports individual levels (with active search/filters) or 'all' levels in separate sheets.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    level = request.GET.get('level', 'province')
    search = request.GET.get('q', '').strip()
    province_id = request.GET.get('province_id', '')
    district_id = request.GET.get('district_id', '')
    commune_id = request.GET.get('commune_id', '')

    wb = openpyxl.Workbook()
    default_sheet = wb.active

    def style_sheet(ws, title, meta_text, headers, rows, header_bg='1E40AF'):
        title_font = Font(name='Khmer OS Siemreap', size=13, bold=True, color='1E3A8A')
        meta_font = Font(name='Khmer OS Siemreap', size=9, italic=True, color='475569')
        header_font = Font(name='Khmer OS Siemreap', size=9, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
        data_font = Font(name='Khmer OS Siemreap', size=9)
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        total_cols = len(headers)
        end_col = get_column_letter(total_cols)

        # Title
        ws.merge_cells(f'A1:{end_col}1')
        ws['A1'] = title
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Metadata
        ws.merge_cells(f'A2:{end_col}2')
        ws['A2'] = meta_text
        ws['A2'].font = meta_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        # Empty row 3
        ws.row_dimensions[3].height = 6

        # Header Row (Row 4)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[4].height = 24

        # Append data rows
        for row in rows:
            ws.append(row)

        # If <= 3000 rows (provinces, districts, communes, or filtered lists), apply full styling
        if len(rows) <= 3000:
            for r_idx in range(5, 5 + len(rows)):
                is_even = (r_idx % 2 == 0)
                ws.row_dimensions[r_idx].height = 20
                for c_idx in range(1, total_cols + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if is_even:
                        cell.fill = zebra_fill
                    if c_idx in [1, 2, 4, 6, 8] or isinstance(cell.value, int):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto width
        for col_idx, h in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(h))
            for r in rows[:100]:
                if col_idx - 1 < len(r):
                    max_len = max(max_len, len(str(r[col_idx - 1] or '')))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        ws.freeze_panes = 'A5'

    now_str = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')

    # Build data for Provinces
    def get_province_rows():
        district_counts = dict(District.objects.values('province_id').annotate(c=Count('id')).values_list('province_id', 'c'))
        commune_counts = dict(Commune.objects.values('district__province_id').annotate(c=Count('id')).values_list('district__province_id', 'c'))
        village_counts = dict(Village.objects.values('commune__district__province_id').annotate(c=Count('id')).values_list('commune__district__province_id', 'c'))

        qs = Province.objects.all().order_by('code')
        if search and level == 'province':
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search))
        rows = []
        for idx, p in enumerate(qs, 1):
            rows.append([
                idx,
                p.code,
                p.name_kh,
                p.name_en or '-',
                district_counts.get(p.id, 0),
                commune_counts.get(p.id, 0),
                village_counts.get(p.id, 0)
            ])
        return rows, len(rows)

    # Build data for Districts
    def get_district_rows():
        commune_counts = dict(Commune.objects.values('district_id').annotate(c=Count('id')).values_list('district_id', 'c'))
        village_counts = dict(Village.objects.values('commune__district_id').annotate(c=Count('id')).values_list('commune__district_id', 'c'))

        qs = District.objects.select_related('province').all().order_by('province__code', 'code')
        if province_id:
            qs = qs.filter(province_id=province_id)
        if search and (level == 'district' or not province_id):
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search) | Q(province__name_kh__icontains=search))
        rows = []
        for idx, d in enumerate(qs, 1):
            rows.append([
                idx,
                d.province.code,
                d.province.name_kh,
                d.code,
                d.name_kh,
                d.name_en or '-',
                commune_counts.get(d.id, 0),
                village_counts.get(d.id, 0)
            ])
        return rows, len(rows)

    # Build data for Communes
    def get_commune_rows():
        village_counts = dict(Village.objects.values('commune_id').annotate(c=Count('id')).values_list('commune_id', 'c'))

        qs = Commune.objects.select_related('district__province').all().order_by('district__province__code', 'district__code', 'code')
        if commune_id:
            qs = qs.filter(id=commune_id)
        elif district_id:
            qs = qs.filter(district_id=district_id)
        elif province_id:
            qs = qs.filter(district__province_id=province_id)
        if search and level == 'commune':
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search) | Q(district__name_kh__icontains=search) | Q(district__province__name_kh__icontains=search))
        rows = []
        for idx, c in enumerate(qs, 1):
            rows.append([
                idx,
                c.district.province.code,
                c.district.province.name_kh,
                c.district.code,
                c.district.name_kh,
                c.code,
                c.name_kh,
                c.name_en or '-',
                village_counts.get(c.id, 0)
            ])
        return rows, len(rows)

    # Build data for Villages
    def get_village_rows():
        qs = Village.objects.select_related('commune__district__province').all().order_by('commune__district__province__code', 'commune__district__code', 'commune__code', 'code')
        if commune_id:
            qs = qs.filter(commune_id=commune_id)
        elif district_id:
            qs = qs.filter(commune__district_id=district_id)
        elif province_id:
            qs = qs.filter(commune__district__province_id=province_id)
        if search and level == 'village':
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search) | Q(commune__name_kh__icontains=search) | Q(commune__district__name_kh__icontains=search) | Q(commune__district__province__name_kh__icontains=search))

        val_list = qs.values_list(
            'commune__district__province__code',
            'commune__district__province__name_kh',
            'commune__district__code',
            'commune__district__name_kh',
            'commune__code',
            'commune__name_kh',
            'code',
            'name_kh',
            'name_en'
        )
        rows = []
        for idx, v in enumerate(val_list, 1):
            rows.append([
                idx,
                v[0],
                v[1],
                v[2],
                v[3],
                v[4],
                v[5],
                v[6],
                v[7],
                v[8] or '-'
            ])
        return rows, len(rows)

    if level == 'province':
        ws = wb.create_sheet(title="ខេត្ត-រាជធានី")
        rows, count = get_province_rows()
        headers = ["ល.រ", "កូដខេត្ត", "ឈ្មោះខេត្ត/រាជធានី (ខ្មែរ)", "ឈ្មោះខេត្ត/រាជធានី (ឡាតាំង)", "ចំនួនស្រុក/ខណ្ឌ", "ចំនួនឃុំ/សង្កាត់", "ចំនួនភូមិ"]
        meta = f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុបខេត្ត/រាជធានី៖ {count}"
        style_sheet(ws, "តារាងបញ្ជីខេត្ត/រាជធានីនៃព្រះរាជាណាចក្រកម្ពុជា (Provinces List)", meta, headers, rows, header_bg='1E40AF')

    elif level == 'district':
        ws = wb.create_sheet(title="ស្រុក-ខណ្ឌ")
        rows, count = get_district_rows()
        headers = ["ល.រ", "កូដខេត្ត", "ខេត្ត/រាជធានី", "កូដស្រុក", "ឈ្មោះស្រុក/ខណ្ឌ/ក្រុង (ខ្មែរ)", "ឈ្មោះស្រុក/ខណ្ឌ/ក្រុង (ឡាតាំង)", "ចំនួនឃុំ/សង្កាត់", "ចំនួនភូមិ"]
        meta = f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុបស្រុក/ខណ្ឌ៖ {count}"
        style_sheet(ws, "តារាងបញ្ជីស្រុក/ខណ្ឌ/ក្រុងនៃព្រះរាជាណាចក្រកម្ពុជា (Districts List)", meta, headers, rows, header_bg='047857')

    elif level == 'commune':
        ws = wb.create_sheet(title="ឃុំ-សង្កាត់")
        rows, count = get_commune_rows()
        headers = ["ល.រ", "កូដខេត្ត", "ខេត្ត/រាជធានី", "កូដស្រុក", "ស្រុក/ខណ្ឌ/ក្រុង", "កូដឃុំ", "ឈ្មោះឃុំ/សង្កាត់ (ខ្មែរ)", "ឈ្មោះឃុំ/សង្កាត់ (ឡាតាំង)", "ចំនួនភូមិ"]
        meta = f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុបឃុំ/សង្កាត់៖ {count}"
        style_sheet(ws, "តារាងបញ្ជីឃុំ/សង្កាត់នៃព្រះរាជាណាចក្រកម្ពុជា (Communes List)", meta, headers, rows, header_bg='D97706')

    elif level == 'village':
        ws = wb.create_sheet(title="ភូមិ")
        rows, count = get_village_rows()
        headers = ["ល.រ", "កូដខេត្ត", "ខេត្ត/រាជធានី", "កូដស្រុក", "ស្រុក/ខណ្ឌ/ក្រុង", "កូដឃុំ", "ឃុំ/សង្កាត់", "កូដភូមិ", "ឈ្មោះភូមិ (ខ្មែរ)", "ឈ្មោះភូមិ (ឡាតាំង)"]
        meta = f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុបភូមិ៖ {count}"
        style_sheet(ws, "តារាងបញ្ជីភូមិនៃព្រះរាជាណាចក្រកម្ពុជា (Villages List)", meta, headers, rows, header_bg='0284C7')

    elif level == 'all':
        # Sheet 1: Provinces
        ws1 = wb.create_sheet(title="១. ខេត្ត-រាជធានី")
        p_rows, p_cnt = get_province_rows()
        p_headers = ["ល.រ", "កូដខេត្ត", "ឈ្មោះខេត្ត/រាជធានី (ខ្មែរ)", "ឈ្មោះខេត្ត/រាជធានី (ឡាតាំង)", "ចំនួនស្រុក/ខណ្ឌ", "ចំនួនឃុំ/សង្កាត់", "ចំនួនភូមិ"]
        style_sheet(ws1, "១. តារាងបញ្ជីខេត្ត/រាជធានីនៃព្រះរាជាណាចក្រកម្ពុជា (Provinces)", f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុប៖ {p_cnt}", p_headers, p_rows, header_bg='1E40AF')

        # Sheet 2: Districts
        ws2 = wb.create_sheet(title="២. ស្រុក-ខណ្ឌ")
        d_rows, d_cnt = get_district_rows()
        d_headers = ["ល.រ", "កូដខេត្ត", "ខេត្ត/រាជធានី", "កូដស្រុក", "ឈ្មោះស្រុក/ខណ្ឌ/ក្រុង (ខ្មែរ)", "ឈ្មោះស្រុក/ខណ្ឌ/ក្រុង (ឡាតាំង)", "ចំនួនឃុំ/សង្កាត់", "ចំនួនភូមិ"]
        style_sheet(ws2, "២. តារាងបញ្ជីស្រុក/ខណ្ឌ/ក្រុង (Districts)", f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុប៖ {d_cnt}", d_headers, d_rows, header_bg='047857')

        # Sheet 3: Communes
        ws3 = wb.create_sheet(title="៣. ឃុំ-សង្កាត់")
        c_rows, c_cnt = get_commune_rows()
        c_headers = ["ល.រ", "កូដខេត្ត", "ខេត្ត/រាជធានី", "កូដស្រុក", "ស្រុក/ខណ្ឌ/ក្រុង", "កូដឃុំ", "ឈ្មោះឃុំ/សង្កាត់ (ខ្មែរ)", "ឈ្មោះឃុំ/សង្កាត់ (ឡាតាំង)", "ចំនួនភូមិ"]
        style_sheet(ws3, "៣. តារាងបញ្ជីឃុំ/សង្កាត់ (Communes)", f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុប៖ {c_cnt}", c_headers, c_rows, header_bg='D97706')

        # Sheet 4: Villages
        ws4 = wb.create_sheet(title="៤. ភូមិ")
        v_rows, v_cnt = get_village_rows()
        v_headers = ["ល.រ", "កូដខេត្ត", "ខេត្ត/រាជធានី", "កូដស្រុក", "ស្រុក/ខណ្ឌ/ក្រុង", "កូដឃុំ", "ឃុំ/សង្កាត់", "កូដភូមិ", "ឈ្មោះភូមិ (ខ្មែរ)", "ឈ្មោះភូមិ (ឡាតាំង)"]
        style_sheet(ws4, "៤. តារាងបញ្ជីភូមិ (Villages)", f"កាលបរិច្ឆេទ Export៖ {now_str} | សរុប៖ {v_cnt}", v_headers, v_rows, header_bg='0284C7')

    # Remove default sheet
    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    filename = f"Cambodia_Locations_{level}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@role_required(['ADMIN'])
def location_export_csv(request):
    """
    Export Cambodia Administrative Divisions to UTF-8 CSV (with BOM).
    """
    import csv

    level = request.GET.get('level', 'province')
    search = request.GET.get('q', '').strip()
    province_id = request.GET.get('province_id', '')
    district_id = request.GET.get('district_id', '')
    commune_id = request.GET.get('commune_id', '')

    filename = f"Cambodia_Locations_{level}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # UTF-8 BOM

    writer = csv.writer(response)

    if level == 'province':
        writer.writerow(["ល.រ", "កូដខេត្ត", "ឈ្មោះខេត្ត_ខ្មែរ", "ឈ្មោះខេត្ត_ឡាតាំង", "ចំនួនស្រុក", "ចំនួនឃុំ", "ចំនួនភូមិ"])
        district_counts = dict(District.objects.values('province_id').annotate(c=Count('id')).values_list('province_id', 'c'))
        commune_counts = dict(Commune.objects.values('district__province_id').annotate(c=Count('id')).values_list('district__province_id', 'c'))
        village_counts = dict(Village.objects.values('commune__district__province_id').annotate(c=Count('id')).values_list('commune__district__province_id', 'c'))

        qs = Province.objects.all().order_by('code')
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search))
        for idx, p in enumerate(qs, 1):
            writer.writerow([idx, p.code, p.name_kh, p.name_en or '', district_counts.get(p.id, 0), commune_counts.get(p.id, 0), village_counts.get(p.id, 0)])

    elif level == 'district':
        writer.writerow(["ល.រ", "កូដខេត្ត", "ខេត្ត_ខ្មែរ", "កូដស្រុក", "ឈ្មោះស្រុក_ខ្មែរ", "ឈ្មោះស្រុក_ឡាតាំង", "ចំនួនឃុំ", "ចំនួនភូមិ"])
        commune_counts = dict(Commune.objects.values('district_id').annotate(c=Count('id')).values_list('district_id', 'c'))
        village_counts = dict(Village.objects.values('commune__district_id').annotate(c=Count('id')).values_list('commune__district_id', 'c'))

        qs = District.objects.select_related('province').all().order_by('province__code', 'code')
        if province_id:
            qs = qs.filter(province_id=province_id)
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search) | Q(province__name_kh__icontains=search))
        for idx, d in enumerate(qs, 1):
            writer.writerow([idx, d.province.code, d.province.name_kh, d.code, d.name_kh, d.name_en or '', commune_counts.get(d.id, 0), village_counts.get(d.id, 0)])

    elif level == 'commune':
        writer.writerow(["ល.រ", "កូដខេត្ត", "ខេត្ត_ខ្មែរ", "កូដស្រុក", "ស្រុក_ខ្មែរ", "កូដឃុំ", "ឈ្មោះឃុំ_ខ្មែរ", "ឈ្មោះឃុំ_ឡាតាំង", "ចំនួនភូមិ"])
        village_counts = dict(Village.objects.values('commune_id').annotate(c=Count('id')).values_list('commune_id', 'c'))

        qs = Commune.objects.select_related('district__province').all().order_by('district__province__code', 'district__code', 'code')
        if commune_id:
            qs = qs.filter(id=commune_id)
        elif district_id:
            qs = qs.filter(district_id=district_id)
        elif province_id:
            qs = qs.filter(district__province_id=province_id)
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search) | Q(district__name_kh__icontains=search) | Q(district__province__name_kh__icontains=search))
        for idx, c in enumerate(qs, 1):
            writer.writerow([idx, c.district.province.code, c.district.province.name_kh, c.district.code, c.district.name_kh, c.code, c.name_kh, c.name_en or '', village_counts.get(c.id, 0)])

    elif level == 'village':
        writer.writerow(["ល.រ", "កូដខេត្ត", "ខេត្ត_ខ្មែរ", "កូដស្រុក", "ស្រុក_ខ្មែរ", "កូដឃុំ", "ឃុំ_ខ្មែរ", "កូដភូមិ", "ឈ្មោះភូមិ_ខ្មែរ", "ឈ្មោះភូមិ_ឡាតាំង"])
        qs = Village.objects.select_related('commune__district__province').all().order_by('commune__district__province__code', 'commune__district__code', 'commune__code', 'code')
        if commune_id:
            qs = qs.filter(commune_id=commune_id)
        elif district_id:
            qs = qs.filter(commune__district_id=district_id)
        elif province_id:
            qs = qs.filter(commune__district__province_id=province_id)
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search) | Q(commune__name_kh__icontains=search) | Q(commune__district__name_kh__icontains=search) | Q(commune__district__province__name_kh__icontains=search))
        
        val_list = qs.values_list(
            'commune__district__province__code',
            'commune__district__province__name_kh',
            'commune__district__code',
            'commune__district__name_kh',
            'commune__code',
            'commune__name_kh',
            'code',
            'name_kh',
            'name_en'
        )
        for idx, v in enumerate(val_list, 1):
            writer.writerow([idx, v[0], v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8] or ''])

    elif level in ['all', 'full_flat']:
        writer.writerow(["ល.រ", "កូដខេត្ត", "ខេត្ត_ខ្មែរ", "ខេត្ត_ឡាតាំង", "កូដស្រុក", "ស្រុក_ខ្មែរ", "ស្រុក_ឡាតាំង", "កូដឃុំ", "ឃុំ_ខ្មែរ", "ឃុំ_ឡាតាំង", "កូដភូមិ", "ភូមិ_ខ្មែរ", "ភូមិ_ឡាតាំង"])
        qs = Village.objects.select_related('commune__district__province').all().order_by('commune__district__province__code', 'commune__district__code', 'commune__code', 'code')
        if commune_id:
            qs = qs.filter(commune_id=commune_id)
        elif district_id:
            qs = qs.filter(commune__district_id=district_id)
        elif province_id:
            qs = qs.filter(commune__district__province_id=province_id)
        if search:
            qs = qs.filter(Q(name_kh__icontains=search) | Q(name_en__icontains=search) | Q(code__icontains=search) | Q(commune__name_kh__icontains=search) | Q(commune__district__name_kh__icontains=search) | Q(commune__district__province__name_kh__icontains=search))
        
        val_list = qs.values_list(
            'commune__district__province__code',
            'commune__district__province__name_kh',
            'commune__district__province__name_en',
            'commune__district__code',
            'commune__district__name_kh',
            'commune__district__name_en',
            'commune__code',
            'commune__name_kh',
            'commune__name_en',
            'code',
            'name_kh',
            'name_en'
        )
        for idx, v in enumerate(val_list, 1):
            writer.writerow([
                idx,
                v[0],
                v[1],
                v[2] or '',
                v[3],
                v[4],
                v[5] or '',
                v[6],
                v[7],
                v[8] or '',
                v[9],
                v[10],
                v[11] or ''
            ])

    return response
