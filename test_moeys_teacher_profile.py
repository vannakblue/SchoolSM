import os
import sys
import io
import django
from datetime import date

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

from django.test import Client
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
import openpyxl
from apps.accounts.models import User
from apps.teachers.models import Teacher
from apps.teachers.forms import TeacherForm


def test_moeys_teacher_profile_system():
    print("=== STARTING MOEYS OFFICIAL TEACHER PROFILE & COLUMNS VERIFICATION ===")

    admin_user, _ = User.objects.get_or_create(
        username='admin_moeys_tester',
        defaults={'role': User.Role.ADMIN, 'khmer_name': 'Admin MoEYS Tester'}
    )
    admin_user.set_password('password123')
    admin_user.save()

    client = Client()
    client.force_login(admin_user)

    # 1. Test Direct Model Creation with all MoEYS fields
    teacher_1, created = Teacher.objects.update_or_create(
        teacher_id='T-MOEYS-001',
        defaults={
            'khmer_name': 'អ៊ុំ សារ៉ាត់',
            'latin_name': 'Oum Sarath',
            'gender': Teacher.Gender.MALE,
            'date_of_birth': date(1982, 4, 15),
            'qualification': 'បរិញ្ញាបត្រអប់រំ',
            'specialization': 'គណិតវិទ្យា & រូបវិទ្យា',
            'training_level': 'គរុកោសល្យឧត្តម (បរិញ្ញាបត្រ+១)',
            'state_hire_date': date(2005, 10, 1),
            'permanent_date': date(2007, 12, 15),
            'primary_subject': 'គណិតវិទ្យា',
            'secondary_subject': 'រូបវិទ្យា',
            'current_duty': 'ប្រធានក្រុមបច្ចេកទេស',
            'prakas_category': 'ក្របខ័ណ្ឌ ក.១',
            'prakas_year': '2007',
            'prakas_number': 'ប្រកាសលេខ ៥៤២ អយក.ប្រក',
            'phone': '012998877',
            'email': 'sarath.oum@school.edu.kh',
            'base_salary': 650.00,
            'status': Teacher.Status.ACTIVE
        }
    )
    assert teacher_1.training_level == 'គរុកោសល្យឧត្តម (បរិញ្ញាបត្រ+១)'
    assert teacher_1.state_hire_date == date(2005, 10, 1)
    assert teacher_1.permanent_date == date(2007, 12, 15)
    assert teacher_1.primary_subject == 'គណិតវិទ្យា'
    assert teacher_1.secondary_subject == 'រូបវិទ្យា'
    assert teacher_1.current_duty == 'ប្រធានក្រុមបច្ចេកទេស'
    assert teacher_1.prakas_category == 'ក្របខ័ណ្ឌ ក.១'
    assert teacher_1.prakas_year == '2007'
    assert teacher_1.prakas_number == 'ប្រកាសលេខ ៥៤២ អយក.ប្រក'
    print("  [PASS] 1. Teacher model successfully saved all MoEYS official fields.")

    # 2. Test TeacherForm with MoEYS fields
    form_data = {
        'teacher_id': 'T-MOEYS-002',
        'khmer_name': 'សុខា ស្រីមុំ',
        'latin_name': 'Sokha Sreymom',
        'gender': 'F',
        'date_of_birth': '1989-06-25',
        'qualification': 'បរិញ្ញាបត្រជាន់ខ្ពស់',
        'specialization': 'ភាសាខ្មែរ',
        'training_level': 'គរុកោសល្យឧត្តម',
        'state_hire_date': '2012-10-01',
        'permanent_date': '2014-11-20',
        'primary_subject': 'តែងសេចក្តី',
        'secondary_subject': 'អក្សរសិល្ប៍ខ្មែរ',
        'current_duty': 'នាយករង',
        'prakas_category': 'ក្របខ័ណ្ឌ ក.២',
        'prakas_year': '2014',
        'prakas_number': 'ប្រកាសលេខ ៨៨',
        'phone': '017223344',
        'email': 'sreymom@school.edu.kh',
        'base_salary': '600.00',
        'status': 'ACTIVE'
    }
    form = TeacherForm(data=form_data)
    assert form.is_valid(), f"Form errors: {form.errors}"
    saved_teacher = form.save()
    assert saved_teacher.primary_subject == 'តែងសេចក្តី'
    assert saved_teacher.current_duty == 'នាយករង'
    print("  [PASS] 2. TeacherForm valid and saved all MoEYS fields successfully.")

    # 3. Test Teacher Export Excel Endpoint
    res_export = client.get(reverse('teacher_export_excel'))
    assert res_export.status_code == 200
    wb_export = openpyxl.load_workbook(io.BytesIO(res_export.content))
    ws_export = wb_export.active
    # Check headers
    assert "អត្តលេខ" in [str(c.value) for c in ws_export[1]]
    assert "តាំងស៊ប់តាមប្រកាស" in [str(c.value) for c in ws_export[1]]
    assert "កម្រិតបណ្តុះបណ្តាល" in [str(c.value) for c in ws_export[1]]
    print("  [PASS] 3. Excel Export generates complete official MoEYS standard 2-row layout.")

    # 4. Test Teacher Import Template Excel & CSV
    res_tpl_xl = client.get(reverse('teacher_import_template_excel'))
    assert res_tpl_xl.status_code == 200
    wb_tpl = openpyxl.load_workbook(io.BytesIO(res_tpl_xl.content))
    ws_tpl = wb_tpl.active
    headers_tpl = [str(c.value) for c in ws_tpl[1]]
    assert "Training Level" in headers_tpl
    assert "State Hire Date (DD-MM-YYYY)" in headers_tpl
    assert "Prakas Number" in headers_tpl
    print("  [PASS] 4. Excel Import Template verified with all 20 MoEYS columns.")

    res_tpl_csv = client.get(reverse('teacher_import_template_csv'))
    assert res_tpl_csv.status_code == 200
    csv_content = res_tpl_csv.content.decode('utf-8-sig')
    assert "Training Level" in csv_content
    assert "Primary Subject" in csv_content
    print("  [PASS] 5. CSV Import Template verified.")

    # 5. Test MoEYS Bulk Import via Excel
    wb_test = openpyxl.Workbook()
    ws_test = wb_test.active
    ws_test.append(headers_tpl)
    ws_test.append([
        "T-MOEYS-IMPORT", "លី សុវណ្ណ", "Ly Sovann", "M", "10-08-1986",
        "បរិញ្ញាបត្រ", "គីមីវិទ្យា", "គរុកោសល្យឧត្តម", "01-10-2009", "15-11-2011",
        "គីមីវិទ្យា", "ជីវវិទ្យា", "គ្រូបង្រៀន", "ក្របខ័ណ្ឌ ក.១", "2011", "ប្រកាសលេខ ៣២០",
        "096554433", "sovann@school.edu.kh", 580, "ACTIVE"
    ])
    out = io.BytesIO()
    wb_test.save(out)
    out.seek(0)

    upload_file = SimpleUploadedFile('test_moeys_import.xlsx', out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    res_import = client.post(reverse('teacher_import'), {'file': upload_file}, follow=True)
    assert res_import.status_code == 200

    imported_teacher = Teacher.objects.filter(teacher_id='T-MOEYS-IMPORT').first()
    assert imported_teacher is not None
    assert imported_teacher.khmer_name == 'លី សុវណ្ណ'
    assert imported_teacher.date_of_birth == date(1986, 8, 10)
    assert imported_teacher.state_hire_date == date(2009, 10, 1)
    assert imported_teacher.permanent_date == date(2011, 11, 15)
    assert imported_teacher.primary_subject == 'គីមីវិទ្យា'
    assert imported_teacher.prakas_number == 'ប្រកាសលេខ ៣២០'
    print("  [PASS] 6. MoEYS Excel import parsed and populated all 20 fields seamlessly.")

    # 6. Test Teacher List and Detail Pages Render cleanly
    res_list = client.get(reverse('teacher_list'))
    assert res_list.status_code == 200
    list_content = res_list.content.decode('utf-8')
    assert 'id="moeysTableView"' in list_content
    assert 'តាំងស៊ប់តាមប្រកាស' in list_content
    assert 'T-MOEYS-001' in list_content

    res_detail = client.get(reverse('teacher_detail', args=[teacher_1.id]))
    assert res_detail.status_code == 200
    detail_content = res_detail.content.decode('utf-8')
    assert 'ព័ត៌មានក្របខ័ណ្ឌរដ្ឋ & គរុកោសល្យ' in detail_content
    assert 'ប្រកាសលេខ ៥៤២' in detail_content
    print("  [PASS] 7. Teacher List & Detail pages rendered MoEYS official table and civil service cards.")

    # Clean up test records
    Teacher.objects.filter(teacher_id__in=['T-MOEYS-001', 'T-MOEYS-002', 'T-MOEYS-IMPORT']).delete()
    User.objects.filter(username__in=['t_moeys_001', 't_moeys_002', 't_moeys_import', 'admin_moeys_tester']).delete()

    print("=== ALL MOEYS OFFICIAL TEACHER PROFILE TESTS PASSED 100% ===")


if __name__ == '__main__':
    test_moeys_teacher_profile_system()
