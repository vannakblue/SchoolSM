import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from apps.teachers.models import Teacher
from apps.academics.models import Classroom, Subject, AcademicYear

file_path = r'E:\SchoolSM\Data.xlsx'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"Sheets in Data.xlsx: {wb.sheetnames}")

if 'duty' in wb.sheetnames:
    ws = wb['duty']
    print(f"\n=== SHEET: 'duty' (Rows: {ws.max_row}, Cols: {ws.max_column}) ===")
    for idx, row in enumerate(list(ws.iter_rows(values_only=True))[:30]):
        non_empty = [c for c in row if c is not None]
        if non_empty:
            print(f"Row {idx+1:2d}: {row[:15]}")
else:
    print("Sheet 'duty' not found! Checking case-insensitive matches...")
    for s in wb.sheetnames:
        if 'duty' in s.lower():
            print(f"Found match: '{s}'")

print("\n=== ALL TEACHERS IN DATABASE ===")
teachers = list(Teacher.objects.all().order_by('id'))
print(f"Total Teachers in DB: {len(teachers)}")
for t in teachers:
    print(f"ID: {t.id:2d} | Code: '{t.code:8s}' | Name: '{t.khmer_name:25s}' | Latin: '{t.latin_name or '':20s}' | Gender: {t.gender} | Phone: {t.phone_number or ''}")
