import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl

file_path = r'E:\SchoolSM\2026-2027.xlsm'
print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"Sheet names in 2026-2027.xlsm ({len(wb.sheetnames)} sheets):")
print(wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    non_empty_rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    print(f"\n--- Sheet: '{sname}' (Total Rows: {len(rows)}, Non-empty: {len(non_empty_rows)}) ---")
    for r in non_empty_rows[:6]:
        # truncate long cells
        print("  ", [str(c)[:25] if c is not None else '' for c in r[:12]])
