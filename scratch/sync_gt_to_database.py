import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath('.'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import datetime
import openpyxl
from django.db import transaction
from apps.academics.models import AcademicYear, Classroom, Subject, Timetable, ClassSubject
from apps.teachers.models import Teacher
from apps.academics.views import STANDARD_PERIOD_TIMES

file_path = r'E:\SchoolSM\Data.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws_duty = wb['duty']
ws_gt = wb['GT']

# 1. Map symbols to teachers and subjects
def get_subject_code_from_symbol(sym):
    sym_upper = sym.upper().strip()
    if sym_upper.startswith('ED'):
        return 'ED'
    elif sym_upper.startswith('AG'):
        return 'AG'
    elif sym_upper.startswith('ES'):
        return 'Es'
    elif sym_upper.startswith('EC'):
        return 'Ec'
    elif sym_upper.startswith('HE'):
        return 'He'
    elif sym_upper.startswith('COM') or sym_upper.startswith('IT'):
        return 'IT'
    elif sym_upper.startswith('M'):
        return 'M'
    elif sym_upper.startswith('P'):
        return 'P'
    elif sym_upper.startswith('C'):
        return 'C'
    elif sym_upper.startswith('B'):
        return 'B'
    elif sym_upper.startswith('H'):
        return 'H'
    elif sym_upper.startswith('G'):
        return 'G'
    elif sym_upper.startswith('I'):
        return 'I'
    elif sym_upper.startswith('K'):
        return 'K'
    elif sym_upper.startswith('E'):
        return 'E'
    return None

# Load teacher map from duty sheet and database
symbol_to_teacher = {}
all_teachers = list(Teacher.objects.all())
teachers_by_name = {t.khmer_name.strip(): t for t in all_teachers}

for r in list(ws_duty.iter_rows(values_only=True))[1:]:
    sym = str(r[1]).strip() if r[1] is not None else ''
    name = str(r[2]).strip() if r[2] is not None else ''
    if sym and name:
        t_obj = teachers_by_name.get(name)
        if t_obj:
            symbol_to_teacher[sym.upper()] = t_obj

print(f"Mapped {len(symbol_to_teacher)} symbols to active DB teachers.")

# Subjects map
subjects_map = {s.code.upper().strip(): s for s in Subject.objects.all()}
for s in Subject.objects.all():
    subjects_map[s.code] = s

# Target Academic Year
ay = AcademicYear.objects.filter(name='2026-2027').first() or AcademicYear.objects.filter(is_current=True).first()
classrooms_map = {c.code.upper().strip(): c for c in Classroom.objects.filter(academic_year=ay)}

day_configs = [
    # (DayNumber, DayName, [period_cols], room_col_idx)
    (1, 'Monday', [2, 3, 4, 5, 7, 8, 9, 10], 6),
    (2, 'Tuesday', [12, 13, 14, 15, 17, 18, 19, 20], 16),
    (3, 'Wednesday', [22, 23, 24, 25, 27, 28, 29, 30], 26),
    (4, 'Thursday', [32, 33, 34, 35, 37, 38, 39, 40], 36),
    (5, 'Friday', [42, 43, 44, 45, 47, 48, 49, 50], 46),
    (6, 'Saturday', [52, 53, 54, 55, 57, 58, 59, 60], 56),
]

timetable_entries_to_create = []
unresolved_slots = []
total_slots_read = 0

for r_idx, row in enumerate(list(ws_gt.iter_rows(values_only=True))[1:], start=2):
    class_code = str(row[0]).strip().upper() if row[0] is not None else ''
    classroom = classrooms_map.get(class_code)
    if not classroom:
        continue

    for day_num, day_name, p_cols, room_col in day_configs:
        room_val = str(row[room_col - 1]).strip() if (room_col - 1 < len(row) and row[room_col - 1] is not None) else ''
        for p_idx, col_num in enumerate(p_cols, start=1):
            if col_num - 1 < len(row):
                cell_val = str(row[col_num - 1]).strip() if row[col_num - 1] is not None else ''
                if not cell_val:
                    continue

                total_slots_read += 1
                sym_upper = cell_val.upper()
                teacher = symbol_to_teacher.get(sym_upper)
                sub_code = get_subject_code_from_symbol(cell_val)
                subject = subjects_map.get(sub_code.upper()) if sub_code else None

                if not teacher:
                    # Try finding teacher from specialization or duty
                    for t in all_teachers:
                        if t.current_duty and t.current_duty.upper() == sym_upper:
                            teacher = t
                            break

                if not subject and teacher and teacher.primary_subject:
                    subject = teacher.primary_subject

                st_time, et_time = STANDARD_PERIOD_TIMES.get(p_idx, (datetime.time(7, 0), datetime.time(7, 50)))

                if classroom and subject and teacher:
                    entry = Timetable(
                        classroom=classroom,
                        subject=subject,
                        teacher=teacher,
                        day_of_week=day_num,
                        period_number=p_idx,
                        start_time=st_time,
                        end_time=et_time,
                        room=room_val or classroom.room_number or ''
                    )
                    timetable_entries_to_create.append(entry)
                else:
                    unresolved_slots.append({
                        'class': class_code,
                        'day': day_name,
                        'period': p_idx,
                        'symbol': cell_val,
                        'teacher': teacher.khmer_name if teacher else 'Missing',
                        'subject': subject.name_kh if subject else 'Missing'
                    })

print(f"Read {total_slots_read} timetable slots from sheet 'GT'.")
print(f"Valid timetable entries resolved: {len(timetable_entries_to_create)}")
print(f"Unresolved slots: {len(unresolved_slots)}")
if unresolved_slots:
    print("Sample unresolved slots:", unresolved_slots[:10])

# Clean and recreate timetable entries for 2026-2027
with transaction.atomic():
    Timetable.objects.filter(classroom__academic_year=ay).delete()
    # Bulk create entries in batches
    Timetable.objects.bulk_create(timetable_entries_to_create, batch_size=500)

print(f"Successfully synced {len(timetable_entries_to_create)} timetable entries into Academic Year '{ay.name}'!")
