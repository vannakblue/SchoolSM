import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath('.'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from apps.students.views import _normalize_header

wb = openpyxl.load_workbook(r'E:\SchoolSM\2026-2027.xlsm', data_only=True)
for s_name in wb.sheetnames:
    ws = wb[s_name]
    print(f"\n=== Sheet '{s_name}' ===")
    found_header = False
    header_map = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not found_header:
            row_str = ' '.join([str(c) for c in row if c is not None])
            if 'ឈ្មោះ' in row_str or 'name' in row_str.lower() or 'អត្តលេខ' in row_str:
                header_map = [_normalize_header(c) for c in row]
                found_header = True
                print(f"Header found at Row {r_idx}: {header_map}")
        else:
            if any(row):
                if r_idx <= 6:
                    print(f"Row {r_idx}: {[str(c) if c is not None else '' for c in row]}")
