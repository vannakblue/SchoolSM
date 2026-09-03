import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from apps.teachers.models import Teacher

file_path = r'E:\SchoolSM\Data.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['duty']

db_teachers = list(Teacher.objects.all().order_by('khmer_name'))
print(f"=== ALL {len(db_teachers)} TEACHERS IN DATABASE (Sorted by Khmer Name) ===")
for t in db_teachers:
    print(f"ID: {t.id:3d} | TID: '{t.teacher_id:12s}' | Name: '{t.khmer_name:25s}' | Spec: '{t.specialization:20s}' | Duty: '{t.current_duty or ''}'")

print(f"\n=== ALL ROWS IN SHEET 'duty' ({ws.max_row} rows) ===")
for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if r_idx == 1:
        continue
    no = row[0]
    symbol = str(row[1]).strip() if row[1] is not None else ''
    name = str(row[2]).strip() if row[2] is not None else ''
    gender = str(row[3]).strip() if row[3] is not None else ''
    classes = str(row[4]).strip() if row[4] is not None else ''
    if symbol or name:
        print(f"Row {r_idx:3d} | No: {str(no):3s} | Symbol: {symbol:8s} | Name: '{name:25s}' | G: {gender:2s} | Classes: '{classes}'")
