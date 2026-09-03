import os, sys
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

wb = openpyxl.load_workbook(r'E:\SchoolSM\Data.xlsx', data_only=True)
ws = wb['GT']

print(f"Max Columns in GT: {ws.max_column}")
headers = [str(c) if c is not None else '' for c in list(ws.iter_rows(values_only=True))[0]]
print("Headers in GT:")
for idx, h in enumerate(headers, start=1):
    print(f"Col {idx:2d}: '{h}'")
