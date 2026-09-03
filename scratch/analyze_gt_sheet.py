import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath('.'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from apps.teachers.models import Teacher

wb = openpyxl.load_workbook(r'E:\SchoolSM\Data.xlsx', data_only=True)
ws_duty = wb['duty']
ws_gt = wb['GT']

# Collect duty symbols and teacher mapping
duty_symbols = {}
for r in list(ws_duty.iter_rows(values_only=True))[1:]:
    sym = str(r[1]).strip() if r[1] is not None else ''
    name = str(r[2]).strip() if r[2] is not None else ''
    classes = str(r[4]).strip() if r[4] is not None else ''
    if sym:
        duty_symbols[sym.upper()] = {
            'symbol': sym,
            'name': name,
            'classes': classes
        }

print(f"Total symbols in sheet 'duty': {len(duty_symbols)}")
print("Duty symbols:", sorted(duty_symbols.keys()))

# Analyze all cells in GT
gt_symbols = set()
day_cols = [
    # (DayName, [period_cols])
    ('Monday', [2, 3, 4, 5, 7, 8, 9, 10]),
    ('Tuesday', [12, 13, 14, 15, 17, 18, 19, 20]),
    ('Wednesday', [22, 23, 24, 25, 27, 28, 29, 30]),
    ('Thursday', [32, 33, 34, 35, 37, 38, 39, 40]),
    ('Friday', [42, 43, 44, 45, 47, 48, 49, 50]),
    ('Saturday', [52, 53, 54, 55, 57, 58, 59, 60]),
]

unknown_symbols = {}
known_symbols = {}

for r_idx, row in enumerate(list(ws_gt.iter_rows(values_only=True))[1:], start=2):
    class_code = str(row[0]).strip() if row[0] is not None else ''
    if not class_code:
        continue
    
    for day_name, p_cols in day_cols:
        for p_idx, col_num in enumerate(p_cols, start=1):
            if col_num - 1 < len(row):
                cell_val = str(row[col_num - 1]).strip() if row[col_num - 1] is not None else ''
                if cell_val:
                    gt_symbols.add(cell_val)
                    if cell_val.upper() in duty_symbols:
                        known_symbols[cell_val] = known_symbols.get(cell_val, 0) + 1
                    else:
                        if cell_val not in unknown_symbols:
                            unknown_symbols[cell_val] = []
                        unknown_symbols[cell_val].append(f"{class_code} {day_name} P{p_idx}")

print(f"\nTotal distinct symbols in GT: {len(gt_symbols)}")
print(f"Known symbols matching duty sheet: {len(known_symbols)}")
print(f"Unknown / Unmatched symbols in GT: {len(unknown_symbols)}")
print("\n=== UNMATCHED SYMBOLS IN GT ===")
for sym, occurrences in sorted(unknown_symbols.items()):
    print(f"Symbol: '{sym:15s}' | Count: {len(occurrences):2d} | Occurrences: {occurrences[:5]}")
