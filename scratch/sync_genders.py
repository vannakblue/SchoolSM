import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)
sys.stdout.reconfigure(encoding='utf-8')

import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from apps.students.models import Student
from apps.examinations.models import StandardizedExam, ExamCandidate

def main():
    excel_path = 'e:/SchoolSM/2026-2027.xlsm'
    print(f"Reading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheets = ['7', '8', '9', '10', '11', '12']

    female_ids = set()
    male_ids = set()

    for sname in sheets:
        ws = wb[sname]
        for r in range(4, ws.max_row + 1):
            stu_id = ws.cell(row=r, column=2).value
            name = ws.cell(row=r, column=3).value
            gender_raw = ws.cell(row=r, column=4).value
            if not name or str(name).strip() == '':
                continue
            g_val = str(gender_raw).strip() if gender_raw else ''
            sid = str(stu_id).strip() if stu_id else None
            if sid:
                if g_val in ['ស', 'ស.', 'ស្រី', 'ស្រី្ត', 'ស្ត្រី', 'កញ្ញា', 'f', 'F', 'female', 'Female']:
                    female_ids.add(sid)
                else:
                    male_ids.add(sid)

    print(f"Total female IDs in Excel: {len(female_ids)}")
    print(f"Total male IDs in Excel: {len(male_ids)}")

    # Detailed mapping
    female_stu_pks = set()
    male_stu_pks = set()

    for sname in sheets:
        ws = wb[sname]
        for r in range(4, ws.max_row + 1):
            stu_id = ws.cell(row=r, column=2).value
            name = ws.cell(row=r, column=3).value
            gender_raw = ws.cell(row=r, column=4).value
            if not name or str(name).strip() == '':
                continue
            name_str = str(name).strip()
            sid_str = str(stu_id).strip() if stu_id else None
            g_val = str(gender_raw).strip() if gender_raw else ''
            is_female = g_val in ['ស', 'ស.', 'ស្រី', 'ស្រី្ត', 'ស្ត្រី', 'កញ្ញា', 'f', 'F', 'female', 'Female']

            # Find matching student
            stu = None
            if sid_str and name_str:
                stu = Student.objects.filter(student_id=sid_str, khmer_name=name_str).first()
            if not stu and name_str:
                stu = Student.objects.filter(khmer_name=name_str).first()
            if not stu and sid_str:
                # Only use ID if name is empty or matches
                potential = Student.objects.filter(student_id=sid_str).first()
                if potential and (not name_str or potential.khmer_name == name_str):
                    stu = potential

            if stu:
                if is_female:
                    female_stu_pks.add(stu.pk)
                else:
                    male_stu_pks.add(stu.pk)

    print(f"Matched female student PKs: {len(female_stu_pks)}")
    print(f"Matched male student PKs: {len(male_stu_pks)}")

    # 1. Update Student table
    f_updated = Student.objects.filter(pk__in=female_stu_pks).update(gender='F')
    m_updated = Student.objects.filter(pk__in=male_stu_pks).update(gender='M')
    print(f"Student table updated: {f_updated} females, {m_updated} males.")

    # 2. Update ExamCandidate table
    cands_with_student = ExamCandidate.objects.filter(student__isnull=False)
    cand_f_from_stu = 0
    cand_m_from_stu = 0
    for cand in cands_with_student.select_related('student'):
        if cand.gender != cand.student.gender:
            cand.gender = cand.student.gender
            cand.save(update_fields=['gender'])
            if cand.gender == 'F':
                cand_f_from_stu += 1
            else:
                cand_m_from_stu += 1

    # Candidates linked via student FK
    ExamCandidate.objects.filter(student_id__in=female_stu_pks).update(gender='F')
    ExamCandidate.objects.filter(student_id__in=male_stu_pks).update(gender='M')

    print(f"ExamCandidate sync complete.")

    # 3. Print verification for all standardized exams
    print("\n--- Verification for November monthly exams ---")
    for ex in StandardizedExam.objects.filter(name__icontains='វិច្ឆិកា').order_by('grade_level'):
        total = ex.candidates.count()
        females = ex.candidates.filter(gender='F').count()
        males = ex.candidates.filter(gender='M').count()
        print(f"Grade {ex.grade_level:02d} (Exam {ex.id:03d}): {ex.name} -> Total: {total}, Female: {females}, Male: {males}")

if __name__ == '__main__':
    main()
