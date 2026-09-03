import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from datetime import datetime
from collections import Counter

file_path = r'E:\SchoolSM\2026-2027.xlsm'
wb = openpyxl.load_workbook(file_path, data_only=True)

print("=== DETAILED ANALYSIS OF E:\\SchoolSM\\2026-2027.xlsm ===")

grand_total = 0
sheet_summaries = {}

for sname in ['7', '8', '9', '10', '11', '12']:
    if sname not in wb.sheetnames:
        continue
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    
    header = rows[0]
    data_rows = rows[1:]
    
    class_counter = Counter()
    students_in_sheet = 0
    
    for r in data_rows:
        if not r or not any(r):
            continue
        st_id = str(r[1]).strip() if r[1] is not None else ''
        st_name = str(r[2]).strip() if r[2] is not None else ''
        if not st_id or not st_name or st_id.lower() in ['none', '', 'ល.រ', 'អត្តលេខ']:
            continue
        
        gender_raw = str(r[3]).strip() if r[3] is not None else ''
        dob_raw = r[4]
        grade_raw = str(r[5]).strip() if r[5] is not None else sname
        class_letter = str(r[6]).strip() if r[6] is not None else ''
        
        class_code = f"{grade_raw}{class_letter}".upper().replace(' ', '')
        class_counter[class_code] += 1
        students_in_sheet += 1
        grand_total += 1
        
    print(f"\n--- Sheet '{sname}' (Total Valid Students: {students_in_sheet}) ---")
    for ccode, count in sorted(class_counter.items()):
        print(f"    Class {ccode:6s}: {count:3d} students")

print(f"\n========================================================")
print(f"GRAND TOTAL STUDENTS IN 2026-2027.xlsm: {grand_total}")
print(f"========================================================")
