import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import re
import openpyxl
from apps.teachers.models import Teacher
from apps.academics.models import AcademicYear, Classroom, Subject, ClassSubject

file_path = r'E:\SchoolSM\Data.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['duty']

# 1. Subject code mapping by symbol prefix
def get_subject_code_from_symbol(sym):
    sym_upper = sym.upper().strip()
    if sym_upper.startswith('ED'):
        return 'ED'
    elif sym_upper.startswith('AG'):
        return 'AG'
    elif sym_upper.startswith('ES'):
        return 'Es'
    elif sym_upper.startswith('EC'):
        return 'Ec'
    elif sym_upper.startswith('HE'):
        return 'He'
    elif sym_upper.startswith('M'):
        return 'M'
    elif sym_upper.startswith('P'):
        return 'P'
    elif sym_upper.startswith('C'):
        return 'C'
    elif sym_upper.startswith('B'):
        return 'B'
    elif sym_upper.startswith('H'):
        return 'H'
    elif sym_upper.startswith('G'):
        return 'G'
    elif sym_upper.startswith('I'):
        return 'I'
    elif sym_upper.startswith('K'):
        return 'K'
    elif sym_upper.startswith('E'):
        return 'E'
    return None

def parse_class_tokens(classes_str):
    if not classes_str or not classes_str.strip():
        return []
    pattern = re.compile(r'(\d+)([A-Za-z]+)')
    matches = pattern.findall(classes_str)
    result = []
    for grade_num, letters in matches:
        for letter in letters.upper():
            result.append(f"{grade_num}{letter}")
    return result

MANUAL_REPLACEMENTS = {
    'ផាត ស៊្រុន': 'ផាត់ ស្រ៊ុន',
    'សុង ភស្ស': 'សុង ភ័ស',
    'សាន់ ភឿន': 'សាន់ កឿន',
    'ជួង សុភក្រ័': 'ជួង សុភ័ក្រ',
    'បូរ កញ្ញា': 'បូ កញ្ញា',
    'អឹម សំអុល': 'អ៊ឹម សំអុល',
    'ចេង ប៊ុណ្ណវេទ': 'ចេង បុណ្ណវេទ',
    'ជួ សូរិយា': 'ជួ សូរីយា',
    'សួន ស្រីរត្ត័': 'សួន ស្រីរ័ត្ន',
    'ឃុត  បូរ៉ាមី': 'ឃុត បូរាមី',
    'ឃុត\u200b បូរ៉ាមី': 'ឃុត បូរាមី',
    'ជឹង សុចាន់': 'ជឹង សុខចាន់',
    'មាស ស្រីល័ក្ខ': 'មាស ស្រីលក្ខ័',
    'លន ស្រីល័ក្ខ': 'លន ស្រីលក្ខ័',
    'ទិត សាម៉នវីរ:': 'ទិត សោម៉នវីរៈ',
    'ប្រាក់ សុភារ:': 'ប្រាក់ សុភារៈ',
    'ផាត់ ចាន់សុផាណា': 'ផាត់ ចាន់សុផាន់ណា',
    'ស៊ីន ម៉ូនីដា': 'ស៊ិន ម៉ូនីដា',
    'សូ វណ្ណ:': 'សូ វណ្ណៈ',
    'ទុន វណ្ណ:': 'ទុន វណ្ណៈ',
}

db_teachers = list(Teacher.objects.all())
db_map = {}
for t in db_teachers:
    norm = t.khmer_name.strip().replace('\u200b', ' ').replace('\xa0', ' ')
    db_map[norm] = t
    db_map[norm.replace(' ', '')] = t

# Target Academic Year
ay = AcademicYear.objects.filter(name='2026-2027').first()
if not ay:
    ay = AcademicYear.objects.filter(is_current=True).first()

print(f"Assigning ClassSubjects for Academic Year: '{ay.name}'")

# Map classrooms in 2026-2027 by code
classrooms_2026 = {c.code.upper().strip(): c for c in Classroom.objects.filter(academic_year=ay)}

# Subjects map by code
subjects_map = {s.code.upper().strip(): s for s in Subject.objects.all()}
# Also case-preserving mapping
for s in Subject.objects.all():
    subjects_map[s.code] = s

print(f"Found {len(classrooms_2026)} classrooms in '{ay.name}' and {len(subjects_map)} subjects.")

updated_excel_count = 0
assigned_class_subjects = 0
teacher_duty_updated = 0

for r_idx in range(2, ws.max_row + 1):
    no = ws.cell(row=r_idx, column=1).value
    symbol_val = ws.cell(row=r_idx, column=2).value
    name_val = ws.cell(row=r_idx, column=3).value
    gender_val = ws.cell(row=r_idx, column=4).value
    classes_val = ws.cell(row=r_idx, column=5).value
    
    symbol = str(symbol_val).strip() if symbol_val is not None else ''
    name = str(name_val).strip().replace('\u200b', ' ').replace('\xa0', ' ') if name_val is not None else ''
    classes_str = str(classes_val).strip() if classes_val is not None else ''
    
    if not symbol and not name:
        continue
        
    lookup_name = MANUAL_REPLACEMENTS.get(name, name)
    clean_no_sp = lookup_name.replace(' ', '')
    
    t_obj = db_map.get(lookup_name) or db_map.get(clean_no_sp)
    if not t_obj:
        for k, v in db_map.items():
            if clean_no_sp in k or k in clean_no_sp:
                t_obj = v
                break
                
    if not t_obj:
        print(f"❌ Error: Teacher '{name}' could not be matched!")
        continue

    # 1. Update Excel cell with exact clean DB name
    correct_name = t_obj.khmer_name
    if name != correct_name:
        ws.cell(row=r_idx, column=3).value = correct_name
        updated_excel_count += 1
        print(f"Excel Row {r_idx:3d}: Name updated '{name}' -> '{correct_name}'")

    # 2. Update Teacher current_duty in DB
    duty_title = f"{symbol} ({t_obj.specialization})" if symbol else t_obj.specialization
    t_obj.current_duty = symbol if symbol else t_obj.current_duty
    t_obj.save(update_fields=['current_duty'])
    teacher_duty_updated += 1

    # 3. Assign ClassSubjects
    sub_code = get_subject_code_from_symbol(symbol)
    parsed_classes = parse_class_tokens(classes_str)
    
    if sub_code and parsed_classes:
        sub_obj = subjects_map.get(sub_code.upper()) or subjects_map.get(sub_code)
        if sub_obj:
            for ccode in parsed_classes:
                c_obj = classrooms_2026.get(ccode.upper())
                if c_obj:
                    cs_obj, created_cs = ClassSubject.objects.update_or_create(
                        classroom=c_obj,
                        subject=sub_obj,
                        defaults={
                            'teacher': t_obj,
                            'weekly_hours': 2
                        }
                    )
                    assigned_class_subjects += 1
                else:
                    print(f"⚠️ Warning: Classroom '{ccode}' in 2026-2027 not found for teacher '{t_obj.khmer_name}'!")
        else:
            print(f"⚠️ Warning: Subject '{sub_code}' not found in Subject table!")

# Save Excel
wb.save(file_path)
print(f"\n========================================================")
print(f"EXCEL UPDATE & DB ASSIGNMENT COMPLETED!")
print(f"- Corrected Teacher Names in Excel: {updated_excel_count}")
print(f"- Updated Teacher Duties in DB: {teacher_duty_updated}")
print(f"- Created/Updated ClassSubject Assignments: {assigned_class_subjects}")
print(f"- Saved updated Excel file: '{file_path}'")
print(f"========================================================")
