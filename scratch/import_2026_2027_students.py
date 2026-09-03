import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from datetime import datetime, date
from django.db import transaction
from apps.academics.models import AcademicYear, Classroom
from apps.students.models import Student
from apps.students.khmer_romanizer import romanize_khmer_name

file_path = r'E:\SchoolSM\2026-2027.xlsm'

print("=== STARTING FULL IMPORT FROM E:\\SchoolSM\\2026-2027.xlsm ===")

# 1. Get or setup AcademicYear '2026-2027'
academic_year = AcademicYear.objects.filter(name='2026-2027').first()
if not academic_year:
    academic_year = AcademicYear.objects.create(
        name='2026-2027',
        start_date=date(2026, 9, 1),
        end_date=date(2027, 7, 15),
        is_current=True
    )
academic_year.is_current = True
academic_year.save()
AcademicYear.objects.exclude(id=academic_year.id).update(is_current=False)
print(f"Academic Year set to: '{academic_year.name}' (is_current={academic_year.is_current})")

# 2. Ensure all 40 classrooms exist in 2026-2027
CLASS_CONFIGS = {
    # Grade 7
    '7A': ('ថ្នាក់ទី ៧A', 7, 'GENERAL'),
    '7B': ('ថ្នាក់ទី ៧B', 7, 'GENERAL'),
    '7C': ('ថ្នាក់ទី ៧C', 7, 'GENERAL'),
    '7D': ('ថ្នាក់ទី ៧D', 7, 'GENERAL'),
    '7E': ('ថ្នាក់ទី ៧E', 7, 'GENERAL'),
    # Grade 8
    '8A': ('ថ្នាក់ទី ៨A', 8, 'GENERAL'),
    '8B': ('ថ្នាក់ទី ៨B', 8, 'GENERAL'),
    '8C': ('ថ្នាក់ទី ៨C', 8, 'GENERAL'),
    '8D': ('ថ្នាក់ទី ៨D', 8, 'GENERAL'),
    # Grade 9
    '9A': ('ថ្នាក់ទី ៩A', 9, 'GENERAL'),
    '9B': ('ថ្នាក់ទី ៩B', 9, 'GENERAL'),
    '9C': ('ថ្នាក់ទី ៩C', 9, 'GENERAL'),
    '9D': ('ថ្នាក់ទី ៩D', 9, 'GENERAL'),
    # Grade 10
    '10A': ('ថ្នាក់ទី ១០A', 10, 'GENERAL'),
    '10B': ('ថ្នាក់ទី ១០B', 10, 'GENERAL'),
    '10C': ('ថ្នាក់ទី ១០C', 10, 'GENERAL'),
    '10D': ('ថ្នាក់ទី ១០D', 10, 'GENERAL'),
    '10E': ('ថ្នាក់ទី ១០E', 10, 'GENERAL'),
    '10F': ('ថ្នាក់ទី ១០F', 10, 'GENERAL'),
    '10G': ('ថ្នាក់ទី ១០G', 10, 'GENERAL'),
    '10H': ('ថ្នាក់ទី ១០H', 10, 'GENERAL'),
    '10I': ('ថ្នាក់ទី ១០I', 10, 'GENERAL'),
    # Grade 11
    '11A': ('ថ្នាក់ទី ១១A', 11, 'SCIENCE'),
    '11B': ('ថ្នាក់ទី ១១B', 11, 'SCIENCE'),
    '11C': ('ថ្នាក់ទី ១១C', 11, 'SCIENCE'),
    '11D': ('ថ្នាក់ទី ១១D', 11, 'SCIENCE'),
    '11E': ('ថ្នាក់ទី ១១E', 11, 'SCIENCE'),
    '11F': ('ថ្នាក់ទី ១១F', 11, 'SOCIAL'),
    '11G': ('ថ្នាក់ទី ១១G', 11, 'SOCIAL'),
    '11H': ('ថ្នាក់ទី ១១H', 11, 'SOCIAL'),
    '11I': ('ថ្នាក់ទី ១១I', 11, 'SOCIAL'),
    # Grade 12
    '12A': ('ថ្នាក់ទី ១២A', 12, 'SCIENCE'),
    '12B': ('ថ្នាក់ទី ១២B', 12, 'SCIENCE'),
    '12C': ('ថ្នាក់ទី ១២C', 12, 'SCIENCE'),
    '12D': ('ថ្នាក់ទី ១២D', 12, 'SCIENCE'),
    '12E': ('ថ្នាក់ទី ១២E', 12, 'SCIENCE'),
    '12F': ('ថ្នាក់ទី ១២F', 12, 'SOCIAL'),
    '12G': ('ថ្នាក់ទី ១២G', 12, 'SOCIAL'),
    '12H': ('ថ្នាក់ទី ១២H', 12, 'SOCIAL'),
    '12I': ('ថ្នាក់ទី ១២I', 12, 'SOCIAL'),
}

classroom_map = {}
for code, (cname, gr, tr) in CLASS_CONFIGS.items():
    c_obj, _ = Classroom.objects.get_or_create(
        academic_year=academic_year,
        code=code,
        defaults={
            'name': cname,
            'grade_level': gr,
            'track': tr,
            'capacity': 50
        }
    )
    c_obj.name = cname
    c_obj.grade_level = gr
    c_obj.track = tr
    c_obj.save()
    classroom_map[code] = c_obj

print(f"Ensured all {len(classroom_map)} classrooms are active in 2026-2027.")

# 3. Read and import students from all sheets
wb = openpyxl.load_workbook(file_path, data_only=True)

def parse_dob(val):
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    val_str = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val_str.split()[0], fmt).date()
        except ValueError:
            pass
    return None

def parse_gender(val):
    v = str(val).strip().upper() if val else ''
    if v in ['ស', 'ស្រី', 'F', 'FEMALE', 'GIRL']:
        return 'F'
    return 'M'

created_count = 0
updated_count = 0
errors = []

with transaction.atomic():
    for sname in ['7', '8', '9', '10', '11', '12']:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        sheet_created = 0
        sheet_updated = 0
        
        for r_idx, r in enumerate(rows[1:], start=2):
            if not r or not any(r):
                continue
            st_id = str(r[1]).strip() if r[1] is not None else ''
            st_name = str(r[2]).strip() if r[2] is not None else ''
            
            if not st_id or not st_name or st_id.lower() in ['none', '', 'ល.រ', 'អត្តលេខ']:
                continue
            
            gender = parse_gender(r[3])
            dob = parse_dob(r[4])
            grade_val = str(r[5]).strip() if r[5] is not None else sname
            class_let = str(r[6]).strip() if r[6] is not None else ''
            
            class_code = f"{grade_val}{class_let}".upper().replace(' ', '')
            target_classroom = classroom_map.get(class_code)
            
            if not target_classroom:
                errors.append(f"Sheet {sname}, Row {r_idx}: Class code '{class_code}' not found!")
                continue
            
            # Generate Romanized Latin Name (All Caps)
            latin_name = romanize_khmer_name(st_name)
            
            # Find or Create Student
            student = Student.objects.filter(student_id=st_id).first()
            if student:
                student.khmer_name = st_name
                student.latin_name = latin_name
                student.gender = gender
                if dob:
                    student.date_of_birth = dob
                student.classroom = target_classroom
                student.academic_year = academic_year
                student.status = 'ACTIVE'
                student.save()
                sheet_updated += 1
                updated_count += 1
            else:
                Student.objects.create(
                    student_id=st_id,
                    khmer_name=st_name,
                    latin_name=latin_name,
                    gender=gender,
                    date_of_birth=dob or date(2010, 1, 1),
                    classroom=target_classroom,
                    academic_year=academic_year,
                    status='ACTIVE'
                )
                sheet_created += 1
                created_count += 1
                
        print(f"✅ Sheet '{sname}': {sheet_created} Created, {sheet_updated} Updated (Total: {sheet_created + sheet_updated})")

print(f"\n========================================================")
print(f"IMPORT SUMMARY:")
print(f"Total Created: {created_count}")
print(f"Total Updated: {updated_count}")
print(f"Grand Total Processed: {created_count + updated_count}")
if errors:
    print(f"Errors ({len(errors)}):", errors[:10])
else:
    print("Zero errors encountered!")
print(f"========================================================")

# Verify counts per grade in 2026-2027
print("\n=== CLASSROOM ROSTER SUMMARY IN 2026-2027 ===")
for grade in [7, 8, 9, 10, 11, 12]:
    classes = Classroom.objects.filter(academic_year=academic_year, grade_level=grade).order_by('code')
    total_in_grade = Student.objects.filter(classroom__in=classes, status='ACTIVE').count()
    female_in_grade = Student.objects.filter(classroom__in=classes, status='ACTIVE', gender='F').count()
    class_details = []
    for c in classes:
        c_tot = Student.objects.filter(classroom=c, status='ACTIVE').count()
        c_fem = Student.objects.filter(classroom=c, status='ACTIVE', gender='F').count()
        class_details.append(f"{c.code}: {c_tot} (F: {c_fem})")
    print(f"Grade {grade:2d} (Total: {total_in_grade:3d}, F: {female_in_grade:3d}) -> " + ", ".join(class_details))
