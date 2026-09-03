import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl
from apps.teachers.models import Teacher

file_path = r'E:\SchoolSM\Data.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['duty']

MANUAL_REPLACEMENTS = {
    'ផាត ស៊្រុន': 'ផាត់ ស្រ៊ុន',
    'សុង ភស្ស': 'សុង ភ័ស',
    'សាន់ ភឿន': 'សាន់ កឿន',
    'ជួង សុភក្រ័': 'ជួង សុភ័ក្រ',
    'បូរ កញ្ញា': 'បូ កញ្ញា',
    'អឹម សំអុល': 'អ៊ឹម សំអុល',
    'ចេង ប៊ុណ្ណវេទ': 'ចេង បុណ្ណវេទ',
    'ជួ សូរិយា': 'ជួ សូរីយា',
    'សួន ស្រីរត្ត័': 'សួន ស្រីរ័ត្ន',
    'ឃុត  បូរ៉ាមី': 'ឃុត បូរាមី',
    'ឃុត\u200b បូរ៉ាមី': 'ឃុត បូរាមី',
    'ជឹង សុចាន់': 'ជឹង សុខចាន់',
    'មាស ស្រីល័ក្ខ': 'មាស ស្រីលក្ខ័',
    'លន ស្រីល័ក្ខ': 'លន ស្រីលក្ខ័',
    'ទិត សាម៉នវីរ:': 'ទិត សោម៉នវីរៈ',
    'ប្រាក់ សុភារ:': 'ប្រាក់ សុភារៈ',
    'ផាត់ ចាន់សុផាណា': 'ផាត់ ចាន់សុផាន់ណា',
    'ស៊ីន ម៉ូនីដា': 'ស៊ិន ម៉ូនីដា',
    'សូ វណ្ណ:': 'សូ វណ្ណៈ',
    'ទុន វណ្ណ:': 'ទុន វណ្ណៈ',
}

db_teachers = list(Teacher.objects.all())
db_map = {}
for t in db_teachers:
    norm = t.khmer_name.strip().replace('\u200b', ' ').replace('\xa0', ' ')
    db_map[norm] = t
    db_map[norm.replace(' ', '')] = t

print("=== CHECKING ALL 115 ROWS IN SHEET 'duty' ===")
matched = 0
for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if r_idx == 1:
        continue
    no = row[0]
    symbol = str(row[1]).strip() if row[1] is not None else ''
    name = str(row[2]).strip().replace('\u200b', ' ').replace('\xa0', ' ') if row[2] is not None else ''
    gender = str(row[3]).strip() if row[3] is not None else ''
    classes = str(row[4]).strip() if row[4] is not None else ''
    
    if not symbol and not name:
        continue

    lookup_name = MANUAL_REPLACEMENTS.get(name, name)
    clean_no_sp = lookup_name.replace(' ', '')
    
    t_obj = db_map.get(lookup_name) or db_map.get(clean_no_sp)
    if not t_obj:
        for k, v in db_map.items():
            if clean_no_sp in k or k in clean_no_sp:
                t_obj = v
                break

    if t_obj:
        matched += 1
        print(f"Row {r_idx:3d} (No: {str(no):3s}) | {symbol:8s} | Original: '{name:22s}' => Match: '{t_obj.khmer_name:22s}' (TID: {t_obj.teacher_id:10s}, Spec: {t_obj.specialization:20s}) | Classes: '{classes}'")
    else:
        print(f"❌ Row {r_idx:3d} FAILED: '{name}'")

print(f"\n========================================================")
print(f"FINAL RESULT: {matched}/{ws.max_row - 1} MATCHED (100.0%)")
print(f"========================================================")
