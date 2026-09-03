import os, sys, django
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath('.'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

import openpyxl

wb = openpyxl.load_workbook(r'E:\SchoolSM\Data.xlsx', data_only=True)
print("Sheet names in Data.xlsx:", wb.sheetnames)

if 'GT' in wb.sheetnames:
    ws = wb['GT']
    print(f"\n=== SHEET 'GT' (Rows: {ws.max_row}, Cols: {ws.max_column}) ===")
    for idx, r in enumerate(list(ws.iter_rows(values_only=True))[:50], start=1):
        if any(r):
            print(f"Row {idx:2d}: {[str(c) if c is not None else '' for c in r[:15]]}")
else:
    print("Sheet 'GT' not found!")
