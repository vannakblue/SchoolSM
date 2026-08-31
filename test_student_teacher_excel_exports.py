import os
import django
import io
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'school_management.settings')
django.setup()

from django.test import Client
from apps.accounts.models import User
from apps.academics.models import Classroom
from apps.teachers.models import Teacher

def run_tests():
    print("=== TESTING STUDENT & TEACHER TIMETABLE EXCEL EXPORT (INDIVIDUAL & ALL SHEETS) ===")

    admin_user = User.objects.filter(role='ADMIN').first()
    if not admin_user:
        admin_user = User.objects.create_superuser('admin_tt_test', 'admin_tt@test.com', 'adminpass123')

    client = Client()
    client.force_login(admin_user)

    # 1. Test All Classrooms Multi-Sheet Excel Export
    resp_all_cls = client.get('/academics/timetable/student-teacher/export-excel/?mode=class&id=all')
    assert resp_all_cls.status_code == 200, f"Failed: {resp_all_cls.status_code}"
    assert resp_all_cls['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    wb_all_cls = openpyxl.load_workbook(io.BytesIO(resp_all_cls.content))
    print(f"1. [PASS] All Classrooms Multi-Sheet Export: Generated {len(wb_all_cls.sheetnames)} sheets -> {wb_all_cls.sheetnames[:5]}...")
    assert len(wb_all_cls.sheetnames) > 0

    # 2. Test Single Classroom Excel Export
    cls = Classroom.objects.first()
    if cls:
        resp_single_cls = client.get(f'/academics/timetable/student-teacher/export-excel/?mode=class&id={cls.id}')
        assert resp_single_cls.status_code == 200
        wb_single_cls = openpyxl.load_workbook(io.BytesIO(resp_single_cls.content))
        assert len(wb_single_cls.sheetnames) == 1
        ws = wb_single_cls.active
        assert "កាលវិភាគបង្រៀន និងរៀនប្រចាំសប្តាហ៍" in str(ws['A2'].value)
        print(f"2. [PASS] Single Class Export for [{cls.name}]: Sheet [{ws.title}] formatted successfully!")

    # 3. Test All Teachers Multi-Sheet Excel Export
    resp_all_tch = client.get('/academics/timetable/student-teacher/export-excel/?mode=teacher&id=all')
    assert resp_all_tch.status_code == 200
    wb_all_tch = openpyxl.load_workbook(io.BytesIO(resp_all_tch.content))
    print(f"3. [PASS] All Teachers Multi-Sheet Export: Generated {len(wb_all_tch.sheetnames)} sheets -> {wb_all_tch.sheetnames[:5]}...")
    assert len(wb_all_tch.sheetnames) > 0

    # 4. Test Single Teacher Excel Export
    tch = Teacher.objects.filter(status='ACTIVE').first()
    if tch:
        resp_single_tch = client.get(f'/academics/timetable/student-teacher/export-excel/?mode=teacher&id={tch.id}')
        assert resp_single_tch.status_code == 200
        wb_single_tch = openpyxl.load_workbook(io.BytesIO(resp_single_tch.content))
        assert len(wb_single_tch.sheetnames) == 1
        ws_tch = wb_single_tch.active
        assert "កាលវិភាគបង្រៀនប្រចាំសប្តាហ៍" in str(ws_tch['A2'].value)
        print(f"4. [PASS] Single Teacher Export for [{tch.khmer_name}]: Sheet [{ws_tch.title}] formatted successfully!")

    print("\n=== ALL STUDENT & TEACHER TIMETABLE EXCEL TESTS PASSED 100%! ===")

if __name__ == '__main__':
    run_tests()
